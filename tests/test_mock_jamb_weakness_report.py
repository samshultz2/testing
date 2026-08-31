"""Printable failure/weakness report: the questions, topics and sub-topics the
cohort got wrong, rendered server-side to PDF / Excel / image."""
from utils.mock_deep_report import (weakness_pdf, weakness_xlsx, weakness_png,
                                    weakness_filename, _weakness_lists)


def _payload():
    return {
        'meta': {'exam_name': 'Mock 2', 'session_name': '2024/2025',
                 'exam_date': '10 May 2025', 'sitters': 42, 'cohort_mastery': 48},
        'subject_mastery': [{'subject': 'Mathematics', 'items': 40, 'served': 400,
                             'correct': 150, 'mastery': 37.5, 'band': 'critical',
                             'band_label': 'Critical'}],
        'items': [
            {'subject': 'Mathematics', 'topic': 'Algebra', 'subtopic': 'Quadratic equations',
             'text': 'Solve \\(3x^2+2x-5=0\\)', 'correct_option': 'B', 'served': 40,
             'answered': 38, 'blank': 2, 'correct': 9, 'p_value': 22.5, 'p_raw': 0.225,
             'blank_rate': 5.0},
            {'subject': 'Physics', 'topic': 'Mechanics', 'subtopic': 'Projectile motion',
             'text': 'A ball is thrown at 20 m/s', 'correct_option': 'C', 'served': 40,
             'answered': 30, 'blank': 10, 'correct': 12, 'p_value': 30.0, 'p_raw': 0.30,
             'blank_rate': 25.0},
            {'subject': 'Biology', 'topic': 'Cell biology', 'subtopic': 'Organelles',
             'text': 'The powerhouse of the cell', 'correct_option': 'A', 'served': 40,
             'answered': 40, 'blank': 0, 'correct': 36, 'p_value': 90.0, 'p_raw': 0.90,
             'blank_rate': 0.0},           # mastered — must NOT appear as failed
        ],
        'topics': [
            {'subject': 'Mathematics', 'topic': 'Algebra', 'items': 10, 'served': 100,
             'correct': 30, 'mastery': 30.0, 'band': 'critical', 'band_label': 'Critical',
             'subtopics': [{'subtopic': 'Quadratic equations', 'items': 4, 'served': 40,
                            'correct': 9, 'mastery': 22.5, 'band': 'critical',
                            'band_label': 'Critical'}]},
            {'subject': 'Biology', 'topic': 'Cell biology', 'items': 6, 'served': 60,
             'correct': 54, 'mastery': 90.0, 'band': 'strong', 'band_label': 'Mastered',
             'subtopics': [{'subtopic': 'Organelles', 'items': 2, 'served': 40, 'correct': 36,
                            'mastery': 90.0, 'band': 'strong', 'band_label': 'Mastered'}]},
        ],
        'flagged': [], 'blank_heavy': [],
        'recommendations': {
            'students': [{'tone': 'negative', 'title': 'Re-drill algebra', 'text': 'Focus on quadratics.'}],
            'teachers': [{'tone': 'warning', 'title': 'Clinic', 'text': 'Run a quadratics clinic.'}],
            'management': []},
    }


def test_weakness_lists_pick_failures_only():
    most_failed, weak_topics, weak_subs = _weakness_lists(_payload())
    # the mastered Biology item / topic / sub-topic must be excluded
    assert [it['subject'] for it in most_failed] == ['Mathematics', 'Physics']  # weakest first
    assert all(t['mastery'] < 70 for t in weak_topics)
    assert [t['topic'] for t in weak_topics] == ['Algebra']
    assert [s['subtopic'] for s in weak_subs] == ['Quadratic equations']


def test_weakness_pdf_renders_and_converts_latex(app):
    with app.app_context():
        pdf = weakness_pdf(_payload())
    assert pdf[:4] == b'%PDF' and len(pdf) > 1500


def test_weakness_xlsx_has_three_failure_sheets(app):
    from openpyxl import load_workbook
    from io import BytesIO
    with app.app_context():
        xl = weakness_xlsx(_payload())
    wb = load_workbook(BytesIO(xl))
    assert wb.sheetnames == ['Questions failed', 'Topics failed', 'Sub-topics failed']
    # LaTeX in the question is converted to readable text (no raw \\( … \\))
    q_cell = wb['Questions failed']['G2'].value
    assert '3x²' in q_cell and '\\(' not in q_cell


def test_weakness_png_renders(app):
    with app.app_context():
        png = weakness_png(_payload())
    assert png[:4] == b'\x89PNG' and len(png) > 2000


def test_weakness_filename():
    assert weakness_filename({'exam_name': 'Mock 2'}, 'pdf') == 'mock_jamb_weaknesses_Mock_2.pdf'
