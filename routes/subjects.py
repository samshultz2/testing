"""
Subjects and Score Management routes
"""
from flask import Blueprint, render_template, request, redirect, url_for, flash, jsonify
from utils.helpers import get_active_term
from utils.web_exports import xlsx_response
from utils.db_tx import safe_transaction
from utils.branch_scope import require_branch_access
from models import (
    db, Subject, ClassSubject, AssessmentType, SubjectAssessmentOverride,
    StudentScore, StudentEnrollment, ClassArmAssignment, Term, SchoolClass,
    ClassArm, Student, GradeScale, SchoolSettings, TermSummary
)
from utils.access_control import (
    login_required, can_access_class, can_enter_results,
    filter_classes_for_user, is_admin, result_card_required
)

subjects_bp = Blueprint('subjects', __name__, url_prefix='/subjects')

SUBJECT_CATEGORIES = ['Science', 'Arts', 'Commercial', 'General', 'Languages', 'Vocational']


# --- SPA helpers (no-reload React shell + JSON-aware action responses) ---
from utils.spa import section_responders
_wants_json, _render, _ok, _err = section_responders(
    'subjects/app.html', 'subj_json', 'subjects.subjects_list')


def _nav_urls():
    return {'subjects': url_for('subjects.subjects_list'),
            'class_subjects': url_for('subjects.class_subjects_list')}


# ============================================================================
# SUBJECTS CRUD
# ============================================================================

@subjects_bp.route('/')
@login_required
def subjects_list():
    """List all subjects"""
    from utils.org_scope import scope_subjects
    subjects = scope_subjects(
        Subject.query.filter_by(is_active=True), Subject
    ).order_by(Subject.category, Subject.name).all()
    
    # Group by category
    categories = {}
    for subject in subjects:
        cat = subject.category or 'General'
        categories.setdefault(cat, []).append({
            'id': subject.id, 'name': subject.name, 'short_name': subject.short_name or '',
            'edit_url': url_for('subjects.edit_subject', subject_id=subject.id),
            'delete_url': url_for('subjects.delete_subject', subject_id=subject.id),
        })

    return _render({
        'page': 'list', 'nav': _nav_urls(),
        'categories': [{'name': k, 'subjects': v} for k, v in categories.items()],
        'urls': {'add': url_for('subjects.add_subject'),
                 'bulk_add': url_for('subjects.bulk_add_subjects')},
    })


@subjects_bp.route('/add', methods=['GET', 'POST'])
@login_required
def add_subject():
    """Add a new subject"""
    if request.method == 'POST':
        try:
            name = request.form.get('name', '').strip()
            short_name = request.form.get('short_name', '').strip().upper()
            category = request.form.get('category', '').strip()

            if not name:
                return _err('Subject name is required.', url_for('subjects.add_subject'))
            if Subject.query.filter_by(name=name).first():
                return _err('Subject already exists.', url_for('subjects.add_subject'))

            subject = Subject(
                name=name,
                short_name=short_name or name[:3].upper(),
                category=category or 'General',
                has_practical=bool(request.form.get('has_practical'))
            )
            db.session.add(subject)
            db.session.flush()
            from utils.assessments import apply_practical
            apply_practical(db, subject)
            db.session.commit()
            return _ok(f'Subject "{name}" added!', url_for('subjects.subjects_list'))
        except Exception as e:
            db.session.rollback()
            return _err(f'Error: {str(e)}', url_for('subjects.add_subject'))

    return _render({
        'page': 'add', 'nav': _nav_urls(), 'categories': SUBJECT_CATEGORIES,
        'submit_url': url_for('subjects.add_subject'), 'cancel_url': url_for('subjects.subjects_list'),
    })


@subjects_bp.route('/<int:subject_id>/edit', methods=['GET', 'POST'])
@login_required
def edit_subject(subject_id):
    """Edit a subject"""
    subject = db.get_or_404(Subject, subject_id)

    if request.method == 'POST':
        try:
            subject.name = request.form.get('name', '').strip()
            subject.short_name = request.form.get('short_name', '').strip().upper()
            subject.category = request.form.get('category', '').strip()
            subject.has_practical = bool(request.form.get('has_practical'))
            from utils.assessments import apply_practical
            apply_practical(db, subject)
            db.session.commit()
            return _ok('Subject updated!', url_for('subjects.subjects_list'))
        except Exception as e:
            db.session.rollback()
            return _err(f'Error: {str(e)}', url_for('subjects.edit_subject', subject_id=subject_id))

    return _render({
        'page': 'edit', 'nav': _nav_urls(), 'categories': SUBJECT_CATEGORIES,
        'subject': {'id': subject.id, 'name': subject.name, 'short_name': subject.short_name or '',
                    'category': subject.category or 'General', 'has_practical': bool(subject.has_practical)},
        'submit_url': url_for('subjects.edit_subject', subject_id=subject.id),
        'cancel_url': url_for('subjects.subjects_list'),
    })


@subjects_bp.route('/<int:subject_id>/delete', methods=['POST'])
@login_required
def delete_subject(subject_id):
    """Delete (deactivate) a subject"""
    subject = db.get_or_404(Subject, subject_id)
    try:
        subject.is_active = False
        db.session.commit()
        return _ok('Subject deleted!', url_for('subjects.subjects_list'))
    except Exception as e:
        db.session.rollback()
        return _err(f'Error: {str(e)}', url_for('subjects.subjects_list'))


@subjects_bp.route('/bulk-add', methods=['GET', 'POST'])
@login_required
def bulk_add_subjects():
    """Bulk add subjects"""
    if request.method == 'POST':
        try:
            subjects_text = request.form.get('subjects', '')
            category = request.form.get('category', 'General')

            lines = [line.strip() for line in subjects_text.split('\n') if line.strip()]
            added = 0
            for line in lines:
                if not Subject.query.filter_by(name=line).first():
                    db.session.add(Subject(name=line, short_name=line[:3].upper(), category=category))
                    added += 1

            db.session.commit()
            return _ok(f'{added} subjects added!', url_for('subjects.subjects_list'))
        except Exception as e:
            db.session.rollback()
            return _err(f'Error: {str(e)}', url_for('subjects.bulk_add_subjects'))

    default_subjects = '\n'.join([
        'Mathematics', 'English Language', 'Civic Education', 'Physics', 'Chemistry', 'Biology',
        'Further Mathematics', 'Agricultural Science', 'Economics', 'Geography', 'Government',
        'Literature in English', 'Commerce', 'Accounting', 'Computer Studies',
        'Christian Religious Studies', 'Islamic Religious Studies', 'French', 'Yoruba', 'Igbo',
        'Hausa', 'Technical Drawing', 'Food and Nutrition', 'Home Economics', 'Physical Education',
        'History', 'Music', 'Visual Arts'])
    return _render({
        'page': 'bulk_add', 'nav': _nav_urls(), 'categories': SUBJECT_CATEGORIES,
        'default_subjects': default_subjects,
        'submit_url': url_for('subjects.bulk_add_subjects'), 'cancel_url': url_for('subjects.subjects_list'),
    })


# ============================================================================
# CLASS-SUBJECT ASSIGNMENT
# ============================================================================

@subjects_bp.route('/class-subjects')
@login_required
def class_subjects_list():
    """List class-subject assignments for a term"""
    term_id = request.args.get('term_id', type=int)
    class_id = request.args.get('class_id', type=int)
    
    terms = Term.query.order_by(Term.id.desc()).all()
    classes = SchoolClass.query.order_by(SchoolClass.level).all()
    
    # Get active term if not specified
    if not term_id:
        active_term = get_active_term()
        if active_term:
            term_id = active_term.id
    
    class_subjects = []
    selected_term = None
    selected_class = None
    
    if term_id:
        selected_term = db.session.get(Term, term_id)
        query = ClassSubject.query.filter_by(term_id=term_id, is_active=True)
        
        if class_id:
            selected_class = db.session.get(SchoolClass, class_id)
            query = query.filter_by(class_id=class_id)
        
        class_subjects = query.join(Subject).order_by(Subject.name).all()

    return _render({
        'page': 'class_subjects', 'nav': _nav_urls(),
        'term_id': term_id or '', 'class_id': class_id or '',
        'selected_term': selected_term.full_name if selected_term else '',
        'terms': [{'id': t.id, 'full_name': t.full_name} for t in terms],
        'classes': [{'id': c.id, 'name': c.name} for c in classes],
        'class_subjects': [{'id': cs.id, 'subject': cs.subject.name,
                            'class_name': cs.school_class.name if cs.school_class else '',
                            'arm': cs.arm.name if cs.arm else '', 'teacher_name': cs.teacher_name or '',
                            'edit_url': url_for('subjects.edit_class_subject', cs_id=cs.id),
                            'delete_url': url_for('subjects.delete_class_subject', cs_id=cs.id)}
                           for cs in class_subjects],
        'self_url': url_for('subjects.class_subjects_list'),
        'urls': {'assign': url_for('subjects.assign_class_subjects'),
                 'copy': url_for('subjects.copy_class_subjects')},
    })


@subjects_bp.route('/class-subjects/copy', methods=['POST'])
@login_required
def copy_class_subjects():
    """Copy subject assignments from one term into another (modifiable later)."""
    from_term_id = request.form.get('from_term_id', type=int)
    to_term_id = request.form.get('to_term_id', type=int)
    class_id = request.form.get('class_id', type=int)

    if not from_term_id or not to_term_id or from_term_id == to_term_id:
        return _err('Choose two different terms to copy between.',
                    url_for('subjects.class_subjects_list', term_id=to_term_id or '', class_id=class_id or ''))

    source = ClassSubject.query.filter_by(term_id=from_term_id, is_active=True)
    if class_id:
        source = source.filter_by(class_id=class_id)
    source = source.all()

    copied, skipped = 0, 0
    for cs in source:
        exists = ClassSubject.query.filter_by(
            term_id=to_term_id, class_id=cs.class_id, arm_id=cs.arm_id,
            subject_id=cs.subject_id
        ).first()
        if exists:
            skipped += 1
            continue
        db.session.add(ClassSubject(
            subject_id=cs.subject_id, class_id=cs.class_id, arm_id=cs.arm_id,
            term_id=to_term_id, teacher_name=cs.teacher_name, is_active=True))
        copied += 1
    db.session.commit()
    return _ok(f'Copied {copied} subject assignment(s){" (" + str(skipped) + " already existed)" if skipped else ""}.',
               url_for('subjects.class_subjects_list', term_id=to_term_id, class_id=class_id or ''))


@subjects_bp.route('/class-subjects/assign', methods=['GET', 'POST'])
@login_required
def assign_class_subjects():
    """Assign subjects to a class for a term"""
    if request.method == 'POST':
        try:
            term_id = request.form.get('term_id', type=int)
            class_id = request.form.get('class_id', type=int)
            arm_id = request.form.get('arm_id', type=int) or None
            subject_ids = request.form.getlist('subject_ids[]')
            teacher_names = request.form.getlist('teacher_names[]')
            
            if not term_id or not class_id:
                return _err('Term and class are required.', url_for('subjects.assign_class_subjects'))

            added = 0
            for i, subject_id in enumerate(subject_ids):
                if subject_id:
                    # Check if already exists
                    existing = ClassSubject.query.filter_by(
                        subject_id=int(subject_id),
                        class_id=class_id,
                        arm_id=arm_id,
                        term_id=term_id
                    ).first()
                    
                    if existing:
                        # Update teacher name
                        existing.teacher_name = teacher_names[i] if i < len(teacher_names) else None
                        existing.is_active = True
                    else:
                        cs = ClassSubject(
                            subject_id=int(subject_id),
                            class_id=class_id,
                            arm_id=arm_id,
                            term_id=term_id,
                            teacher_name=teacher_names[i] if i < len(teacher_names) else None
                        )
                        db.session.add(cs)
                        added += 1
            
            db.session.commit()
            return _ok(f'{added} subjects assigned!',
                       url_for('subjects.class_subjects_list', term_id=term_id, class_id=class_id))
        except Exception as e:
            db.session.rollback()
            return _err(f'Error: {str(e)}', url_for('subjects.assign_class_subjects'))

    terms = Term.query.order_by(Term.id.desc()).all()
    classes = SchoolClass.query.order_by(SchoolClass.level).all()
    arms = ClassArm.query.order_by(ClassArm.name).all()
    subjects = Subject.query.filter_by(is_active=True).order_by(Subject.name).all()

    return _render({
        'page': 'assign', 'nav': _nav_urls(),
        'terms': [{'id': t.id, 'full_name': t.full_name, 'is_active': bool(t.is_active)} for t in terms],
        'classes': [{'id': c.id, 'name': c.name} for c in classes],
        'arms': [{'id': a.id, 'name': a.name} for a in arms],
        'subjects': [{'id': s.id, 'name': s.name, 'category': s.category or ''} for s in subjects],
        'submit_url': url_for('subjects.assign_class_subjects'),
        'cancel_url': url_for('subjects.class_subjects_list'),
    })


@subjects_bp.route('/class-subjects/<int:cs_id>/edit', methods=['GET', 'POST'])
@login_required
def edit_class_subject(cs_id):
    """Edit class-subject assignment (mainly teacher name)"""
    cs = db.get_or_404(ClassSubject, cs_id)
    
    if request.method == 'POST':
        try:
            cs.teacher_name = request.form.get('teacher_name', '').strip() or None
            db.session.commit()
            return _ok('Updated!', url_for('subjects.class_subjects_list', term_id=cs.term_id, class_id=cs.class_id))
        except Exception as e:
            db.session.rollback()
            return _err(f'Error: {str(e)}', url_for('subjects.edit_class_subject', cs_id=cs.id))

    return _render({
        'page': 'edit_class_subject', 'nav': _nav_urls(),
        'cs': {'id': cs.id, 'subject': cs.subject.name,
               'class_name': cs.school_class.name if cs.school_class else '',
               'arm': cs.arm.name if cs.arm else 'All Arms',
               'term': cs.term.full_name if cs.term else '', 'teacher_name': cs.teacher_name or ''},
        'submit_url': url_for('subjects.edit_class_subject', cs_id=cs.id),
        'cancel_url': url_for('subjects.class_subjects_list', term_id=cs.term_id, class_id=cs.class_id),
    })


@subjects_bp.route('/class-subjects/<int:cs_id>/delete', methods=['POST'])
@login_required
def delete_class_subject(cs_id):
    """Remove class-subject assignment"""
    cs = db.get_or_404(ClassSubject, cs_id)
    term_id = cs.term_id
    class_id = cs.class_id
    
    try:
        cs.is_active = False
        db.session.commit()
        return _ok('Subject removed from class!',
                   url_for('subjects.class_subjects_list', term_id=term_id, class_id=class_id))
    except Exception as e:
        db.session.rollback()
        return _err(f'Error: {str(e)}',
                    url_for('subjects.class_subjects_list', term_id=term_id, class_id=class_id))


# ============================================================================
# SCORE ENTRY
# ============================================================================

@subjects_bp.route('/scores')
@login_required
def scores_entry():
    """Score entry page"""
    term_id = request.args.get('term_id', type=int)
    assignment_id = request.args.get('assignment_id', type=int)
    class_subject_id = request.args.get('class_subject_id', type=int)
    assessment_type_id = request.args.get('assessment_type_id', type=int)
    
    # Check if user can enter results
    if not can_enter_results() and not is_admin():
        flash('You do not have permission to enter scores.', 'error')
        return redirect(url_for('main.dashboard'))
    
    # Check class access
    if assignment_id and not can_access_class(assignment_id):
        flash('You do not have access to this class.', 'error')
        return redirect(url_for('subjects.scores_entry'))
    
    terms = Term.query.order_by(Term.id.desc()).all()
    
    # Get active term if not specified
    if not term_id:
        active_term = get_active_term()
        if active_term:
            term_id = active_term.id
    
    selected_term = db.session.get(Term, term_id) if term_id else None
    
    # Get class arm assignments for selected term (filtered for teachers)
    assignments = []
    if term_id:
        all_assignments = ClassArmAssignment.query.filter_by(term_id=term_id).all()
        assignments = filter_classes_for_user(all_assignments)
    
    selected_assignment = db.session.get(ClassArmAssignment, assignment_id) if assignment_id else None
    
    # Get subjects for selected class (filter by teacher's assigned subjects if not admin)
    class_subjects = []
    if selected_assignment:
        from utils.access_control import get_teacher_profile
        teacher = get_teacher_profile()
        
        class_subjects_query = ClassSubject.query.filter_by(
            term_id=term_id,
            class_id=selected_assignment.class_id,
            is_active=True
        ).filter(
            (ClassSubject.arm_id == None) | (ClassSubject.arm_id == selected_assignment.arm_id)
        ).join(Subject).order_by(Subject.name)
        
        all_class_subjects = class_subjects_query.all()
        
        # Filter by teacher's assigned subjects if not admin
        if teacher and not is_admin():
            teacher_subject_ids = [
                a.subject_id for a in teacher.subject_assignments.filter_by(
                    class_arm_assignment_id=assignment_id,
                    is_active=True
                ).all()
            ]
            class_subjects = [cs for cs in all_class_subjects if cs.subject_id in teacher_subject_ids]
        else:
            class_subjects = all_class_subjects
    
    selected_class_subject = db.session.get(ClassSubject, class_subject_id) if class_subject_id else None
    
    # Get assessment types
    assessment_types = AssessmentType.query.filter_by(is_active=True).order_by(AssessmentType.order).all()
    selected_assessment = db.session.get(AssessmentType, assessment_type_id) if assessment_type_id else None
    
    # Get students and existing scores
    students_data = []
    if selected_assignment and selected_class_subject and selected_assessment:
        enrollments = StudentEnrollment.query.filter_by(
            class_arm_assignment_id=assignment_id,
            is_active=True
        ).join(Student).order_by(Student.surname, Student.first_name).all()
        
        for enrollment in enrollments:
            # Get existing score
            existing_score = StudentScore.query.filter_by(
                student_id=enrollment.student_id,
                class_subject_id=class_subject_id,
                assessment_type_id=assessment_type_id
            ).first()
            
            students_data.append({
                'enrollment': enrollment,
                'student': enrollment.student,
                'score': existing_score.score if existing_score else None
            })
    
    # Get max score (check for override)
    max_score = selected_assessment.max_score if selected_assessment else 0
    if selected_class_subject and selected_assessment:
        override = SubjectAssessmentOverride.query.filter_by(
            subject_id=selected_class_subject.subject_id,
            assessment_type_id=assessment_type_id,
            is_active=True
        ).first()
        if override:
            max_score = override.max_score
    
    return _render({
        'page': 'scores', 'nav': _nav_urls(),
        'term_id': term_id or '', 'assignment_id': assignment_id or '',
        'class_subject_id': class_subject_id or '', 'assessment_type_id': assessment_type_id or '',
        'terms': [{'id': t.id, 'full_name': t.full_name} for t in terms],
        'assignments': [{'id': a.id, 'display_name': a.display_name} for a in assignments],
        'class_subjects': [{'id': cs.id, 'subject_name': cs.subject.name} for cs in class_subjects],
        'assessment_types': [{'id': at.id, 'name': at.name, 'max_score': at.max_score} for at in assessment_types],
        'selected_subject': selected_class_subject.subject.name if selected_class_subject else '',
        'selected_assessment': selected_assessment.name if selected_assessment else '',
        'max_score': max_score,
        'has_selection': bool(selected_assignment and selected_class_subject and selected_assessment),
        'students_data': [{'id': it['student'].id, 'full_name': it['student'].full_name,
                           'gender': it['student'].gender or '',
                           'score': it['score'] if it['score'] is not None else ''}
                          for it in students_data],
        'self_url': url_for('subjects.scores_entry'),
        'save_url': url_for('subjects.save_scores'),
        'urls': {'scan': url_for('subjects.scoresheet_scan', term_id=term_id or '', assignment_id=assignment_id or '', class_subject_id=class_subject_id or ''),
                 'paste': url_for('subjects.scoresheet_paste', term_id=term_id or '', assignment_id=assignment_id or '', class_subject_id=class_subject_id or ''),
                 'import': url_for('subjects.import_scores', term_id=term_id or '', assignment_id=assignment_id or '', class_subject_id=class_subject_id or '')},
    })


@subjects_bp.route('/scores/save', methods=['POST'])
@login_required
def save_scores():
    """Save student scores"""
    try:
        class_subject_id = request.form.get('class_subject_id', type=int)
        assessment_type_id = request.form.get('assessment_type_id', type=int)
        assignment_id = request.form.get('assignment_id', type=int)
        term_id = request.form.get('term_id', type=int)
        
        student_ids = request.form.getlist('student_id[]')
        scores = request.form.getlist('score[]')

        # A teacher may only save scores for a subject they actually teach in
        # this class (admins pass). Resolve the subject from the class-subject.
        cs = db.session.get(ClassSubject, class_subject_id) if class_subject_id else None
        if not can_enter_results(assignment_id, cs.subject_id if cs else None):
            return _err('You can only enter scores for the subjects you teach in this class.',
                        url_for('subjects.scores_entry', term_id=term_id, assignment_id=assignment_id))

        at = db.session.get(AssessmentType, assessment_type_id)
        max_score = at.max_score if at else None
        saved = 0
        rejected = 0
        for i, student_id in enumerate(student_ids):
            score_value = scores[i].strip() if i < len(scores) else ''

            if score_value:
                try:
                    score_float = float(score_value)
                except ValueError:
                    continue
                if max_score and score_float > max_score:
                    rejected += 1
                    continue
                # Get or create score record
                existing = StudentScore.query.filter_by(
                    student_id=int(student_id),
                    class_subject_id=class_subject_id,
                    assessment_type_id=assessment_type_id
                ).first()
                
                if existing:
                    existing.score = score_float
                else:
                    score = StudentScore(
                        student_id=int(student_id),
                        class_subject_id=class_subject_id,
                        assessment_type_id=assessment_type_id,
                        score=score_float
                    )
                    db.session.add(score)
                saved += 1
            else:
                # Remove score if empty
                StudentScore.query.filter_by(
                    student_id=int(student_id),
                    class_subject_id=class_subject_id,
                    assessment_type_id=assessment_type_id
                ).delete()
        
        db.session.commit()
        # Keep term results/positions fresh as scores change.
        if term_id and assignment_id:
            asg = db.session.get(ClassArmAssignment, assignment_id)
            if asg:
                from utils.report_card import compute_term_summaries
                compute_term_summaries(term_id, asg.class_id)
        msg = f'{saved} scores saved!'
        if rejected:
            msg += f' {rejected} skipped (above the {max_score:g} maximum).'
        return _ok(msg, url_for('subjects.scores_entry',
            term_id=term_id, assignment_id=assignment_id,
            class_subject_id=class_subject_id, assessment_type_id=assessment_type_id))
    except Exception as e:
        db.session.rollback()
        return _err(f'Error: {str(e)}', url_for('subjects.scores_entry',
            term_id=term_id, assignment_id=assignment_id,
            class_subject_id=class_subject_id, assessment_type_id=assessment_type_id))


# ============================================================================
# VIEW SCORES / BROADSHEET
# ============================================================================

@subjects_bp.route('/workflow')
@login_required
def workflow():
    """Guided results checklist for a class+term: setup → entry → finalize → print."""
    term_id = request.args.get('term_id', type=int)
    assignment_id = request.args.get('assignment_id', type=int)
    if assignment_id and not can_access_class(assignment_id):
        flash('You do not have access to this class.', 'error')
        return redirect(url_for('subjects.workflow'))
    if not term_id:
        active = get_active_term()
        term_id = active.id if active else None
    terms = Term.query.order_by(Term.id.desc()).all()
    assignments = (filter_classes_for_user(
        ClassArmAssignment.query.filter_by(term_id=term_id).all()) if term_id else [])
    selected = db.session.get(ClassArmAssignment, assignment_id) if assignment_id else None

    steps = None
    if selected and term_id:
        enr = StudentEnrollment.query.filter_by(
            class_arm_assignment_id=assignment_id, is_active=True).all()
        sids = [e.student_id for e in enr]
        css = ClassSubject.query.filter_by(
            term_id=term_id, class_id=selected.class_id, is_active=True).all()
        n_assess = AssessmentType.query.filter_by(is_active=True).count()
        entered = 0
        if sids and css:
            entered = (StudentScore.query.join(ClassSubject).filter(
                ClassSubject.term_id == term_id,
                ClassSubject.class_id == selected.class_id,
                StudentScore.student_id.in_(sids)).count())
        ts_rows = (TermSummary.query.filter(
            TermSummary.term_id == term_id,
            TermSummary.student_id.in_(sids or [-1])).all())
        steps = {
            'students': len(sids),
            'subjects': len(css),
            'scores_entered': entered,
            'scores_expected': len(sids) * len(css) * (n_assess or 1),
            'positions': sum(1 for t in ts_rows if t.position_in_class),
            'comments': sum(1 for t in ts_rows if t.teacher_comment),
            'behaviour': sum(1 for t in ts_rows if t.affective),
        }
    selected_term = db.session.get(Term, term_id) if term_id else None
    return _render({
        'page': 'workflow', 'nav': _nav_urls(),
        'term_id': term_id or '', 'assignment_id': assignment_id or '',
        'terms': [{'id': t.id, 'full_name': t.full_name} for t in terms],
        'assignments': [{'id': a.id, 'display_name': a.display_name} for a in assignments],
        'steps': steps, 'published': bool(selected_term and selected_term.results_published),
        'self_url': url_for('subjects.workflow'),
        'urls': {
            'class_subjects': url_for('subjects.class_subjects_list'),
            'enrol': url_for('academics.assignments_list'),
            'bulk_entry': url_for('subjects.bulk_entry', term_id=term_id or '', assignment_id=assignment_id or ''),
            'broadsheet': url_for('subjects.broadsheet', term_id=term_id or '', assignment_id=assignment_id or ''),
            'comments': url_for('subjects.comments', term_id=term_id or '', assignment_id=assignment_id or ''),
            'affective': url_for('subjects.affective', term_id=term_id or '', assignment_id=assignment_id or ''),
            'compute': url_for('subjects.compute_summaries'),
            'print_all': url_for('subjects.print_all_report_cards', term_id=term_id or '', assignment_id=assignment_id or ''),
            'publish': url_for('scratchcards.publish', term_id=term_id) if term_id else '',
        },
    })


@subjects_bp.route('/bulk-entry', methods=['GET', 'POST'])
@login_required
def bulk_entry():
    """Enter every subject's scores for a whole class on one screen."""
    term_id = request.values.get('term_id', type=int)
    assignment_id = request.values.get('assignment_id', type=int)
    if assignment_id and not can_access_class(assignment_id):
        flash('You do not have access to this class.', 'error')
        return redirect(url_for('subjects.bulk_entry'))
    if not term_id:
        active = get_active_term()
        term_id = active.id if active else None
    terms = Term.query.order_by(Term.id.desc()).all()
    assignments = (filter_classes_for_user(
        ClassArmAssignment.query.filter_by(term_id=term_id).all()) if term_id else [])
    selected = db.session.get(ClassArmAssignment, assignment_id) if assignment_id else None
    assessment_types = AssessmentType.query.filter_by(is_active=True).order_by(AssessmentType.order).all()

    class_subjects, enrollments = [], []
    if selected:
        class_subjects = (ClassSubject.query.filter_by(
            term_id=term_id, class_id=selected.class_id, is_active=True)
            .filter((ClassSubject.arm_id == None) | (ClassSubject.arm_id == selected.arm_id))
            .join(Subject).order_by(Subject.name).all())
        # A teacher only sees/saves the subjects they actually teach in this class
        # (admins/eligible staff keep all). Scopes both the grid and the save loop.
        class_subjects = [cs for cs in class_subjects
                          if can_enter_results(assignment_id, cs.subject_id)]
        enrollments = (StudentEnrollment.query.filter_by(
            class_arm_assignment_id=assignment_id, is_active=True)
            .join(Student).order_by(Student.surname, Student.first_name).all())

    if request.method == 'POST' and selected and class_subjects and enrollments:
        sids = [e.student_id for e in enrollments]
        cs_ids = [cs.id for cs in class_subjects]
        existing = {(s.student_id, s.class_subject_id, s.assessment_type_id): s
                    for s in StudentScore.query.filter(
                        StudentScore.student_id.in_(sids),
                        StudentScore.class_subject_id.in_(cs_ids)).all()}
        changed = 0
        rejected = 0
        for e in enrollments:
            for cs in class_subjects:
                for at in assessment_types:
                    raw = (request.form.get(f's_{e.student_id}_{cs.id}_{at.id}') or '').strip()
                    obj = existing.get((e.student_id, cs.id, at.id))
                    if raw:
                        try:
                            val = float(raw)
                        except ValueError:
                            continue
                        if at.max_score and val > at.max_score:
                            rejected += 1
                            continue
                        if obj:
                            if obj.score != val:
                                obj.score = val; changed += 1
                        else:
                            db.session.add(StudentScore(student_id=e.student_id,
                                class_subject_id=cs.id, assessment_type_id=at.id, score=val))
                            changed += 1
                    elif obj:
                        db.session.delete(obj); changed += 1
        db.session.commit()
        from utils.report_card import compute_term_summaries
        compute_term_summaries(term_id, selected.class_id)
        msg = f'Saved — {changed} change(s).'
        if rejected:
            msg += f' {rejected} skipped (above the subject maximum).'
        return _ok(msg, url_for('subjects.bulk_entry', term_id=term_id, assignment_id=assignment_id))

    scores = {}
    if selected and class_subjects and enrollments:
        sids = [e.student_id for e in enrollments]
        cs_ids = [cs.id for cs in class_subjects]
        for s in StudentScore.query.filter(StudentScore.student_id.in_(sids),
                                           StudentScore.class_subject_id.in_(cs_ids)).all():
            scores[(s.student_id, s.class_subject_id, s.assessment_type_id)] = s.score

    students = [e.student for e in enrollments]
    return _render({
        'page': 'bulk_entry', 'nav': _nav_urls(),
        'term_id': term_id or '', 'assignment_id': assignment_id or '',
        'terms': [{'id': t.id, 'full_name': t.full_name} for t in terms],
        'assignments': [{'id': a.id, 'display_name': a.display_name} for a in assignments],
        'has_grid': bool(selected and class_subjects and enrollments),
        'class_subjects': [{'id': cs.id, 'subject_name': cs.subject.name} for cs in class_subjects],
        'assessment_types': [{'id': at.id, 'name': at.name, 'short_name': at.short_name or at.name,
                              'max_score': at.max_score} for at in assessment_types],
        'students': [{'id': s.id, 'full_name': s.full_name} for s in students],
        'scores': {f'{sid}_{csid}_{atid}': v for (sid, csid, atid), v in scores.items()},
        'self_url': url_for('subjects.bulk_entry'), 'submit_url': url_for('subjects.bulk_entry'),
        'broadsheet_url': url_for('subjects.broadsheet', term_id=term_id or '', assignment_id=assignment_id or ''),
    })


@subjects_bp.route('/broadsheet')
@login_required
def broadsheet():
    """View broadsheet (all scores for a class)"""
    term_id = request.args.get('term_id', type=int)
    assignment_id = request.args.get('assignment_id', type=int)
    
    # Check class access
    if assignment_id and not can_access_class(assignment_id):
        flash('You do not have access to this class.', 'error')
        return redirect(url_for('subjects.broadsheet'))
    
    terms = Term.query.order_by(Term.id.desc()).all()
    
    if not term_id:
        active_term = get_active_term()
        if active_term:
            term_id = active_term.id
    
    selected_term = db.session.get(Term, term_id) if term_id else None
    
    # Filter assignments for teachers
    assignments = []
    if term_id:
        all_assignments = ClassArmAssignment.query.filter_by(term_id=term_id).all()
        assignments = filter_classes_for_user(all_assignments)
    
    selected_assignment = db.session.get(ClassArmAssignment, assignment_id) if assignment_id else None
    
    # Build broadsheet data
    broadsheet_data = []
    class_subjects = []
    assessment_types = AssessmentType.query.filter_by(is_active=True).order_by(AssessmentType.order).all()
    
    if selected_assignment:
        # Get subjects for this class
        class_subjects = ClassSubject.query.filter_by(
            term_id=term_id,
            class_id=selected_assignment.class_id,
            is_active=True
        ).filter(
            (ClassSubject.arm_id == None) | (ClassSubject.arm_id == selected_assignment.arm_id)
        ).join(Subject).order_by(Subject.name).all()
        
        # Get enrolled students
        enrollments = StudentEnrollment.query.filter_by(
            class_arm_assignment_id=assignment_id,
            is_active=True
        ).join(Student).order_by(Student.surname, Student.first_name).all()
        
        pass_mark = SchoolSettings.get('pass_mark', 50)
        
        for enrollment in enrollments:
            student_row = {
                'student': enrollment.student,
                'subjects': {},
                'total': 0,
                'average': 0,
                'subjects_passed': 0,
                'subjects_failed': 0
            }
            
            for cs in class_subjects:
                subject_data = {'assessments': {}, 'total': 0}
                
                # Get all scores for this student-subject
                scores = StudentScore.query.filter_by(
                    student_id=enrollment.student_id,
                    class_subject_id=cs.id
                ).all()
                
                scores_dict = {s.assessment_type_id: s.score for s in scores}
                
                subject_total = 0
                for at in assessment_types:
                    score = scores_dict.get(at.id)
                    subject_data['assessments'][at.id] = score
                    if score:
                        subject_total += score
                
                subject_data['total'] = subject_total
                subject_data['grade'] = GradeScale.get_grade(subject_total) if subject_total else '-'
                
                if subject_total >= pass_mark:
                    student_row['subjects_passed'] += 1
                elif subject_total > 0:
                    student_row['subjects_failed'] += 1
                
                student_row['subjects'][cs.id] = subject_data
                student_row['total'] += subject_total
            
            if class_subjects:
                student_row['average'] = round(student_row['total'] / len(class_subjects), 2)
            
            broadsheet_data.append(student_row)
        
        # Sort by average (descending) for ranking
        broadsheet_data.sort(key=lambda x: x['average'], reverse=True)
        
        # Add positions
        for i, row in enumerate(broadsheet_data):
            row['position'] = i + 1
    
    return _render({
        'page': 'broadsheet', 'nav': _nav_urls(),
        'term_id': term_id or '', 'assignment_id': assignment_id or '',
        'selected_assignment': selected_assignment.display_name if selected_assignment else '',
        'has_selection': bool(selected_assignment),
        'terms': [{'id': t.id, 'full_name': t.full_name} for t in terms],
        'assignments': [{'id': a.id, 'display_name': a.display_name} for a in assignments],
        'class_subjects': [{'id': cs.id, 'short': cs.subject.short_name or cs.subject.name[:3],
                            'name': cs.subject.name} for cs in class_subjects],
        'rows': [{'position': r['position'], 'student': r['student'].full_name,
                  'subjects': {str(cs.id): (round(r['subjects'].get(cs.id, {}).get('total', 0), 1)
                                            if r['subjects'].get(cs.id, {}).get('total') else None)
                               for cs in class_subjects},
                  'total': round(r['total'], 1), 'average': r['average'],
                  'passed': r['subjects_passed'], 'failed': r['subjects_failed']}
                 for r in broadsheet_data],
        'self_url': url_for('subjects.broadsheet'),
        'urls': {'compute': url_for('subjects.compute_summaries'),
                 'bulk_entry': url_for('subjects.bulk_entry', term_id=term_id or '', assignment_id=assignment_id or ''),
                 'affective': url_for('subjects.affective', term_id=term_id or '', assignment_id=assignment_id or ''),
                 'comments': url_for('subjects.comments', term_id=term_id or '', assignment_id=assignment_id or ''),
                 'export': url_for('subjects.export_broadsheet', term_id=term_id or '', assignment_id=assignment_id or ''),
                 'scores': url_for('subjects.scores_entry', term_id=term_id or '', assignment_id=assignment_id or '')},
    })


@subjects_bp.route('/broadsheet/compute', methods=['POST'])
@login_required
def compute_summaries():
    """Compute & persist term results + class/arm positions for a class."""
    term_id = request.form.get('term_id', type=int)
    assignment_id = request.form.get('assignment_id', type=int)
    asg = db.session.get(ClassArmAssignment, assignment_id) if assignment_id else None
    if not (term_id and asg):
        return _err('Select a term and class first.', url_for('subjects.broadsheet'))
    from utils.report_card import compute_term_summaries
    from utils.audit import log_action
    count = compute_term_summaries(term_id, asg.class_id)
    log_action('results.compute_summaries',
               detail=f'term {term_id}, class {asg.class_id}: {count} student(s)')
    return _ok(f'Computed results and positions for {count} student(s).',
               url_for('subjects.broadsheet', term_id=term_id, assignment_id=assignment_id))


@subjects_bp.route('/affective', methods=['GET', 'POST'])
@login_required
def affective():
    """Enter behavioural / affective ratings (1–5) for a class arm in a term."""
    from utils.report_card import active_traits
    AFFECTIVE_TRAITS = active_traits()
    AFFECTIVE_KEYS = [k for k, _ in AFFECTIVE_TRAITS]
    term_id = request.values.get('term_id', type=int)
    assignment_id = request.values.get('assignment_id', type=int)
    if assignment_id and not can_access_class(assignment_id):
        flash('You do not have access to this class.', 'error')
        return redirect(url_for('subjects.affective'))
    if not term_id:
        active = get_active_term()
        term_id = active.id if active else None
    terms = Term.query.order_by(Term.id.desc()).all()
    assignments = (filter_classes_for_user(
        ClassArmAssignment.query.filter_by(term_id=term_id).all()) if term_id else [])
    selected_assignment = db.session.get(ClassArmAssignment, assignment_id) if assignment_id else None

    if request.method == 'POST' and selected_assignment and term_id:
        enrollments = StudentEnrollment.query.filter_by(
            class_arm_assignment_id=assignment_id, is_active=True).all()
        for e in enrollments:
            mapping = {k: request.form.get(f'r_{e.student_id}_{k}', type=int)
                       for k in AFFECTIVE_KEYS}
            ts = TermSummary.query.filter_by(student_id=e.student_id, term_id=term_id).first()
            if not ts:
                ts = TermSummary(student_id=e.student_id, term_id=term_id, enrollment_id=e.id)
                db.session.add(ts)
            ts.set_affective(mapping)
        db.session.commit()
        from utils.audit import log_action
        log_action('results.affective',
                   detail=f'term {term_id}, {selected_assignment.display_name}')
        return _ok('Behavioural ratings saved.',
                   url_for('subjects.affective', term_id=term_id, assignment_id=assignment_id))

    students = []
    if selected_assignment:
        enrollments = (StudentEnrollment.query.filter_by(
            class_arm_assignment_id=assignment_id, is_active=True)
            .join(Student).order_by(Student.surname, Student.first_name).all())
        ratings = {ts.student_id: ts.affective_map
                   for ts in TermSummary.query.filter_by(term_id=term_id).all()}
        for e in enrollments:
            students.append({'student': e.student, 'ratings': ratings.get(e.student_id, {})})

    return _render({
        'page': 'affective', 'nav': _nav_urls(),
        'term_id': term_id or '', 'assignment_id': assignment_id or '',
        'has_students': bool(selected_assignment and students),
        'selected': bool(selected_assignment),
        'terms': [{'id': t.id, 'full_name': t.full_name} for t in terms],
        'assignments': [{'id': a.id, 'display_name': a.display_name} for a in assignments],
        'traits': [{'key': k, 'label': lbl} for k, lbl in AFFECTIVE_TRAITS],
        'students': [{'id': r['student'].id, 'full_name': r['student'].full_name,
                      'ratings': r['ratings']} for r in students],
        'self_url': url_for('subjects.affective'), 'submit_url': url_for('subjects.affective'),
        'broadsheet_url': url_for('subjects.broadsheet', term_id=term_id or '', assignment_id=assignment_id or ''),
    })


@subjects_bp.route('/comments', methods=['GET', 'POST'])
@login_required
def comments():
    """Enter form-teacher & principal comments for a class arm in a term."""
    term_id = request.values.get('term_id', type=int)
    assignment_id = request.values.get('assignment_id', type=int)
    if assignment_id and not can_access_class(assignment_id):
        flash('You do not have access to this class.', 'error')
        return redirect(url_for('subjects.comments'))
    if not term_id:
        active = get_active_term()
        term_id = active.id if active else None
    terms = Term.query.order_by(Term.id.desc()).all()
    assignments = (filter_classes_for_user(
        ClassArmAssignment.query.filter_by(term_id=term_id).all()) if term_id else [])
    selected_assignment = db.session.get(ClassArmAssignment, assignment_id) if assignment_id else None

    if request.method == 'POST' and selected_assignment and term_id:
        enrollments = StudentEnrollment.query.filter_by(
            class_arm_assignment_id=assignment_id, is_active=True).all()
        for e in enrollments:
            ts = TermSummary.query.filter_by(student_id=e.student_id, term_id=term_id).first()
            if not ts:
                ts = TermSummary(student_id=e.student_id, term_id=term_id, enrollment_id=e.id)
                db.session.add(ts)
            ts.teacher_comment = (request.form.get(f't_{e.student_id}') or '').strip() or None
            ts.principal_comment = (request.form.get(f'p_{e.student_id}') or '').strip() or None
        db.session.commit()
        from utils.audit import log_action
        log_action('results.comments', detail=f'term {term_id}, {selected_assignment.display_name}')
        return _ok('Comments saved.',
                   url_for('subjects.comments', term_id=term_id, assignment_id=assignment_id))

    students = []
    if selected_assignment:
        enrollments = (StudentEnrollment.query.filter_by(
            class_arm_assignment_id=assignment_id, is_active=True)
            .join(Student).order_by(Student.surname, Student.first_name).all())
        summ = {ts.student_id: ts for ts in TermSummary.query.filter_by(term_id=term_id).all()}
        for e in enrollments:
            ts = summ.get(e.student_id)
            students.append({'student': e.student,
                             'teacher_comment': ts.teacher_comment if ts else '',
                             'principal_comment': ts.principal_comment if ts else ''})

    return _render({
        'page': 'comments', 'nav': _nav_urls(),
        'term_id': term_id or '', 'assignment_id': assignment_id or '',
        'has_students': bool(selected_assignment and students),
        'selected': bool(selected_assignment),
        'terms': [{'id': t.id, 'full_name': t.full_name} for t in terms],
        'assignments': [{'id': a.id, 'display_name': a.display_name} for a in assignments],
        'students': [{'id': r['student'].id, 'full_name': r['student'].full_name,
                      'teacher_comment': r['teacher_comment'] or '',
                      'principal_comment': r['principal_comment'] or ''} for r in students],
        'self_url': url_for('subjects.comments'), 'submit_url': url_for('subjects.comments'),
        'broadsheet_url': url_for('subjects.broadsheet', term_id=term_id or '', assignment_id=assignment_id or ''),
    })


# ============================================================================
# STUDENT REPORT CARD
# ============================================================================

@subjects_bp.route('/report-card/<int:student_id>')
@login_required
@result_card_required
def student_report_card(student_id):
    """View student report card"""
    student = db.get_or_404(Student, student_id)
    from utils.access_control import assert_student_access
    assert_student_access(student)   # branch + form-teacher scope
    term_id = request.args.get('term_id', type=int)

    terms = Term.query.order_by(Term.id.desc()).all()
    
    if not term_id:
        active_term = get_active_term()
        if active_term:
            term_id = active_term.id
    
    selected_term = db.session.get(Term, term_id) if term_id else None

    report_data = None
    enrollment = None

    from utils.report_card import active_traits, RATING_LABELS
    if selected_term:
        from utils.report_card import build_report_card
        enrollment, report_data = build_report_card(student_id, term_id)

    return render_template('subjects/report_card.html',
        student=student, terms=terms, term_id=term_id, selected_term=selected_term,
        report_data=report_data, enrollment=enrollment,
        affective_traits=active_traits(), rating_labels=RATING_LABELS
    )


@subjects_bp.route('/report-card/<int:student_id>/pdf')
@login_required
@result_card_required
def report_card_pdf(student_id):
    """Download the student's term report card as a PDF."""
    from flask import send_file
    from utils.report_card import build_report_card, active_traits, RATING_LABELS
    from utils.report_pdf import report_card_pdf as build_pdf
    from utils.access_control import assert_student_access
    student = db.get_or_404(Student, student_id)
    assert_student_access(student)   # branch + form-teacher scope
    term_id = request.args.get('term_id', type=int) or (
        get_active_term().id if get_active_term() else None)
    _, report_data = build_report_card(student_id, term_id) if term_id else (None, None)
    if not report_data:
        flash('No results to export for this term.', 'error')
        return redirect(url_for('subjects.student_report_card', student_id=student_id, term_id=term_id))
    term = db.session.get(Term, term_id)
    buf = build_pdf(student, report_data, term,
                    SchoolSettings.get('school_name', 'School'),
                    active_traits(), RATING_LABELS)
    name = f"{student.student_id}_{term.name.replace(' ', '_')}_report.pdf"
    return send_file(buf, mimetype='application/pdf', as_attachment=True, download_name=name)


# ============================================================================
# API ENDPOINTS
# ============================================================================

@subjects_bp.route('/api/class-subjects/<int:term_id>/<int:class_id>')
@login_required
def api_class_subjects(term_id, class_id):
    """Get subjects for a class in a term"""
    class_subjects = ClassSubject.query.filter_by(
        term_id=term_id,
        class_id=class_id,
        is_active=True
    ).join(Subject).order_by(Subject.name).all()
    
    return jsonify([{
        'id': cs.id,
        'subject_id': cs.subject_id,
        'subject_name': cs.subject.name,
        'teacher_name': cs.teacher_name
    } for cs in class_subjects])


@subjects_bp.route('/api/student-scores/<int:student_id>/<int:term_id>')
@login_required
def api_student_scores(student_id, term_id):
    """Get all scores for a student in a term"""
    require_branch_access(db.get_or_404(Student, student_id).branch_id)
    scores = StudentScore.query.join(ClassSubject).filter(
        StudentScore.student_id == student_id,
        ClassSubject.term_id == term_id
    ).all()
    
    return jsonify([{
        'subject': s.class_subject.subject.name,
        'assessment': s.assessment_type.name,
        'score': s.score
    } for s in scores])


# ============================================================================
# EXPORT BROADSHEET TO EXCEL
# ============================================================================

@subjects_bp.route('/broadsheet/export')
@login_required
def export_broadsheet():
    """Export broadsheet to Excel"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from flask import Response
    import io
    
    term_id = request.args.get('term_id', type=int)
    assignment_id = request.args.get('assignment_id', type=int)
    
    if not term_id or not assignment_id:
        flash('Select term and class first.', 'error')
        return redirect(url_for('subjects.broadsheet'))
    
    selected_term = db.session.get(Term, term_id)
    selected_assignment = db.session.get(ClassArmAssignment, assignment_id)
    
    if not selected_term or not selected_assignment:
        flash('Invalid selection.', 'error')
        return redirect(url_for('subjects.broadsheet'))
    if not can_access_class(assignment_id):
        flash('You do not have access to that class.', 'error')
        return redirect(url_for('subjects.broadsheet'))

    # Get data (same as broadsheet view)
    class_subjects = ClassSubject.query.filter_by(
        term_id=term_id,
        class_id=selected_assignment.class_id,
        is_active=True
    ).filter(
        (ClassSubject.arm_id == None) | (ClassSubject.arm_id == selected_assignment.arm_id)
    ).join(Subject).order_by(Subject.name).all()
    
    assessment_types = AssessmentType.query.filter_by(is_active=True).order_by(AssessmentType.order).all()
    
    enrollments = StudentEnrollment.query.filter_by(
        class_arm_assignment_id=assignment_id,
        is_active=True
    ).join(Student).order_by(Student.surname, Student.first_name).all()
    
    pass_mark = SchoolSettings.get('pass_mark', 50)
    
    # Build data
    broadsheet_data = []
    for enrollment in enrollments:
        student_row = {
            'student': enrollment.student,
            'subjects': {},
            'total': 0,
            'average': 0
        }
        
        for cs in class_subjects:
            scores = StudentScore.query.filter_by(
                student_id=enrollment.student_id,
                class_subject_id=cs.id
            ).all()
            
            subject_total = sum(s.score for s in scores)
            student_row['subjects'][cs.id] = subject_total
            student_row['total'] += subject_total
        
        if class_subjects:
            student_row['average'] = round(student_row['total'] / len(class_subjects), 2)
        
        broadsheet_data.append(student_row)
    
    # Sort by average
    broadsheet_data.sort(key=lambda x: x['average'], reverse=True)
    
    # Create Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "Broadsheet"
    
    # Styles
    header_font = Font(bold=True, size=12)
    header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
    header_font_white = Font(bold=True, color='FFFFFF')
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    
    # Title
    ws['A1'] = f"{selected_assignment.display_name} - Broadsheet"
    ws['A1'].font = Font(bold=True, size=14)
    ws['A2'] = f"{selected_term.full_name}"
    
    # Headers
    row = 4
    headers = ['Pos', 'Student Name', 'Student ID']
    for cs in class_subjects:
        headers.append(cs.subject.short_name or cs.subject.name[:5])
    headers.extend(['Total', 'Average', 'Grade'])
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.font = header_font_white
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center')
    
    # Data rows
    for idx, data in enumerate(broadsheet_data, 1):
        row += 1
        ws.cell(row=row, column=1, value=idx).border = thin_border
        ws.cell(row=row, column=2, value=data['student'].full_name).border = thin_border
        ws.cell(row=row, column=3, value=data['student'].student_id).border = thin_border
        
        col = 4
        for cs in class_subjects:
            score = data['subjects'].get(cs.id, 0)
            cell = ws.cell(row=row, column=col, value=score if score else '')
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='center')
            col += 1
        
        ws.cell(row=row, column=col, value=data['total']).border = thin_border
        ws.cell(row=row, column=col+1, value=data['average']).border = thin_border
        ws.cell(row=row, column=col+2, value=GradeScale.get_grade(data['average']) if data['average'] else '').border = thin_border
    
    # Auto-width columns
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except Exception:
                pass
        ws.column_dimensions[column].width = min(max_length + 2, 30)
    
    filename = f"broadsheet_{selected_assignment.school_class.name}_{selected_assignment.arm.name}_{selected_term.name}.xlsx"

    return xlsx_response(wb, filename)


# ============================================================================
# BULK SCORE IMPORT FROM EXCEL
# ============================================================================

@subjects_bp.route('/scores/import', methods=['GET', 'POST'])
@login_required
def import_scores():
    """Import scores from Excel"""
    if request.method == 'POST':
        try:
            from openpyxl import load_workbook
            
            term_id = request.form.get('term_id', type=int)
            assignment_id = request.form.get('assignment_id', type=int)
            class_subject_id = request.form.get('class_subject_id', type=int)

            # Teachers may only import scores for a subject they teach here.
            cs = db.session.get(ClassSubject, class_subject_id) if class_subject_id else None
            if not can_enter_results(assignment_id, cs.subject_id if cs else None):
                flash('You can only enter scores for the subjects you teach in this class.', 'error')
                return redirect(url_for('subjects.import_scores'))

            if 'file' not in request.files:
                flash('No file selected.', 'error')
                return redirect(url_for('subjects.import_scores'))
            
            file = request.files['file']
            if not file.filename.endswith(('.xlsx', '.xls')):
                flash('Please upload an Excel file.', 'error')
                return redirect(url_for('subjects.import_scores'))
            
            wb = load_workbook(file)
            ws = wb.active
            
            # Get assessment types
            assessment_types = AssessmentType.query.filter_by(is_active=True).order_by(AssessmentType.order).all()
            
            # Expected columns: Student ID, then assessment type names
            imported = 0
            errors = []
            
            for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), 2):
                if not row[0]:
                    continue
                
                student_id_str = str(row[0]).strip()
                student = Student.query.filter_by(student_id=student_id_str).first()
                
                if not student:
                    errors.append(f"Row {row_num}: Student {student_id_str} not found")
                    continue
                
                # Import each assessment score
                for col_idx, at in enumerate(assessment_types, 1):
                    if col_idx < len(row) and row[col_idx] is not None:
                        try:
                            score_value = float(row[col_idx])
                            
                            # Get or create score
                            existing = StudentScore.query.filter_by(
                                student_id=student.id,
                                class_subject_id=class_subject_id,
                                assessment_type_id=at.id
                            ).first()
                            
                            if existing:
                                existing.score = score_value
                            else:
                                score = StudentScore(
                                    student_id=student.id,
                                    class_subject_id=class_subject_id,
                                    assessment_type_id=at.id,
                                    score=score_value
                                )
                                db.session.add(score)
                            imported += 1
                        except ValueError:
                            pass
            
            db.session.commit()
            flash(f'Imported {imported} scores!', 'success')
            
            if errors:
                for err in errors[:5]:
                    flash(err, 'warning')
                    
        except Exception as e:
            db.session.rollback()
            flash(f'Error: {str(e)}', 'error')
        
        return redirect(url_for('subjects.scores_entry', term_id=term_id, assignment_id=assignment_id, class_subject_id=class_subject_id))
    
    # GET - show form
    term_id = request.args.get('term_id', type=int)
    assignment_id = request.args.get('assignment_id', type=int)
    class_subject_id = request.args.get('class_subject_id', type=int)
    
    terms = Term.query.order_by(Term.id.desc()).all()
    assignments = filter_classes_for_user(
        ClassArmAssignment.query.filter_by(term_id=term_id).all()) if term_id else []
    class_subjects = []
    
    if assignment_id:
        assignment = db.session.get(ClassArmAssignment, assignment_id)
        if assignment:
            class_subjects = ClassSubject.query.filter_by(
                term_id=term_id,
                class_id=assignment.class_id,
                is_active=True
            ).join(Subject).order_by(Subject.name).all()
    
    return render_template('subjects/import_scores.html',
        terms=terms, term_id=term_id,
        assignments=assignments, assignment_id=assignment_id,
        class_subjects=class_subjects, class_subject_id=class_subject_id
    )


@subjects_bp.route('/scores/import/template')
@login_required
def score_import_template():
    """Download score import template"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    from flask import Response
    import io
    
    assignment_id = request.args.get('assignment_id', type=int)
    class_subject_id = request.args.get('class_subject_id', type=int)
    
    if not assignment_id:
        flash('Select a class first.', 'error')
        return redirect(url_for('subjects.import_scores'))
    
    assignment = db.session.get(ClassArmAssignment, assignment_id)
    assessment_types = AssessmentType.query.filter_by(is_active=True).order_by(AssessmentType.order).all()
    
    # Get students
    enrollments = StudentEnrollment.query.filter_by(
        class_arm_assignment_id=assignment_id,
        is_active=True
    ).join(Student).order_by(Student.surname, Student.first_name).all()
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Score Import"
    
    # Headers
    headers = ['Student ID', 'Student Name'] + [at.short_name or at.name for at in assessment_types]
    header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF')
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
    
    # Student rows
    for row_num, enrollment in enumerate(enrollments, 2):
        ws.cell(row=row_num, column=1, value=enrollment.student.student_id)
        ws.cell(row=row_num, column=2, value=enrollment.student.full_name)
    
    return xlsx_response(wb, 'score_import_template.xlsx')


# ============================================================================
# SCORE-SHEET (BROADSHEET) IMAGE IMPORT — Tesseract OCR
# ============================================================================

def _sheet_columns(class_subject):
    """Assessment columns for a subject, ordered as they appear on a printed
    broadsheet. Returns [(assessment_type, max_score), ...]."""
    from utils.assessments import subject_columns
    from utils.waec_ocr import SHEET_COLUMN_ORDER
    cols = subject_columns(class_subject.subject)  # [(at, max)] in storage order
    present = {at.short_name: (at, mx) for at, mx in cols}
    ordered = [present[sn] for sn in SHEET_COLUMN_ORDER if sn in present]
    # Append any columns not covered by the canonical order (defensive).
    seen = {at.id for at, _ in ordered}
    ordered += [(at, mx) for at, mx in cols if at.id not in seen]
    return ordered


def _scan_selector_context():
    """Shared term/class/subject selector context for the scan pages."""
    term_id = request.values.get('term_id', type=int)
    assignment_id = request.values.get('assignment_id', type=int)
    class_subject_id = request.values.get('class_subject_id', type=int)

    if not term_id:
        active_term = get_active_term()
        if active_term:
            term_id = active_term.id

    terms = Term.query.order_by(Term.id.desc()).all()
    assignments = []
    if term_id:
        all_assignments = ClassArmAssignment.query.filter_by(term_id=term_id).all()
        assignments = filter_classes_for_user(all_assignments)

    class_subjects = []
    assignment = db.session.get(ClassArmAssignment, assignment_id) if assignment_id else None
    if assignment:
        class_subjects = ClassSubject.query.filter_by(
            term_id=term_id, class_id=assignment.class_id, is_active=True
        ).filter(
            (ClassSubject.arm_id == None) | (ClassSubject.arm_id == assignment.arm_id)
        ).join(Subject).order_by(Subject.name).all()

    return {
        'terms': terms, 'term_id': term_id,
        'assignments': assignments, 'assignment_id': assignment_id,
        'class_subjects': class_subjects, 'class_subject_id': class_subject_id,
    }


@subjects_bp.route('/scores/scan', methods=['GET', 'POST'])
@login_required
def scoresheet_scan():
    """Upload a photographed score sheet and OCR it into an editable grid."""
    if not can_enter_results() and not is_admin():
        flash('You do not have permission to enter scores.', 'error')
        return redirect(url_for('main.dashboard'))

    from utils.waec_ocr import (
        tesseract_available, extract_text, extract_text_from_pdf,
        parse_score_sheet, match_student,
    )

    ctx = _scan_selector_context()

    if request.method == 'POST':
        term_id = ctx['term_id']
        assignment_id = ctx['assignment_id']
        class_subject_id = ctx['class_subject_id']

        assignment = db.session.get(ClassArmAssignment, assignment_id) if assignment_id else None
        class_subject = db.session.get(ClassSubject, class_subject_id) if class_subject_id else None

        if not (assignment and class_subject):
            flash('Select a class and subject before uploading.', 'error')
            return render_template('subjects/scoresheet_scan.html', **ctx)

        if not can_access_class(assignment_id):
            flash('You do not have access to this class.', 'error')
            return redirect(url_for('subjects.scoresheet_scan'))

        if not tesseract_available():
            flash('OCR engine (Tesseract) is not available on the server.', 'error')
            return render_template('subjects/scoresheet_scan.html', **ctx)

        upload = request.files.get('file')
        if not upload or not upload.filename:
            flash('No file selected.', 'error')
            return render_template('subjects/scoresheet_scan.html', **ctx)

        data = upload.read()
        try:
            if upload.filename.lower().endswith('.pdf'):
                text = extract_text_from_pdf(data)
            else:
                text = extract_text(data)
        except Exception as e:
            flash(f'Could not read the image: {e}', 'error')
            return render_template('subjects/scoresheet_scan.html', **ctx)

        sheet_cols = _sheet_columns(class_subject)
        parsed = parse_score_sheet(text, num_columns=len(sheet_cols))

        if not parsed:
            flash('No student rows could be detected. Try a clearer, straight photo.', 'warning')
            return render_template('subjects/scoresheet_scan.html', **ctx)

        # Students enrolled in this class (for matching + the picker).
        enrollments = StudentEnrollment.query.filter_by(
            class_arm_assignment_id=assignment_id, is_active=True
        ).join(Student).order_by(Student.surname, Student.first_name).all()
        students = [e.student for e in enrollments]
        by_student_id = {s.student_id: s for s in students}

        rows = []
        for p in parsed:
            matched = None
            # 1) exact match on a scanned student number
            if p['student_num'] and p['student_num'] in by_student_id:
                matched = by_student_id[p['student_num']]
            # 2) fuzzy match on the name
            if not matched:
                matched, _ = match_student(p['name'], students)
            # Map the positional cells to assessment-type ids.
            cell_map = {}
            for (at, _mx), value in zip(sheet_cols, p['cells']):
                cell_map[at.id] = value
            rows.append({
                'student_num': p['student_num'],
                'name': p['name'],
                'matched_id': matched.id if matched else None,
                'cells': cell_map,
            })

        return render_template('subjects/scoresheet_review.html',
            term_id=term_id, assignment_id=assignment_id, class_subject_id=class_subject_id,
            assignment=assignment, class_subject=class_subject,
            columns=sheet_cols, rows=rows, students=students,
        )

    return render_template('subjects/scoresheet_scan.html', **ctx)


# Tokens that mean "no score" in a pasted cell.
_PASTE_BLANKS = {'', '-', '--', '–', 'nil', 'absent', 'a', 'x', 'na', 'n/a'}


def _parse_pasted_scores(text, num_columns):
    """Parse pasted comma/tab-separated rows into ``[{'identifier', 'cells'}]``.

    One student per line: the first field is the identifier (admission number or
    name), the remaining fields are scores in the sheet's column order. Blank /
    dash / 'absent' cells become empty. A leading header row is skipped."""
    import re as _re
    rows = []
    for raw in text.splitlines():
        line = raw.strip().strip(',').strip()
        if not line:
            continue
        parts = [p.strip() for p in _re.split(r'[,\t]', line)]
        ident = parts[0].strip() if parts else ''
        if not ident:
            continue
        # Skip an obvious header line — the first field is a header label, not a
        # real student (column headers like CA1/CA2 legitimately contain digits,
        # so we key off the identifier word, not the trailing cells).
        key = _re.sub(r'[^a-z]', '', ident.lower())
        if key in {'name', 'names', 'fullname', 'student', 'students', 'studentname',
                   'adm', 'admno', 'admission', 'admissionnumber', 'admissionno',
                   'sn', 'sno', 'no', 'reg', 'regno', 'serial'}:
            continue
        cells = []
        for c in parts[1:1 + num_columns]:
            c = c.strip()
            if c.lower() in _PASTE_BLANKS:
                cells.append('')
            else:
                m = _re.search(r'\d+(?:\.\d+)?', c)
                cells.append(m.group(0) if m else '')
        rows.append({'identifier': ident, 'cells': cells})
    return rows


@subjects_bp.route('/scores/paste', methods=['GET', 'POST'])
@login_required
def scoresheet_paste():
    """Paste comma-separated scores (e.g. produced by asking an external AI to read
    a photographed sheet) into an editable, student-matched grid — no OCR, no API
    keys. Reuses the same review/confirm grid and save path as the scanner."""
    if not can_enter_results() and not is_admin():
        flash('You do not have permission to enter scores.', 'error')
        return redirect(url_for('main.dashboard'))

    from utils.waec_ocr import match_student

    ctx = _scan_selector_context()
    cs = db.session.get(ClassSubject, ctx['class_subject_id']) if ctx['class_subject_id'] else None
    sheet_cols = _sheet_columns(cs) if cs else []
    ctx['columns'] = sheet_cols

    if request.method == 'POST':
        assignment_id = ctx['assignment_id']
        assignment = db.session.get(ClassArmAssignment, assignment_id) if assignment_id else None
        if not (assignment and cs):
            flash('Select a class and subject before pasting.', 'error')
            return render_template('subjects/scoresheet_paste.html', **ctx)
        if not can_access_class(assignment_id):
            flash('You do not have access to this class.', 'error')
            return redirect(url_for('subjects.scoresheet_paste'))

        pasted = (request.form.get('data') or '').strip()
        if not pasted:
            flash('Paste the comma-separated rows first.', 'error')
            return render_template('subjects/scoresheet_paste.html', pasted=pasted, **ctx)

        parsed = _parse_pasted_scores(pasted, len(sheet_cols))
        if not parsed:
            flash('No rows could be read from the pasted text. Check the format.', 'warning')
            return render_template('subjects/scoresheet_paste.html', pasted=pasted, **ctx)

        enrollments = StudentEnrollment.query.filter_by(
            class_arm_assignment_id=assignment_id, is_active=True
        ).join(Student).order_by(Student.surname, Student.first_name).all()
        students = [e.student for e in enrollments]
        by_student_id = {s.student_id: s for s in students}

        rows = []
        for p in parsed:
            matched = by_student_id.get(p['identifier'])
            student_num = p['identifier'] if matched else ''
            if not matched:                                   # fall back to fuzzy name match
                matched, _ = match_student(p['identifier'], students)
            cell_map = {}
            for (at, _mx), value in zip(sheet_cols, p['cells']):
                cell_map[at.id] = value
            rows.append({
                'student_num': student_num,
                'name': p['identifier'],
                'matched_id': matched.id if matched else None,
                'cells': cell_map,
            })

        return render_template('subjects/scoresheet_review.html',
            term_id=ctx['term_id'], assignment_id=assignment_id, class_subject_id=ctx['class_subject_id'],
            assignment=assignment, class_subject=cs,
            columns=sheet_cols, rows=rows, students=students,
        )

    return render_template('subjects/scoresheet_paste.html', **ctx)


@subjects_bp.route('/scores/scan/save', methods=['POST'])
@login_required
def scoresheet_save():
    """Persist the reviewed score-sheet grid as StudentScores."""
    import re as _re

    if not can_enter_results() and not is_admin():
        flash('You do not have permission to enter scores.', 'error')
        return redirect(url_for('main.dashboard'))

    term_id = request.form.get('term_id', type=int)
    assignment_id = request.form.get('assignment_id', type=int)
    class_subject_id = request.form.get('class_subject_id', type=int)
    row_count = request.form.get('row_count', type=int) or 0

    assignment = db.session.get(ClassArmAssignment, assignment_id) if assignment_id else None
    class_subject = db.session.get(ClassSubject, class_subject_id) if class_subject_id else None

    if not (assignment and class_subject):
        flash('Missing class/subject context.', 'error')
        return redirect(url_for('subjects.scoresheet_scan'))

    # Teachers may only save scores for a subject they teach in this class.
    if not can_enter_results(assignment_id, class_subject.subject_id):
        flash('You can only enter scores for the subjects you teach in this class.', 'error')
        return redirect(url_for('subjects.scoresheet_scan'))

    sheet_cols = _sheet_columns(class_subject)
    auto_id_re = _re.compile(r'^STU\d{3,}$')

    saved_rows = 0
    saved_scores = 0
    adopted = 0
    warnings = []

    try:
        for r in range(row_count):
            student_pk = request.form.get(f'student_{r}', type=int)
            if not student_pk:
                continue  # row skipped by the user
            student = db.session.get(Student, student_pk)
            if not student:
                continue

            # Adopt the scanned student number only when the student currently
            # has an auto-generated STU##### id (never overwrite a manual id).
            scanned = (request.form.get(f'studentnum_{r}') or '').strip()
            if scanned and scanned != student.student_id and auto_id_re.match(student.student_id or ''):
                clash = Student.query.filter(
                    Student.student_id == scanned, Student.id != student.id).first()
                if clash:
                    warnings.append(f"{student.full_name}: number {scanned} already used by {clash.full_name}")
                else:
                    student.student_id = scanned
                    adopted += 1

            row_has_score = False
            for at, max_score in sheet_cols:
                raw = (request.form.get(f'cell_{r}_{at.id}') or '').strip()
                if raw == '':
                    continue
                try:
                    value = float(raw)
                except ValueError:
                    continue
                if value < 0:
                    value = 0
                if max_score and value > max_score:
                    value = max_score

                existing = StudentScore.query.filter_by(
                    student_id=student.id,
                    class_subject_id=class_subject_id,
                    assessment_type_id=at.id,
                ).first()
                if existing:
                    existing.score = value
                else:
                    db.session.add(StudentScore(
                        student_id=student.id,
                        class_subject_id=class_subject_id,
                        assessment_type_id=at.id,
                        score=value,
                    ))
                saved_scores += 1
                row_has_score = True

            if row_has_score:
                saved_rows += 1

        db.session.commit()
        flash(f'Saved {saved_scores} scores for {saved_rows} students.', 'success')
        if adopted:
            flash(f'Adopted scanned student number for {adopted} student(s).', 'info')
        for w in warnings[:5]:
            flash(w, 'warning')
    except Exception as e:
        db.session.rollback()
        flash(f'Error saving scores: {e}', 'error')

    return redirect(url_for('subjects.scores_entry',
        term_id=term_id, assignment_id=assignment_id, class_subject_id=class_subject_id))


# ============================================================================
# PRINT ALL REPORT CARDS
# ============================================================================

@subjects_bp.route('/report-cards/print-all')
@login_required
@result_card_required
def print_all_report_cards():
    """Print all report cards for a class"""
    term_id = request.args.get('term_id', type=int)
    assignment_id = request.args.get('assignment_id', type=int)
    
    terms = Term.query.order_by(Term.id.desc()).all()
    
    if not term_id:
        active_term = get_active_term()
        if active_term:
            term_id = active_term.id
    
    selected_term = db.session.get(Term, term_id) if term_id else None

    # A user may only print report cards for classes they can access.
    if assignment_id and not can_access_class(assignment_id):
        flash('You do not have access to that class.', 'error')
        return redirect(url_for('subjects.print_all_report_cards', term_id=term_id))

    assignments = []
    if term_id:
        assignments = filter_classes_for_user(
            ClassArmAssignment.query.filter_by(term_id=term_id).all())

    selected_assignment = db.session.get(ClassArmAssignment, assignment_id) if assignment_id else None
    
    all_reports = []
    
    if selected_assignment and selected_term:
        assessment_types = AssessmentType.query.filter_by(is_active=True).order_by(AssessmentType.order).all()
        pass_mark = SchoolSettings.get('pass_mark', 50)
        
        class_subjects = ClassSubject.query.filter_by(
            term_id=term_id,
            class_id=selected_assignment.class_id,
            is_active=True
        ).filter(
            (ClassSubject.arm_id == None) | (ClassSubject.arm_id == selected_assignment.arm_id)
        ).join(Subject).order_by(Subject.name).all()
        
        enrollments = StudentEnrollment.query.filter_by(
            class_arm_assignment_id=assignment_id,
            is_active=True
        ).join(Student).order_by(Student.surname, Student.first_name).all()
        
        for enrollment in enrollments:
            student = enrollment.student
            subjects_data = []
            total_score = 0
            subjects_passed = 0
            subjects_failed = 0
            
            for cs in class_subjects:
                subject_row = {
                    'subject': cs.subject,
                    'assessments': {},
                    'total': 0,
                    'grade': '-',
                    'remark': '-'
                }
                
                scores = StudentScore.query.filter_by(
                    student_id=student.id,
                    class_subject_id=cs.id
                ).all()
                
                scores_dict = {s.assessment_type_id: s.score for s in scores}
                
                subject_total = 0
                for at in assessment_types:
                    score = scores_dict.get(at.id)
                    subject_row['assessments'][at.id] = score
                    if score:
                        subject_total += score
                
                subject_row['total'] = subject_total
                if subject_total > 0:
                    subject_row['grade'] = GradeScale.get_grade(subject_total)
                    subject_row['remark'] = GradeScale.get_remark(subject_total)
                    total_score += subject_total
                    
                    if subject_total >= pass_mark:
                        subjects_passed += 1
                    else:
                        subjects_failed += 1
                
                subjects_data.append(subject_row)
            
            average = round(total_score / len(class_subjects), 2) if class_subjects else 0
            
            all_reports.append({
                'student': student,
                'enrollment': enrollment,
                'subjects': subjects_data,
                'assessment_types': assessment_types,
                'total_score': total_score,
                'average': average,
                'overall_grade': GradeScale.get_grade(average) if average else '-',
                'subjects_passed': subjects_passed,
                'subjects_failed': subjects_failed,
                'total_subjects': len(class_subjects)
            })
    
    school_name = SchoolSettings.get('school_name', 'School Name')
    
    return render_template('subjects/print_all_report_cards.html',
        terms=terms, term_id=term_id, selected_term=selected_term,
        assignments=assignments, assignment_id=assignment_id, selected_assignment=selected_assignment,
        all_reports=all_reports, school_name=school_name
    )
