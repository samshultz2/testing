"""Sales & Inventory routes (Stage 5).

Bursars manage a branch's products/stock and ring up sales. All views are
branch-scoped via utils.branch_scope.
"""

from flask import (Blueprint, render_template, request, redirect, url_for,
                   flash, session, jsonify)
from sqlalchemy import func

from models import (db, Product, Sale, SaleItem, StockMovement, Student,
                    StudentEnrollment, ClassArmAssignment, SchoolClass, ClassArm,
                    Supplier, PurchaseOrder, PurchaseOrderItem, SupplierPayment, PromoCode,
                    StockAudit, StockAuditItem, FixedAsset, StockBatch)
from models.models_sales import (PRODUCT_CATEGORIES, SALE_METHODS, CUSTOMER_TYPES,
                                 UNITS, STOCK_IN_REASONS, STOCK_OUT_REASONS,
                                 PO_STATUSES, PURCHASE_METHODS,
                                 FIXED_ASSET_CATEGORIES, FIXED_ASSET_STATUSES)
from utils.access_control import (login_required, filter_classes_for_user,
                                  can_approve_purchase, can_sign_off_count)
from utils.branch_scope import scope_query, branch_for_new, can_access_branch
from utils import timeutil
from utils.helpers import get_active_term, parse_date
from utils.search import like_term
from utils.web_exports import xlsx_response, csv_response

sales_bp = Blueprint('sales', __name__, url_prefix='/sales')


def _wants_json():
    """The React UI posts via fetch with this header; reply JSON to it while
    keeping the classic flash+redirect for any plain form submit."""
    return request.headers.get('X-Requested-With') == 'fetch' or request.is_json


def _ok(message, redirect_url=None):
    if _wants_json():
        return jsonify({'ok': True, 'message': message, 'redirect': redirect_url})
    flash(message, 'success')
    return redirect(redirect_url or url_for('sales.dashboard'))


def _err(message, redirect_url=None):
    if _wants_json():
        return jsonify({'ok': False, 'error': message}), 400
    flash(message, 'error')
    return redirect(redirect_url or url_for('sales.dashboard'))


def _render(payload):
    """Render the shared sales React shell with an embedded payload."""
    from utils.spa import render_or_json
    return render_or_json('sales/app.html', 'sales_json', payload)


# ---------------------------------------------------------------------------
# Dashboard
# ---------------------------------------------------------------------------

def _sale_row(s, with_when_year=False):
    fmt = '%d %b %Y %H:%M' if with_when_year else '%d %b %H:%M'
    return {
        'id': s.id, 'receipt_no': s.receipt_no, 'buyer': s.buyer,
        'payment_method': s.payment_method, 'total': s.total or 0,
        'item_count': s.items.count(),
        'when': s.created_at.strftime(fmt) if s.created_at else '',
        'receipt_url': url_for('sales.receipt', sale_id=s.id),
    }


@sales_bp.route('/')
@login_required
def dashboard():
    from datetime import timedelta
    from collections import defaultdict
    today = timeutil.today()
    week_start = today - timedelta(days=6)
    month_start = today - timedelta(days=29)

    # One 30-day pass powers today/week/month totals + the breakdowns/trend.
    month_sales = (scope_query(Sale.query, Sale)
                   .filter(func.date(Sale.created_at) >= month_start).all())
    today_total = week_total = month_total = 0.0
    today_count = 0
    by_method = defaultdict(float); by_cashier = defaultdict(float); trend_map = defaultdict(float)
    for s in month_sales:
        amt = s.total or 0
        month_total += amt
        d = s.created_at.date() if s.created_at else today
        if d >= week_start:
            week_total += amt
        if d == today:
            today_total += amt; today_count += 1
        by_method[s.payment_method or 'Cash'] += amt
        by_cashier[s.sold_by or 'Unknown'] += amt
        trend_map[d] += amt

    sale_ids = [s.id for s in month_sales]
    by_category = defaultdict(float); prod = defaultdict(lambda: {'revenue': 0.0, 'units': 0, 'name': ''})
    cogs = 0.0
    if sale_ids:
        for it, p in (db.session.query(SaleItem, Product)
                      .outerjoin(Product, SaleItem.product_id == Product.id)
                      .filter(SaleItem.sale_id.in_(sale_ids)).all()):
            by_category[(p.category if p else None) or 'Other'] += it.line_total or 0
            key = it.product_id or f'd:{it.description}'
            prod[key]['revenue'] += it.line_total or 0
            prod[key]['units'] += it.quantity or 0
            prod[key]['name'] = (p.name if p else None) or it.description or 'Unknown'
            cogs += (it.quantity or 0) * ((p.cost_price or 0) if p else 0)

    products = scope_query(Product.query.filter_by(is_active=True), Product).all()
    low_stock = [p for p in products if p.low_stock]
    out_stock = [p for p in products if p.out_of_stock]
    inv_value = round(sum(p.stock_value for p in products), 2)
    awaiting = scope_query(PurchaseOrder.query, PurchaseOrder).filter(
        PurchaseOrder.status.in_(['Approved', 'Ordered', 'Partially Received'])).count()
    payables = 0.0
    for s in scope_query(Supplier.query, Supplier).all():
        payables += _supplier_stats(s.id)['outstanding']
    recent = (scope_query(Sale.query, Sale).order_by(Sale.created_at.desc()).limit(8).all())

    def _pack(dmap):
        return sorted([{'label': k, 'revenue': round(v, 2)} for k, v in dmap.items()],
                      key=lambda x: x['revenue'], reverse=True)[:6]
    trend = []
    dcur = month_start
    while dcur <= today:
        trend.append({'label': dcur.strftime('%d %b'), 'revenue': round(trend_map.get(dcur, 0.0), 2)})
        dcur += timedelta(days=1)
    top_products = sorted(prod.values(), key=lambda x: x['revenue'], reverse=True)[:6]
    for r in top_products:
        r['revenue'] = round(r['revenue'], 2)

    return _render({
        'page': 'dashboard',
        'today_total': today_total, 'today_count': today_count,
        'week_total': round(week_total, 2), 'month_total': round(month_total, 2),
        'month_profit': round(month_total - cogs, 2),
        'product_count': len(products), 'inventory_value': inv_value,
        'out_of_stock_count': len(out_stock), 'awaiting_delivery': awaiting,
        'expiring_soon': _expiring_soon_count(),
        'supplier_payables': round(payables, 2),
        'by_method': _pack(by_method), 'by_cashier': _pack(by_cashier),
        'by_category': _pack(by_category), 'top_products': top_products, 'trend': trend,
        'low_stock': [{'name': p.name, 'category': p.category,
                       'stock_qty': p.stock_qty, 'reorder_level': p.reorder_level} for p in low_stock],
        'recent': [_sale_row(s) for s in recent],
        'urls': {'new_sale': url_for('sales.new_sale'), 'products': url_for('sales.products'),
                 'history': url_for('sales.history'), 'analytics': url_for('sales.analytics'),
                 'movements': url_for('sales.movements'), 'suppliers': url_for('sales.suppliers'),
                 'purchases': url_for('sales.purchases'), 'reports': url_for('sales.reports'),
                 'promos': url_for('sales.promos'), 'audits': url_for('sales.audits'),
                 'assets': url_for('sales.assets')},
    })


# ---------------------------------------------------------------------------
# Products / stock
# ---------------------------------------------------------------------------

_PRODUCT_STR = ('sku', 'barcode', 'brand', 'description', 'image_url', 'unit',
                'pack_size', 'preferred_supplier', 'storage_location', 'warranty_period')
_PRODUCT_FLOAT = ('unit_price', 'cost_price', 'discount_price', 'wholesale_price',
                  'staff_price', 'student_price', 'parent_price', 'vat_rate')
_PRODUCT_INT = ('stock_qty', 'reorder_level', 'opening_stock', 'max_stock', 'reorder_qty')
_PRODUCT_REQUIRED_NUM = {'unit_price', 'cost_price', 'stock_qty', 'reorder_level'}


def _apply_product_fields(p, form, is_new=False):
    """Copy the (many, mostly optional) product fields from a form onto a Product.
    Only keys present in the POST are touched, so a partial edit never blanks a
    field it didn't submit. Required numerics default to 0; the rest may be None."""
    from utils.security import strip_tags
    if 'name' in form:
        p.name = strip_tags(form.get('name') or '').strip() or p.name
    if 'category' in form:
        p.category = (form.get('category') or '').strip() or p.category or 'Other'
    for f in _PRODUCT_STR:
        if f in form:
            setattr(p, f, strip_tags(form.get(f) or '').strip() or None)
    for f in _PRODUCT_FLOAT:
        if f in form:
            v = form.get(f, type=float)
            setattr(p, f, (v or 0) if f in _PRODUCT_REQUIRED_NUM else v)
    for f in _PRODUCT_INT:
        if f in form:
            v = form.get(f, type=int)
            setattr(p, f, (v or 0) if f in _PRODUCT_REQUIRED_NUM else v)
    if 'taxable' in form:
        p.taxable = form.get('taxable') in ('on', 'true', '1', 'yes')
    if 'batch_tracked' in form:
        p.batch_tracked = form.get('batch_tracked') in ('on', 'true', '1', 'yes')
    if 'expiry_date' in form:
        p.expiry_date = parse_date(form.get('expiry_date'))
    if 'is_active' in form:
        p.is_active = form.get('is_active') in ('on', 'true', '1', 'yes')
    if is_new and p.opening_stock is None:
        p.opening_stock = p.stock_qty or 0


@sales_bp.route('/api/isbn-lookup')
@login_required
def product_isbn_lookup():
    """Auto-fill a bookshop product from an ISBN (typed or scanned), and flag any
    existing product with the same barcode/name so stock can be topped up instead
    of duplicated — the same flow as the library catalogue."""
    from utils.isbn_lookup import lookup_isbn, normalise_isbn
    raw = request.args.get('isbn') or ''
    isbn = normalise_isbn(raw)
    if not isbn:
        return jsonify({'error': 'Enter a 10- or 13-digit ISBN.'}), 400
    meta = lookup_isbn(isbn)
    product = None
    if meta:
        bits = []
        if meta.get('author'):
            bits.append(f"by {meta['author']}")
        if meta.get('publisher'):
            bits.append(meta['publisher'])
        if meta.get('publication_year'):
            bits.append(str(meta['publication_year']))
        product = {'name': meta.get('title') or '', 'barcode': isbn,
                   'brand': meta.get('publisher') or '', 'category': 'Textbooks',
                   'description': ' · '.join(bits)}
    conds = [Product.barcode == isbn]
    if meta and meta.get('title'):
        conds.append(func.lower(Product.name) == meta['title'].strip().lower())
    matches = scope_query(Product.query.filter_by(is_active=True), Product).filter(
        db.or_(*conds)).all()
    existing = [{'id': p.id, 'name': p.name, 'stock_qty': p.stock_qty or 0,
                 'restock_url': url_for('sales.restock', product_id=p.id)} for p in matches]
    return jsonify({'found': bool(meta), 'isbn': isbn, 'product': product or {}, 'existing': existing})


def _product_dict(p):
    """Full product detail for the list + edit form."""
    return {
        'id': p.id, 'name': p.name, 'category': p.category, 'sku': p.sku,
        'barcode': p.barcode, 'brand': p.brand, 'description': p.description,
        'unit_price': p.unit_price or 0, 'cost_price': p.cost_price or 0,
        'discount_price': p.discount_price, 'wholesale_price': p.wholesale_price,
        'staff_price': p.staff_price, 'student_price': p.student_price,
        'parent_price': p.parent_price,
        'stock_qty': p.stock_qty or 0, 'reorder_level': p.reorder_level or 0,
        'opening_stock': p.opening_stock, 'max_stock': p.max_stock, 'reorder_qty': p.reorder_qty,
        'unit': p.unit, 'pack_size': p.pack_size, 'taxable': bool(p.taxable),
        'vat_rate': p.vat_rate, 'preferred_supplier': p.preferred_supplier,
        'storage_location': p.storage_location,
        'expiry_date': p.expiry_date.isoformat() if p.expiry_date else '',
        'warranty_period': p.warranty_period, 'is_active': bool(p.is_active),
        'batch_tracked': bool(p.batch_tracked),
        'low_stock': bool(p.low_stock), 'out_of_stock': bool(p.out_of_stock),
        'stock_value': p.stock_value, 'margin_pct': p.margin_pct,
        'restock_url': url_for('sales.restock', product_id=p.id),
        'edit_url': url_for('sales.edit_product', product_id=p.id),
        'adjust_url': url_for('sales.adjust_stock', product_id=p.id),
        'convert_url': url_for('sales.convert_to_asset', product_id=p.id),
    }


@sales_bp.route('/products')
@login_required
def products():
    q = (request.args.get('q') or '').strip()
    category = (request.args.get('category') or '').strip()
    stock = (request.args.get('stock') or '').strip()   # 'low' | 'out'
    query = scope_query(Product.query.filter_by(is_active=True), Product)
    if q:
        query = query.filter(db.or_(Product.name.ilike(like_term(q), escape='\\'),
                                    Product.sku.ilike(like_term(q), escape='\\'),
                                    Product.barcode.ilike(like_term(q), escape='\\')))
    if category:
        query = query.filter(Product.category == category)
    rows = query.order_by(Product.category, Product.name).all()
    if stock == 'low':
        rows = [p for p in rows if p.low_stock]
    elif stock == 'out':
        rows = [p for p in rows if p.out_of_stock]
    return _render({
        'page': 'products', 'q': q, 'category': category, 'stock': stock,
        'categories': PRODUCT_CATEGORIES, 'units': UNITS,
        'in_reasons': STOCK_IN_REASONS, 'out_reasons': STOCK_OUT_REASONS,
        'products': [_product_dict(p) for p in rows],
        'add_url': url_for('sales.add_product'),
        'isbn_lookup_url': url_for('sales.product_isbn_lookup'),
        'labels_url': url_for('sales.product_labels'),
        'generate_barcodes_url': url_for('sales.generate_barcodes'),
        'asset_categories': FIXED_ASSET_CATEGORIES,
        'urls': {'new_sale': url_for('sales.new_sale'), 'dashboard': url_for('sales.dashboard'),
                 'movements': url_for('sales.movements'), 'assets': url_for('sales.assets'),
                 'batches': url_for('sales.batches')},
    })


@sales_bp.route('/products/add', methods=['POST'])
@login_required
def add_product():
    name = (request.form.get('name') or '').strip()
    if not name:
        return _err('Product name is required.', url_for('sales.products'))
    p = Product(branch_id=branch_for_new())
    _apply_product_fields(p, request.form, is_new=True)
    db.session.add(p)
    db.session.commit()
    return _ok(f'Added "{p.name}".', url_for('sales.products'))


@sales_bp.route('/products/<int:product_id>/edit', methods=['POST'])
@login_required
def edit_product(product_id):
    p = db.get_or_404(Product, product_id)
    if not can_access_branch(p.branch_id):
        return _err('That product belongs to another branch.', url_for('sales.products'))
    _apply_product_fields(p, request.form)
    db.session.commit()
    return _ok('Product updated.', url_for('sales.products'))


def _internal_barcode(p):
    """A stable internal code for a product that has no scannable barcode of its
    own, so a printed label still scans back to it (matched by the barcode field
    in search)."""
    return f'PB{p.id:06d}'


@sales_bp.route('/products/generate-barcodes', methods=['POST'])
@login_required
def generate_barcodes():
    """Assign an internal barcode to every in-scope product that lacks one, so
    labels can be printed and scanned. Optionally limited to a category or a set
    of ids. Existing barcodes are never overwritten."""
    ids = request.form.getlist('ids', type=int)
    category = (request.form.get('category') or '').strip()
    q = scope_query(Product.query.filter_by(is_active=True), Product)
    if ids:
        q = q.filter(Product.id.in_(ids))
    if category:
        q = q.filter(Product.category == category)
    assigned = 0
    for p in q.all():
        if not (p.barcode or '').strip():
            p.barcode = _internal_barcode(p)
            assigned += 1
    db.session.commit()
    return _ok(f'Assigned {assigned} barcode(s).', url_for('sales.products'))


@sales_bp.route('/products/labels')
@login_required
def product_labels():
    """A printable sheet of product labels (name, price, Code 128 barcode).
    Filter by ids / category / stock and set how many copies of each label."""
    from utils.barcode_svg import code128_svg, encodable
    ids = request.args.getlist('ids', type=int)
    category = (request.args.get('category') or '').strip()
    stock = (request.args.get('stock') or '').strip()
    copies = max(1, min(request.args.get('copies', type=int) or 1, 50))
    price_tier = (request.args.get('tier') or '').strip() or None
    q = scope_query(Product.query.filter_by(is_active=True), Product)
    if ids:
        q = q.filter(Product.id.in_(ids))
    if category:
        q = q.filter(Product.category == category)
    rows = q.order_by(Product.category, Product.name).all()
    if stock == 'low':
        rows = [p for p in rows if p.low_stock]
    elif stock == 'out':
        rows = [p for p in rows if p.out_of_stock]

    labels = []
    for p in rows:
        code = (p.barcode or '').strip() or _internal_barcode(p)
        svg = code128_svg(code, height=44) if encodable(code) else None
        price = p.price_for(price_tier) if price_tier else (p.unit_price or 0)
        labels.append({'name': p.name, 'code': code, 'svg': svg,
                       'price': price, 'sku': p.sku or ''})
    # Each label repeated `copies` times, flattened in catalogue order.
    labels = [lab for lab in labels for _ in range(copies)]
    return render_template('sales/labels.html', labels=labels, copies=copies,
                           back_url=url_for('sales.products'))


def _consume_batches(product, quantity):
    """Draw ``quantity`` units out of a batch-tracked product's lots, earliest
    expiry first (FEFO), then earliest received. Best-effort: if the lots don't
    cover the quantity (data drift), consume what's there and stop."""
    import datetime as _dt
    batches = [b for b in StockBatch.query.filter_by(product_id=product.id).all()
               if (b.quantity or 0) > 0]
    batches.sort(key=lambda b: (b.expiry_date or _dt.date.max,
                                b.received_on or _dt.date.max, b.id))
    remaining = quantity
    for b in batches:
        if remaining <= 0:
            break
        take = min(b.quantity or 0, remaining)
        b.quantity = (b.quantity or 0) - take
        remaining -= take


def _record_movement(product, direction, quantity, reason, *, unit_cost=None,
                     reference=None, note=None, sale_id=None, apply=True,
                     batch_no=None, expiry_date=None, serial_number=None, supplier=None):
    """Write one stock-ledger row and (by default) apply the change to the
    product's on-hand quantity. Pass apply=False when the caller has already
    adjusted stock (e.g. the sale loop). For batch-tracked products a stock-in
    opens a new lot and a stock-out draws down lots FEFO. Returns the
    StockMovement."""
    quantity = abs(int(quantity or 0))
    if apply:
        product.stock_qty = (product.stock_qty or 0) + (quantity if direction == 'in' else -quantity)
    if getattr(product, 'batch_tracked', False) and quantity:
        if direction == 'in':
            db.session.add(StockBatch(
                branch_id=product.branch_id, product_id=product.id,
                batch_no=batch_no or None, serial_number=serial_number or None,
                quantity=quantity, original_qty=quantity,
                unit_cost=unit_cost if unit_cost is not None else product.cost_price,
                expiry_date=expiry_date or product.expiry_date, supplier=supplier,
                reference=reference, note=note))
        else:
            _consume_batches(product, quantity)
    mv = StockMovement(
        branch_id=product.branch_id, product_id=product.id, direction=direction,
        reason=reason, quantity=quantity, qty_after=product.stock_qty,
        unit_cost=unit_cost, reference=reference, note=note, sale_id=sale_id,
        performed_by=session.get('user') or session.get('username') or 'System')
    db.session.add(mv)
    return mv


@sales_bp.route('/products/<int:product_id>/restock', methods=['POST'])
@login_required
def restock(product_id):
    p = db.get_or_404(Product, product_id)
    if not can_access_branch(p.branch_id):
        return _err('That product belongs to another branch.', url_for('sales.products'))
    qty = request.form.get('qty', type=int) or 0
    if qty <= 0:
        return _err('Enter a quantity to add.', url_for('sales.products'))
    _record_movement(p, 'in', qty, request.form.get('reason') or 'Stock In (Purchase/GRN)',
                     unit_cost=request.form.get('unit_cost', type=float),
                     reference=(request.form.get('reference') or '').strip() or None,
                     batch_no=(request.form.get('batch_no') or '').strip() or None,
                     expiry_date=parse_date(request.form.get('expiry_date')),
                     serial_number=(request.form.get('serial_number') or '').strip() or None)
    db.session.commit()
    return _ok(f'Added {qty} to {p.name} (now {p.stock_qty}).', url_for('sales.products'))


@sales_bp.route('/products/<int:product_id>/adjust', methods=['POST'])
@login_required
def adjust_stock(product_id):
    """Record a manual stock movement (issue, usage, damage, return, transfer,
    correction…) or set an exact physical count. Every change is ledgered."""
    p = db.get_or_404(Product, product_id)
    if not can_access_branch(p.branch_id):
        return _err('That product belongs to another branch.', url_for('sales.products'))
    mode = request.form.get('mode') or 'move'
    note = (request.form.get('note') or '').strip() or None
    reference = (request.form.get('reference') or '').strip() or None

    if mode == 'count':
        counted = request.form.get('counted', type=int)
        if counted is None or counted < 0:
            return _err('Enter the counted quantity.', url_for('sales.products'))
        delta = counted - (p.stock_qty or 0)
        if delta == 0:
            return _ok('Count matches — no change.', url_for('sales.products'))
        p.stock_qty = counted
        _record_movement(p, 'in' if delta > 0 else 'out', abs(delta),
                         'Physical Stock Count', note=note, reference=reference, apply=False)
        db.session.commit()
        return _ok(f'Stock count set to {counted} for {p.name}.', url_for('sales.products'))

    direction = request.form.get('direction')
    reason = request.form.get('reason') or ''
    qty = request.form.get('quantity', type=int) or 0
    if direction not in ('in', 'out') or qty <= 0:
        return _err('Choose a direction and quantity.', url_for('sales.products'))
    allowed = STOCK_IN_REASONS if direction == 'in' else STOCK_OUT_REASONS
    if reason not in allowed:
        return _err('Invalid reason for this movement.', url_for('sales.products'))
    if direction == 'out' and qty > (p.stock_qty or 0):
        return _err(f'Only {p.stock_qty} in stock — cannot remove {qty}.', url_for('sales.products'))
    _record_movement(p, direction, qty, reason,
                     unit_cost=request.form.get('unit_cost', type=float),
                     reference=reference, note=note)
    db.session.commit()
    return _ok(f'Recorded {qty} {direction} for {p.name} (now {p.stock_qty}).',
               url_for('sales.products'))


@sales_bp.route('/movements')
@login_required
def movements():
    """The inventory movement ledger — every stock change, filterable."""
    a = request.args
    product_id = a.get('product_id', type=int)
    direction = (a.get('direction') or '').strip()
    reason = (a.get('reason') or '').strip()
    from_d = parse_date(a.get('from'))
    to_d = parse_date(a.get('to'))
    query = scope_query(StockMovement.query, StockMovement)
    if product_id:
        query = query.filter(StockMovement.product_id == product_id)
    if direction in ('in', 'out'):
        query = query.filter(StockMovement.direction == direction)
    if reason:
        query = query.filter(StockMovement.reason == reason)
    if from_d:
        query = query.filter(func.date(StockMovement.created_at) >= from_d)
    if to_d:
        query = query.filter(func.date(StockMovement.created_at) <= to_d)
    rows = query.order_by(StockMovement.created_at.desc()).limit(500).all()
    name_by_id = {p.id: p.name for p in scope_query(Product.query, Product).all()}
    total_in = sum(m.quantity for m in rows if m.direction == 'in')
    total_out = sum(m.quantity for m in rows if m.direction == 'out')
    return _render({
        'page': 'movements',
        'movements': [{'id': m.id, 'product': name_by_id.get(m.product_id, '—'),
                       'direction': m.direction, 'reason': m.reason,
                       'quantity': m.quantity, 'qty_after': m.qty_after,
                       'unit_cost': m.unit_cost, 'reference': m.reference,
                       'note': m.note, 'by': m.performed_by,
                       'when': m.created_at.strftime('%d %b %Y %H:%M') if m.created_at else ''}
                      for m in rows],
        'summary': {'count': len(rows), 'total_in': total_in, 'total_out': total_out},
        'options': {'products': [{'id': p.id, 'name': p.name} for p in scope_query(
            Product.query, Product).order_by(Product.name).all()],
            'reasons': STOCK_IN_REASONS + STOCK_OUT_REASONS + ['Sale', 'Physical Stock Count']},
        'applied': {'product_id': product_id or '', 'direction': direction,
                    'reason': reason, 'from': a.get('from', ''), 'to': a.get('to', '')},
        'self_url': url_for('sales.movements'),
        'urls': {'dashboard': url_for('sales.dashboard'), 'products': url_for('sales.products'),
                 'batches': url_for('sales.batches')},
    })


@sales_bp.route('/batches')
@login_required
def batches():
    """Live view of stock lots for batch-tracked products — remaining quantity
    and expiry per lot, filterable by product and hiding empty lots by default."""
    a = request.args
    product_id = a.get('product_id', type=int)
    show_empty = a.get('empty') in ('1', 'true', 'yes')
    query = scope_query(StockBatch.query, StockBatch)
    if product_id:
        query = query.filter(StockBatch.product_id == product_id)
    if not show_empty:
        query = query.filter(StockBatch.quantity > 0)
    rows = query.all()
    today = timeutil.today()
    rows.sort(key=lambda b: (b.expiry_date or today.replace(year=today.year + 50),
                             b.received_on or today, b.id))
    name_by_id = {p.id: p.name for p in scope_query(Product.query, Product).all()}

    def _status(b):
        if b.is_empty:
            return 'empty'
        if b.is_expired(today):
            return 'expired'
        from datetime import timedelta
        if b.expiry_date and b.expiry_date <= today + timedelta(days=30):
            return 'expiring'
        return 'ok'
    tracked = scope_query(Product.query.filter_by(is_active=True, batch_tracked=True), Product)\
        .order_by(Product.name).all()
    return _render({
        'page': 'batches',
        'batches': [{'id': b.id, 'product': name_by_id.get(b.product_id, '—'),
                     'product_id': b.product_id, 'batch_no': b.batch_no or '',
                     'serial_number': b.serial_number or '',
                     'quantity': b.quantity or 0, 'original_qty': b.original_qty or 0,
                     'unit_cost': b.unit_cost or 0,
                     'expiry_date': b.expiry_date.isoformat() if b.expiry_date else '',
                     'supplier': b.supplier or '', 'reference': b.reference or '',
                     'received_on': b.received_on.isoformat() if b.received_on else '',
                     'status': _status(b)} for b in rows],
        'options': {'products': [{'id': p.id, 'name': p.name} for p in tracked]},
        'applied': {'product_id': product_id or '', 'empty': show_empty},
        'self_url': url_for('sales.batches'),
        'urls': {'dashboard': url_for('sales.dashboard'), 'products': url_for('sales.products'),
                 'movements': url_for('sales.movements')},
    })


# ---------------------------------------------------------------------------
# Selling
# ---------------------------------------------------------------------------

@sales_bp.route('/new', methods=['GET', 'POST'])
@login_required
def new_sale():
    if request.method == 'POST':
        student_id = request.form.get('student_id', type=int) or None
        # Buyer type drives tiered pricing. A chosen student is always 'Student';
        # otherwise use the submitted type (Staff/Parent/Visitor/Walk-in).
        customer_type = (request.form.get('customer_type') or '').strip() or None
        if student_id:
            customer_type = 'Student'
        product_ids = request.form.getlist('product_id', type=int)
        quantities = request.form.getlist('quantity', type=int)
        lines = []
        total = 0.0
        for pid, qty in zip(product_ids, quantities):
            if not pid or not qty or qty <= 0:
                continue
            p = db.session.get(Product, pid)
            if not p or not can_access_branch(p.branch_id):
                continue
            if qty > (p.stock_qty or 0):
                return _err(f'Only {p.stock_qty} of "{p.name}" in stock.', url_for('sales.new_sale'))
            unit = p.price_for(customer_type)              # tiered price
            line_total = round((unit or 0) * qty, 2)
            total += line_total
            lines.append((p, qty, unit, line_total))
        if not lines:
            return _err('Add at least one item with a quantity.', url_for('sales.new_sale'))

        # Discount: an optional promo code plus/or a manual amount, capped at the
        # subtotal. A bad promo code fails the sale so the cashier notices.
        subtotal = round(total, 2)
        discount = 0.0
        discount_code = None
        used_promo = None
        code = (request.form.get('promo_code') or '').strip()
        if code:
            promo = scope_query(PromoCode.query.filter(
                func.upper(PromoCode.code) == code.upper()), PromoCode).first()
            if not promo:
                return _err('Unknown promo code.', url_for('sales.new_sale'))
            ok, reason = promo.usable(subtotal, timeutil.today())
            if not ok:
                return _err(reason, url_for('sales.new_sale'))
            if promo.category and not any(p.category == promo.category for p, _, _, _ in lines):
                return _err(f'That code only applies to {promo.category} purchases.',
                            url_for('sales.new_sale'))
            discount += promo.discount_for(subtotal)
            discount_code = promo.code
            used_promo = promo
        manual = request.form.get('discount_amount', type=float) or 0
        if manual > 0:
            discount += manual
        discount = round(min(discount, subtotal), 2)
        total = round(subtotal - discount, 2)

        sale = Sale(
            branch_id=branch_for_new(),
            student_id=student_id,
            customer_name=(request.form.get('customer_name') or '').strip() or None,
            customer_type=customer_type,
            payment_method=request.form.get('payment_method') or 'Cash',
            subtotal=subtotal, discount=discount, discount_code=discount_code,
            total=total,
            amount_paid=request.form.get('amount_paid', type=float) or total,
            sold_by=session.get('user') or session.get('username') or 'Bursar',
            notes=(request.form.get('notes') or '').strip() or None)
        if used_promo is not None:
            used_promo.used_count = (used_promo.used_count or 0) + 1
        # Stamp the revenue bucket (Bookshop / Uniform / …) from the line items so
        # the finance ledger categorises this sale when its after_insert fires —
        # the SaleItem rows don't exist in the DB yet at that point.
        from utils.finance_ledger import sale_category
        sale._ledger_category = sale_category([p.category for p, _, _, _ in lines])
        db.session.add(sale)
        db.session.flush()
        sale.receipt_no = f'SL{sale.id:05d}'
        for p, qty, unit, line_total in lines:
            db.session.add(SaleItem(sale_id=sale.id, product_id=p.id,
                                    description=p.name, quantity=qty,
                                    unit_price=unit, line_total=line_total))
            p.stock_qty = (p.stock_qty or 0) - qty      # decrement stock
            # Ledger the outward movement (stock already decremented above).
            _record_movement(p, 'out', qty, 'Sale', sale_id=sale.id,
                             reference=sale.receipt_no, apply=False)
        # Book cost of goods sold as an expense so the finance ledger reflects
        # true gross profit (revenue − COGS), recognised at the point of sale.
        from utils import finance_ledger as _fl
        _fl.post_sale_cogs(sale, lines)
        db.session.commit()
        from utils.audit import log_action
        log_action('sales.sale', detail=f'{sale.total:g} ({len(lines)} item(s))',
                   target=sale)
        return _ok(f'Sale recorded — receipt {sale.receipt_no}.',
                   url_for('sales.receipt', sale_id=sale.id))

    products = scope_query(
        Product.query.filter_by(is_active=True), Product
    ).filter(Product.stock_qty > 0).order_by(Product.category, Product.name).all()
    classes, arms = _sale_class_arm_options()
    return _render({
        'page': 'new_sale', 'methods': SALE_METHODS, 'customer_types': CUSTOMER_TYPES,
        'submit_url': url_for('sales.new_sale'),
        # Tier prices ride along so the form can price the cart live per buyer type.
        'products': [{'id': p.id, 'name': p.name, 'category': p.category,
                      'unit_price': p.unit_price or 0, 'stock_qty': p.stock_qty,
                      'student_price': p.student_price, 'staff_price': p.staff_price,
                      'parent_price': p.parent_price} for p in products],
        # Large schools can't scroll one flat dropdown of every student — the
        # buyer is chosen by class (+ optional arm) then searched on demand.
        'classes': classes, 'arms': arms,
        'student_search_url': url_for('sales.api_students'),
        'promo_check_url': url_for('sales.check_promo'),
        'urls': {'dashboard': url_for('sales.dashboard'), 'products': url_for('sales.products')},
    })


def _active_term_assignments():
    """ClassArmAssignments in the active term the current user may access."""
    active_term = get_active_term()
    if not active_term:
        return []
    return filter_classes_for_user(
        ClassArmAssignment.query.filter_by(term_id=active_term.id)
        .join(SchoolClass, ClassArmAssignment.class_id == SchoolClass.id).all())


def _sale_class_arm_options():
    """Distinct class + arm options for the sale buyer picker, scoped to what the
    user may see in the active term. Empty when no term is active."""
    assignments = _active_term_assignments()
    seen_c, seen_a, classes, arms = set(), set(), [], []
    for a in sorted(assignments, key=lambda x: ((x.school_class.level if x.school_class else 0),
                                                (x.school_class.name if x.school_class else ''))):
        if a.class_id not in seen_c and a.school_class:
            seen_c.add(a.class_id)
            classes.append({'id': a.class_id, 'name': a.school_class.name})
        arm = a.arm
        if arm and not arm.is_default and arm.id not in seen_a:
            seen_a.add(arm.id)
            arms.append({'id': arm.id, 'name': arm.name})
    arms.sort(key=lambda x: x['name'])
    return classes, arms


@sales_bp.route('/api/students')
@login_required
def api_students():
    """Students in a chosen class (+ optional arm) for the active term, matching
    an optional search — the buyer picker's on-demand lookup, so the sale form
    never ships every student. Branch/term/teacher scoped; capped."""
    class_id = request.args.get('class_id', type=int)
    arm_id = request.args.get('arm_id', type=int)
    q = (request.args.get('q') or '').strip()
    if not class_id:
        return jsonify({'students': []})
    assignments = [a for a in _active_term_assignments() if a.class_id == class_id
                   and (not arm_id or a.arm_id == arm_id)]
    aids = [a.id for a in assignments]
    if not aids:
        return jsonify({'students': []})
    query = (Student.query.filter_by(is_active=True)
             .join(StudentEnrollment, StudentEnrollment.student_id == Student.id)
             .filter(StudentEnrollment.class_arm_assignment_id.in_(aids),
                     StudentEnrollment.is_active == True))
    if q:
        term = like_term(q)
        query = query.filter(db.or_(
            Student.first_name.ilike(term, escape='\\'),
            Student.surname.ilike(term, escape='\\'),
            Student.middle_name.ilike(term, escape='\\'),
            Student.student_id.ilike(term, escape='\\')))
    rows = query.order_by(Student.surname, Student.first_name).limit(60).all()
    return jsonify({'students': [
        {'id': s.id, 'label': f'{s.full_name} ({s.student_id})',
         'student_id': s.student_id} for s in rows]})


def _history_context(limit=500):
    """Filtered, detailed sales history shared by the page and the export. Reads
    from/to, method, cashier, buyer_type, product_id, category and a text query
    from request.args; returns (rows, summary, options, applied)."""
    from datetime import timedelta
    from collections import defaultdict
    a = request.args
    to_d = parse_date(a.get('to'))
    from_d = parse_date(a.get('from'))
    method = (a.get('method') or '').strip()
    cashier = (a.get('cashier') or '').strip()
    buyer_type = (a.get('buyer_type') or '').strip()   # 'student' | 'other'
    product_id = a.get('product_id', type=int)
    category = (a.get('category') or '').strip()
    q = (a.get('q') or '').strip()

    query = scope_query(Sale.query, Sale)
    if from_d:
        query = query.filter(func.date(Sale.created_at) >= from_d)
    if to_d:
        query = query.filter(func.date(Sale.created_at) <= to_d)
    if method:
        query = query.filter(Sale.payment_method == method)
    if cashier:
        query = query.filter(Sale.sold_by == cashier)
    if buyer_type == 'student':
        query = query.filter(Sale.student_id.isnot(None))
    elif buyer_type == 'other':
        query = query.filter(Sale.student_id.is_(None))
    if q:
        term = like_term(q)
        query = query.filter(db.or_(Sale.receipt_no.ilike(term, escape='\\'),
                                    Sale.customer_name.ilike(term, escape='\\')))
    # Product / category filters compose via a SaleItem subquery.
    if product_id or category:
        sub = db.session.query(SaleItem.sale_id)
        if category:
            sub = sub.join(Product, SaleItem.product_id == Product.id).filter(Product.category == category)
        if product_id:
            sub = sub.filter(SaleItem.product_id == product_id)
        query = query.filter(Sale.id.in_(sub))

    sales = query.order_by(Sale.created_at.desc()).limit(limit).all()
    sale_ids = [s.id for s in sales]

    # Batch line items + buyer placements (no N+1).
    items_by_sale = defaultdict(list)
    if sale_ids:
        for it in SaleItem.query.filter(SaleItem.sale_id.in_(sale_ids)).all():
            items_by_sale[it.sale_id].append(it)
    placement = {}
    active_term = get_active_term()
    stud_ids = [s.student_id for s in sales if s.student_id]
    if stud_ids and active_term:
        enr = (db.session.query(StudentEnrollment.student_id, SchoolClass.name,
                                ClassArm.name, ClassArm.is_default)
               .join(ClassArmAssignment, StudentEnrollment.class_arm_assignment_id == ClassArmAssignment.id)
               .join(SchoolClass, ClassArmAssignment.class_id == SchoolClass.id)
               .join(ClassArm, ClassArmAssignment.arm_id == ClassArm.id)
               .filter(StudentEnrollment.student_id.in_(stud_ids),
                       StudentEnrollment.is_active == True,
                       ClassArmAssignment.term_id == active_term.id).all())
        for sid, cname, aname, adef in enr:
            placement[sid] = f'{cname}{"" if adef else " " + (aname or "")}'.strip()

    rows = []
    for s in sales:
        its = items_by_sale.get(s.id, [])
        paid = s.amount_paid if s.amount_paid is not None else (s.total or 0)
        rows.append({
            'id': s.id, 'receipt_no': s.receipt_no, 'buyer': s.buyer,
            'buyer_type': 'Student' if s.student_id else 'Staff / Walk-in',
            'class_arm': placement.get(s.student_id, '') if s.student_id else '',
            'cashier': s.sold_by or '', 'payment_method': s.payment_method or 'Cash',
            'total': round(s.total or 0, 2), 'amount_paid': round(paid, 2),
            'balance': round((s.total or 0) - paid, 2),
            'item_count': len(its),
            'items': [{'name': it.description or '', 'quantity': it.quantity or 0,
                       'unit_price': round(it.unit_price or 0, 2),
                       'line_total': round(it.line_total or 0, 2)} for it in its],
            'when': s.created_at.strftime('%d %b %Y %H:%M') if s.created_at else '',
            'receipt_url': url_for('sales.receipt', sale_id=s.id),
        })
    summary = {'count': len(rows), 'revenue': round(sum(r['total'] for r in rows), 2),
               'units': sum(sum(i['quantity'] for i in r['items']) for r in rows)}
    cashiers = sorted({r[0] for r in scope_query(
        db.session.query(Sale.sold_by).distinct(), Sale).all() if r[0]})
    options = {'methods': SALE_METHODS, 'categories': list(PRODUCT_CATEGORIES),
               'cashiers': cashiers,
               'products': [{'id': p.id, 'name': p.name} for p in scope_query(
                   Product.query, Product).order_by(Product.name).all()]}
    applied = {'from': a.get('from', ''), 'to': a.get('to', ''), 'method': method,
               'cashier': cashier, 'buyer_type': buyer_type,
               'product_id': product_id or '', 'category': category, 'q': q}
    return rows, summary, options, applied


@sales_bp.route('/history')
@login_required
def history():
    rows, summary, options, applied = _history_context()
    return _render({
        'page': 'history', 'total': summary['revenue'], 'summary': summary,
        'sales': rows, 'options': options, 'applied': applied,
        'self_url': url_for('sales.history'),
        'export_url': url_for('sales.history_export'),
        'urls': {'dashboard': url_for('sales.dashboard'),
                 'analytics': url_for('sales.analytics')},
    })


@sales_bp.route('/history/export')
@login_required
def history_export():
    """Export the filtered sales history as CSV or Excel (one row per sale)."""
    fmt = (request.args.get('format') or 'csv').lower()
    rows, _summary, _opts, _applied = _history_context(limit=5000)
    headers = ['Receipt', 'Date', 'Buyer', 'Type', 'Class/Arm', 'Cashier',
               'Method', 'Items', 'Total', 'Paid', 'Balance']

    def _record(r):
        return [r['receipt_no'], r['when'], r['buyer'], r['buyer_type'], r['class_arm'],
                r['cashier'], r['payment_method'], r['item_count'],
                r['total'], r['amount_paid'], r['balance']]

    if fmt in ('excel', 'xlsx'):
        from openpyxl import Workbook
        wb = Workbook(); ws = wb.active; ws.title = 'Sales'
        ws.append(headers)
        for r in rows:
            ws.append(_record(r))
        return xlsx_response(wb, 'sales_history.xlsx')
    # CSV (default)
    import csv
    import io
    from flask import Response
    out = io.StringIO(); w = csv.writer(out)
    w.writerow(headers)
    for r in rows:
        w.writerow(_record(r))
    return csv_response(out.getvalue(), 'sales_history.csv')


@sales_bp.route('/analytics')
@login_required
def analytics():
    """Sales analytics for management: revenue / units / profit over a date
    range, broken down by category, product, payment method and cashier, with a
    daily/weekly trend — plus a product drill-down showing how many students in
    each class/arm bought a chosen product. Branch-scoped."""
    from datetime import timedelta
    from collections import defaultdict

    to_d = parse_date(request.args.get('to')) or timeutil.today()
    from_d = parse_date(request.args.get('from')) or (to_d - timedelta(days=29))
    if from_d > to_d:
        from_d, to_d = to_d, from_d
    product_id = request.args.get('product_id', type=int)

    sales = (scope_query(Sale.query, Sale)
             .filter(func.date(Sale.created_at) >= from_d,
                     func.date(Sale.created_at) <= to_d).all())
    sale_ids = [s.id for s in sales]
    items = []
    if sale_ids:
        items = (db.session.query(SaleItem, Product)
                 .outerjoin(Product, SaleItem.product_id == Product.id)
                 .filter(SaleItem.sale_id.in_(sale_ids)).all())

    revenue = sum(s.total or 0 for s in sales)
    units = sum((it.quantity or 0) for it, _ in items)
    cogs = sum((it.quantity or 0) * ((p.cost_price or 0) if p else 0) for it, p in items)
    count = len(sales)

    cat = defaultdict(lambda: {'units': 0, 'revenue': 0.0})
    prod = defaultdict(lambda: {'units': 0, 'revenue': 0.0, 'name': ''})
    for it, p in items:
        c = (p.category if p else None) or 'Other'
        cat[c]['units'] += it.quantity or 0
        cat[c]['revenue'] += it.line_total or 0
        key = it.product_id or f'desc:{it.description}'
        prod[key]['units'] += it.quantity or 0
        prod[key]['revenue'] += it.line_total or 0
        prod[key]['name'] = (p.name if p else None) or it.description or 'Unknown'

    method = defaultdict(lambda: {'count': 0, 'revenue': 0.0})
    cashier = defaultdict(lambda: {'count': 0, 'revenue': 0.0})
    trend_map = defaultdict(float)
    for s in sales:
        method[s.payment_method or 'Cash']['count'] += 1
        method[s.payment_method or 'Cash']['revenue'] += s.total or 0
        cashier[s.sold_by or 'Unknown']['count'] += 1
        cashier[s.sold_by or 'Unknown']['revenue'] += s.total or 0
        if s.created_at:
            trend_map[s.created_at.date()] += s.total or 0

    # Trend: daily, or weekly buckets when the range is long, so the chart stays
    # readable.
    span = (to_d - from_d).days + 1
    trend = []
    if span > 62:
        wk_start = from_d
        while wk_start <= to_d:
            wk_end = min(wk_start + timedelta(days=6), to_d)
            rv = sum(v for dd, v in trend_map.items() if wk_start <= dd <= wk_end)
            trend.append({'label': wk_start.strftime('%d %b'), 'revenue': round(rv, 2)})
            wk_start += timedelta(days=7)
    else:
        dcur = from_d
        while dcur <= to_d:
            trend.append({'label': dcur.strftime('%d %b'), 'revenue': round(trend_map.get(dcur, 0.0), 2)})
            dcur += timedelta(days=1)

    def _pack(dmap, label_key='label'):
        out = [{label_key: k, **v} for k, v in dmap.items()]
        out.sort(key=lambda x: x.get('revenue', 0), reverse=True)
        for r in out:
            r['revenue'] = round(r['revenue'], 2)
        return out

    top_products = sorted(prod.values(), key=lambda x: x['revenue'], reverse=True)[:10]
    for r in top_products:
        r['revenue'] = round(r['revenue'], 2)

    drill = _product_drilldown(sale_ids, product_id) if product_id else None
    products_list = [{'id': p.id, 'name': p.name} for p in scope_query(
        Product.query, Product).order_by(Product.name).all()]

    return _render({
        'page': 'analytics',
        'from': from_d.isoformat(), 'to': to_d.isoformat(),
        'summary': {'revenue': round(revenue, 2), 'count': count, 'units': units,
                    'profit': round(revenue - cogs, 2),
                    'avg_sale': round(revenue / count, 2) if count else 0.0},
        'by_category': _pack(cat), 'top_products': top_products,
        'by_method': _pack(method), 'by_cashier': _pack(cashier),
        'trend': trend,
        'products': products_list, 'product_id': product_id, 'drill': drill,
        'self_url': url_for('sales.analytics'),
        'urls': {'dashboard': url_for('sales.dashboard'), 'history': url_for('sales.history')},
    })


def _product_drilldown(sale_ids, product_id):
    """For a chosen product, how many students in each class/arm bought it (and
    units/revenue), using each buyer's current active enrolment for attribution.
    Non-student buyers (staff/parent/walk-in) are summarised separately."""
    from collections import defaultdict
    if not sale_ids:
        return {'product_id': product_id, 'rows': [], 'total_units': 0,
                'total_students': 0, 'non_student_units': 0}
    rows = (db.session.query(SaleItem, Sale)
            .join(Sale, SaleItem.sale_id == Sale.id)
            .filter(SaleItem.sale_id.in_(sale_ids),
                    SaleItem.product_id == product_id).all())
    active_term = get_active_term()
    student_ids = [sale.student_id for _it, sale in rows if sale.student_id]
    placement = {}
    if student_ids and active_term:
        enr = (db.session.query(StudentEnrollment.student_id, SchoolClass.name,
                                ClassArm.name, ClassArm.is_default)
               .join(ClassArmAssignment, StudentEnrollment.class_arm_assignment_id == ClassArmAssignment.id)
               .join(SchoolClass, ClassArmAssignment.class_id == SchoolClass.id)
               .join(ClassArm, ClassArmAssignment.arm_id == ClassArm.id)
               .filter(StudentEnrollment.student_id.in_(student_ids),
                       StudentEnrollment.is_active == True,
                       ClassArmAssignment.term_id == active_term.id).all())
        for sid, cname, aname, adef in enr:
            placement[sid] = (cname, '' if adef else (aname or ''))
    by_class = defaultdict(lambda: {'students': set(), 'units': 0, 'revenue': 0.0})
    total_units = 0
    total_students = set()
    non_student_units = 0
    for it, sale in rows:
        total_units += it.quantity or 0
        if sale.student_id and sale.student_id in placement:
            cname, aname = placement[sale.student_id]
            key = f'{cname} {aname}'.strip()
            by_class[key]['students'].add(sale.student_id)
            by_class[key]['units'] += it.quantity or 0
            by_class[key]['revenue'] += it.line_total or 0
            total_students.add(sale.student_id)
        else:
            non_student_units += it.quantity or 0
    out = [{'label': k, 'students': len(v['students']), 'units': v['units'],
            'revenue': round(v['revenue'], 2)} for k, v in by_class.items()]
    out.sort(key=lambda x: x['units'], reverse=True)
    return {'product_id': product_id, 'rows': out, 'total_units': total_units,
            'total_students': len(total_students), 'non_student_units': non_student_units}


# ---------------------------------------------------------------------------
# Suppliers
# ---------------------------------------------------------------------------

_SUPPLIER_FIELDS = ('company_name', 'contact_person', 'phone', 'email', 'address',
                    'tax_id', 'bank_details', 'products_supplied', 'notes')


def _supplier_dict(s):
    return {'id': s.id, 'company_name': s.company_name, 'contact_person': s.contact_person,
            'phone': s.phone, 'email': s.email, 'address': s.address, 'tax_id': s.tax_id,
            'bank_details': s.bank_details, 'products_supplied': s.products_supplied,
            'notes': s.notes,
            'edit_url': url_for('sales.edit_supplier', supplier_id=s.id),
            'url': url_for('sales.supplier_detail', supplier_id=s.id)}


def _supplier_stats(supplier_id):
    pos = scope_query(PurchaseOrder.query.filter_by(supplier_id=supplier_id), PurchaseOrder).all()
    received = round(sum(po.received_value for po in pos), 2)
    ordered = round(sum(po.total or 0 for po in pos if po.status != 'Cancelled'), 2)
    paid = round(sum(sp.amount or 0 for sp in scope_query(
        SupplierPayment.query.filter_by(supplier_id=supplier_id), SupplierPayment).all()), 2)
    return {'orders': len(pos), 'ordered_value': ordered, 'received_value': received,
            'paid': paid, 'outstanding': round(received - paid, 2)}


@sales_bp.route('/suppliers')
@login_required
def suppliers():
    rows = scope_query(Supplier.query.filter_by(is_active=True), Supplier).order_by(
        Supplier.company_name).all()
    return _render({
        'page': 'suppliers',
        'suppliers': [{**_supplier_dict(s), **_supplier_stats(s.id)} for s in rows],
        'add_url': url_for('sales.add_supplier'),
        'urls': {'dashboard': url_for('sales.dashboard'), 'purchases': url_for('sales.purchases')},
    })


def _apply_supplier_fields(s, form):
    from utils.security import strip_tags
    for f in _SUPPLIER_FIELDS:
        if f in form:
            setattr(s, f, strip_tags(form.get(f) or '').strip() or None)


@sales_bp.route('/suppliers/add', methods=['POST'])
@login_required
def add_supplier():
    name = (request.form.get('company_name') or '').strip()
    if not name:
        return _err('Company name is required.', url_for('sales.suppliers'))
    s = Supplier(branch_id=branch_for_new())
    _apply_supplier_fields(s, request.form)
    db.session.add(s)
    db.session.commit()
    return _ok(f'Added supplier "{s.company_name}".', url_for('sales.suppliers'))


@sales_bp.route('/suppliers/<int:supplier_id>/edit', methods=['POST'])
@login_required
def edit_supplier(supplier_id):
    s = db.get_or_404(Supplier, supplier_id)
    if not can_access_branch(s.branch_id):
        return _err('That supplier belongs to another branch.', url_for('sales.suppliers'))
    _apply_supplier_fields(s, request.form)
    db.session.commit()
    return _ok('Supplier updated.', url_for('sales.suppliers'))


@sales_bp.route('/suppliers/<int:supplier_id>')
@login_required
def supplier_detail(supplier_id):
    s = db.get_or_404(Supplier, supplier_id)
    if not can_access_branch(s.branch_id):
        flash('That supplier belongs to another branch.', 'error')
        return redirect(url_for('sales.suppliers'))
    pos = (scope_query(PurchaseOrder.query.filter_by(supplier_id=supplier_id), PurchaseOrder)
           .order_by(PurchaseOrder.created_at.desc()).all())
    pays = (scope_query(SupplierPayment.query.filter_by(supplier_id=supplier_id), SupplierPayment)
            .order_by(SupplierPayment.created_at.desc()).all())
    return _render({
        'page': 'supplier_detail', 'supplier': _supplier_dict(s),
        'stats': _supplier_stats(supplier_id),
        'orders': [_po_row(po) for po in pos],
        'payments': [{'id': p.id, 'amount': p.amount or 0, 'method': p.method,
                      'reference': p.reference, 'note': p.note, 'by': p.paid_by,
                      'when': p.created_at.strftime('%d %b %Y') if p.created_at else ''} for p in pays],
        'pay_url': url_for('sales.pay_supplier', supplier_id=supplier_id),
        'methods': PURCHASE_METHODS,
        'urls': {'suppliers': url_for('sales.suppliers'), 'purchases': url_for('sales.purchases')},
    })


@sales_bp.route('/suppliers/<int:supplier_id>/pay', methods=['POST'])
@login_required
def pay_supplier(supplier_id):
    s = db.get_or_404(Supplier, supplier_id)
    if not can_access_branch(s.branch_id):
        return _err('That supplier belongs to another branch.', url_for('sales.suppliers'))
    amount = request.form.get('amount', type=float) or 0
    if amount <= 0:
        return _err('Enter a payment amount.', url_for('sales.supplier_detail', supplier_id=supplier_id))
    db.session.add(SupplierPayment(
        branch_id=s.branch_id, supplier_id=supplier_id,
        po_id=request.form.get('po_id', type=int) or None,
        amount=round(amount, 2), method=request.form.get('method') or 'Cash',
        reference=(request.form.get('reference') or '').strip() or None,
        note=(request.form.get('note') or '').strip() or None,
        paid_by=session.get('user') or session.get('username') or 'Bursar'))
    db.session.commit()
    return _ok(f'Recorded payment of {amount:g}.',
               url_for('sales.supplier_detail', supplier_id=supplier_id))


# ---------------------------------------------------------------------------
# Purchase orders
# ---------------------------------------------------------------------------

def _po_row(po):
    return {'id': po.id, 'po_number': po.po_number, 'supplier': po.supplier.company_name if po.supplier else '—',
            'status': po.status, 'total': po.total or 0, 'received_value': po.received_value,
            'expected_date': po.expected_date.isoformat() if po.expected_date else '',
            'created_at': po.created_at.strftime('%d %b %Y') if po.created_at else '',
            'url': url_for('sales.purchase_detail', po_id=po.id)}


@sales_bp.route('/purchases')
@login_required
def purchases():
    status = (request.args.get('status') or '').strip()
    supplier_id = request.args.get('supplier_id', type=int)
    query = scope_query(PurchaseOrder.query, PurchaseOrder)
    if status:
        query = query.filter(PurchaseOrder.status == status)
    if supplier_id:
        query = query.filter(PurchaseOrder.supplier_id == supplier_id)
    rows = query.order_by(PurchaseOrder.created_at.desc()).limit(300).all()
    sups = scope_query(Supplier.query.filter_by(is_active=True), Supplier).order_by(Supplier.company_name).all()
    awaiting = [po for po in scope_query(PurchaseOrder.query, PurchaseOrder).filter(
        PurchaseOrder.status.in_(['Approved', 'Ordered', 'Partially Received'])).all()]
    return _render({
        'page': 'purchases',
        'orders': [_po_row(po) for po in rows],
        'statuses': PO_STATUSES,
        'applied': {'status': status, 'supplier_id': supplier_id or ''},
        'awaiting_delivery': len(awaiting),
        'suppliers': [{'id': s.id, 'company_name': s.company_name} for s in sups],
        'products': [{'id': p.id, 'name': p.name, 'cost_price': p.cost_price or 0}
                     for p in scope_query(Product.query.filter_by(is_active=True), Product).order_by(Product.name).all()],
        'new_url': url_for('sales.new_purchase'), 'self_url': url_for('sales.purchases'),
        'urls': {'dashboard': url_for('sales.dashboard'), 'suppliers': url_for('sales.suppliers')},
    })


@sales_bp.route('/purchases/new', methods=['POST'])
@login_required
def new_purchase():
    data = request.get_json(silent=True) or request.form
    supplier_id = data.get('supplier_id')
    try:
        supplier_id = int(supplier_id)
    except (TypeError, ValueError):
        return _err('Choose a supplier.', url_for('sales.purchases'))
    supplier = db.session.get(Supplier, supplier_id)
    if not supplier or not can_access_branch(supplier.branch_id):
        return _err('Unknown supplier.', url_for('sales.purchases'))
    items = data.get('items') or []
    if isinstance(items, str):
        import json as _json
        try:
            items = _json.loads(items)
        except ValueError:
            items = []
    clean = []
    for it in items:
        qty = int(it.get('quantity') or 0)
        cost = float(it.get('unit_cost') or 0)
        pid = it.get('product_id') or None
        desc = (it.get('description') or '').strip()
        if pid:
            p = db.session.get(Product, int(pid))
            if p:
                desc = desc or p.name
        if qty > 0 and desc:
            clean.append({'product_id': int(pid) if pid else None, 'description': desc,
                          'quantity': qty, 'unit_cost': cost})
    if not clean:
        return _err('Add at least one item.', url_for('sales.purchases'))
    submit = (data.get('submit') or 'draft')
    po = PurchaseOrder(
        branch_id=branch_for_new(), supplier_id=supplier_id,
        status='Pending Approval' if submit == 'submit' else 'Draft',
        expected_date=parse_date(data.get('expected_date')),
        invoice_number=(data.get('invoice_number') or '').strip() or None,
        notes=(data.get('notes') or '').strip() or None,
        total=round(sum(i['quantity'] * i['unit_cost'] for i in clean), 2),
        created_by=session.get('user') or session.get('username') or 'Bursar')
    db.session.add(po)
    db.session.flush()
    po.po_number = f'PO{po.id:05d}'
    for i in clean:
        db.session.add(PurchaseOrderItem(po_id=po.id, product_id=i['product_id'],
                                         description=i['description'], quantity=i['quantity'],
                                         unit_cost=i['unit_cost']))
    db.session.commit()
    from utils.audit import log_action
    log_action('sales.purchase_order', detail=f'{po.po_number} · {po.total:g}', target=po)
    return _ok(f'Purchase order {po.po_number} created.',
               url_for('sales.purchase_detail', po_id=po.id))


@sales_bp.route('/purchases/<int:po_id>')
@login_required
def purchase_detail(po_id):
    po = db.get_or_404(PurchaseOrder, po_id)
    if not can_access_branch(po.branch_id):
        flash('That order belongs to another branch.', 'error')
        return redirect(url_for('sales.purchases'))
    items = po.items.all()
    return _render({
        'page': 'purchase_detail',
        'po': {**_po_row(po), 'invoice_number': po.invoice_number, 'notes': po.notes,
               'created_by': po.created_by, 'approved_by': po.approved_by,
               'supplier_id': po.supplier_id, 'is_open': po.is_open},
        'can_approve': can_approve_purchase(),
        'items': [{'id': i.id, 'description': i.description, 'quantity': i.quantity,
                   'unit_cost': i.unit_cost or 0, 'quantity_received': i.quantity_received or 0,
                   'outstanding': i.outstanding_qty, 'line_total': i.line_total} for i in items],
        'urls': {'purchases': url_for('sales.purchases'),
                 'supplier': url_for('sales.supplier_detail', supplier_id=po.supplier_id),
                 'approve': url_for('sales.approve_purchase', po_id=po.id),
                 'receive': url_for('sales.receive_purchase', po_id=po.id),
                 'cancel': url_for('sales.cancel_purchase', po_id=po.id)},
    })


@sales_bp.route('/purchases/<int:po_id>/approve', methods=['POST'])
@login_required
def approve_purchase(po_id):
    po = db.get_or_404(PurchaseOrder, po_id)
    if not can_access_branch(po.branch_id):
        return _err('That order belongs to another branch.', url_for('sales.purchases'))
    if not can_approve_purchase():
        return _err('You are not authorised to approve purchase orders.',
                    url_for('sales.purchase_detail', po_id=po.id))
    if po.status not in ('Draft', 'Pending Approval'):
        return _err('This order can no longer be approved.', url_for('sales.purchase_detail', po_id=po.id))
    from datetime import datetime as _dt
    po.status = 'Approved'
    po.approved_by = session.get('user') or session.get('username') or 'Admin'
    po.approved_at = _dt.now()
    db.session.commit()
    return _ok(f'{po.po_number} approved.', url_for('sales.purchase_detail', po_id=po.id))


@sales_bp.route('/purchases/<int:po_id>/cancel', methods=['POST'])
@login_required
def cancel_purchase(po_id):
    po = db.get_or_404(PurchaseOrder, po_id)
    if not can_access_branch(po.branch_id):
        return _err('That order belongs to another branch.', url_for('sales.purchases'))
    if po.status == 'Received':
        return _err('A fully received order cannot be cancelled.', url_for('sales.purchase_detail', po_id=po.id))
    po.status = 'Cancelled'
    db.session.commit()
    return _ok(f'{po.po_number} cancelled.', url_for('sales.purchase_detail', po_id=po.id))


@sales_bp.route('/purchases/<int:po_id>/receive', methods=['POST'])
@login_required
def receive_purchase(po_id):
    """Goods Received: record received quantities, add them to stock (ledgered)
    and advance the order status. Receiving is blocked until the PO is approved."""
    po = db.get_or_404(PurchaseOrder, po_id)
    if not can_access_branch(po.branch_id):
        return _err('That order belongs to another branch.', url_for('sales.purchases'))
    if po.status in ('Draft', 'Pending Approval'):
        return _err('Approve the order before receiving goods.', url_for('sales.purchase_detail', po_id=po.id))
    if po.status in ('Received', 'Cancelled'):
        return _err('This order is closed.', url_for('sales.purchase_detail', po_id=po.id))
    data = request.get_json(silent=True) or request.form
    recv = data.get('items') or []
    if isinstance(recv, str):
        import json as _json
        try:
            recv = _json.loads(recv)
        except ValueError:
            recv = []
    by_id = {int(r['item_id']): int(r.get('receive_qty') or 0) for r in recv if r.get('item_id')}
    invoice = (data.get('invoice_number') or '').strip() or None
    if invoice:
        po.invoice_number = invoice
    received_any = False
    for it in po.items.all():
        take = by_id.get(it.id, 0)
        if take <= 0:
            continue
        take = min(take, it.outstanding_qty)
        if take <= 0:
            continue
        it.quantity_received = (it.quantity_received or 0) + take
        received_any = True
        if it.product_id:
            p = db.session.get(Product, it.product_id)
            if p:
                _record_movement(p, 'in', take, 'Stock In (Purchase/GRN)',
                                 unit_cost=it.unit_cost, reference=po.po_number,
                                 batch_no=po.po_number, supplier=(po.supplier.company_name if po.supplier else None))
                if it.unit_cost:
                    p.cost_price = it.unit_cost      # keep valuation current
    if not received_any:
        return _err('Enter quantities to receive.', url_for('sales.purchase_detail', po_id=po.id))
    fully = all(i.outstanding_qty == 0 for i in po.items.all())
    po.status = 'Received' if fully else 'Partially Received'
    db.session.commit()
    return _ok(f'Goods received for {po.po_number}.', url_for('sales.purchase_detail', po_id=po.id))


# ---------------------------------------------------------------------------
# Discounts & promo codes
# ---------------------------------------------------------------------------

@sales_bp.route('/api/promo')
@login_required
def check_promo():
    """Validate a promo code against a subtotal and preview the discount."""
    code = (request.args.get('code') or '').strip()
    subtotal = request.args.get('subtotal', type=float) or 0
    if not code:
        return jsonify({'ok': False, 'error': 'Enter a code.'})
    promo = scope_query(PromoCode.query.filter(
        func.upper(PromoCode.code) == code.upper()), PromoCode).first()
    if not promo:
        return jsonify({'ok': False, 'error': 'Unknown code.'})
    ok, reason = promo.usable(subtotal, timeutil.today())
    if not ok:
        return jsonify({'ok': False, 'error': reason})
    return jsonify({'ok': True, 'code': promo.code, 'discount': promo.discount_for(subtotal),
                    'description': promo.description or '', 'category': promo.category or ''})


@sales_bp.route('/promos')
@login_required
def promos():
    rows = scope_query(PromoCode.query, PromoCode).order_by(PromoCode.created_at.desc()).all()
    return _render({
        'page': 'promos',
        'promos': [{'id': p.id, 'code': p.code, 'description': p.description,
                    'kind': p.kind, 'value': p.value or 0, 'min_purchase': p.min_purchase or 0,
                    'category': p.category, 'expires_on': p.expires_on.isoformat() if p.expires_on else '',
                    'usage_limit': p.usage_limit, 'used_count': p.used_count or 0,
                    'is_active': bool(p.is_active),
                    'toggle_url': url_for('sales.toggle_promo', promo_id=p.id)} for p in rows],
        'categories': PRODUCT_CATEGORIES, 'add_url': url_for('sales.add_promo'),
        'urls': {'dashboard': url_for('sales.dashboard')},
    })


@sales_bp.route('/promos/add', methods=['POST'])
@login_required
def add_promo():
    from utils.security import strip_tags
    code = (request.form.get('code') or '').strip().upper()
    if not code:
        return _err('A code is required.', url_for('sales.promos'))
    if scope_query(PromoCode.query.filter(func.upper(PromoCode.code) == code), PromoCode).first():
        return _err('That code already exists.', url_for('sales.promos'))
    kind = request.form.get('kind') if request.form.get('kind') in ('percent', 'fixed') else 'percent'
    db.session.add(PromoCode(
        branch_id=branch_for_new(), code=code,
        description=strip_tags(request.form.get('description') or '').strip() or None,
        kind=kind, value=request.form.get('value', type=float) or 0,
        min_purchase=request.form.get('min_purchase', type=float) or 0,
        category=(request.form.get('category') or '').strip() or None,
        expires_on=parse_date(request.form.get('expires_on')),
        usage_limit=request.form.get('usage_limit', type=int),
        created_by=session.get('user') or session.get('username') or 'Admin'))
    db.session.commit()
    return _ok(f'Promo code "{code}" added.', url_for('sales.promos'))


@sales_bp.route('/promos/<int:promo_id>/toggle', methods=['POST'])
@login_required
def toggle_promo(promo_id):
    p = db.get_or_404(PromoCode, promo_id)
    if not can_access_branch(p.branch_id):
        return _err('That code belongs to another branch.', url_for('sales.promos'))
    p.is_active = not p.is_active
    db.session.commit()
    return _ok(f'{p.code} {"activated" if p.is_active else "deactivated"}.', url_for('sales.promos'))


# ---------------------------------------------------------------------------
# Stock audits (physical count + variance sign-off)
# ---------------------------------------------------------------------------

def _audit_row(a):
    counted = a.counted_list
    return {
        'id': a.id, 'reference': a.reference, 'status': a.status,
        'scope': a.scope_category or a.scope_location or 'All products',
        'note': a.note or '', 'started_by': a.started_by or '',
        'approved_by': a.approved_by or '',
        'items': a.items.count(), 'counted': len(counted),
        'variance_value': round(a.variance_value or 0, 2),
        'when': a.created_at.strftime('%d %b %Y') if a.created_at else '',
        'url': url_for('sales.audit_detail', audit_id=a.id),
    }


@sales_bp.route('/audits')
@login_required
def audits():
    rows = scope_query(StockAudit.query, StockAudit).order_by(StockAudit.created_at.desc()).all()
    cats = sorted({c for (c,) in scope_query(
        db.session.query(Product.category), Product).distinct().all() if c})
    locs = sorted({l for (l,) in scope_query(
        db.session.query(Product.storage_location), Product).distinct().all() if l})
    return _render({
        'page': 'audits', 'audits': [_audit_row(a) for a in rows],
        'categories': cats, 'locations': locs,
        'new_url': url_for('sales.new_audit'),
        'urls': {'dashboard': url_for('sales.dashboard')},
    })


@sales_bp.route('/audits/new', methods=['POST'])
@login_required
def new_audit():
    """Open a count session: snapshot the on-hand quantity of every active
    product (optionally filtered to a category or storage location) so the
    physical count can be entered against it."""
    category = (request.form.get('category') or '').strip() or None
    location = (request.form.get('location') or '').strip() or None
    q = scope_query(Product.query.filter_by(is_active=True), Product)
    if category:
        q = q.filter(Product.category == category)
    if location:
        q = q.filter(Product.storage_location == location)
    products = q.order_by(Product.category, Product.name).all()
    if not products:
        return _err('No matching products to count.', url_for('sales.audits'))
    audit = StockAudit(
        branch_id=branch_for_new(), scope_category=category, scope_location=location,
        status='Counting', note=(request.form.get('note') or '').strip() or None,
        started_by=session.get('user') or session.get('username') or 'Bursar')
    db.session.add(audit)
    db.session.flush()
    audit.reference = f'SA{audit.id:05d}'
    for p in products:
        db.session.add(StockAuditItem(
            audit_id=audit.id, product_id=p.id, product_name=p.name,
            system_qty=p.stock_qty or 0, unit_cost=p.cost_price or 0))
    db.session.commit()
    return _ok(f'Count sheet {audit.reference} ready ({len(products)} items).',
               url_for('sales.audit_detail', audit_id=audit.id))


def _audit_detail_payload(a):
    items = a.items.join(Product, StockAuditItem.product_id == Product.id, isouter=True)\
             .order_by(Product.category, StockAuditItem.product_name).all()
    net = round(sum(i.variance_value for i in a.counted_list), 2)
    return {
        'page': 'audit_detail', 'id': a.id, 'reference': a.reference,
        'status': a.status, 'scope': a.scope_category or a.scope_location or 'All products',
        'note': a.note or '', 'started_by': a.started_by or '',
        'approved_by': a.approved_by or '',
        'variance_value': round(a.variance_value or 0, 2),
        'net_preview': net, 'can_signoff': can_sign_off_count(),
        'items': [{
            'id': i.id, 'product': i.product_name,
            'category': (i.product.category if i.product else '') or '',
            'system_qty': i.system_qty or 0, 'counted_qty': i.counted_qty,
            'unit_cost': i.unit_cost or 0,
            'variance_qty': i.variance_qty, 'variance_value': i.variance_value,
            'note': i.note or '',
        } for i in items],
        'save_url': url_for('sales.save_audit', audit_id=a.id),
        'complete_url': url_for('sales.complete_audit', audit_id=a.id),
        'cancel_url': url_for('sales.cancel_audit', audit_id=a.id),
        'export_url': url_for('sales.audit_export', audit_id=a.id),
        'urls': {'audits': url_for('sales.audits')},
    }


@sales_bp.route('/audits/<int:audit_id>')
@login_required
def audit_detail(audit_id):
    a = db.get_or_404(StockAudit, audit_id)
    if not can_access_branch(a.branch_id):
        return _err('That count belongs to another branch.', url_for('sales.audits'))
    return _render(_audit_detail_payload(a))


def _load_audit_counts(data):
    counts = data.get('counts') or data.get('items') or []
    if isinstance(counts, str):
        import json as _json
        try:
            counts = _json.loads(counts)
        except ValueError:
            counts = []
    out = {}
    for r in counts:
        try:
            iid = int(r['item_id'])
        except (KeyError, ValueError, TypeError):
            continue
        raw = r.get('counted_qty')
        out[iid] = None if raw in (None, '') else int(raw)
    return out


@sales_bp.route('/audits/<int:audit_id>/save', methods=['POST'])
@login_required
def save_audit(audit_id):
    """Persist entered counts without closing the session (save-as-you-go)."""
    a = db.get_or_404(StockAudit, audit_id)
    if not can_access_branch(a.branch_id):
        return _err('That count belongs to another branch.', url_for('sales.audits'))
    if not a.is_open:
        return _err('This count is closed.', url_for('sales.audit_detail', audit_id=a.id))
    counts = _load_audit_counts(request.get_json(silent=True) or request.form)
    for it in a.items.all():
        if it.id in counts:
            it.counted_qty = counts[it.id]
    db.session.commit()
    if _wants_json():
        return jsonify({'ok': True, 'net_preview': round(
            sum(i.variance_value for i in a.counted_list), 2)})
    return _ok('Counts saved.', url_for('sales.audit_detail', audit_id=a.id))


@sales_bp.route('/audits/<int:audit_id>/complete', methods=['POST'])
@login_required
def complete_audit(audit_id):
    """Sign off the count: apply each variance to stock (ledgered as a Physical
    Stock Count movement) and post the net cost value to the finance ledger —
    shrinkage as an expense, an overage as an inventory gain."""
    a = db.get_or_404(StockAudit, audit_id)
    if not can_access_branch(a.branch_id):
        return _err('That count belongs to another branch.', url_for('sales.audits'))
    if not can_sign_off_count():
        return _err('You are not authorised to sign off stock counts.',
                    url_for('sales.audit_detail', audit_id=a.id))
    if not a.is_open:
        return _err('This count is already closed.', url_for('sales.audit_detail', audit_id=a.id))
    # Absorb any counts submitted alongside the sign-off.
    counts = _load_audit_counts(request.get_json(silent=True) or request.form)
    for it in a.items.all():
        if it.id in counts:
            it.counted_qty = counts[it.id]
    db.session.flush()

    counted = a.counted_list
    if not counted:
        return _err('Enter at least one count before signing off.',
                    url_for('sales.audit_detail', audit_id=a.id))
    net_value = 0.0
    for it in counted:
        delta = it.variance_qty
        if not delta:
            continue
        p = it.product or db.session.get(Product, it.product_id)
        if not p:
            continue
        p.stock_qty = it.counted_qty
        _record_movement(p, 'in' if delta > 0 else 'out', abs(delta),
                         'Physical Stock Count', reference=a.reference,
                         note=f'Audit {a.reference}', apply=False)
        net_value += it.variance_value
    net_value = round(net_value, 2)
    a.variance_value = net_value
    a.status = 'Completed'
    a.approved_by = session.get('user') or session.get('username') or 'Bursar'
    a.completed_at = timeutil.now()
    # Post the net inventory variance to the finance ledger (idempotent per audit).
    if net_value and not a.ledger_posted:
        from utils import finance_ledger as _fl
        if net_value < 0:
            _fl.post(_fl.EXPENSE, abs(net_value), source_module='inventory',
                     category='Inventory Shrinkage', branch_id=a.branch_id,
                     origin_type='stock_audit', origin_id=a.id, reference=a.reference,
                     description=f'Stock count shrinkage ({a.reference})',
                     created_by=a.approved_by)
        else:
            _fl.post(_fl.REVENUE, net_value, source_module='inventory',
                     category='Inventory Gain', branch_id=a.branch_id,
                     origin_type='stock_audit', origin_id=a.id, reference=a.reference,
                     description=f'Stock count overage ({a.reference})',
                     created_by=a.approved_by)
        a.ledger_posted = True
    db.session.commit()
    from utils.audit import log_action
    log_action('sales.stock_audit', detail=f'{a.reference} net {net_value:g}', target=a)
    return _ok(f'{a.reference} signed off — net variance {net_value:g}.',
               url_for('sales.audit_detail', audit_id=a.id))


@sales_bp.route('/audits/<int:audit_id>/cancel', methods=['POST'])
@login_required
def cancel_audit(audit_id):
    a = db.get_or_404(StockAudit, audit_id)
    if not can_access_branch(a.branch_id):
        return _err('That count belongs to another branch.', url_for('sales.audits'))
    if not a.is_open:
        return _err('This count is already closed.', url_for('sales.audit_detail', audit_id=a.id))
    a.status = 'Cancelled'
    db.session.commit()
    return _ok(f'{a.reference} cancelled — no stock changed.',
               url_for('sales.audit_detail', audit_id=a.id))


@sales_bp.route('/audits/<int:audit_id>/export')
@login_required
def audit_export(audit_id):
    a = db.get_or_404(StockAudit, audit_id)
    if not can_access_branch(a.branch_id):
        return _err('That count belongs to another branch.', url_for('sales.audits'))
    from openpyxl import Workbook
    wb = Workbook(); ws = wb.active; ws.title = 'Stock Count'
    ws.append(['Product', 'Category', 'System Qty', 'Counted', 'Variance',
               'Unit Cost', 'Variance Value', 'Note'])
    for i in a.items.order_by(StockAuditItem.product_name).all():
        ws.append([i.product_name, (i.product.category if i.product else ''),
                   i.system_qty or 0,
                   '' if i.counted_qty is None else i.counted_qty,
                   '' if i.variance_qty is None else i.variance_qty,
                   i.unit_cost or 0, i.variance_value, i.note or ''])
    return xlsx_response(wb, f'{a.reference}_count.xlsx')


# ---------------------------------------------------------------------------
# Fixed assets (register + conversion from inventory)
# ---------------------------------------------------------------------------

def _asset_dict(a):
    return {
        'id': a.id, 'name': a.name, 'asset_tag': a.asset_tag or '',
        'category': a.category, 'description': a.description or '',
        'serial_number': a.serial_number or '',
        'acquisition_cost': a.acquisition_cost or 0,
        'acquisition_date': a.acquisition_date.isoformat() if a.acquisition_date else '',
        'supplier': a.supplier or '', 'location': a.location or '',
        'custodian': a.custodian or '', 'status': a.status,
        'useful_life_years': a.useful_life_years, 'salvage_value': a.salvage_value or 0,
        'quantity': a.quantity or 1,
        'annual_depreciation': a.annual_depreciation,
        'accumulated_depreciation': a.accumulated_depreciation,
        'book_value': a.book_value, 'is_disposed': a.is_disposed,
        'disposed_on': a.disposed_on.isoformat() if a.disposed_on else '',
        'disposal_amount': a.disposal_amount,
        'from_product': bool(a.source_product_id),
        'edit_url': url_for('sales.edit_asset', asset_id=a.id),
        'dispose_url': url_for('sales.dispose_asset', asset_id=a.id),
    }


def _apply_asset_fields(a, form):
    from utils.security import strip_tags
    a.name = strip_tags(form.get('name') or '').strip() or a.name
    a.asset_tag = (form.get('asset_tag') or '').strip() or None
    cat = (form.get('category') or '').strip()
    if cat:
        a.category = cat
    a.description = strip_tags(form.get('description') or '').strip() or None
    a.serial_number = (form.get('serial_number') or '').strip() or None
    if form.get('acquisition_cost') is not None:
        a.acquisition_cost = form.get('acquisition_cost', type=float) or 0
    d = parse_date(form.get('acquisition_date'))
    if d:
        a.acquisition_date = d
    a.supplier = (form.get('supplier') or '').strip() or None
    a.location = (form.get('location') or '').strip() or None
    a.custodian = (form.get('custodian') or '').strip() or None
    st = (form.get('status') or '').strip()
    if st in FIXED_ASSET_STATUSES:
        a.status = st
    a.useful_life_years = form.get('useful_life_years', type=int)
    a.salvage_value = form.get('salvage_value', type=float) or 0


@sales_bp.route('/assets')
@login_required
def assets():
    q = (request.args.get('q') or '').strip()
    category = (request.args.get('category') or '').strip()
    status = (request.args.get('status') or '').strip()
    query = scope_query(FixedAsset.query, FixedAsset)
    if q:
        query = query.filter(db.or_(FixedAsset.name.ilike(like_term(q), escape='\\'),
                                    FixedAsset.asset_tag.ilike(like_term(q), escape='\\'),
                                    FixedAsset.serial_number.ilike(like_term(q), escape='\\')))
    if category:
        query = query.filter(FixedAsset.category == category)
    if status:
        query = query.filter(FixedAsset.status == status)
    rows = query.order_by(FixedAsset.created_at.desc()).all()
    live = [a for a in rows if not a.is_disposed]
    from collections import defaultdict
    by_cat = defaultdict(lambda: {'count': 0, 'cost': 0.0, 'book': 0.0})
    for a in live:
        b = by_cat[a.category or 'Other']
        b['count'] += 1; b['cost'] += a.acquisition_cost or 0; b['book'] += a.book_value
    summary = {
        'count': len(live),
        'total_cost': round(sum(a.acquisition_cost or 0 for a in live), 2),
        'total_book': round(sum(a.book_value for a in live), 2),
        'disposed': sum(1 for a in rows if a.is_disposed),
        'by_category': sorted(
            [{'category': k, **{kk: round(vv, 2) if isinstance(vv, float) else vv
                                for kk, vv in v.items()}} for k, v in by_cat.items()],
            key=lambda r: -r['cost']),
    }
    return _render({
        'page': 'assets', 'assets': [_asset_dict(a) for a in rows], 'summary': summary,
        'q': q, 'category': category, 'status': status,
        'categories': FIXED_ASSET_CATEGORIES, 'statuses': FIXED_ASSET_STATUSES,
        'add_url': url_for('sales.add_asset'),
        'export_url': url_for('sales.assets_export'),
        'urls': {'dashboard': url_for('sales.dashboard'), 'products': url_for('sales.products')},
    })


@sales_bp.route('/assets/add', methods=['POST'])
@login_required
def add_asset():
    name = (request.form.get('name') or '').strip()
    if not name:
        return _err('Asset name is required.', url_for('sales.assets'))
    a = FixedAsset(branch_id=branch_for_new(), name=name, quantity=1,
                   acquisition_date=timeutil.today(),
                   created_by=session.get('user') or session.get('username') or 'Admin')
    _apply_asset_fields(a, request.form)
    db.session.add(a)
    db.session.commit()
    return _ok(f'Registered "{a.name}".', url_for('sales.assets'))


@sales_bp.route('/assets/<int:asset_id>/edit', methods=['POST'])
@login_required
def edit_asset(asset_id):
    a = db.get_or_404(FixedAsset, asset_id)
    if not can_access_branch(a.branch_id):
        return _err('That asset belongs to another branch.', url_for('sales.assets'))
    _apply_asset_fields(a, request.form)
    db.session.commit()
    return _ok('Asset updated.', url_for('sales.assets'))


@sales_bp.route('/products/<int:product_id>/convert-asset', methods=['POST'])
@login_required
def convert_to_asset(product_id):
    """Convert on-hand units of a product into a tracked fixed asset. The units
    are drawn out of stock (ledgered as a movement); inventory value simply
    becomes fixed-asset value, so there's no profit-and-loss impact."""
    p = db.get_or_404(Product, product_id)
    if not can_access_branch(p.branch_id):
        return _err('That product belongs to another branch.', url_for('sales.products'))
    qty = request.form.get('quantity', type=int) or 1
    if qty <= 0:
        return _err('Enter how many units to convert.', url_for('sales.products'))
    if qty > (p.stock_qty or 0):
        return _err(f'Only {p.stock_qty} in stock.', url_for('sales.products'))
    cost = request.form.get('acquisition_cost', type=float)
    if cost is None:
        cost = round((p.cost_price or 0) * qty, 2)      # default to stock cost
    a = FixedAsset(
        branch_id=p.branch_id,
        name=(request.form.get('name') or '').strip() or p.name,
        asset_tag=(request.form.get('asset_tag') or '').strip() or None,
        category=(request.form.get('category') or '').strip() or 'Other',
        serial_number=(request.form.get('serial_number') or '').strip() or None,
        acquisition_cost=cost, acquisition_date=timeutil.today(),
        supplier=p.preferred_supplier, location=p.storage_location,
        custodian=(request.form.get('custodian') or '').strip() or None,
        status='In Use', quantity=qty, source_product_id=p.id,
        useful_life_years=request.form.get('useful_life_years', type=int),
        salvage_value=request.form.get('salvage_value', type=float) or 0,
        created_by=session.get('user') or session.get('username') or 'Admin')
    p.stock_qty = (p.stock_qty or 0) - qty
    _record_movement(p, 'out', qty, 'Converted to Fixed Asset',
                     note=f'Fixed asset: {a.name}', apply=False)
    db.session.add(a)
    db.session.commit()
    from utils.audit import log_action
    log_action('sales.asset_convert', detail=f'{a.name} x{qty}', target=a)
    return _ok(f'Converted {qty} × {p.name} into fixed asset "{a.name}".',
               url_for('sales.assets'))


@sales_bp.route('/assets/<int:asset_id>/dispose', methods=['POST'])
@login_required
def dispose_asset(asset_id):
    """Retire an asset. Any sale proceeds are posted to the finance ledger as
    'Asset Disposal' income (idempotent per asset)."""
    a = db.get_or_404(FixedAsset, asset_id)
    if not can_access_branch(a.branch_id):
        return _err('That asset belongs to another branch.', url_for('sales.assets'))
    if a.is_disposed:
        return _err('That asset is already disposed.', url_for('sales.assets'))
    a.status = 'Disposed'
    a.disposed_on = parse_date(request.form.get('disposed_on')) or timeutil.today()
    a.disposal_amount = request.form.get('disposal_amount', type=float) or 0
    a.disposal_note = (request.form.get('disposal_note') or '').strip() or None
    db.session.commit()
    if a.disposal_amount and a.disposal_amount > 0:
        from utils import finance_ledger as _fl
        _fl.post(_fl.REVENUE, a.disposal_amount, source_module='assets',
                 category='Asset Disposal', branch_id=a.branch_id,
                 method=(request.form.get('method') or 'Cash'),
                 origin_type='asset_disposal', origin_id=a.id,
                 reference=a.asset_tag or a.name, description=f'Disposal of {a.name}',
                 created_by=session.get('user') or session.get('username') or 'Admin')
        db.session.commit()
    return _ok(f'"{a.name}" marked disposed.', url_for('sales.assets'))


@sales_bp.route('/assets/export')
@login_required
def assets_export():
    rows = scope_query(FixedAsset.query, FixedAsset).order_by(FixedAsset.category, FixedAsset.name).all()
    from openpyxl import Workbook
    wb = Workbook(); ws = wb.active; ws.title = 'Fixed Assets'
    ws.append(['Tag', 'Name', 'Category', 'Serial', 'Acquired', 'Cost', 'Life (yrs)',
               'Accum. Depr.', 'Book Value', 'Custodian', 'Location', 'Status'])
    for a in rows:
        ws.append([a.asset_tag or '', a.name, a.category, a.serial_number or '',
                   a.acquisition_date.isoformat() if a.acquisition_date else '',
                   a.acquisition_cost or 0, a.useful_life_years or '',
                   a.accumulated_depreciation, a.book_value,
                   a.custodian or '', a.location or '', a.status])
    return xlsx_response(wb, 'fixed_assets.xlsx')


# ---------------------------------------------------------------------------
# Reports
# ---------------------------------------------------------------------------

REPORT_KINDS = [
    ('inventory_valuation', 'Inventory Valuation'),
    ('low_stock', 'Low / Out of Stock'),
    ('dead_stock', 'Dead Stock'),
    ('fast_moving', 'Fast-Moving Items'),
    ('product_sales', 'Product Sales'),
    ('profit', 'Profit by Category'),
    ('purchases', 'Purchase Orders'),
    ('suppliers', 'Suppliers'),
    ('stock_movements', 'Stock Movements'),
    ('damaged_returned', 'Damaged / Returned'),
    ('expiry', 'Expiry Tracking'),
]
_REPORT_KEYS = {k for k, _ in REPORT_KINDS}
# Reports keyed off a date range (vs a current-snapshot report).
_PERIOD_REPORTS = {'fast_moving', 'product_sales', 'profit', 'purchases',
                   'stock_movements', 'damaged_returned'}


def _col(key, label, money=False, align=None):
    return {'key': key, 'label': label, 'money': money,
            'align': align or ('right' if money else 'left')}


def _period_items(from_d, to_d):
    sales = (scope_query(Sale.query, Sale)
             .filter(func.date(Sale.created_at) >= from_d,
                     func.date(Sale.created_at) <= to_d).all())
    ids = [s.id for s in sales]
    items = []
    if ids:
        items = (db.session.query(SaleItem, Product)
                 .outerjoin(Product, SaleItem.product_id == Product.id)
                 .filter(SaleItem.sale_id.in_(ids)).all())
    return sales, items


def _sales_report(kind, from_d, to_d, category):
    """Build one report as {title, columns, rows, totals}. Snapshot reports use
    the current catalogue; period reports use the from/to window."""
    from collections import defaultdict
    prods = scope_query(Product.query, Product)
    if category:
        prods = prods.filter(Product.category == category)
    prods = prods.all()
    active = [p for p in prods if p.is_active]

    if kind == 'inventory_valuation':
        rows = [{'name': p.name, 'category': p.category, 'stock_qty': p.stock_qty or 0,
                 'cost_price': p.cost_price or 0, 'stock_value': p.stock_value,
                 'sell_value': round((p.stock_qty or 0) * (p.unit_price or 0), 2)} for p in active]
        rows.sort(key=lambda r: r['stock_value'], reverse=True)
        return {'title': 'Inventory Valuation',
                'columns': [_col('name', 'Product'), _col('category', 'Category'),
                            _col('stock_qty', 'Qty', align='right'), _col('cost_price', 'Cost', money=True),
                            _col('stock_value', 'Stock value', money=True), _col('sell_value', 'Sell value', money=True)],
                'rows': rows,
                'totals': {'stock_value': round(sum(r['stock_value'] for r in rows), 2),
                           'sell_value': round(sum(r['sell_value'] for r in rows), 2)}}

    if kind == 'low_stock':
        rows = [{'name': p.name, 'category': p.category, 'stock_qty': p.stock_qty or 0,
                 'reorder_level': p.reorder_level or 0,
                 'status': 'Out' if p.out_of_stock else 'Low'} for p in active if p.low_stock]
        rows.sort(key=lambda r: r['stock_qty'])
        return {'title': 'Low / Out of Stock',
                'columns': [_col('name', 'Product'), _col('category', 'Category'),
                            _col('stock_qty', 'In stock', align='right'),
                            _col('reorder_level', 'Reorder at', align='right'), _col('status', 'Status')],
                'rows': rows, 'totals': None}

    if kind == 'dead_stock':
        _sales, items = _period_items(from_d, to_d)
        sold_ids = {it.product_id for it, _ in items if it.product_id}
        rows = [{'name': p.name, 'category': p.category, 'stock_qty': p.stock_qty or 0,
                 'stock_value': p.stock_value} for p in active
                if (p.stock_qty or 0) > 0 and p.id not in sold_ids]
        rows.sort(key=lambda r: r['stock_value'], reverse=True)
        return {'title': 'Dead Stock (no sales in range)',
                'columns': [_col('name', 'Product'), _col('category', 'Category'),
                            _col('stock_qty', 'Qty', align='right'), _col('stock_value', 'Tied-up value', money=True)],
                'rows': rows, 'totals': {'stock_value': round(sum(r['stock_value'] for r in rows), 2)}}

    if kind in ('fast_moving', 'product_sales'):
        _sales, items = _period_items(from_d, to_d)
        agg = defaultdict(lambda: {'units': 0, 'revenue': 0.0, 'name': '', 'category': ''})
        for it, p in items:
            if category and (not p or p.category != category):
                continue
            key = it.product_id or f'd:{it.description}'
            agg[key]['units'] += it.quantity or 0
            agg[key]['revenue'] += it.line_total or 0
            agg[key]['name'] = (p.name if p else None) or it.description or 'Unknown'
            agg[key]['category'] = (p.category if p else None) or '—'
        rows = [{'name': v['name'], 'category': v['category'], 'units': v['units'],
                 'revenue': round(v['revenue'], 2)} for v in agg.values()]
        rows.sort(key=lambda r: r['units'] if kind == 'fast_moving' else r['revenue'], reverse=True)
        if kind == 'fast_moving':
            rows = rows[:50]
        return {'title': 'Fast-Moving Items' if kind == 'fast_moving' else 'Product Sales',
                'columns': [_col('name', 'Product'), _col('category', 'Category'),
                            _col('units', 'Units', align='right'), _col('revenue', 'Revenue', money=True)],
                'rows': rows, 'totals': {'units': sum(r['units'] for r in rows),
                                         'revenue': round(sum(r['revenue'] for r in rows), 2)}}

    if kind == 'profit':
        _sales, items = _period_items(from_d, to_d)
        agg = defaultdict(lambda: {'revenue': 0.0, 'cogs': 0.0})
        for it, p in items:
            cat = (p.category if p else None) or 'Other'
            if category and cat != category:
                continue
            agg[cat]['revenue'] += it.line_total or 0
            agg[cat]['cogs'] += (it.quantity or 0) * ((p.cost_price or 0) if p else 0)
        rows = [{'category': c, 'revenue': round(v['revenue'], 2), 'cogs': round(v['cogs'], 2),
                 'profit': round(v['revenue'] - v['cogs'], 2)} for c, v in agg.items()]
        rows.sort(key=lambda r: r['profit'], reverse=True)
        return {'title': 'Profit by Category',
                'columns': [_col('category', 'Category'), _col('revenue', 'Revenue', money=True),
                            _col('cogs', 'Cost of goods', money=True), _col('profit', 'Profit', money=True)],
                'rows': rows, 'totals': {'revenue': round(sum(r['revenue'] for r in rows), 2),
                                         'cogs': round(sum(r['cogs'] for r in rows), 2),
                                         'profit': round(sum(r['profit'] for r in rows), 2)}}

    if kind == 'purchases':
        pos = (scope_query(PurchaseOrder.query, PurchaseOrder)
               .filter(func.date(PurchaseOrder.created_at) >= from_d,
                       func.date(PurchaseOrder.created_at) <= to_d)
               .order_by(PurchaseOrder.created_at.desc()).all())
        rows = [{'po_number': po.po_number, 'supplier': po.supplier.company_name if po.supplier else '—',
                 'status': po.status, 'total': po.total or 0, 'received_value': po.received_value,
                 'created_at': po.created_at.strftime('%d %b %Y') if po.created_at else ''} for po in pos]
        return {'title': 'Purchase Orders',
                'columns': [_col('po_number', 'PO'), _col('supplier', 'Supplier'), _col('status', 'Status'),
                            _col('total', 'Ordered', money=True), _col('received_value', 'Received', money=True),
                            _col('created_at', 'Date')],
                'rows': rows, 'totals': {'total': round(sum(r['total'] for r in rows), 2),
                                         'received_value': round(sum(r['received_value'] for r in rows), 2)}}

    if kind == 'suppliers':
        sups = scope_query(Supplier.query.filter_by(is_active=True), Supplier).order_by(Supplier.company_name).all()
        rows = []
        for s in sups:
            st = _supplier_stats(s.id)
            rows.append({'company_name': s.company_name, 'orders': st['orders'],
                         'received_value': st['received_value'], 'paid': st['paid'],
                         'outstanding': st['outstanding']})
        return {'title': 'Suppliers',
                'columns': [_col('company_name', 'Supplier'), _col('orders', 'Orders', align='right'),
                            _col('received_value', 'Received', money=True), _col('paid', 'Paid', money=True),
                            _col('outstanding', 'Outstanding', money=True)],
                'rows': rows, 'totals': {'outstanding': round(sum(r['outstanding'] for r in rows), 2),
                                         'paid': round(sum(r['paid'] for r in rows), 2)}}

    if kind in ('stock_movements', 'damaged_returned'):
        q = (scope_query(StockMovement.query, StockMovement)
             .filter(func.date(StockMovement.created_at) >= from_d,
                     func.date(StockMovement.created_at) <= to_d))
        if kind == 'damaged_returned':
            q = q.filter(StockMovement.reason.in_(['Damage', 'Theft', 'Loss', 'Expired',
                                                   'Customer Return', 'Supplier Return']))
        moves = q.order_by(StockMovement.created_at.desc()).limit(2000).all()
        name_by = {p.id: p.name for p in prods}
        rows = [{'when': m.created_at.strftime('%d %b %Y') if m.created_at else '',
                 'product': name_by.get(m.product_id, '—'), 'direction': m.direction,
                 'reason': m.reason, 'quantity': m.quantity, 'reference': m.reference or '',
                 'by': m.performed_by or ''} for m in moves]
        return {'title': 'Damaged / Returned' if kind == 'damaged_returned' else 'Stock Movements',
                'columns': [_col('when', 'Date'), _col('product', 'Product'), _col('direction', 'Dir'),
                            _col('reason', 'Reason'), _col('quantity', 'Qty', align='right'),
                            _col('reference', 'Ref'), _col('by', 'By')],
                'rows': rows, 'totals': None}

    if kind == 'expiry':
        today = timeutil.today()
        rows = []
        for p in active:
            if not p.expiry_date or (p.stock_qty or 0) <= 0:
                continue
            days = (p.expiry_date - today).days
            if days > 90:
                continue                       # only surface the next 90 days + expired
            status = ('Expired' if days < 0 else '≤30 days' if days <= 30
                      else '≤60 days' if days <= 60 else '≤90 days')
            rows.append({'name': p.name, 'category': p.category,
                         'expiry_date': p.expiry_date.strftime('%d %b %Y'), 'days_left': days,
                         'stock_qty': p.stock_qty or 0, 'stock_value': p.stock_value, 'status': status})
        rows.sort(key=lambda r: r['days_left'])
        return {'title': 'Expiry Tracking (next 90 days + expired)',
                'columns': [_col('name', 'Product'), _col('category', 'Category'),
                            _col('expiry_date', 'Expires'), _col('days_left', 'Days left', align='right'),
                            _col('stock_qty', 'Qty', align='right'), _col('stock_value', 'Value', money=True),
                            _col('status', 'Status')],
                'rows': rows, 'totals': {'stock_value': round(sum(r['stock_value'] for r in rows), 2)}}

    return {'title': 'Report', 'columns': [], 'rows': [], 'totals': None}


def _expiring_soon_count():
    """Products with stock that are expired or expiring within 30 days."""
    try:
        from datetime import timedelta
        today = timeutil.today()
        horizon = today + timedelta(days=30)
        return scope_query(Product.query.filter(
            Product.is_active == True, Product.stock_qty > 0,
            Product.expiry_date.isnot(None), Product.expiry_date <= horizon), Product).count()
    except Exception:
        return 0


def _report_range():
    from datetime import timedelta
    to_d = parse_date(request.args.get('to')) or timeutil.today()
    from_d = parse_date(request.args.get('from')) or (to_d - timedelta(days=29))
    if from_d > to_d:
        from_d, to_d = to_d, from_d
    return from_d, to_d, (request.args.get('category') or '').strip()


@sales_bp.route('/reports')
@login_required
def reports():
    kind = request.args.get('kind') or 'inventory_valuation'
    if kind not in _REPORT_KEYS:
        kind = 'inventory_valuation'
    from_d, to_d, category = _report_range()
    report = _sales_report(kind, from_d, to_d, category)
    return _render({
        'page': 'reports', 'kind': kind,
        'report_kinds': [{'key': k, 'label': lbl, 'period': k in _PERIOD_REPORTS} for k, lbl in REPORT_KINDS],
        'is_period': kind in _PERIOD_REPORTS,
        'from': from_d.isoformat(), 'to': to_d.isoformat(), 'category': category,
        'categories': PRODUCT_CATEGORIES, 'report': report,
        'self_url': url_for('sales.reports'), 'export_url': url_for('sales.reports_export'),
        'urls': {'dashboard': url_for('sales.dashboard')},
    })


@sales_bp.route('/reports/export')
@login_required
def reports_export():
    kind = request.args.get('kind') or 'inventory_valuation'
    if kind not in _REPORT_KEYS:
        kind = 'inventory_valuation'
    fmt = (request.args.get('format') or 'csv').lower()
    from_d, to_d, category = _report_range()
    report = _sales_report(kind, from_d, to_d, category)
    cols = report['columns']
    headers = [c['label'] for c in cols]

    def _rec(r):
        return [r.get(c['key'], '') for c in cols]

    fname = f'sales_report_{kind}'
    if fmt in ('excel', 'xlsx'):
        from openpyxl import Workbook
        wb = Workbook(); ws = wb.active; ws.title = 'Report'
        ws.append(headers)
        for r in report['rows']:
            ws.append(_rec(r))
        return xlsx_response(wb, f'{fname}.xlsx')
    import csv
    import io
    from flask import Response
    out = io.StringIO(); w = csv.writer(out)
    w.writerow(headers)
    for r in report['rows']:
        w.writerow(_rec(r))
    return csv_response(out.getvalue(), f'{fname}.csv')


@sales_bp.route('/receipt/<int:sale_id>')
@login_required
def receipt(sale_id):
    sale = db.get_or_404(Sale, sale_id)
    if not can_access_branch(sale.branch_id):
        flash('That sale belongs to another branch.', 'error')
        return redirect(url_for('sales.history'))
    return render_template('sales/receipt.html', sale=sale,
                           items=sale.items.all())
