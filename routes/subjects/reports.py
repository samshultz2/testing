"""subjects blueprint — reports routes (split from the former routes/subjects.py)."""
from routes.subjects import *  # noqa: F401,F403
from utils.security import strip_tags


@subjects_bp.route('/broadsheet')
@login_required
def broadsheet():
    """View broadsheet (all scores for a class)"""
    term_id = request.args.get('term_id', type=int)
    assignment_id = request.args.get('assignment_id', type=int)
    
    # Check class access
    if assignment_id and not can_access_class(assignment_id):
        flash('You do not have access to this class.', 'error')
        return redirect(url_for('subjects.broadsheet'))
    
    terms = session_terms()
    
    if not term_id:
        active_term = get_active_term()
        if active_term:
            term_id = active_term.id
    
    selected_term = db.session.get(Term, term_id) if term_id else None
    
    # Filter assignments for teachers. SSS3 write no internal exams in third term
    # (only WAEC/NECO/JAMB), so third-term internal broadsheets exclude SSS3.
    assignments = []
    if term_id:
        all_assignments = ClassArmAssignment.query.filter_by(term_id=term_id).all()
        assignments = strip_sss3_third_term(filter_classes_for_user(all_assignments), term_id)

    selected_assignment = db.session.get(ClassArmAssignment, assignment_id) if assignment_id else None
    
    # Build broadsheet data
    broadsheet_data = []
    class_subjects = []
    assessment_types = AssessmentType.query.filter_by(is_active=True).order_by(AssessmentType.order).all()
    
    if selected_assignment:
        # Get subjects for this class
        class_subjects = ClassSubject.query.filter_by(
            term_id=term_id,
            class_id=selected_assignment.class_id,
            is_active=True
        ).filter(
            (ClassSubject.arm_id == None) | (ClassSubject.arm_id == selected_assignment.arm_id)
        ).join(Subject).order_by(Subject.name).all()
        
        # Get enrolled students (eager-load the student to avoid a lazy load each)
        from sqlalchemy.orm import joinedload
        enrollments = (StudentEnrollment.query
                       .options(joinedload(StudentEnrollment.student))
                       .filter_by(class_arm_assignment_id=assignment_id, is_active=True)
                       .join(Student).order_by(Student.surname, Student.first_name).all())

        pass_mark = SchoolSettings.get('pass_mark', 50)

        # Every score for the whole class in one query, indexed by (student,
        # class-subject, assessment) — replaces a query per student × subject.
        cs_ids = [cs.id for cs in class_subjects]
        sids = [e.student_id for e in enrollments]
        score_map = {}
        if cs_ids and sids:
            for s in StudentScore.query.filter(
                    StudentScore.student_id.in_(sids),
                    StudentScore.class_subject_id.in_(cs_ids)).all():
                score_map[(s.student_id, s.class_subject_id, s.assessment_type_id)] = s.score

        for enrollment in enrollments:
            student_row = {
                'student': enrollment.student,
                'subjects': {},
                'total': 0,
                'average': 0,
                'subjects_passed': 0,
                'subjects_failed': 0
            }

            for cs in class_subjects:
                subject_data = {'assessments': {}, 'total': 0}

                subject_total = 0
                for at in assessment_types:
                    score = score_map.get((enrollment.student_id, cs.id, at.id))
                    subject_data['assessments'][at.id] = score
                    if score:
                        subject_total += score
                
                subject_data['total'] = subject_total
                subject_data['grade'] = GradeScale.get_grade(subject_total) if subject_total else '-'
                
                if subject_total >= pass_mark:
                    student_row['subjects_passed'] += 1
                elif subject_total > 0:
                    student_row['subjects_failed'] += 1
                
                student_row['subjects'][cs.id] = subject_data
                student_row['total'] += subject_total
            
            if class_subjects:
                student_row['average'] = round(student_row['total'] / len(class_subjects), 2)
            
            broadsheet_data.append(student_row)
        
        # Sort by average (descending) for ranking
        broadsheet_data.sort(key=lambda x: x['average'], reverse=True)
        
        # Add positions
        for i, row in enumerate(broadsheet_data):
            row['position'] = i + 1
    
    return _render({
        'page': 'broadsheet', 'nav': _nav_urls(),
        'term_id': term_id or '', 'assignment_id': assignment_id or '',
        'selected_assignment': selected_assignment.display_name if selected_assignment else '',
        'has_selection': bool(selected_assignment),
        'terms': [{'id': t.id, 'full_name': t.full_name} for t in terms],
        'assignments': [{'id': a.id, 'display_name': a.display_name} for a in assignments],
        'class_subjects': [{'id': cs.id, 'short': cs.subject.short_name or cs.subject.name[:3],
                            'name': cs.subject.name} for cs in class_subjects],
        'rows': [{'position': r['position'], 'student': r['student'].full_name,
                  'subjects': {str(cs.id): (round(r['subjects'].get(cs.id, {}).get('total', 0), 1)
                                            if r['subjects'].get(cs.id, {}).get('total') else None)
                               for cs in class_subjects},
                  'total': round(r['total'], 1), 'average': r['average'],
                  'passed': r['subjects_passed'], 'failed': r['subjects_failed']}
                 for r in broadsheet_data],
        'self_url': url_for('subjects.broadsheet'),
        'urls': {'compute': url_for('subjects.compute_summaries'),
                 'bulk_entry': url_for('subjects.bulk_entry', term_id=term_id or '', assignment_id=assignment_id or ''),
                 'affective': url_for('subjects.affective', term_id=term_id or '', assignment_id=assignment_id or ''),
                 'comments': url_for('subjects.comments', term_id=term_id or '', assignment_id=assignment_id or ''),
                 'export': url_for('subjects.export_broadsheet', term_id=term_id or '', assignment_id=assignment_id or ''),
                 'export_pdf': url_for('subjects.export_broadsheet', term_id=term_id or '', assignment_id=assignment_id or '', format='pdf'),
                 'export_word': url_for('subjects.export_broadsheet', term_id=term_id or '', assignment_id=assignment_id or '', format='word'),
                 'export_image': url_for('subjects.export_broadsheet', term_id=term_id or '', assignment_id=assignment_id or '', format='image'),
                 'blank_sheet': url_for('subjects.blank_score_sheet', term_id=term_id or '', assignment_id=assignment_id or ''),
                 'scores': url_for('subjects.scores_entry', term_id=term_id or '', assignment_id=assignment_id or ''),
                 'analytics': url_for('subjects.analytics_dashboard', term_id=term_id or '', assignment_id=assignment_id or ''),
                 'explore': url_for('subjects.broadsheet_explore', term_id=term_id or '', scopes=assignment_id or '')},
    })


def _explore_dataset(term_id, scope_ids):
    """Shared cross-class / cross-arm dataset for the Results Explorer.

    Returns everything the view *and* the exports need, keyed by the underlying
    Subject so the same subject lines up across classes. Access is re-checked per
    scope (``filter_classes_for_user``), so a teacher only ever pulls their own
    classes even if they hand-craft the ``scopes`` query string.
    """
    from collections import OrderedDict
    from sqlalchemy.orm import joinedload

    scope_options = []
    assignments_by_id = {}
    allowed_ids = set()
    if term_id:
        all_assignments = ClassArmAssignment.query.filter_by(term_id=term_id).all()
        # SSS3 have no internal results in third term (only WAEC/NECO/JAMB), so
        # the explorer/combination datasets omit SSS3 arms in third term.
        allowed = strip_sss3_third_term(filter_classes_for_user(all_assignments), term_id)
        allowed_ids = {a.id for a in allowed}
        by_class = OrderedDict()
        for a in sorted(allowed, key=lambda x: (
                (x.school_class.level if x.school_class else 0),
                (x.school_class.name if x.school_class else ''),
                (x.arm.name if x.arm else ''))):
            assignments_by_id[a.id] = a
            if a.class_id not in by_class:
                by_class[a.class_id] = {
                    'class_id': a.class_id,
                    'class_name': a.school_class.name if a.school_class else '',
                    'arms': []}
            by_class[a.class_id]['arms'].append(
                {'assignment_id': a.id,
                 'arm_name': (a.arm.name if a.arm else ''),
                 'label': a.display_name})
        scope_options = list(by_class.values())

    scope_ids = [i for i in scope_ids if i in allowed_ids]
    selected = [assignments_by_id[i] for i in scope_ids]

    pass_mark = SchoolSettings.get('pass_mark', 50)
    at_ids = [at.id for at in AssessmentType.query.filter_by(is_active=True).all()]

    subjects_union = {}
    scope_meta = []
    rows = []

    if selected:
        class_ids = list({a.class_id for a in selected})
        cs_rows = (ClassSubject.query
                   .filter(ClassSubject.term_id == term_id,
                           ClassSubject.class_id.in_(class_ids),
                           ClassSubject.is_active == True)  # noqa: E712
                   .join(Subject).all())
        cs_by_class = {}
        cs_subject = {}
        for cs in cs_rows:
            cs_by_class.setdefault(cs.class_id, []).append(cs)
            cs_subject[cs.id] = cs.subject

        def subjects_for(asg):
            return [cs for cs in cs_by_class.get(asg.class_id, [])
                    if cs.arm_id is None or cs.arm_id == asg.arm_id]

        asg_ids = [a.id for a in selected]
        enrollments = (StudentEnrollment.query
                       .options(joinedload(StudentEnrollment.student))
                       .filter(StudentEnrollment.class_arm_assignment_id.in_(asg_ids),
                               StudentEnrollment.is_active == True)  # noqa: E712
                       .join(Student).order_by(Student.surname, Student.first_name).all())
        student_ids = [e.student_id for e in enrollments]

        all_cs_ids = [cs.id for cs in cs_rows]
        score_map = {}
        comp_map = {}            # (student_id, cs_id, at_id) -> component score
        if student_ids and all_cs_ids and at_ids:
            for s in StudentScore.query.filter(
                    StudentScore.student_id.in_(student_ids),
                    StudentScore.class_subject_id.in_(all_cs_ids),
                    StudentScore.assessment_type_id.in_(at_ids)).all():
                key = (s.student_id, s.class_subject_id)
                score_map[key] = score_map.get(key, 0) + (s.score or 0)
                comp_map[(s.student_id, s.class_subject_id, s.assessment_type_id)] = (s.score or 0)

        for a in selected:
            for cs in subjects_for(a):
                subj = cs_subject[cs.id]
                subjects_union.setdefault(subj.id, {
                    'id': subj.id, 'name': subj.name,
                    'short': subj.short_name or subj.name[:3]})
            scope_meta.append({
                'assignment_id': a.id, 'class_id': a.class_id,
                'class_name': a.school_class.name if a.school_class else '',
                'arm_name': a.arm.name if a.arm else '',
                'label': a.display_name})

        for e in enrollments:
            a = assignments_by_id[e.class_arm_assignment_id]
            asg_cs = subjects_for(a)
            subj_totals = {}
            components = {}          # {subject_id: {at_id: score}} for component filters
            total = 0.0
            passed = failed = 0
            for cs in asg_cs:
                st = score_map.get((e.student_id, cs.id))
                if st is None:
                    continue
                sid_key = str(cs_subject[cs.id].id)
                subj_totals[sid_key] = round(st, 1)
                comp = {str(at): round(comp_map[(e.student_id, cs.id, at)], 1)
                        for at in at_ids if (e.student_id, cs.id, at) in comp_map}
                if comp:
                    components[sid_key] = comp
                total += st
                if st >= pass_mark:
                    passed += 1
                elif st > 0:
                    failed += 1
            avg = round(total / len(asg_cs), 2) if asg_cs else 0
            rows.append({
                'student': e.student.full_name,
                'assignment_id': a.id,
                'class_name': a.school_class.name if a.school_class else '',
                'arm_name': a.arm.name if a.arm else '',
                'scope_label': a.display_name,
                'subjects': subj_totals,
                'components': components,
                'total': round(total, 1), 'average': avg,
                'passed': passed, 'failed': failed})
        rows.sort(key=lambda r: r['average'], reverse=True)

    assessment_types = [{'id': at.id, 'name': at.name}
                        for at in AssessmentType.query.filter_by(is_active=True)
                        .order_by(AssessmentType.order).all()]

    return {
        'scope_options': scope_options,
        'selected_ids': scope_ids,
        'scope_meta': scope_meta,
        'subjects': sorted(subjects_union.values(), key=lambda s: s['name']),
        'assessment_types': assessment_types,
        'rows': rows,
        'pass_mark': pass_mark,
    }


def _explore_filter(rows, field, op, v1, v2):
    """Apply the Explorer's field × condition × value(s) filter server-side, so an
    export mirrors exactly what the user sees. ``field`` is 'average', 'total' or
    a subject id; ``op`` is gte|lte|eq|between. Returns (filtered_rows, active,
    label)."""
    def val(r):
        if field == 'average':
            return r['average']
        if field == 'total':
            return r['total']
        return r['subjects'].get(str(field))

    active = ((op == 'between' and v1 is not None and v2 is not None)
              or (op != 'between' and v1 is not None))
    if not active:
        return rows, False, ''

    def ok(r):
        x = val(r)
        if x is None:
            return False
        if op == 'gte':
            return x >= v1
        if op == 'lte':
            return x <= v1
        if op == 'eq':
            return x == v1
        if op == 'between':
            return min(v1, v2) <= x <= max(v1, v2)
        return True

    op_text = {'gte': '≥', 'lte': '≤', 'eq': '=', 'between': 'between'}.get(op, '')
    rng = f'{min(v1, v2):g}–{max(v1, v2):g}' if op == 'between' else f'{v1:g}'
    return [r for r in rows if ok(r)], True, f'{op_text} {rng}'


def _explore_scope_ids():
    raw = request.args.get('scopes', '') or ''
    return [int(x) for x in raw.replace(' ', '').split(',') if x.isdigit()]


@subjects_bp.route('/broadsheet/explore')
@login_required
def broadsheet_explore():
    """Cross-class / cross-arm results explorer (the interactive screen).

    Query params:
      term_id  — the term.
      scopes   — comma-separated ClassArmAssignment ids to include.
    """
    term_id = request.args.get('term_id', type=int)
    if not term_id:
        active_term = get_active_term()
        if active_term:
            term_id = active_term.id
    ds = _explore_dataset(term_id, _explore_scope_ids())
    terms = session_terms()
    return _render({
        'page': 'explore', 'nav': _nav_urls(),
        'term_id': term_id or '', 'scopes': ds['selected_ids'],
        'terms': [{'id': t.id, 'full_name': t.full_name} for t in terms],
        'scope_options': ds['scope_options'],
        'scope_meta': ds['scope_meta'],
        'subjects_union': ds['subjects'],
        'rows': ds['rows'], 'pass_mark': ds['pass_mark'],
        'self_url': url_for('subjects.broadsheet_explore'),
        'urls': {'broadsheet': url_for('subjects.broadsheet', term_id=term_id or ''),
                 'combine': url_for('subjects.broadsheet_combine', term_id=term_id or ''),
                 'export': url_for('subjects.export_explore')},
    })


@subjects_bp.route('/broadsheet/explore/export')
@login_required
def export_explore():
    """Export the filtered cross-class Explorer view — Excel or a combined,
    print-ready PDF (``format`` = excel (default) | pdf). Filter params
    (field/op/v1/v2) mirror the on-screen condition so the file matches the view.
    """
    term_id = request.args.get('term_id', type=int)
    if not term_id:
        active_term = get_active_term()
        if active_term:
            term_id = active_term.id
    ds = _explore_dataset(term_id, _explore_scope_ids())
    if not ds['rows']:
        flash('Select a term and at least one class arm with scores to export.', 'error')
        return redirect(url_for('subjects.broadsheet_explore', term_id=term_id or ''))

    field = request.args.get('field', 'average')
    op = request.args.get('op', 'gte')
    v1 = request.args.get('v1', type=float)
    v2 = request.args.get('v2', type=float)
    rows, active, cond = _explore_filter(ds['rows'], field, op, v1, v2)
    field_label = ('Average (%)' if field == 'average' else 'Total' if field == 'total'
                   else next((s['name'] for s in ds['subjects'] if str(s['id']) == str(field)), 'Subject'))
    filter_label = f'{field_label} {cond}' if active else ''

    term = db.session.get(Term, term_id) if term_id else None
    term_name = term.full_name if term else ''
    fmt = (request.args.get('format') or 'excel').lower()
    from utils import broadsheet_export as bx
    if fmt == 'pdf':
        data = bx.explore_pdf(ds['subjects'], ds['scope_meta'], rows, term_name,
                              pass_mark=ds['pass_mark'], filter_label=filter_label)
        from flask import Response
        return Response(data, mimetype='application/pdf', headers={
            'Content-Disposition': 'attachment; filename="results_explorer.pdf"'})
    wb = bx.explore_xlsx(ds['subjects'], ds['scope_meta'], rows, term_name,
                         pass_mark=ds['pass_mark'], filter_label=filter_label)
    return xlsx_response(wb, 'results_explorer.xlsx')


def _combine_rows(ds, subject_ids, metric, op, value):
    """Compute a combined total/average over just the chosen subjects for each
    student, then filter by the condition. Missing subject scores count as 0 and
    the average divides by the number of chosen subjects (matches the literal
    '(a+b+c)/3'). Returns (rows, active, label)."""
    sids = [str(s) for s in subject_ids]
    n = len(sids)
    out = []
    for r in ds['rows']:
        present = [r['subjects'].get(s) for s in sids]
        got = [p for p in present if p is not None]
        total = round(sum(got), 1)
        avg = round(total / n, 2) if n else 0
        row = dict(r)
        row['combo_total'] = total
        row['combo_average'] = avg
        row['combo_missing'] = n - len(got)
        out.append(row)

    active = value is not None and n > 0
    if active:
        key = 'combo_average' if metric == 'average' else 'combo_total'

        def ok(r):
            x = r[key]
            if op == 'gte':
                return x >= value
            if op == 'lte':
                return x <= value
            if op == 'eq':
                return x == value
            return True
        out = [r for r in out if ok(r)]

    op_text = {'gte': '≥', 'lte': '≤', 'eq': '='}.get(op, '')
    metric_text = 'Average' if metric == 'average' else 'Total'
    label = f'{metric_text} {op_text} {value:g}' if active else ''
    # Sort by the combined metric, best first.
    out.sort(key=lambda r: (r['combo_average'] if metric == 'average' else r['combo_total']), reverse=True)
    return out, active, label


def _cond_value(r, subject_ids, basis, component):
    """The value a single condition tests for one student row.

    basis: all_total | all_average | combo_total | combo_average.
    component: '' (whole-subject total) or an assessment-type id (e.g. Exam) — it
    only applies to the two ``combo_*`` bases. Missing subject scores count as 0;
    the combined average divides by the number of chosen subjects.
    """
    if basis == 'all_total':
        return r.get('total', 0)
    if basis == 'all_average':
        return r.get('average', 0)
    sids = [str(s) for s in subject_ids]
    n = len(sids)
    if not n:
        return None
    tot = 0.0
    for sid in sids:
        if component:
            v = (r.get('components', {}).get(sid, {}) or {}).get(str(component))
        else:
            v = r.get('subjects', {}).get(sid)
        tot += (v or 0)
    tot = round(tot, 1)
    return tot if basis == 'combo_total' else round(tot / n, 2)


def _combine_multi(rows_in, subject_ids, conditions, at_names=None):
    """Filter rows by ANY number of AND-combined conditions. Each condition is
    ``{basis, component, op, value}``. Returns ``(rows, active, labels)`` where
    labels is a human list like ['All-subject average ≥ 50', 'Exam average ≥ 67']."""
    at_names = at_names or {}
    valid = []
    for c in conditions:
        try:
            v = float(c.get('value'))
        except (TypeError, ValueError):
            continue
        basis = c.get('basis') or 'combo_total'
        op = c.get('op') or 'gte'
        comp = str(c.get('component') or '')
        if basis in ('all_total', 'all_average'):
            comp = ''
        valid.append({'basis': basis, 'op': op, 'value': v, 'component': comp})

    def passes(r, c):
        x = _cond_value(r, subject_ids, c['basis'], c['component'])
        if x is None:
            return False
        if c['op'] == 'gte':
            return x >= c['value']
        if c['op'] == 'lte':
            return x <= c['value']
        if c['op'] == 'eq':
            return x == c['value']
        return True

    rows = rows_in
    if valid:
        rows = [r for r in rows if all(passes(r, c) for c in valid)]

    op_text = {'gte': '≥', 'lte': '≤', 'eq': '='}
    basis_text = {'all_total': 'All-subject total', 'all_average': 'All-subject average',
                  'combo_total': 'Combined total', 'combo_average': 'Combined average'}
    labels = []
    for c in valid:
        name = basis_text.get(c['basis'], c['basis'])
        if c['component'] and c['basis'].startswith('combo'):
            comp_name = at_names.get(str(c['component']), 'component')
            name += f' ({comp_name})'
        labels.append(f"{name} {op_text.get(c['op'], '')} {c['value']:g}")
    return rows, bool(valid), labels


@subjects_bp.route('/broadsheet/combine')
@login_required
def broadsheet_combine():
    """Subject-combination explorer: pick any set of subjects and filter students
    by their combined total/average across just those subjects."""
    term_id = request.args.get('term_id', type=int)
    if not term_id:
        active_term = get_active_term()
        if active_term:
            term_id = active_term.id
    ds = _explore_dataset(term_id, _explore_scope_ids())
    terms = session_terms()
    return _render({
        'page': 'combine', 'nav': _nav_urls(),
        'term_id': term_id or '', 'scopes': ds['selected_ids'],
        'terms': [{'id': t.id, 'full_name': t.full_name} for t in terms],
        'scope_options': ds['scope_options'],
        'scope_meta': ds['scope_meta'],
        'subjects_union': ds['subjects'],
        'assessment_types': ds['assessment_types'],
        'rows': ds['rows'], 'pass_mark': ds['pass_mark'],
        'self_url': url_for('subjects.broadsheet_combine'),
        'urls': {'broadsheet': url_for('subjects.broadsheet', term_id=term_id or ''),
                 'explore': url_for('subjects.broadsheet_explore', term_id=term_id or ''),
                 'export': url_for('subjects.combine_export')},
    })


@subjects_bp.route('/broadsheet/combine/export')
@login_required
def combine_export():
    """Export the subject-combination view (pdf | image | excel | csv) with a
    caller-chosen column set. Params mirror the on-screen tool so the file
    matches exactly what the user sees."""
    term_id = request.args.get('term_id', type=int)
    if not term_id:
        active_term = get_active_term()
        if active_term:
            term_id = active_term.id
    ds = _explore_dataset(term_id, _explore_scope_ids())

    subj_raw = request.args.get('subjects', '') or ''
    subject_ids = [x for x in subj_raw.replace(' ', '').split(',') if x.isdigit()]
    subj_lookup = {str(s['id']): s for s in ds['subjects']}
    chosen_subjects = [subj_lookup[s] for s in subject_ids if s in subj_lookup]

    metric = request.args.get('metric', 'total')
    op = request.args.get('op', 'gte')
    value = request.args.get('value', type=float)
    # All rows with the combined columns computed (no single-metric filtering).
    all_combo, _, _ = _combine_rows(ds, subject_ids, metric, op, None)

    # Multi-condition (AND) filter — the primary path. Falls back to the legacy
    # single metric/op/value when no `conditions` JSON is supplied.
    import json as _json
    conditions = []
    raw_cond = request.args.get('conditions')
    if raw_cond:
        try:
            parsed = _json.loads(raw_cond)
            if isinstance(parsed, list):
                conditions = parsed
        except (ValueError, TypeError):
            conditions = []
    if conditions:
        at_names = {str(at['id']): at['name'] for at in ds['assessment_types']}
        rows, active, labels = _combine_multi(all_combo, subject_ids, conditions, at_names)
        cond = '; '.join(labels)
    else:
        rows, active, cond = _combine_rows(ds, subject_ids, metric, op, value)

    if not chosen_subjects:
        flash('Pick at least one subject to combine before exporting.', 'error')
        return redirect(url_for('subjects.broadsheet_combine', term_id=term_id or '',
                                scopes=','.join(str(i) for i in ds['selected_ids'])))

    # Column selection from the modal. Keys: sn, student, class, arm,
    # gender is not carried; subj:<id>, total, average, missing.
    cols_raw = request.args.get('columns', '') or ''
    keys = [k for k in cols_raw.split(',') if k]
    if not keys:
        keys = ['sn', 'student', 'class', 'arm'] + [f'subj:{s["id"]}' for s in chosen_subjects] + ['total', 'average']

    def header_for(k):
        if k == 'sn':
            return 'S/N'
        if k == 'student':
            return 'Student'
        if k == 'class':
            return 'Class'
        if k == 'arm':
            return 'Arm'
        if k == 'total':
            return 'Combined total'
        if k == 'average':
            return 'Combined average'
        if k == 'missing':
            return 'Missing'
        if k.startswith('subj:'):
            s = subj_lookup.get(k.split(':', 1)[1])
            return s['name'] if s else k
        return k

    from utils.numfmt import fmt_num as _n

    def value_for(k, r, i):
        if k == 'sn':
            return str(i)
        if k == 'student':
            return r['student']
        if k == 'class':
            return r['class_name']
        if k == 'arm':
            return r['arm_name']
        if k == 'total':
            return _n(r['combo_total'])
        if k == 'average':
            return _n(r['combo_average'])
        if k == 'missing':
            return str(r['combo_missing'])
        if k.startswith('subj:'):
            v = r['subjects'].get(k.split(':', 1)[1])
            return _n(v) if v is not None else '–'
        return ''

    headers = [header_for(k) for k in keys]
    data_rows = [[value_for(k, r, i) for k in keys] for i, r in enumerate(rows, 1)]

    term = db.session.get(Term, term_id) if term_id else None
    scope_names = ', '.join(m['label'] for m in ds['scope_meta']) or '—'
    subj_names = ', '.join(s['name'] for s in chosen_subjects)
    bits = [subj_names, scope_names]
    if term:
        bits.append(term.full_name)
    if active:
        bits.append('Filter: ' + cond)
    subtitle = 'Subject Combination · ' + ' · '.join(bits)
    title = 'Subject Combination Results'

    fmt = (request.args.get('format') or 'pdf').lower()
    from utils import broadsheet_export as bx
    from flask import Response
    if fmt in ('image', 'png'):
        data = bx.combo_png(headers, data_rows, title, subtitle)
        return Response(data, mimetype='image/png', headers={
            'Content-Disposition': 'attachment; filename="subject_combination.png"'})
    if fmt in ('excel', 'xlsx'):
        wb = bx.combo_xlsx(headers, data_rows, title, subtitle)
        return xlsx_response(wb, 'subject_combination.xlsx')
    if fmt == 'csv':
        import csv as _csv
        from io import StringIO
        buf = StringIO(); w = _csv.writer(buf); w.writerow(headers)
        for dr in data_rows:
            w.writerow(dr)
        return Response(buf.getvalue(), mimetype='text/csv', headers={
            'Content-Disposition': 'attachment; filename="subject_combination.csv"'})
    data = bx.combo_pdf(headers, data_rows, title, subtitle, numeric_from=_combo_numeric_from(keys))
    return Response(data, mimetype='application/pdf', headers={
        'Content-Disposition': 'attachment; filename="subject_combination.pdf"'})


def _combo_numeric_from(keys):
    """Index of the first numeric column (everything up to and including the last
    of student/class/arm stays left-aligned)."""
    left = {'sn', 'student', 'class', 'arm'}
    idx = 0
    for i, k in enumerate(keys):
        if k in left:
            idx = i + 1
    return max(idx, 1)


@subjects_bp.route('/broadsheet/compute', methods=['POST'])
@login_required
def compute_summaries():
    """Compute & persist term results + class/arm positions for a class."""
    term_id = request.form.get('term_id', type=int)
    assignment_id = request.form.get('assignment_id', type=int)
    asg = db.session.get(ClassArmAssignment, assignment_id) if assignment_id else None
    if not (term_id and asg):
        return _err('Select a term and class first.', url_for('subjects.broadsheet'))
    from utils.report_card import compute_term_summaries
    from utils.audit import log_action
    count = compute_term_summaries(term_id, asg.class_id)
    log_action('results.compute_summaries',
               detail=f'term {term_id}, class {asg.class_id}: {count} student(s)')
    return _ok(f'Computed results and positions for {count} student(s).',
               url_for('subjects.broadsheet', term_id=term_id, assignment_id=assignment_id))


@subjects_bp.route('/analytics')
@login_required
def analytics_dashboard():
    """Academic analytics for a class in a term — grade distribution, subject
    difficulty, pass/fail rate, intervention list and a term-on-term trend.
    Computed from entered scores (cached; ?refresh=1 forces recomputation)."""
    term_id = request.args.get('term_id', type=int)
    assignment_id = request.args.get('assignment_id', type=int)
    if not term_id:
        active = get_active_term()
        term_id = active.id if active else None
    if assignment_id and not can_access_class(assignment_id):
        flash('You do not have access to this class.', 'error')
        return redirect(url_for('subjects.analytics_dashboard'))
    terms = session_terms()
    assignments = (strip_sss3_third_term(filter_classes_for_user(
        ClassArmAssignment.query.filter_by(term_id=term_id).all()), term_id) if term_id else [])
    data = None
    if term_id and assignment_id:
        from utils.results_analytics import class_analytics
        data = class_analytics(term_id, assignment_id,
                               use_cache=(request.args.get('refresh') != '1'))
    return _render({
        'page': 'analytics', 'nav': _nav_urls(),
        'term_id': term_id or '', 'assignment_id': assignment_id or '',
        'terms': [{'id': t.id, 'full_name': t.full_name} for t in terms],
        'assignments': [{'id': a.id, 'display_name': a.display_name} for a in assignments],
        'has_selection': bool(term_id and assignment_id),
        'analytics': data,
        'self_url': url_for('subjects.analytics_dashboard'),
        'refresh_url': url_for('subjects.analytics_dashboard', term_id=term_id or '',
                               assignment_id=assignment_id or '', refresh=1),
        'report_card_base': url_for('subjects.student_report_card', student_id=0)[:-1],
        'urls': {'broadsheet': url_for('subjects.broadsheet', term_id=term_id or '', assignment_id=assignment_id or ''),
                 'scores': url_for('subjects.scores_entry', term_id=term_id or '', assignment_id=assignment_id or ''),
                 'report_pdf': url_for('subjects.analytics_report', term_id=term_id or '', assignment_id=assignment_id or ''),
                 'institution': url_for('subjects.institution_analytics', term_id=term_id or '')},
    })


@subjects_bp.route('/analytics/report.pdf')
@login_required
def analytics_report():
    """Download the class analytics as a formatted PDF report."""
    from utils.broadsheet_export import analytics_pdf, analytics_filename
    from utils.web_exports import pdf_response

    term_id = request.args.get('term_id', type=int)
    assignment_id = request.args.get('assignment_id', type=int)
    if not term_id or not assignment_id:
        flash('Select term and class first.', 'error')
        return redirect(url_for('subjects.analytics_dashboard'))
    if not can_access_class(assignment_id):
        flash('You do not have access to this class.', 'error')
        return redirect(url_for('subjects.analytics_dashboard'))
    term = db.session.get(Term, term_id)
    asg = db.session.get(ClassArmAssignment, assignment_id)
    if not (term and asg):
        flash('Invalid selection.', 'error')
        return redirect(url_for('subjects.analytics_dashboard'))
    data = analytics_pdf(term_id, assignment_id)
    if not data:
        flash('No scores entered for this class yet.', 'warning')
        return redirect(url_for('subjects.analytics_dashboard', term_id=term_id, assignment_id=assignment_id))
    return pdf_response(data, analytics_filename(asg, term), inline=False)


def _org_allowed_ids(term_id):
    """Assignment ids the current user may roll up for a term — None for admins
    (meaning 'everything', the fast path)."""
    if is_admin():
        return None
    asgs = strip_sss3_third_term(filter_classes_for_user(
        ClassArmAssignment.query.filter_by(term_id=term_id).all()), term_id) if term_id else []
    return {a.id for a in asgs}


@subjects_bp.route('/analytics/institution')
@login_required
def institution_analytics():
    """Institution-wide academic analytics — roll the term's entered scores up to
    an arm, a class, a section or the whole school, with subject & teacher
    leagues and decision-oriented recommendations. Cached (?refresh=1 forces)."""
    term_id = request.args.get('term_id', type=int)
    scope = request.args.get('scope') or 'school'
    scope_id = request.args.get('scope_id')
    if not term_id:
        active = get_active_term()
        term_id = active.id if active else None
    # An arm/class scope is guarded to the caller's access.
    if scope == 'arm' and scope_id and not can_access_class(int(scope_id)):
        flash('You do not have access to this class.', 'error')
        scope, scope_id = 'school', None
    terms = session_terms()
    data = None
    if term_id:
        from utils.results_analytics_org import org_analytics
        data = org_analytics(term_id, scope, scope_id, _org_allowed_ids(term_id),
                             use_cache=(request.args.get('refresh') != '1'))
    return _render({
        'page': 'institution', 'nav': _nav_urls(), 'is_admin': is_admin(),
        'term_id': term_id or '', 'scope': scope, 'scope_id': scope_id or '',
        'auto_board_pack': bool(SchoolSettings.get('auto_board_pack')),
        'terms': [{'id': t.id, 'full_name': t.full_name} for t in terms],
        'analytics': data,
        'self_url': url_for('subjects.institution_analytics'),
        'report_card_base': url_for('subjects.student_report_card', student_id=0)[:-1],
        'urls': {
            'report_base': url_for('subjects.institution_report',
                                   term_id=term_id or '', scope=scope, scope_id=scope_id or ''),
            'class_analytics_base': url_for('subjects.analytics_dashboard'),
            'teacher_base': url_for('subjects.teacher_scorecard_view',
                                    term_id=term_id or '', scope=scope, scope_id=scope_id or ''),
            'subject_base': url_for('subjects.subject_scorecard_view',
                                    term_id=term_id or '', scope=scope, scope_id=scope_id or ''),
            'compose': url_for('comms.compose'),
            'email_report': url_for('subjects.institution_email'),
            'toggle_auto': url_for('subjects.institution_auto_email'),
        },
    })


@subjects_bp.route('/analytics/institution/email', methods=['POST'])
@login_required
def institution_email():
    """Email the board pack for a scope to the school's owners/admins now."""
    if not is_admin():
        return _err('Only administrators can email the board pack.',
                    url_for('subjects.institution_analytics'))
    from utils.results_notify import deliver_board_pack, board_pack_recipients
    from utils import mailer
    term_id = request.form.get('term_id', type=int)
    scope = request.form.get('scope') or 'school'
    scope_id = request.form.get('scope_id') or None
    if not term_id:
        return _err('Select a term first.', url_for('subjects.institution_analytics'))
    if not mailer.is_configured():
        return _err('Email is not configured — set up SMTP in Settings first.',
                    url_for('subjects.institution_analytics'))
    if not board_pack_recipients():
        return _err('No owner/admin email addresses on file to send to.',
                    url_for('subjects.institution_analytics'))
    res = deliver_board_pack(term_id=term_id, scope=scope, scope_id=scope_id, force=True)
    if res.get('ok'):
        return _ok(f"Board pack emailed to {res['sent']} recipient(s).",
                   url_for('subjects.institution_analytics'))
    reason = {'no scores': 'No scores entered for this scope yet.'}.get(
        res.get('reason'), 'Could not send the board pack.')
    return _err(reason, url_for('subjects.institution_analytics'))


@subjects_bp.route('/analytics/institution/auto-email', methods=['POST'])
@login_required
def institution_auto_email():
    """Toggle the once-per-term automatic board-pack delivery to owners."""
    if not is_admin():
        return _err('Only administrators can change this.',
                    url_for('subjects.institution_analytics'))
    on = (request.form.get('enabled') in ('1', 'true', 'on', 'yes'))
    SchoolSettings.set('auto_board_pack', 'yes' if on else '', 'string',
                       'Auto-email the institution board pack to owners each published term')
    state = 'on' if on else 'off'
    return _ok(f'Automatic term board-pack email is now {state}.',
               url_for('subjects.institution_analytics'))


@subjects_bp.route('/analytics/institution/report.pdf')
@subjects_bp.route('/analytics/institution/report')
@login_required
def institution_report():
    """Institution analytics export for a scope in a term.
    ``format`` = pdf (default) | excel | image."""
    from utils.analytics_org_pdf import (institution_pdf, institution_png,
                                         institution_xlsx, institution_filename)
    from utils.web_exports import pdf_response, xlsx_response, png_response
    from utils.results_analytics_org import org_analytics

    term_id = request.args.get('term_id', type=int)
    scope = request.args.get('scope') or 'school'
    scope_id = request.args.get('scope_id')
    fmt = (request.args.get('format') or 'pdf').lower()
    if not term_id:
        flash('Select a term first.', 'error')
        return redirect(url_for('subjects.institution_analytics'))
    if scope == 'arm' and scope_id and not can_access_class(int(scope_id)):
        flash('You do not have access to this class.', 'error')
        return redirect(url_for('subjects.institution_analytics'))
    data = org_analytics(term_id, scope, scope_id, _org_allowed_ids(term_id))
    if not data or not (data.get('summary') or {}).get('assessed'):
        flash('No scores entered for this scope yet.', 'warning')
        return redirect(url_for('subjects.institution_analytics',
                                term_id=term_id, scope=scope, scope_id=scope_id or ''))
    term = db.session.get(Term, term_id)
    if fmt in ('excel', 'xlsx'):
        return xlsx_response(institution_xlsx(data, term), institution_filename(data, term, 'xlsx'))
    if fmt in ('image', 'png'):
        return png_response(institution_png(data, term), institution_filename(data, term, 'png'), inline=False)
    return pdf_response(institution_pdf(data, term), institution_filename(data, term, 'pdf'), inline=False)


@subjects_bp.route('/analytics/teacher')
@login_required
def teacher_scorecard_view():
    """Per-class, per-subject scorecard for one teacher in a term (the drill-down
    behind the teacher-effectiveness league)."""
    from utils.results_analytics_org import teacher_scorecard
    term_id = request.args.get('term_id', type=int)
    name = (request.args.get('name') or '').strip()
    scope = request.args.get('scope') or 'school'
    scope_id = request.args.get('scope_id')
    if not term_id:
        active = get_active_term()
        term_id = active.id if active else None
    data = None
    if term_id and name:
        data = teacher_scorecard(term_id, name, _org_allowed_ids(term_id))
    terms = session_terms()
    staff_id = (data or {}).get('staff_id') if data else None
    return _render({
        'page': 'teacher', 'nav': _nav_urls(), 'is_admin': is_admin(),
        'term_id': term_id or '', 'teacher_name': name, 'staff_id': staff_id,
        'terms': [{'id': t.id, 'full_name': t.full_name} for t in terms],
        'scorecard': data,
        'self_url': url_for('subjects.teacher_scorecard_view'),
        'back_url': url_for('subjects.institution_analytics', term_id=term_id or '',
                            scope=scope, scope_id=scope_id or ''),
        'urls': {
            'report_base': url_for('subjects.teacher_report', term_id=term_id or '', name=name),
            'compose': url_for('comms.compose'),
        },
    })


@subjects_bp.route('/analytics/teacher/report')
@login_required
def teacher_report():
    """Teacher scorecard export. ``format`` = pdf (default) | excel | image."""
    from utils.analytics_org_pdf import (teacher_pdf, teacher_png, teacher_xlsx,
                                         teacher_filename)
    from utils.web_exports import pdf_response, xlsx_response, png_response
    from utils.results_analytics_org import teacher_scorecard
    term_id = request.args.get('term_id', type=int)
    name = (request.args.get('name') or '').strip()
    fmt = (request.args.get('format') or 'pdf').lower()
    if not term_id or not name:
        flash('Select a term and teacher first.', 'error')
        return redirect(url_for('subjects.institution_analytics'))
    data = teacher_scorecard(term_id, name, _org_allowed_ids(term_id))
    if not data or not (data.get('summary') or {}).get('entries'):
        flash('No scores attributed to that teacher yet.', 'warning')
        return redirect(url_for('subjects.teacher_scorecard_view', term_id=term_id, name=name))
    term = db.session.get(Term, term_id)
    if fmt in ('excel', 'xlsx'):
        return xlsx_response(teacher_xlsx(data, term), teacher_filename(data, term, 'xlsx'))
    if fmt in ('image', 'png'):
        return png_response(teacher_png(data, term), teacher_filename(data, term, 'png'), inline=False)
    return pdf_response(teacher_pdf(data, term), teacher_filename(data, term, 'pdf'), inline=False)


@subjects_bp.route('/analytics/subject')
@login_required
def subject_scorecard_view():
    """Per-class-arm, per-teacher scorecard for one subject in a term (the
    drill-down behind the subject league)."""
    from utils.results_analytics_org import subject_scorecard
    term_id = request.args.get('term_id', type=int)
    subject_id = request.args.get('subject_id', type=int)
    scope = request.args.get('scope') or 'school'
    scope_id = request.args.get('scope_id')
    if not term_id:
        active = get_active_term()
        term_id = active.id if active else None
    data = None
    if term_id and subject_id:
        data = subject_scorecard(term_id, subject_id, _org_allowed_ids(term_id))
    terms = session_terms()
    return _render({
        'page': 'subject', 'nav': _nav_urls(),
        'term_id': term_id or '', 'subject_id': subject_id or '',
        'terms': [{'id': t.id, 'full_name': t.full_name} for t in terms],
        'scorecard': data,
        'self_url': url_for('subjects.subject_scorecard_view'),
        'back_url': url_for('subjects.institution_analytics', term_id=term_id or '',
                            scope=scope, scope_id=scope_id or ''),
        'urls': {'report_base': url_for('subjects.subject_report',
                                        term_id=term_id or '', subject_id=subject_id or '')},
    })


@subjects_bp.route('/analytics/subject/report')
@login_required
def subject_report():
    """Subject scorecard export. ``format`` = pdf (default) | excel | image."""
    from utils.analytics_org_pdf import (subject_pdf, subject_png, subject_xlsx,
                                         subject_filename)
    from utils.web_exports import pdf_response, xlsx_response, png_response
    from utils.results_analytics_org import subject_scorecard
    term_id = request.args.get('term_id', type=int)
    subject_id = request.args.get('subject_id', type=int)
    fmt = (request.args.get('format') or 'pdf').lower()
    if not term_id or not subject_id:
        flash('Select a term and subject first.', 'error')
        return redirect(url_for('subjects.institution_analytics'))
    data = subject_scorecard(term_id, subject_id, _org_allowed_ids(term_id))
    if not data or not (data.get('summary') or {}).get('entries'):
        flash('No scores for that subject yet.', 'warning')
        return redirect(url_for('subjects.subject_scorecard_view',
                                term_id=term_id, subject_id=subject_id))
    term = db.session.get(Term, term_id)
    if fmt in ('excel', 'xlsx'):
        return xlsx_response(subject_xlsx(data, term), subject_filename(data, term, 'xlsx'))
    if fmt in ('image', 'png'):
        return png_response(subject_png(data, term), subject_filename(data, term, 'png'), inline=False)
    return pdf_response(subject_pdf(data, term), subject_filename(data, term, 'pdf'), inline=False)


@subjects_bp.route('/affective', methods=['GET', 'POST'])
@login_required
def affective():
    """Enter behavioural / affective ratings (1–5) for a class arm in a term."""
    from utils.report_card import active_traits
    AFFECTIVE_TRAITS = active_traits()
    AFFECTIVE_KEYS = [k for k, _ in AFFECTIVE_TRAITS]
    term_id = request.values.get('term_id', type=int)
    assignment_id = request.values.get('assignment_id', type=int)
    if assignment_id and not can_access_class(assignment_id):
        flash('You do not have access to this class.', 'error')
        return redirect(url_for('subjects.affective'))
    if not term_id:
        active = get_active_term()
        term_id = active.id if active else None
    terms = session_terms()
    assignments = (strip_sss3_third_term(filter_classes_for_user(
        ClassArmAssignment.query.filter_by(term_id=term_id).all()), term_id) if term_id else [])
    selected_assignment = db.session.get(ClassArmAssignment, assignment_id) if assignment_id else None

    if request.method == 'POST' and selected_assignment and term_id:
        enrollments = StudentEnrollment.query.filter_by(
            class_arm_assignment_id=assignment_id, is_active=True).all()
        for e in enrollments:
            mapping = {k: request.form.get(f'r_{e.student_id}_{k}', type=int)
                       for k in AFFECTIVE_KEYS}
            ts = TermSummary.query.filter_by(student_id=e.student_id, term_id=term_id).first()
            if not ts:
                ts = TermSummary(student_id=e.student_id, term_id=term_id, enrollment_id=e.id)
                db.session.add(ts)
            ts.set_affective(mapping)
        db.session.commit()
        from utils.audit import log_action
        log_action('results.affective',
                   detail=f'term {term_id}, {selected_assignment.display_name}')
        return _ok('Behavioural ratings saved.',
                   url_for('subjects.affective', term_id=term_id, assignment_id=assignment_id))

    students = []
    if selected_assignment:
        enrollments = (StudentEnrollment.query.filter_by(
            class_arm_assignment_id=assignment_id, is_active=True)
            .join(Student).order_by(Student.surname, Student.first_name).all())
        ratings = {ts.student_id: ts.affective_map
                   for ts in TermSummary.query.filter_by(term_id=term_id).all()}
        for e in enrollments:
            students.append({'student': e.student, 'ratings': ratings.get(e.student_id, {})})

    return _render({
        'page': 'affective', 'nav': _nav_urls(),
        'term_id': term_id or '', 'assignment_id': assignment_id or '',
        'has_students': bool(selected_assignment and students),
        'selected': bool(selected_assignment),
        'terms': [{'id': t.id, 'full_name': t.full_name} for t in terms],
        'assignments': [{'id': a.id, 'display_name': a.display_name} for a in assignments],
        'traits': [{'key': k, 'label': lbl} for k, lbl in AFFECTIVE_TRAITS],
        'students': [{'id': r['student'].id, 'full_name': r['student'].full_name,
                      'ratings': r['ratings']} for r in students],
        'self_url': url_for('subjects.affective'), 'submit_url': url_for('subjects.affective'),
        'broadsheet_url': url_for('subjects.broadsheet', term_id=term_id or '', assignment_id=assignment_id or ''),
    })


@subjects_bp.route('/comments', methods=['GET', 'POST'])
@login_required
def comments():
    """Enter form-teacher & principal comments for a class arm in a term."""
    term_id = request.values.get('term_id', type=int)
    assignment_id = request.values.get('assignment_id', type=int)
    if assignment_id and not can_access_class(assignment_id):
        flash('You do not have access to this class.', 'error')
        return redirect(url_for('subjects.comments'))
    if not term_id:
        active = get_active_term()
        term_id = active.id if active else None
    terms = session_terms()
    assignments = (strip_sss3_third_term(filter_classes_for_user(
        ClassArmAssignment.query.filter_by(term_id=term_id).all()), term_id) if term_id else [])
    selected_assignment = db.session.get(ClassArmAssignment, assignment_id) if assignment_id else None

    if request.method == 'POST' and selected_assignment and term_id:
        enrollments = StudentEnrollment.query.filter_by(
            class_arm_assignment_id=assignment_id, is_active=True).all()
        for e in enrollments:
            ts = TermSummary.query.filter_by(student_id=e.student_id, term_id=term_id).first()
            if not ts:
                ts = TermSummary(student_id=e.student_id, term_id=term_id, enrollment_id=e.id)
                db.session.add(ts)
            ts.teacher_comment = strip_tags(request.form.get(f't_{e.student_id}')) or None
            ts.principal_comment = strip_tags(request.form.get(f'p_{e.student_id}')) or None
        db.session.commit()
        from utils.audit import log_action
        log_action('results.comments', detail=f'term {term_id}, {selected_assignment.display_name}')
        return _ok('Comments saved.',
                   url_for('subjects.comments', term_id=term_id, assignment_id=assignment_id))

    students = []
    if selected_assignment:
        enrollments = (StudentEnrollment.query.filter_by(
            class_arm_assignment_id=assignment_id, is_active=True)
            .join(Student).order_by(Student.surname, Student.first_name).all())
        summ = {ts.student_id: ts for ts in TermSummary.query.filter_by(term_id=term_id).all()}
        for e in enrollments:
            ts = summ.get(e.student_id)
            students.append({'student': e.student,
                             'teacher_comment': ts.teacher_comment if ts else '',
                             'principal_comment': ts.principal_comment if ts else ''})

    return _render({
        'page': 'comments', 'nav': _nav_urls(),
        'term_id': term_id or '', 'assignment_id': assignment_id or '',
        'has_students': bool(selected_assignment and students),
        'selected': bool(selected_assignment),
        'terms': [{'id': t.id, 'full_name': t.full_name} for t in terms],
        'assignments': [{'id': a.id, 'display_name': a.display_name} for a in assignments],
        'students': [{'id': r['student'].id, 'full_name': r['student'].full_name,
                      'teacher_comment': r['teacher_comment'] or '',
                      'principal_comment': r['principal_comment'] or ''} for r in students],
        'self_url': url_for('subjects.comments'), 'submit_url': url_for('subjects.comments'),
        'broadsheet_url': url_for('subjects.broadsheet', term_id=term_id or '', assignment_id=assignment_id or ''),
    })


@subjects_bp.route('/report-card/<int:student_id>')
@login_required
@result_card_required
def student_report_card(student_id):
    """View student report card"""
    student = db.get_or_404(Student, student_id)
    from utils.access_control import assert_student_access
    assert_student_access(student)   # branch + form-teacher scope
    term_id = request.args.get('term_id', type=int)

    terms = session_terms()
    
    if not term_id:
        active_term = get_active_term()
        if active_term:
            term_id = active_term.id
    
    selected_term = db.session.get(Term, term_id) if term_id else None

    report_data = None
    enrollment = None

    from utils.report_card import active_traits, RATING_LABELS
    if selected_term:
        from utils.report_card import build_report_card
        enrollment, report_data = build_report_card(student_id, term_id)

    from utils.school import school_profile
    return render_template('subjects/report_card.html',
        student=student, terms=terms, term_id=term_id, selected_term=selected_term,
        report_data=report_data, enrollment=enrollment, school=school_profile(),
        affective_traits=active_traits(), rating_labels=RATING_LABELS
    )


@subjects_bp.route('/report-card/<int:student_id>/pdf')
@login_required
@result_card_required
def report_card_pdf(student_id):
    """Download the student's term report card as a PDF."""
    from flask import send_file
    from utils.report_card import build_report_card, active_traits, RATING_LABELS
    from utils.report_pdf import report_card_pdf as build_pdf
    from utils.access_control import assert_student_access
    student = db.get_or_404(Student, student_id)
    assert_student_access(student)   # branch + form-teacher scope
    term_id = request.args.get('term_id', type=int) or (
        get_active_term().id if get_active_term() else None)
    _, report_data = build_report_card(student_id, term_id) if term_id else (None, None)
    if not report_data:
        flash('No results to export for this term.', 'error')
        return redirect(url_for('subjects.student_report_card', student_id=student_id, term_id=term_id))
    term = db.session.get(Term, term_id)
    from utils.school import school_profile
    buf = build_pdf(student, report_data, term, school_profile(),
                    active_traits(), RATING_LABELS)
    name = f"{student.student_id}_{term.name.replace(' ', '_')}_report.pdf"
    return send_file(buf, mimetype='application/pdf', as_attachment=True, download_name=name)


@subjects_bp.route('/report-cards/pdf')
@login_required
@result_card_required
def report_cards_pdf_batch():
    """Download every report card for a class+term as one PDF (one per page).
    Reuses the single-card renderer, so batch pages match the individual PDFs."""
    from flask import send_file
    from utils.report_card import build_report_card, active_traits, RATING_LABELS
    from utils.report_pdf import batch_report_cards_pdf
    from sqlalchemy.orm import joinedload
    term_id = request.args.get('term_id', type=int)
    assignment_id = request.args.get('assignment_id', type=int)
    if not term_id or not assignment_id:
        flash('Select a term and class first.', 'error')
        return redirect(url_for('subjects.print_all_report_cards', term_id=term_id or ''))
    if not can_access_class(assignment_id):
        flash('You do not have access to that class.', 'error')
        return redirect(url_for('subjects.print_all_report_cards'))
    term = db.session.get(Term, term_id)
    asg = db.session.get(ClassArmAssignment, assignment_id)
    if not term or not asg:
        flash('Select a term and class first.', 'error')
        return redirect(url_for('subjects.print_all_report_cards'))
    enrollments = (StudentEnrollment.query
                   .options(joinedload(StudentEnrollment.student))
                   .filter_by(class_arm_assignment_id=assignment_id, is_active=True)
                   .join(Student).order_by(Student.surname, Student.first_name).all())
    cards = []
    for e in enrollments:
        _, rc = build_report_card(e.student_id, term_id)
        if rc:
            cards.append((e.student, rc, term))
    if not cards:
        flash('No results to export for this class yet.', 'error')
        return redirect(url_for('subjects.print_all_report_cards', term_id=term_id, assignment_id=assignment_id))
    from utils.school import school_profile
    buf = batch_report_cards_pdf(cards, school_profile(),
                                 active_traits(), RATING_LABELS,
                                 title=f'{asg.display_name} — {term.full_name}')
    from utils.audit import log_action
    log_action('results.report_cards_pdf', detail=f'{asg.display_name} {term.name}: {len(cards)} card(s)')
    name = f"{asg.display_name.replace(' ', '_')}_{term.name.replace(' ', '_')}_report_cards.pdf"
    return send_file(buf, mimetype='application/pdf', as_attachment=True, download_name=name)


@subjects_bp.route('/broadsheet/export')
@login_required
def export_broadsheet():
    """Export a class broadsheet. ``format`` = excel (default) | pdf | word | image."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from flask import Response
    import io

    term_id = request.args.get('term_id', type=int)
    assignment_id = request.args.get('assignment_id', type=int)
    fmt = (request.args.get('format') or 'excel').lower()
    min_score = request.args.get('min_score', type=float)
    filter_field = (request.args.get('filter_field') or 'average').strip() or 'average'

    if not term_id or not assignment_id:
        flash('Select term and class first.', 'error')
        return redirect(url_for('subjects.broadsheet'))

    selected_term = db.session.get(Term, term_id)
    selected_assignment = db.session.get(ClassArmAssignment, assignment_id)

    if not selected_term or not selected_assignment:
        flash('Invalid selection.', 'error')
        return redirect(url_for('subjects.broadsheet'))
    if not can_access_class(assignment_id):
        flash('You do not have access to that class.', 'error')
        return redirect(url_for('subjects.broadsheet'))

    # Non-Excel formats are rendered by the shared exporter (PDF / Word / HD image).
    if fmt in ('pdf', 'word', 'docx', 'image', 'png'):
        from utils import broadsheet_export as bx
        from utils.web_exports import pdf_response, docx_response, png_response
        base = (selected_assignment.display_name or 'broadsheet').replace(' ', '_')
        stem = f"broadsheet_{base}_{selected_term.name.replace(' ', '_')}"
        if fmt == 'pdf':
            data = bx.broadsheet_pdf(term_id, assignment_id, min_score=min_score, filter_field=filter_field)
            return pdf_response(data, f'{stem}.pdf', inline=False)
        if fmt in ('word', 'docx'):
            data = bx.broadsheet_docx(term_id, assignment_id, min_score=min_score, filter_field=filter_field)
            return docx_response(data, f'{stem}.docx')
        data = bx.broadsheet_png(term_id, assignment_id, min_score=min_score, filter_field=filter_field)
        return png_response(data, f'{stem}.png', inline=False)

    # Get data (same as broadsheet view)
    class_subjects = ClassSubject.query.filter_by(
        term_id=term_id,
        class_id=selected_assignment.class_id,
        is_active=True
    ).filter(
        (ClassSubject.arm_id == None) | (ClassSubject.arm_id == selected_assignment.arm_id)
    ).join(Subject).order_by(Subject.name).all()
    
    assessment_types = AssessmentType.query.filter_by(is_active=True).order_by(AssessmentType.order).all()
    
    enrollments = StudentEnrollment.query.filter_by(
        class_arm_assignment_id=assignment_id,
        is_active=True
    ).join(Student).order_by(Student.surname, Student.first_name).all()
    
    pass_mark = SchoolSettings.get('pass_mark', 50)
    
    # Build data
    broadsheet_data = []
    for enrollment in enrollments:
        student_row = {
            'student': enrollment.student,
            'subjects': {},
            'total': 0,
            'average': 0
        }
        
        for cs in class_subjects:
            scores = StudentScore.query.filter_by(
                student_id=enrollment.student_id,
                class_subject_id=cs.id
            ).all()
            
            subject_total = sum(s.score for s in scores)
            student_row['subjects'][cs.id] = subject_total
            student_row['total'] += subject_total
        
        if class_subjects:
            student_row['average'] = round(student_row['total'] / len(class_subjects), 2)
        
        broadsheet_data.append(student_row)
    
    # Sort by average and stamp the (arm) position BEFORE any filtering so a
    # filtered export still shows each student's true position in the arm.
    broadsheet_data.sort(key=lambda x: x['average'], reverse=True)
    for i, r in enumerate(broadsheet_data, 1):
        r['position'] = i
    if min_score is not None:
        def _passes(r):
            if filter_field == 'average':
                v = r['average']
            elif filter_field == 'total':
                v = r['total']
            else:
                try:
                    v = r['subjects'].get(int(filter_field))
                except (TypeError, ValueError):
                    v = None
            return v is not None and v >= min_score
        broadsheet_data = [r for r in broadsheet_data if _passes(r)]

    # Create Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "Broadsheet"
    
    # Styles
    header_font = Font(bold=True, size=12)
    header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
    header_font_white = Font(bold=True, color='FFFFFF')
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    
    # Title
    ws['A1'] = f"{selected_assignment.display_name} - Broadsheet"
    ws['A1'].font = Font(bold=True, size=14)
    ws['A2'] = f"{selected_term.full_name}"
    
    # Headers
    row = 4
    headers = ['Pos', 'Student Name', 'Student ID']
    for cs in class_subjects:
        headers.append(cs.subject.short_name or cs.subject.name[:5])
    headers.extend(['Total', 'Average', 'Grade'])
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.font = header_font_white
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center')
    
    # Data rows
    for idx, data in enumerate(broadsheet_data, 1):
        row += 1
        ws.cell(row=row, column=1, value=data.get('position', idx)).border = thin_border
        ws.cell(row=row, column=2, value=data['student'].full_name).border = thin_border
        ws.cell(row=row, column=3, value=data['student'].student_id).border = thin_border
        
        col = 4
        for cs in class_subjects:
            score = data['subjects'].get(cs.id, 0)
            cell = ws.cell(row=row, column=col, value=score if score else '')
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='center')
            col += 1
        
        ws.cell(row=row, column=col, value=data['total']).border = thin_border
        ws.cell(row=row, column=col+1, value=data['average']).border = thin_border
        ws.cell(row=row, column=col+2, value=GradeScale.get_grade(data['average']) if data['average'] else '').border = thin_border
    
    # Auto-width columns
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except Exception:
                pass
        ws.column_dimensions[column].width = min(max_length + 2, 30)
    
    filename = f"broadsheet_{selected_assignment.display_name.replace(' ', '_')}_{selected_term.name}.xlsx"

    return xlsx_response(wb, filename)


@subjects_bp.route('/broadsheet/blank-sheet')
@login_required
def blank_score_sheet():
    """A printable, A4 blank score-entry sheet for a class arm: the school's own
    assessment columns (CAs, optional HA, optional P.E/M.E, CBT, PBT/Theory, Exam
    Total, General Total) with the roster pre-printed in First / Middle / Surname
    columns and a blank space for the subject name."""
    from utils.broadsheet_export import blank_sheet_pdf, blank_sheet_filename
    from utils.web_exports import pdf_response

    term_id = request.args.get('term_id', type=int)
    assignment_id = request.args.get('assignment_id', type=int)
    subject_name = (request.args.get('subject') or '').strip()[:60]

    if not term_id or not assignment_id:
        flash('Select term and class first.', 'error')
        return redirect(url_for('subjects.broadsheet'))
    selected_term = db.session.get(Term, term_id)
    selected_assignment = db.session.get(ClassArmAssignment, assignment_id)
    if not selected_term or not selected_assignment:
        flash('Invalid selection.', 'error')
        return redirect(url_for('subjects.broadsheet'))
    if not can_access_class(assignment_id):
        flash('You do not have access to that class.', 'error')
        return redirect(url_for('subjects.broadsheet'))

    data = blank_sheet_pdf(term_id, assignment_id, subject_name=subject_name)
    if not data:
        flash('Could not build the score sheet.', 'error')
        return redirect(url_for('subjects.broadsheet', term_id=term_id, assignment_id=assignment_id))
    return pdf_response(data, blank_sheet_filename(selected_assignment, selected_term), inline=True)


@subjects_bp.route('/report-cards/print-all')
@login_required
@result_card_required
def print_all_report_cards():
    """Print all report cards for a class"""
    term_id = request.args.get('term_id', type=int)
    assignment_id = request.args.get('assignment_id', type=int)
    
    terms = session_terms()
    
    if not term_id:
        active_term = get_active_term()
        if active_term:
            term_id = active_term.id
    
    selected_term = db.session.get(Term, term_id) if term_id else None

    # A user may only print report cards for classes they can access.
    if assignment_id and not can_access_class(assignment_id):
        flash('You do not have access to that class.', 'error')
        return redirect(url_for('subjects.print_all_report_cards', term_id=term_id))

    assignments = []
    if term_id:
        assignments = strip_sss3_third_term(filter_classes_for_user(
            ClassArmAssignment.query.filter_by(term_id=term_id).all()), term_id)

    selected_assignment = db.session.get(ClassArmAssignment, assignment_id) if assignment_id else None

    all_reports = []
    
    if selected_assignment and selected_term:
        assessment_types = AssessmentType.query.filter_by(is_active=True).order_by(AssessmentType.order).all()
        pass_mark = SchoolSettings.get('pass_mark', 50)
        
        class_subjects = ClassSubject.query.filter_by(
            term_id=term_id,
            class_id=selected_assignment.class_id,
            is_active=True
        ).filter(
            (ClassSubject.arm_id == None) | (ClassSubject.arm_id == selected_assignment.arm_id)
        ).join(Subject).order_by(Subject.name).all()
        
        from sqlalchemy.orm import joinedload
        enrollments = (StudentEnrollment.query
                       .options(joinedload(StudentEnrollment.student))
                       .filter_by(class_arm_assignment_id=assignment_id, is_active=True)
                       .join(Student).order_by(Student.surname, Student.first_name).all())

        # Every score for the class in one query, indexed by (student,
        # class-subject, assessment) — was a query per student × subject, the
        # main cost when printing a whole class.
        cs_ids = [cs.id for cs in class_subjects]
        sids = [e.student_id for e in enrollments]
        score_map = {}
        if cs_ids and sids:
            for s in StudentScore.query.filter(
                    StudentScore.student_id.in_(sids),
                    StudentScore.class_subject_id.in_(cs_ids)).all():
                score_map[(s.student_id, s.class_subject_id, s.assessment_type_id)] = s.score

        for enrollment in enrollments:
            student = enrollment.student
            subjects_data = []
            total_score = 0
            subjects_passed = 0
            subjects_failed = 0

            for cs in class_subjects:
                subject_row = {
                    'subject': cs.subject,
                    'assessments': {},
                    'total': 0,
                    'grade': '-',
                    'remark': '-'
                }

                subject_total = 0
                for at in assessment_types:
                    score = score_map.get((student.id, cs.id, at.id))
                    subject_row['assessments'][at.id] = score
                    if score:
                        subject_total += score
                
                subject_row['total'] = subject_total
                if subject_total > 0:
                    subject_row['grade'] = GradeScale.get_grade(subject_total)
                    subject_row['remark'] = GradeScale.get_remark(subject_total)
                    total_score += subject_total
                    
                    if subject_total >= pass_mark:
                        subjects_passed += 1
                    else:
                        subjects_failed += 1
                
                subjects_data.append(subject_row)
            
            average = round(total_score / len(class_subjects), 2) if class_subjects else 0
            
            all_reports.append({
                'student': student,
                'enrollment': enrollment,
                'subjects': subjects_data,
                'assessment_types': assessment_types,
                'total_score': total_score,
                'average': average,
                'overall_grade': GradeScale.get_grade(average) if average else '-',
                'subjects_passed': subjects_passed,
                'subjects_failed': subjects_failed,
                'total_subjects': len(class_subjects)
            })
    
    school_name = SchoolSettings.get('school_name', 'School Name')
    
    return render_template('subjects/print_all_report_cards.html',
        terms=terms, term_id=term_id, selected_term=selected_term,
        assignments=assignments, assignment_id=assignment_id, selected_assignment=selected_assignment,
        all_reports=all_reports, school_name=school_name
    )
