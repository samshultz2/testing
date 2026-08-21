"""subjects blueprint — scores routes (split from the former routes/subjects.py)."""
from routes.subjects import *  # noqa: F401,F403


@subjects_bp.route('/scores')
@login_required
def scores_entry():
    """Score entry page"""
    term_id = request.args.get('term_id', type=int)
    assignment_id = request.args.get('assignment_id', type=int)
    class_subject_id = request.args.get('class_subject_id', type=int)
    assessment_type_id = request.args.get('assessment_type_id', type=int)
    
    # Check if user can enter results
    if not can_enter_results() and not is_admin():
        flash('You do not have permission to enter scores.', 'error')
        return redirect(url_for('main.dashboard'))
    
    # Check class access
    if assignment_id and not can_access_class(assignment_id):
        flash('You do not have access to this class.', 'error')
        return redirect(url_for('subjects.scores_entry'))
    
    terms = session_terms()
    
    # Get active term if not specified
    if not term_id:
        active_term = get_active_term()
        if active_term:
            term_id = active_term.id
    
    selected_term = db.session.get(Term, term_id) if term_id else None
    
    # Get class arm assignments for selected term (filtered for teachers)
    assignments = []
    if term_id:
        all_assignments = ClassArmAssignment.query.filter_by(term_id=term_id).all()
        # SSS3 write no internal exams in third term (only WAEC/NECO/JAMB), so
        # they never appear in third-term internal score entry.
        assignments = strip_sss3_third_term(filter_classes_for_user(all_assignments), term_id)

    selected_assignment = db.session.get(ClassArmAssignment, assignment_id) if assignment_id else None
    # Refuse a directly-passed SSS3 third-term class (e.g. a stale bookmark).
    if is_sss3_third_term_assignment(selected_assignment, term_id):
        flash('SSS3 sit no internal exams in third term — only WAEC/NECO/JAMB.', 'info')
        return redirect(url_for('subjects.scores_entry', term_id=term_id))
    
    # Get subjects for selected class (filter by teacher's assigned subjects if not admin)
    class_subjects = []
    if selected_assignment:
        from utils.access_control import get_teacher_profile
        teacher = get_teacher_profile()
        
        class_subjects_query = ClassSubject.query.filter_by(
            term_id=term_id,
            class_id=selected_assignment.class_id,
            is_active=True
        ).filter(
            (ClassSubject.arm_id == None) | (ClassSubject.arm_id == selected_assignment.arm_id)
        ).join(Subject).order_by(Subject.name)
        
        all_class_subjects = class_subjects_query.all()
        
        # Filter by teacher's assigned subjects if not admin
        if teacher and not is_admin():
            teacher_subject_ids = [
                a.subject_id for a in teacher.subject_assignments.filter_by(
                    class_arm_assignment_id=assignment_id,
                    is_active=True
                ).all()
            ]
            class_subjects = [cs for cs in all_class_subjects if cs.subject_id in teacher_subject_ids]
        else:
            class_subjects = all_class_subjects
    
    selected_class_subject = db.session.get(ClassSubject, class_subject_id) if class_subject_id else None
    
    # Get assessment types
    assessment_types = AssessmentType.query.filter_by(is_active=True).order_by(AssessmentType.order).all()
    selected_assessment = db.session.get(AssessmentType, assessment_type_id) if assessment_type_id else None
    
    # Get students and existing scores
    students_data = []
    if selected_assignment and selected_class_subject and selected_assessment:
        from sqlalchemy.orm import joinedload
        enrollments = (StudentEnrollment.query
                       .options(joinedload(StudentEnrollment.student))
                       .filter_by(class_arm_assignment_id=assignment_id, is_active=True)
                       .join(Student).order_by(Student.surname, Student.first_name).all())
        # One query for every existing score in this class-subject + assessment,
        # indexed by student, instead of a lookup per student (removes the N+1).
        existing = {s.student_id: s.score for s in StudentScore.query.filter_by(
            class_subject_id=class_subject_id,
            assessment_type_id=assessment_type_id).all()}
        for enrollment in enrollments:
            students_data.append({
                'enrollment': enrollment,
                'student': enrollment.student,
                'score': existing.get(enrollment.student_id),
            })
    
    # Max score: per-term setting > subject override > global default.
    max_score = selected_assessment.max_score if selected_assessment else 0
    if selected_class_subject and selected_assessment:
        from utils.assessments import effective_max
        max_score = effective_max(selected_class_subject.subject, selected_assessment, term=term_id)

    # Roster for the "by student" entry mode: every student in the selected
    # class arm, so the UI can offer a live-filterable single-student picker
    # without an extra round-trip.
    roster = []
    if selected_assignment:
        from sqlalchemy.orm import joinedload as _jl
        for e in (StudentEnrollment.query
                  .options(_jl(StudentEnrollment.student))
                  .filter_by(class_arm_assignment_id=assignment_id, is_active=True)
                  .join(Student).order_by(Student.surname, Student.first_name).all()):
            roster.append({'id': e.student.id, 'full_name': e.student.full_name,
                           'gender': e.student.gender or ''})

    return _render({
        'page': 'scores', 'nav': _nav_urls(),
        'term_id': term_id or '', 'assignment_id': assignment_id or '',
        'class_subject_id': class_subject_id or '', 'assessment_type_id': assessment_type_id or '',
        'terms': [{'id': t.id, 'full_name': t.full_name} for t in terms],
        'assignments': [{'id': a.id, 'display_name': a.display_name} for a in assignments],
        'class_subjects': [{'id': cs.id, 'subject_name': cs.subject.name} for cs in class_subjects],
        'assessment_types': [{'id': at.id, 'name': at.name, 'max_score': at.max_score} for at in assessment_types],
        'selected_subject': selected_class_subject.subject.name if selected_class_subject else '',
        'selected_assessment': selected_assessment.name if selected_assessment else '',
        'max_score': max_score,
        'has_selection': bool(selected_assignment and selected_class_subject and selected_assessment),
        'students_data': [{'id': it['student'].id, 'full_name': it['student'].full_name,
                           'gender': it['student'].gender or '',
                           'score': it['score'] if it['score'] is not None else ''}
                          for it in students_data],
        'roster': roster,
        'self_url': url_for('subjects.scores_entry'),
        'save_url': url_for('subjects.save_scores'),
        'save_student_url': url_for('subjects.save_student_scores'),
        'student_scores_api': url_for('subjects.api_student_subject_scores'),
        'urls': {'scan': url_for('subjects.scoresheet_scan', term_id=term_id or '', assignment_id=assignment_id or '', class_subject_id=class_subject_id or ''),
                 'paste': url_for('subjects.scoresheet_paste', term_id=term_id or '', assignment_id=assignment_id or '', class_subject_id=class_subject_id or ''),
                 'import': url_for('subjects.import_scores', term_id=term_id or '', assignment_id=assignment_id or '', class_subject_id=class_subject_id or ''),
                 'broadsheet_import': url_for('subjects.broadsheet_import', term_id=term_id or '', assignment_id=assignment_id or ''),
                 'blank_sheet': (url_for('subjects.blank_score_sheet', term_id=term_id or '', assignment_id=assignment_id or '',
                                         subject=selected_class_subject.subject.name if selected_class_subject else '')
                                 if term_id and assignment_id else '')},
    })


@subjects_bp.route('/scores/save', methods=['POST'])
@login_required
def save_scores():
    """Save student scores"""
    try:
        class_subject_id = request.form.get('class_subject_id', type=int)
        assessment_type_id = request.form.get('assessment_type_id', type=int)
        assignment_id = request.form.get('assignment_id', type=int)
        term_id = request.form.get('term_id', type=int)
        
        student_ids = request.form.getlist('student_id[]')
        scores = request.form.getlist('score[]')

        # A teacher may only save scores for a subject they actually teach in
        # this class (admins pass). Resolve the subject from the class-subject.
        cs = db.session.get(ClassSubject, class_subject_id) if class_subject_id else None
        if not can_enter_results(assignment_id, cs.subject_id if cs else None):
            return _err('You can only enter scores for the subjects you teach in this class.',
                        url_for('subjects.scores_entry', term_id=term_id, assignment_id=assignment_id))

        at = db.session.get(AssessmentType, assessment_type_id)
        # Per-term setting > subject override > global default.
        if at and cs:
            from utils.assessments import effective_max
            max_score = effective_max(cs.subject, at, term=term_id)
        else:
            max_score = at.max_score if at else None
        items = [(int(sid), assessment_type_id,
                  scores[i].strip() if i < len(scores) else '', max_score)
                 for i, sid in enumerate(student_ids)]
        counts = persist_scores(term_id, assignment_id, class_subject_id,
                                cs.subject_id if cs else None, items)
        if counts is None:
            return _err('Results for this term are published — ask an admin to '
                        'unlock them before editing scores.',
                        url_for('subjects.scores_entry', term_id=term_id,
                                assignment_id=assignment_id))

        db.session.commit()
        # Keep term results/positions fresh as scores change.
        if term_id and assignment_id:
            asg = db.session.get(ClassArmAssignment, assignment_id)
            if asg:
                from utils.report_card import compute_term_summaries
                compute_term_summaries(term_id, asg.class_id)
            from utils.results_analytics import bust as _bust_analytics
            _bust_analytics(term_id, assignment_id)
            from utils.results_analytics_org import bust_all as _bust_org
            _bust_org()
        msg = f'{counts["saved"]} scores saved!'
        if counts['rejected']:
            msg += f' {counts["rejected"]} skipped (outside the 0–{max_score:g} range).'
        if counts['blocked']:
            msg += (f' {counts["blocked"]} left unchanged (you can enter new '
                    f'scores but not edit existing ones).')
        return _ok(msg, url_for('subjects.scores_entry',
            term_id=term_id, assignment_id=assignment_id,
            class_subject_id=class_subject_id, assessment_type_id=assessment_type_id))
    except Exception as e:
        db.session.rollback()
        return _err(f'Error: {str(e)}', url_for('subjects.scores_entry',
            term_id=term_id, assignment_id=assignment_id,
            class_subject_id=class_subject_id, assessment_type_id=assessment_type_id))


@subjects_bp.route('/workflow')
@login_required
def workflow():
    """Guided results checklist for a class+term: setup → entry → finalize → print."""
    term_id = request.args.get('term_id', type=int)
    assignment_id = request.args.get('assignment_id', type=int)
    if assignment_id and not can_access_class(assignment_id):
        flash('You do not have access to this class.', 'error')
        return redirect(url_for('subjects.workflow'))
    if not term_id:
        active = get_active_term()
        term_id = active.id if active else None
    terms = session_terms()
    assignments = (strip_sss3_third_term(filter_classes_for_user(
        ClassArmAssignment.query.filter_by(term_id=term_id).all()), term_id) if term_id else [])
    selected = db.session.get(ClassArmAssignment, assignment_id) if assignment_id else None

    steps = None
    if selected and term_id:
        enr = StudentEnrollment.query.filter_by(
            class_arm_assignment_id=assignment_id, is_active=True).all()
        sids = [e.student_id for e in enr]
        css = ClassSubject.query.filter_by(
            term_id=term_id, class_id=selected.class_id, is_active=True).all()
        n_assess = AssessmentType.query.filter_by(is_active=True).count()
        entered = 0
        if sids and css:
            entered = (StudentScore.query.join(ClassSubject).filter(
                ClassSubject.term_id == term_id,
                ClassSubject.class_id == selected.class_id,
                StudentScore.student_id.in_(sids)).count())
        ts_rows = (TermSummary.query.filter(
            TermSummary.term_id == term_id,
            TermSummary.student_id.in_(sids or [-1])).all())
        steps = {
            'students': len(sids),
            'subjects': len(css),
            'scores_entered': entered,
            'scores_expected': len(sids) * len(css) * (n_assess or 1),
            'positions': sum(1 for t in ts_rows if t.position_in_class),
            'comments': sum(1 for t in ts_rows if t.teacher_comment),
            'behaviour': sum(1 for t in ts_rows if t.affective),
        }
    selected_term = db.session.get(Term, term_id) if term_id else None
    return _render({
        'page': 'workflow', 'nav': _nav_urls(),
        'term_id': term_id or '', 'assignment_id': assignment_id or '',
        'terms': [{'id': t.id, 'full_name': t.full_name} for t in terms],
        'assignments': [{'id': a.id, 'display_name': a.display_name} for a in assignments],
        'steps': steps, 'published': bool(selected_term and selected_term.results_published),
        'self_url': url_for('subjects.workflow'),
        'urls': {
            'class_subjects': url_for('subjects.class_subjects_list'),
            'enrol': url_for('academics.assignments_list'),
            'bulk_entry': url_for('subjects.bulk_entry', term_id=term_id or '', assignment_id=assignment_id or ''),
            'broadsheet': url_for('subjects.broadsheet', term_id=term_id or '', assignment_id=assignment_id or ''),
            'comments': url_for('subjects.comments', term_id=term_id or '', assignment_id=assignment_id or ''),
            'affective': url_for('subjects.affective', term_id=term_id or '', assignment_id=assignment_id or ''),
            'compute': url_for('subjects.compute_summaries'),
            'print_all': url_for('subjects.print_all_report_cards', term_id=term_id or '', assignment_id=assignment_id or ''),
            'publish': url_for('scratchcards.publish', term_id=term_id) if term_id else '',
        },
    })


@subjects_bp.route('/bulk-entry', methods=['GET', 'POST'])
@login_required
def bulk_entry():
    """Enter every subject's scores for a whole class on one screen."""
    term_id = request.values.get('term_id', type=int)
    assignment_id = request.values.get('assignment_id', type=int)
    if assignment_id and not can_access_class(assignment_id):
        flash('You do not have access to this class.', 'error')
        return redirect(url_for('subjects.bulk_entry'))
    if not term_id:
        active = get_active_term()
        term_id = active.id if active else None
    terms = session_terms()
    assignments = (strip_sss3_third_term(filter_classes_for_user(
        ClassArmAssignment.query.filter_by(term_id=term_id).all()), term_id) if term_id else [])
    selected = db.session.get(ClassArmAssignment, assignment_id) if assignment_id else None
    assessment_types = AssessmentType.query.filter_by(is_active=True).order_by(AssessmentType.order).all()

    class_subjects, enrollments = [], []
    if selected:
        class_subjects = (ClassSubject.query.filter_by(
            term_id=term_id, class_id=selected.class_id, is_active=True)
            .filter((ClassSubject.arm_id == None) | (ClassSubject.arm_id == selected.arm_id))
            .join(Subject).order_by(Subject.name).all())
        # A teacher only sees/saves the subjects they actually teach in this class
        # (admins/eligible staff keep all). Scopes both the grid and the save loop.
        class_subjects = [cs for cs in class_subjects
                          if can_enter_results(assignment_id, cs.subject_id)]
        enrollments = (StudentEnrollment.query.filter_by(
            class_arm_assignment_id=assignment_id, is_active=True)
            .join(Student).order_by(Student.surname, Student.first_name).all())

    if request.method == 'POST' and selected and class_subjects and enrollments:
        total = {'saved': 0, 'rejected': 0, 'blocked': 0}
        for cs in class_subjects:
            items = [(e.student_id, at.id,
                      request.form.get(f's_{e.student_id}_{cs.id}_{at.id}'), at.max_score)
                     for e in enrollments for at in assessment_types]
            counts = persist_scores(term_id, assignment_id, cs.id, cs.subject_id, items)
            if counts is None:
                return _err('Results for this term are published — ask an admin to '
                            'unlock them before editing scores.',
                            url_for('subjects.bulk_entry', term_id=term_id,
                                    assignment_id=assignment_id))
            for k in total:
                total[k] += counts[k]
        db.session.commit()
        from utils.report_card import compute_term_summaries
        compute_term_summaries(term_id, selected.class_id)
        from utils.results_analytics import bust as _bust_analytics
        _bust_analytics(term_id, assignment_id)
        from utils.results_analytics_org import bust_all as _bust_org
        _bust_org()
        msg = f'Saved — {total["saved"]} change(s).'
        if total['rejected']:
            msg += f' {total["rejected"]} skipped (outside the allowed range).'
        if total['blocked']:
            msg += (f' {total["blocked"]} left unchanged (edit permission '
                    f'required to alter existing scores).')
        return _ok(msg, url_for('subjects.bulk_entry', term_id=term_id, assignment_id=assignment_id))

    scores = {}
    if selected and class_subjects and enrollments:
        sids = [e.student_id for e in enrollments]
        cs_ids = [cs.id for cs in class_subjects]
        for s in StudentScore.query.filter(StudentScore.student_id.in_(sids),
                                           StudentScore.class_subject_id.in_(cs_ids)).all():
            scores[(s.student_id, s.class_subject_id, s.assessment_type_id)] = s.score

    students = [e.student for e in enrollments]
    return _render({
        'page': 'bulk_entry', 'nav': _nav_urls(),
        'term_id': term_id or '', 'assignment_id': assignment_id or '',
        'terms': [{'id': t.id, 'full_name': t.full_name} for t in terms],
        'assignments': [{'id': a.id, 'display_name': a.display_name} for a in assignments],
        'has_grid': bool(selected and class_subjects and enrollments),
        'class_subjects': [{'id': cs.id, 'subject_name': cs.subject.name} for cs in class_subjects],
        'assessment_types': [{'id': at.id, 'name': at.name, 'short_name': at.short_name or at.name,
                              'max_score': at.max_score} for at in assessment_types],
        'students': [{'id': s.id, 'full_name': s.full_name} for s in students],
        'scores': {f'{sid}_{csid}_{atid}': v for (sid, csid, atid), v in scores.items()},
        'self_url': url_for('subjects.bulk_entry'), 'submit_url': url_for('subjects.bulk_entry'),
        'broadsheet_url': url_for('subjects.broadsheet', term_id=term_id or '', assignment_id=assignment_id or ''),
    })


@subjects_bp.route('/api/student-subject-scores')
@login_required
def api_student_subject_scores():
    """All assessment rows (CA1, CA2, Exam…) for one student in one class-subject,
    with each row's effective max and any existing score — powers the single-
    student entry mode (enter the whole subject for a student at once)."""
    term_id = request.args.get('term_id', type=int)
    assignment_id = request.args.get('assignment_id', type=int)
    class_subject_id = request.args.get('class_subject_id', type=int)
    student_id = request.args.get('student_id', type=int)
    if not (assignment_id and class_subject_id and student_id):
        return jsonify({'error': 'Missing selection.'}), 400
    if not can_access_class(assignment_id):
        return jsonify({'error': 'No access to this class.'}), 403
    cs = db.session.get(ClassSubject, class_subject_id)
    student = db.session.get(Student, student_id)
    if not cs or not student:
        return jsonify({'error': 'Not found.'}), 404
    require_branch_access(student.branch_id)

    from utils.assessments import effective_max
    ats = AssessmentType.query.filter_by(is_active=True).order_by(AssessmentType.order).all()
    existing = {s.assessment_type_id: s.score for s in StudentScore.query.filter_by(
        student_id=student_id, class_subject_id=class_subject_id).all()}
    rows = [{'assessment_type_id': at.id, 'name': at.name,
             'max_score': effective_max(cs.subject, at, term=term_id),
             'score': existing.get(at.id) if existing.get(at.id) is not None else ''}
            for at in ats]
    return jsonify({'subject': cs.subject.name, 'student': student.full_name, 'rows': rows})


@subjects_bp.route('/scores/save-student', methods=['POST'])
@login_required
def save_student_scores():
    """Save every assessment score for a single student in one class-subject."""
    try:
        term_id = request.form.get('term_id', type=int)
        assignment_id = request.form.get('assignment_id', type=int)
        class_subject_id = request.form.get('class_subject_id', type=int)
        student_id = request.form.get('student_id', type=int)
        at_ids = request.form.getlist('assessment_type_id[]')
        scores = request.form.getlist('score[]')

        cs = db.session.get(ClassSubject, class_subject_id) if class_subject_id else None
        if not (cs and student_id):
            return _err('Select a student and subject first.',
                        url_for('subjects.scores_entry', term_id=term_id, assignment_id=assignment_id))
        if not can_enter_results(assignment_id, cs.subject_id):
            return _err('You can only enter scores for the subjects you teach in this class.',
                        url_for('subjects.scores_entry', term_id=term_id, assignment_id=assignment_id))

        from utils.assessments import effective_max
        items = []
        for i, at_raw in enumerate(at_ids):
            try:
                at_id = int(at_raw)
            except (TypeError, ValueError):
                continue
            at = db.session.get(AssessmentType, at_id)
            mx = effective_max(cs.subject, at, term=term_id) if at else None
            items.append((int(student_id), at_id,
                          scores[i].strip() if i < len(scores) else '', mx))

        counts = persist_scores(term_id, assignment_id, class_subject_id, cs.subject_id, items)
        if counts is None:
            return _err('Results for this term are published — ask an admin to '
                        'unlock them before editing scores.',
                        url_for('subjects.scores_entry', term_id=term_id, assignment_id=assignment_id))
        db.session.commit()
        if term_id and assignment_id:
            asg = db.session.get(ClassArmAssignment, assignment_id)
            if asg:
                from utils.report_card import compute_term_summaries
                compute_term_summaries(term_id, asg.class_id)
            from utils.results_analytics import bust as _bust_analytics
            _bust_analytics(term_id, assignment_id)
            from utils.results_analytics_org import bust_all as _bust_org
            _bust_org()
        msg = f'{counts["saved"]} score(s) saved for this student.'
        if counts['rejected']:
            msg += f' {counts["rejected"]} skipped (out of range).'
        if counts['blocked']:
            msg += f' {counts["blocked"]} left unchanged (existing scores locked).'
        return _ok(msg, url_for('subjects.scores_entry', term_id=term_id,
                   assignment_id=assignment_id, class_subject_id=class_subject_id))
    except Exception as e:
        db.session.rollback()
        return _err(f'Error: {str(e)}', url_for('subjects.scores_entry',
                    term_id=request.form.get('term_id', type=int),
                    assignment_id=request.form.get('assignment_id', type=int)))


@subjects_bp.route('/api/student-scores/<int:student_id>/<int:term_id>')
@login_required
def api_student_scores(student_id, term_id):
    """Get all scores for a student in a term"""
    require_branch_access(db.get_or_404(Student, student_id).branch_id)
    scores = StudentScore.query.join(ClassSubject).filter(
        StudentScore.student_id == student_id,
        ClassSubject.term_id == term_id
    ).all()
    
    return jsonify([{
        'subject': s.class_subject.subject.name,
        'assessment': s.assessment_type.name,
        'score': s.score
    } for s in scores])


@subjects_bp.route('/scores/import', methods=['GET', 'POST'])
@login_required
def import_scores():
    """Import scores from Excel"""
    if request.method == 'POST':
        try:
            from openpyxl import load_workbook
            
            term_id = request.form.get('term_id', type=int)
            assignment_id = request.form.get('assignment_id', type=int)
            class_subject_id = request.form.get('class_subject_id', type=int)

            # Teachers may only import scores for a subject they teach here.
            cs = db.session.get(ClassSubject, class_subject_id) if class_subject_id else None
            if not can_enter_results(assignment_id, cs.subject_id if cs else None):
                flash('You can only enter scores for the subjects you teach in this class.', 'error')
                return redirect(url_for('subjects.import_scores'))

            if 'file' not in request.files:
                flash('No file selected.', 'error')
                return redirect(url_for('subjects.import_scores'))
            
            file = request.files['file']
            if not file.filename.endswith(('.xlsx', '.xls')):
                flash('Please upload an Excel file.', 'error')
                return redirect(url_for('subjects.import_scores'))
            
            wb = load_workbook(file)
            ws = wb.active
            
            # Get assessment types
            assessment_types = AssessmentType.query.filter_by(is_active=True).order_by(AssessmentType.order).all()
            
            # Expected columns: Student ID, then assessment type names
            errors = []
            items = []            # (student_id, at_id, raw, max_score)

            for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), 2):
                if not row[0]:
                    continue

                student_id_str = str(row[0]).strip()
                student = Student.query.filter_by(student_id=student_id_str).first()

                if not student:
                    errors.append(f"Row {row_num}: Student {student_id_str} not found")
                    continue

                # Import each assessment score
                for col_idx, at in enumerate(assessment_types, 1):
                    if col_idx < len(row) and row[col_idx] is not None:
                        items.append((student.id, at.id, row[col_idx], at.max_score))

            counts = persist_scores(term_id, assignment_id, class_subject_id,
                                    cs.subject_id if cs else None, items,
                                    allow_delete=False)
            if counts is None:
                flash('Results for this term are published — ask an admin to '
                      'unlock them before importing scores.', 'error')
                return redirect(url_for('subjects.import_scores'))

            db.session.commit()
            imported = counts['saved']
            flash(f'Imported {imported} scores!', 'success')
            
            if errors:
                for err in errors[:5]:
                    flash(err, 'warning')
                    
        except Exception as e:
            db.session.rollback()
            flash(f'Error: {str(e)}', 'error')
        
        return redirect(url_for('subjects.scores_entry', term_id=term_id, assignment_id=assignment_id, class_subject_id=class_subject_id))
    
    # GET - show form
    term_id = request.args.get('term_id', type=int)
    assignment_id = request.args.get('assignment_id', type=int)
    class_subject_id = request.args.get('class_subject_id', type=int)
    
    terms = session_terms()
    assignments = strip_sss3_third_term(filter_classes_for_user(
        ClassArmAssignment.query.filter_by(term_id=term_id).all()), term_id) if term_id else []
    class_subjects = []

    if assignment_id:
        assignment = db.session.get(ClassArmAssignment, assignment_id)
        if assignment:
            class_subjects = ClassSubject.query.filter_by(
                term_id=term_id,
                class_id=assignment.class_id,
                is_active=True
            ).join(Subject).order_by(Subject.name).all()
    
    return render_template('subjects/import_scores.html',
        terms=terms, term_id=term_id,
        assignments=assignments, assignment_id=assignment_id,
        class_subjects=class_subjects, class_subject_id=class_subject_id
    )


@subjects_bp.route('/scores/import/template')
@login_required
def score_import_template():
    """Download score import template"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    from flask import Response
    import io
    
    assignment_id = request.args.get('assignment_id', type=int)
    class_subject_id = request.args.get('class_subject_id', type=int)
    
    if not assignment_id:
        flash('Select a class first.', 'error')
        return redirect(url_for('subjects.import_scores'))
    
    assignment = db.session.get(ClassArmAssignment, assignment_id)
    assessment_types = AssessmentType.query.filter_by(is_active=True).order_by(AssessmentType.order).all()
    
    # Get students
    enrollments = StudentEnrollment.query.filter_by(
        class_arm_assignment_id=assignment_id,
        is_active=True
    ).join(Student).order_by(Student.surname, Student.first_name).all()
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Score Import"
    
    # Headers
    headers = ['Student ID', 'Student Name'] + [at.short_name or at.name for at in assessment_types]
    header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF')
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
    
    # Student rows
    for row_num, enrollment in enumerate(enrollments, 2):
        ws.cell(row=row_num, column=1, value=enrollment.student.student_id)
        ws.cell(row=row_num, column=2, value=enrollment.student.full_name)
    
    return xlsx_response(wb, 'score_import_template.xlsx')


@subjects_bp.route('/scores/scan', methods=['GET', 'POST'])
@login_required
def scoresheet_scan():
    """Upload a photographed score sheet and OCR it into an editable grid."""
    if not can_enter_results() and not is_admin():
        flash('You do not have permission to enter scores.', 'error')
        return redirect(url_for('main.dashboard'))

    from utils.waec_ocr import (
        tesseract_available, extract_text, extract_text_from_pdf,
        parse_score_sheet, match_students_unique, vision_available, vision_extract_scoresheet,
    )

    ctx = _scan_selector_context()
    ctx['vision_on'] = vision_available()      # Claude reads handwriting when a key is set

    if request.method == 'POST':
        term_id = ctx['term_id']
        assignment_id = ctx['assignment_id']
        class_subject_id = ctx['class_subject_id']

        assignment = db.session.get(ClassArmAssignment, assignment_id) if assignment_id else None
        class_subject = db.session.get(ClassSubject, class_subject_id) if class_subject_id else None

        if not (assignment and class_subject):
            flash('Select a class and subject before uploading.', 'error')
            return render_template('subjects/scoresheet_scan.html', **ctx)

        if not can_access_class(assignment_id):
            flash('You do not have access to this class.', 'error')
            return redirect(url_for('subjects.scoresheet_scan'))

        if not tesseract_available() and not vision_available():
            flash('OCR engine (Tesseract) is not available on the server.', 'error')
            return render_template('subjects/scoresheet_scan.html', **ctx)

        upload = request.files.get('file')
        if not upload or not upload.filename:
            flash('No file selected.', 'error')
            return render_template('subjects/scoresheet_scan.html', **ctx)
        from utils.uploads import ext_ok, SCAN_EXTS
        if not ext_ok(upload.filename, SCAN_EXTS):
            flash('Please upload an image or PDF.', 'error')
            return render_template('subjects/scoresheet_scan.html', **ctx)

        data = upload.read()
        is_pdf = upload.filename.lower().endswith('.pdf')
        sheet_cols = _sheet_columns(class_subject, term=term_id)

        # Read with the school's chosen engine first, then fall back through the
        # others that are available. PDFs always go through the text path.
        from utils.ocr_engine import engine_order
        col_labels = [(at.short_name or at.name) for at, _mx in sheet_cols]
        max_scores = [mx for _at, mx in sheet_cols]
        parsed = None
        order = engine_order() if not is_pdf else ['tesseract']
        for eng in order:
            try:
                if eng == 'claude' and vision_available():
                    parsed = vision_extract_scoresheet(data, col_labels, upload.mimetype or 'image/png')
                elif eng == 'tesseract':
                    text = extract_text_from_pdf(data) if is_pdf else extract_text(data)
                    parsed = parse_score_sheet(text, num_columns=len(sheet_cols))
            except Exception:
                parsed = None
            if parsed:
                break
        if parsed is None and not order:
            # Nothing available; last-resort tesseract text path.
            try:
                text = extract_text_from_pdf(data) if is_pdf else extract_text(data)
                parsed = parse_score_sheet(text, num_columns=len(sheet_cols))
            except Exception as e:
                flash(f'Could not read the image: {e}', 'error')
                return render_template('subjects/scoresheet_scan.html', **ctx)

        if not parsed:
            flash('No student rows could be detected. Try a clearer, straight photo.', 'warning')
            return render_template('subjects/scoresheet_scan.html', **ctx)

        # Students enrolled in this class (for matching + the picker).
        enrollments = StudentEnrollment.query.filter_by(
            class_arm_assignment_id=assignment_id, is_active=True
        ).join(Student).order_by(Student.surname, Student.first_name).all()
        students = [e.student for e in enrollments]
        by_student_id = {s.student_id: s for s in students}

        # 1) Exact matches on a scanned student number reserve those pupils.
        matched = [None] * len(parsed)
        used = set()
        for i, p in enumerate(parsed):
            s = by_student_id.get(p['student_num']) if p['student_num'] else None
            if s and s.id not in used:
                matched[i] = s
                used.add(s.id)
        # 2) Greedy unique fuzzy match for the rest — no two rows share one pupil.
        rest_idx = [i for i in range(len(parsed)) if matched[i] is None]
        pool = [s for s in students if s.id not in used]
        fuzzy = match_students_unique([parsed[i]['name'] for i in rest_idx], pool)
        for i, (s, _sc) in zip(rest_idx, fuzzy):
            matched[i] = s

        rows = []
        for i, p in enumerate(parsed):
            # Map the positional cells to assessment-type ids.
            cell_map = {}
            for (at, _mx), value in zip(sheet_cols, p['cells']):
                cell_map[at.id] = value
            rows.append({
                'student_num': p['student_num'],
                'name': p['name'],
                'matched_id': matched[i].id if matched[i] else None,
                'cells': cell_map,
            })

        return render_template('subjects/scoresheet_review.html',
            term_id=term_id, assignment_id=assignment_id, class_subject_id=class_subject_id,
            assignment=assignment, class_subject=class_subject,
            columns=sheet_cols, rows=rows, students=students,
        )

    return render_template('subjects/scoresheet_scan.html', **ctx)


@subjects_bp.route('/scores/paste', methods=['GET', 'POST'])
@login_required
def scoresheet_paste():
    """Paste comma-separated scores (e.g. produced by asking an external AI to read
    a photographed sheet) into an editable, student-matched grid — no OCR, no API
    keys. Reuses the same review/confirm grid and save path as the scanner."""
    if not can_enter_results() and not is_admin():
        flash('You do not have permission to enter scores.', 'error')
        return redirect(url_for('main.dashboard'))

    from utils.waec_ocr import match_students_unique

    ctx = _scan_selector_context()
    cs = db.session.get(ClassSubject, ctx['class_subject_id']) if ctx['class_subject_id'] else None
    sheet_cols = _sheet_columns(cs, term=ctx['term_id']) if cs else []
    ctx['columns'] = sheet_cols

    if request.method == 'POST':
        assignment_id = ctx['assignment_id']
        assignment = db.session.get(ClassArmAssignment, assignment_id) if assignment_id else None
        if not (assignment and cs):
            flash('Select a class and subject before pasting.', 'error')
            return render_template('subjects/scoresheet_paste.html', **ctx)
        if not can_access_class(assignment_id):
            flash('You do not have access to this class.', 'error')
            return redirect(url_for('subjects.scoresheet_paste'))

        pasted = (request.form.get('data') or '').strip()
        if not pasted:
            flash('Paste the comma-separated rows first.', 'error')
            return render_template('subjects/scoresheet_paste.html', pasted=pasted, **ctx)

        parsed = _parse_pasted_scores(pasted, len(sheet_cols))
        if not parsed:
            flash('No rows could be read from the pasted text. Check the format.', 'warning')
            return render_template('subjects/scoresheet_paste.html', pasted=pasted, **ctx)

        enrollments = StudentEnrollment.query.filter_by(
            class_arm_assignment_id=assignment_id, is_active=True
        ).join(Student).order_by(Student.surname, Student.first_name).all()
        students = [e.student for e in enrollments]
        by_student_id = {s.student_id: s for s in students}

        # 1) Exact matches on the school's own student number reserve those pupils.
        matched = [None] * len(parsed)
        exact_num = [False] * len(parsed)
        used = set()
        for i, p in enumerate(parsed):
            s = by_student_id.get(p['identifier'])
            if s and s.id not in used:
                matched[i], exact_num[i] = s, True
                used.add(s.id)
        # 2) Greedy unique fuzzy match for the rest, so two pasted rows never claim
        #    the same pupil (that collision silently dropped scores at save time).
        rest_idx = [i for i in range(len(parsed)) if matched[i] is None]
        pool = [s for s in students if s.id not in used]
        fuzzy = match_students_unique([parsed[i]['identifier'] for i in rest_idx], pool)
        for i, (s, _sc) in zip(rest_idx, fuzzy):
            matched[i] = s

        rows = []
        for i, p in enumerate(parsed):
            cell_map = {}
            for (at, _mx), value in zip(sheet_cols, p['cells']):
                cell_map[at.id] = value
            rows.append({
                'student_num': p['identifier'] if exact_num[i] else '',
                'name': p['identifier'],
                'matched_id': matched[i].id if matched[i] else None,
                'cells': cell_map,
            })

        return render_template('subjects/scoresheet_review.html',
            term_id=ctx['term_id'], assignment_id=assignment_id, class_subject_id=ctx['class_subject_id'],
            assignment=assignment, class_subject=cs,
            columns=sheet_cols, rows=rows, students=students,
        )

    return render_template('subjects/scoresheet_paste.html', **ctx)


@subjects_bp.route('/scores/scan/save', methods=['POST'])
@login_required
def scoresheet_save():
    """Persist the reviewed score-sheet grid as StudentScores."""
    import re as _re

    if not can_enter_results() and not is_admin():
        flash('You do not have permission to enter scores.', 'error')
        return redirect(url_for('main.dashboard'))

    term_id = request.form.get('term_id', type=int)
    assignment_id = request.form.get('assignment_id', type=int)
    class_subject_id = request.form.get('class_subject_id', type=int)
    row_count = request.form.get('row_count', type=int) or 0

    assignment = db.session.get(ClassArmAssignment, assignment_id) if assignment_id else None
    class_subject = db.session.get(ClassSubject, class_subject_id) if class_subject_id else None

    if not (assignment and class_subject):
        flash('Missing class/subject context.', 'error')
        return redirect(url_for('subjects.scoresheet_scan'))

    # Teachers may only save scores for a subject they teach in this class.
    if not can_enter_results(assignment_id, class_subject.subject_id):
        flash('You can only enter scores for the subjects you teach in this class.', 'error')
        return redirect(url_for('subjects.scoresheet_scan'))

    if results_locked(term_id):
        flash('Results for this term are published — ask an admin to unlock them '
              'before saving scanned scores.', 'error')
        return redirect(url_for('subjects.scores_entry', term_id=term_id,
                                assignment_id=assignment_id, class_subject_id=class_subject_id))

    auto_id_re = _re.compile(r'^STU\d{3,}$')

    # Read the cell values straight from the submitted field names rather than
    # re-deriving which assessment-type ids "should" be present. The review grid
    # rendered inputs named ``cell_<row>_<assessment_type_id>``; a tenant with two
    # active assessment types sharing a short name (e.g. two "EXAM") could make a
    # re-derived column list resolve to a *different* id here than the one the grid
    # used, so every cell read back empty and the save reported 0 (or a partial
    # count) scores. Trusting the posted keys makes the save match the grid exactly.
    cell_key_re = _re.compile(r'^cell_(\d+)_(\d+)$')
    row_cells = {}                                   # row_index -> {at_id: raw_value}
    for key, val in request.form.items():
        m = cell_key_re.match(key)
        if m:
            row_cells.setdefault(int(m.group(1)), {})[int(m.group(2))] = val

    # Effective max per assessment-type id (per-term settings honoured), so an
    # out-of-range value is still rejected correctly however the column was named.
    from utils.assessments import effective_max
    from models import AssessmentType
    at_by_id = {at.id: at for at in AssessmentType.query.filter_by(is_active=True).all()}

    def _max_for(at_id):
        at = at_by_id.get(at_id)
        return effective_max(class_subject.subject, at, term=term_id) if at else None

    adopted = 0
    warnings = []
    dropped = []          # rows that carried scores but weren't linked to a student
    items = []            # (student_id, at_id, raw, max_score)
    matched_rows = 0
    cells_with_values = 0
    bad_max = 0           # matched cells whose column had no resolvable max (unknown at id)

    try:
        for r in range(row_count):
            cells = row_cells.get(r, {})
            student_pk = request.form.get(f'student_{r}', type=int)
            if not student_pk:
                # A row the user left unmatched: warn (don't silently lose it) if it
                # actually carried any scores.
                if any((v or '').strip() for v in cells.values()):
                    dropped.append((request.form.get(f'rowname_{r}') or '').strip() or f'row {r + 1}')
                continue
            student = db.session.get(Student, student_pk)
            if not student:
                continue
            matched_rows += 1

            # Adopt the scanned student number only when the student currently
            # has an auto-generated STU##### id (never overwrite a manual id).
            scanned = (request.form.get(f'studentnum_{r}') or '').strip()
            if scanned and scanned != student.student_id and auto_id_re.match(student.student_id or ''):
                clash = Student.query.filter(
                    Student.student_id == scanned, Student.id != student.id).first()
                if clash:
                    warnings.append(f"{student.full_name}: number {scanned} already used by {clash.full_name}")
                else:
                    student.student_id = scanned
                    adopted += 1

            for at_id, raw in cells.items():
                if (raw or '').strip():
                    cells_with_values += 1
                mx = _max_for(at_id)
                if mx is None:
                    bad_max += 1
                items.append((student.id, at_id, raw, mx))

        counts = persist_scores(term_id, assignment_id, class_subject_id,
                                class_subject.subject_id, items, allow_delete=False)
        if counts is None:                       # term became locked between load and save
            db.session.rollback()
            flash('Results for this term are locked — nothing was saved.', 'error')
            return redirect(url_for('subjects.scores_entry', term_id=term_id,
                                    assignment_id=assignment_id, class_subject_id=class_subject_id))
        db.session.commit()

        # --- server-side diagnostic (terminal / server log, not the page) --------
        # A compact breakdown of exactly what the save did, so a "Saved N scores"
        # that looks too low can be traced without guessing. Includes the column
        # ids the grid posted vs the ids the sheet re-derives, to spot any drift,
        # plus a few sample cells.
        try:
            from flask import current_app
            posted_at_ids = sorted({aid for cs in row_cells.values() for aid in cs})
            sheet_at_ids = sorted(at.id for at, _ in _sheet_columns(class_subject, term=term_id))
            sample = []
            for r in range(min(row_count, 3)):
                sample.append({'row': r, 'student': request.form.get(f'student_{r}'),
                               'name': request.form.get(f'rowname_{r}'),
                               'cells': row_cells.get(r, {})})
            unchanged = max(cells_with_values - counts['rejected'] - counts['blocked'] - counts['saved'], 0)
            current_app.logger.warning(
                'SCORESHEET_SAVE_DIAG term=%s asg=%s cs=%s subject_id=%s | rows=%s matched=%s '
                'cells_with_values=%s | saved=%s rejected=%s blocked=%s unknown_col=%s '
                'already_had_value=%s | posted_col_ids=%s sheet_col_ids=%s | sample=%s',
                term_id, assignment_id, class_subject_id, class_subject.subject_id,
                row_count, matched_rows, cells_with_values, counts['saved'], counts['rejected'],
                counts['blocked'], bad_max, unchanged, posted_at_ids, sheet_at_ids, sample)
        except Exception:
            pass

        msg = f'Saved {counts["saved"]} scores.'
        if counts['rejected']:
            msg += f' {counts["rejected"]} skipped (outside the allowed range).'
        if counts['blocked']:
            msg += f' {counts["blocked"]} left unchanged (edit permission required).'
        flash(msg, 'success' if (counts['saved'] and not dropped) else 'warning')
        if adopted:
            flash(f'Adopted scanned student number for {adopted} student(s).', 'info')
        if dropped:
            names = ', '.join(dropped[:10]) + ('…' if len(dropped) > 10 else '')
            flash(f'{len(dropped)} row(s) had scores but were NOT saved because no student was '
                  f'selected: {names}. Go back, pick the student for each, and save again.', 'error')
        for w in warnings[:5]:
            flash(w, 'warning')
    except Exception as e:
        db.session.rollback()
        flash(f'Error saving scores: {e}', 'error')

    return redirect(url_for('subjects.scores_entry',
        term_id=term_id, assignment_id=assignment_id, class_subject_id=class_subject_id))


# ============================================================================
# BROADSHEET IMPORT — upload an Excel/CSV of per-subject TOTALS for a class,
# map columns to subjects + rows to students, auto-break each total into the
# term's assessment components, preview/edit, then save.
# ============================================================================

@subjects_bp.route('/scores/broadsheet-import', methods=['GET', 'POST'])
@login_required
def broadsheet_import():
    """Upload a whole-class broadsheet (subject totals) and review the mapping."""
    if not can_enter_results() and not is_admin():
        flash('You do not have permission to enter scores.', 'error')
        return redirect(url_for('main.dashboard'))

    ctx = _scan_selector_context()
    if request.method == 'POST':
        term_id = ctx['term_id']
        assignment_id = ctx['assignment_id']
        assignment = db.session.get(ClassArmAssignment, assignment_id) if assignment_id else None
        if not (assignment and term_id):
            flash('Select a term and class before uploading.', 'error')
            return render_template('subjects/broadsheet_import.html', **ctx)
        if not can_access_class(assignment_id):
            flash('You do not have access to this class.', 'error')
            return redirect(url_for('subjects.broadsheet_import'))

        upload = request.files.get('file')
        if not upload or not upload.filename:
            flash('No file selected.', 'error')
            return render_template('subjects/broadsheet_import.html', **ctx)
        from utils.uploads import ext_ok, IMAGE_EXTS
        data = upload.read()
        is_image = ext_ok(upload.filename, IMAGE_EXTS)
        if not (is_image or ext_ok(upload.filename, {'.xlsx', '.xlsm', '.xls', '.csv'})):
            flash('Please upload an Excel/CSV file or a photo (JPG/PNG) of the broadsheet.', 'error')
            return render_template('subjects/broadsheet_import.html', **ctx)

        from utils.broadsheet_import import parse_table, guess_name_column, match_subject
        ocr_flags, ocr_review_count, image_data_uri = {}, 0, None
        if is_image:
            # OCR the photographed broadsheet into the same headers+rows table,
            # using the school's chosen engine (Claude vision / Tesseract).
            from utils.table_ocr import ocr_table_rich
            from utils.ocr_engine import engine_order
            if not engine_order():
                flash('No OCR engine is available. Configure one in Settings → AI Vision OCR, '
                      'or upload an Excel/CSV file instead.', 'error')
                return render_template('subjects/broadsheet_import.html', **ctx)
            # Subject names hint the table reconstructor's header detection.
            _subs = ClassSubject.query.filter_by(
                term_id=term_id, class_id=assignment.class_id, is_active=True
            ).filter((ClassSubject.arm_id == None) | (ClassSubject.arm_id == assignment.arm_id)  # noqa: E711
                     ).join(Subject).all()
            expected = ['Student Name'] + [cs.subject.name for cs in _subs] \
                + [cs.subject.short_name for cs in _subs if cs.subject.short_name]
            tried = engine_order()
            rich = ocr_table_rich(data, upload.mimetype or 'image/png', expected_headers=expected)
            if not rich:
                names = {'claude': 'Claude vision', 'tesseract': 'Tesseract'}
                who = ', '.join(names.get(e, e) for e in tried) or 'the configured engine'
                flash(f'{who} could not read a table from that image. Try a clearer, straight '
                      f'photo, switch the engine in Settings → AI Vision OCR, or upload the '
                      f'Excel/CSV instead.', 'warning')
                return render_template('subjects/broadsheet_import.html', **ctx)
            parsed = {'headers': rich['headers'], 'rows': rich['rows']}
            ocr_flags = rich.get('cell_flags') or {}
            ocr_review_count = rich.get('review_count') or 0
            import base64
            image_data_uri = 'data:%s;base64,%s' % (
                upload.mimetype or 'image/png', base64.b64encode(data).decode())
        else:
            try:
                parsed = parse_table(data, upload.filename)
            except Exception as e:
                flash(f'Could not read the file: {e}', 'error')
                return render_template('subjects/broadsheet_import.html', **ctx)
        headers = parsed['headers']
        rows = parsed['rows']
        if not headers or not rows:
            flash('No table with a header row and data was found. Check the file/photo and try again.', 'warning')
            return render_template('subjects/broadsheet_import.html', **ctx)

        # Class subjects (mapping targets) + their term component config.
        class_subjects = ClassSubject.query.filter_by(
            term_id=term_id, class_id=assignment.class_id, is_active=True
        ).filter((ClassSubject.arm_id == None) | (ClassSubject.arm_id == assignment.arm_id)  # noqa: E711
                 ).join(Subject).order_by(Subject.name).all()
        subjects = [cs.subject for cs in class_subjects]
        cs_by_subject = {cs.subject_id: cs for cs in class_subjects}

        # Auto-guess column→subject and the name column.
        name_col = guess_name_column(headers)
        col_map = []
        for idx, hdr in enumerate(headers):
            s = None if idx == name_col else match_subject(hdr, subjects)
            col_map.append(cs_by_subject[s.id].id if s and s.id in cs_by_subject else '')

        # Enrolled students + auto name match.
        from utils.waec_ocr import match_students_unique
        enrollments = StudentEnrollment.query.filter_by(
            class_arm_assignment_id=assignment_id, is_active=True
        ).join(Student).order_by(Student.surname, Student.first_name).all()
        students = [e.student for e in enrollments]
        row_names = [(r[name_col] if name_col is not None and name_col < len(r) else '') for r in rows]
        matches = match_students_unique(row_names, students)
        matched_ids = [(s.id if s else '') for s, _sc in matches]

        # Component config per class-subject for the JS breakdown preview.
        subject_components = {}
        for cs in class_subjects:
            cols = _sheet_columns(cs, term=term_id)
            subject_components[cs.id] = {
                'subject': cs.subject.name,
                'components': [{'at_id': at.id, 'name': (at.short_name or at.name), 'max': mx}
                               for at, mx in cols],
            }

        return render_template('subjects/broadsheet_review.html',
            term_id=term_id, assignment_id=assignment_id, assignment=assignment,
            headers=headers, rows=rows, name_col=(name_col if name_col is not None else ''),
            col_map=col_map, class_subjects=class_subjects,
            students=students, matched_ids=matched_ids,
            subject_components=subject_components,
            cell_flags=ocr_flags, review_count=ocr_review_count, image_data_uri=image_data_uri,
            save_url=url_for('subjects.broadsheet_save'))

    return render_template('subjects/broadsheet_import.html', **ctx)


@subjects_bp.route('/scores/broadsheet-import/save', methods=['POST'])
@login_required
def broadsheet_save():
    """Persist a reviewed broadsheet: for each (student, mapped subject) total,
    break it into the term's components and write the component scores."""
    if not can_enter_results() and not is_admin():
        flash('You do not have permission to enter scores.', 'error')
        return redirect(url_for('main.dashboard'))

    term_id = request.form.get('term_id', type=int)
    assignment_id = request.form.get('assignment_id', type=int)
    assignment = db.session.get(ClassArmAssignment, assignment_id) if assignment_id else None
    if not (assignment and term_id) or not can_access_class(assignment_id):
        flash('Invalid class/term, or no access.', 'error')
        return redirect(url_for('subjects.broadsheet_import'))

    from utils.score_breakdown import breakdown_for_subject
    import json as _json
    # payload: [{student_id, cells: {class_subject_id: total}}]
    try:
        payload = _json.loads(request.form.get('payload') or '[]')
    except ValueError:
        payload = []

    cs_ids = {int(k) for row in payload for k in (row.get('cells') or {}).keys()}
    cs_by_id = {cs.id: cs for cs in ClassSubject.query.filter(ClassSubject.id.in_(cs_ids)).all()
                if cs_ids} if cs_ids else {}

    saved = subjects_touched = skipped = 0
    blocked = False
    for row in payload:
        sid = row.get('student_id')
        if not sid:
            continue
        student = db.session.get(Student, int(sid))
        if not student:
            continue
        for cs_id_s, total in (row.get('cells') or {}).items():
            cs = cs_by_id.get(int(cs_id_s))
            if not cs:
                continue
            if str(total).strip() == '':
                continue
            comps = breakdown_for_subject(cs.subject, term_id, total)
            items = [(student.id, at.id, str(sc), mx) for at, mx, sc in comps]
            counts = persist_scores(term_id, assignment_id, cs.id, cs.subject_id, items,
                                    allow_delete=False)
            if counts is None:
                blocked = True
                continue
            saved += counts['saved']
            subjects_touched += 1
    if not blocked:
        db.session.commit()
        # Recompute term results/positions for the class.
        from utils.report_card import compute_term_summaries
        compute_term_summaries(term_id, assignment.class_id)
        from utils.results_analytics import bust as _bust
        _bust(term_id, assignment_id)
        flash(f'Imported {saved} component score(s) across {subjects_touched} '
              f'student-subject total(s).', 'success')
    else:
        db.session.rollback()
        flash('Results for this term are published — ask an admin to unlock them first.', 'error')
    return redirect(url_for('subjects.scores_entry', term_id=term_id, assignment_id=assignment_id))
