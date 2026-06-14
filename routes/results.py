"""
Results management routes - WAEC, JAMB, and Analytics Dashboard
Comprehensive academic performance tracking and analysis
"""
from flask import Blueprint, render_template, request, redirect, url_for, flash, jsonify, Response, abort
from utils.helpers import get_active_term, get_active_session
from collections import defaultdict
from models import (db, Student, WAECResult, JAMBResult, UniversityCutoff, SchoolSettings, StudentEnrollment,
                    ClassArmAssignment, TermSummary)
import json as _json
from utils.access_control import login_required, admin_required
from utils.branch_scope import require_branch_access
from utils.audit import log_action
from utils.helpers import (
    WAEC_SUBJECTS, WAEC_GRADES, WAEC_DEFAULT_SUBJECTS, STREAM_WAEC_SUBJECTS, get_sss3_students,
    student_subject_map,
)
from datetime import date as _date
from sqlalchemy import func
from sqlalchemy.orm import joinedload
from utils.web_exports import xlsx_response
from utils.analytics_service import AcademicAnalytics

results_bp = Blueprint('results', __name__, url_prefix='/results')


@results_bp.route('/')
@login_required
def index():
    """Results main page with overview"""
    waec_count = WAECResult.query.count()
    jamb_count = JAMBResult.query.count()
    
    waec_years = db.session.query(WAECResult.exam_year).distinct().order_by(WAECResult.exam_year.desc()).all()
    jamb_years = db.session.query(JAMBResult.exam_year).distinct().order_by(JAMBResult.exam_year.desc()).all()
    
    return render_template('results/index.html',
        waec_count=waec_count,
        jamb_count=jamb_count,
        waec_years=[y[0] for y in waec_years],
        jamb_years=[y[0] for y in jamb_years]
    )


# ============================================================================
# WAEC RESULTS - COMPREHENSIVE DASHBOARD
# ============================================================================

@results_bp.route('/waec')
@login_required
def waec_list():
    """Comprehensive WAEC dashboard with filtering and analytics"""
    exam_year = request.args.get('year', type=int)
    min_a1 = request.args.get('min_a1', type=int)
    min_credits = request.args.get('min_credits', type=int)
    subject_filter = request.args.get('subject', '')
    grade_filter = request.args.get('grade', '')
    search = request.args.get('search', '').strip()
    sort_by = request.args.get('sort', 'name')
    sort_order = request.args.get('order', 'asc')
    view_mode = request.args.get('view', 'students')
    
    years = db.session.query(WAECResult.exam_year).distinct().order_by(WAECResult.exam_year.desc()).all()
    years = [y[0] for y in years]
    
    if not exam_year and years:
        exam_year = years[0]
    
    students_data = []
    subject_stats = []
    school_stats = None
    grade_distribution = {g: 0 for g in WAEC_GRADES}
    top_by_grade = {g: [] for g in WAEC_GRADES}
    subjects_by_grade = []
    
    if exam_year:
        from utils.branch_scope import viewing_branch_id
        school_stats = AcademicAnalytics.get_waec_school_statistics(exam_year, viewing_branch_id())

        if school_stats:
            grade_distribution = school_stats['grade_distribution']
            subject_stats = school_stats['subject_analysis']
        
        # Get all results for the year for detailed analysis (branch-scoped)
        from utils.branch_scope import scope_by_student
        all_results = scope_by_student(
            WAECResult.query.filter_by(exam_year=exam_year), WAECResult).all()
        
        # Build subject-grade matrix
        subject_grade_counts = defaultdict(lambda: {g: 0 for g in WAEC_GRADES})
        for r in all_results:
            subject_grade_counts[r.subject][r.grade] += 1
        
        # Convert to list for template
        subjects_by_grade = []
        for subject, grades in subject_grade_counts.items():
            total = sum(grades.values())
            subjects_by_grade.append({
                'subject': subject,
                'grades': grades,
                'total': total,
                'a1_count': grades.get('A1', 0),
                'b2_count': grades.get('B2', 0),
                'b3_count': grades.get('B3', 0),
                'credit_count': sum(grades.get(g, 0) for g in ['A1', 'B2', 'B3', 'C4', 'C5', 'C6']),
                'pass_rate': round(sum(grades.get(g, 0) for g in ['A1', 'B2', 'B3', 'C4', 'C5', 'C6']) / total * 100, 1) if total > 0 else 0
            })
        
        # Sort subjects based on parameters
        if sort_by == 'subject_a1':
            subjects_by_grade.sort(key=lambda x: x['a1_count'], reverse=(sort_order == 'desc'))
        elif sort_by == 'subject_b2':
            subjects_by_grade.sort(key=lambda x: x['b2_count'], reverse=(sort_order == 'desc'))
        elif sort_by == 'subject_credits':
            subjects_by_grade.sort(key=lambda x: x['credit_count'], reverse=(sort_order == 'desc'))
        elif sort_by == 'subject_pass':
            subjects_by_grade.sort(key=lambda x: x['pass_rate'], reverse=(sort_order == 'desc'))
        else:
            subjects_by_grade.sort(key=lambda x: x['a1_count'], reverse=True)
        
        from utils.branch_scope import scope_query
        base_query = scope_query(
            db.session.query(Student).join(WAECResult).filter(
                WAECResult.exam_year == exam_year),
            Student)

        if search:
            term = f"%{search}%"
            base_query = base_query.filter(
                db.or_(
                    Student.first_name.ilike(term),
                    Student.surname.ilike(term),
                    Student.middle_name.ilike(term),
                    Student.student_id.ilike(term),
                )
            )

        if subject_filter and grade_filter:
            base_query = base_query.filter(
                WAECResult.subject == subject_filter,
                WAECResult.grade == grade_filter
            )
        elif subject_filter:
            base_query = base_query.filter(WAECResult.subject == subject_filter)
        elif grade_filter:
            base_query = base_query.filter(WAECResult.grade == grade_filter)
        
        students = base_query.distinct().all()

        # Batch-load every matching student's results for the year in one query
        # (instead of one query per student).
        results_by_student = {}
        if students:
            for r in WAECResult.query.filter(
                    WAECResult.student_id.in_([s.id for s in students]),
                    WAECResult.exam_year == exam_year).all():
                results_by_student.setdefault(r.student_id, []).append(r)

        for student in students:
            results = results_by_student.get(student.id, [])

            # Count each grade
            grade_counts = {g: sum(1 for r in results if r.grade == g) for g in WAEC_GRADES}
            a1_count = grade_counts['A1']
            b2_count = grade_counts['B2']
            b3_count = grade_counts['B3']
            credit_count = sum(grade_counts[g] for g in ['A1', 'B2', 'B3', 'C4', 'C5', 'C6'])
            total_points = sum(WAECResult.grade_to_points(r.grade) for r in results)
            
            if min_a1 and a1_count < min_a1:
                continue
            if min_credits and credit_count < min_credits:
                continue
            
            student_data = {
                'id': student.id,
                'student_id': student.student_id,
                'name': student.full_name,
                'total_subjects': len(results),
                'a1_count': a1_count,
                'b2_count': b2_count,
                'b3_count': b3_count,
                'grade_counts': grade_counts,
                'credit_count': credit_count,
                'total_points': total_points,
                'avg_points': round(total_points / len(results), 2) if results else 0,
                'results': [{'subject': r.subject, 'grade': r.grade} for r in results]
            }
            students_data.append(student_data)
        
        # Build top performers by each grade
        for grade in WAEC_GRADES:
            grade_key = f'{grade.lower()}_count'
            sorted_students = sorted(
                [s for s in students_data if s['grade_counts'].get(grade, 0) > 0],
                key=lambda x: x['grade_counts'].get(grade, 0),
                reverse=True
            )[:10]
            top_by_grade[grade] = sorted_students
        
        # Sort students based on parameters
        if sort_by == 'a1':
            students_data.sort(key=lambda x: x['a1_count'], reverse=(sort_order == 'desc'))
        elif sort_by == 'b2':
            students_data.sort(key=lambda x: x['b2_count'], reverse=(sort_order == 'desc'))
        elif sort_by == 'b3':
            students_data.sort(key=lambda x: x['b3_count'], reverse=(sort_order == 'desc'))
        elif sort_by == 'credits':
            students_data.sort(key=lambda x: x['credit_count'], reverse=(sort_order == 'desc'))
        elif sort_by == 'points':
            students_data.sort(key=lambda x: x['avg_points'], reverse=(sort_order != 'desc'))
        else:
            students_data.sort(key=lambda x: x['name'], reverse=(sort_order == 'desc'))
    
    from utils.branch_scope import viewing_branch_id
    yoy_data = AcademicAnalytics.get_year_over_year_comparison(viewing_branch_id())
    
    return render_template('results/waec_dashboard.html',
        students=students_data,
        years=years,
        selected_year=exam_year,
        subjects=WAEC_SUBJECTS,
        grades=WAEC_GRADES,
        subject_filter=subject_filter,
        grade_filter=grade_filter,
        search=search,
        min_a1=min_a1,
        min_credits=min_credits,
        sort_by=sort_by,
        sort_order=sort_order,
        view_mode=view_mode,
        school_stats=school_stats,
        subject_stats=subject_stats,
        grade_distribution=grade_distribution,
        top_by_grade=top_by_grade,
        subjects_by_grade=subjects_by_grade,
        yoy_data=yoy_data
    )


@results_bp.route('/waec/add', methods=['GET', 'POST'])
@login_required
def add_waec():
    """Add WAEC results for a student"""
    students = get_sss3_students()

    if request.method == 'POST':
        try:
            student_id = request.form.get('student_id', type=int)
            exam_year = request.form.get('exam_year', type=int)
            
            if not student_id or not exam_year:
                flash('Student and exam year are required.', 'error')
                return redirect(url_for('results.add_waec'))
            
            subjects = request.form.getlist('subject[]')
            grades = request.form.getlist('grade[]')
            
            results_added = 0
            for i, subject in enumerate(subjects):
                if subject and i < len(grades) and grades[i]:
                    existing = WAECResult.query.filter_by(
                        student_id=student_id,
                        exam_year=exam_year,
                        subject=subject
                    ).first()
                    
                    if existing:
                        existing.grade = grades[i]
                    else:
                        result = WAECResult(
                            student_id=student_id,
                            exam_year=exam_year,
                            subject=subject,
                            grade=grades[i]
                        )
                        db.session.add(result)
                    
                    results_added += 1
            
            db.session.commit()
            flash(f'{results_added} WAEC results saved!', 'success')
            return redirect(url_for('results.view_waec_student', student_id=student_id))
            
        except Exception as e:
            db.session.rollback()
            flash(f'Error saving results: {str(e)}', 'error')
    
    return render_template('results/add_waec.html',
        students=students,
        subjects=WAEC_SUBJECTS,
        grades=WAEC_GRADES,
        default_subjects=WAEC_DEFAULT_SUBJECTS,
        stream_defaults=STREAM_WAEC_SUBJECTS,
        subject_map=student_subject_map(students),
        current_year=_date.today().year
    )


@results_bp.route('/waec/scan', methods=['GET', 'POST'])
@login_required
def scan_waec():
    """Upload a WAEC result image, OCR it, and review before saving."""
    from utils.waec_ocr import (
        tesseract_available, extract_text, parse_waec_result, match_student,
        pdf_available, extract_text_from_pdf, pdf_first_page_png,
        vision_available, vision_extract,
    )
    import base64

    students = get_sss3_students()

    if request.method == 'POST':
        file = request.files.get('result_image')
        if not file or not file.filename:
            flash('Please choose a file to upload.', 'error')
            return redirect(url_for('results.scan_waec'))

        filename = (file.filename or '').lower()
        is_pdf = (file.mimetype == 'application/pdf') or filename.endswith('.pdf')
        file_bytes = file.read()

        try:
            if is_pdf:
                if not pdf_available():
                    flash('PDF support (PyMuPDF) is not installed on the server.', 'error')
                    return redirect(url_for('results.scan_waec'))
                text = extract_text_from_pdf(file_bytes)
                try:
                    preview_bytes = pdf_first_page_png(file_bytes)
                    preview = 'data:image/png;base64,' + base64.b64encode(preview_bytes).decode()
                except Exception:
                    preview = ''
            else:
                if not tesseract_available():
                    flash('OCR engine (Tesseract) is not installed on the server. '
                          'Install "tesseract-ocr" to scan images.', 'error')
                    return redirect(url_for('results.scan_waec'))
                text = extract_text(file_bytes)
                mime = file.mimetype or 'image/png'
                preview = f"data:{mime};base64,{base64.b64encode(file_bytes).decode()}"
        except Exception as e:
            flash(f'Could not read the file: {e}', 'error')
            return redirect(url_for('results.scan_waec'))

        # Prefer the Claude-vision reading when enabled; fall back to Tesseract.
        parsed = None
        if not is_pdf and vision_available():
            parsed = vision_extract(file_bytes, 'waec', file.mimetype or 'image/png')
            if parsed:
                text = '(read by Claude vision)\n' + text
        if not parsed:
            parsed = parse_waec_result(text)
        matched, score = match_student(parsed['name'], students)

        return render_template('results/waec_scan_review.html',
            students=students,
            subjects=WAEC_SUBJECTS,
            grades=WAEC_GRADES,
            parsed=parsed,
            matched=matched,
            match_score=score,
            preview=preview,
            current_year=parsed.get('year') or _date.today().year,
            raw_text=text
        )

    return render_template('results/waec_scan.html',
        ocr_ready=tesseract_available(),
        pdf_ready=pdf_available()
    )


@results_bp.route('/waec/student/<int:student_id>')
@login_required
def view_waec_student(student_id):
    """View comprehensive WAEC profile for a student"""
    student = db.get_or_404(Student, student_id)
    require_branch_access(student.branch_id)
    waec_summary = AcademicAnalytics.get_student_waec_summary(student_id)
    risk_assessment = AcademicAnalytics.calculate_student_risk_score(student_id)
    jamb_prediction = AcademicAnalytics.predict_jamb_score(student_id)
    recommendations = AcademicAnalytics.get_subject_recommendations(student_id)
    
    return render_template('results/view_waec_student.html',
        student=student,
        waec_summary=waec_summary,
        risk_assessment=risk_assessment,
        jamb_prediction=jamb_prediction,
        recommendations=recommendations,
        grades=WAEC_GRADES
    )


@results_bp.route('/waec/student/<int:student_id>/edit/<int:year>', methods=['GET', 'POST'])
@login_required
def edit_waec(student_id, year):
    """Edit WAEC results for a student/year"""
    student = db.get_or_404(Student, student_id)
    require_branch_access(student.branch_id)
    
    if request.method == 'POST':
        try:
            WAECResult.query.filter_by(student_id=student_id, exam_year=year).delete()
            
            subjects = request.form.getlist('subject[]')
            grades = request.form.getlist('grade[]')
            
            for i, subject in enumerate(subjects):
                if subject and i < len(grades) and grades[i]:
                    result = WAECResult(
                        student_id=student_id,
                        exam_year=year,
                        subject=subject,
                        grade=grades[i]
                    )
                    db.session.add(result)
            
            db.session.commit()
            log_action('results.waec_edit', detail=f'{year}', target=student)
            flash('WAEC results updated!', 'success')
            return redirect(url_for('results.view_waec_student', student_id=student_id))
            
        except Exception as e:
            db.session.rollback()
            flash(f'Error: {str(e)}', 'error')
    
    results = WAECResult.query.filter_by(student_id=student_id, exam_year=year).all()
    results_dict = {r.subject: r.grade for r in results}
    
    return render_template('results/edit_waec.html',
        student=student,
        year=year,
        results_dict=results_dict,
        subjects=WAEC_SUBJECTS,
        grades=WAEC_GRADES
    )


# ============================================================================
# WAEC DELETE ROUTES
# ============================================================================

@results_bp.route('/waec/student/<int:student_id>/delete/<int:year>', methods=['POST'])
@admin_required
def delete_waec(student_id, year):
    """Delete all WAEC results for a student in a given year"""
    student = db.get_or_404(Student, student_id)
    require_branch_access(student.branch_id)
    
    try:
        deleted = WAECResult.query.filter_by(student_id=student_id, exam_year=year).delete()
        db.session.commit()
        
        if deleted:
            log_action('results.waec_delete', detail=f'{deleted} result(s), {year}', target=student)
            flash(f'Deleted {deleted} WAEC results for {student.full_name} ({year}).', 'success')
        else:
            flash('No results found to delete.', 'warning')
            
    except Exception as e:
        db.session.rollback()
        flash(f'Error deleting results: {str(e)}', 'error')
    
    return redirect(url_for('results.waec_list', year=year))


@results_bp.route('/waec/result/<int:result_id>/delete', methods=['POST'])
@admin_required
def delete_waec_single(result_id):
    """Delete a single WAEC result entry"""
    result = db.get_or_404(WAECResult, result_id)
    student_id = result.student_id
    year = result.exam_year
    subject = result.subject
    
    try:
        db.session.delete(result)
        db.session.commit()
        flash(f'Deleted {subject} result.', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Error: {str(e)}', 'error')
    
    return redirect(url_for('results.view_waec_student', student_id=student_id))


# ============================================================================
# JAMB RESULTS - COMPREHENSIVE DASHBOARD
# ============================================================================

@results_bp.route('/jamb')
@login_required
def jamb_list():
    """Comprehensive JAMB dashboard with filtering and analytics"""
    exam_year = request.args.get('year', type=int)
    min_score = request.args.get('min_score', type=int)
    max_score = request.args.get('max_score', type=int)
    search = request.args.get('search', '').strip()
    sort_by = request.args.get('sort', 'score')
    sort_order = request.args.get('order', 'desc')
    view_mode = request.args.get('view', 'students')
    
    years = db.session.query(JAMBResult.exam_year).distinct().order_by(JAMBResult.exam_year.desc()).all()
    years = [y[0] for y in years]
    
    if not exam_year and years:
        exam_year = years[0]
    
    students_data = []
    school_stats = None
    correlation = None
    subject_performance = []
    
    if exam_year:
        from utils.branch_scope import viewing_branch_id
        _bid = viewing_branch_id()
        school_stats = AcademicAnalytics.get_jamb_school_statistics(exam_year, _bid)
        correlation = AcademicAnalytics.calculate_waec_jamb_correlation(exam_year, _bid)
        
        # Get all results for subject analysis
        all_results = JAMBResult.query.filter_by(exam_year=exam_year).all()
        
        # Build subject performance data
        subject_scores = defaultdict(list)
        for r in all_results:
            if r.subject1 and r.subject1_score:
                subject_scores[r.subject1].append(r.subject1_score)
            if r.subject2 and r.subject2_score:
                subject_scores[r.subject2].append(r.subject2_score)
            if r.subject3 and r.subject3_score:
                subject_scores[r.subject3].append(r.subject3_score)
            if r.subject4 and r.subject4_score:
                subject_scores[r.subject4].append(r.subject4_score)
        
        # Calculate subject statistics
        for subject, scores in subject_scores.items():
            if scores:
                avg_score = sum(scores) / len(scores)
                above_50 = sum(1 for s in scores if s >= 50)
                above_70 = sum(1 for s in scores if s >= 70)
                subject_performance.append({
                    'subject': subject,
                    'count': len(scores),
                    'avg_score': round(avg_score, 1),
                    'max_score': max(scores),
                    'min_score': min(scores),
                    'above_50': above_50,
                    'above_50_pct': round(above_50 / len(scores) * 100, 1),
                    'above_70': above_70,
                    'above_70_pct': round(above_70 / len(scores) * 100, 1)
                })
        
        # Sort subjects by performance
        if sort_by == 'subject_avg':
            subject_performance.sort(key=lambda x: x['avg_score'], reverse=(sort_order == 'desc'))
        elif sort_by == 'subject_above50':
            subject_performance.sort(key=lambda x: x['above_50_pct'], reverse=(sort_order == 'desc'))
        elif sort_by == 'subject_above70':
            subject_performance.sort(key=lambda x: x['above_70_pct'], reverse=(sort_order == 'desc'))
        else:
            subject_performance.sort(key=lambda x: x['avg_score'], reverse=True)
        
        query = (JAMBResult.query.filter_by(exam_year=exam_year)
                 .options(joinedload(JAMBResult.student)).join(Student))

        if search:
            term = f"%{search}%"
            query = query.filter(
                db.or_(
                    Student.first_name.ilike(term),
                    Student.surname.ilike(term),
                    Student.middle_name.ilike(term),
                    Student.student_id.ilike(term),
                )
            )

        if min_score:
            query = query.filter(JAMBResult.total_score >= min_score)
        if max_score:
            query = query.filter(JAMBResult.total_score <= max_score)
        
        if sort_by == 'score':
            if sort_order == 'desc':
                query = query.order_by(JAMBResult.total_score.desc())
            else:
                query = query.order_by(JAMBResult.total_score.asc())
        elif sort_by == 'name':
            if sort_order == 'desc':
                query = query.order_by(Student.surname.desc())
            else:
                query = query.order_by(Student.surname.asc())
        else:
            query = query.order_by(JAMBResult.total_score.desc())
        
        results = query.all()
        
        for idx, r in enumerate(results):
            subjects = []
            if r.subject1: subjects.append({'name': r.subject1, 'score': r.subject1_score or 0})
            if r.subject2: subjects.append({'name': r.subject2, 'score': r.subject2_score or 0})
            if r.subject3: subjects.append({'name': r.subject3, 'score': r.subject3_score or 0})
            if r.subject4: subjects.append({'name': r.subject4, 'score': r.subject4_score or 0})
            
            students_data.append({
                'id': r.student.id,
                'jamb_id': r.id,
                'student_id': r.student.student_id,
                'name': r.student.full_name,
                'total_score': r.total_score,
                'rank': idx + 1,
                'subjects': subjects,
                'performance_level': AcademicAnalytics._jamb_performance_level(r.total_score)
            })
    
    from utils.branch_scope import viewing_branch_id
    yoy_data = AcademicAnalytics.get_year_over_year_comparison(viewing_branch_id())
    
    return render_template('results/jamb_dashboard.html',
        students=students_data,
        years=years,
        selected_year=exam_year,
        min_score=min_score,
        max_score=max_score,
        search=search,
        sort_by=sort_by,
        sort_order=sort_order,
        view_mode=view_mode,
        school_stats=school_stats,
        correlation=correlation,
        subject_performance=subject_performance,
        yoy_data=yoy_data
    )


@results_bp.route('/jamb/add', methods=['GET', 'POST'])
@login_required
def add_jamb():
    """Add JAMB results for a student"""
    students = get_sss3_students()
    
    if request.method == 'POST':
        try:
            student_id = request.form.get('student_id', type=int)
            exam_year = request.form.get('exam_year', type=int)
            total_score = request.form.get('total_score', type=int)
            
            if not student_id or not exam_year or total_score is None:
                flash('All required fields must be filled.', 'error')
                return redirect(url_for('results.add_jamb'))
            
            if total_score < 0 or total_score > 400:
                flash('Total score must be between 0 and 400.', 'error')
                return redirect(url_for('results.add_jamb'))
            
            existing = JAMBResult.query.filter_by(student_id=student_id, exam_year=exam_year).first()
            
            if existing:
                existing.total_score = total_score
                existing.subject1 = request.form.get('subject1')
                existing.subject1_score = request.form.get('subject1_score', type=int)
                existing.subject2 = request.form.get('subject2')
                existing.subject2_score = request.form.get('subject2_score', type=int)
                existing.subject3 = request.form.get('subject3')
                existing.subject3_score = request.form.get('subject3_score', type=int)
                existing.subject4 = request.form.get('subject4')
                existing.subject4_score = request.form.get('subject4_score', type=int)
            else:
                result = JAMBResult(
                    student_id=student_id,
                    exam_year=exam_year,
                    total_score=total_score,
                    subject1=request.form.get('subject1'),
                    subject1_score=request.form.get('subject1_score', type=int),
                    subject2=request.form.get('subject2'),
                    subject2_score=request.form.get('subject2_score', type=int),
                    subject3=request.form.get('subject3'),
                    subject3_score=request.form.get('subject3_score', type=int),
                    subject4=request.form.get('subject4'),
                    subject4_score=request.form.get('subject4_score', type=int)
                )
                db.session.add(result)
            
            db.session.commit()
            flash('JAMB result saved!', 'success')
            return redirect(url_for('results.jamb_list'))
            
        except Exception as e:
            db.session.rollback()
            flash(f'Error: {str(e)}', 'error')
    
    return render_template('results/add_jamb.html',
        students=students,
        subjects=WAEC_SUBJECTS,
        subject_map=student_subject_map(students),
        current_year=_date.today().year
    )


@results_bp.route('/jamb/scan', methods=['GET', 'POST'])
@login_required
def scan_jamb():
    """Upload a JAMB result image/PDF, OCR it, and review before saving."""
    from utils.waec_ocr import (
        tesseract_available, extract_text, parse_jamb_result, match_student,
        pdf_available, extract_text_from_pdf, pdf_first_page_png,
        vision_available, vision_extract,
    )
    import base64

    students = get_sss3_students()

    if request.method == 'POST':
        file = request.files.get('result_image')
        if not file or not file.filename:
            flash('Please choose a file to upload.', 'error')
            return redirect(url_for('results.scan_jamb'))

        filename = (file.filename or '').lower()
        is_pdf = (file.mimetype == 'application/pdf') or filename.endswith('.pdf')
        file_bytes = file.read()

        try:
            if is_pdf:
                if not pdf_available():
                    flash('PDF support (PyMuPDF) is not installed on the server.', 'error')
                    return redirect(url_for('results.scan_jamb'))
                text = extract_text_from_pdf(file_bytes)
                try:
                    preview = 'data:image/png;base64,' + base64.b64encode(pdf_first_page_png(file_bytes)).decode()
                except Exception:
                    preview = ''
            else:
                if not tesseract_available():
                    flash('OCR engine (Tesseract) is not installed on the server. '
                          'Install "tesseract-ocr" to scan images.', 'error')
                    return redirect(url_for('results.scan_jamb'))
                text = extract_text(file_bytes)
                mime = file.mimetype or 'image/png'
                preview = f"data:{mime};base64,{base64.b64encode(file_bytes).decode()}"
        except Exception as e:
            flash(f'Could not read the file: {e}', 'error')
            return redirect(url_for('results.scan_jamb'))

        parsed = None
        if not is_pdf and vision_available():
            parsed = vision_extract(file_bytes, 'jamb', file.mimetype or 'image/png')
            if parsed:
                text = '(read by Claude vision)\n' + text
        if not parsed:
            parsed = parse_jamb_result(text)
        matched, score = match_student(parsed['name'], students)

        return render_template('results/jamb_scan_review.html',
            students=students,
            subjects=WAEC_SUBJECTS,
            parsed=parsed,
            matched=matched,
            match_score=score,
            preview=preview,
            current_year=parsed.get('year') or _date.today().year,
            raw_text=text
        )

    return render_template('results/jamb_scan.html',
        ocr_ready=tesseract_available(),
        pdf_ready=pdf_available()
    )


def _read_uploaded_text(file):
    """Return OCR/PDF text + a flag for whether the upload was usable."""
    from utils.waec_ocr import (tesseract_available, extract_text, pdf_available, extract_text_from_pdf)
    filename = (file.filename or '').lower()
    data = file.read()
    if (file.mimetype == 'application/pdf') or filename.endswith('.pdf'):
        if not pdf_available():
            return None
        return extract_text_from_pdf(data)
    if not tesseract_available():
        return None
    return extract_text(data)


@results_bp.route('/scan/batch', methods=['GET', 'POST'])
@login_required
def scan_batch():
    """Batch-scan several WAEC or JAMB results, each routed to its student."""
    from utils.waec_ocr import parse_waec_result, parse_jamb_result, match_student
    import json

    students = get_sss3_students()
    exam = request.values.get('exam', 'waec')
    exam = 'jamb' if exam == 'jamb' else 'waec'

    if request.method == 'POST' and request.form.get('action') == 'save':
        count = int(request.form.get('count', 0))
        saved = 0
        for i in range(count):
            if not request.form.get(f'include_{i}'):
                continue
            student_id = request.form.get(f'student_id_{i}', type=int)
            year = request.form.get(f'year_{i}', type=int)
            if not student_id or not year:
                continue
            data = json.loads(request.form.get(f'data_{i}', '{}'))
            if exam == 'waec':
                for row in data.get('subjects', []):
                    if not row.get('subject') or not row.get('grade'):
                        continue
                    existing = WAECResult.query.filter_by(student_id=student_id, exam_year=year, subject=row['subject']).first()
                    if existing:
                        existing.grade = row['grade']
                    else:
                        db.session.add(WAECResult(student_id=student_id, exam_year=year, subject=row['subject'], grade=row['grade']))
                saved += 1
            else:
                subs = data.get('subjects', [])[:4]
                total = data.get('total_score') or sum(r.get('score', 0) for r in subs)
                existing = JAMBResult.query.filter_by(student_id=student_id, exam_year=year).first()
                target = existing or JAMBResult(student_id=student_id, exam_year=year, total_score=total)
                target.total_score = total
                for n, row in enumerate(subs, 1):
                    setattr(target, f'subject{n}', row.get('subject'))
                    setattr(target, f'subject{n}_score', row.get('score'))
                if not existing:
                    db.session.add(target)
                saved += 1
        db.session.commit()
        flash(f'Saved {saved} {exam.upper()} result(s) from batch scan.', 'success')
        return redirect(url_for('results.waec_list' if exam == 'waec' else 'results.jamb_list'))

    if request.method == 'POST':
        files = request.files.getlist('result_images')
        items = []
        for f in files:
            if not f or not f.filename:
                continue
            try:
                text = _read_uploaded_text(f)
            except Exception:
                text = None
            if text is None:
                items.append({'filename': f.filename, 'error': 'Unreadable / engine missing', 'data': {}})
                continue
            parsed = parse_jamb_result(text) if exam == 'jamb' else parse_waec_result(text)
            matched, score = match_student(parsed.get('name'), students)
            items.append({
                'filename': f.filename,
                'parsed': parsed,
                'matched_id': matched.id if matched else None,
                'matched_name': matched.full_name if matched else None,
                'score': score,
                'data_json': json.dumps(parsed),
            })
        return render_template('results/scan_batch_review.html',
            exam=exam, items=items, students=students, current_year=_date.today().year)

    return render_template('results/scan_batch.html', exam=exam)


@results_bp.route('/jamb/student/<int:student_id>')
@login_required
def view_jamb_student(student_id):
    """View JAMB results for a student"""
    student = db.get_or_404(Student, student_id)
    require_branch_access(student.branch_id)
    jamb_summary = AcademicAnalytics.get_student_jamb_summary(student_id)
    waec_summary = AcademicAnalytics.get_student_waec_summary(student_id)
    
    return render_template('results/view_jamb_student.html',
        student=student,
        jamb_summary=jamb_summary,
        waec_summary=waec_summary
    )


@results_bp.route('/subject-enrolment')
@login_required
def subject_enrolment():
    """Report: how many students are enrolled for each WAEC / JAMB subject."""
    only_sss3 = request.args.get('scope', 'sss3') != 'all'
    if only_sss3:
        students = get_sss3_students()
    else:
        students = Student.query.filter_by(is_active=True).order_by(Student.surname).all()

    waec_counts = {}
    jamb_counts = {}
    waec_enrolled = 0
    jamb_enrolled = 0
    for s in students:
        wl = s.waec_subject_list
        jl = s.jamb_subject_list
        if wl:
            waec_enrolled += 1
        if jl:
            jamb_enrolled += 1
        for subj in wl:
            waec_counts[subj] = waec_counts.get(subj, 0) + 1
        for subj in jl:
            jamb_counts[subj] = jamb_counts.get(subj, 0) + 1

    waec_rows = sorted(waec_counts.items(), key=lambda x: (-x[1], x[0]))
    jamb_rows = sorted(jamb_counts.items(), key=lambda x: (-x[1], x[0]))

    return render_template('results/subject_enrolment.html',
        waec_rows=waec_rows,
        jamb_rows=jamb_rows,
        waec_enrolled=waec_enrolled,
        jamb_enrolled=jamb_enrolled,
        student_count=len(students),
        only_sss3=only_sss3
    )


@results_bp.route('/student/<int:student_id>/report')
@login_required
def student_report(student_id):
    """A consolidated, print/PDF-ready exam report for one student."""
    from models.mock_jamb import MockJAMBResult, MockJAMBExam, MockJAMBAnalytics

    student = db.get_or_404(Student, student_id)
    require_branch_access(student.branch_id)
    from utils.access_control import teacher_form_student_ids
    tids = teacher_form_student_ids()
    if tids is not None and student.id not in tids:
        abort(403)
    pass_grades = set(AcademicAnalytics.PASS_GRADES)
    distinction_grades = set(AcademicAnalytics.DISTINCTION_GRADES)

    # WAEC grouped by year with summary counts.
    waec = student.waec_results.order_by(WAECResult.exam_year.desc(), WAECResult.subject).all()
    waec_by_year = {}
    for r in waec:
        y = waec_by_year.setdefault(r.exam_year, {'rows': [], 'credits': 0, 'distinctions': 0})
        y['rows'].append(r)
        if r.grade in pass_grades:
            y['credits'] += 1
        if r.grade in distinction_grades:
            y['distinctions'] += 1
    waec_years = sorted(waec_by_year.items(), reverse=True)

    # JAMB results (most recent first), with subject breakdown.
    jamb = student.jamb_results.order_by(JAMBResult.exam_year.desc()).all()
    jamb_list = []
    for r in jamb:
        subs = []
        for n, sc in [(r.subject1, r.subject1_score), (r.subject2, r.subject2_score),
                      (r.subject3, r.subject3_score), (r.subject4, r.subject4_score)]:
            if n:
                subs.append({'subject': n, 'score': sc or 0})
        jamb_list.append({'year': r.exam_year, 'total': r.total_score, 'subjects': subs,
                          'level': AcademicAnalytics._jamb_performance_level(r.total_score)})

    # Mock JAMB progression + prediction (active session).
    mock_rows = MockJAMBResult.query.filter_by(student_id=student_id).join(MockJAMBExam).order_by(
        MockJAMBExam.exam_number).all()
    mock_progress = [{'name': m.exam.display_name, 'date': m.exam.exam_date, 'score': m.total_score}
                     for m in mock_rows]
    prediction = None
    active_session = get_active_session()
    if active_session:
        prediction = MockJAMBAnalytics.predict_real_jamb(student_id, active_session.id)

    from utils.admission import assess_admission
    admission = assess_admission(student)

    return render_template('results/student_report.html',
        student=student,
        waec_years=waec_years,
        jamb_list=jamb_list,
        mock_progress=mock_progress,
        prediction=prediction,
        admission=admission,
        generated=_date.today()
    )


@results_bp.route('/import')
@login_required
def import_results():
    return render_template('results/import_results.html')


@results_bp.route('/import/template/<exam>')
@login_required
def import_template(exam):
    from openpyxl import Workbook
    from io import BytesIO
    wb = Workbook()
    ws = wb.active
    if exam == 'jamb':
        ws.title = 'JAMB'
        ws.append(['student_id', 'exam_year', 'total_score', 'subject1', 'subject1_score',
                   'subject2', 'subject2_score', 'subject3', 'subject3_score', 'subject4', 'subject4_score'])
        ws.append(['STU00001', 2025, 250, 'English Language', 62, 'Mathematics', 70, 'Physics', 60, 'Chemistry', 58])
        fn = 'jamb_import_template.xlsx'
    else:
        ws.title = 'WAEC'
        ws.append(['student_id', 'exam_year', 'subject', 'grade'])
        ws.append(['STU00001', 2025, 'Mathematics', 'B2'])
        ws.append(['STU00001', 2025, 'English Language', 'C4'])
        fn = 'waec_import_template.xlsx'
    return xlsx_response(wb, fn)


@results_bp.route('/import/<exam>', methods=['POST'])
@login_required
def import_results_run(exam):
    import pandas as pd
    exam = 'jamb' if exam == 'jamb' else 'waec'
    file = request.files.get('file')
    if not file or not file.filename:
        flash('Please choose a file.', 'error')
        return redirect(url_for('results.import_results'))

    try:
        name = file.filename.lower()
        df = pd.read_csv(file) if name.endswith('.csv') else pd.read_excel(file)
    except Exception as e:
        flash(f'Could not read the file: {e}', 'error')
        return redirect(url_for('results.import_results'))

    df.columns = [str(c).strip().lower() for c in df.columns]

    students = Student.query.all()
    by_code = {s.student_id.lower(): s for s in students if s.student_id}
    by_name = {s.full_name.lower(): s for s in students}

    def find_student(val):
        v = str(val).strip().lower()
        return by_code.get(v) or by_name.get(v)

    imported, skipped, errors = 0, 0, []
    for i, row in df.iterrows():
        line = i + 2  # header is row 1
        try:
            stu = find_student(row.get('student_id'))
            if not stu:
                skipped += 1
                errors.append(f'Row {line}: student "{row.get("student_id")}" not found')
                continue
            year = int(row.get('exam_year'))
            if exam == 'waec':
                subject = str(row.get('subject') or '').strip()
                grade = str(row.get('grade') or '').strip().upper()
                if not subject or not grade:
                    skipped += 1
                    continue
                existing = WAECResult.query.filter_by(student_id=stu.id, exam_year=year, subject=subject).first()
                if existing:
                    existing.grade = grade
                else:
                    db.session.add(WAECResult(student_id=stu.id, exam_year=year, subject=subject, grade=grade))
                imported += 1
            else:
                total = int(row.get('total_score') or 0)
                existing = JAMBResult.query.filter_by(student_id=stu.id, exam_year=year).first()
                obj = existing or JAMBResult(student_id=stu.id, exam_year=year, total_score=total)
                obj.total_score = total
                for n in range(1, 5):
                    sname = row.get(f'subject{n}')
                    sscore = row.get(f'subject{n}_score')
                    if pd.notna(sname):
                        setattr(obj, f'subject{n}', str(sname).strip())
                        setattr(obj, f'subject{n}_score', int(sscore) if pd.notna(sscore) else None)
                if not existing:
                    db.session.add(obj)
                imported += 1
        except Exception as e:
            skipped += 1
            errors.append(f'Row {line}: {e}')

    db.session.commit()
    log_action(f'import_{exam}', f'{imported} imported, {skipped} skipped')
    flash(f'{exam.upper()} import complete: {imported} saved, {skipped} skipped.',
          'success' if imported else 'warning')
    for e in errors[:8]:
        flash(e, 'error')
    return redirect(url_for('results.import_results'))


@results_bp.route('/readiness')
@login_required
def readiness():
    """Actionable exam-readiness checklist for the SSS3 cohort."""
    students = get_sss3_students()
    total = len(students)

    no_stream, no_jamb, no_waec, no_jamb_subjects, no_waec_subjects = [], [], [], [], []
    below_target = []
    for s in students:
        if not s.stream:
            no_stream.append(s)
        if s.jamb_results.count() == 0:
            no_jamb.append(s)
        if s.waec_results.count() == 0:
            no_waec.append(s)
        if not s.jamb_subject_list:
            no_jamb_subjects.append(s)
        if not s.waec_subject_list:
            no_waec_subjects.append(s)
        if s.jamb_target:
            mocks = [m.total_score for m in s.mock_jamb_results.all()]
            best = max(mocks) if mocks else 0
            if best < s.jamb_target:
                below_target.append(s)

    groups = [
        {'key': 'no_jamb', 'title': 'No JAMB result entered', 'icon': 'fa-file-contract', 'students': no_jamb},
        {'key': 'no_waec', 'title': 'No WAEC result entered', 'icon': 'fa-file-alt', 'students': no_waec},
        {'key': 'below_target', 'title': 'Below their JAMB target', 'icon': 'fa-bullseye', 'students': below_target},
        {'key': 'no_stream', 'title': 'No stream / track set', 'icon': 'fa-route', 'students': no_stream},
        {'key': 'no_jamb_subjects', 'title': 'No JAMB subjects on profile', 'icon': 'fa-list', 'students': no_jamb_subjects},
        {'key': 'no_waec_subjects', 'title': 'No WAEC subjects on profile', 'icon': 'fa-list-check', 'students': no_waec_subjects},
    ]
    ready = total - len({s.id for g in groups for s in g['students']})
    return render_template('results/readiness.html', total=total, ready=ready, groups=groups)


@results_bp.route('/analytics')
@login_required
def analytics_hub():
    """One-stop analytics hub: every WAEC/JAMB stat, correlation and projection."""
    waec_years = [y[0] for y in db.session.query(WAECResult.exam_year).distinct().all()]
    jamb_years = [y[0] for y in db.session.query(JAMBResult.exam_year).distinct().all()]
    years = sorted(set(waec_years + jamb_years), reverse=True)

    year = request.args.get('year', type=int)
    if not year and years:
        year = years[0]

    from utils.branch_scope import viewing_branch_id
    bid = viewing_branch_id()
    waec_stats = AcademicAnalytics.get_waec_school_statistics(year, bid) if year else None
    jamb_stats = AcademicAnalytics.get_jamb_school_statistics(year, bid) if year else None
    correlation = AcademicAnalytics.calculate_waec_jamb_correlation(year, bid) if year else None
    yoy = AcademicAnalytics.get_year_over_year_comparison(bid)

    # Gender breakdowns for the selected year.
    def gender_split(model):
        rows = db.session.query(Student.gender, func.count(func.distinct(Student.id))).join(
            model, Student.id == model.student_id
        ).filter(model.exam_year == year).group_by(Student.gender).all()
        return {g or 'Unknown': c for g, c in rows}

    waec_gender = gender_split(WAECResult) if (year and waec_stats) else {}
    jamb_gender = gender_split(JAMBResult) if (year and jamb_stats) else {}

    # Gender-comparison breakdown: pass/distinction rates (WAEC) and mean score
    # / >=200 rate (JAMB) split by gender for the selected year.
    waec_gender_stats = []
    if year and waec_stats:
        rows = db.session.query(Student.gender, WAECResult.grade).join(
            WAECResult, Student.id == WAECResult.student_id
        ).filter(WAECResult.exam_year == year).all()
        by_gender = defaultdict(list)
        for g, grade in rows:
            by_gender[g or 'Unknown'].append(grade)
        for g, grades in by_gender.items():
            total = len(grades)
            passes = sum(1 for x in grades if x in AcademicAnalytics.PASS_GRADES)
            dist = sum(1 for x in grades if x in AcademicAnalytics.DISTINCTION_GRADES)
            pts = [AcademicAnalytics.GRADE_POINTS.get(x, 9) for x in grades]
            waec_gender_stats.append({
                'gender': g,
                'entries': total,
                'pass_rate': round(passes / total * 100, 1) if total else 0,
                'distinction_rate': round(dist / total * 100, 1) if total else 0,
                'mean_points': round(sum(pts) / len(pts), 2) if pts else 0,
            })
        waec_gender_stats.sort(key=lambda x: x['gender'])

    jamb_gender_stats = []
    if year and jamb_stats:
        rows = db.session.query(Student.gender, JAMBResult.total_score).join(
            JAMBResult, Student.id == JAMBResult.student_id
        ).filter(JAMBResult.exam_year == year).all()
        by_gender = defaultdict(list)
        for g, score in rows:
            by_gender[g or 'Unknown'].append(score)
        for g, scores in by_gender.items():
            total = len(scores)
            jamb_gender_stats.append({
                'gender': g,
                'candidates': total,
                'mean_score': round(sum(scores) / total, 1) if total else 0,
                'above_200': sum(1 for s in scores if s >= 200),
                'above_200_rate': round(sum(1 for s in scores if s >= 200) / total * 100, 1) if total else 0,
                'max_score': max(scores) if scores else 0,
            })
        jamb_gender_stats.sort(key=lambda x: x['gender'])

    # JAMB mean projection (simple linear fit) computed directly from JAMB years.
    projection = None
    jamb_means = []
    for jy in sorted(set(jamb_years)):
        rs = JAMBResult.query.filter_by(exam_year=jy).all()
        if rs:
            jamb_means.append({'year': jy, 'mean_score': round(sum(r.total_score for r in rs) / len(rs), 1)})
    if len(jamb_means) >= 2:
        xs = [t['year'] for t in jamb_means]
        ys = [t['mean_score'] for t in jamb_means]
        n = len(xs)
        mean_x = sum(xs) / n
        mean_y = sum(ys) / n
        denom = sum((x - mean_x) ** 2 for x in xs)
        slope = (sum((xs[i] - mean_x) * (ys[i] - mean_y) for i in range(n)) / denom) if denom else 0
        next_year = max(xs) + 1
        projected = round(mean_y + slope * (next_year - mean_x), 1)
        projected = max(0, min(400, projected))
        projection = {
            'next_year': next_year,
            'projected_mean': projected,
            'direction': 'up' if slope > 0.5 else 'down' if slope < -0.5 else 'flat',
            'slope_per_year': round(slope, 1),
            'latest_mean': ys[-1],
        }

    # University-cutoff readiness (JAMB >= 200) for the selected year.
    cutoff = None
    if year and jamb_stats:
        total = jamb_stats['total_students']
        cutoff = {
            'eligible_200': jamb_stats['above_200'],
            'eligible_200_pct': round(jamb_stats['above_200'] / total * 100, 1) if total else 0,
            'competitive_250': jamb_stats['above_250'],
            'competitive_250_pct': round(jamb_stats['above_250'] / total * 100, 1) if total else 0,
            'elite_300': jamb_stats['above_300'],
            'elite_300_pct': round(jamb_stats['above_300'] / total * 100, 1) if total else 0,
        }

    # Class/arm comparison + internal-vs-JAMB correlation for the selected year.
    class_compare = []
    internal_corr = None
    if year:
        active_term = get_active_term()
        arm_map = {}
        if active_term:
            enrs = StudentEnrollment.query.join(ClassArmAssignment).filter(
                ClassArmAssignment.term_id == active_term.id,
                StudentEnrollment.is_active == True
            ).all()
            for e in enrs:
                arm_map.setdefault(e.student_id, e.class_arm_assignment.display_name)

        jamb_by_arm = defaultdict(list)
        for r in JAMBResult.query.filter_by(exam_year=year).all():
            arm = arm_map.get(r.student_id)
            if arm:
                jamb_by_arm[arm].append(r.total_score)
        waec_by_arm = defaultdict(lambda: {'pass': 0, 'total': 0})
        for r in WAECResult.query.filter_by(exam_year=year).all():
            arm = arm_map.get(r.student_id)
            if arm:
                waec_by_arm[arm]['total'] += 1
                if r.grade in AcademicAnalytics.PASS_GRADES:
                    waec_by_arm[arm]['pass'] += 1
        for arm in sorted(set(list(jamb_by_arm) + list(waec_by_arm))):
            js = jamb_by_arm.get(arm, [])
            w = waec_by_arm.get(arm, {'pass': 0, 'total': 0})
            class_compare.append({
                'arm': arm,
                'jamb_count': len(js),
                'jamb_mean': round(sum(js) / len(js), 1) if js else 0,
                'waec_pass_rate': round(w['pass'] / w['total'] * 100, 1) if w['total'] else 0,
                'waec_entries': w['total'],
            })

        pairs = []
        for r in JAMBResult.query.filter_by(exam_year=year).all():
            ts = TermSummary.query.filter_by(student_id=r.student_id).order_by(
                TermSummary.term_id.desc()).first()
            if ts and ts.average_score is not None:
                pairs.append((ts.average_score, r.total_score))
        if len(pairs) >= 5:
            xs = [p[0] for p in pairs]
            ys = [p[1] for p in pairs]
            internal_corr = {
                'n': len(pairs),
                'r': round(AcademicAnalytics._pearson_correlation(xs, ys), 3),
                'mean_internal': round(sum(xs) / len(xs), 1),
                'mean_jamb': round(sum(ys) / len(ys), 1),
            }

    return render_template('results/analytics_hub.html',
        years=years,
        selected_year=year,
        class_compare=class_compare,
        internal_corr=internal_corr,
        waec_stats=waec_stats,
        jamb_stats=jamb_stats,
        correlation=correlation,
        yoy=yoy,
        waec_gender=waec_gender,
        jamb_gender=jamb_gender,
        waec_gender_stats=waec_gender_stats,
        jamb_gender_stats=jamb_gender_stats,
        projection=projection,
        cutoff=cutoff
    )


@results_bp.route('/analytics/export')
@login_required
def analytics_export():
    """Export the analytics hub for a year to a multi-sheet Excel workbook."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from io import BytesIO

    year = request.args.get('year', type=int)
    if not year:
        flash('Select a year to export.', 'error')
        return redirect(url_for('results.analytics_hub'))

    waec_stats = AcademicAnalytics.get_waec_school_statistics(year)
    jamb_stats = AcademicAnalytics.get_jamb_school_statistics(year)
    correlation = AcademicAnalytics.calculate_waec_jamb_correlation(year)

    wb = Workbook()
    head_font = Font(bold=True, color='FFFFFF')
    head_fill = PatternFill(start_color='1a5f4a', end_color='1a5f4a', fill_type='solid')
    title_font = Font(bold=True, size=13)

    def style_header(ws, row=1):
        for cell in ws[row]:
            cell.font = head_font
            cell.fill = head_fill
            cell.alignment = Alignment(horizontal='center')

    # --- Overview sheet ---
    ws = wb.active
    ws.title = 'Overview'
    ws['A1'] = f'Exam Analytics — {year}'
    ws['A1'].font = title_font
    r = 3
    ws.cell(row=r, column=1, value='JAMB'); ws.cell(row=r, column=1).font = Font(bold=True); r += 1
    if jamb_stats:
        for label, key in [('Candidates', 'total_students'), ('Mean', 'mean_score'),
                           ('Median', 'median_score'), ('Highest', 'max_score'),
                           ('Lowest', 'min_score'), ('Std Dev', 'std_deviation'),
                           ('>=200', 'above_200'), ('>=250', 'above_250'), ('>=300', 'above_300')]:
            ws.cell(row=r, column=1, value=label); ws.cell(row=r, column=2, value=jamb_stats[key]); r += 1
    else:
        ws.cell(row=r, column=1, value='No JAMB data'); r += 1
    r += 1
    ws.cell(row=r, column=1, value='WAEC'); ws.cell(row=r, column=1).font = Font(bold=True); r += 1
    if waec_stats:
        for label, key in [('Students', 'unique_students'), ('Subject Entries', 'total_results'),
                           ('Pass Rate %', 'overall_pass_rate'), ('Distinction Rate %', 'overall_distinction_rate')]:
            ws.cell(row=r, column=1, value=label); ws.cell(row=r, column=2, value=waec_stats[key]); r += 1
    else:
        ws.cell(row=r, column=1, value='No WAEC data'); r += 1
    r += 1
    ws.cell(row=r, column=1, value='WAEC↔JAMB Correlation'); ws.cell(row=r, column=1).font = Font(bold=True); r += 1
    if correlation and not correlation.get('error'):
        for label, key in [('Pearson r', 'correlation_coefficient'), ('Predictive Power', 'predictive_power'),
                           ('Paired Students', 'sample_size')]:
            ws.cell(row=r, column=1, value=label); ws.cell(row=r, column=2, value=correlation[key]); r += 1
    else:
        ws.cell(row=r, column=1, value='Insufficient paired data'); r += 1
    ws.column_dimensions['A'].width = 22
    ws.column_dimensions['B'].width = 16

    # --- JAMB subjects ---
    if jamb_stats and jamb_stats['subject_analysis']:
        ws = wb.create_sheet('JAMB Subjects')
        ws.append(['Subject', 'Count', 'Mean', 'Max', 'Min', '>=50', '>=70'])
        style_header(ws)
        for s in jamb_stats['subject_analysis']:
            ws.append([s['subject'], s['count'], s['mean_score'], s['max_score'], s['min_score'], s['above_50'], s['above_70']])
        ws.column_dimensions['A'].width = 24

    # --- JAMB top 10 ---
    if jamb_stats and jamb_stats['top_10']:
        ws = wb.create_sheet('JAMB Top 10')
        ws.append(['Rank', 'Student', 'Score'])
        style_header(ws)
        for i, t in enumerate(jamb_stats['top_10'], 1):
            ws.append([i, t['student_name'], t['score']])
        ws.column_dimensions['B'].width = 28

    # --- WAEC subjects ---
    if waec_stats and waec_stats['subject_analysis']:
        ws = wb.create_sheet('WAEC Subjects')
        ws.append(['Subject', 'Entries', 'A1 %', 'Pass %', 'Fail %'])
        style_header(ws)
        for s in sorted(waec_stats['subject_analysis'], key=lambda x: x['pass_rate'], reverse=True):
            ws.append([s['subject'], s['total_entries'], s['a1_rate'], s['pass_rate'], s['fail_rate']])
        ws.column_dimensions['A'].width = 24

    return xlsx_response(wb, f'exam_analytics_{year}.xlsx')


@results_bp.route('/cutoffs')
@login_required
def cutoffs_list():
    """Manage university/course admission cut-offs used by the advisor."""
    universities = [u[0] for u in db.session.query(UniversityCutoff.university_name)
                    .distinct().order_by(UniversityCutoff.university_name).all()]
    if 'General Requirements' not in universities:
        universities.insert(0, 'General Requirements')
    selected = request.args.get('university') or universities[0]
    rows = UniversityCutoff.query.filter_by(university_name=selected).order_by(
        UniversityCutoff.faculty, UniversityCutoff.course_name).all()

    edit_id = request.args.get('edit', type=int)
    editing = db.session.get(UniversityCutoff, edit_id) if edit_id else None
    editing_subjects = _json.loads(editing.required_subjects or '[]') if editing else []
    reference = SchoolSettings.get('admission_reference', 'General Requirements')

    return render_template('results/cutoffs.html',
        universities=universities, selected=selected, rows=rows,
        editing=editing, editing_subjects=editing_subjects,
        reference=reference, subjects=WAEC_SUBJECTS)


@results_bp.route('/cutoffs/save', methods=['POST'])
@admin_required
def cutoffs_save():
    cid = request.form.get('id', type=int)
    uni = (request.form.get('university_name') or '').strip() or 'General Requirements'
    course = (request.form.get('course_name') or '').strip()
    if not course:
        flash('Course name is required.', 'error')
        return redirect(url_for('results.cutoffs_list', university=uni))

    year = request.form.get('exam_year', type=int) or 0
    obj = db.session.get(UniversityCutoff, cid) if cid else None
    if not obj:
        obj = UniversityCutoff.query.filter_by(university_name=uni, course_name=course, exam_year=year).first()
    if not obj:
        obj = UniversityCutoff(university_name=uni, course_name=course, exam_year=year)
        db.session.add(obj)

    obj.university_name = uni
    obj.course_name = course
    obj.exam_year = year
    obj.faculty = (request.form.get('faculty') or '').strip() or None
    obj.jamb_cutoff = request.form.get('jamb_cutoff', type=int)
    obj.min_credits = request.form.get('min_credits', type=int) or 5
    obj.required_subjects = _json.dumps(request.form.getlist('required_subjects[]'))
    try:
        db.session.commit()
        log_action('cutoff_save', f'{uni} / {course}')
        flash(f'Saved cut-off for {course}.', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Error: {str(e)}', 'error')
    return redirect(url_for('results.cutoffs_list', university=uni))


@results_bp.route('/cutoffs/<int:cid>/delete', methods=['POST'])
@admin_required
def cutoffs_delete(cid):
    obj = db.session.get(UniversityCutoff, cid)
    if obj:
        uni = obj.university_name
        db.session.delete(obj)
        db.session.commit()
        flash('Cut-off deleted.', 'success')
        return redirect(url_for('results.cutoffs_list', university=uni))
    return redirect(url_for('results.cutoffs_list'))


@results_bp.route('/cutoffs/reference', methods=['POST'])
@admin_required
def cutoffs_reference():
    ref = (request.form.get('reference') or 'General Requirements').strip()
    SchoolSettings.set('admission_reference', ref, 'string',
                       'University reference used by the admission advisor')
    flash(f'Admission advisor now uses "{ref}" cut-offs.', 'success')
    return redirect(url_for('results.cutoffs_list', university=ref))


@results_bp.route('/subject-enrolment/<exam>/<path:subject>')
@login_required
def subject_enrolment_detail(exam, subject):
    """List the students enrolled for a particular WAEC/JAMB subject."""
    exam = 'jamb' if exam.lower() == 'jamb' else 'waec'
    only_sss3 = request.args.get('scope', 'sss3') != 'all'
    if only_sss3:
        students = get_sss3_students()
    else:
        students = Student.query.filter_by(is_active=True).order_by(Student.surname).all()

    matched = []
    for s in students:
        enrolled = s.jamb_subject_list if exam == 'jamb' else s.waec_subject_list
        if subject in enrolled:
            matched.append(s)
    matched.sort(key=lambda s: (s.surname or '', s.first_name or ''))

    return render_template('results/subject_enrolment_detail.html',
        exam=exam,
        exam_label='JAMB' if exam == 'jamb' else 'WAEC',
        subject=subject,
        students=matched,
        only_sss3=only_sss3
    )


@results_bp.route('/jamb/student/<int:student_id>/edit/<int:year>', methods=['GET', 'POST'])
@login_required
def edit_jamb(student_id, year):
    """Edit JAMB results for a student/year"""
    student = db.get_or_404(Student, student_id)
    require_branch_access(student.branch_id)
    result = JAMBResult.query.filter_by(student_id=student_id, exam_year=year).first_or_404()
    
    if request.method == 'POST':
        try:
            result.total_score = request.form.get('total_score', type=int)
            result.subject1 = request.form.get('subject1')
            result.subject1_score = request.form.get('subject1_score', type=int)
            result.subject2 = request.form.get('subject2')
            result.subject2_score = request.form.get('subject2_score', type=int)
            result.subject3 = request.form.get('subject3')
            result.subject3_score = request.form.get('subject3_score', type=int)
            result.subject4 = request.form.get('subject4')
            result.subject4_score = request.form.get('subject4_score', type=int)
            
            # Validate total score
            if result.total_score < 0 or result.total_score > 400:
                flash('Total score must be between 0 and 400.', 'error')
                return redirect(url_for('results.edit_jamb', student_id=student_id, year=year))
            
            db.session.commit()
            log_action('results.jamb_edit', detail=f'{year}', target=student)
            flash('JAMB result updated!', 'success')
            return redirect(url_for('results.view_jamb_student', student_id=student_id))
            
        except Exception as e:
            db.session.rollback()
            flash(f'Error: {str(e)}', 'error')
    
    return render_template('results/edit_jamb.html',
        student=student,
        result=result,
        year=year,
        subjects=WAEC_SUBJECTS
    )


@results_bp.route('/jamb/student/<int:student_id>/delete/<int:year>', methods=['POST'])
@admin_required
def delete_jamb(student_id, year):
    """Delete JAMB result for a student in a given year"""
    student = db.get_or_404(Student, student_id)
    require_branch_access(student.branch_id)
    
    try:
        result = JAMBResult.query.filter_by(student_id=student_id, exam_year=year).first()
        
        if result:
            db.session.delete(result)
            db.session.commit()
            log_action('results.jamb_delete', detail=f'{year}', target=student)
            flash(f'Deleted JAMB result for {student.full_name} ({year}).', 'success')
        else:
            flash('No result found to delete.', 'warning')
            
    except Exception as e:
        db.session.rollback()
        flash(f'Error deleting result: {str(e)}', 'error')
    
    return redirect(url_for('results.jamb_list', year=year))


# ============================================================================
# ANALYTICS API ENDPOINTS
# ============================================================================

@results_bp.route('/api/waec/grade-distribution/<int:year>')
@login_required
def api_waec_grade_distribution(year):
    """Get WAEC grade distribution for charts"""
    results = WAECResult.query.filter_by(exam_year=year).all()
    distribution = {g: sum(1 for r in results if r.grade == g) for g in WAEC_GRADES}
    return jsonify(distribution)


@results_bp.route('/api/waec/subject-stats/<int:year>')
@login_required
def api_waec_subject_stats(year):
    """Get subject-level statistics"""
    stats = AcademicAnalytics.get_waec_school_statistics(year)
    if stats:
        return jsonify(stats['subject_analysis'])
    return jsonify([])


@results_bp.route('/api/jamb/score-distribution/<int:year>')
@login_required
def api_jamb_score_distribution(year):
    """Get JAMB score distribution for charts"""
    results = JAMBResult.query.filter_by(exam_year=year).all()
    scores = [r.total_score for r in results]
    
    distribution = {
        '0-100': sum(1 for s in scores if s <= 100),
        '101-150': sum(1 for s in scores if 101 <= s <= 150),
        '151-200': sum(1 for s in scores if 151 <= s <= 200),
        '201-250': sum(1 for s in scores if 201 <= s <= 250),
        '251-300': sum(1 for s in scores if 251 <= s <= 300),
        '301-350': sum(1 for s in scores if 301 <= s <= 350),
        '351-400': sum(1 for s in scores if 351 <= s <= 400)
    }
    
    return jsonify(distribution)


@results_bp.route('/api/yoy-trends')
@login_required
def api_yoy_trends():
    """Get year-over-year performance trends"""
    data = AcademicAnalytics.get_year_over_year_comparison()
    return jsonify(data)


@results_bp.route('/api/student-risk/<int:student_id>')
@login_required
def api_student_risk(student_id):
    """Get risk assessment for a student"""
    require_branch_access(db.get_or_404(Student, student_id).branch_id)
    risk = AcademicAnalytics.calculate_student_risk_score(student_id)
    return jsonify(risk)


@results_bp.route('/api/predict-jamb/<int:student_id>')
@login_required
def api_predict_jamb(student_id):
    """Get JAMB prediction for a student"""
    require_branch_access(db.get_or_404(Student, student_id).branch_id)
    prediction = AcademicAnalytics.predict_jamb_score(student_id)
    return jsonify(prediction)


@results_bp.route('/api/waec-jamb-correlation/<int:year>')
@login_required
def api_waec_jamb_correlation(year):
    """Get WAEC-JAMB correlation data"""
    correlation = AcademicAnalytics.calculate_waec_jamb_correlation(year)
    return jsonify(correlation)


@results_bp.route('/api/top-performers/<int:year>')
@login_required
def api_top_performers(year):
    """Get top performing students"""
    waec_stats = AcademicAnalytics.get_waec_school_statistics(year)
    
    from utils.branch_scope import scope_by_student
    from utils.access_control import teacher_form_student_ids
    _jq = scope_by_student(JAMBResult.query.filter_by(exam_year=year).options(
        joinedload(JAMBResult.student)), JAMBResult)
    _tids = teacher_form_student_ids()
    if _tids is not None:
        _jq = _jq.filter(JAMBResult.student_id.in_(_tids or [-1]))
    jamb_results = _jq.order_by(JAMBResult.total_score.desc()).limit(10).all()
    
    jamb_top = [{
        'student_id': r.student_id,
        'name': r.student.full_name,
        'score': r.total_score
    } for r in jamb_results]
    
    return jsonify({
        'waec_top': waec_stats['top_performers'] if waec_stats else [],
        'jamb_top': jamb_top
    })


# ============================================================================
# EXPORT FUNCTIONALITY
# ============================================================================

@results_bp.route('/waec/export')
@login_required
def export_waec():
    """Export WAEC results to Excel"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from io import BytesIO
    
    year = request.args.get('year', type=int)
    if not year:
        flash('Please select a year to export.', 'error')
        return redirect(url_for('results.waec_list'))
    
    from utils.branch_scope import scope_query
    students_with_results = scope_query(
        db.session.query(Student).join(WAECResult).filter(WAECResult.exam_year == year),
        Student).distinct().order_by(Student.surname).all()
    
    wb = Workbook()
    ws = wb.active
    ws.title = f"WAEC {year}"
    
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF')
    border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    
    ws.merge_cells('A1:L1')
    ws['A1'] = f'WAEC RESULTS - {year}'
    ws['A1'].font = Font(bold=True, size=14)
    ws['A1'].alignment = Alignment(horizontal='center')
    
    headers = ['S/N', 'Student ID', 'Name'] + WAEC_SUBJECTS[:9] + ['Total Points', 'Credits', 'A1s']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
    
    for idx, student in enumerate(students_with_results, 1):
        row = idx + 3
        results = {r.subject: r.grade for r in student.waec_results.filter_by(exam_year=year).all()}
        
        ws.cell(row=row, column=1, value=idx).border = border
        ws.cell(row=row, column=2, value=student.student_id).border = border
        ws.cell(row=row, column=3, value=student.full_name).border = border
        
        total_points = 0
        credit_count = 0
        a1_count = 0
        
        for col, subject in enumerate(WAEC_SUBJECTS[:9], 4):
            grade = results.get(subject, '-')
            cell = ws.cell(row=row, column=col, value=grade)
            cell.border = border
            
            if grade != '-':
                points = WAECResult.grade_to_points(grade)
                total_points += points
                if grade in ['A1', 'B2', 'B3', 'C4', 'C5', 'C6']:
                    credit_count += 1
                if grade == 'A1':
                    a1_count += 1
        
        ws.cell(row=row, column=13, value=total_points).border = border
        ws.cell(row=row, column=14, value=credit_count).border = border
        ws.cell(row=row, column=15, value=a1_count).border = border
    
    ws.column_dimensions['A'].width = 5
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 25
    
    return xlsx_response(wb, f'waec_results_{year}.xlsx')


@results_bp.route('/jamb/export')
@login_required
def export_jamb():
    """Export JAMB results to Excel"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Border, Side
    from io import BytesIO
    
    year = request.args.get('year', type=int)
    if not year:
        flash('Please select a year to export.', 'error')
        return redirect(url_for('results.jamb_list'))
    
    results = (JAMBResult.query.filter_by(exam_year=year)
               .options(joinedload(JAMBResult.student)).join(Student)
               .order_by(JAMBResult.total_score.desc()).all())

    wb = Workbook()
    ws = wb.active
    ws.title = f"JAMB {year}"
    
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF')
    border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    
    ws.merge_cells('A1:L1')
    ws['A1'] = f'JAMB RESULTS - {year}'
    ws['A1'].font = Font(bold=True, size=14)
    
    headers = ['Rank', 'Student ID', 'Name', 'Total Score', 'Subject 1', 'Score', 'Subject 2', 'Score', 'Subject 3', 'Score', 'Subject 4', 'Score']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
    
    for idx, r in enumerate(results, 1):
        row = idx + 3
        ws.cell(row=row, column=1, value=idx).border = border
        ws.cell(row=row, column=2, value=r.student.student_id).border = border
        ws.cell(row=row, column=3, value=r.student.full_name).border = border
        ws.cell(row=row, column=4, value=r.total_score).border = border
        ws.cell(row=row, column=5, value=r.subject1 or '-').border = border
        ws.cell(row=row, column=6, value=r.subject1_score or '-').border = border
        ws.cell(row=row, column=7, value=r.subject2 or '-').border = border
        ws.cell(row=row, column=8, value=r.subject2_score or '-').border = border
        ws.cell(row=row, column=9, value=r.subject3 or '-').border = border
        ws.cell(row=row, column=10, value=r.subject3_score or '-').border = border
        ws.cell(row=row, column=11, value=r.subject4 or '-').border = border
        ws.cell(row=row, column=12, value=r.subject4_score or '-').border = border
    
    return xlsx_response(wb, f'jamb_results_{year}.xlsx')


# ============================================================================
# ENHANCED WAEC ANALYTICS
# ============================================================================

@results_bp.route('/waec/analytics')
@login_required
def waec_analytics():
    """Enhanced WAEC analytics dashboard"""
    from utils.exam_analytics import WAECAnalytics
    
    year = request.args.get('year', type=int)
    
    years = db.session.query(WAECResult.exam_year).distinct().order_by(WAECResult.exam_year.desc()).all()
    years = [y[0] for y in years]
    
    if not year and years:
        year = years[0]
    
    stats = None
    year_comparison = None
    
    if year:
        stats = WAECAnalytics.get_year_statistics(year)
        year_comparison = WAECAnalytics.compare_years()
    
    return render_template('results/waec_analytics.html',
        years=years,
        selected_year=year,
        stats=stats,
        year_comparison=year_comparison
    )


@results_bp.route('/waec/student/<int:student_id>')
@login_required
def waec_student_analysis(student_id):
    """Detailed WAEC analysis for a specific student"""
    from utils.exam_analytics import WAECAnalytics, WAECJAMBCorrelation
    
    student = db.get_or_404(Student, student_id)
    require_branch_access(student.branch_id)
    year = request.args.get('year', type=int)
    
    waec_analysis = WAECAnalytics.get_student_waec_analysis(student_id, year)
    jamb_prediction = WAECJAMBCorrelation.predict_jamb_from_waec(student_id, year)
    
    return render_template('results/waec_student.html',
        student=student,
        waec_analysis=waec_analysis,
        jamb_prediction=jamb_prediction
    )


# ============================================================================
# WAEC-JAMB CORRELATION & PREDICTIONS
# ============================================================================

@results_bp.route('/predictions')
@login_required
def predictions_dashboard():
    """WAEC-JAMB correlation and predictions dashboard"""
    from utils.exam_analytics import WAECJAMBCorrelation
    
    correlation_data = WAECJAMBCorrelation.get_correlation_analysis()
    
    return render_template('results/predictions_dashboard.html',
        correlation_data=correlation_data
    )


@results_bp.route('/predictions/student/<int:student_id>')
@login_required  
def student_predictions(student_id):
    """Comprehensive predictions for a specific student"""
    from utils.exam_analytics import WAECJAMBCorrelation, MockJAMBAnalytics
    from models.mock_jamb import MockJAMBResult
    
    student = db.get_or_404(Student, student_id)
    require_branch_access(student.branch_id)
    
    # Get mock exam history first to check data
    mock_results = MockJAMBResult.query.filter_by(student_id=student_id).all()
    
    # Debug: Print mock results info
    debug_info = []
    for mr in mock_results:
        debug_info.append({
            'id': mr.id,
            'total_score': mr.total_score,
            'subject1': mr.subject1,
            'subject1_score': mr.subject1_score,
            'subject2': mr.subject2,
            'subject2_score': mr.subject2_score,
            'subject3': mr.subject3,
            'subject3_score': mr.subject3_score,
            'subject4': mr.subject4,
            'subject4_score': mr.subject4_score,
        })
    
    # Get all available predictions
    waec_from_mock = None
    jamb_prediction = None
    
    if mock_results:
        try:
            waec_from_mock = WAECJAMBCorrelation.predict_waec_from_mock_jamb(student_id)
        except Exception as e:
            print(f"WAEC prediction error: {e}")
        
        try:
            jamb_prediction = MockJAMBAnalytics.predict_real_jamb(student_id)
        except Exception as e:
            print(f"JAMB prediction error: {e}")
    
    # Check if student has actual WAEC results
    waec_years = db.session.query(WAECResult.exam_year).filter_by(student_id=student_id).distinct().all()
    waec_years = [y[0] for y in waec_years]
    
    jamb_from_waec = None
    if waec_years:
        try:
            jamb_from_waec = WAECJAMBCorrelation.predict_jamb_from_waec(student_id, waec_years[0])
        except Exception as e:
            print(f"JAMB from WAEC error: {e}")
    
    return render_template('results/student_predictions.html',
        student=student,
        waec_from_mock=waec_from_mock,
        jamb_prediction=jamb_prediction,
        jamb_from_waec=jamb_from_waec,
        waec_years=waec_years,
        has_mock_results=len(mock_results) > 0,
        mock_count=len(mock_results),
        debug_info=debug_info
    )


@results_bp.route('/api/predictions/<int:student_id>')
@login_required
def api_student_predictions(student_id):
    """API endpoint for student predictions"""
    require_branch_access(db.get_or_404(Student, student_id).branch_id)
    from utils.exam_analytics import WAECJAMBCorrelation, MockJAMBAnalytics
    
    waec_from_mock = WAECJAMBCorrelation.predict_waec_from_mock_jamb(student_id)
    jamb_prediction = MockJAMBAnalytics.predict_real_jamb(student_id)
    
    return jsonify({
        'student_id': student_id,
        'waec_prediction': waec_from_mock,
        'jamb_prediction': jamb_prediction
    })

