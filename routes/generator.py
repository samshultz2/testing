"""
Timetable Generator routes - Complete Version with all features
"""
from flask import Blueprint, render_template, request, redirect, url_for, flash, jsonify, Response
from models import (
    db, GenSubject, GenTeacher, GenTeacherSubject, GenTeacherAvailability, GenSubjectConfig,
    GenClassSubjectConfig, GenClassStreamSubject, GenStream, GenStreamSubject, GenClassConfig, GenClassArmStream, GenRoom,
    GenTeacherAssignment, GenTimetableRule, GenTimetableResult, GenSettings
)
from utils.helpers import login_required
from io import BytesIO
import uuid

generator_bp = Blueprint('generator', __name__, url_prefix='/generator')


@generator_bp.route('/import-school-data', methods=['POST'])
@login_required
def import_school_data():
    """Populate the generator from the school's existing Subjects / Class Subjects
    / classes / teachers so they don't have to be re-entered here. Purely additive:
    it creates anything missing and never deletes or overwrites your tweaks."""
    from models import Subject, SchoolClass, ClassArm, ClassSubject, ClassArmAssignment
    from utils.helpers import get_active_term
    term = get_active_term()
    if not term:
        flash('Set an active term first, then import.', 'error')
        return redirect(url_for('generator.index'))

    def level_of(section):
        s = (section or '').lower()
        return 'jss' if s == 'junior' else 'sss' if s == 'senior' else None

    # Arms used by each class this term (for GenClassConfig.arm_names).
    arms_by_class = {}
    for caa in ClassArmAssignment.query.filter_by(term_id=term.id).all():
        if caa.arm:
            arms_by_class.setdefault(caa.class_id, set()).add(caa.arm.name)

    gsub, gcls, gtea = {}, {}, {}
    added = {'subjects': 0, 'classes': 0, 'teachers': 0, 'class_subjects': 0, 'assignments': 0}

    def get_subject(name, level):
        key = (name.lower(), level)
        if key in gsub:
            return gsub[key]
        o = GenSubject.query.filter_by(name=name, school_level=level).first()
        if not o:
            o = GenSubject(name=name, school_level=level, is_active=True)
            db.session.add(o); db.session.flush(); added['subjects'] += 1
        gsub[key] = o
        return o

    def get_class(sc, level):
        key = (sc.name.lower(), level)
        if key in gcls:
            return gcls[key]
        o = GenClassConfig.query.filter_by(class_name=sc.name, school_level=level).first()
        if not o:
            arms = sorted(arms_by_class.get(sc.id, []))
            o = GenClassConfig(class_name=sc.name, school_level=level,
                               num_arms=len(arms) or 1,
                               arm_names=','.join(arms) or None, is_active=True)
            db.session.add(o); db.session.flush(); added['classes'] += 1
        gcls[key] = o
        return o

    def get_teacher(name, level):
        key = (name.lower(), level)
        if key in gtea:
            return gtea[key]
        o = GenTeacher.query.filter_by(name=name, school_level=level).first()
        if not o:
            o = GenTeacher(name=name, school_level=level, is_active=True)
            db.session.add(o); db.session.flush(); added['teachers'] += 1
        gtea[key] = o
        return o

    try:
        rows = ClassSubject.query.filter_by(term_id=term.id, is_active=True).all()
        for cs in rows:
            sc = db.session.get(SchoolClass, cs.class_id)
            subj = db.session.get(Subject, cs.subject_id)
            if not (sc and subj):
                continue
            level = level_of(getattr(sc, 'section', None))
            if not level:
                continue
            gs = get_subject(subj.name, level)
            gc = get_class(sc, level)
            arm = db.session.get(ClassArm, cs.arm_id) if cs.arm_id else None
            arm_name = arm.name if arm else None

            # Class takes this subject.
            if not GenClassSubjectConfig.query.filter_by(
                    class_config_id=gc.id, subject_id=gs.id).first():
                db.session.add(GenClassSubjectConfig(
                    class_config_id=gc.id, subject_id=gs.id, is_enabled=True, is_active=True))
                added['class_subjects'] += 1

            # Teacher who teaches it -> can-teach + assignment.
            tname = (cs.teacher_name or '').strip()
            if tname:
                gt = get_teacher(tname, level)
                if not GenTeacherSubject.query.filter_by(
                        teacher_id=gt.id, subject_id=gs.id).first():
                    db.session.add(GenTeacherSubject(teacher_id=gt.id, subject_id=gs.id))
                if not GenTeacherAssignment.query.filter_by(
                        teacher_id=gt.id, subject_id=gs.id,
                        class_config_id=gc.id, arm_name=arm_name).first():
                    db.session.add(GenTeacherAssignment(
                        teacher_id=gt.id, subject_id=gs.id, class_config_id=gc.id,
                        arm_name=arm_name, is_active=True))
                    added['assignments'] += 1
        db.session.commit()
        flash('Imported from school setup — added {subjects} subjects, {classes} classes, '
              '{teachers} teachers, {class_subjects} class-subjects, {assignments} assignments. '
              'Adjust periods/availability below, then generate.'.format(**added), 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Import failed: {e}', 'error')
    return redirect(url_for('generator.index'))

DAYS_OF_WEEK = ['Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday']
SUBJECT_COLORS = [
    '#FF6B6B', '#4ECDC4', '#45B7D1', '#96CEB4', '#FFEAA7',
    '#DDA0DD', '#98D8C8', '#F7DC6F', '#BB8FCE', '#85C1E9'
]


# ============================================================================
# MAIN DASHBOARD - LEVEL SELECTOR
# ============================================================================

@generator_bp.route('/')
@login_required
def index():
    """Timetable generator - Level selector"""
    
    # Get counts for each level
    jss_teachers = GenTeacher.query.filter_by(is_active=True, school_level='jss').count()
    sss_teachers = GenTeacher.query.filter_by(is_active=True, school_level='sss').count()
    
    # Count actual subjects (GenSubject table)
    jss_subjects = GenSubject.query.filter_by(is_active=True, school_level='jss').count()
    sss_subjects = GenSubject.query.filter_by(is_active=True, school_level='sss').count()
    
    # Class counts
    jss_class_ids = [c.id for c in GenClassConfig.query.filter_by(is_active=True, school_level='jss').all()]
    sss_class_ids = [c.id for c in GenClassConfig.query.filter_by(is_active=True, school_level='sss').all()]
    
    jss_classes = len(jss_class_ids)
    sss_classes = len(sss_class_ids)
    
    jss_assignments = GenTeacherAssignment.query.join(GenClassConfig).filter(
        GenTeacherAssignment.is_active == True,
        GenClassConfig.school_level == 'jss'
    ).count()
    sss_assignments = GenTeacherAssignment.query.join(GenClassConfig).filter(
        GenTeacherAssignment.is_active == True,
        GenClassConfig.school_level == 'sss'
    ).count()
    
    # Get latest timetables for each level
    jss_latest = GenTimetableResult.query.filter_by(school_level='jss').order_by(GenTimetableResult.generated_at.desc()).first()
    sss_latest = GenTimetableResult.query.filter_by(school_level='sss').order_by(GenTimetableResult.generated_at.desc()).first()
    
    return render_template('generator/index.html',
        jss_stats={
            'teachers': jss_teachers,
            'subjects': jss_subjects,
            'classes': jss_classes,
            'assignments': jss_assignments,
            'latest': jss_latest,
            'ready': jss_teachers > 0 and jss_subjects > 0 and jss_classes > 0
        },
        sss_stats={
            'teachers': sss_teachers,
            'subjects': sss_subjects,
            'classes': sss_classes,
            'assignments': sss_assignments,
            'latest': sss_latest,
            'ready': sss_teachers > 0 and sss_subjects > 0 and sss_classes > 0
        }
    )


@generator_bp.route('/level/<level>')
@login_required
def level_dashboard(level):
    """Dashboard for a specific level (JSS or SSS)"""
    if level not in ['jss', 'sss']:
        flash('Invalid level', 'error')
        return redirect(url_for('generator.index'))
    
    # Store current level in session
    from flask import session
    session['generator_level'] = level
    
    teachers_count = GenTeacher.query.filter_by(is_active=True, school_level=level).count()
    
    # Count actual subjects from GenSubject table
    subjects_count = GenSubject.query.filter_by(is_active=True, school_level=level).count()
    
    # Count configured subjects (for "ready to generate" check)
    configured_subjects = GenSubjectConfig.query.filter_by(is_active=True, school_level=level).count()
    
    classes_count = GenClassConfig.query.filter_by(is_active=True, school_level=level).count()
    streams_count = GenStream.query.filter_by(is_active=True).count() if level == 'sss' else 0
    rooms_count = GenRoom.query.filter_by(is_active=True).count()
    
    assignments_count = GenTeacherAssignment.query.join(GenClassConfig).filter(
        GenTeacherAssignment.is_active == True,
        GenClassConfig.school_level == level
    ).count()
    
    latest_batch = GenTimetableResult.query.filter_by(school_level=level).order_by(
        GenTimetableResult.generated_at.desc()
    ).first()
    
    # Ready to generate if we have teachers, subjects (configured), and classes
    ready_to_generate = (teachers_count > 0 and configured_subjects > 0 and classes_count > 0)
    
    return render_template('generator/level_dashboard.html',
        level=level,
        level_name='Junior Secondary (JSS)' if level == 'jss' else 'Senior Secondary (SSS)',
        teachers_count=teachers_count,
        subjects_count=subjects_count,
        configured_subjects=configured_subjects,
        streams_count=streams_count,
        classes_count=classes_count,
        rooms_count=rooms_count,
        assignments_count=assignments_count,
        ready_to_generate=ready_to_generate,
        latest_batch=latest_batch
    )


def get_current_level():
    """Get current school level from session"""
    from flask import session
    return session.get('generator_level', 'sss')


# ============================================================================
# TEACHERS
# ============================================================================

@generator_bp.route('/teachers')
@login_required
def teachers_list():
    level = get_current_level()
    teachers = GenTeacher.query.filter_by(is_active=True, school_level=level).order_by(GenTeacher.name).all()
    return render_template('generator/teachers.html', teachers=teachers, level=level)


@generator_bp.route('/teachers/add', methods=['GET', 'POST'])
@login_required
def add_teacher():
    level = get_current_level()
    if request.method == 'POST':
        try:
            name = request.form.get('name', '').strip()
            if not name:
                flash('Teacher name is required.', 'error')
                return redirect(url_for('generator.add_teacher'))
            
            teacher = GenTeacher(
                name=name,
                school_level=level,
                staff_id=request.form.get('staff_id', '').strip() or None,
                phone=request.form.get('phone', '').strip() or None,
                email=request.form.get('email', '').strip() or None,
                max_periods_per_day=request.form.get('max_periods_per_day', type=int) or 6,
                max_periods_per_week=request.form.get('max_periods_per_week', type=int) or 30,
                preferred_time=request.form.get('preferred_time', 'any'),
                is_part_time=request.form.get('is_part_time') == 'on',
                available_days=','.join(request.form.getlist('available_days[]')) or '0,1,2,3,4'
            )
            db.session.add(teacher)
            db.session.commit()
            flash(f'Teacher "{name}" added!', 'success')
            return redirect(url_for('generator.edit_teacher', teacher_id=teacher.id))
        except Exception as e:
            db.session.rollback()
            flash(f'Error: {str(e)}', 'error')
    
    return render_template('generator/add_teacher.html', days=DAYS_OF_WEEK)


@generator_bp.route('/teachers/<int:teacher_id>')
@login_required
def edit_teacher(teacher_id):
    teacher = db.get_or_404(GenTeacher, teacher_id)
    level = teacher.school_level
    all_subjects = GenSubject.query.filter_by(is_active=True, school_level=level).order_by(GenSubject.name).all()
    assigned_ids = [ts.subject_id for ts in teacher.subjects]
    rules = {r.rule_type: r.value for r in GenTimetableRule.query.filter_by(is_active=True, school_level=level).all()}
    periods_per_day = int(rules.get('periods_per_day', 8))
    
    # Get unavailable slots
    unavailable = set()
    for ua in teacher.availability.filter_by(is_available=False).all():
        unavailable.add(f"{ua.day_of_week}_{ua.period_number}")
    
    return render_template('generator/edit_teacher.html',
        teacher=teacher, all_subjects=all_subjects, assigned_ids=assigned_ids,
        days=DAYS_OF_WEEK, periods_per_day=periods_per_day, unavailable=unavailable
    )


@generator_bp.route('/teachers/<int:teacher_id>/update', methods=['POST'])
@login_required
def update_teacher(teacher_id):
    teacher = db.get_or_404(GenTeacher, teacher_id)
    try:
        teacher.name = request.form.get('name', '').strip()
        teacher.staff_id = request.form.get('staff_id', '').strip() or None
        teacher.phone = request.form.get('phone', '').strip() or None
        teacher.email = request.form.get('email', '').strip() or None
        teacher.max_periods_per_day = request.form.get('max_periods_per_day', type=int) or 6
        teacher.max_periods_per_week = request.form.get('max_periods_per_week', type=int) or 30
        teacher.preferred_time = request.form.get('preferred_time', 'any')
        teacher.is_part_time = request.form.get('is_part_time') == 'on'
        teacher.available_days = ','.join(request.form.getlist('available_days[]')) or '0,1,2,3,4'
        db.session.commit()
        flash('Teacher updated!', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Error: {str(e)}', 'error')
    return redirect(url_for('generator.edit_teacher', teacher_id=teacher_id))


@generator_bp.route('/teachers/<int:teacher_id>/subjects', methods=['POST'])
@login_required
def update_teacher_subjects(teacher_id):
    teacher = db.get_or_404(GenTeacher, teacher_id)
    try:
        subject_ids = [int(x) for x in request.form.getlist('subject_ids[]') if x]
        GenTeacherSubject.query.filter_by(teacher_id=teacher_id).delete()
        for subject_id in subject_ids:
            db.session.add(GenTeacherSubject(teacher_id=teacher_id, subject_id=subject_id))
        db.session.commit()
        flash(f'Updated subjects for {teacher.name}!', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Error: {str(e)}', 'error')
    return redirect(url_for('generator.edit_teacher', teacher_id=teacher_id))


@generator_bp.route('/teachers/<int:teacher_id>/availability', methods=['POST'])
@login_required
def update_teacher_availability(teacher_id):
    teacher = db.get_or_404(GenTeacher, teacher_id)
    try:
        GenTeacherAvailability.query.filter_by(teacher_id=teacher_id).delete()
        for slot in request.form.getlist('unavailable[]'):
            day, period = slot.split('_')
            db.session.add(GenTeacherAvailability(
                teacher_id=teacher_id, day_of_week=int(day),
                period_number=int(period), is_available=False
            ))
        db.session.commit()
        flash('Availability updated!', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Error: {str(e)}', 'error')
    return redirect(url_for('generator.edit_teacher', teacher_id=teacher_id))


@generator_bp.route('/teachers/<int:teacher_id>/delete', methods=['POST'])
@login_required
def delete_teacher(teacher_id):
    teacher = db.get_or_404(GenTeacher, teacher_id)
    try:
        teacher.is_active = False
        db.session.commit()
        flash(f'Teacher "{teacher.name}" removed.', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Error: {str(e)}', 'error')
    return redirect(url_for('generator.teachers_list'))


@generator_bp.route('/teachers/import', methods=['GET', 'POST'])
@login_required
def import_teachers():
    if request.method == 'POST':
        try:
            import csv
            from io import StringIO
            file = request.files.get('file')
            if not file:
                flash('No file uploaded.', 'error')
                return redirect(url_for('generator.import_teachers'))
            
            content = file.read().decode('utf-8')
            reader = csv.DictReader(StringIO(content))
            count = 0
            for row in reader:
                name = row.get('name', '').strip()
                if not name or GenTeacher.query.filter_by(name=name, is_active=True).first():
                    continue
                db.session.add(GenTeacher(
                    name=name,
                    staff_id=row.get('staff_id', '').strip() or None,
                    phone=row.get('phone', '').strip() or None,
                    email=row.get('email', '').strip() or None,
                    max_periods_per_day=int(row.get('max_per_day', 6) or 6),
                    max_periods_per_week=int(row.get('max_per_week', 30) or 30)
                ))
                count += 1
            db.session.commit()
            flash(f'Imported {count} teachers!', 'success')
            return redirect(url_for('generator.teachers_list'))
        except Exception as e:
            db.session.rollback()
            flash(f'Error: {str(e)}', 'error')
    return render_template('generator/import_teachers.html')


# ============================================================================
# SUBJECT CONFIGURATION
# ============================================================================

@generator_bp.route('/subjects')
@login_required
def subjects_config():
    level = get_current_level()
    # Get subjects for this level only
    all_subjects = GenSubject.query.filter_by(is_active=True, school_level=level).order_by(GenSubject.name).all()
    configs = {c.subject_id: c for c in GenSubjectConfig.query.filter_by(school_level=level).all()}
    subjects_data = []
    for i, subject in enumerate(all_subjects):
        subjects_data.append({
            'subject': subject,
            'config': configs.get(subject.id),
            'color': SUBJECT_COLORS[i % len(SUBJECT_COLORS)]
        })
    rooms = GenRoom.query.filter_by(is_active=True).all()
    return render_template('generator/subjects_config.html',
        subjects_data=subjects_data,
        categories=['core', 'science', 'arts', 'commercial', 'vocational'],
        rooms=rooms,
        level=level
    )


@generator_bp.route('/subjects/add', methods=['GET', 'POST'])
@login_required
def add_gen_subject():
    """Add a new subject for timetable generation"""
    level = get_current_level()
    if request.method == 'POST':
        try:
            name = request.form.get('name', '').strip()
            if not name:
                flash('Subject name is required.', 'error')
                return redirect(url_for('generator.add_gen_subject'))
            
            # Check if already exists for this level
            if GenSubject.query.filter_by(name=name, school_level=level, is_active=True).first():
                flash(f'Subject "{name}" already exists for {level.upper()}.', 'error')
                return redirect(url_for('generator.add_gen_subject'))
            
            subject = GenSubject(
                name=name,
                short_name=request.form.get('short_name', '').strip() or None,
                school_level=level,
                category=request.form.get('category', 'core')
            )
            db.session.add(subject)
            db.session.commit()
            flash(f'Subject "{name}" added!', 'success')
            return redirect(url_for('generator.subjects_config'))
        except Exception as e:
            db.session.rollback()
            flash(f'Error: {str(e)}', 'error')
    
    return render_template('generator/add_subject.html', level=level)


@generator_bp.route('/subjects/<int:subject_id>/delete', methods=['POST'])
@login_required
def delete_gen_subject(subject_id):
    """Delete a subject"""
    subject = db.get_or_404(GenSubject, subject_id)
    try:
        subject.is_active = False
        # Also deactivate related configs
        GenSubjectConfig.query.filter_by(subject_id=subject_id).update({'is_active': False})
        db.session.commit()
        flash(f'Subject "{subject.name}" removed.', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Error: {str(e)}', 'error')
    return redirect(url_for('generator.subjects_config'))


@generator_bp.route('/subjects/<int:subject_id>/edit', methods=['GET', 'POST'])
@login_required
def edit_gen_subject(subject_id):
    """Edit a subject"""
    subject = db.get_or_404(GenSubject, subject_id)
    level = subject.school_level
    
    if request.method == 'POST':
        try:
            name = request.form.get('name', '').strip()
            if not name:
                flash('Subject name is required.', 'error')
                return redirect(url_for('generator.edit_gen_subject', subject_id=subject_id))
            
            # Check if name already exists for another subject at this level
            existing = GenSubject.query.filter(
                GenSubject.name == name,
                GenSubject.school_level == level,
                GenSubject.id != subject_id,
                GenSubject.is_active == True
            ).first()
            if existing:
                flash(f'Subject "{name}" already exists for {level.upper()}.', 'error')
                return redirect(url_for('generator.edit_gen_subject', subject_id=subject_id))
            
            subject.name = name
            subject.short_name = request.form.get('short_name', '').strip() or None
            subject.category = request.form.get('category', 'core')
            db.session.commit()
            flash(f'Subject "{name}" updated!', 'success')
            return redirect(url_for('generator.subjects_config'))
        except Exception as e:
            db.session.rollback()
            flash(f'Error: {str(e)}', 'error')
    
    return render_template('generator/edit_subject.html', subject=subject, level=level)


@generator_bp.route('/subjects/save', methods=['POST'])
@login_required
def save_subjects_config():
    level = get_current_level()
    try:
        for subject_id in request.form.getlist('subject_id[]'):
            subject_id = int(subject_id)
            periods = request.form.get(f'periods_{subject_id}', type=int) or 4
            needs_double = request.form.get(f'double_{subject_id}') == 'on'
            double_count = request.form.get(f'double_count_{subject_id}', type=int) or 0
            
            config = GenSubjectConfig.query.filter_by(subject_id=subject_id, school_level=level).first()
            data = {
                'periods_per_week': periods,
                'needs_double_period': needs_double,
                'double_period_count': double_count if needs_double else 0,
                'preferred_time': request.form.get(f'preferred_{subject_id}', 'any'),
                'category': request.form.get(f'category_{subject_id}', 'core'),
                'not_first_period': request.form.get(f'not_first_{subject_id}') == 'on',
                'not_last_period': request.form.get(f'not_last_{subject_id}') == 'on',
                'is_active': True
            }
            if config:
                for k, v in data.items():
                    setattr(config, k, v)
            else:
                db.session.add(GenSubjectConfig(subject_id=subject_id, school_level=level, **data))
        db.session.commit()
        flash('Subject configurations saved!', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Error: {str(e)}', 'error')
    return redirect(url_for('generator.subjects_config'))


# ============================================================================
# PER-CLASS SUBJECT CONFIGURATION
# ============================================================================

@generator_bp.route('/class-subjects')
@login_required
def class_subjects_list():
    """List classes for per-class subject configuration"""
    level = get_current_level()
    classes = GenClassConfig.query.filter_by(is_active=True, school_level=level).order_by(GenClassConfig.class_name).all()
    return render_template('generator/class_subjects_list.html', classes=classes, level=level)


@generator_bp.route('/class-subjects/<int:class_id>')
@login_required
def class_subjects_config(class_id):
    """Configure subject periods for a specific class"""
    class_config = db.get_or_404(GenClassConfig, class_id)
    level = class_config.school_level
    
    # Get all subjects for this level with global defaults
    all_subjects = GenSubject.query.filter_by(is_active=True, school_level=level).order_by(GenSubject.name).all()
    global_configs = {c.subject_id: c for c in GenSubjectConfig.query.filter_by(is_active=True, school_level=level).all()}
    
    # Get class-specific overrides
    class_configs = {c.subject_id: c for c in GenClassSubjectConfig.query.filter_by(
        class_config_id=class_id, is_active=True
    ).all()}
    
    # Build subjects data with effective values
    subjects_data = []
    for subject in all_subjects:
        global_cfg = global_configs.get(subject.id)
        class_cfg = class_configs.get(subject.id)
        
        # Determine category
        category = global_cfg.category if global_cfg else 'core'
        
        # Class config overrides global
        if class_cfg:
            periods = class_cfg.periods_per_week
            needs_double = class_cfg.needs_double_period
            double_count = class_cfg.double_period_count
            is_enabled = class_cfg.is_enabled
            has_override = True
        elif global_cfg:
            periods = global_cfg.periods_per_week
            needs_double = global_cfg.needs_double_period
            double_count = global_cfg.double_period_count
            is_enabled = True  # Default enabled if global config exists
            has_override = False
        else:
            periods = 4
            needs_double = False
            double_count = 0
            is_enabled = True  # Default enabled
            has_override = False
        
        subjects_data.append({
            'subject': subject,
            'periods': periods,
            'needs_double': needs_double,
            'double_count': double_count,
            'is_enabled': is_enabled,
            'has_override': has_override,
            'category': category
        })
    
    return render_template('generator/class_subjects_config.html',
        class_config=class_config,
        subjects_data=subjects_data
    )


@generator_bp.route('/class-subjects/<int:class_id>/save', methods=['POST'])
@login_required
def save_class_subjects_config(class_id):
    """Save per-class subject configuration"""
    class_config = db.get_or_404(GenClassConfig, class_id)
    
    try:
        for subject_id in request.form.getlist('subject_id[]'):
            subject_id = int(subject_id)
            is_enabled = request.form.get(f'enabled_{subject_id}') == 'on'
            periods = request.form.get(f'periods_{subject_id}', type=int) or 4
            needs_double = request.form.get(f'double_{subject_id}') == 'on'
            double_count = request.form.get(f'double_count_{subject_id}', type=int) or 0
            
            existing = GenClassSubjectConfig.query.filter_by(
                class_config_id=class_id, subject_id=subject_id
            ).first()
            
            if existing:
                existing.is_enabled = is_enabled
                existing.periods_per_week = periods
                existing.needs_double_period = needs_double
                existing.double_period_count = double_count if needs_double else 0
                existing.is_active = True
            else:
                db.session.add(GenClassSubjectConfig(
                    class_config_id=class_id,
                    subject_id=subject_id,
                    is_enabled=is_enabled,
                    periods_per_week=periods,
                    needs_double_period=needs_double,
                    double_period_count=double_count if needs_double else 0
                ))
        
        db.session.commit()
        flash(f'Subject configuration for {class_config.class_name} saved!', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Error: {str(e)}', 'error')
    
    return redirect(url_for('generator.class_subjects_config', class_id=class_id))


@generator_bp.route('/classes/<int:class_id>/stream-subjects')
@login_required
def class_stream_subjects(class_id):
    """Configure subject periods for each stream in this class"""
    class_config = db.get_or_404(GenClassConfig, class_id)
    
    if not class_config.has_streams:
        flash('This class does not use streams.', 'warning')
        return redirect(url_for('generator.edit_class_config', class_id=class_id))
    
    # Get streams used by this class
    stream_ids = set()
    for cas in GenClassArmStream.query.filter_by(class_config_id=class_id).all():
        if cas.stream_id:
            stream_ids.add(cas.stream_id)
    
    streams_data = []
    for stream_id in stream_ids:
        stream = db.session.get(GenStream, stream_id)
        if not stream:
            continue
        
        # Get subjects in this stream
        stream_subjects = GenStreamSubject.query.filter_by(stream_id=stream_id).all()
        
        # Get class-stream configs
        class_stream_configs = {
            css.subject_id: css 
            for css in GenClassStreamSubject.query.filter_by(
                class_config_id=class_id, stream_id=stream_id
            ).all()
        }
        
        subjects = []
        total_periods = 0
        for ss in stream_subjects:
            cfg = GenSubjectConfig.query.filter_by(subject_id=ss.subject_id).first()
            class_stream_cfg = class_stream_configs.get(ss.subject_id)
            
            # Priority: class-stream > stream > global
            if class_stream_cfg:
                periods = class_stream_cfg.periods_per_week
                needs_double = class_stream_cfg.needs_double_period
                is_enabled = class_stream_cfg.is_enabled
            elif ss.periods_per_week is not None:
                periods = ss.periods_per_week
                needs_double = ss.needs_double_period or False
                is_enabled = True
            elif cfg:
                periods = cfg.periods_per_week
                needs_double = cfg.needs_double_period
                is_enabled = True
            else:
                periods = 2
                needs_double = False
                is_enabled = True
            
            if is_enabled:
                total_periods += periods
            
            subjects.append({
                'id': ss.subject_id,
                'name': ss.subject.name if ss.subject else f'[Subject {ss.subject_id}]',
                'periods': periods,
                'needs_double': needs_double,
                'is_enabled': is_enabled,
                'global_periods': cfg.periods_per_week if cfg else 2,
                'stream_periods': ss.periods_per_week
            })
        
        streams_data.append({
            'stream': stream,
            'subjects': subjects,
            'total_periods': total_periods
        })
    
    return render_template('generator/class_stream_subjects.html',
        class_config=class_config,
        streams_data=streams_data
    )


@generator_bp.route('/classes/<int:class_id>/stream-subjects/save', methods=['POST'])
@login_required
def save_class_stream_subjects(class_id):
    """Save class-stream subject configuration"""
    class_config = db.get_or_404(GenClassConfig, class_id)
    
    try:
        stream_id = int(request.form.get('stream_id'))
        subject_ids = [int(x) for x in request.form.getlist('subject_ids[]') if x]
        
        for subject_id in subject_ids:
            periods = int(request.form.get(f'periods_{subject_id}', 2))
            needs_double = request.form.get(f'double_{subject_id}') == '1'
            is_enabled = request.form.get(f'enabled_{subject_id}') == '1'
            
            existing = GenClassStreamSubject.query.filter_by(
                class_config_id=class_id,
                stream_id=stream_id,
                subject_id=subject_id
            ).first()
            
            # Get double_period_count from per-class settings
            double_count = 0
            if needs_double:
                per_class_config = GenClassSubjectConfig.query.filter_by(
                    class_config_id=class_id,
                    subject_id=subject_id
                ).first()
                double_count = per_class_config.double_period_count if per_class_config else 1
            
            if existing:
                existing.periods_per_week = periods
                existing.needs_double_period = needs_double
                existing.double_period_count = double_count
                existing.is_enabled = is_enabled
            else:
                db.session.add(GenClassStreamSubject(
                    class_config_id=class_id,
                    stream_id=stream_id,
                    subject_id=subject_id,
                    periods_per_week=periods,
                    needs_double_period=needs_double,
                    double_period_count=double_count,
                    is_enabled=is_enabled
                ))
        
        db.session.commit()
        stream = db.session.get(GenStream, stream_id)
        flash(f'Saved {stream.name} subjects for {class_config.class_name}!', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Error: {str(e)}', 'error')
    
    return redirect(url_for('generator.class_stream_subjects', class_id=class_id))


# ============================================================================
# ROOMS
# ============================================================================

@generator_bp.route('/rooms')
@login_required
def rooms_list():
    rooms = GenRoom.query.filter_by(is_active=True).order_by(GenRoom.name).all()
    return render_template('generator/rooms.html', rooms=rooms)


@generator_bp.route('/rooms/add', methods=['GET', 'POST'])
@login_required
def add_room():
    if request.method == 'POST':
        try:
            name = request.form.get('name', '').strip()
            if not name:
                flash('Room name is required.', 'error')
                return redirect(url_for('generator.add_room'))
            db.session.add(GenRoom(
                name=name,
                short_name=request.form.get('short_name', '').strip() or None,
                room_type=request.form.get('room_type', 'classroom'),
                capacity=request.form.get('capacity', type=int)
            ))
            db.session.commit()
            flash(f'Room "{name}" added!', 'success')
            return redirect(url_for('generator.rooms_list'))
        except Exception as e:
            db.session.rollback()
            flash(f'Error: {str(e)}', 'error')
    return render_template('generator/add_room.html')


@generator_bp.route('/rooms/<int:room_id>/delete', methods=['POST'])
@login_required
def delete_room(room_id):
    room = db.get_or_404(GenRoom, room_id)
    try:
        room.is_active = False
        db.session.commit()
        flash(f'Room "{room.name}" removed.', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Error: {str(e)}', 'error')
    return redirect(url_for('generator.rooms_list'))


# ============================================================================
# STREAMS
# ============================================================================

@generator_bp.route('/streams')
@login_required
def streams_list():
    streams = GenStream.query.filter_by(is_active=True).order_by(GenStream.name).all()
    return render_template('generator/streams.html', streams=streams)


@generator_bp.route('/streams/add', methods=['GET', 'POST'])
@login_required
def add_stream():
    if request.method == 'POST':
        try:
            name = request.form.get('name', '').strip()
            if not name:
                flash('Stream name is required.', 'error')
                return redirect(url_for('generator.add_stream'))
            if GenStream.query.filter_by(name=name).first():
                flash('Stream exists.', 'error')
                return redirect(url_for('generator.add_stream'))
            stream = GenStream(
                name=name,
                short_name=request.form.get('short_name', '').strip() or None,
                description=request.form.get('description', '').strip() or None
            )
            db.session.add(stream)
            db.session.commit()
            flash(f'Stream "{name}" added!', 'success')
            return redirect(url_for('generator.edit_stream', stream_id=stream.id))
        except Exception as e:
            db.session.rollback()
            flash(f'Error: {str(e)}', 'error')
    return render_template('generator/add_stream.html')


@generator_bp.route('/streams/<int:stream_id>')
@login_required
def edit_stream(stream_id):
    stream = db.get_or_404(GenStream, stream_id)
    # Streams are SSS-only
    all_subjects = GenSubject.query.filter_by(is_active=True, school_level='sss').order_by(GenSubject.name).all()
    assigned = {ss.subject_id: ss for ss in stream.subjects}
    # Get global subject configs for default periods display
    global_configs = {cfg.subject_id: cfg for cfg in GenSubjectConfig.query.filter_by(is_active=True, school_level='sss').all()}
    return render_template('generator/edit_stream.html',
        stream=stream, all_subjects=all_subjects, assigned=assigned, global_configs=global_configs
    )


@generator_bp.route('/streams/<int:stream_id>/update', methods=['POST'])
@login_required
def update_stream(stream_id):
    stream = db.get_or_404(GenStream, stream_id)
    try:
        stream.name = request.form.get('name', '').strip()
        stream.short_name = request.form.get('short_name', '').strip() or None
        stream.description = request.form.get('description', '').strip() or None
        db.session.commit()
        flash('Stream updated!', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Error: {str(e)}', 'error')
    return redirect(url_for('generator.edit_stream', stream_id=stream_id))


@generator_bp.route('/streams/<int:stream_id>/subjects', methods=['POST'])
@login_required
def update_stream_subjects(stream_id):
    stream = db.get_or_404(GenStream, stream_id)
    try:
        subject_ids = [int(x) for x in request.form.getlist('subject_ids[]') if x]
        compulsory_ids = [int(x) for x in request.form.getlist('compulsory_ids[]') if x]
        GenStreamSubject.query.filter_by(stream_id=stream_id).delete()
        for subject_id in subject_ids:
            # Get stream-specific periods (if set)
            periods_str = request.form.get(f'periods_{subject_id}', '').strip()
            periods = int(periods_str) if periods_str else None
            needs_double = request.form.get(f'double_{subject_id}') == '1'
            
            db.session.add(GenStreamSubject(
                stream_id=stream_id, 
                subject_id=subject_id,
                is_compulsory=subject_id in compulsory_ids,
                periods_per_week=periods,
                needs_double_period=needs_double if needs_double else None,
                double_period_count=1 if needs_double else None
            ))
        db.session.commit()
        flash(f'Updated subjects for {stream.name}!', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Error: {str(e)}', 'error')
    return redirect(url_for('generator.edit_stream', stream_id=stream_id))


@generator_bp.route('/streams/<int:stream_id>/delete', methods=['POST'])
@login_required
def delete_stream(stream_id):
    stream = db.get_or_404(GenStream, stream_id)
    try:
        stream.is_active = False
        db.session.commit()
        flash(f'Stream "{stream.name}" removed.', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Error: {str(e)}', 'error')
    return redirect(url_for('generator.streams_list'))


# ============================================================================
# CLASSES
# ============================================================================

@generator_bp.route('/classes')
@login_required
def classes_config():
    level = get_current_level()
    classes = GenClassConfig.query.filter_by(is_active=True, school_level=level).order_by(GenClassConfig.class_name).all()
    streams = GenStream.query.filter_by(is_active=True).all()
    return render_template('generator/classes_config.html', classes=classes, streams=streams, level=level)


@generator_bp.route('/classes/add', methods=['GET', 'POST'])
@login_required
def add_class_config():
    level = get_current_level()
    if request.method == 'POST':
        try:
            class_name = request.form.get('class_name', '').strip()
            if not class_name:
                flash('Class name is required.', 'error')
                return redirect(url_for('generator.add_class_config'))
            if GenClassConfig.query.filter_by(class_name=class_name, school_level=level, is_active=True).first():
                flash('Class exists.', 'error')
                return redirect(url_for('generator.add_class_config'))
            has_streams = request.form.get('has_streams') == 'on'
            class_config = GenClassConfig(
                class_name=class_name,
                school_level=level,
                num_arms=request.form.get('num_arms', type=int) or 1,
                arm_names=request.form.get('arm_names', '').strip() or None,
                has_streams=has_streams
            )
            db.session.add(class_config)
            db.session.commit()
            flash(f'Class "{class_name}" added!', 'success')
            if has_streams:
                return redirect(url_for('generator.edit_class_config', class_id=class_config.id))
            return redirect(url_for('generator.classes_config'))
        except Exception as e:
            db.session.rollback()
            flash(f'Error: {str(e)}', 'error')
    return render_template('generator/add_class_config.html', level=level)


@generator_bp.route('/classes/<int:class_id>')
@login_required
def edit_class_config(class_id):
    class_config = db.get_or_404(GenClassConfig, class_id)
    streams = GenStream.query.filter_by(is_active=True).all()
    arm_streams = {cas.arm_name: cas for cas in class_config.arm_streams}
    return render_template('generator/edit_class_config.html',
        class_config=class_config, streams=streams, arm_streams=arm_streams, level=class_config.school_level
    )


@generator_bp.route('/classes/<int:class_id>/update', methods=['POST'])
@login_required
def update_class_config(class_id):
    class_config = db.get_or_404(GenClassConfig, class_id)
    try:
        class_config.class_name = request.form.get('class_name', '').strip()
        class_config.num_arms = request.form.get('num_arms', type=int) or 1
        class_config.arm_names = request.form.get('arm_names', '').strip() or None
        class_config.has_streams = request.form.get('has_streams') == 'on'
        db.session.commit()
        flash('Class configuration updated!', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Error: {str(e)}', 'error')
    return redirect(url_for('generator.edit_class_config', class_id=class_id))


@generator_bp.route('/classes/<int:class_id>/arm-streams', methods=['POST'])
@login_required
def update_arm_streams(class_id):
    class_config = db.get_or_404(GenClassConfig, class_id)
    try:
        GenClassArmStream.query.filter_by(class_config_id=class_id).delete()
        for arm_name in class_config.arm_list:
            stream_id = request.form.get(f'stream_{arm_name}', type=int)
            if stream_id:
                db.session.add(GenClassArmStream(
                    class_config_id=class_id, arm_name=arm_name, stream_id=stream_id
                ))
        db.session.commit()
        flash('Arm-stream assignments updated!', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Error: {str(e)}', 'error')
    return redirect(url_for('generator.edit_class_config', class_id=class_id))


@generator_bp.route('/classes/<int:class_id>/delete', methods=['POST'])
@login_required
def delete_class_config(class_id):
    class_config = db.get_or_404(GenClassConfig, class_id)
    try:
        class_config.is_active = False
        db.session.commit()
        flash(f'Class "{class_config.class_name}" removed.', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Error: {str(e)}', 'error')
    return redirect(url_for('generator.classes_config'))


# ============================================================================
# TEACHER ASSIGNMENTS
# ============================================================================

@generator_bp.route('/assignments')
@login_required
def teacher_assignments():
    level = get_current_level()
    classes = GenClassConfig.query.filter_by(is_active=True, school_level=level).order_by(GenClassConfig.class_name).all()
    teachers = GenTeacher.query.filter_by(is_active=True, school_level=level).order_by(GenTeacher.name).all()
    
    # Get subjects for this level
    subjects = GenSubject.query.filter_by(is_active=True, school_level=level).order_by(GenSubject.name).all()
    
    assignments_by_class = {}
    for cc in classes:
        assignments_by_class[cc.id] = GenTeacherAssignment.query.filter_by(
            class_config_id=cc.id, is_active=True
        ).all()
    return render_template('generator/teacher_assignments.html',
        classes=classes, teachers=teachers, subjects=subjects,
        assignments_by_class=assignments_by_class, level=level
    )


@generator_bp.route('/assignments/add', methods=['POST'])
@login_required
def add_teacher_assignment():
    try:
        teacher_id = request.form.get('teacher_id', type=int)
        subject_id = request.form.get('subject_id', type=int)
        class_config_id = request.form.get('class_config_id', type=int)
        arm_name = request.form.get('arm_name', '').strip() or None
        
        if not all([teacher_id, subject_id, class_config_id]):
            flash('Teacher, subject, and class are required.', 'error')
            return redirect(url_for('generator.teacher_assignments'))
        
        if GenTeacherAssignment.query.filter_by(
            teacher_id=teacher_id, subject_id=subject_id,
            class_config_id=class_config_id, arm_name=arm_name, is_active=True
        ).first():
            flash('Assignment exists.', 'warning')
            return redirect(url_for('generator.teacher_assignments'))
        
        db.session.add(GenTeacherAssignment(
            teacher_id=teacher_id, subject_id=subject_id,
            class_config_id=class_config_id, arm_name=arm_name
        ))
        db.session.commit()
        flash('Teacher assignment added!', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Error: {str(e)}', 'error')
    return redirect(url_for('generator.teacher_assignments'))


@generator_bp.route('/assignments/<int:assignment_id>/delete', methods=['POST'])
@login_required
def delete_teacher_assignment(assignment_id):
    assignment = db.get_or_404(GenTeacherAssignment, assignment_id)
    try:
        assignment.is_active = False
        db.session.commit()
        flash('Assignment removed.', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Error: {str(e)}', 'error')
    return redirect(url_for('generator.teacher_assignments'))


# ============================================================================
# RULES
# ============================================================================

@generator_bp.route('/rules')
@login_required
def rules_config():
    level = get_current_level()
    rules = {r.rule_type: r.value for r in GenTimetableRule.query.filter_by(is_active=True, school_level=level).all()}
    return render_template('generator/rules_config.html', rules=rules, level=level)


@generator_bp.route('/rules/save', methods=['POST'])
@login_required
def save_rules():
    level = get_current_level()
    try:
        rules_to_save = [
            ('periods_per_day', request.form.get('periods_per_day', '8')),
            ('break_after_period', request.form.get('break_after_period', '5')),
            ('no_repeat_same_day', 'true' if request.form.get('no_repeat_same_day') == 'on' else 'false'),
            ('max_consecutive', request.form.get('max_consecutive', '3')),
            ('distribute_evenly', 'true' if request.form.get('distribute_evenly') == 'on' else 'false'),
        ]
        
        for rule_type, value in rules_to_save:
            # Find existing rule for this level or create new
            existing = GenTimetableRule.query.filter_by(rule_type=rule_type, school_level=level).first()
            if existing:
                existing.value = value
                existing.is_active = True
            else:
                db.session.add(GenTimetableRule(rule_type=rule_type, value=value, school_level=level, is_active=True))
        
        db.session.commit()
        flash('Rules saved!', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Error: {str(e)}', 'error')
    return redirect(url_for('generator.rules_config'))


# ============================================================================
# SETTINGS
# ============================================================================

@generator_bp.route('/settings')
@login_required
def generator_settings():
    return render_template('generator/settings.html',
        school_name=GenSettings.get('school_name', ''),
        school_address=GenSettings.get('school_address', ''),
        academic_year=GenSettings.get('academic_year', ''),
        term_name=GenSettings.get('term_name', '')
    )


@generator_bp.route('/settings/save', methods=['POST'])
@login_required
def save_generator_settings():
    try:
        for key in ['school_name', 'school_address', 'academic_year', 'term_name']:
            GenSettings.set(key, request.form.get(key, ''))
        flash('Settings saved!', 'success')
    except Exception as e:
        flash(f'Error: {str(e)}', 'error')
    return redirect(url_for('generator.generator_settings'))


# ============================================================================
# GENERATION
# ============================================================================

@generator_bp.route('/generate')
@login_required
def generate_page():
    level = get_current_level()
    classes = GenClassConfig.query.filter_by(is_active=True, school_level=level).order_by(GenClassConfig.class_name).all()
    rules = {r.rule_type: r.value for r in GenTimetableRule.query.filter_by(is_active=True, school_level=level).all()}
    
    # Validation
    issues = []
    if GenTeacher.query.filter_by(is_active=True, school_level=level).count() == 0:
        issues.append({'type': 'error', 'msg': 'No teachers configured'})
    if GenSubjectConfig.query.filter_by(is_active=True, school_level=level).count() == 0:
        issues.append({'type': 'error', 'msg': 'No subjects configured'})
    if not classes:
        issues.append({'type': 'error', 'msg': 'No classes configured'})
    
    # Check for assignments
    assignment_count = GenTeacherAssignment.query.join(GenClassConfig).filter(
        GenTeacherAssignment.is_active == True,
        GenClassConfig.school_level == level
    ).count()
    if assignment_count == 0:
        issues.append({'type': 'error', 'msg': 'No teacher assignments'})
    
    # Check subjects without teachers
    for config in GenSubjectConfig.query.filter_by(is_active=True, school_level=level, category='core').all():
        if not GenTeacherAssignment.query.join(GenClassConfig).filter(
            GenTeacherAssignment.subject_id == config.subject_id,
            GenTeacherAssignment.is_active == True,
            GenClassConfig.school_level == level
        ).first():
            subj_name = config.subject.name if config.subject else f'Subject {config.subject_id}'
            issues.append({'type': 'warning', 'msg': f'{subj_name} has no teacher'})
    
    return render_template('generator/generate.html',
        classes=classes, rules=rules, issues=issues[:15], level=level,
        teachers_count=GenTeacher.query.filter_by(is_active=True, school_level=level).count(),
        subjects_count=GenSubjectConfig.query.filter_by(is_active=True, school_level=level).count(),
        assignments_count=assignment_count
    )


@generator_bp.route('/generate/run', methods=['POST'])
@login_required
def run_generation():
    """Generate timetable using global scheduling approach"""
    level = get_current_level()
    try:
        class_ids = [int(x) for x in request.form.getlist('class_ids[]') if x]
        if not class_ids:
            flash('Select at least one class.', 'error')
            return redirect(url_for('generator.generate_page'))
        
        rules = {r.rule_type: r.value for r in GenTimetableRule.query.filter_by(is_active=True, school_level=level).all()}
        periods_per_day = int(rules.get('periods_per_day', 8))
        break_after = int(rules.get('break_after_period', 5))
        
        # Use global generation
        from routes.generator_global import run_global_generation
        
        result = run_global_generation(class_ids, periods_per_day, break_after, num_attempts=25)
        
        # Save results
        batch_id = str(uuid.uuid4())[:8]
        
        # Clear old results
        for class_name, arm, _ in result['class_arms']:
            GenTimetableResult.query.filter_by(
                class_name=class_name, arm_name=arm, school_level=level, is_locked=False
            ).delete()
        
        # Save new results
        for (class_name, arm), days in result['timetables'].items():
            for day, periods in days.items():
                for period, entry in periods.items():
                    if entry:
                        db.session.add(GenTimetableResult(
                            batch_id=batch_id,
                            school_level=level,
                            class_name=class_name,
                            arm_name=arm,
                            day_of_week=day,
                            period_number=period,
                            subject_id=entry.get('subject_id'),
                            teacher_id=entry.get('teacher_id'),
                            is_double_period=entry.get('is_double', False)
                        ))
        
        db.session.commit()
        
        empty = result['empty_count']
        total = result['total_slots']
        filled_pct = ((total - empty) / total) * 100 if total > 0 else 0
        
        if empty > 0:
            flash(f'Generated {len(result["class_arms"])} timetables. {empty} empty slots ({filled_pct:.1f}% filled).', 'warning')
        else:
            flash(f'Successfully generated {len(result["class_arms"])} timetables with 100% slots filled!', 'success')
        
        return redirect(url_for('generator.view_results', batch_id=batch_id))
        
    except Exception as e:
        db.session.rollback()
        import traceback
        traceback.print_exc()
        flash(f'Error: {str(e)}', 'error')
        return redirect(url_for('generator.generate_page'))


@generator_bp.route('/generate/ortools', methods=['POST'])
@login_required
def run_ortools_generation():
    """Generate timetable using OR-Tools constraint programming solver (optimal)"""
    try:
        class_ids = [int(x) for x in request.form.getlist('class_ids[]') if x]
        if not class_ids:
            flash('Select at least one class.', 'error')
            return redirect(url_for('generator.generate_page'))
        
        # Check if OR-Tools is available
        from routes.generator_ortools import check_ortools_available, generate_with_ortools, save_ortools_result
        
        if not check_ortools_available():
            flash('OR-Tools not installed. Run: pip install ortools --break-system-packages', 'error')
            return redirect(url_for('generator.generate_page'))
        
        rules = {r.rule_type: r.value for r in GenTimetableRule.query.filter_by(is_active=True).all()}
        periods_per_day = int(rules.get('periods_per_day', 8))
        
        # Get time limit from form or default to 300 seconds
        time_limit = int(request.form.get('time_limit', 300))
        
        # Run OR-Tools solver
        # Get break_after from rules
        break_after = int(rules.get('break_after_period', 5))
        
        # Run OR-Tools solver
        result = generate_with_ortools(class_ids, periods_per_day, time_limit, break_after)
        
        if not result['success']:
            flash(f'Generation failed: {result["message"]}', 'error')
            return redirect(url_for('generator.generate_page'))
        
        # Save results
        batch_id = save_ortools_result(result)
        
        empty = result['empty_count']
        assigned = result['assigned_count']
        total = result['total_requirements']
        total_slots = len(result['class_arms']) * 40
        filled_pct = ((total_slots - empty) / total_slots) * 100 if total_slots > 0 else 0
        
        if empty > 0:
            flash(f'OR-Tools: Assigned {assigned}/{total} requirements. {empty} empty slots ({filled_pct:.1f}% filled).', 'warning')
        else:
            flash(f'OR-Tools: Perfect! Assigned all {assigned} requirements with 100% slots filled!', 'success')
        
        return redirect(url_for('generator.view_results', batch_id=batch_id))
        
    except Exception as e:
        db.session.rollback()
        import traceback
        traceback.print_exc()
        flash(f'Error: {str(e)}', 'error')
        return redirect(url_for('generator.generate_page'))


def generate_for_class(batch_id, cc, arm, periods_per_day, break_after, no_repeat, distribute, teacher_schedule):
    from random import shuffle
    conflicts = []
    
    # Get stream
    stream_id = None
    if cc.has_streams:
        cas = GenClassArmStream.query.filter_by(class_config_id=cc.id, arm_name=arm).first()
        if cas:
            stream_id = cas.stream_id
    
    # Get per-class subject configs (the "Takes" checkbox settings)
    class_configs = {c.subject_id: c for c in GenClassSubjectConfig.query.filter_by(
        class_config_id=cc.id, is_active=True
    ).all()}
    
    # Determine which subjects this class arm should take
    subjects = []
    
    if cc.has_streams and stream_id:
        # Class has streams - ONLY use subjects defined in the stream
        # (User already selected all subjects including core in the stream setup)
        for ss in GenStreamSubject.query.filter_by(stream_id=stream_id).all():
            # Check per-class "Takes" checkbox - if explicitly disabled, skip
            class_cfg = class_configs.get(ss.subject_id)
            if class_cfg and not class_cfg.is_enabled:
                continue  # Explicitly unchecked "Takes" for this class
            
            # Get subject config for periods info
            cfg = GenSubjectConfig.query.filter_by(subject_id=ss.subject_id, is_active=True).first()
            subj_name = ss.subject.name if ss.subject else f'[Subject {ss.subject_id}]'
            if cfg:
                subjects.append({'id': ss.subject_id, 'name': subj_name, 'category': cfg.category})
            else:
                # Subject not configured in Subject Settings, use defaults
                subjects.append({'id': ss.subject_id, 'name': subj_name, 'category': 'general'})
    else:
        # Class doesn't have streams (e.g., JSS) - include all enabled subjects
        for cfg in GenSubjectConfig.query.filter_by(is_active=True).all():
            # Check per-class "Takes" checkbox
            class_cfg = class_configs.get(cfg.subject_id)
            if class_cfg and not class_cfg.is_enabled:
                continue  # Explicitly unchecked "Takes" for this class
            
            subj_name = cfg.subject.name if cfg.subject else f'[Subject {cfg.subject_id}]'
            subjects.append({'id': cfg.subject_id, 'name': subj_name, 'category': cfg.category})
    
    # Get teacher assignments
    assignments = {}
    for ta in GenTeacherAssignment.query.filter_by(class_config_id=cc.id, is_active=True).filter(
        (GenTeacherAssignment.arm_name == None) | (GenTeacherAssignment.arm_name == arm)
    ).all():
        assignments[ta.subject_id] = ta.teacher_id
    
    # Get max consecutive periods from rules
    max_consecutive_rule = GenTimetableRule.query.filter_by(rule_type='max_consecutive').first()
    max_consecutive = int(max_consecutive_rule.value) if max_consecutive_rule else 3
    
    # Initialize timetable - 8 periods (5 morning + 3 afternoon), break is between period 5 and 6
    # Period 1-5 = morning, Period 6-8 = afternoon
    timetable = {d: {p: None for p in range(1, periods_per_day + 1)} for d in range(5)}
    # Note: break_after indicates break comes after this period (e.g., 5 means break after period 5)
    
    # Build needs - using per-class config if available, else global
    needs = []
    for subj in subjects:
        global_cfg = GenSubjectConfig.query.filter_by(subject_id=subj['id']).first()
        class_cfg = class_configs.get(subj['id'])
        
        # Use class-specific config if available, otherwise global
        if class_cfg:
            periods = class_cfg.periods_per_week
            needs_double = class_cfg.needs_double_period
            double_count = class_cfg.double_period_count if needs_double else 0
        elif global_cfg:
            periods = global_cfg.periods_per_week
            needs_double = global_cfg.needs_double_period
            double_count = global_cfg.double_period_count if needs_double else 0
        else:
            periods = 4
            needs_double = False
            double_count = 0
        
        teacher_id = assignments.get(subj['id'])
        single_count = periods - (double_count * 2)
        
        # Get constraints from global config
        not_first = global_cfg.not_first_period if global_cfg else False
        not_last = global_cfg.not_last_period if global_cfg else False
        preferred = global_cfg.preferred_time if global_cfg else 'any'
        
        for _ in range(double_count):
            needs.append({
                'subject_id': subj['id'], 'name': subj['name'], 'teacher_id': teacher_id,
                'not_first': not_first, 'not_last': not_last, 'preferred': preferred,
                'is_double': True
            })
        for _ in range(single_count):
            needs.append({
                'subject_id': subj['id'], 'name': subj['name'], 'teacher_id': teacher_id,
                'not_first': not_first, 'not_last': not_last, 'preferred': preferred,
                'is_double': False
            })
    
    # Count how busy each teacher already is (from teacher_schedule)
    def teacher_busyness(tid):
        if not tid or tid not in teacher_schedule:
            return 0
        return len(teacher_schedule[tid])
    
    # Group needs by subject and count total periods per subject
    subject_periods = {}
    for need in needs:
        sid = need['subject_id']
        if sid not in subject_periods:
            subject_periods[sid] = 0
        subject_periods[sid] += 1
    
    # Sort by:
    # 1. Subjects with MORE periods per week first (harder to fit, need more days)
    # 2. Double periods before singles (harder to place)
    # 3. Subjects with constraints (not first/last)
    # 4. Subjects with teachers over those without
    needs.sort(key=lambda x: (
        -subject_periods.get(x['subject_id'], 0),  # More periods first
        -1 if x['is_double'] else 0,  # Doubles first
        -1 if x['not_first'] or x['not_last'] else 0,  # Constrained subjects
        0 if x['teacher_id'] else 1,  # Subjects with teachers
    ))
    
    # Calculate total periods needed vs available slots
    total_periods_needed = len(needs)  # Each need is 1 period (doubles counted as 2 separate needs)
    total_slots_available = periods_per_day * 5  # 5 days
    
    if total_periods_needed > total_slots_available:
        conflicts.append(f"WARNING: {cc.class_name} {arm} needs {total_periods_needed} periods but only {total_slots_available} slots available!")
    
    # Track teacher periods per day and week
    teacher_day_count = {}  # {teacher_id: {day: count}}
    teacher_week_count = {}  # {teacher_id: total_count}
    
    def get_teacher_day_count(tid, day):
        if not tid:
            return 0
        if tid not in teacher_day_count:
            teacher_day_count[tid] = {d: 0 for d in range(5)}
        return teacher_day_count[tid].get(day, 0)
    
    def get_teacher_week_count(tid):
        if not tid:
            return 0
        return teacher_week_count.get(tid, 0)
    
    def increment_teacher_count(tid, day, count=1):
        if not tid:
            return
        if tid not in teacher_day_count:
            teacher_day_count[tid] = {d: 0 for d in range(5)}
        teacher_day_count[tid][day] = teacher_day_count[tid].get(day, 0) + count
        teacher_week_count[tid] = teacher_week_count.get(tid, 0) + count
    
    def is_teacher_free(tid, day, period):
        if not tid:
            return True
        # Check availability table
        if GenTeacherAvailability.query.filter_by(
            teacher_id=tid, day_of_week=day, period_number=period, is_available=False
        ).first():
            return False
        # Check clash with other classes
        if tid in teacher_schedule and (day, period) in teacher_schedule[tid]:
            return False
        return True
    
    def can_teacher_take_more(tid, day):
        """Check if teacher can take more periods on this day and week"""
        if not tid:
            return True
        teacher = db.session.get(GenTeacher, tid)
        if not teacher:
            return True
        # Check daily limit
        if get_teacher_day_count(tid, day) >= teacher.max_periods_per_day:
            return False
        # Check weekly limit
        if get_teacher_week_count(tid) >= teacher.max_periods_per_week:
            return False
        return True
    
    def get_teacher_consecutive_at(tid, day, period):
        """Count how many consecutive periods teacher already has ending at this period"""
        if not tid:
            return 0
        count = 0
        p = period - 1
        while p >= 1:
            # Break resets consecutive count
            if p == break_after:
                break
            if tid in teacher_schedule and (day, p) in teacher_schedule[tid]:
                count += 1
                p -= 1
            else:
                break
        return count
    
    def would_exceed_consecutive(tid, day, period, adding=1):
        """Check if assigning this period would exceed max consecutive for teacher"""
        if not tid:
            return False
        
        # Count consecutive before this period
        before = get_teacher_consecutive_at(tid, day, period)
        
        # Count consecutive after this period
        after = 0
        p = period + adding  # If adding double, start check from period + 2
        while p <= periods_per_day:
            if p == break_after + 1:  # Just after break
                break  # Break resets
            if tid in teacher_schedule and (day, p) in teacher_schedule[tid]:
                after += 1
                p += 1
            else:
                break
        
        total_consecutive = before + adding + after
        return total_consecutive > max_consecutive
    
    def count_on_day(sid, day):
        """Count how many times a subject appears on a day for THIS class arm"""
        return sum(1 for p in range(1, periods_per_day + 1) 
                   if timetable[day][p] and not timetable[day][p].get('is_break') 
                   and timetable[day][p].get('subject_id') == sid)
    
    def subject_already_on_day(sid, day):
        """Check if subject has ANY appearance on this day (single or double)"""
        for p in range(1, periods_per_day + 1):
            if timetable[day][p] and timetable[day][p].get('subject_id') == sid:
                return True
        return False
    
    def teacher_day_load(tid, day):
        """Count how many periods teacher has on this day across ALL classes"""
        if not tid or tid not in teacher_schedule:
            return 0
        return sum(1 for (d, p) in teacher_schedule[tid] if d == day)
    
    def sort_slots_for_teacher(slots, tid):
        """Sort slots to prefer days where teacher has fewer periods (better distribution)"""
        if not tid:
            return slots
        # Sort by: (teacher's load on that day, period number)
        return sorted(slots, key=lambda s: (teacher_day_load(tid, s[0]), s[1]))
    
    # First pass - assign with all constraints
    unassigned = []
    for need in needs:
        assigned = False
        slots = [(d, p) for d in range(5) for p in range(1, periods_per_day + 1) if timetable[d][p] is None]
        
        # Sort slots to:
        # 1. Prefer days where this subject hasn't appeared yet (spread across week)
        # 2. Then by teacher's load on that day (spread teacher's work)
        # 3. Then by period number
        def slot_score(slot):
            day, period = slot
            subject_on_day = 1 if subject_already_on_day(need['subject_id'], day) else 0
            teacher_load = teacher_day_load(need['teacher_id'], day) if need['teacher_id'] else 0
            return (subject_on_day, teacher_load, period)
        
        slots.sort(key=slot_score)
        
        for day, period in slots:
            if timetable[day][period]:
                continue
            if need['not_first'] and period == 1:
                continue
            if need['not_last'] and period == periods_per_day:
                continue
            if need['preferred'] == 'morning' and period > break_after:
                continue
            if need['preferred'] == 'afternoon' and period <= break_after:
                continue
            if not is_teacher_free(need['teacher_id'], day, period):
                continue
            if not can_teacher_take_more(need['teacher_id'], day):
                continue
            
            # NO subject repeat on same day for same arm - if subject already appeared, skip this day
            if subject_already_on_day(need['subject_id'], day):
                continue
            
            # Check consecutive periods for teacher
            if would_exceed_consecutive(need['teacher_id'], day, period, 1):
                continue
            
            if need['is_double']:
                np = period + 1
                # Double period cannot span across break
                if period == break_after:
                    continue  # Can't start double at last morning period
                if np > periods_per_day or timetable[day][np]:
                    continue
                if not is_teacher_free(need['teacher_id'], day, np):
                    continue
                if would_exceed_consecutive(need['teacher_id'], day, period, 2):
                    continue
                
                timetable[day][period] = {'subject_id': need['subject_id'], 'teacher_id': need['teacher_id'], 'is_double': True}
                timetable[day][np] = {'subject_id': need['subject_id'], 'teacher_id': need['teacher_id'], 'is_double': True}
                
                if need['teacher_id']:
                    if need['teacher_id'] not in teacher_schedule:
                        teacher_schedule[need['teacher_id']] = {}
                    teacher_schedule[need['teacher_id']][(day, period)] = (cc.class_name, arm)
                    teacher_schedule[need['teacher_id']][(day, np)] = (cc.class_name, arm)
                    increment_teacher_count(need['teacher_id'], day, 2)
            else:
                timetable[day][period] = {'subject_id': need['subject_id'], 'teacher_id': need['teacher_id'], 'is_double': False}
                if need['teacher_id']:
                    if need['teacher_id'] not in teacher_schedule:
                        teacher_schedule[need['teacher_id']] = {}
                    teacher_schedule[need['teacher_id']][(day, period)] = (cc.class_name, arm)
                    increment_teacher_count(need['teacher_id'], day, 1)
            
            assigned = True
            break
        
        if not assigned:
            unassigned.append(need)
    
    # Second pass - try unassigned with relaxed constraints (ignore preferred time, teacher limits, consecutive)
    # But STILL enforce: no subject repeat same day, no teacher clash, no double across break
    still_unassigned = []
    for need in unassigned:
        assigned = False
        slots = [(d, p) for d in range(5) for p in range(1, periods_per_day + 1) if timetable[d][p] is None]
        
        # Sort by teacher's availability on that day (prefer days where teacher is less busy)
        if need['teacher_id']:
            slots = sort_slots_for_teacher(slots, need['teacher_id'])
        else:
            shuffle(slots)
        
        for day, period in slots:
            if timetable[day][period]:
                continue
            # Must still check: teacher free (no clash)
            if not is_teacher_free(need['teacher_id'], day, period):
                continue
            # Must still check: no subject repeat same day AT ALL
            if subject_already_on_day(need['subject_id'], day):
                continue
            
            if need['is_double']:
                np = period + 1
                # No double across break
                if period == break_after:
                    continue
                if np > periods_per_day or timetable[day][np]:
                    continue
                if not is_teacher_free(need['teacher_id'], day, np):
                    continue
                
                timetable[day][period] = {'subject_id': need['subject_id'], 'teacher_id': need['teacher_id'], 'is_double': True}
                timetable[day][np] = {'subject_id': need['subject_id'], 'teacher_id': need['teacher_id'], 'is_double': True}
                
                if need['teacher_id']:
                    if need['teacher_id'] not in teacher_schedule:
                        teacher_schedule[need['teacher_id']] = {}
                    teacher_schedule[need['teacher_id']][(day, period)] = (cc.class_name, arm)
                    teacher_schedule[need['teacher_id']][(day, np)] = (cc.class_name, arm)
            else:
                timetable[day][period] = {'subject_id': need['subject_id'], 'teacher_id': need['teacher_id'], 'is_double': False}
                if need['teacher_id']:
                    if need['teacher_id'] not in teacher_schedule:
                        teacher_schedule[need['teacher_id']] = {}
                    teacher_schedule[need['teacher_id']][(day, period)] = (cc.class_name, arm)
            
            assigned = True
            break
        
        if not assigned:
            still_unassigned.append(need)
    
    # Log conflicts for truly unassigned subjects
    for need in still_unassigned:
        conflicts.append(f"Could not assign {need['name']} for {cc.class_name} {arm}")
    
    # Save
    for day in range(5):
        for period in range(1, periods_per_day + 1):
            entry = timetable[day][period]
            if entry and not entry.get('is_break'):
                db.session.add(GenTimetableResult(
                    batch_id=batch_id, class_name=cc.class_name, arm_name=arm,
                    day_of_week=day, period_number=period,
                    subject_id=entry.get('subject_id'), teacher_id=entry.get('teacher_id'),
                    is_double_period=entry.get('is_double', False)
                ))
    
    return {'success': len(conflicts) == 0, 'conflicts': conflicts}


# ============================================================================
# VIEW & OUTPUT
# ============================================================================

@generator_bp.route('/results')
@login_required
def results_list():
    level = get_current_level()
    batches = db.session.query(
        GenTimetableResult.batch_id,
        db.func.min(GenTimetableResult.generated_at).label('generated_at'),
        db.func.count(db.distinct(GenTimetableResult.class_name + GenTimetableResult.arm_name)).label('class_count')
    ).filter(GenTimetableResult.school_level == level).group_by(GenTimetableResult.batch_id).order_by(db.desc('generated_at')).all()
    return render_template('generator/results_list.html', batches=batches, level=level)


@generator_bp.route('/results/<batch_id>')
@login_required
def view_results(batch_id):
    level = get_current_level()
    results = GenTimetableResult.query.filter_by(batch_id=batch_id, school_level=level).all()
    if not results:
        flash('No results found.', 'error')
        return redirect(url_for('generator.results_list'))
    
    timetables = {}
    for r in results:
        key = f"{r.class_name}_{r.arm_name}"
        if key not in timetables:
            timetables[key] = {'class_name': r.class_name, 'arm_name': r.arm_name, 'grid': {d: {} for d in range(5)}}
        timetables[key]['grid'][r.day_of_week][r.period_number] = r
    
    rules = {r.rule_type: r.value for r in GenTimetableRule.query.filter_by(is_active=True, school_level=level).all()}
    
    return render_template('generator/view_results.html',
        batch_id=batch_id, timetables=dict(sorted(timetables.items())),
        periods_per_day=int(rules.get('periods_per_day', 8)),
        break_after_period=int(rules.get('break_after_period', 4)),
        days=DAYS_OF_WEEK, level=level
    )


@generator_bp.route('/results/<batch_id>/edit/<class_name>/<arm_name>')
@login_required
def edit_timetable(batch_id, class_name, arm_name):
    level = get_current_level()
    results = GenTimetableResult.query.filter_by(batch_id=batch_id, class_name=class_name, arm_name=arm_name, school_level=level).all()
    grid = {d: {} for d in range(5)}
    for r in results:
        grid[r.day_of_week][r.period_number] = r
    
    rules = {r.rule_type: r.value for r in GenTimetableRule.query.filter_by(is_active=True, school_level=level).all()}
    
    # Get subjects for this level
    subjects = GenSubject.query.filter_by(is_active=True, school_level=level).order_by(GenSubject.name).all()
    
    teachers = GenTeacher.query.filter_by(is_active=True, school_level=level).order_by(GenTeacher.name).all()
    
    return render_template('generator/edit_timetable.html',
        batch_id=batch_id, class_name=class_name, arm_name=arm_name, grid=grid,
        periods_per_day=int(rules.get('periods_per_day', 8)),
        break_after_period=int(rules.get('break_after_period', 4)),
        days=DAYS_OF_WEEK, subjects=subjects, teachers=teachers, level=level
    )


@generator_bp.route('/results/<batch_id>/edit/<class_name>/<arm_name>/save', methods=['POST'])
@login_required
def save_timetable_edit(batch_id, class_name, arm_name):
    level = get_current_level()
    try:
        rules = {r.rule_type: r.value for r in GenTimetableRule.query.filter_by(is_active=True, school_level=level).all()}
        periods_per_day = int(rules.get('periods_per_day', 8))
        break_after = int(rules.get('break_after_period', 4))
        
        for day in range(5):
            for period in range(1, periods_per_day + 1):
                if period == break_after:
                    continue
                
                subject_id = request.form.get(f'subject_{day}_{period}', type=int)
                teacher_id = request.form.get(f'teacher_{day}_{period}', type=int)
                is_locked = request.form.get(f'locked_{day}_{period}') == 'on'
                
                existing = GenTimetableResult.query.filter_by(
                    batch_id=batch_id, class_name=class_name, arm_name=arm_name,
                    day_of_week=day, period_number=period, school_level=level
                ).first()
                
                if subject_id:
                    if existing:
                        existing.subject_id = subject_id
                        existing.teacher_id = teacher_id
                        existing.is_locked = is_locked
                    else:
                        db.session.add(GenTimetableResult(
                            batch_id=batch_id, class_name=class_name, arm_name=arm_name,
                            day_of_week=day, period_number=period, school_level=level,
                            subject_id=subject_id, teacher_id=teacher_id, is_locked=is_locked
                        ))
                elif existing:
                    db.session.delete(existing)
        
        db.session.commit()
        flash('Timetable updated!', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Error: {str(e)}', 'error')
    return redirect(url_for('generator.view_results', batch_id=batch_id))


@generator_bp.route('/results/<batch_id>/print')
@login_required
def print_results(batch_id):
    level = get_current_level()
    results = GenTimetableResult.query.filter_by(batch_id=batch_id, school_level=level).all()
    if not results:
        flash('No results.', 'error')
        return redirect(url_for('generator.results_list'))
    
    timetables = {}
    for r in results:
        key = f"{r.class_name}_{r.arm_name}"
        if key not in timetables:
            timetables[key] = {'class_name': r.class_name, 'arm_name': r.arm_name, 'grid': {d: {} for d in range(5)}}
        timetables[key]['grid'][r.day_of_week][r.period_number] = r
    
    rules = {r.rule_type: r.value for r in GenTimetableRule.query.filter_by(is_active=True).all()}
    
    return render_template('generator/print_results.html',
        batch_id=batch_id, timetables=dict(sorted(timetables.items())),
        periods_per_day=int(rules.get('periods_per_day', 8)),
        break_after_period=int(rules.get('break_after_period', 4)),
        days=DAYS_OF_WEEK,
        school_name=GenSettings.get('school_name', ''),
        academic_year=GenSettings.get('academic_year', ''),
        term_name=GenSettings.get('term_name', '')
    )


@generator_bp.route('/results/<batch_id>/print/<class_name>/<arm_name>')
@login_required
def print_single_timetable(batch_id, class_name, arm_name):
    results = GenTimetableResult.query.filter_by(batch_id=batch_id, class_name=class_name, arm_name=arm_name).all()
    if not results:
        flash('Not found.', 'error')
        return redirect(url_for('generator.view_results', batch_id=batch_id))
    
    grid = {d: {} for d in range(5)}
    for r in results:
        grid[r.day_of_week][r.period_number] = r
    
    rules = {r.rule_type: r.value for r in GenTimetableRule.query.filter_by(is_active=True).all()}
    
    return render_template('generator/print_single.html',
        batch_id=batch_id, class_name=class_name, arm_name=arm_name, grid=grid,
        periods_per_day=int(rules.get('periods_per_day', 8)),
        break_after_period=int(rules.get('break_after_period', 4)),
        days=DAYS_OF_WEEK,
        school_name=GenSettings.get('school_name', ''),
        academic_year=GenSettings.get('academic_year', ''),
        term_name=GenSettings.get('term_name', '')
    )


@generator_bp.route('/results/<batch_id>/export')
@login_required
def export_results(batch_id):
    """Export timetable by class - Each class MAXIMIZES A4 landscape page, no teacher names, NO COLORS for B&W printing"""
    import openpyxl
    from openpyxl.styles import Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    
    results = GenTimetableResult.query.filter_by(batch_id=batch_id).all()
    if not results:
        flash('No results.', 'error')
        return redirect(url_for('generator.results_list'))
    
    # Get school info
    school_name = GenSettings.get('school_name', 'School')
    school_address = GenSettings.get('school_address', '')
    
    timetables = {}
    for r in results:
        key = f"{r.class_name}_{r.arm_name}"
        if key not in timetables:
            timetables[key] = {'class_name': r.class_name, 'arm_name': r.arm_name, 'grid': {d: {} for d in range(5)}}
        timetables[key]['grid'][r.day_of_week][r.period_number] = r
    
    rules = {r.rule_type: r.value for r in GenTimetableRule.query.filter_by(is_active=True).all()}
    periods_per_day = int(rules.get('periods_per_day', 8))
    
    # Period time slots (8:20 AM start, 40 min each, 30 min break after period 5)
    period_times = []
    break_time = ""
    start_hour, start_min = 8, 20
    for p in range(1, periods_per_day + 1):
        start_total = start_hour * 60 + start_min
        end_total = start_total + 40
        start_h, start_m = start_total // 60, start_total % 60
        end_h, end_m = end_total // 60, end_total % 60
        period_times.append(f"P{p}\n{start_h}:{start_m:02d}-{end_h}:{end_m:02d}")
        start_hour, start_min = end_h, end_m
        if p == 5:
            break_start = f"{end_h}:{end_m:02d}"
            start_total = start_hour * 60 + start_min + 30
            start_hour, start_min = start_total // 60, start_total % 60
            break_end = f"{start_hour}:{start_min:02d}"
            break_time = f"BREAK\n{break_start}\n-\n{break_end}"
    
    # Abbreviations
    abbrev_map = {
        'Mathematics': 'Maths', 'English Language': 'Eng', 'Physics': 'Phy',
        'Chemistry': 'Chem', 'Biology': 'Bio', 'Economics': 'Econs',
        'Government': 'Govt', 'Literature in English': 'Lit', 'Agricultural Science': 'Agric',
        'Christian Religious Studies': 'CRS', 'Civic Education': 'Civic',
        'Computer Studies': 'Comp', 'Commerce': 'Comm', 'Geography': 'Geo',
        'Further Mathematics': 'F/Mth', 'Livestock Farming': 'Livst',
        'History': 'Hist', 'Phonics': 'Phon',
    }
    
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    
    # NO COLORS - Pure black text on white, maximum contrast for B&W printing
    school_font = Font(bold=True, size=32, color='000000')
    address_font = Font(bold=False, size=14, color='000000')
    class_title_font = Font(bold=True, size=28, color='000000')
    day_font = Font(bold=True, size=24, color='000000')
    header_font = Font(bold=True, size=11, color='000000')
    cell_font = Font(bold=True, size=32, color='000000')
    break_header_font = Font(bold=True, size=8, color='000000')
    
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    thick_border = Border(
        left=Side(style='medium'), right=Side(style='medium'),
        top=Side(style='medium'), bottom=Side(style='medium')
    )
    
    center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    days = ['Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday']
    
    # Total columns: Day label + P1-P5 + BREAK + P6-P8 = 1 + 5 + 1 + 3 = 10
    total_cols = 1 + 5 + 1 + (periods_per_day - 5)
    
    # A4 Landscape: 297mm x 210mm
    # Safe margins for most printers: 0.5" (12.7mm) each side
    # Usable: ~272mm x 185mm
    total_page_height = 500  # points for A4 landscape height (conservative)
    school_header_height = 38
    address_header_height = 20 if school_address else 0
    class_title_height = 32
    period_header_height = 45
    # 5 day rows get ALL remaining space
    fixed_height = school_header_height + address_header_height + class_title_height + period_header_height
    day_row_height = (total_page_height - fixed_height) / 5
    
    for key in sorted(timetables.keys()):
        tt = timetables[key]
        ws = wb.create_sheet(title=f"{tt['class_name']} {tt['arm_name']}"[:31])
        
        # Page setup for A4 landscape - fit on ONE page
        ws.page_setup.paperSize = ws.PAPERSIZE_A4
        ws.page_setup.orientation = 'landscape'
        ws.page_setup.fitToPage = True
        ws.page_setup.fitToWidth = 1
        ws.page_setup.fitToHeight = 1
        ws.print_options.horizontalCentered = True
        ws.print_options.verticalCentered = True
        
        # Safe margins for most printers (0.5 inch = 12.7mm)
        ws.page_margins.left = 0.5
        ws.page_margins.right = 0.5
        ws.page_margins.top = 0.4
        ws.page_margins.bottom = 0.4
        ws.page_margins.header = 0
        ws.page_margins.footer = 0
        
        current_row = 1
        
        # School name
        ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=total_cols)
        cell = ws.cell(row=current_row, column=1, value=school_name.upper())
        cell.font = school_font
        cell.alignment = center_align
        ws.row_dimensions[current_row].height = school_header_height
        current_row += 1
        
        # School address
        if school_address:
            ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=total_cols)
            cell = ws.cell(row=current_row, column=1, value=school_address)
            cell.font = address_font
            cell.alignment = center_align
            ws.row_dimensions[current_row].height = address_header_height
            current_row += 1
        
        # Class title
        ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=total_cols)
        cell = ws.cell(row=current_row, column=1, value=f"{tt['class_name']} {tt['arm_name']} - TIMETABLE")
        cell.font = class_title_font
        cell.alignment = center_align
        cell.border = thick_border
        for col in range(1, total_cols + 1):
            ws.cell(row=current_row, column=col).border = thick_border
        ws.row_dimensions[current_row].height = class_title_height
        current_row += 1
        
        # Period header row (times)
        col = 1
        cell = ws.cell(row=current_row, column=col, value="Day")
        cell.font = header_font
        cell.border = thick_border
        cell.alignment = center_align
        col += 1
        
        # P1-P5 headers
        for i in range(5):
            cell = ws.cell(row=current_row, column=col, value=period_times[i])
            cell.font = header_font
            cell.border = thick_border
            cell.alignment = center_align
            col += 1
        
        # Break header
        cell = ws.cell(row=current_row, column=col, value=break_time)
        cell.font = break_header_font
        cell.border = thick_border
        cell.alignment = center_align
        col += 1
        
        # P6-P8 headers
        for i in range(5, periods_per_day):
            cell = ws.cell(row=current_row, column=col, value=period_times[i])
            cell.font = header_font
            cell.border = thick_border
            cell.alignment = center_align
            col += 1
        
        ws.row_dimensions[current_row].height = period_header_height
        current_row += 1
        
        # Data rows (one per day)
        for day_idx, day_name in enumerate(days):
            col = 1
            
            # Day name
            cell = ws.cell(row=current_row, column=col, value=day_name)
            cell.font = day_font
            cell.border = thick_border
            cell.alignment = center_align
            col += 1
            
            # P1-P5
            for p in range(1, 6):
                entry = tt['grid'][day_idx].get(p)
                value = ""
                if entry and entry.subject:
                    subj_name = entry.subject.name
                    value = abbrev_map.get(subj_name, subj_name[:6])
                
                cell = ws.cell(row=current_row, column=col, value=value)
                cell.font = cell_font
                cell.border = thin_border
                cell.alignment = center_align
                col += 1
            
            # Break column - empty with border
            cell = ws.cell(row=current_row, column=col, value="")
            cell.border = thin_border
            col += 1
            
            # P6-P8
            for p in range(6, periods_per_day + 1):
                entry = tt['grid'][day_idx].get(p)
                value = ""
                if entry and entry.subject:
                    subj_name = entry.subject.name
                    value = abbrev_map.get(subj_name, subj_name[:6])
                
                cell = ws.cell(row=current_row, column=col, value=value)
                cell.font = cell_font
                cell.border = thin_border
                cell.alignment = center_align
                col += 1
            
            ws.row_dimensions[current_row].height = day_row_height
            current_row += 1
        
        # Column widths - fit within safe printable area
        # A4 landscape with 0.5" margins: ~250mm usable width
        ws.column_dimensions['A'].width = 12  # Day column
        period_col_width = 28  # Period columns
        break_col_width = 8   # Break column
        
        for col in range(2, total_cols + 1):
            if col == 7:  # Break column
                ws.column_dimensions[get_column_letter(col)].width = break_col_width
            else:
                ws.column_dimensions[get_column_letter(col)].width = period_col_width
    
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    return Response(output.getvalue(),
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={'Content-Disposition': f'attachment; filename=timetables_{batch_id}.xlsx'}
    )


@generator_bp.route('/results/<batch_id>/delete', methods=['POST'])
@login_required
def delete_results(batch_id):
    try:
        GenTimetableResult.query.filter_by(batch_id=batch_id).delete()
        db.session.commit()
        flash('Deleted.', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Error: {str(e)}', 'error')
    return redirect(url_for('generator.results_list'))


# ============================================================================
# REPORTS
# ============================================================================

@generator_bp.route('/reports')
@login_required
def reports_index():
    latest = GenTimetableResult.query.order_by(GenTimetableResult.generated_at.desc()).first()
    return render_template('generator/reports_index.html', latest_batch=latest)


@generator_bp.route('/reports/period-count/<batch_id>')
@login_required
def period_count_report(batch_id):
    results = GenTimetableResult.query.filter_by(batch_id=batch_id).all()
    counts = {}
    for r in results:
        key = f"{r.class_name} {r.arm_name}"
        if key not in counts:
            counts[key] = {}
        if r.subject_id:
            counts[key][r.subject_id] = counts[key].get(r.subject_id, 0) + 1
    
    level = get_current_level()
    configs = {c.subject_id: c for c in GenSubjectConfig.query.filter_by(school_level=level).all()}
    subjects = GenSubject.query.filter_by(is_active=True, school_level=level).order_by(GenSubject.name).all()
    
    return render_template('generator/report_period_count.html',
        batch_id=batch_id, counts=counts, configs=configs, subjects=subjects
    )


@generator_bp.route('/reports/teacher-workload/<batch_id>')
@login_required
def teacher_workload_report(batch_id):
    results = GenTimetableResult.query.filter_by(batch_id=batch_id).all()
    workload = {}
    for r in results:
        if r.teacher_id:
            if r.teacher_id not in workload:
                workload[r.teacher_id] = {
                    'teacher': r.teacher, 'total': 0,
                    'per_day': {d: 0 for d in range(5)}, 'classes': set()
                }
            workload[r.teacher_id]['total'] += 1
            workload[r.teacher_id]['per_day'][r.day_of_week] += 1
            workload[r.teacher_id]['classes'].add(f"{r.class_name} {r.arm_name}")
    
    return render_template('generator/report_teacher_workload.html',
        batch_id=batch_id, workload=workload, days=DAYS_OF_WEEK
    )


@generator_bp.route('/reports/clashes/<batch_id>')
@login_required
def clash_report(batch_id):
    results = GenTimetableResult.query.filter_by(batch_id=batch_id).all()
    teacher_slots = {}
    for r in results:
        if r.teacher_id:
            if r.teacher_id not in teacher_slots:
                teacher_slots[r.teacher_id] = {}
            key = (r.day_of_week, r.period_number)
            if key not in teacher_slots[r.teacher_id]:
                teacher_slots[r.teacher_id][key] = []
            teacher_slots[r.teacher_id][key].append({
                'class': f"{r.class_name} {r.arm_name}",
                'subject': r.subject.name if r.subject else '-'
            })
    
    clashes = []
    for tid, slots in teacher_slots.items():
        teacher = db.session.get(GenTeacher, tid)
        for (day, period), entries in slots.items():
            if len(entries) > 1:
                clashes.append({
                    'teacher': teacher, 'day': DAYS_OF_WEEK[day],
                    'period': period, 'entries': entries
                })
    
    return render_template('generator/report_clashes.html', batch_id=batch_id, clashes=clashes)


@generator_bp.route('/reports/unassigned/<batch_id>')
@login_required
def unassigned_report(batch_id):
    rules = {r.rule_type: r.value for r in GenTimetableRule.query.filter_by(is_active=True).all()}
    periods_per_day = int(rules.get('periods_per_day', 8))
    break_after = int(rules.get('break_after_period', 4))
    
    results = GenTimetableResult.query.filter_by(batch_id=batch_id).all()
    class_arms = set((r.class_name, r.arm_name) for r in results)
    
    empty = []
    for cn, an in class_arms:
        filled = {(r.day_of_week, r.period_number) for r in results if r.class_name == cn and r.arm_name == an}
        for day in range(5):
            for period in range(1, periods_per_day + 1):
                if period != break_after and (day, period) not in filled:
                    empty.append({'class': cn, 'arm': an, 'day': DAYS_OF_WEEK[day], 'period': period})
    
    return render_template('generator/report_unassigned.html', batch_id=batch_id, empty_slots=empty)


# ============================================================================
# SPECIAL VIEWS
# ============================================================================

@generator_bp.route('/teacher-timetable')
@generator_bp.route('/teacher-timetable')
@login_required
def teacher_timetable():
    """View individual teacher's timetable with workload breakdown"""
    from collections import defaultdict
    
    teacher_id = request.args.get('teacher_id', type=int)
    batch_id = request.args.get('batch_id')
    
    teachers = GenTeacher.query.filter_by(is_active=True).order_by(GenTeacher.name).all()
    
    if not batch_id:
        latest = GenTimetableResult.query.order_by(GenTimetableResult.generated_at.desc()).first()
        if latest:
            batch_id = latest.batch_id
    
    batches = db.session.query(
        GenTimetableResult.batch_id,
        db.func.min(GenTimetableResult.generated_at).label('generated_at')
    ).group_by(GenTimetableResult.batch_id).order_by(db.desc('generated_at')).all()
    
    teacher_grid = None
    selected_teacher = None
    total_periods = 0
    breakdown = {
        'by_subject': {},
        'by_class': {},
        'by_class_arm': {},
        'by_day': {},
        'by_stream': {},
        'details': []
    }
    
    if teacher_id and batch_id:
        selected_teacher = db.session.get(GenTeacher, teacher_id)
        results = GenTimetableResult.query.filter_by(batch_id=batch_id, teacher_id=teacher_id).all()
        
        rules = {r.rule_type: r.value for r in GenTimetableRule.query.filter_by(is_active=True).all()}
        teacher_grid = {d: {} for d in range(5)}
        
        # For breakdown statistics
        by_subject = defaultdict(int)
        by_class = defaultdict(int)
        by_class_arm = defaultdict(int)
        by_day = defaultdict(int)
        by_stream = defaultdict(int)
        class_arm_subject_periods = defaultdict(int)
        day_names = ['Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday']
        
        for r in results:
            teacher_grid[r.day_of_week][r.period_number] = r
            total_periods += 1
            
            subject_name = r.subject.name if r.subject else 'Unknown'
            class_name = r.class_name
            arm_name = r.arm_name
            class_arm = f"{class_name} {arm_name}"
            
            # Count by different categories
            by_subject[subject_name] += 1
            by_class[class_name] += 1
            by_class_arm[class_arm] += 1
            by_day[day_names[r.day_of_week]] += 1
            
            # Track for detailed breakdown
            class_arm_subject_periods[(class_name, arm_name, subject_name)] += 1
            
            # Get stream for this class-arm
            class_config = GenClassConfig.query.filter_by(class_name=class_name).first()
            if class_config and class_config.has_streams:
                arm_stream = GenClassArmStream.query.filter_by(
                    class_config_id=class_config.id, 
                    arm_name=arm_name
                ).first()
                if arm_stream and arm_stream.stream:
                    by_stream[arm_stream.stream.name] += 1
        
        # Build detailed breakdown list
        for (class_name, arm_name, subject_name), periods in sorted(class_arm_subject_periods.items()):
            stream_name = '-'
            class_config = GenClassConfig.query.filter_by(class_name=class_name).first()
            if class_config and class_config.has_streams:
                arm_stream = GenClassArmStream.query.filter_by(
                    class_config_id=class_config.id, 
                    arm_name=arm_name
                ).first()
                if arm_stream and arm_stream.stream:
                    stream_name = arm_stream.stream.name
            
            breakdown['details'].append({
                'class': class_name,
                'arm': arm_name,
                'subject': subject_name,
                'stream': stream_name,
                'periods': periods
            })
        
        # Convert to regular dicts
        breakdown['by_subject'] = dict(by_subject)
        breakdown['by_class'] = dict(by_class)
        breakdown['by_class_arm'] = dict(by_class_arm)
        breakdown['by_day'] = dict(by_day)
        breakdown['by_stream'] = dict(by_stream)
    
    rules = {r.rule_type: r.value for r in GenTimetableRule.query.filter_by(is_active=True).all()}
    
    return render_template('generator/teacher_timetable.html',
        teachers=teachers, teacher_id=teacher_id, selected_teacher=selected_teacher,
        batches=batches, batch_id=batch_id, teacher_grid=teacher_grid,
        total_periods=total_periods,
        breakdown=breakdown,
        periods_per_day=int(rules.get('periods_per_day', 8)),
        break_after_period=int(rules.get('break_after_period', 4)),
        days=DAYS_OF_WEEK
    )


@generator_bp.route('/master-timetable')
@login_required
def master_timetable():
    batch_id = request.args.get('batch_id')
    selected_day = request.args.get('day', type=int, default=0)
    selected_period = request.args.get('period', type=int, default=1)
    
    if not batch_id:
        latest = GenTimetableResult.query.order_by(GenTimetableResult.generated_at.desc()).first()
        if latest:
            batch_id = latest.batch_id
    
    batches = db.session.query(
        GenTimetableResult.batch_id,
        db.func.min(GenTimetableResult.generated_at).label('generated_at')
    ).group_by(GenTimetableResult.batch_id).order_by(db.desc('generated_at')).all()
    
    rules = {r.rule_type: r.value for r in GenTimetableRule.query.filter_by(is_active=True).all()}
    
    master_data = []
    if batch_id:
        results = GenTimetableResult.query.filter_by(
            batch_id=batch_id, day_of_week=selected_day, period_number=selected_period
        ).order_by(GenTimetableResult.class_name, GenTimetableResult.arm_name).all()
        
        for r in results:
            master_data.append({
                'class_name': r.class_name, 'arm_name': r.arm_name,
                'subject': r.subject.name if r.subject else '-',
                'teacher': r.teacher.name if r.teacher else '-'
            })
    
    return render_template('generator/master_timetable.html',
        batches=batches, batch_id=batch_id,
        selected_day=selected_day, selected_period=selected_period,
        periods_per_day=int(rules.get('periods_per_day', 8)),
        break_after_period=int(rules.get('break_after_period', 4)),
        days=DAYS_OF_WEEK, master_data=master_data
    )


# ============================================================================
# API
# ============================================================================

@generator_bp.route('/api/teacher/<int:teacher_id>/subjects')
@login_required
def api_teacher_subjects(teacher_id):
    teacher = db.get_or_404(GenTeacher, teacher_id)
    return jsonify([{'id': ts.subject_id, 'name': ts.subject.name} for ts in teacher.subjects])


@generator_bp.route('/api/class/<int:class_id>/arms')
@login_required
def api_class_arms(class_id):
    cc = db.get_or_404(GenClassConfig, class_id)
    return jsonify(cc.arm_list)


@generator_bp.route('/api/stream/<int:stream_id>/subjects')
@login_required
def api_stream_subjects(stream_id):
    stream = db.get_or_404(GenStream, stream_id)
    return jsonify([{'id': ss.subject_id, 'name': ss.subject.name, 'compulsory': ss.is_compulsory} for ss in stream.subjects])


@generator_bp.route('/api/swap', methods=['POST'])
@login_required
def api_swap_slots():
    try:
        data = request.get_json()
        batch_id = data['batch_id']
        class_name = data['class_name']
        arm_name = data['arm_name']
        slot1, slot2 = data['slot1'], data['slot2']
        
        e1 = GenTimetableResult.query.filter_by(
            batch_id=batch_id, class_name=class_name, arm_name=arm_name,
            day_of_week=slot1['day'], period_number=slot1['period']
        ).first()
        e2 = GenTimetableResult.query.filter_by(
            batch_id=batch_id, class_name=class_name, arm_name=arm_name,
            day_of_week=slot2['day'], period_number=slot2['period']
        ).first()
        
        if e1 and e2:
            e1.subject_id, e2.subject_id = e2.subject_id, e1.subject_id
            e1.teacher_id, e2.teacher_id = e2.teacher_id, e1.teacher_id
            db.session.commit()
        
        return jsonify({'success': True})
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)})


@generator_bp.route('/results/<batch_id>/export_by_day')
@login_required
def export_results_by_day(batch_id):
    """Export timetable in day-wise format - MAXIMUM SIZE for A4 landscape printing, NO COLORS"""
    import openpyxl
    from openpyxl.styles import Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    
    results = GenTimetableResult.query.filter_by(batch_id=batch_id).all()
    if not results:
        flash('No results.', 'error')
        return redirect(url_for('generator.results_list'))
    
    # Determine school level from results
    school_level = results[0].school_level if results else 'sss'
    
    rules = {r.rule_type: r.value for r in GenTimetableRule.query.filter_by(is_active=True, school_level=school_level).all()}
    periods_per_day = int(rules.get('periods_per_day', 8))
    
    # Get school info
    school_name = GenSettings.get('school_name', 'School')
    school_address = GenSettings.get('school_address', '')
    
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    
    days = ['Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday']
    
    # Period time slots (8:20 AM start, 40 min each, 30 min break after period 5)
    period_times_before_break = []  # P1-P5
    period_times_after_break = []   # P6-P8
    break_time = ""
    
    start_hour, start_min = 8, 20
    for p in range(1, periods_per_day + 1):
        start_total = start_hour * 60 + start_min
        end_total = start_total + 40
        start_h, start_m = start_total // 60, start_total % 60
        end_h, end_m = end_total // 60, end_total % 60
        start_str = f"{start_h}:{start_m:02d}"
        end_str = f"{end_h}:{end_m:02d}"
        
        if p <= 5:
            period_times_before_break.append(f"P{p}\n{start_str}-{end_str}")
        else:
            period_times_after_break.append(f"P{p}\n{start_str}-{end_str}")
        
        start_hour, start_min = end_h, end_m
        
        # Capture break time after period 5
        if p == 5:
            break_start = f"{end_h}:{end_m:02d}"
            start_total = start_hour * 60 + start_min + 30
            start_hour, start_min = start_total // 60, start_total % 60
            break_end = f"{start_hour}:{start_min:02d}"
            break_time = f"BREAK\n{break_start}\n-\n{break_end}"
    
    class_arms = sorted(set((r.class_name, r.arm_name) for r in results))
    
    def get_short_code(class_name, arm):
        if class_name.startswith('JSS'):
            class_num = class_name.replace('JSS', '')
            arm_letter = arm[0].upper()
            return f"J{class_num}{arm_letter}"
        else:
            class_num = class_name.replace('SSS', '')
            arm_letter = arm[0].upper()
            return f"{class_num}{arm_letter}"
    
    abbrev_map = {
        'Mathematics': 'Maths',
        'English Language': 'Eng',
        'Physics': 'Phy',
        'Chemistry': 'Chem',
        'Biology': 'Bio',
        'Economics': 'Econs',
        'Government': 'Govt',
        'Literature in English': 'Lit',
        'Agricultural Science': 'Agric',
        'Christian Religious Studies': 'CRS',
        'Civic Education': 'Civic',
        'Computer Studies': 'Comp',
        'Commerce': 'Comm',
        'Geography': 'Geo',
        'Further Mathematics': 'F/Mth',
        'Livestock Farming': 'Livst',
        'History': 'Hist',
        'Phonics': 'Phon',
        # JSS subjects
        'Basic Science': 'B.Sci',
        'Basic Technology': 'B.Tech',
        'Social Studies': 'Soc',
        'Home Economics': 'H/Eco',
        'Business Studies': 'Bus',
        'Physical Health Education': 'PHE',
        'Islamic Religious Studies': 'IRS',
        'French': 'Fren',
        'Yoruba': 'Yor',
        'Igbo': 'Igbo',
        'Hausa': 'Hau',
        'Music': 'Music',
        'Fine Art': 'Art',
        'Data Processing': 'D.Pro',
        'Marketing': 'Mkt',
        'Accounting': 'Acct',
        'Food and Nutrition': 'F&N',
    }
    
    # Build a lookup dict for subjects
    subject_lookup = {s.id: s for s in GenSubject.query.filter_by(is_active=True, school_level=school_level).all()}
    
    # NO COLORS - Pure black text on white for B&W printing
    school_font = Font(bold=True, size=24, color='000000')
    address_font = Font(bold=False, size=12, color='000000')
    day_font = Font(bold=True, size=32, color='000000')
    header_font = Font(bold=True, size=14, color='000000')
    class_font = Font(bold=True, size=20, color='000000')
    cell_font = Font(bold=True, size=24, color='000000')
    break_header_font = Font(bold=True, size=9, color='000000')
    
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    thick_border = Border(
        left=Side(style='medium'),
        right=Side(style='medium'),
        top=Side(style='medium'),
        bottom=Side(style='medium')
    )
    
    center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    # Total columns: Class + P1-P5 + BREAK + P6-P8 = 1 + 5 + 1 + 3 = 10
    total_cols = 1 + 5 + 1 + (periods_per_day - 5)
    
    # Calculate heights (conservative for safe printing)
    num_data_rows = len(class_arms)
    school_header_height = 35
    address_header_height = 20
    day_header_height = 45
    period_header_height = 45
    total_page_height = 500  # Conservative for safe printing
    
    for d, day_name in enumerate(days):
        ws = wb.create_sheet(title=day_name)
        
        ws.page_setup.paperSize = ws.PAPERSIZE_A4
        ws.page_setup.orientation = 'landscape'
        ws.page_setup.fitToPage = True
        ws.page_setup.fitToWidth = 1
        ws.page_setup.fitToHeight = 1
        ws.print_options.horizontalCentered = True
        ws.print_options.verticalCentered = True
        
        # Safe margins for most printers (0.5 inch)
        ws.page_margins.left = 0.5
        ws.page_margins.right = 0.5
        ws.page_margins.top = 0.4
        ws.page_margins.bottom = 0.4
        ws.page_margins.header = 0
        ws.page_margins.footer = 0
        
        current_row = 1
        
        # School name and address ONLY on Monday
        if d == 0:
            # School name header
            ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=total_cols)
            cell = ws.cell(row=current_row, column=1, value=school_name.upper())
            cell.font = school_font
            cell.alignment = center_align
            ws.row_dimensions[current_row].height = school_header_height
            current_row += 1
            
            if school_address:
                ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=total_cols)
                cell = ws.cell(row=current_row, column=1, value=school_address)
                cell.font = address_font
                cell.alignment = center_align
                ws.row_dimensions[current_row].height = address_header_height
                current_row += 1
            
            # Recalculate data row height for Monday
            header_rows_height = school_header_height + (address_header_height if school_address else 0) + day_header_height + period_header_height
            remaining_height = total_page_height - header_rows_height
            data_row_height = remaining_height / num_data_rows
            data_row_height = max(data_row_height, 28)
        else:
            # Other days - just day header and period header
            remaining_height = total_page_height - day_header_height - period_header_height
            data_row_height = remaining_height / num_data_rows
            data_row_height = max(data_row_height, 30)
        
        # Day header
        ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=total_cols)
        cell = ws.cell(row=current_row, column=1, value=day_name.upper())
        cell.font = day_font
        cell.alignment = center_align
        cell.border = thick_border
        for col in range(1, total_cols + 1):
            ws.cell(row=current_row, column=col).border = thick_border
        ws.row_dimensions[current_row].height = day_header_height
        current_row += 1
        
        # Period headers with BREAK column
        col = 1
        # Class column
        cell = ws.cell(row=current_row, column=col, value="Class")
        cell.font = header_font
        cell.border = thick_border
        cell.alignment = center_align
        col += 1
        
        # P1-P5
        for i, p in enumerate(range(1, 6)):
            header_value = period_times_before_break[i] if d == 0 else f"P{p}"
            cell = ws.cell(row=current_row, column=col, value=header_value)
            cell.font = header_font
            cell.border = thick_border
            cell.alignment = center_align
            col += 1
        
        # BREAK column header
        cell = ws.cell(row=current_row, column=col, value=break_time if d == 0 else "BREAK")
        cell.font = break_header_font
        cell.border = thick_border
        cell.alignment = center_align
        col += 1
        
        # P6-P8
        for i, p in enumerate(range(6, periods_per_day + 1)):
            header_value = period_times_after_break[i] if d == 0 else f"P{p}"
            cell = ws.cell(row=current_row, column=col, value=header_value)
            cell.font = header_font
            cell.border = thick_border
            cell.alignment = center_align
            col += 1
        
        ws.row_dimensions[current_row].height = period_header_height
        current_row += 1
        
        # Data rows
        for class_name, arm in class_arms:
            short_code = get_short_code(class_name, arm)
            col = 1
            
            # Class code
            cell = ws.cell(row=current_row, column=col, value=short_code)
            cell.font = class_font
            cell.border = thick_border
            cell.alignment = center_align
            col += 1
            
            arm_results = [r for r in results if r.class_name == class_name and r.arm_name == arm and r.day_of_week == d]
            
            # P1-P5
            for p in range(1, 6):
                slot_result = next((r for r in arm_results if r.period_number == p), None)
                value = ""
                if slot_result and slot_result.subject_id:
                    subj = subject_lookup.get(slot_result.subject_id)
                    if subj:
                        subj_name = subj.name
                        value = abbrev_map.get(subj_name, subj_name[:5])
                
                cell = ws.cell(row=current_row, column=col, value=value)
                cell.font = cell_font
                cell.border = thin_border
                cell.alignment = center_align
                col += 1
            
            # BREAK column - empty with border
            cell = ws.cell(row=current_row, column=col, value="")
            cell.border = thin_border
            cell.alignment = center_align
            col += 1
            
            # P6-P8
            for p in range(6, periods_per_day + 1):
                slot_result = next((r for r in arm_results if r.period_number == p), None)
                value = ""
                if slot_result and slot_result.subject_id:
                    subj = subject_lookup.get(slot_result.subject_id)
                    if subj:
                        subj_name = subj.name
                        value = abbrev_map.get(subj_name, subj_name[:5])
                
                cell = ws.cell(row=current_row, column=col, value=value)
                cell.font = cell_font
                cell.border = thin_border
                cell.alignment = center_align
                col += 1
            
            ws.row_dimensions[current_row].height = data_row_height
            current_row += 1
        
        # Column widths - fit within safe printable area
        ws.column_dimensions['A'].width = 8  # Class
        col_width = 26 if periods_per_day <= 8 else 22
        for col in range(2, total_cols + 1):
            if col == 7:  # Break column
                ws.column_dimensions[get_column_letter(col)].width = 8
            else:
                ws.column_dimensions[get_column_letter(col)].width = col_width
    
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    return Response(output.getvalue(),
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={'Content-Disposition': f'attachment; filename=timetables_by_day_{batch_id}.xlsx'}
    )


@generator_bp.route('/results/<batch_id>/export_by_day_pdf')
@login_required
def export_results_by_day_pdf(batch_id):
    """Export timetable as PDF - Each day fits EXACTLY on one A4 landscape page - NO COLORS"""
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.units import mm
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, PageBreak
    
    results = GenTimetableResult.query.filter_by(batch_id=batch_id).all()
    if not results:
        flash('No results.', 'error')
        return redirect(url_for('generator.results_list'))
    
    rules = {r.rule_type: r.value for r in GenTimetableRule.query.filter_by(is_active=True).all()}
    periods_per_day = int(rules.get('periods_per_day', 8))
    
    # Get school info
    school_name = GenSettings.get('school_name', 'School')
    school_address = GenSettings.get('school_address', '')
    
    days = ['Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday']
    
    # Period times with break tracking
    period_times_before = []  # P1-P5
    period_times_after = []   # P6-P8
    break_time = ""
    
    start_hour, start_min = 8, 20
    for p in range(1, periods_per_day + 1):
        start_total = start_hour * 60 + start_min
        end_total = start_total + 40
        start_h, start_m = start_total // 60, start_total % 60
        end_h, end_m = end_total // 60, end_total % 60
        
        time_str = f"P{p}\n{start_h}:{start_m:02d}\n{end_h}:{end_m:02d}"
        if p <= 5:
            period_times_before.append(time_str)
        else:
            period_times_after.append(time_str)
        
        start_hour, start_min = end_h, end_m
        if p == 5:
            break_start = f"{end_h}:{end_m:02d}"
            start_total = start_hour * 60 + start_min + 30
            start_hour, start_min = start_total // 60, start_total % 60
            break_end = f"{start_hour}:{start_min:02d}"
            break_time = f"BREAK\n{break_start}\n-\n{break_end}"
    
    class_arms = sorted(set((r.class_name, r.arm_name) for r in results))
    num_data_rows = len(class_arms)
    
    def get_short_code(class_name, arm):
        class_num = class_name.replace('SSS', '')
        arm_letter = arm[0].upper()
        return f"{class_num}{arm_letter}"
    
    abbrev_map = {
        'Mathematics': 'Maths', 'English Language': 'Eng', 'Physics': 'Phy',
        'Chemistry': 'Chem', 'Biology': 'Bio', 'Economics': 'Econs',
        'Government': 'Govt', 'Literature in English': 'Lit', 'Agricultural Science': 'Agric',
        'Christian Religious Studies': 'CRS', 'Civic Education': 'Civic',
        'Computer Studies': 'Comp', 'Commerce': 'Comm', 'Geography': 'Geo',
        'Further Mathematics': 'F/Mth', 'Livestock Farming': 'Livst',
        'History': 'Hist', 'Phonics': 'Phon',
    }
    
    output = BytesIO()
    
    # A4 landscape: 297mm x 210mm - safe margins for printers
    margin = 10*mm
    page_w, page_h = landscape(A4)
    usable_width = page_w - 2*margin
    usable_height = page_h - 2*margin
    
    doc = SimpleDocTemplate(
        output,
        pagesize=landscape(A4),
        leftMargin=margin,
        rightMargin=margin,
        topMargin=margin,
        bottomMargin=margin
    )
    
    elements = []
    
    # Total columns: Class + P1-P5 + BREAK + P6-P8
    num_periods_before = 5
    num_periods_after = periods_per_day - 5
    total_cols = 1 + num_periods_before + 1 + num_periods_after
    
    # Column widths - fill entire width
    first_col_width = 12*mm
    break_col_width = 10*mm
    remaining_width = usable_width - first_col_width - break_col_width
    period_col_width = remaining_width / periods_per_day
    
    col_widths = [first_col_width]
    col_widths += [period_col_width] * num_periods_before
    col_widths += [break_col_width]
    col_widths += [period_col_width] * num_periods_after
    
    for d, day_name in enumerate(days):
        table_data = []
        row_heights = []
        
        # Calculate row heights to fit EXACTLY on one page
        if d == 0:  # Monday - include school name and address
            school_header_height = 8*mm
            address_header_height = 4*mm if school_address else 0
            day_header_height = 10*mm
            period_header_height = 12*mm
            fixed_height = school_header_height + address_header_height + day_header_height + period_header_height
        else:
            day_header_height = 10*mm
            period_header_height = 10*mm
            fixed_height = day_header_height + period_header_height
        
        # Data rows get remaining height divided equally
        available_for_data = usable_height - fixed_height - 2*mm
        data_row_height = available_for_data / num_data_rows
        
        # Build table data
        if d == 0:
            # School name row
            school_row = [school_name.upper()] + [''] * (total_cols - 1)
            table_data.append(school_row)
            row_heights.append(school_header_height)
            
            if school_address:
                address_row = [school_address] + [''] * (total_cols - 1)
                table_data.append(address_row)
                row_heights.append(address_header_height)
        
        # Day header row
        day_row = [day_name.upper()] + [''] * (total_cols - 1)
        table_data.append(day_row)
        row_heights.append(day_header_height)
        
        # Period header row
        if d == 0:  # Monday - show times
            header_row = ['Class'] + period_times_before + [break_time] + period_times_after
        else:
            header_row = ['Class'] + [f'P{p}' for p in range(1, 6)] + ['BREAK'] + [f'P{p}' for p in range(6, periods_per_day + 1)]
        table_data.append(header_row)
        row_heights.append(period_header_height)
        
        # Data rows
        for class_name, arm in class_arms:
            short_code = get_short_code(class_name, arm)
            row = [short_code]
            
            arm_results = [r for r in results if r.class_name == class_name and r.arm_name == arm and r.day_of_week == d]
            
            # P1-P5
            for p in range(1, 6):
                slot_result = next((r for r in arm_results if r.period_number == p), None)
                if slot_result and slot_result.subject:
                    row.append(abbrev_map.get(slot_result.subject.name, slot_result.subject.name[:5]))
                else:
                    row.append('')
            
            # BREAK column - empty
            row.append('')
            
            # P6-P8
            for p in range(6, periods_per_day + 1):
                slot_result = next((r for r in arm_results if r.period_number == p), None)
                if slot_result and slot_result.subject:
                    row.append(abbrev_map.get(slot_result.subject.name, slot_result.subject.name[:5]))
                else:
                    row.append('')
            
            table_data.append(row)
            row_heights.append(data_row_height)
        
        # Create table with exact dimensions
        table = Table(table_data, colWidths=col_widths, rowHeights=row_heights)
        
        # Break column index
        break_col = 6
        
        # Calculate row indices
        if d == 0:
            if school_address:
                day_row_idx = 2
                header_row_idx = 3
                data_start_row = 4
            else:
                day_row_idx = 1
                header_row_idx = 2
                data_start_row = 3
        else:
            day_row_idx = 0
            header_row_idx = 1
            data_start_row = 2
        
        # NO COLORS - just black text on white, with borders
        style_commands = [
            # Day header row
            ('SPAN', (0, day_row_idx), (-1, day_row_idx)),
            ('TEXTCOLOR', (0, day_row_idx), (-1, day_row_idx), colors.black),
            ('FONTNAME', (0, day_row_idx), (-1, day_row_idx), 'Helvetica-Bold'),
            ('FONTSIZE', (0, day_row_idx), (-1, day_row_idx), 22),
            ('ALIGN', (0, day_row_idx), (-1, day_row_idx), 'CENTER'),
            ('VALIGN', (0, day_row_idx), (-1, day_row_idx), 'MIDDLE'),
            
            # Period header row
            ('FONTNAME', (0, header_row_idx), (-1, header_row_idx), 'Helvetica-Bold'),
            ('FONTSIZE', (0, header_row_idx), (-1, header_row_idx), 7 if d == 0 else 12),
            ('ALIGN', (0, header_row_idx), (-1, header_row_idx), 'CENTER'),
            ('VALIGN', (0, header_row_idx), (-1, header_row_idx), 'MIDDLE'),
            
            # Break column header
            ('FONTSIZE', (break_col, header_row_idx), (break_col, header_row_idx), 6),
            
            # Class codes column
            ('FONTNAME', (0, data_start_row), (0, -1), 'Helvetica-Bold'),
            ('FONTSIZE', (0, data_start_row), (0, -1), 14),
            
            # Subject cells - LARGE
            ('FONTNAME', (1, data_start_row), (-1, -1), 'Helvetica-Bold'),
            ('FONTSIZE', (1, data_start_row), (-1, -1), 18),
            
            # All cells alignment
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            
            # Borders only - no background colors
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('BOX', (0, 0), (-1, -1), 1.5, colors.black),
        ]
        
        # Add school name styling for Monday
        if d == 0:
            style_commands.extend([
                ('SPAN', (0, 0), (-1, 0)),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 16),
                ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
                ('VALIGN', (0, 0), (-1, 0), 'MIDDLE'),
            ])
            if school_address:
                style_commands.extend([
                    ('SPAN', (0, 1), (-1, 1)),
                    ('FONTNAME', (0, 1), (-1, 1), 'Helvetica'),
                    ('FONTSIZE', (0, 1), (-1, 1), 8),
                    ('TEXTCOLOR', (0, 1), (-1, 1), colors.black),
                    ('ALIGN', (0, 1), (-1, 1), 'CENTER'),
                    ('VALIGN', (0, 1), (-1, 1), 'MIDDLE'),
                ])
        
        table.setStyle(TableStyle(style_commands))
        elements.append(table)
        
        if d < len(days) - 1:
            elements.append(PageBreak())
    
    doc.build(elements)
    output.seek(0)
    
    return Response(output.getvalue(),
        mimetype='application/pdf',
        headers={'Content-Disposition': f'attachment; filename=timetables_by_day_{batch_id}.pdf'}
    )


# ============================================================================
# IMAGE EXPORT
# ============================================================================

@generator_bp.route('/results/<batch_id>/export_image')
@login_required
def export_image(batch_id):
    """Export timetable as PNG image (Ultra HD - 8x scale)"""
    from routes.generator_image import generate_timetable_image, image_to_response
    
    quality = request.args.get('quality', 'ultra')  # 'hd' or 'ultra'
    img = generate_timetable_image(batch_id, quality=quality)
    if not img:
        flash('No results found.', 'error')
        return redirect(url_for('generator.results_list'))
    
    quality_suffix = '_hd' if quality == 'hd' else '_ultrahd'
    return image_to_response(img, f'timetable_{batch_id}{quality_suffix}.png')


@generator_bp.route('/results/<batch_id>/export_image_hd')
@login_required
def export_image_hd(batch_id):
    """Export timetable as PNG image (HD - 4x scale)"""
    from routes.generator_image import generate_timetable_image, image_to_response
    
    img = generate_timetable_image(batch_id, quality='hd')
    if not img:
        flash('No results found.', 'error')
        return redirect(url_for('generator.results_list'))
    
    return image_to_response(img, f'timetable_{batch_id}_hd.png')


@generator_bp.route('/results/<batch_id>/export_image_ultra')
@login_required
def export_image_ultra(batch_id):
    """Export timetable as PNG image (Ultra HD - 8x scale)"""
    from routes.generator_image import generate_timetable_image, image_to_response
    
    img = generate_timetable_image(batch_id, quality='ultra')
    if not img:
        flash('No results found.', 'error')
        return redirect(url_for('generator.results_list'))
    
    return image_to_response(img, f'timetable_{batch_id}_ultrahd.png')


@generator_bp.route('/results/<batch_id>/teacher/<int:teacher_id>/image')
@login_required
def export_teacher_image(batch_id, teacher_id):
    """Export individual teacher timetable as PNG image"""
    from routes.generator_image import generate_teacher_timetable_image, image_to_response
    
    img = generate_teacher_timetable_image(batch_id, teacher_id)
    if not img:
        flash('No results found for this teacher.', 'error')
        return redirect(url_for('generator.teacher_timetable', batch_id=batch_id))
    
    teacher = db.session.get(GenTeacher, teacher_id)
    filename = f'timetable_{teacher.name.replace(" ", "_")}_{batch_id}.png'
    
    return image_to_response(img, filename)


@generator_bp.route('/results/<batch_id>/teacher/<int:teacher_id>/print')
@login_required
def print_teacher_timetable(batch_id, teacher_id):
    """Printable view for individual teacher timetable"""
    teacher = db.get_or_404(GenTeacher, teacher_id)
    results = GenTimetableResult.query.filter_by(batch_id=batch_id, teacher_id=teacher_id).all()
    
    if not results:
        flash('No timetable found for this teacher.', 'error')
        return redirect(url_for('generator.teacher_timetable', batch_id=batch_id))
    
    rules = {r.rule_type: r.value for r in GenTimetableRule.query.filter_by(is_active=True).all()}
    periods_per_day = int(rules.get('periods_per_day', 8))
    
    # Build timetable grid
    timetable = {d: {p: None for p in range(1, periods_per_day + 1)} for d in range(5)}
    for r in results:
        timetable[r.day_of_week][r.period_number] = {
            'subject': r.subject,
            'class_name': r.class_name,
            'arm_name': r.arm_name
        }
    
    return render_template('generator/print_teacher_timetable.html',
        teacher=teacher,
        timetable=timetable,
        batch_id=batch_id,
        periods=range(1, periods_per_day + 1),
        days=DAYS_OF_WEEK
    )


# ============================================================================
# SUBJECT CLASH RULES
# ============================================================================

@generator_bp.route('/clash-rules')
@login_required
def clash_rules_list():
    """List all subject clash rules"""
    from models import GenSubjectClashRule, GenCombinedClassRule
    
    clash_rules = GenSubjectClashRule.query.order_by(GenSubjectClashRule.id).all()
    combined_rules = GenCombinedClassRule.query.order_by(GenCombinedClassRule.id).all()
    
    return render_template('generator/clash_rules.html',
        clash_rules=clash_rules,
        combined_rules=combined_rules
    )


@generator_bp.route('/clash-rules/add', methods=['GET', 'POST'])
@login_required
def add_clash_rule():
    """Add a new subject clash rule"""
    from models import GenSubjectClashRule, GenClassConfig
    
    if request.method == 'POST':
        try:
            rule = GenSubjectClashRule(
                name=request.form.get('name', '').strip(),
                description=request.form.get('description', '').strip() or None,
                source_subject_id=int(request.form.get('source_subject_id')),
                source_class_name=request.form.get('source_class_name'),
                source_arm_name=request.form.get('source_arm_name') or None,
                target_subject_id=int(request.form.get('target_subject_id')),
                target_class_name=request.form.get('target_class_name') or None,
                target_arm_name=request.form.get('target_arm_name') or None,
                is_active=True
            )
            db.session.add(rule)
            db.session.commit()
            flash('Subject clash rule added successfully', 'success')
            return redirect(url_for('generator.clash_rules_list'))
        except Exception as e:
            db.session.rollback()
            flash(f'Error adding rule: {str(e)}', 'error')
    
    level = get_current_level()
    subjects = GenSubject.query.filter_by(is_active=True, school_level=level).order_by(GenSubject.name).all()
    classes = GenClassConfig.query.filter_by(is_active=True, school_level=level).order_by(GenClassConfig.class_name).all()
    
    return render_template('generator/add_clash_rule.html',
        subjects=subjects,
        classes=classes
    )


@generator_bp.route('/clash-rules/<int:rule_id>/toggle', methods=['POST'])
@login_required
def toggle_clash_rule(rule_id):
    """Toggle a clash rule active/inactive"""
    from models import GenSubjectClashRule
    
    rule = db.get_or_404(GenSubjectClashRule, rule_id)
    rule.is_active = not rule.is_active
    db.session.commit()
    
    status = 'activated' if rule.is_active else 'deactivated'
    flash(f'Rule {status}', 'success')
    return redirect(url_for('generator.clash_rules_list'))


@generator_bp.route('/clash-rules/<int:rule_id>/delete', methods=['POST'])
@login_required
def delete_clash_rule(rule_id):
    """Delete a clash rule"""
    from models import GenSubjectClashRule
    
    rule = db.get_or_404(GenSubjectClashRule, rule_id)
    db.session.delete(rule)
    db.session.commit()
    
    flash('Rule deleted', 'success')
    return redirect(url_for('generator.clash_rules_list'))


@generator_bp.route('/combined-rules/add', methods=['GET', 'POST'])
@login_required
def add_combined_rule():
    """Add a new combined class rule"""
    from models import GenCombinedClassRule, GenClassConfig
    
    if request.method == 'POST':
        try:
            rule = GenCombinedClassRule(
                name=request.form.get('name', '').strip(),
                description=request.form.get('description', '').strip() or None,
                shadow_subject_id=int(request.form.get('shadow_subject_id')),
                shadow_class_name=request.form.get('shadow_class_name'),
                shadow_arm_name=request.form.get('shadow_arm_name') or None,
                teacher_subject_id=int(request.form.get('teacher_subject_id')),
                teacher_class_name=request.form.get('teacher_class_name'),
                teacher_arm_name=request.form.get('teacher_arm_name') or None,
                is_active=True
            )
            db.session.add(rule)
            db.session.commit()
            flash('Combined class rule added successfully', 'success')
            return redirect(url_for('generator.clash_rules_list'))
        except Exception as e:
            db.session.rollback()
            flash(f'Error adding rule: {str(e)}', 'error')
    
    level = get_current_level()
    subjects = GenSubject.query.filter_by(is_active=True, school_level=level).order_by(GenSubject.name).all()
    classes = GenClassConfig.query.filter_by(is_active=True, school_level=level).order_by(GenClassConfig.class_name).all()
    
    return render_template('generator/add_combined_rule.html',
        subjects=subjects,
        classes=classes
    )


@generator_bp.route('/combined-rules/<int:rule_id>/toggle', methods=['POST'])
@login_required
def toggle_combined_rule(rule_id):
    """Toggle a combined rule active/inactive"""
    from models import GenCombinedClassRule
    
    rule = db.get_or_404(GenCombinedClassRule, rule_id)
    rule.is_active = not rule.is_active
    db.session.commit()
    
    status = 'activated' if rule.is_active else 'deactivated'
    flash(f'Rule {status}', 'success')
    return redirect(url_for('generator.clash_rules_list'))


@generator_bp.route('/combined-rules/<int:rule_id>/delete', methods=['POST'])
@login_required
def delete_combined_rule(rule_id):
    """Delete a combined rule"""
    from models import GenCombinedClassRule
    
    rule = db.get_or_404(GenCombinedClassRule, rule_id)
    db.session.delete(rule)
    db.session.commit()
    
    flash('Rule deleted', 'success')
    return redirect(url_for('generator.clash_rules_list'))


# ============================================================================
# JSS QUICK SETUP
# ============================================================================

@generator_bp.route('/setup/jss', methods=['GET', 'POST'])
@login_required
def setup_jss():
    """Quick setup for JSS1, JSS2, JSS3 classes"""
    if request.method == 'POST':
        try:
            arm_names = request.form.get('arm_names', '').strip()
            num_arms = len([a.strip() for a in arm_names.split(',') if a.strip()]) if arm_names else 1
            
            created = []
            for class_name in ['JSS1', 'JSS2', 'JSS3']:
                # Check if already exists for JSS level
                existing = GenClassConfig.query.filter_by(class_name=class_name, school_level='jss', is_active=True).first()
                if existing:
                    # Update existing
                    existing.num_arms = num_arms
                    existing.arm_names = arm_names or None
                    existing.has_streams = False  # JSS doesn't have streams
                else:
                    # Create new
                    class_config = GenClassConfig(
                        class_name=class_name,
                        school_level='jss',
                        num_arms=num_arms,
                        arm_names=arm_names or None,
                        has_streams=False
                    )
                    db.session.add(class_config)
                created.append(class_name)
            
            db.session.commit()
            
            # Set session to JSS level
            from flask import session
            session['generator_level'] = 'jss'
            
            flash(f'JSS classes configured: {", ".join(created)}', 'success')
            return redirect(url_for('generator.classes_config'))
            
        except Exception as e:
            db.session.rollback()
            flash(f'Error: {str(e)}', 'error')
    
    # Get SSS arm names as default suggestion
    sss_class = GenClassConfig.query.filter(
        GenClassConfig.class_name.like('SSS%'),
        GenClassConfig.school_level == 'sss',
        GenClassConfig.is_active == True
    ).first()
    
    suggested_arms = sss_class.arm_names if sss_class else 'Iris,Rose,Lily,Tulip'
    
    return render_template('generator/setup_jss.html', suggested_arms=suggested_arms)


@generator_bp.route('/setup/copy-subjects', methods=['POST'])
@login_required
def copy_subject_config():
    """Copy subject configuration from one class level to another"""
    try:
        source_level = request.form.get('source_level')  # 'SSS' or 'JSS'
        target_level = request.form.get('target_level')  # 'JSS' or 'SSS'
        
        if not source_level or not target_level:
            flash('Select source and target levels.', 'error')
            return redirect(url_for('generator.classes_config'))
        
        # Get source classes
        source_classes = GenClassConfig.query.filter(
            GenClassConfig.class_name.like(f'{source_level}%'),
            GenClassConfig.is_active == True
        ).all()
        
        # Get target classes
        target_classes = GenClassConfig.query.filter(
            GenClassConfig.class_name.like(f'{target_level}%'),
            GenClassConfig.is_active == True
        ).all()
        
        if not source_classes or not target_classes:
            flash('No classes found for source or target level.', 'error')
            return redirect(url_for('generator.classes_config'))
        
        # Copy subject configs from first source class to all target classes
        source_id = source_classes[0].id
        source_configs = GenClassSubjectConfig.query.filter_by(
            class_config_id=source_id, is_active=True
        ).all()
        
        copied = 0
        for target_class in target_classes:
            for src_cfg in source_configs:
                # Check if exists
                existing = GenClassSubjectConfig.query.filter_by(
                    class_config_id=target_class.id,
                    subject_id=src_cfg.subject_id
                ).first()
                
                if existing:
                    existing.periods_per_week = src_cfg.periods_per_week
                    existing.needs_double_period = src_cfg.needs_double_period
                    existing.double_period_count = src_cfg.double_period_count
                    existing.is_enabled = src_cfg.is_enabled
                    existing.is_active = True
                else:
                    new_cfg = GenClassSubjectConfig(
                        class_config_id=target_class.id,
                        subject_id=src_cfg.subject_id,
                        periods_per_week=src_cfg.periods_per_week,
                        needs_double_period=src_cfg.needs_double_period,
                        double_period_count=src_cfg.double_period_count,
                        is_enabled=src_cfg.is_enabled,
                        is_active=True
                    )
                    db.session.add(new_cfg)
                copied += 1
        
        db.session.commit()
        flash(f'Copied {copied} subject configurations to {target_level} classes.', 'success')
        
    except Exception as e:
        db.session.rollback()
        flash(f'Error: {str(e)}', 'error')
    
    return redirect(url_for('generator.classes_config'))
