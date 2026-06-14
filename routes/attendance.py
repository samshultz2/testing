"""
Attendance management routes
"""
from flask import Blueprint, render_template, request, redirect, url_for, flash, jsonify, Response
from utils.helpers import get_active_term
from utils.web_exports import xlsx_response
from datetime import datetime, date, timedelta
from models import (
    db, Student, StudentEnrollment, ClassArmAssignment, 
    Week, Holiday, Attendance, Term, AcademicSession
)
from utils.access_control import (
    login_required, can_access_class, can_mark_attendance,
    filter_classes_for_user
)
from utils.branch_scope import require_branch_access
from utils.helpers import is_school_day
from utils.calculations import (
    get_daily_attendance_summary, get_weekly_attendance_summary,
    get_termly_attendance_summary, mark_attendance_bulk, mark_all_present
)
from utils.excel_utils import export_attendance_to_excel

attendance_bp = Blueprint('attendance', __name__, url_prefix='/attendance')


@attendance_bp.route('/')
@login_required
def index():
    """Attendance main page"""
    # Get active term
    active_term = get_active_term()
    
    # Get all sessions and terms for selection
    sessions = AcademicSession.query.order_by(AcademicSession.name.desc()).all()
    
    return render_template('attendance/index.html',
        active_term=active_term,
        sessions=sessions
    )


@attendance_bp.route('/mark')
@login_required
def mark_attendance_page():
    """Page for marking daily attendance"""
    # Get parameters
    term_id = request.args.get('term_id', type=int)
    assignment_id = request.args.get('assignment_id', type=int)
    target_date = request.args.get('date')
    session_type = request.args.get('session', 'morning')
    
    # Check attendance permission
    if not can_mark_attendance():
        flash('You do not have permission to mark attendance.', 'error')
        return redirect(url_for('main.dashboard'))
    
    # Check class access if class is selected
    if assignment_id and not can_access_class(assignment_id):
        flash('You do not have access to this class.', 'error')
        return redirect(url_for('attendance.mark_attendance_page'))
    
    # Get active term if not specified
    if not term_id:
        active_term = get_active_term()
        if active_term:
            term_id = active_term.id
    
    # Get data for dropdowns
    terms = Term.query.join(AcademicSession).order_by(
        AcademicSession.name.desc(),
        Term.term_number
    ).all()
    
    assignments = []
    weeks = []
    holidays = []
    selected_term = None
    selected_assignment = None
    enrollments = []
    existing_attendance = {}
    current_week = None
    
    if term_id:
        selected_term = db.session.get(Term, term_id)
        all_assignments = ClassArmAssignment.query.filter_by(term_id=term_id).all()
        # Filter classes for teachers
        assignments = filter_classes_for_user(all_assignments, form_only=True)
        weeks = Week.query.filter_by(term_id=term_id).order_by(Week.week_number).all()
        holidays = Holiday.query.filter_by(term_id=term_id).all()
    
    if assignment_id:
        selected_assignment = db.session.get(ClassArmAssignment, assignment_id)
        enrollments = StudentEnrollment.query.filter_by(
            class_arm_assignment_id=assignment_id,
            is_active=True
        ).join(Student).order_by(Student.surname).all()
    
    # Parse date
    if target_date:
        try:
            target_date = datetime.strptime(target_date, '%Y-%m-%d').date()
        except Exception:
            target_date = date.today()
    else:
        target_date = date.today()
    
    # Find which week this date belongs to
    if weeks:
        for week in weeks:
            if week.start_date <= target_date <= week.end_date:
                current_week = week
                break
    
    # Check if it's a school day, and if not, why (so the page can be specific).
    holiday_dates = set(h.date for h in holidays)
    is_valid_school_day = is_school_day(target_date, holidays)
    holiday_for_date = next((h for h in holidays if h.date == target_date), None)
    is_weekend = target_date.weekday() >= 5

    # Get existing attendance for this date
    if enrollments and target_date:
        for enrollment in enrollments:
            att = Attendance.query.filter_by(
                enrollment_id=enrollment.id,
                date=target_date
            ).first()
            if att:
                existing_attendance[enrollment.id] = {
                    'morning': att.morning_present,
                    'afternoon': att.afternoon_present
                }
    
    return render_template('attendance/mark.html',
        terms=terms,
        selected_term=selected_term,
        assignments=assignments,
        selected_assignment=selected_assignment,
        enrollments=enrollments,
        target_date=target_date,
        session_type=session_type,
        existing_attendance=existing_attendance,
        current_week=current_week,
        is_valid_school_day=is_valid_school_day,
        holiday_for_date=holiday_for_date,
        is_weekend=is_weekend,
        holidays=holidays
    )


@attendance_bp.route('/mark/save', methods=['POST'])
@login_required
def save_attendance():
    """Save attendance for a session"""
    try:
        assignment_id = request.form.get('assignment_id', type=int)
        target_date = request.form.get('date')
        session_type = request.form.get('session_type')
        week_id = request.form.get('week_id', type=int)
        
        # Check access permission
        if not can_mark_attendance(assignment_id):
            flash('You do not have permission to mark attendance for this class.', 'error')
            return redirect(url_for('attendance.mark_attendance_page'))
        
        # Parse date
        target_date = datetime.strptime(target_date, '%Y-%m-%d').date()
        
        # Get all enrollment IDs for this class
        enrollments = StudentEnrollment.query.filter_by(
            class_arm_assignment_id=assignment_id,
            is_active=True
        ).all()
        all_enrollment_ids = [e.id for e in enrollments]
        
        # Get IDs of students marked present (checked checkboxes)
        present_ids = [int(x) for x in request.form.getlist('present[]')]
        
        # Mark attendance
        # Check if auto-copy to afternoon is enabled (default True for morning)
        auto_copy = request.form.get('auto_copy_afternoon', 'on') == 'on'
        
        count = mark_attendance_bulk(
            all_enrollment_ids,
            target_date,
            week_id,
            session_type,
            present_ids,
            marked_by='Admin',
            auto_copy_to_afternoon=auto_copy if session_type == 'morning' else False
        )
        
        if session_type == 'morning' and auto_copy:
            flash(f'Attendance saved for {count} students! (Morning & Afternoon)', 'success')
        else:
            flash(f'Attendance saved for {count} students!', 'success')
        
    except Exception as e:
        db.session.rollback()
        flash(f'Error saving attendance: {str(e)}', 'error')
    
    return redirect(url_for('attendance.mark_attendance_page',
        term_id=request.form.get('term_id'),
        assignment_id=assignment_id,
        date=request.form.get('date'),
        session=session_type
    ))


def _week_school_days(week, holidays):
    """The school weekdays (Mon–Fri, minus holidays) within a week."""
    holiday_dates = {h.date for h in holidays}
    days, d = [], week.start_date
    while d <= week.end_date:
        if d.weekday() < 5 and d not in holiday_dates:
            days.append(d)
        d += timedelta(days=1)
    return days


@attendance_bp.route('/week')
@login_required
def week_grid():
    """Mark a whole week at once: a students × school-days grid (backlog-friendly)."""
    if not can_mark_attendance():
        flash('You do not have permission to mark attendance.', 'error')
        return redirect(url_for('main.dashboard'))

    term_id = request.args.get('term_id', type=int)
    assignment_id = request.args.get('assignment_id', type=int)
    week_id = request.args.get('week_id', type=int)
    split = request.args.get('sessions') == 'split'

    if not term_id:
        active = get_active_term()
        term_id = active.id if active else None

    terms = Term.query.join(AcademicSession).order_by(
        AcademicSession.name.desc(), Term.term_number).all()
    selected_term = db.session.get(Term, term_id) if term_id else None
    assignments, weeks, holidays = [], [], []
    if term_id:
        assignments = filter_classes_for_user(
            ClassArmAssignment.query.filter_by(term_id=term_id).all(), form_only=True)
        weeks = Week.query.filter_by(term_id=term_id).order_by(Week.week_number).all()
        holidays = Holiday.query.filter_by(term_id=term_id).all()

    if assignment_id and not can_access_class(assignment_id):
        flash('You do not have access to this class.', 'error')
        return redirect(url_for('attendance.week_grid'))

    selected_assignment = db.session.get(ClassArmAssignment, assignment_id) if assignment_id else None
    selected_week = db.session.get(Week, week_id) if week_id else (weeks[0] if weeks else None)

    enrollments, school_days, existing = [], [], {}
    if selected_assignment and selected_week:
        enrollments = (StudentEnrollment.query
                       .filter_by(class_arm_assignment_id=assignment_id, is_active=True)
                       .join(Student).order_by(Student.surname, Student.first_name).all())
        school_days = _week_school_days(selected_week, holidays)
        if enrollments and school_days:
            recs = Attendance.query.filter(
                Attendance.enrollment_id.in_([e.id for e in enrollments]),
                Attendance.date.in_(school_days)).all()
            # existing[enrollment_id][iso_date] = (morning_present, afternoon_present)
            for r in recs:
                existing.setdefault(r.enrollment_id, {})[r.date.isoformat()] = (
                    bool(r.morning_present), bool(r.afternoon_present))

    return render_template('attendance/week.html',
        terms=terms, selected_term=selected_term, assignments=assignments,
        selected_assignment=selected_assignment, weeks=weeks, split=split,
        selected_week=selected_week, enrollments=enrollments,
        school_days=school_days, existing=existing)


@attendance_bp.route('/week/save', methods=['POST'])
@login_required
def week_save():
    """Batch-save a whole week's attendance from the grid (one transaction)."""
    assignment_id = request.form.get('assignment_id', type=int)
    week_id = request.form.get('week_id', type=int)
    term_id = request.form.get('term_id', type=int)
    if not can_mark_attendance(assignment_id):
        flash('You do not have permission to mark attendance for this class.', 'error')
        return redirect(url_for('attendance.week_grid'))

    week = db.session.get(Week, week_id)
    if not week:
        flash('Select a week first.', 'error')
        return redirect(url_for('attendance.week_grid', term_id=term_id, assignment_id=assignment_id))

    holidays = Holiday.query.filter_by(term_id=week.term_id).all()
    school_days = _week_school_days(week, holidays)
    enrollments = StudentEnrollment.query.filter_by(
        class_arm_assignment_id=assignment_id, is_active=True).all()
    day_isos = [d.isoformat() for d in school_days]

    # Existing records keyed by (enrollment_id, date) so we upsert.
    existing = {(r.enrollment_id, r.date): r for r in Attendance.query.filter(
        Attendance.enrollment_id.in_([e.id for e in enrollments]),
        Attendance.date.in_(school_days)).all()}

    marked_by = request.form.get('marked_by') or 'Admin'
    split = request.form.get('sessions') == 'split'   # AM/PM ticks vs whole-day
    saved = 0
    try:
        for e in enrollments:
            for d, iso in zip(school_days, day_isos):
                if split:
                    morning = request.form.get(f'am_{e.id}_{iso}') == 'on'
                    afternoon = request.form.get(f'pm_{e.id}_{iso}') == 'on'
                else:
                    morning = afternoon = request.form.get(f'p_{e.id}_{iso}') == 'on'
                rec = existing.get((e.id, d))
                if rec is None:
                    rec = Attendance(enrollment_id=e.id, week_id=week.id, date=d)
                    db.session.add(rec)
                rec.week_id = week.id
                rec.morning_present = morning
                rec.afternoon_present = afternoon
                rec.marked_by = marked_by
                saved += 1
        db.session.commit()
        flash(f'Saved attendance for {len(enrollments)} student(s) across '
              f'{len(school_days)} day(s) of {week.week_number and "Week %d" % week.week_number}.',
              'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Error saving attendance: {str(e)}', 'error')

    return redirect(url_for('attendance.week_grid',
        term_id=term_id, assignment_id=assignment_id, week_id=week_id))


@attendance_bp.route('/mark/all-present', methods=['POST'])
@login_required
def mark_all_present_route():
    """Mark all students as present for a session"""
    try:
        assignment_id = request.form.get('assignment_id', type=int)
        target_date = request.form.get('date')
        session_type = request.form.get('session_type')
        week_id = request.form.get('week_id', type=int)
        
        target_date = datetime.strptime(target_date, '%Y-%m-%d').date()
        
        enrollments = StudentEnrollment.query.filter_by(
            class_arm_assignment_id=assignment_id,
            is_active=True
        ).all()
        enrollment_ids = [e.id for e in enrollments]
        
        count = mark_all_present(enrollment_ids, target_date, week_id, session_type)
        
        flash(f'All {count} students marked as present!', 'success')
        
    except Exception as e:
        db.session.rollback()
        flash(f'Error: {str(e)}', 'error')
    
    return redirect(url_for('attendance.mark_attendance_page',
        term_id=request.form.get('term_id'),
        assignment_id=assignment_id,
        date=request.form.get('date'),
        session=session_type
    ))


@attendance_bp.route('/daily')
@login_required
def daily_summary():
    """View daily attendance summary"""
    assignment_id = request.args.get('assignment_id', type=int)
    target_date = request.args.get('date')
    
    if target_date:
        target_date = datetime.strptime(target_date, '%Y-%m-%d').date()
    else:
        target_date = date.today()
    
    # Get assignments for dropdown
    active_term = get_active_term()
    assignments = []
    if active_term:
        assignments = filter_classes_for_user(
            ClassArmAssignment.query.filter_by(term_id=active_term.id).all(),
            form_only=True)

    summary = None
    selected_assignment = None
    
    if assignment_id:
        selected_assignment = db.session.get(ClassArmAssignment, assignment_id)
        summary = get_daily_attendance_summary(assignment_id, target_date)
    
    return render_template('attendance/daily.html',
        assignments=assignments,
        selected_assignment=selected_assignment,
        target_date=target_date,
        summary=summary
    )


@attendance_bp.route('/copy-previous', methods=['POST'])
@login_required
def copy_previous_attendance():
    """Copy attendance from previous school day"""
    try:
        assignment_id = request.form.get('assignment_id', type=int)
        target_date_str = request.form.get('date')
        term_id = request.form.get('term_id', type=int)
        
        target_date = datetime.strptime(target_date_str, '%Y-%m-%d').date()
        
        # Find previous school day
        holidays = Holiday.query.filter_by(term_id=term_id).all()
        holiday_dates = set(h.date for h in holidays)
        
        prev_date = target_date - timedelta(days=1)
        attempts = 0
        while attempts < 10:
            if prev_date.weekday() < 5 and prev_date not in holiday_dates:
                break
            prev_date -= timedelta(days=1)
            attempts += 1
        
        if attempts >= 10:
            flash('Could not find a previous school day.', 'error')
            return redirect(url_for('attendance.mark_attendance_page',
                term_id=term_id, assignment_id=assignment_id, date=target_date_str))
        
        # Get previous day's attendance
        enrollments = StudentEnrollment.query.filter_by(
            class_arm_assignment_id=assignment_id,
            is_active=True
        ).all()
        
        prev_attendance = {}
        for enrollment in enrollments:
            att = Attendance.query.filter_by(
                enrollment_id=enrollment.id,
                date=prev_date
            ).first()
            if att:
                prev_attendance[enrollment.id] = att
        
        if not prev_attendance:
            flash(f'No attendance found for {prev_date.strftime("%d %b %Y")}.', 'warning')
            return redirect(url_for('attendance.mark_attendance_page',
                term_id=term_id, assignment_id=assignment_id, date=target_date_str))
        
        # Get or find week for target date
        week = Week.query.filter(
            Week.term_id == term_id,
            Week.start_date <= target_date,
            Week.end_date >= target_date
        ).first()
        
        if not week:
            flash('Target date is not within any defined week.', 'error')
            return redirect(url_for('attendance.mark_attendance_page',
                term_id=term_id, assignment_id=assignment_id, date=target_date_str))
        
        # Copy attendance
        copied = 0
        for enrollment in enrollments:
            prev = prev_attendance.get(enrollment.id)
            if prev:
                existing = Attendance.query.filter_by(
                    enrollment_id=enrollment.id,
                    date=target_date
                ).first()
                
                if existing:
                    existing.morning_present = prev.morning_present
                    existing.afternoon_present = prev.afternoon_present
                else:
                    new_att = Attendance(
                        enrollment_id=enrollment.id,
                        week_id=week.id,
                        date=target_date,
                        morning_present=prev.morning_present,
                        afternoon_present=prev.afternoon_present
                    )
                    db.session.add(new_att)
                copied += 1
        
        db.session.commit()
        flash(f'Copied attendance from {prev_date.strftime("%d %b")} for {copied} students!', 'success')
        
    except Exception as e:
        db.session.rollback()
        flash(f'Error: {str(e)}', 'error')
    
    return redirect(url_for('attendance.mark_attendance_page',
        term_id=term_id, assignment_id=assignment_id, date=target_date_str))


@attendance_bp.route('/weekly')
@login_required
def weekly_summary():
    """View weekly attendance summary"""
    assignment_id = request.args.get('assignment_id', type=int)
    week_id = request.args.get('week_id', type=int)
    
    # Check class access
    if assignment_id and not can_access_class(assignment_id):
        flash('You do not have access to this class.', 'error')
        return redirect(url_for('attendance.weekly_summary'))
    
    # Get active term
    active_term = get_active_term()
    
    assignments = []
    weeks = []
    if active_term:
        all_assignments = ClassArmAssignment.query.filter_by(term_id=active_term.id).all()
        assignments = filter_classes_for_user(all_assignments, form_only=True)
        weeks = Week.query.filter_by(term_id=active_term.id).order_by(Week.week_number).all()
    
    summary = None
    selected_assignment = None
    selected_week = None
    
    if assignment_id and week_id:
        selected_assignment = db.session.get(ClassArmAssignment, assignment_id)
        selected_week = db.session.get(Week, week_id)
        summary = get_weekly_attendance_summary(assignment_id, week_id)
    
    return render_template('attendance/weekly.html',
        assignments=assignments,
        weeks=weeks,
        selected_assignment=selected_assignment,
        selected_week=selected_week,
        summary=summary
    )


@attendance_bp.route('/weekly/export')
@login_required
def export_weekly():
    """Export weekly attendance to Excel"""
    assignment_id = request.args.get('assignment_id', type=int)
    week_id = request.args.get('week_id', type=int)
    
    if not assignment_id or not week_id:
        flash('Please select a class and week.', 'error')
        return redirect(url_for('attendance.weekly_summary'))
    
    assignment = db.get_or_404(ClassArmAssignment, assignment_id)
    week = db.get_or_404(Week, week_id)
    
    summary = get_weekly_attendance_summary(assignment_id, week_id)
    
    excel_file = export_attendance_to_excel(
        summary,
        assignment.display_name,
        {
            'week_number': week.week_number,
            'start_date': week.start_date.strftime('%Y-%m-%d'),
            'end_date': week.end_date.strftime('%Y-%m-%d')
        }
    )
    
    filename = f"attendance_{assignment.display_name.replace(' ', '_')}_week{week.week_number}.xlsx"

    return xlsx_response(excel_file, filename)


@attendance_bp.route('/termly')
@login_required
def termly_summary():
    """View termly attendance summary"""
    assignment_id = request.args.get('assignment_id', type=int)
    term_id = request.args.get('term_id', type=int)
    
    # Check class access
    if assignment_id and not can_access_class(assignment_id):
        flash('You do not have access to this class.', 'error')
        return redirect(url_for('attendance.termly_summary'))
    
    # Get terms
    terms = Term.query.join(AcademicSession).order_by(
        AcademicSession.name.desc(),
        Term.term_number
    ).all()
    
    if not term_id:
        active_term = get_active_term()
        if active_term:
            term_id = active_term.id
    
    assignments = []
    selected_term = None
    if term_id:
        selected_term = db.session.get(Term, term_id)
        all_assignments = ClassArmAssignment.query.filter_by(term_id=term_id).all()
        assignments = filter_classes_for_user(all_assignments, form_only=True)
    
    summary = None
    selected_assignment = None
    
    if assignment_id and term_id:
        selected_assignment = db.session.get(ClassArmAssignment, assignment_id)
        summary = get_termly_attendance_summary(assignment_id, term_id)
    
    return render_template('attendance/termly.html',
        terms=terms,
        assignments=assignments,
        selected_term=selected_term,
        selected_assignment=selected_assignment,
        summary=summary
    )


@attendance_bp.route('/termly/export')
@login_required
def export_termly():
    """Export termly attendance to Excel"""
    assignment_id = request.args.get('assignment_id', type=int)
    term_id = request.args.get('term_id', type=int)
    
    if not assignment_id or not term_id:
        flash('Please select a class and term.', 'error')
        return redirect(url_for('attendance.termly_summary'))
    
    assignment = db.get_or_404(ClassArmAssignment, assignment_id)
    term = db.get_or_404(Term, term_id)
    
    summary = get_termly_attendance_summary(assignment_id, term_id)
    
    if not summary:
        flash('No attendance data found.', 'error')
        return redirect(url_for('attendance.termly_summary'))
    
    # Create Excel file
    import openpyxl
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    from io import BytesIO
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Termly Attendance"
    
    # Styles
    header_font = Font(bold=True, size=14)
    subheader_font = Font(bold=True, size=11)
    center_align = Alignment(horizontal='center', vertical='center')
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    green_fill = PatternFill(start_color='90EE90', end_color='90EE90', fill_type='solid')
    yellow_fill = PatternFill(start_color='FFFF99', end_color='FFFF99', fill_type='solid')
    red_fill = PatternFill(start_color='FFB6C1', end_color='FFB6C1', fill_type='solid')
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    header_font_white = Font(bold=True, size=11, color='FFFFFF')
    
    term_info = summary.get('term_info', {})
    total_school_days = term_info.get('total_school_days', 0)
    total_times_opened = term_info.get('total_times_opened', 0)
    class_totals = summary.get('class_totals', {})
    
    # Title
    ws.merge_cells('A1:I1')
    ws['A1'] = f"TERMLY ATTENDANCE REPORT - {assignment.display_name}"
    ws['A1'].font = header_font
    ws['A1'].alignment = center_align
    
    ws.merge_cells('A2:I2')
    ws['A2'] = f"{term.session.name} - Term {term.term_number}"
    ws['A2'].font = subheader_font
    ws['A2'].alignment = center_align
    
    # Term info row
    ws.merge_cells('A3:I3')
    ws['A3'] = f"Total School Days: {total_school_days} | Times Opened (AM+PM): {total_times_opened} | Total Students: {summary.get('total_students', 0)}"
    ws['A3'].alignment = center_align
    
    # Headers
    headers = ['S/N', 'Student Name', 'Gender', 'AM Present', 'PM Present', 'Total Present', 'Times Opened', 'Attendance %', 'Status']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=5, column=col, value=header)
        cell.font = header_font_white
        cell.alignment = center_align
        cell.border = thin_border
        cell.fill = header_fill
    
    # Data rows
    row = 6
    for idx, student in enumerate(summary.get('students', []), 1):
        morning = student.get('morning_total', 0)
        afternoon = student.get('afternoon_total', 0)
        total_present = student.get('termly_total', morning + afternoon)
        percentage = student.get('percentage', 0)
        
        ws.cell(row=row, column=1, value=idx).border = thin_border
        ws.cell(row=row, column=1).alignment = center_align
        
        ws.cell(row=row, column=2, value=student.get('student_name', '')).border = thin_border
        
        ws.cell(row=row, column=3, value=student.get('gender', '')).border = thin_border
        ws.cell(row=row, column=3).alignment = center_align
        
        ws.cell(row=row, column=4, value=morning).border = thin_border
        ws.cell(row=row, column=4).alignment = center_align
        
        ws.cell(row=row, column=5, value=afternoon).border = thin_border
        ws.cell(row=row, column=5).alignment = center_align
        
        ws.cell(row=row, column=6, value=total_present).border = thin_border
        ws.cell(row=row, column=6).alignment = center_align
        
        ws.cell(row=row, column=7, value=total_times_opened).border = thin_border
        ws.cell(row=row, column=7).alignment = center_align
        
        pct_cell = ws.cell(row=row, column=8, value=f"{percentage:.1f}%")
        pct_cell.border = thin_border
        pct_cell.alignment = center_align
        
        # Status based on percentage
        if percentage >= 75:
            status = "Good"
            status_fill = green_fill
        elif percentage >= 50:
            status = "Fair"
            status_fill = yellow_fill
        else:
            status = "At Risk"
            status_fill = red_fill
        
        status_cell = ws.cell(row=row, column=9, value=status)
        status_cell.border = thin_border
        status_cell.alignment = center_align
        status_cell.fill = status_fill
        
        row += 1
    
    # Summary rows
    row += 1
    ws.merge_cells(f'A{row}:C{row}')
    ws.cell(row=row, column=1, value="CLASS TOTALS").font = subheader_font
    ws.cell(row=row, column=4, value=class_totals.get('total_morning', 0)).font = subheader_font
    ws.cell(row=row, column=4).alignment = center_align
    ws.cell(row=row, column=5, value=class_totals.get('total_afternoon', 0)).font = subheader_font
    ws.cell(row=row, column=5).alignment = center_align
    ws.cell(row=row, column=6, value=class_totals.get('total_attendance', 0)).font = subheader_font
    ws.cell(row=row, column=6).alignment = center_align
    
    row += 1
    ws.merge_cells(f'A{row}:C{row}')
    ws.cell(row=row, column=1, value="CLASS AVERAGE").font = subheader_font
    ws.cell(row=row, column=8, value=f"{class_totals.get('termly_percentage', 0):.1f}%").font = subheader_font
    ws.cell(row=row, column=8).alignment = center_align
    
    # Gender breakdown
    row += 2
    ws.cell(row=row, column=1, value="GENDER BREAKDOWN:").font = subheader_font
    row += 1
    ws.cell(row=row, column=1, value=f"Male Students: {summary.get('total_male_students', 0)}")
    ws.cell(row=row, column=3, value=f"Male Attendance: {class_totals.get('male_attendance', 0)}")
    row += 1
    ws.cell(row=row, column=1, value=f"Female Students: {summary.get('total_female_students', 0)}")
    ws.cell(row=row, column=3, value=f"Female Attendance: {class_totals.get('female_attendance', 0)}")
    
    # Column widths
    ws.column_dimensions['A'].width = 6
    ws.column_dimensions['B'].width = 28
    ws.column_dimensions['C'].width = 10
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['F'].width = 14
    ws.column_dimensions['G'].width = 14
    ws.column_dimensions['H'].width = 14
    ws.column_dimensions['I'].width = 12
    
    filename = f"attendance_{assignment.display_name.replace(' ', '_')}_term{term.term_number}.xlsx"

    return xlsx_response(wb, filename)


# ============================================================================
# API ENDPOINTS
# ============================================================================

@attendance_bp.route('/api/summary/<int:assignment_id>/<date_str>')
@login_required
def api_daily_summary(assignment_id, date_str):
    """API endpoint for daily attendance summary"""
    require_branch_access(db.get_or_404(ClassArmAssignment, assignment_id).branch_id)
    try:
        target_date = datetime.strptime(date_str, '%Y-%m-%d').date()
        summary = get_daily_attendance_summary(assignment_id, target_date)
        return jsonify(summary)
    except Exception as e:
        return jsonify({'error': str(e)}), 400


@attendance_bp.route('/api/check-attendance')
@login_required
def api_check_attendance():
    """Check if attendance exists for a date/assignment"""
    assignment_id = request.args.get('assignment_id', type=int)
    target_date = request.args.get('date')
    
    if not assignment_id or not target_date:
        return jsonify({'exists': False})
    
    target_date = datetime.strptime(target_date, '%Y-%m-%d').date()
    
    # Check if any attendance records exist
    exists = Attendance.query.join(StudentEnrollment).filter(
        StudentEnrollment.class_arm_assignment_id == assignment_id,
        Attendance.date == target_date
    ).first() is not None
    
    return jsonify({'exists': exists})


@attendance_bp.route('/api/school-days/<int:week_id>')
@login_required
def api_school_days(week_id):
    """Get school days for a week"""
    week = db.get_or_404(Week, week_id)
    holidays = Holiday.query.filter_by(term_id=week.term_id).all()
    holiday_dates = set(h.date for h in holidays)
    
    school_days = []
    current = week.start_date
    while current <= week.end_date:
        if current.weekday() < 5 and current not in holiday_dates:
            school_days.append({
                'date': current.isoformat(),
                'day_name': current.strftime('%A')
            })
        current += timedelta(days=1)
    
    return jsonify(school_days)


# ============================================================================
# JSON API for the React offline pilot (roster read + idempotent mark write)
# ============================================================================

def _week_for_date(term_id, target_date):
    """The Week (within term_id) that contains target_date, or None."""
    return Week.query.filter(
        Week.term_id == term_id,
        Week.start_date <= target_date,
        Week.end_date >= target_date,
    ).first()


@attendance_bp.route('/app')
@attendance_bp.route('/react')   # keep the earlier pilot URL working
@login_required
def attendance_app():
    """The attendance mini-SPA (React). Self-loads its data from /api/context;
    the classic server-rendered pages remain available as a fallback."""
    if not can_mark_attendance():
        flash('You do not have permission to mark attendance.', 'error')
        return redirect(url_for('main.dashboard'))
    return render_template('attendance/app.html')


@attendance_bp.route('/api/roster')
@login_required
def api_roster():
    """Roster + existing marks for a class arm on a date (JSON).

    Branch-scoped and permission-checked; the React pilot caches this in
    IndexedDB so the register opens offline.
    """
    assignment_id = request.args.get('assignment_id', type=int)
    if not assignment_id:
        return jsonify({'error': 'assignment_id required'}), 400
    caa = db.get_or_404(ClassArmAssignment, assignment_id)
    require_branch_access(caa.branch_id)
    if not (can_mark_attendance(assignment_id) and can_access_class(assignment_id)):
        return jsonify({'error': 'forbidden'}), 403

    ds = request.args.get('date')
    try:
        target = datetime.strptime(ds, '%Y-%m-%d').date() if ds else date.today()
    except Exception:
        target = date.today()

    holidays = Holiday.query.filter_by(term_id=caa.term_id).all()
    week = _week_for_date(caa.term_id, target)
    enrollments = (StudentEnrollment.query
                   .filter_by(class_arm_assignment_id=assignment_id, is_active=True)
                   .join(Student).order_by(Student.surname, Student.first_name).all())
    existing = {a.enrollment_id: a for a in Attendance.query.filter(
        Attendance.enrollment_id.in_([e.id for e in enrollments] or [-1]),
        Attendance.date == target).all()}
    students = [{
        'enrollment_id': e.id,
        'student_id': e.student.student_id,
        'name': e.student.full_name,
        'morning_present': existing[e.id].morning_present if e.id in existing else True,
        'afternoon_present': existing[e.id].afternoon_present if e.id in existing else True,
    } for e in enrollments]
    return jsonify({
        'assignment_id': assignment_id,
        'class_name': caa.display_name,
        'date': target.isoformat(),
        'week_id': week.id if week else None,
        'school_day': is_school_day(target, holidays),
        'students': students,
    })


@attendance_bp.route('/api/mark', methods=['POST'])
@login_required
def api_mark():
    """Record attendance from JSON — the React pilot's offline outbox flushes
    here. Idempotent upsert (mark_attendance_bulk), branch-scoped + CSRF."""
    from utils.calculations import mark_attendance_bulk
    data = request.get_json(silent=True) or {}
    assignment_id = data.get('assignment_id')
    caa = db.session.get(ClassArmAssignment, assignment_id) if assignment_id else None
    if not caa:
        return jsonify({'error': 'invalid assignment'}), 400
    require_branch_access(caa.branch_id)
    if not can_mark_attendance(assignment_id):
        return jsonify({'error': 'forbidden'}), 403
    try:
        target = datetime.strptime(data['date'], '%Y-%m-%d').date()
    except Exception:
        return jsonify({'error': 'bad date'}), 400
    week = _week_for_date(caa.term_id, target)
    if not week:
        return jsonify({'error': 'no school week for that date'}), 400

    session_type = 'afternoon' if data.get('session_type') == 'afternoon' else 'morning'
    auto_copy = bool(data.get('auto_copy', session_type == 'morning'))
    enroll_ids = [e.id for e in StudentEnrollment.query.filter_by(
        class_arm_assignment_id=assignment_id, is_active=True).all()]
    valid = set(enroll_ids)
    present = [int(x) for x in data.get('present', []) if int(x) in valid]
    try:
        count = mark_attendance_bulk(
            enroll_ids, target, week.id, session_type, present, marked_by='React',
            auto_copy_to_afternoon=auto_copy if session_type == 'morning' else False)
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': str(e)}), 500
    return jsonify({'ok': True, 'count': count, 'date': target.isoformat(),
                    'session_type': session_type})


@attendance_bp.route('/api/context')
@login_required
def api_context():
    """Selector data for the attendance SPA (term, terms, markable classes,
    weeks, holidays) — cacheable so the picker works offline."""
    if not can_mark_attendance():
        return jsonify({'error': 'forbidden'}), 403
    term = get_active_term()
    terms = [{'id': t.id, 'name': t.full_name, 'active': bool(t.is_active)}
             for t in Term.query.join(AcademicSession)
             .order_by(AcademicSession.name.desc(), Term.term_number).all()]
    classes, weeks, holidays = [], [], []
    if term:
        classes = [{'id': a.id, 'name': a.display_name} for a in filter_classes_for_user(
            ClassArmAssignment.query.filter_by(term_id=term.id).all(), form_only=True)]
        weeks = [{'id': w.id, 'number': w.week_number,
                  'start': w.start_date.isoformat(), 'end': w.end_date.isoformat()}
                 for w in Week.query.filter_by(term_id=term.id).order_by(Week.week_number).all()]
        holidays = [h.date.isoformat() for h in Holiday.query.filter_by(term_id=term.id).all()]
    return jsonify({
        'term': {'id': term.id, 'name': term.full_name} if term else None,
        'terms': terms, 'classes': classes, 'weeks': weeks, 'holidays': holidays,
        'today': date.today().isoformat(),
    })


@attendance_bp.route('/api/week')
@login_required
def api_week():
    """Weekly grid (students × school-days) with existing AM/PM marks."""
    assignment_id = request.args.get('assignment_id', type=int)
    week_id = request.args.get('week_id', type=int)
    if not assignment_id:
        return jsonify({'error': 'assignment_id required'}), 400
    caa = db.get_or_404(ClassArmAssignment, assignment_id)
    require_branch_access(caa.branch_id)
    if not (can_mark_attendance(assignment_id) and can_access_class(assignment_id)):
        return jsonify({'error': 'forbidden'}), 403
    week = db.session.get(Week, week_id) if week_id else None
    if not week:
        return jsonify({'error': 'week_id required'}), 400
    holidays = Holiday.query.filter_by(term_id=week.term_id).all()
    days = _week_school_days(week, holidays)
    enrollments = (StudentEnrollment.query
                   .filter_by(class_arm_assignment_id=assignment_id, is_active=True)
                   .join(Student).order_by(Student.surname, Student.first_name).all())
    recs = {}
    for r in Attendance.query.filter(
            Attendance.enrollment_id.in_([e.id for e in enrollments] or [-1]),
            Attendance.date.in_(days)).all():
        recs.setdefault(r.enrollment_id, {})[r.date.isoformat()] = {
            'am': bool(r.morning_present), 'pm': bool(r.afternoon_present)}
    students = [{
        'enrollment_id': e.id, 'student_id': e.student.student_id, 'name': e.student.full_name,
        'days': {d.isoformat(): recs.get(e.id, {}).get(d.isoformat(), {'am': True, 'pm': True})
                 for d in days},
    } for e in enrollments]
    return jsonify({
        'assignment_id': assignment_id, 'class_name': caa.display_name,
        'week_id': week.id, 'week_number': week.week_number,
        'days': [{'date': d.isoformat(), 'label': d.strftime('%a %d/%m')} for d in days],
        'students': students,
    })


@attendance_bp.route('/api/week/mark', methods=['POST'])
@login_required
def api_week_mark():
    """Idempotent bulk upsert for the weekly grid. Body: {assignment_id, week_id,
    marks:[{enrollment_id, date, am, pm}]}. The offline outbox flushes here."""
    data = request.get_json(silent=True) or {}
    caa = db.session.get(ClassArmAssignment, data.get('assignment_id'))
    week = db.session.get(Week, data.get('week_id'))
    if not caa or not week:
        return jsonify({'error': 'invalid assignment/week'}), 400
    require_branch_access(caa.branch_id)
    if not can_mark_attendance(caa.id):
        return jsonify({'error': 'forbidden'}), 403
    holidays = Holiday.query.filter_by(term_id=week.term_id).all()
    valid_days = {d.isoformat() for d in _week_school_days(week, holidays)}
    enroll_ids = {e.id for e in StudentEnrollment.query.filter_by(
        class_arm_assignment_id=caa.id, is_active=True).all()}
    existing = {(r.enrollment_id, r.date.isoformat()): r for r in Attendance.query.filter(
        Attendance.enrollment_id.in_(list(enroll_ids) or [-1]),
        Attendance.week_id == week.id).all()}
    saved = 0
    try:
        for m in data.get('marks', []):
            eid, iso = m.get('enrollment_id'), m.get('date')
            if eid not in enroll_ids or iso not in valid_days:
                continue
            rec = existing.get((eid, iso))
            if rec is None:
                rec = Attendance(enrollment_id=eid, week_id=week.id,
                                 date=datetime.strptime(iso, '%Y-%m-%d').date())
                db.session.add(rec)
                existing[(eid, iso)] = rec
            rec.week_id = week.id
            rec.morning_present = bool(m.get('am'))
            rec.afternoon_present = bool(m.get('pm'))
            rec.marked_by = 'React'
            saved += 1
        db.session.commit()
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': str(e)}), 500
    return jsonify({'ok': True, 'saved': saved})


@attendance_bp.route('/api/copy-previous', methods=['POST'])
@login_required
def api_copy_previous():
    """Copy the previous school day's marks into `date`. Body: {assignment_id, date}."""
    data = request.get_json(silent=True) or {}
    caa = db.session.get(ClassArmAssignment, data.get('assignment_id'))
    if not caa:
        return jsonify({'error': 'invalid assignment'}), 400
    require_branch_access(caa.branch_id)
    if not can_mark_attendance(caa.id):
        return jsonify({'error': 'forbidden'}), 403
    try:
        target = datetime.strptime(data['date'], '%Y-%m-%d').date()
    except Exception:
        return jsonify({'error': 'bad date'}), 400
    week = _week_for_date(caa.term_id, target)
    if not week:
        return jsonify({'error': 'date not in any school week'}), 400
    holiday_dates = {h.date for h in Holiday.query.filter_by(term_id=caa.term_id).all()}
    prev, n = target - timedelta(days=1), 0
    while n < 10 and (prev.weekday() >= 5 or prev in holiday_dates):
        prev -= timedelta(days=1); n += 1
    enrollments = StudentEnrollment.query.filter_by(
        class_arm_assignment_id=caa.id, is_active=True).all()
    prevmap = {a.enrollment_id: a for a in Attendance.query.filter(
        Attendance.enrollment_id.in_([e.id for e in enrollments] or [-1]),
        Attendance.date == prev).all()}
    if not prevmap:
        return jsonify({'ok': True, 'copied': 0, 'from': prev.isoformat(),
                        'note': 'No attendance on the previous school day.'})
    copied = 0
    try:
        for e in enrollments:
            p = prevmap.get(e.id)
            if not p:
                continue
            rec = Attendance.query.filter_by(enrollment_id=e.id, date=target).first()
            if rec:
                rec.morning_present, rec.afternoon_present = p.morning_present, p.afternoon_present
            else:
                db.session.add(Attendance(
                    enrollment_id=e.id, week_id=week.id, date=target,
                    morning_present=p.morning_present, afternoon_present=p.afternoon_present,
                    marked_by='React'))
            copied += 1
        db.session.commit()
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': str(e)}), 500
    return jsonify({'ok': True, 'copied': copied, 'from': prev.isoformat()})


# ============================================================================
# ATTENDANCE ALERTS
# ============================================================================

@attendance_bp.route('/alerts')
@login_required
def attendance_alerts():
    """View students with poor attendance"""
    
    term_id = request.args.get('term_id', type=int)
    threshold = request.args.get('threshold', type=float)
    
    terms = Term.query.order_by(Term.id.desc()).all()
    
    # Get active term if not specified
    if not term_id:
        active_term = get_active_term()
        if active_term:
            term_id = active_term.id
    
    selected_term = db.session.get(Term, term_id) if term_id else None
    
    # Default threshold from settings or 75%
    if threshold is None:
        threshold = 75.0
    
    alerts = []
    
    if selected_term:
        # Get all enrollments for this term
        enrollments = StudentEnrollment.query.join(ClassArmAssignment).filter(
            ClassArmAssignment.term_id == term_id,
            StudentEnrollment.is_active == True
        ).all()
        
        # Get weeks for this term
        weeks = Week.query.filter_by(term_id=term_id).all()
        week_ids = [w.id for w in weeks]
        
        if week_ids:
            # Get holidays
            holidays = Holiday.query.filter_by(term_id=term_id).all()
            holiday_dates = set(h.date for h in holidays)
            
            # Calculate total times opened
            total_school_days = 0
            for week in weeks:
                current = week.start_date
                while current <= week.end_date:
                    if current.weekday() < 5 and current not in holiday_dates:
                        total_school_days += 1
                    current += timedelta(days=1)
            
            total_times_opened = total_school_days * 2  # Morning + Afternoon
            
            if total_times_opened > 0:
                for enrollment in enrollments:
                    # Count attendance
                    attendance_records = Attendance.query.filter(
                        Attendance.enrollment_id == enrollment.id,
                        Attendance.week_id.in_(week_ids)
                    ).all()
                    
                    total_present = sum(
                        (1 if a.morning_present else 0) + (1 if a.afternoon_present else 0)
                        for a in attendance_records
                    )
                    
                    percentage = round((total_present / total_times_opened) * 100, 2)
                    
                    if percentage < threshold:
                        alerts.append({
                            'student': enrollment.student,
                            'class': enrollment.class_arm_assignment.display_name,
                            'present': total_present,
                            'total': total_times_opened,
                            'percentage': percentage,
                            'status': 'critical' if percentage < 50 else 'warning'
                        })
        
        # Sort by percentage (lowest first)
        alerts.sort(key=lambda x: x['percentage'])
    
    return render_template('attendance/alerts.html',
        terms=terms, term_id=term_id, selected_term=selected_term,
        threshold=threshold, alerts=alerts
    )


@attendance_bp.route('/alerts/export')
@login_required
def export_alerts():
    """Export attendance alerts to Excel"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    import io
    
    term_id = request.args.get('term_id', type=int)
    threshold = request.args.get('threshold', type=float) or 75.0
    
    if not term_id:
        flash('Select a term first.', 'error')
        return redirect(url_for('attendance.attendance_alerts'))
    
    selected_term = db.session.get(Term, term_id)
    
    # Get alerts (same logic as above)
    enrollments = StudentEnrollment.query.join(ClassArmAssignment).filter(
        ClassArmAssignment.term_id == term_id,
        StudentEnrollment.is_active == True
    ).all()
    
    weeks = Week.query.filter_by(term_id=term_id).all()
    week_ids = [w.id for w in weeks]
    
    holidays = Holiday.query.filter_by(term_id=term_id).all()
    holiday_dates = set(h.date for h in holidays)
    
    total_school_days = 0
    for week in weeks:
        current = week.start_date
        while current <= week.end_date:
            if current.weekday() < 5 and current not in holiday_dates:
                total_school_days += 1
            current += timedelta(days=1)
    
    total_times_opened = total_school_days * 2
    
    alerts = []
    if total_times_opened > 0:
        for enrollment in enrollments:
            attendance_records = Attendance.query.filter(
                Attendance.enrollment_id == enrollment.id,
                Attendance.week_id.in_(week_ids)
            ).all()
            
            total_present = sum(
                (1 if a.morning_present else 0) + (1 if a.afternoon_present else 0)
                for a in attendance_records
            )
            
            percentage = round((total_present / total_times_opened) * 100, 2)
            
            if percentage < threshold:
                alerts.append({
                    'student': enrollment.student,
                    'class': enrollment.class_arm_assignment.display_name,
                    'present': total_present,
                    'total': total_times_opened,
                    'percentage': percentage
                })
    
    alerts.sort(key=lambda x: x['percentage'])
    
    # Create Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "Attendance Alerts"
    
    # Title
    ws['A1'] = f"Attendance Alerts - {selected_term.full_name}"
    ws['A1'].font = Font(bold=True, size=14)
    ws['A2'] = f"Students below {threshold}% attendance"
    
    # Headers
    headers = ['S/N', 'Student Name', 'Student ID', 'Class', 'Present', 'Total', 'Percentage']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col, value=header)
        cell.font = Font(bold=True)
    
    # Data
    for idx, alert in enumerate(alerts, 1):
        row = idx + 4
        ws.cell(row=row, column=1, value=idx)
        ws.cell(row=row, column=2, value=alert['student'].full_name)
        ws.cell(row=row, column=3, value=alert['student'].student_id)
        ws.cell(row=row, column=4, value=alert['class'])
        ws.cell(row=row, column=5, value=alert['present'])
        ws.cell(row=row, column=6, value=alert['total'])
        ws.cell(row=row, column=7, value=f"{alert['percentage']}%")
        
        # Highlight critical
        if alert['percentage'] < 50:
            for col in range(1, 8):
                ws.cell(row=row, column=col).fill = PatternFill(start_color='FFCDD2', fill_type='solid')
    
    return xlsx_response(wb, f'attendance_alerts_{selected_term.name}.xlsx')


# ============================================================================
# PRINT-READY ATTENDANCE REGISTER
# ============================================================================

@attendance_bp.route('/print-register')
@login_required
def print_register():
    """Print-ready attendance register"""
    from models import SchoolSettings
    
    term_id = request.args.get('term_id', type=int)
    assignment_id = request.args.get('assignment_id', type=int)
    week_id = request.args.get('week_id', type=int)
    
    terms = Term.query.order_by(Term.id.desc()).all()
    
    # Get active term if not specified
    if not term_id:
        active_term = get_active_term()
        if active_term:
            term_id = active_term.id
    
    selected_term = db.session.get(Term, term_id) if term_id else None
    
    # Get assignments for term
    assignments = []
    if term_id:
        assignments = ClassArmAssignment.query.filter_by(term_id=term_id).all()
    
    selected_assignment = db.session.get(ClassArmAssignment, assignment_id) if assignment_id else None
    
    # Get weeks for term
    weeks = []
    if term_id:
        weeks = Week.query.filter_by(term_id=term_id).order_by(Week.week_number).all()
    
    selected_week = db.session.get(Week, week_id) if week_id else None
    
    register_data = []
    school_days = []
    school_name = SchoolSettings.get('school_name', 'School Name')
    
    if selected_assignment and selected_week:
        # Get holidays
        holidays = Holiday.query.filter_by(term_id=term_id).all()
        holiday_dates = set(h.date for h in holidays)
        
        # Calculate school days in the week
        current = selected_week.start_date
        while current <= selected_week.end_date:
            if current.weekday() < 5 and current not in holiday_dates:
                school_days.append(current)
            current += timedelta(days=1)
        
        # Get enrolled students
        enrollments = StudentEnrollment.query.filter_by(
            class_arm_assignment_id=assignment_id,
            is_active=True
        ).join(Student).order_by(Student.surname, Student.first_name).all()
        
        for enrollment in enrollments:
            # Get attendance for the week
            attendance_records = Attendance.query.filter_by(
                enrollment_id=enrollment.id,
                week_id=week_id
            ).all()
            
            # Build attendance lookup by date
            attendance_lookup = {}
            total_present = 0
            
            for record in attendance_records:
                attendance_lookup[record.date.isoformat()] = {
                    'morning_present': record.morning_present,
                    'afternoon_present': record.afternoon_present
                }
                if record.morning_present:
                    total_present += 1
                if record.afternoon_present:
                    total_present += 1
            
            register_data.append({
                'student': enrollment.student,
                'attendance': attendance_lookup,
                'total_present': total_present
            })
    
    return render_template('attendance/print_register.html',
        terms=terms, term_id=term_id, selected_term=selected_term,
        assignments=assignments, assignment_id=assignment_id, selected_assignment=selected_assignment,
        weeks=weeks, week_id=week_id, selected_week=selected_week,
        register_data=register_data, school_days=school_days, school_name=school_name
    )
