"""generator_bp — exports routes (split from the former routes/generator.py)."""
from routes.generator import *  # noqa: F401,F403


@generator_bp.route('/results/<batch_id>/print')
@login_required
def print_results(batch_id):
    level = get_current_level()
    results = GenTimetableResult.query.filter_by(batch_id=batch_id, school_level=level, branch_id=gen_bid()).all()
    if not results:
        flash('No results.', 'error')
        return redirect(url_for('generator.results_list'))
    
    timetables = {}
    for r in results:
        key = f"{r.class_name}_{r.arm_name}"
        if key not in timetables:
            timetables[key] = {'class_name': r.class_name, 'arm_name': r.arm_name, 'grid': {d: {} for d in range(5)}}
        timetables[key]['grid'][r.day_of_week][r.period_number] = r
    
    rules = {r.rule_type: r.value for r in GenTimetableRule.query.filter_by(is_active=True, branch_id=gen_bid()).all()}
    
    return render_template('generator/print_results.html',
        batch_id=batch_id, timetables=dict(sorted(timetables.items())),
        periods_per_day=int(rules.get('periods_per_day', 8)),
        break_after_period=int(rules.get('break_after_period', 4)),
        days=DAYS_OF_WEEK,
        school_name=GenSettings.get('school_name', ''),
        academic_year=GenSettings.get('academic_year', ''),
        term_name=GenSettings.get('term_name', '')
    )


@generator_bp.route('/results/<batch_id>/print/<class_name>/<arm_name>')
@login_required
def print_single_timetable(batch_id, class_name, arm_name):
    results = GenTimetableResult.query.filter_by(batch_id=batch_id, class_name=class_name, arm_name=arm_name, branch_id=gen_bid()).all()
    if not results:
        flash('Not found.', 'error')
        return redirect(url_for('generator.view_results', batch_id=batch_id))
    
    grid = {d: {} for d in range(5)}
    for r in results:
        grid[r.day_of_week][r.period_number] = r
    
    rules = {r.rule_type: r.value for r in GenTimetableRule.query.filter_by(is_active=True, branch_id=gen_bid()).all()}
    
    return render_template('generator/print_single.html',
        batch_id=batch_id, class_name=class_name, arm_name=arm_name, grid=grid,
        periods_per_day=int(rules.get('periods_per_day', 8)),
        break_after_period=int(rules.get('break_after_period', 4)),
        days=DAYS_OF_WEEK,
        school_name=GenSettings.get('school_name', ''),
        academic_year=GenSettings.get('academic_year', ''),
        term_name=GenSettings.get('term_name', '')
    )


@generator_bp.route('/results/<batch_id>/export')
@login_required
def export_results(batch_id):
    """Export timetable by class - Each class MAXIMIZES A4 landscape page, no teacher names, NO COLORS for B&W printing"""
    import openpyxl
    from openpyxl.styles import Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    
    results = GenTimetableResult.query.filter_by(batch_id=batch_id, branch_id=gen_bid()).all()
    if not results:
        flash('No results.', 'error')
        return redirect(url_for('generator.results_list'))
    
    # Get school info
    school_name = GenSettings.get('school_name', 'School')
    school_address = GenSettings.get('school_address', '')
    
    timetables = {}
    for r in results:
        key = f"{r.class_name}_{r.arm_name}"
        if key not in timetables:
            timetables[key] = {'class_name': r.class_name, 'arm_name': r.arm_name, 'grid': {d: {} for d in range(5)}}
        timetables[key]['grid'][r.day_of_week][r.period_number] = r
    
    rules = {r.rule_type: r.value for r in GenTimetableRule.query.filter_by(is_active=True, branch_id=gen_bid()).all()}
    periods_per_day = int(rules.get('periods_per_day', 8))
    
    # Period time slots (8:20 AM start, 40 min each, 30 min break after period 5)
    period_times = []
    break_time = ""
    start_hour, start_min = 8, 20
    for p in range(1, periods_per_day + 1):
        start_total = start_hour * 60 + start_min
        end_total = start_total + 40
        start_h, start_m = start_total // 60, start_total % 60
        end_h, end_m = end_total // 60, end_total % 60
        period_times.append(f"P{p}\n{start_h}:{start_m:02d}-{end_h}:{end_m:02d}")
        start_hour, start_min = end_h, end_m
        if p == 5:
            break_start = f"{end_h}:{end_m:02d}"
            start_total = start_hour * 60 + start_min + 30
            start_hour, start_min = start_total // 60, start_total % 60
            break_end = f"{start_hour}:{start_min:02d}"
            break_time = f"BREAK\n{break_start}\n-\n{break_end}"
    
    # Abbreviations
    abbrev_map = {
        'Mathematics': 'Maths', 'English Language': 'Eng', 'Physics': 'Phy',
        'Chemistry': 'Chem', 'Biology': 'Bio', 'Economics': 'Econs',
        'Government': 'Govt', 'Literature in English': 'Lit', 'Agricultural Science': 'Agric',
        'Christian Religious Studies': 'CRS', 'Civic Education': 'Civic',
        'Computer Studies': 'Comp', 'Commerce': 'Comm', 'Geography': 'Geo',
        'Further Mathematics': 'F/Mth', 'Livestock Farming': 'Livst',
        'History': 'Hist', 'Phonics': 'Phon',
    }
    
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    
    # NO COLORS - Pure black text on white, maximum contrast for B&W printing
    school_font = Font(bold=True, size=32, color='000000')
    address_font = Font(bold=False, size=14, color='000000')
    class_title_font = Font(bold=True, size=28, color='000000')
    day_font = Font(bold=True, size=24, color='000000')
    header_font = Font(bold=True, size=11, color='000000')
    cell_font = Font(bold=True, size=32, color='000000')
    break_header_font = Font(bold=True, size=8, color='000000')
    
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    thick_border = Border(
        left=Side(style='medium'), right=Side(style='medium'),
        top=Side(style='medium'), bottom=Side(style='medium')
    )
    
    center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    days = ['Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday']
    
    # Total columns: Day label + P1-P5 + BREAK + P6-P8 = 1 + 5 + 1 + 3 = 10
    total_cols = 1 + 5 + 1 + (periods_per_day - 5)
    
    # A4 Landscape: 297mm x 210mm
    # Safe margins for most printers: 0.5" (12.7mm) each side
    # Usable: ~272mm x 185mm
    total_page_height = 500  # points for A4 landscape height (conservative)
    school_header_height = 38
    address_header_height = 20 if school_address else 0
    class_title_height = 32
    period_header_height = 45
    # 5 day rows get ALL remaining space
    fixed_height = school_header_height + address_header_height + class_title_height + period_header_height
    day_row_height = (total_page_height - fixed_height) / 5
    
    for key in sorted(timetables.keys()):
        tt = timetables[key]
        ws = wb.create_sheet(title=f"{tt['class_name']} {tt['arm_name']}"[:31])
        
        # Page setup for A4 landscape - fit on ONE page
        ws.page_setup.paperSize = ws.PAPERSIZE_A4
        ws.page_setup.orientation = 'landscape'
        ws.page_setup.fitToPage = True
        ws.page_setup.fitToWidth = 1
        ws.page_setup.fitToHeight = 1
        ws.print_options.horizontalCentered = True
        ws.print_options.verticalCentered = True
        
        # Safe margins for most printers (0.5 inch = 12.7mm)
        ws.page_margins.left = 0.5
        ws.page_margins.right = 0.5
        ws.page_margins.top = 0.4
        ws.page_margins.bottom = 0.4
        ws.page_margins.header = 0
        ws.page_margins.footer = 0
        
        current_row = 1
        
        # School name
        ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=total_cols)
        cell = ws.cell(row=current_row, column=1, value=school_name.upper())
        cell.font = school_font
        cell.alignment = center_align
        ws.row_dimensions[current_row].height = school_header_height
        current_row += 1
        
        # School address
        if school_address:
            ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=total_cols)
            cell = ws.cell(row=current_row, column=1, value=school_address)
            cell.font = address_font
            cell.alignment = center_align
            ws.row_dimensions[current_row].height = address_header_height
            current_row += 1
        
        # Class title
        ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=total_cols)
        cell = ws.cell(row=current_row, column=1, value=f"{tt['class_name']} {tt['arm_name']} - TIMETABLE")
        cell.font = class_title_font
        cell.alignment = center_align
        cell.border = thick_border
        for col in range(1, total_cols + 1):
            ws.cell(row=current_row, column=col).border = thick_border
        ws.row_dimensions[current_row].height = class_title_height
        current_row += 1
        
        # Period header row (times)
        col = 1
        cell = ws.cell(row=current_row, column=col, value="Day")
        cell.font = header_font
        cell.border = thick_border
        cell.alignment = center_align
        col += 1
        
        # P1-P5 headers
        for i in range(5):
            cell = ws.cell(row=current_row, column=col, value=period_times[i])
            cell.font = header_font
            cell.border = thick_border
            cell.alignment = center_align
            col += 1
        
        # Break header
        cell = ws.cell(row=current_row, column=col, value=break_time)
        cell.font = break_header_font
        cell.border = thick_border
        cell.alignment = center_align
        col += 1
        
        # P6-P8 headers
        for i in range(5, periods_per_day):
            cell = ws.cell(row=current_row, column=col, value=period_times[i])
            cell.font = header_font
            cell.border = thick_border
            cell.alignment = center_align
            col += 1
        
        ws.row_dimensions[current_row].height = period_header_height
        current_row += 1
        
        # Data rows (one per day)
        for day_idx, day_name in enumerate(days):
            col = 1
            
            # Day name
            cell = ws.cell(row=current_row, column=col, value=day_name)
            cell.font = day_font
            cell.border = thick_border
            cell.alignment = center_align
            col += 1
            
            # P1-P5
            for p in range(1, 6):
                entry = tt['grid'][day_idx].get(p)
                value = ""
                if entry and entry.subject:
                    subj_name = entry.subject.name
                    value = abbrev_map.get(subj_name, subj_name[:6])
                
                cell = ws.cell(row=current_row, column=col, value=value)
                cell.font = cell_font
                cell.border = thin_border
                cell.alignment = center_align
                col += 1
            
            # Break column - empty with border
            cell = ws.cell(row=current_row, column=col, value="")
            cell.border = thin_border
            col += 1
            
            # P6-P8
            for p in range(6, periods_per_day + 1):
                entry = tt['grid'][day_idx].get(p)
                value = ""
                if entry and entry.subject:
                    subj_name = entry.subject.name
                    value = abbrev_map.get(subj_name, subj_name[:6])
                
                cell = ws.cell(row=current_row, column=col, value=value)
                cell.font = cell_font
                cell.border = thin_border
                cell.alignment = center_align
                col += 1
            
            ws.row_dimensions[current_row].height = day_row_height
            current_row += 1
        
        # Column widths - fit within safe printable area
        # A4 landscape with 0.5" margins: ~250mm usable width
        ws.column_dimensions['A'].width = 12  # Day column
        period_col_width = 28  # Period columns
        break_col_width = 8   # Break column
        
        for col in range(2, total_cols + 1):
            if col == 7:  # Break column
                ws.column_dimensions[get_column_letter(col)].width = break_col_width
            else:
                ws.column_dimensions[get_column_letter(col)].width = period_col_width
    
    return xlsx_response(wb, f'timetables_{batch_id}.xlsx')


@generator_bp.route('/results/<batch_id>/export_by_day')
@login_required
def export_results_by_day(batch_id):
    """Export timetable in day-wise format - MAXIMUM SIZE for A4 landscape printing, NO COLORS"""
    import openpyxl
    from openpyxl.styles import Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    
    results = GenTimetableResult.query.filter_by(batch_id=batch_id, branch_id=gen_bid()).all()
    if not results:
        flash('No results.', 'error')
        return redirect(url_for('generator.results_list'))
    
    # Determine school level from results
    school_level = results[0].school_level if results else 'sss'
    
    rules = {r.rule_type: r.value for r in GenTimetableRule.query.filter_by(is_active=True, school_level=school_level, branch_id=gen_bid()).all()}
    periods_per_day = int(rules.get('periods_per_day', 8))
    
    # Get school info
    school_name = GenSettings.get('school_name', 'School')
    school_address = GenSettings.get('school_address', '')
    
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    
    days = ['Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday']
    
    # Period time slots (8:20 AM start, 40 min each, 30 min break after period 5)
    period_times_before_break = []  # P1-P5
    period_times_after_break = []   # P6-P8
    break_time = ""
    
    start_hour, start_min = 8, 20
    for p in range(1, periods_per_day + 1):
        start_total = start_hour * 60 + start_min
        end_total = start_total + 40
        start_h, start_m = start_total // 60, start_total % 60
        end_h, end_m = end_total // 60, end_total % 60
        start_str = f"{start_h}:{start_m:02d}"
        end_str = f"{end_h}:{end_m:02d}"
        
        if p <= 5:
            period_times_before_break.append(f"P{p}\n{start_str}-{end_str}")
        else:
            period_times_after_break.append(f"P{p}\n{start_str}-{end_str}")
        
        start_hour, start_min = end_h, end_m
        
        # Capture break time after period 5
        if p == 5:
            break_start = f"{end_h}:{end_m:02d}"
            start_total = start_hour * 60 + start_min + 30
            start_hour, start_min = start_total // 60, start_total % 60
            break_end = f"{start_hour}:{start_min:02d}"
            break_time = f"BREAK\n{break_start}\n-\n{break_end}"
    
    class_arms = sorted(set((r.class_name, r.arm_name) for r in results))
    
    def get_short_code(class_name, arm):
        if class_name.startswith('JSS'):
            class_num = class_name.replace('JSS', '')
            arm_letter = arm[0].upper()
            return f"J{class_num}{arm_letter}"
        else:
            class_num = class_name.replace('SSS', '')
            arm_letter = arm[0].upper()
            return f"{class_num}{arm_letter}"
    
    abbrev_map = {
        'Mathematics': 'Maths',
        'English Language': 'Eng',
        'Physics': 'Phy',
        'Chemistry': 'Chem',
        'Biology': 'Bio',
        'Economics': 'Econs',
        'Government': 'Govt',
        'Literature in English': 'Lit',
        'Agricultural Science': 'Agric',
        'Christian Religious Studies': 'CRS',
        'Civic Education': 'Civic',
        'Computer Studies': 'Comp',
        'Commerce': 'Comm',
        'Geography': 'Geo',
        'Further Mathematics': 'F/Mth',
        'Livestock Farming': 'Livst',
        'History': 'Hist',
        'Phonics': 'Phon',
        # JSS subjects
        'Basic Science': 'B.Sci',
        'Basic Technology': 'B.Tech',
        'Social Studies': 'Soc',
        'Home Economics': 'H/Eco',
        'Business Studies': 'Bus',
        'Physical Health Education': 'PHE',
        'Islamic Religious Studies': 'IRS',
        'French': 'Fren',
        'Yoruba': 'Yor',
        'Igbo': 'Igbo',
        'Hausa': 'Hau',
        'Music': 'Music',
        'Fine Art': 'Art',
        'Data Processing': 'D.Pro',
        'Marketing': 'Mkt',
        'Accounting': 'Acct',
        'Food and Nutrition': 'F&N',
    }
    
    # Build a lookup dict for subjects
    subject_lookup = {s.id: s for s in GenSubject.query.filter_by(is_active=True, school_level=school_level, branch_id=gen_bid()).all()}
    
    # NO COLORS - Pure black text on white for B&W printing
    school_font = Font(bold=True, size=24, color='000000')
    address_font = Font(bold=False, size=12, color='000000')
    day_font = Font(bold=True, size=32, color='000000')
    header_font = Font(bold=True, size=14, color='000000')
    class_font = Font(bold=True, size=20, color='000000')
    cell_font = Font(bold=True, size=24, color='000000')
    break_header_font = Font(bold=True, size=9, color='000000')
    
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    thick_border = Border(
        left=Side(style='medium'),
        right=Side(style='medium'),
        top=Side(style='medium'),
        bottom=Side(style='medium')
    )
    
    center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    # Total columns: Class + P1-P5 + BREAK + P6-P8 = 1 + 5 + 1 + 3 = 10
    total_cols = 1 + 5 + 1 + (periods_per_day - 5)
    
    # Calculate heights (conservative for safe printing)
    num_data_rows = len(class_arms)
    school_header_height = 35
    address_header_height = 20
    day_header_height = 45
    period_header_height = 45
    total_page_height = 500  # Conservative for safe printing
    
    for d, day_name in enumerate(days):
        ws = wb.create_sheet(title=day_name)
        
        ws.page_setup.paperSize = ws.PAPERSIZE_A4
        ws.page_setup.orientation = 'landscape'
        ws.page_setup.fitToPage = True
        ws.page_setup.fitToWidth = 1
        ws.page_setup.fitToHeight = 1
        ws.print_options.horizontalCentered = True
        ws.print_options.verticalCentered = True
        
        # Safe margins for most printers (0.5 inch)
        ws.page_margins.left = 0.5
        ws.page_margins.right = 0.5
        ws.page_margins.top = 0.4
        ws.page_margins.bottom = 0.4
        ws.page_margins.header = 0
        ws.page_margins.footer = 0
        
        current_row = 1
        
        # School name and address ONLY on Monday
        if d == 0:
            # School name header
            ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=total_cols)
            cell = ws.cell(row=current_row, column=1, value=school_name.upper())
            cell.font = school_font
            cell.alignment = center_align
            ws.row_dimensions[current_row].height = school_header_height
            current_row += 1
            
            if school_address:
                ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=total_cols)
                cell = ws.cell(row=current_row, column=1, value=school_address)
                cell.font = address_font
                cell.alignment = center_align
                ws.row_dimensions[current_row].height = address_header_height
                current_row += 1
            
            # Recalculate data row height for Monday
            header_rows_height = school_header_height + (address_header_height if school_address else 0) + day_header_height + period_header_height
            remaining_height = total_page_height - header_rows_height
            data_row_height = remaining_height / num_data_rows
            data_row_height = max(data_row_height, 28)
        else:
            # Other days - just day header and period header
            remaining_height = total_page_height - day_header_height - period_header_height
            data_row_height = remaining_height / num_data_rows
            data_row_height = max(data_row_height, 30)
        
        # Day header
        ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=total_cols)
        cell = ws.cell(row=current_row, column=1, value=day_name.upper())
        cell.font = day_font
        cell.alignment = center_align
        cell.border = thick_border
        for col in range(1, total_cols + 1):
            ws.cell(row=current_row, column=col).border = thick_border
        ws.row_dimensions[current_row].height = day_header_height
        current_row += 1
        
        # Period headers with BREAK column
        col = 1
        # Class column
        cell = ws.cell(row=current_row, column=col, value="Class")
        cell.font = header_font
        cell.border = thick_border
        cell.alignment = center_align
        col += 1
        
        # P1-P5
        for i, p in enumerate(range(1, 6)):
            header_value = period_times_before_break[i] if d == 0 else f"P{p}"
            cell = ws.cell(row=current_row, column=col, value=header_value)
            cell.font = header_font
            cell.border = thick_border
            cell.alignment = center_align
            col += 1
        
        # BREAK column header
        cell = ws.cell(row=current_row, column=col, value=break_time if d == 0 else "BREAK")
        cell.font = break_header_font
        cell.border = thick_border
        cell.alignment = center_align
        col += 1
        
        # P6-P8
        for i, p in enumerate(range(6, periods_per_day + 1)):
            header_value = period_times_after_break[i] if d == 0 else f"P{p}"
            cell = ws.cell(row=current_row, column=col, value=header_value)
            cell.font = header_font
            cell.border = thick_border
            cell.alignment = center_align
            col += 1
        
        ws.row_dimensions[current_row].height = period_header_height
        current_row += 1
        
        # Data rows
        for class_name, arm in class_arms:
            short_code = get_short_code(class_name, arm)
            col = 1
            
            # Class code
            cell = ws.cell(row=current_row, column=col, value=short_code)
            cell.font = class_font
            cell.border = thick_border
            cell.alignment = center_align
            col += 1
            
            arm_results = [r for r in results if r.class_name == class_name and r.arm_name == arm and r.day_of_week == d]
            
            # P1-P5
            for p in range(1, 6):
                slot_result = next((r for r in arm_results if r.period_number == p), None)
                value = ""
                if slot_result and slot_result.subject_id:
                    subj = subject_lookup.get(slot_result.subject_id)
                    if subj:
                        subj_name = subj.name
                        value = abbrev_map.get(subj_name, subj_name[:5])
                
                cell = ws.cell(row=current_row, column=col, value=value)
                cell.font = cell_font
                cell.border = thin_border
                cell.alignment = center_align
                col += 1
            
            # BREAK column - empty with border
            cell = ws.cell(row=current_row, column=col, value="")
            cell.border = thin_border
            cell.alignment = center_align
            col += 1
            
            # P6-P8
            for p in range(6, periods_per_day + 1):
                slot_result = next((r for r in arm_results if r.period_number == p), None)
                value = ""
                if slot_result and slot_result.subject_id:
                    subj = subject_lookup.get(slot_result.subject_id)
                    if subj:
                        subj_name = subj.name
                        value = abbrev_map.get(subj_name, subj_name[:5])
                
                cell = ws.cell(row=current_row, column=col, value=value)
                cell.font = cell_font
                cell.border = thin_border
                cell.alignment = center_align
                col += 1
            
            ws.row_dimensions[current_row].height = data_row_height
            current_row += 1
        
        # Column widths - fit within safe printable area
        ws.column_dimensions['A'].width = 8  # Class
        col_width = 26 if periods_per_day <= 8 else 22
        for col in range(2, total_cols + 1):
            if col == 7:  # Break column
                ws.column_dimensions[get_column_letter(col)].width = 8
            else:
                ws.column_dimensions[get_column_letter(col)].width = col_width
    
    return xlsx_response(wb, f'timetables_by_day_{batch_id}.xlsx')


@generator_bp.route('/results/<batch_id>/export_by_day_pdf')
@login_required
def export_results_by_day_pdf(batch_id):
    """Export timetable as PDF - Each day fits EXACTLY on one A4 landscape page - NO COLORS"""
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.units import mm
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, PageBreak
    
    results = GenTimetableResult.query.filter_by(batch_id=batch_id, branch_id=gen_bid()).all()
    if not results:
        flash('No results.', 'error')
        return redirect(url_for('generator.results_list'))
    
    rules = {r.rule_type: r.value for r in GenTimetableRule.query.filter_by(is_active=True, branch_id=gen_bid()).all()}
    periods_per_day = int(rules.get('periods_per_day', 8))
    
    # Get school info
    school_name = GenSettings.get('school_name', 'School')
    school_address = GenSettings.get('school_address', '')
    
    days = ['Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday']
    
    # Period times with break tracking
    period_times_before = []  # P1-P5
    period_times_after = []   # P6-P8
    break_time = ""
    
    start_hour, start_min = 8, 20
    for p in range(1, periods_per_day + 1):
        start_total = start_hour * 60 + start_min
        end_total = start_total + 40
        start_h, start_m = start_total // 60, start_total % 60
        end_h, end_m = end_total // 60, end_total % 60
        
        time_str = f"P{p}\n{start_h}:{start_m:02d}\n{end_h}:{end_m:02d}"
        if p <= 5:
            period_times_before.append(time_str)
        else:
            period_times_after.append(time_str)
        
        start_hour, start_min = end_h, end_m
        if p == 5:
            break_start = f"{end_h}:{end_m:02d}"
            start_total = start_hour * 60 + start_min + 30
            start_hour, start_min = start_total // 60, start_total % 60
            break_end = f"{start_hour}:{start_min:02d}"
            break_time = f"BREAK\n{break_start}\n-\n{break_end}"
    
    class_arms = sorted(set((r.class_name, r.arm_name) for r in results))
    num_data_rows = len(class_arms)
    
    def get_short_code(class_name, arm):
        class_num = class_name.replace('SSS', '')
        arm_letter = arm[0].upper()
        return f"{class_num}{arm_letter}"
    
    abbrev_map = {
        'Mathematics': 'Maths', 'English Language': 'Eng', 'Physics': 'Phy',
        'Chemistry': 'Chem', 'Biology': 'Bio', 'Economics': 'Econs',
        'Government': 'Govt', 'Literature in English': 'Lit', 'Agricultural Science': 'Agric',
        'Christian Religious Studies': 'CRS', 'Civic Education': 'Civic',
        'Computer Studies': 'Comp', 'Commerce': 'Comm', 'Geography': 'Geo',
        'Further Mathematics': 'F/Mth', 'Livestock Farming': 'Livst',
        'History': 'Hist', 'Phonics': 'Phon',
    }
    
    output = BytesIO()
    
    # A4 landscape: 297mm x 210mm - safe margins for printers
    margin = 10*mm
    page_w, page_h = landscape(A4)
    usable_width = page_w - 2*margin
    usable_height = page_h - 2*margin
    
    doc = SimpleDocTemplate(
        output,
        pagesize=landscape(A4),
        leftMargin=margin,
        rightMargin=margin,
        topMargin=margin,
        bottomMargin=margin
    )
    
    elements = []
    
    # Total columns: Class + P1-P5 + BREAK + P6-P8
    num_periods_before = 5
    num_periods_after = periods_per_day - 5
    total_cols = 1 + num_periods_before + 1 + num_periods_after
    
    # Column widths - fill entire width
    first_col_width = 12*mm
    break_col_width = 10*mm
    remaining_width = usable_width - first_col_width - break_col_width
    period_col_width = remaining_width / periods_per_day
    
    col_widths = [first_col_width]
    col_widths += [period_col_width] * num_periods_before
    col_widths += [break_col_width]
    col_widths += [period_col_width] * num_periods_after
    
    for d, day_name in enumerate(days):
        table_data = []
        row_heights = []
        
        # Calculate row heights to fit EXACTLY on one page
        if d == 0:  # Monday - include school name and address
            school_header_height = 8*mm
            address_header_height = 4*mm if school_address else 0
            day_header_height = 10*mm
            period_header_height = 12*mm
            fixed_height = school_header_height + address_header_height + day_header_height + period_header_height
        else:
            day_header_height = 10*mm
            period_header_height = 10*mm
            fixed_height = day_header_height + period_header_height
        
        # Data rows get remaining height divided equally
        available_for_data = usable_height - fixed_height - 2*mm
        data_row_height = available_for_data / num_data_rows
        
        # Build table data
        if d == 0:
            # School name row
            school_row = [school_name.upper()] + [''] * (total_cols - 1)
            table_data.append(school_row)
            row_heights.append(school_header_height)
            
            if school_address:
                address_row = [school_address] + [''] * (total_cols - 1)
                table_data.append(address_row)
                row_heights.append(address_header_height)
        
        # Day header row
        day_row = [day_name.upper()] + [''] * (total_cols - 1)
        table_data.append(day_row)
        row_heights.append(day_header_height)
        
        # Period header row
        if d == 0:  # Monday - show times
            header_row = ['Class'] + period_times_before + [break_time] + period_times_after
        else:
            header_row = ['Class'] + [f'P{p}' for p in range(1, 6)] + ['BREAK'] + [f'P{p}' for p in range(6, periods_per_day + 1)]
        table_data.append(header_row)
        row_heights.append(period_header_height)
        
        # Data rows
        for class_name, arm in class_arms:
            short_code = get_short_code(class_name, arm)
            row = [short_code]
            
            arm_results = [r for r in results if r.class_name == class_name and r.arm_name == arm and r.day_of_week == d]
            
            # P1-P5
            for p in range(1, 6):
                slot_result = next((r for r in arm_results if r.period_number == p), None)
                if slot_result and slot_result.subject:
                    row.append(abbrev_map.get(slot_result.subject.name, slot_result.subject.name[:5]))
                else:
                    row.append('')
            
            # BREAK column - empty
            row.append('')
            
            # P6-P8
            for p in range(6, periods_per_day + 1):
                slot_result = next((r for r in arm_results if r.period_number == p), None)
                if slot_result and slot_result.subject:
                    row.append(abbrev_map.get(slot_result.subject.name, slot_result.subject.name[:5]))
                else:
                    row.append('')
            
            table_data.append(row)
            row_heights.append(data_row_height)
        
        # Create table with exact dimensions
        table = Table(table_data, colWidths=col_widths, rowHeights=row_heights)
        
        # Break column index
        break_col = 6
        
        # Calculate row indices
        if d == 0:
            if school_address:
                day_row_idx = 2
                header_row_idx = 3
                data_start_row = 4
            else:
                day_row_idx = 1
                header_row_idx = 2
                data_start_row = 3
        else:
            day_row_idx = 0
            header_row_idx = 1
            data_start_row = 2
        
        # NO COLORS - just black text on white, with borders
        style_commands = [
            # Day header row
            ('SPAN', (0, day_row_idx), (-1, day_row_idx)),
            ('TEXTCOLOR', (0, day_row_idx), (-1, day_row_idx), colors.black),
            ('FONTNAME', (0, day_row_idx), (-1, day_row_idx), 'Helvetica-Bold'),
            ('FONTSIZE', (0, day_row_idx), (-1, day_row_idx), 22),
            ('ALIGN', (0, day_row_idx), (-1, day_row_idx), 'CENTER'),
            ('VALIGN', (0, day_row_idx), (-1, day_row_idx), 'MIDDLE'),
            
            # Period header row
            ('FONTNAME', (0, header_row_idx), (-1, header_row_idx), 'Helvetica-Bold'),
            ('FONTSIZE', (0, header_row_idx), (-1, header_row_idx), 7 if d == 0 else 12),
            ('ALIGN', (0, header_row_idx), (-1, header_row_idx), 'CENTER'),
            ('VALIGN', (0, header_row_idx), (-1, header_row_idx), 'MIDDLE'),
            
            # Break column header
            ('FONTSIZE', (break_col, header_row_idx), (break_col, header_row_idx), 6),
            
            # Class codes column
            ('FONTNAME', (0, data_start_row), (0, -1), 'Helvetica-Bold'),
            ('FONTSIZE', (0, data_start_row), (0, -1), 14),
            
            # Subject cells - LARGE
            ('FONTNAME', (1, data_start_row), (-1, -1), 'Helvetica-Bold'),
            ('FONTSIZE', (1, data_start_row), (-1, -1), 18),
            
            # All cells alignment
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            
            # Borders only - no background colors
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('BOX', (0, 0), (-1, -1), 1.5, colors.black),
        ]
        
        # Add school name styling for Monday
        if d == 0:
            style_commands.extend([
                ('SPAN', (0, 0), (-1, 0)),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 16),
                ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
                ('VALIGN', (0, 0), (-1, 0), 'MIDDLE'),
            ])
            if school_address:
                style_commands.extend([
                    ('SPAN', (0, 1), (-1, 1)),
                    ('FONTNAME', (0, 1), (-1, 1), 'Helvetica'),
                    ('FONTSIZE', (0, 1), (-1, 1), 8),
                    ('TEXTCOLOR', (0, 1), (-1, 1), colors.black),
                    ('ALIGN', (0, 1), (-1, 1), 'CENTER'),
                    ('VALIGN', (0, 1), (-1, 1), 'MIDDLE'),
                ])
        
        table.setStyle(TableStyle(style_commands))
        elements.append(table)
        
        if d < len(days) - 1:
            elements.append(PageBreak())
    
    doc.build(elements)
    return pdf_response(output, f'timetables_by_day_{batch_id}.pdf', inline=False)


@generator_bp.route('/results/<batch_id>/export_image')
@login_required
def export_image(batch_id):
    """Export timetable as PNG image (Ultra HD - 8x scale)"""
    from routes.generator_image import generate_timetable_image, image_to_response
    
    quality = request.args.get('quality', 'ultra')  # 'hd' or 'ultra'
    img = generate_timetable_image(batch_id, quality=quality)
    if not img:
        flash('No results found.', 'error')
        return redirect(url_for('generator.results_list'))
    
    quality_suffix = '_hd' if quality == 'hd' else '_ultrahd'
    return image_to_response(img, f'timetable_{batch_id}{quality_suffix}.png')


@generator_bp.route('/results/<batch_id>/export_image_hd')
@login_required
def export_image_hd(batch_id):
    """Export timetable as PNG image (HD - 4x scale)"""
    from routes.generator_image import generate_timetable_image, image_to_response
    
    img = generate_timetable_image(batch_id, quality='hd')
    if not img:
        flash('No results found.', 'error')
        return redirect(url_for('generator.results_list'))
    
    return image_to_response(img, f'timetable_{batch_id}_hd.png')


@generator_bp.route('/results/<batch_id>/export_image_ultra')
@login_required
def export_image_ultra(batch_id):
    """Export timetable as PNG image (Ultra HD - 8x scale)"""
    from routes.generator_image import generate_timetable_image, image_to_response
    
    img = generate_timetable_image(batch_id, quality='ultra')
    if not img:
        flash('No results found.', 'error')
        return redirect(url_for('generator.results_list'))
    
    return image_to_response(img, f'timetable_{batch_id}_ultrahd.png')


@generator_bp.route('/results/<batch_id>/teacher/<int:teacher_id>/image')
@login_required
def export_teacher_image(batch_id, teacher_id):
    """Export individual teacher timetable as PNG image"""
    from routes.generator_image import generate_teacher_timetable_image, image_to_response
    
    img = generate_teacher_timetable_image(batch_id, teacher_id)
    if not img:
        flash('No results found for this teacher.', 'error')
        return redirect(url_for('generator.teacher_timetable', batch_id=batch_id))
    
    teacher = GenTeacher.query.get(teacher_id)
    filename = f'timetable_{teacher.name.replace(" ", "_")}_{batch_id}.png'
    
    return image_to_response(img, filename)


@generator_bp.route('/results/<batch_id>/teacher/<int:teacher_id>/print')
@login_required
def print_teacher_timetable(batch_id, teacher_id):
    """Printable view for individual teacher timetable"""
    teacher = gen_owned_or_404(GenTeacher, teacher_id)
    results = GenTimetableResult.query.filter_by(batch_id=batch_id, teacher_id=teacher_id, branch_id=gen_bid()).all()
    
    if not results:
        flash('No timetable found for this teacher.', 'error')
        return redirect(url_for('generator.teacher_timetable', batch_id=batch_id))
    
    rules = {r.rule_type: r.value for r in GenTimetableRule.query.filter_by(is_active=True, branch_id=gen_bid()).all()}
    periods_per_day = int(rules.get('periods_per_day', 8))
    
    # Build timetable grid
    timetable = {d: {p: None for p in range(1, periods_per_day + 1)} for d in range(5)}
    for r in results:
        timetable[r.day_of_week][r.period_number] = {
            'subject': r.subject,
            'class_name': r.class_name,
            'arm_name': r.arm_name
        }
    
    return render_template('generator/print_teacher_timetable.html',
        teacher=teacher,
        timetable=timetable,
        batch_id=batch_id,
        periods=range(1, periods_per_day + 1),
        days=DAYS_OF_WEEK
    )
