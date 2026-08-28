"""
Mock JAMB Examination Routes
Full management of mock JAMB exams with analytics and insights
"""
from flask import (Blueprint, request, redirect, url_for, flash, jsonify, Response,
                   render_template, current_app)
from utils.helpers import get_active_term, get_active_session
from datetime import datetime
from io import BytesIO
import os
import secrets
from sqlalchemy import func
from utils.web_exports import xlsx_response

from models import (db, Student, AcademicSession, StudentEnrollment, ClassArmAssignment,
                    SchoolClass, Subject, MockJAMBPassage, MockJAMBQuestion)
from models.mock_jamb import MockJAMBExam, MockJAMBResult, MockJAMBAnalytics
from utils.helpers import login_required, WAEC_SUBJECTS, get_sss3_students, student_subject_map
from utils.branch_scope import require_branch_access, branch_for_new, scope_query
from utils.csrf import csrf_protect
from utils.jamb_config import (
    convert_correct_to_100, question_count_map, COMPULSORY_SUBJECT,
    MAX_TOTAL_SCORE,
)
from utils.search import like_term

mock_jamb_bp = Blueprint('mock_jamb', __name__, url_prefix='/mock-jamb')

# Student-facing online sitting shares the CBT student portal login (student ID +
# portal password) but delivers the JAMB 4-subjects-in-one sitting.
mock_jamb_portal_bp = Blueprint('mock_jamb_portal', __name__, url_prefix='/exam/mock-jamb')


@mock_jamb_bp.before_request
@mock_jamb_portal_bp.before_request
def _ensure_schema():
    """Self-heal the tenant DB's Mock JAMB columns (once per engine), so a DB that
    is behind on Alembic migrations doesn't 500 on pages that load a mock exam."""
    try:
        from utils.mock_jamb_schema import ensure_mock_jamb_schema
        ensure_mock_jamb_schema()
    except Exception:
        pass


def _read_subject_scores(form, suffix=''):
    """
    Read the four subject rows from a submitted form and return a normalised
    list of ``{'name', 'score'}`` dicts.

    Each row may supply either ``subjectN_correct`` (raw number of correct
    answers, which is converted to a score over 100) or, for backwards
    compatibility, an already-computed ``subjectN_score``.
    """
    def safe_int(val):
        try:
            return int(val) if val not in (None, '') else None
        except (TypeError, ValueError):
            return None

    rows = []
    for n in range(1, 5):
        name = (form.get(f'subject{n}{suffix}') or '').strip() or None
        correct = form.get(f'subject{n}_correct{suffix}')
        if correct not in (None, ''):
            score = convert_correct_to_100(name, correct)
        else:
            score = safe_int(form.get(f'subject{n}_score{suffix}'))
        if name is None:
            score = None
        rows.append({'name': name, 'score': score})
    return rows


# =============================================================================
# DASHBOARD & OVERVIEW
# =============================================================================

from utils.spa import section_responders
_wants_json, _render, _ok, _err = section_responders(
    'mock_jamb/app.html', 'mj_json', 'mock_jamb.index')


@mock_jamb_bp.route('/')
@login_required
def index():
    """Mock JAMB main dashboard"""
    active_session = get_active_session()
    sessions = AcademicSession.query.order_by(AcademicSession.name.desc()).all()
    
    session_id = request.args.get('session_id', type=int)
    if not session_id and active_session:
        session_id = active_session.id
    
    exams = []
    comparison_data = None
    
    if session_id:
        exams = scope_query(MockJAMBExam.query.filter_by(session_id=session_id),
                            MockJAMBExam).order_by(MockJAMBExam.exam_number).all()
        comparison_data = MockJAMBAnalytics.compare_mock_exams(session_id)
    
    # A derived SSS3-arm teacher sees only their own arm's figures on each exam.
    from utils.access_control import exam_student_scope
    _scope = exam_student_scope()

    def exam_row(e):
        rows = [r for r in e.results if _scope is None or r.student_id in _scope]
        scored = [r.total_score for r in rows if r.total_score is not None]
        above_200 = sum(1 for v in scored if v >= 200)
        return {
            'id': e.id, 'display_name': e.display_name, 'is_completed': bool(e.is_completed),
            'exam_date': e.exam_date.strftime('%d %B %Y') if e.exam_date else '',
            'student_count': len(rows),
            'average_score': round(sum(scored) / len(scored), 1) if scored else None,
            'above_200': above_200,
            'view_url': url_for('mock_jamb.view_exam', exam_id=e.id),
            'add_url': url_for('mock_jamb.add_result', exam_id=e.id),
            'bulk_url': url_for('mock_jamb.bulk_entry', exam_id=e.id),
            'deep_url': url_for('mock_jamb.deep', exam_id=e.id),
        }
    _rows = [exam_row(e) for e in exams]
    avg_scores = [r['average_score'] for r in _rows if r['average_score'] is not None]
    return _render({
        'page': 'index', 'selected_session_id': session_id or '', 'derived': bool(_scope is not None),
        'sessions': [{'id': s.id, 'name': s.name} for s in sessions],
        'exams': _rows,
        'stats': {'count': len(exams), 'total_results': sum(r['student_count'] for r in _rows),
                  'avg_score': round(sum(avg_scores) / len(avg_scores), 1) if avg_scores else None,
                  'remaining': 4 - len(exams)},
        # No cohort comparison for a derived (arm-only) teacher.
        'comparison': [] if _scope is not None else
                      [{'label': c['exam'].display_name, 'average': round(c['average'], 1),
                        'above_250': round((c.get('above_250_pct') or 0) * c['student_count'] / 100)}
                       for c in (comparison_data or [])],
        'urls': {'create': url_for('mock_jamb.create_exam'), 'analytics': url_for('mock_jamb.analytics'),
                 'predictions': url_for('results.predictions_dashboard'), 'validation': url_for('mock_jamb.validation'),
                 'trends': url_for('mock_jamb.trends'), 'bank': url_for('mock_jamb.bank'),
                 'self': url_for('mock_jamb.index')},
    })


# =============================================================================
# EXAM MANAGEMENT
# =============================================================================

@mock_jamb_bp.route('/exam/create', methods=['GET', 'POST'])
@login_required
@csrf_protect
def create_exam():
    """Create a new mock JAMB exam"""
    sessions = AcademicSession.query.order_by(AcademicSession.name.desc()).all()
    
    if request.method == 'POST':
        try:
            session_id = request.form.get('session_id', type=int)
            exam_number = request.form.get('exam_number', type=int)
            exam_date = request.form.get('exam_date')
            name = request.form.get('name', '').strip()
            description = request.form.get('description', '').strip()
            
            # Validate
            if not session_id or not exam_number or not exam_date:
                return _err('Please fill all required fields.', url_for('mock_jamb.create_exam'))

            # The new exam belongs to the creator's branch; uniqueness of the
            # exam number is per (session, branch).
            new_branch_id = branch_for_new(request.form.get('branch_id', type=int))
            existing = MockJAMBExam.query.filter_by(
                session_id=session_id, exam_number=exam_number, branch_id=new_branch_id).first()
            if existing:
                return _err(f'Mock exam #{exam_number} already exists for this session.', url_for('mock_jamb.create_exam'))

            # Parse date
            try:
                exam_date = datetime.strptime(exam_date, '%Y-%m-%d').date()
            except ValueError:
                return _err('Invalid date format.', url_for('mock_jamb.create_exam'))
            
            # Create exam
            session_obj = db.session.get(AcademicSession, session_id)
            if not name:
                ordinals = {1: 'First', 2: 'Second', 3: 'Third', 4: 'Fourth'}
                name = f"{ordinals.get(exam_number, str(exam_number))} Mock JAMB {session_obj.name}"
            
            exam = MockJAMBExam(
                name=name,
                exam_number=exam_number,
                session_id=session_id,
                exam_date=exam_date,
                description=description,
                branch_id=new_branch_id
            )
            
            db.session.add(exam)
            db.session.commit()
            return _ok(f'{exam.display_name} created successfully!',
                       url_for('mock_jamb.index', session_id=session_id))

        except Exception as e:
            db.session.rollback()
            return _err(f'Error creating exam: {str(e)}', url_for('mock_jamb.create_exam'))

    # Get existing exam numbers for each session
    existing_exams = {}
    for sess in sessions:
        existing_exams[str(sess.id)] = [e.exam_number for e in MockJAMBExam.query.filter_by(session_id=sess.id).all()]

    return _render({
        'page': 'create_exam', 'existing_exams': existing_exams,
        'sessions': [{'id': s.id, 'name': s.name} for s in sessions],
        'submit_url': url_for('mock_jamb.create_exam'),
        'urls': {'index': url_for('mock_jamb.index')},
    })


@mock_jamb_bp.route('/exam/<int:exam_id>/subject/<subject>')
@login_required
def subject_analysis(exam_id, subject):
    """Full per-subject analysis within a Mock JAMB exam — score-band
    distribution, average, ≥50/≥70 rates and the ranked candidate list."""
    from utils.analytics_service import AcademicAnalytics
    exam = db.get_or_404(MockJAMBExam, exam_id)
    require_branch_access(exam.branch_id)
    pairs = []
    for r in MockJAMBResult.query.filter_by(mock_exam_id=exam_id).all():
        for i in (1, 2, 3, 4):
            if getattr(r, f'subject{i}') == subject and getattr(r, f'subject{i}_score') is not None:
                pairs.append((r.student_id, getattr(r, f'subject{i}_score')))
    data = AcademicAnalytics._score_subject_payload(subject, pairs)
    return render_template('results/score_subject.html', subject=subject, data=data,
                           years=[], selected_year=None, exam_label='Mock JAMB',
                           back_url=url_for('mock_jamb.view_exam', exam_id=exam_id),
                           search_endpoint=None)


@mock_jamb_bp.route('/exam/<int:exam_id>')
@login_required
def view_exam(exam_id):
    """View a specific mock exam with detailed statistics"""
    exam = db.get_or_404(MockJAMBExam, exam_id)
    require_branch_access(exam.branch_id)   # no cross-branch exam data
    # finalise any timed-out-but-unsubmitted sittings so their scores show here
    from utils.mock_jamb_sitting import auto_submit_expired
    auto_submit_expired(exam=exam)
    statistics = MockJAMBAnalytics.get_exam_statistics(exam_id)
    
    # Get sort and filter parameters
    sort_by = request.args.get('sort', 'score')
    sort_order = request.args.get('order', 'desc')
    min_score = request.args.get('min_score', type=int)
    max_score = request.args.get('max_score', type=int)
    search = request.args.get('search', '').strip()
    
    # Build query for results
    query = MockJAMBResult.query.filter_by(mock_exam_id=exam_id).join(Student)
    
    # Apply search filter
    if search:
        search_term = like_term(search)
        query = query.filter(
            db.or_(
                Student.first_name.ilike(search_term, escape='\\'),
                Student.surname.ilike(search_term, escape='\\'),
                Student.middle_name.ilike(search_term, escape='\\'),
                Student.student_id.ilike(search_term, escape='\\')
            )
        )
    
    if min_score:
        query = query.filter(MockJAMBResult.total_score >= min_score)
    if max_score:
        query = query.filter(MockJAMBResult.total_score <= max_score)
    
    if sort_by == 'score':
        if sort_order == 'desc':
            query = query.order_by(MockJAMBResult.total_score.desc())
        else:
            query = query.order_by(MockJAMBResult.total_score.asc())
    elif sort_by == 'name':
        if sort_order == 'desc':
            query = query.order_by(Student.surname.desc())
        else:
            query = query.order_by(Student.surname.asc())
    else:
        query = query.order_by(MockJAMBResult.total_score.desc())
    
    results = query.all()

    from utils.broadsheet_export import _abbr_one

    def subj_arr(r):
        out = []
        for i in (1, 2, 3, 4):
            name = getattr(r, f'subject{i}')
            score = getattr(r, f'subject{i}_score')
            # `code` is the short CAPITAL subject abbreviation for the export grid.
            out.append({'name': name, 'code': _abbr_one(name), 'score': score} if name else None)
        return out

    results_payload = []
    for idx, r in enumerate(results):
        st = r.student
        results_payload.append({
            'rank': idx + 1,
            'student': {'id': st.id, 'full_name': st.full_name, 'student_id': st.student_id,
                        'progress_url': url_for('mock_jamb.student_progress', student_id=st.id)},
            'total_score': r.total_score,
            'performance_level': r.performance_level,
            'perf_class': r.performance_level.lower().replace(' ', '_'),
            'subjects': subj_arr(r),
            'edit_url': url_for('mock_jamb.edit_result', result_id=r.id),
            'delete_url': url_for('mock_jamb.delete_result', result_id=r.id),
        })

    stats_payload = None
    if statistics and statistics.get('statistics'):
        stats_payload = {
            'student_count': statistics['student_count'],
            'statistics': statistics['statistics'],
            'distribution': statistics['distribution'],
            'subject_analysis': statistics['subject_analysis'],
        }

    # School + branch for the export masthead (branch = the branch the exam/results
    # belong to).
    from utils.school import school_profile
    from models.models_branch import Branch
    _school_name = (school_profile() or {}).get('name') or ''
    _branch = db.session.get(Branch, exam.branch_id) if exam.branch_id else None
    _branch_name = _branch.name if _branch else ''

    return _render({
        'page': 'view_exam',
        'exam': {'id': exam.id, 'display_name': exam.display_name,
                 'exam_date': exam.exam_date.strftime('%d %B %Y') if exam.exam_date else '',
                 'session_name': exam.session.name if exam.session else '',
                 'school_name': _school_name, 'branch_name': _branch_name},
        'statistics': stats_payload,
        'results': results_payload,
        'filters': {'search': search, 'min_score': min_score or '', 'max_score': max_score or '',
                    'sort': sort_by, 'order': sort_order},
        'has_filter': bool(search or min_score or max_score),
        'urls': {'add': url_for('mock_jamb.add_result', exam_id=exam.id),
                 'bulk': url_for('mock_jamb.bulk_entry', exam_id=exam.id),
                 'export': url_for('mock_jamb.export_results', exam_id=exam.id),
                 'edit': url_for('mock_jamb.edit_exam', exam_id=exam.id),
                 'deep': url_for('mock_jamb.deep', exam_id=exam.id),
                 'questions': url_for('mock_jamb.questions', exam_id=exam.id),
                 'items': url_for('mock_jamb.items', exam_id=exam.id),
                 'index': url_for('mock_jamb.index'),
                 'self': url_for('mock_jamb.view_exam', exam_id=exam.id),
                 'delete_exam': url_for('mock_jamb.delete_exam', exam_id=exam.id)},
    })


@mock_jamb_bp.route('/exam/<int:exam_id>/edit', methods=['GET', 'POST'])
@login_required
@csrf_protect
def edit_exam(exam_id):
    """Edit mock exam details"""
    exam = db.get_or_404(MockJAMBExam, exam_id)
    require_branch_access(exam.branch_id)

    if request.method == 'POST':
        try:
            exam.name = request.form.get('name', '').strip()
            exam.description = request.form.get('description', '').strip()
            
            exam_date = request.form.get('exam_date')
            if exam_date:
                exam.exam_date = datetime.strptime(exam_date, '%Y-%m-%d').date()
            
            exam.is_completed = request.form.get('is_completed') == 'on'

            db.session.commit()
            return _ok('Exam updated successfully!', url_for('mock_jamb.view_exam', exam_id=exam_id))

        except Exception as e:
            db.session.rollback()
            return _err(f'Error updating exam: {str(e)}', url_for('mock_jamb.edit_exam', exam_id=exam_id))

    return _render({
        'page': 'edit_exam',
        'exam': {'id': exam.id, 'name': exam.name or '', 'description': exam.description or '',
                 'exam_date': exam.exam_date.strftime('%Y-%m-%d') if exam.exam_date else '',
                 'is_completed': bool(exam.is_completed), 'display_name': exam.display_name,
                 'session_name': exam.session.name if exam.session else ''},
        'submit_url': url_for('mock_jamb.edit_exam', exam_id=exam.id),
        'delete_url': url_for('mock_jamb.delete_exam', exam_id=exam.id),
        'view_url': url_for('mock_jamb.view_exam', exam_id=exam.id),
    })


@mock_jamb_bp.route('/exam/<int:exam_id>/delete', methods=['POST'])
@login_required
@csrf_protect
def delete_exam(exam_id):
    """Delete a mock exam and all its results"""
    exam = db.get_or_404(MockJAMBExam, exam_id)
    require_branch_access(exam.branch_id)
    session_id = exam.session_id
    exam_name = exam.display_name
    
    try:
        db.session.delete(exam)
        db.session.commit()
        return _ok(f'{exam_name} and all its results have been deleted.',
                   url_for('mock_jamb.index', session_id=session_id))
    except Exception as e:
        db.session.rollback()
        return _err(f'Error deleting exam: {str(e)}', url_for('mock_jamb.index', session_id=session_id))


# =============================================================================
# RESULT ENTRY
# =============================================================================

@mock_jamb_bp.route('/exam/<int:exam_id>/results/add', methods=['GET', 'POST'])
@login_required
@csrf_protect
def add_result(exam_id):
    """Add results for a student"""
    exam = db.get_or_404(MockJAMBExam, exam_id)
    require_branch_access(exam.branch_id)
    
    # SSS3 students who don't already have a result for this exam.
    existing_student_ids = {r.student_id for r in exam.results}
    students = [s for s in get_sss3_students() if s.id not in existing_student_ids]
    
    if request.method == 'POST':
        try:
            student_id = request.form.get('student_id', type=int)

            if not student_id:
                return _err('Please select a student.', url_for('mock_jamb.add_result', exam_id=exam_id))
            from utils.access_control import assert_exam_student
            assert_exam_student(student_id)          # own-arm only for SSS3 teachers

            # Check if result already exists
            existing = MockJAMBResult.query.filter_by(student_id=student_id, mock_exam_id=exam_id).first()
            if existing:
                return _err('Result already exists for this student. Edit instead.', url_for('mock_jamb.add_result', exam_id=exam_id))

            # Read each subject and convert the raw number of correct answers
            # (out of 60 for English, 40 for others) into a score over 100.
            subjects_payload = _read_subject_scores(request.form)
            total_score = sum(s['score'] for s in subjects_payload if s['score'] is not None)

            if total_score < 0 or total_score > MAX_TOTAL_SCORE:
                return _err(f'Total score must be between 0 and {MAX_TOTAL_SCORE}.', url_for('mock_jamb.add_result', exam_id=exam_id))

            result = MockJAMBResult(
                student_id=student_id,
                mock_exam_id=exam_id,
                total_score=total_score,
                subject1=subjects_payload[0]['name'],
                subject1_score=subjects_payload[0]['score'],
                subject2=subjects_payload[1]['name'],
                subject2_score=subjects_payload[1]['score'],
                subject3=subjects_payload[2]['name'],
                subject3_score=subjects_payload[2]['score'],
                subject4=subjects_payload[3]['name'],
                subject4_score=subjects_payload[3]['score'],
            )

            db.session.add(result)
            db.session.commit()
            from utils.access_control import has_sss3_exam_access
            if request.form.get('add_another'):
                dest = url_for('mock_jamb.add_result', exam_id=exam_id)
            elif has_sss3_exam_access():   # can't open the cohort exam view
                dest = url_for('mock_jamb.student_progress', student_id=student_id)
            else:
                dest = url_for('mock_jamb.view_exam', exam_id=exam_id)
            return _ok('Result added successfully!', dest)

        except Exception as e:
            db.session.rollback()
            return _err(f'Error adding result: {str(e)}', url_for('mock_jamb.add_result', exam_id=exam_id))

    return _render({
        'page': 'add_result',
        'exam': {'id': exam.id, 'display_name': exam.display_name,
                 'exam_date': exam.exam_date.strftime('%d %B %Y') if exam.exam_date else ''},
        'students': [{'id': s.id, 'label': f'{s.full_name} ({s.student_id})'} for s in students],
        'subjects': list(WAEC_SUBJECTS),
        'question_counts': question_count_map(WAEC_SUBJECTS),
        'compulsory_subject': COMPULSORY_SUBJECT,
        'subject_map': student_subject_map(students),
        'max_total': MAX_TOTAL_SCORE,
        'submit_url': url_for('mock_jamb.add_result', exam_id=exam.id),
        'view_url': url_for('mock_jamb.view_exam', exam_id=exam.id),
    })


@mock_jamb_bp.route('/exam/<int:exam_id>/results/bulk', methods=['GET', 'POST'])
@login_required
@csrf_protect
def bulk_entry(exam_id):
    """Bulk entry of results for multiple students"""
    exam = db.get_or_404(MockJAMBExam, exam_id)
    require_branch_access(exam.branch_id)
    
    # Get active term to find enrolled students
    active_term = get_active_term()
    
    # Only SSS3 sit Mock JAMB — resolve the class tolerantly of naming.
    from utils.helpers import get_sss3_class
    sss3 = get_sss3_class()
    
    students = []
    if sss3 and active_term:
        assignments = scope_query(ClassArmAssignment.query.filter_by(
            class_id=sss3.id, term_id=active_term.id), ClassArmAssignment).all()
        for assignment in assignments:
            enrollments = StudentEnrollment.query.filter_by(
                class_arm_assignment_id=assignment.id,
                is_active=True
            ).join(Student).order_by(Student.surname).all()
            
            for enrollment in enrollments:
                # Check if result exists
                existing = MockJAMBResult.query.filter_by(
                    student_id=enrollment.student_id,
                    mock_exam_id=exam_id
                ).first()
                
                students.append({
                    'student': enrollment.student,
                    'arm': assignment.arm_label,
                    'existing_result': existing
                })
    
    if request.method == 'POST':
        try:
            added = 0
            updated = 0
            
            def safe_int(val):
                try:
                    return int(val) if val else None
                except Exception:
                    return None
            
            for student_data in students:
                student = student_data['student']
                score_key = f'score_{student.id}'
                
                total_score = request.form.get(score_key, type=int)
                if total_score is None:
                    continue
                
                if total_score < 0 or total_score > 400:
                    continue
                
                existing = student_data['existing_result']
                
                if existing:
                    existing.total_score = total_score
                    existing.subject1 = request.form.get(f'subject1_{student.id}', '').strip() or None
                    existing.subject1_score = safe_int(request.form.get(f'subject1_score_{student.id}'))
                    existing.subject2 = request.form.get(f'subject2_{student.id}', '').strip() or None
                    existing.subject2_score = safe_int(request.form.get(f'subject2_score_{student.id}'))
                    existing.subject3 = request.form.get(f'subject3_{student.id}', '').strip() or None
                    existing.subject3_score = safe_int(request.form.get(f'subject3_score_{student.id}'))
                    existing.subject4 = request.form.get(f'subject4_{student.id}', '').strip() or None
                    existing.subject4_score = safe_int(request.form.get(f'subject4_score_{student.id}'))
                    updated += 1
                else:
                    result = MockJAMBResult(
                        student_id=student.id,
                        mock_exam_id=exam_id,
                        total_score=total_score,
                        subject1=request.form.get(f'subject1_{student.id}', '').strip() or None,
                        subject1_score=safe_int(request.form.get(f'subject1_score_{student.id}')),
                        subject2=request.form.get(f'subject2_{student.id}', '').strip() or None,
                        subject2_score=safe_int(request.form.get(f'subject2_score_{student.id}')),
                        subject3=request.form.get(f'subject3_{student.id}', '').strip() or None,
                        subject3_score=safe_int(request.form.get(f'subject3_score_{student.id}')),
                        subject4=request.form.get(f'subject4_{student.id}', '').strip() or None,
                        subject4_score=safe_int(request.form.get(f'subject4_score_{student.id}'))
                    )
                    db.session.add(result)
                    added += 1
            
            db.session.commit()
            return _ok(f'Results saved! Added: {added}, Updated: {updated}',
                       url_for('mock_jamb.view_exam', exam_id=exam_id))

        except Exception as e:
            db.session.rollback()
            return _err(f'Error saving results: {str(e)}', url_for('mock_jamb.bulk_entry', exam_id=exam_id))

    return _render({
        'page': 'bulk_entry',
        'exam': {'id': exam.id, 'display_name': exam.display_name},
        'students': [{'id': it['student'].id, 'full_name': it['student'].full_name,
                      'student_id': it['student'].student_id, 'arm': it['arm'],
                      'total_score': it['existing_result'].total_score if it['existing_result'] else '',
                      'entered': bool(it['existing_result'])} for it in students],
        'submit_url': url_for('mock_jamb.bulk_entry', exam_id=exam.id),
        'urls': {'view': url_for('mock_jamb.view_exam', exam_id=exam.id),
                 'add': url_for('mock_jamb.add_result', exam_id=exam.id)},
    })


@mock_jamb_bp.route('/result/<int:result_id>/edit', methods=['GET', 'POST'])
@login_required
@csrf_protect
def edit_result(result_id):
    """Edit a specific result"""
    from utils.branch_scope import require_branch_access
    result = db.get_or_404(MockJAMBResult, result_id)
    require_branch_access(result.student.branch_id)   # scope by the result's student
    from utils.access_control import assert_exam_student
    assert_exam_student(result.student_id)            # own-arm only for SSS3 teachers

    if request.method == 'POST':
        try:
            total_score = request.form.get('total_score', type=int)
            
            if total_score is None or total_score < 0 or total_score > 400:
                return _err('Invalid total score. Must be between 0 and 400.',
                            url_for('mock_jamb.edit_result', result_id=result_id))
            
            def safe_int(val):
                try:
                    return int(val) if val else None
                except Exception:
                    return None
            
            result.total_score = total_score
            result.subject1 = request.form.get('subject1', '').strip() or None
            result.subject1_score = safe_int(request.form.get('subject1_score'))
            result.subject2 = request.form.get('subject2', '').strip() or None
            result.subject2_score = safe_int(request.form.get('subject2_score'))
            result.subject3 = request.form.get('subject3', '').strip() or None
            result.subject3_score = safe_int(request.form.get('subject3_score'))
            result.subject4 = request.form.get('subject4', '').strip() or None
            result.subject4_score = safe_int(request.form.get('subject4_score'))
            
            db.session.commit()
            return _ok('Result updated successfully!', url_for('mock_jamb.view_exam', exam_id=result.mock_exam_id))

        except Exception as e:
            db.session.rollback()
            return _err(f'Error updating result: {str(e)}', url_for('mock_jamb.edit_result', result_id=result_id))

    r = result
    return _render({
        'page': 'edit_result',
        'result': {'id': r.id, 'student_name': r.student.full_name, 'student_id': r.student.student_id,
                   'subject1': r.subject1 or '', 'subject1_score': r.subject1_score if r.subject1_score is not None else '',
                   'subject2': r.subject2 or '', 'subject2_score': r.subject2_score if r.subject2_score is not None else '',
                   'subject3': r.subject3 or '', 'subject3_score': r.subject3_score if r.subject3_score is not None else '',
                   'subject4': r.subject4 or '', 'subject4_score': r.subject4_score if r.subject4_score is not None else '',
                   'total_score': r.total_score, 'exam_name': r.exam.display_name,
                   'exam_id': r.mock_exam_id},
        'subjects': list(WAEC_SUBJECTS),
        'submit_url': url_for('mock_jamb.edit_result', result_id=r.id),
        'delete_url': url_for('mock_jamb.delete_result', result_id=r.id),
        'view_url': url_for('mock_jamb.view_exam', exam_id=r.mock_exam_id),
    })


@mock_jamb_bp.route('/result/<int:result_id>/delete', methods=['POST'])
@login_required
@csrf_protect
def delete_result(result_id):
    """Delete a result"""
    from utils.branch_scope import require_branch_access
    result = db.get_or_404(MockJAMBResult, result_id)
    require_branch_access(result.student.branch_id)   # scope by the result's student
    exam_id = result.mock_exam_id
    student_name = result.student.full_name
    
    try:
        db.session.delete(result)
        db.session.commit()
        return _ok(f'Result for {student_name} deleted.', url_for('mock_jamb.view_exam', exam_id=exam_id))
    except Exception as e:
        db.session.rollback()
        return _err(f'Error deleting result: {str(e)}', url_for('mock_jamb.view_exam', exam_id=exam_id))


# =============================================================================
# STUDENT PROGRESS & ANALYTICS
# =============================================================================

@mock_jamb_bp.route('/student/<int:student_id>')
@login_required
def student_progress(student_id):
    """View a student's progress across all mock exams"""
    student = db.get_or_404(Student, student_id)
    from utils.access_control import assert_student_access, assert_exam_student
    assert_student_access(student)   # branch + form-teacher scope
    assert_exam_student(student_id)  # own-arm only for derived SSS3 exam teachers

    session_id = request.args.get('session_id', type=int)
    active_session = get_active_session()
    if not session_id and active_session:
        session_id = active_session.id
    
    progress = MockJAMBAnalytics.get_student_progress(student_id, session_id)
    prediction = MockJAMBAnalytics.predict_real_jamb(student_id, session_id)
    recommendations = MockJAMBAnalytics.get_improvement_recommendations(student_id, session_id)

    sessions = AcademicSession.query.order_by(AcademicSession.name.desc()).all()

    prog_payload = None
    if progress:
        prog_payload = {
            'average_score': progress['average_score'],
            'best_score': progress['best_score'],
            'latest_score': progress['latest_score'],
            'exam_count': progress['exam_count'],
            'progress': [{'exam': p['exam'], 'exam_number': p['exam_number'],
                          'exam_date': p['exam_date'].strftime('%d %b %Y') if p['exam_date'] else '',
                          'score': p['score'], 'change': p['change'],
                          'subjects': p['subjects']} for p in progress['progress']],
        }

    return _render({
        'page': 'student_progress',
        'student': {'id': student.id, 'full_name': student.full_name,
                    'student_id': student.student_id, 'gender': student.gender or '',
                    'jamb_target': student.jamb_target},
        'progress': prog_payload,
        'prediction': prediction,
        'recommendations': recommendations,
        'sessions': [{'id': s.id, 'name': s.name} for s in sessions],
        'selected_session_id': session_id or '',
        'urls': {'index': url_for('mock_jamb.index'),
                 'self': url_for('mock_jamb.student_progress', student_id=student.id),
                 'predictions': url_for('results.student_predictions', student_id=student.id)},
    })


@mock_jamb_bp.route('/analytics')
@login_required
def analytics():
    """Comprehensive analytics dashboard"""
    session_id = request.args.get('session_id', type=int)
    active_session = get_active_session()
    if not session_id and active_session:
        session_id = active_session.id
    
    sessions = AcademicSession.query.order_by(AcademicSession.name.desc()).all()
    
    comparison = None
    exams_stats = []
    
    if session_id:
        comparison = MockJAMBAnalytics.compare_mock_exams(session_id)
        exams = scope_query(MockJAMBExam.query.filter_by(session_id=session_id),
                            MockJAMBExam).order_by(MockJAMBExam.exam_number).all()
        # Only the creator's branch should appear; compare_mock_exams is not
        # branch-scoped, so intersect it with the scoped exam ids.
        scoped_ids = {e.id for e in exams}
        comparison = [c for c in (comparison or []) if c['exam'].id in scoped_ids]
        for exam in exams:
            stats = MockJAMBAnalytics.get_exam_statistics(exam.id)
            if stats:
                exams_stats.append(stats)

    comp_payload = [{'exam': {'id': c['exam'].id, 'display_name': c['exam'].display_name},
                     'student_count': c['student_count'], 'average': c['average'],
                     'max': c['max'], 'min': c['min'], 'above_200': c['above_200'],
                     'above_250_pct': c['above_250_pct']} for c in comparison]
    estats_payload = [{'exam': {'id': s['exam'].id, 'display_name': s['exam'].display_name,
                                'view_url': url_for('mock_jamb.view_exam', exam_id=s['exam'].id),
                                'exam_date': s['exam'].exam_date.strftime('%d %B %Y') if s['exam'].exam_date else ''},
                       'student_count': s['student_count'], 'statistics': s['statistics'],
                       'distribution': s['distribution'], 'subject_analysis': s['subject_analysis']}
                      for s in exams_stats if s.get('statistics')]

    return _render({
        'page': 'analytics',
        'sessions': [{'id': s.id, 'name': s.name} for s in sessions],
        'selected_session_id': session_id or '',
        'comparison': comp_payload,
        'exams_stats': estats_payload,
        'urls': {'index': url_for('mock_jamb.index'), 'create': url_for('mock_jamb.create_exam'),
                 'validation': url_for('mock_jamb.validation'), 'self': url_for('mock_jamb.analytics')},
    })


@mock_jamb_bp.route('/exam/<int:exam_id>/deep')
@login_required
def deep(exam_id):
    """Decision-grade deep analytics for one Mock JAMB exam — per subject, per
    teacher, per class arm, with evidence-based recommendations."""
    from utils.mock_deep_analytics import deep_analytics
    exam = db.get_or_404(MockJAMBExam, exam_id)
    require_branch_access(exam.branch_id)
    data = deep_analytics('jamb', exam_id)
    return _render({
        'page': 'deep',
        'exam': {'id': exam.id, 'display_name': exam.display_name,
                 'exam_date': exam.exam_date.strftime('%d %B %Y') if exam.exam_date else '',
                 'session_name': exam.session.name if exam.session else ''},
        'deep': data,
        'compose_base': url_for('comms.compose'),
        'urls': {'view': url_for('mock_jamb.view_exam', exam_id=exam.id),
                 'index': url_for('mock_jamb.index'),
                 'analytics': url_for('mock_jamb.analytics'),
                 'items': url_for('mock_jamb.items', exam_id=exam.id),
                 'trends': url_for('mock_jamb.trends', session_id=exam.session_id),
                 'self': url_for('mock_jamb.deep', exam_id=exam.id),
                 'export_pdf': url_for('mock_jamb.deep_export', exam_id=exam.id, format='pdf'),
                 'export_excel': url_for('mock_jamb.deep_export', exam_id=exam.id, format='excel'),
                 'export_image': url_for('mock_jamb.deep_export', exam_id=exam.id, format='image')},
    })


@mock_jamb_bp.route('/trends')
@login_required
def trends():
    """Longitudinal deep analytics across many Mock JAMB exams — one session, or
    all sessions (year-over-year progress)."""
    from utils.mock_deep_analytics import deep_trends
    scope = request.args.get('scope')            # 'all' -> every session
    session_id = request.args.get('session_id', type=int)
    active = get_active_session()
    if scope != 'all' and not session_id and active:
        session_id = active.id
    if scope == 'all':
        session_id = None
    data = deep_trends('jamb', session_id=session_id)
    for p in data.get('periods', []):
        p['deep_url'] = url_for('mock_jamb.deep', exam_id=p['exam_id'])
    sessions = AcademicSession.query.order_by(AcademicSession.name.desc()).all()
    return _render({
        'page': 'trends',
        'trends': data,
        'scope': 'all' if session_id is None else 'session',
        'selected_session_id': session_id or '',
        'sessions': [{'id': s.id, 'name': s.name} for s in sessions],
        'compose_base': url_for('comms.compose'),
        'urls': {'index': url_for('mock_jamb.index'),
                 'analytics': url_for('mock_jamb.analytics'),
                 'self': url_for('mock_jamb.trends'),
                 'all': url_for('mock_jamb.trends', scope='all'),
                 'export_pdf': url_for('mock_jamb.trends_export', format='pdf',
                                       scope=('all' if session_id is None else ''), session_id=session_id or ''),
                 'export_excel': url_for('mock_jamb.trends_export', format='excel',
                                         scope=('all' if session_id is None else ''), session_id=session_id or ''),
                 'export_image': url_for('mock_jamb.trends_export', format='image',
                                         scope=('all' if session_id is None else ''), session_id=session_id or '')},
    })


@mock_jamb_bp.route('/trends/export')
@login_required
def trends_export():
    """Export the Mock JAMB progress trends. ``format`` = pdf | excel | image."""
    from utils.mock_deep_analytics import deep_trends
    from utils.mock_deep_report import trends_pdf, trends_xlsx, trends_png, trends_filename
    from utils.web_exports import pdf_response, xlsx_response, png_response
    scope = request.args.get('scope')
    session_id = request.args.get('session_id', type=int)
    if scope != 'all' and not session_id:
        active = get_active_session()
        session_id = active.id if active else None
    if scope == 'all':
        session_id = None
    data = deep_trends('jamb', session_id=session_id)
    if not data or data['meta'].get('insufficient'):
        flash('Need at least two mocks with results to chart progress.', 'warning')
        return redirect(url_for('mock_jamb.trends', scope=scope or '', session_id=session_id or ''))
    fmt = (request.args.get('format') or 'pdf').lower()
    meta = data['meta']
    if fmt in ('excel', 'xlsx'):
        return xlsx_response(trends_xlsx(data), trends_filename(meta, 'xlsx'))
    if fmt in ('image', 'png'):
        return png_response(trends_png(data), trends_filename(meta, 'png'), inline=False)
    return pdf_response(trends_pdf(data), trends_filename(meta, 'pdf'), inline=False)


@mock_jamb_bp.route('/exam/<int:exam_id>/deep/export')
@login_required
def deep_export(exam_id):
    """Export the Mock JAMB deep analytics. ``format`` = pdf | excel | image."""
    from utils.mock_deep_analytics import deep_analytics
    from utils.mock_deep_report import deep_pdf, deep_xlsx, deep_png, deep_filename
    from utils.web_exports import pdf_response, xlsx_response, png_response
    exam = db.get_or_404(MockJAMBExam, exam_id)
    require_branch_access(exam.branch_id)
    data = deep_analytics('jamb', exam_id)
    if not data or data['meta'].get('empty'):
        flash('No results yet — enter scores to unlock deep analytics.', 'warning')
        return redirect(url_for('mock_jamb.deep', exam_id=exam_id))
    fmt = (request.args.get('format') or 'pdf').lower()
    meta = data['meta']
    if fmt in ('excel', 'xlsx'):
        return xlsx_response(deep_xlsx(data), deep_filename(meta, 'xlsx'))
    if fmt in ('image', 'png'):
        return png_response(deep_png(data), deep_filename(meta, 'png'), inline=False)
    return pdf_response(deep_pdf(data), deep_filename(meta, 'pdf'), inline=False)


@mock_jamb_bp.route('/exam/<int:exam_id>/items')
@login_required
def items(exam_id):
    """Item- & topic-level analysis of the ONLINE sitting for one Mock JAMB —
    per-question difficulty/discrimination, distractor analysis, topic and
    sub-topic mastery, flagged items and audience recommendations."""
    from utils.mock_jamb_item_analysis import item_analysis
    from utils.mock_jamb_sitting import auto_submit_expired
    exam = db.get_or_404(MockJAMBExam, exam_id)
    require_branch_access(exam.branch_id)
    auto_submit_expired(exam=exam)          # finalise any timed-out sittings first
    data = item_analysis(exam_id)
    return render_template(
        'mock_jamb/items.html', exam=exam, data=data,
        urls={'view': url_for('mock_jamb.view_exam', exam_id=exam.id),
              'questions': url_for('mock_jamb.questions', exam_id=exam.id),
              'deep': url_for('mock_jamb.deep', exam_id=exam.id),
              'index': url_for('mock_jamb.index'),
              'mastery': url_for('mock_jamb.mastery'),
              'self': url_for('mock_jamb.items', exam_id=exam.id),
              'export_pdf': url_for('mock_jamb.items_export', exam_id=exam.id, format='pdf'),
              'export_excel': url_for('mock_jamb.items_export', exam_id=exam.id, format='excel'),
              'export_image': url_for('mock_jamb.items_export', exam_id=exam.id, format='image'),
              'weakness_pdf': url_for('mock_jamb.weakness_export', exam_id=exam.id, format='pdf'),
              'weakness_excel': url_for('mock_jamb.weakness_export', exam_id=exam.id, format='excel')})


def _mastery_allowed_ids():
    """Branch-scoped student ids for cohort mastery — None (all) for central."""
    from utils.branch_scope import is_central
    if is_central():
        return None
    from models import Student
    return [s.id for s in scope_query(Student.query, Student).all()]


@mock_jamb_bp.route('/mastery')
@login_required
def mastery():
    """Cohort topic/sub-topic mastery across all online Mock JAMB sittings: the
    topics most students get wrong (and excel at), plus a searchable roster to
    drill into any one student's topic mastery and improvement across mocks."""
    from utils.mock_student_mastery import cohort_topic_gaps
    from models import Student, MockJAMBAttempt
    allowed = _mastery_allowed_ids()
    data = cohort_topic_gaps(allowed_ids=allowed)
    # roster of students who have sat at least one online mock
    q = (db.session.query(Student.id, Student.first_name, Student.surname,
                          Student.student_id, func.count(MockJAMBAttempt.id))
         .join(MockJAMBAttempt, MockJAMBAttempt.student_id == Student.id)
         .filter(MockJAMBAttempt.status == 'Submitted'))
    if allowed is not None:
        q = q.filter(Student.id.in_(allowed or [-1]))
    roster = [{'id': i, 'name': f'{sn or ""} {fn or ""}'.strip(), 'code': code, 'mocks': n}
              for (i, fn, sn, code, n) in
              q.group_by(Student.id).order_by(Student.surname, Student.first_name).all()]
    return render_template('mock_jamb/mastery.html', data=data, roster=roster,
                           urls={'index': url_for('mock_jamb.index')})


@mock_jamb_bp.route('/student/<int:student_id>/mastery')
@login_required
def student_mastery_view(student_id):
    """One student's topic/sub-topic mastery across every mock they have sat:
    strengths, weaknesses, per-subject topic tables, and their improvement or
    decline over successive mocks."""
    from utils.mock_student_mastery import student_mastery
    from models import Student
    student = db.get_or_404(Student, student_id)
    require_branch_access(student.branch_id)
    from utils.access_control import assert_exam_student
    assert_exam_student(student_id)   # own-arm only for derived SSS3 exam teachers
    data = student_mastery(student_id)
    return render_template('mock_jamb/student_mastery.html', data=data, student=student,
                           urls={'index': url_for('mock_jamb.index'),
                                 'mastery': url_for('mock_jamb.mastery')})


@mock_jamb_bp.route('/exam/<int:exam_id>/items/export')
@login_required
def items_export(exam_id):
    """Export the Mock JAMB item analysis. ``format`` = pdf | excel | image."""
    from utils.mock_jamb_item_analysis import item_analysis
    from utils.mock_deep_report import items_pdf, items_xlsx, items_png, items_filename
    from utils.web_exports import pdf_response, xlsx_response, png_response
    exam = db.get_or_404(MockJAMBExam, exam_id)
    require_branch_access(exam.branch_id)
    data = item_analysis(exam_id)
    if not data or data['meta'].get('empty'):
        flash('No online sittings yet — publish the mock and let students sit it to unlock item analysis.', 'warning')
        return redirect(url_for('mock_jamb.items', exam_id=exam_id))
    fmt = (request.args.get('format') or 'pdf').lower()
    meta = data['meta']
    if fmt in ('excel', 'xlsx'):
        return xlsx_response(items_xlsx(data), items_filename(meta, 'xlsx'))
    if fmt in ('image', 'png'):
        return png_response(items_png(data), items_filename(meta, 'png'), inline=False)
    return pdf_response(items_pdf(data), items_filename(meta, 'pdf'), inline=False)


@mock_jamb_bp.route('/exam/<int:exam_id>/weakness/export')
@login_required
def weakness_export(exam_id):
    """Printable failure report — the questions, topics and sub-topics the cohort
    got wrong (weakest first). ``format`` = pdf | excel | image."""
    from utils.mock_jamb_item_analysis import item_analysis
    from utils.mock_deep_report import (weakness_pdf, weakness_xlsx, weakness_png,
                                        weakness_filename)
    from utils.web_exports import pdf_response, xlsx_response, png_response
    exam = db.get_or_404(MockJAMBExam, exam_id)
    require_branch_access(exam.branch_id)
    data = item_analysis(exam_id)
    if not data or data['meta'].get('empty'):
        flash('No online sittings yet — publish the mock and let students sit it to unlock the failure report.', 'warning')
        return redirect(url_for('mock_jamb.items', exam_id=exam_id))
    fmt = (request.args.get('format') or 'pdf').lower()
    meta = data['meta']
    if fmt in ('excel', 'xlsx'):
        return xlsx_response(weakness_xlsx(data), weakness_filename(meta, 'xlsx'))
    if fmt in ('image', 'png'):
        return png_response(weakness_png(data), weakness_filename(meta, 'png'), inline=False)
    return pdf_response(weakness_pdf(data), weakness_filename(meta, 'pdf'), inline=False)


@mock_jamb_bp.route('/validation')
@login_required
def validation():
    """Mock→actual validation: how well this session's Mock JAMB predicts the
    real JAMB (correlation, error, bias, calibration, cut-off reliability)."""
    from models import JAMBResult
    from utils.mock_validation import jamb_validation
    session_id = request.args.get('session_id', type=int)
    mock_exam_id = request.args.get('mock_exam_id', type=int)
    year = request.args.get('year', type=int)
    active = get_active_session()
    if not session_id and not mock_exam_id and active:
        session_id = active.id
    sessions = AcademicSession.query.order_by(AcademicSession.name.desc()).all()

    exams = (scope_query(MockJAMBExam.query.filter_by(session_id=session_id), MockJAMBExam)
             .order_by(MockJAMBExam.exam_number).all() if session_id else [])
    scoped = {e.id for e in exams}
    if mock_exam_id and mock_exam_id not in scoped:
        mock_exam_id = None                       # out of the caller's branch scope
    if not mock_exam_id and exams:
        mock_exam_id = exams[-1].id               # default to the final mock
    data = jamb_validation(mock_exam_id=mock_exam_id, year=year) if mock_exam_id else None
    years = sorted({y for (y,) in db.session.query(JAMBResult.exam_year).distinct().all()}, reverse=True)

    return _render({
        'page': 'validation',
        'sessions': [{'id': s.id, 'name': s.name} for s in sessions],
        'selected_session_id': session_id or '',
        'exams': [{'id': e.id, 'display_name': e.display_name, 'number': e.exam_number} for e in exams],
        'selected_mock_exam_id': mock_exam_id or '',
        'years': years, 'selected_year': year or '',
        'validation': data,
        'urls': {'index': url_for('mock_jamb.index'), 'analytics': url_for('mock_jamb.analytics'),
                 'self': url_for('mock_jamb.validation'),
                 'export_pdf': url_for('mock_jamb.validation_export', format='pdf',
                                       mock_exam_id=mock_exam_id or '', year=year or ''),
                 'export_excel': url_for('mock_jamb.validation_export', format='excel',
                                         mock_exam_id=mock_exam_id or '', year=year or '')},
    })


@mock_jamb_bp.route('/validation/export')
@login_required
def validation_export():
    """Export the Mock→actual JAMB validation. ``format`` = pdf | excel."""
    from utils.mock_validation import jamb_validation
    from utils.mock_validation_report import validation_pdf, validation_xlsx
    from utils.web_exports import pdf_response, xlsx_response
    mock_exam_id = request.args.get('mock_exam_id', type=int)
    year = request.args.get('year', type=int)
    if mock_exam_id:
        m = db.session.get(MockJAMBExam, mock_exam_id)
        if m:
            require_branch_access(m.branch_id)
    data = jamb_validation(mock_exam_id=mock_exam_id, year=year) if mock_exam_id else None
    if not data or data['meta'].get('insufficient'):
        flash('Not enough matched candidates to validate this mock yet.', 'warning')
        return redirect(url_for('mock_jamb.validation', mock_exam_id=mock_exam_id or ''))
    stem = 'jamb_mock_validation'
    if (request.args.get('format') or 'pdf').lower() in ('excel', 'xlsx'):
        return xlsx_response(validation_xlsx(data), f'{stem}.xlsx')
    return pdf_response(validation_pdf(data), f'{stem}.pdf', inline=False)


# =============================================================================
# ONLINE QUESTION BANK — JAMB-standard questions (with comprehension passages,
# topics/sub-topics and diagrams) that students will sit in-app (Phase 3).
# =============================================================================

# ext_ok compares against file_ext(), which returns the extension WITH its dot
# (".jpg"), so these must be dotted too — otherwise every upload is rejected.
_MOCK_IMG_EXTS = {'.png', '.jpg', '.jpeg', '.gif', '.webp'}


def _save_mock_image(file):
    """Save an uploaded figure under static/uploads/mock_jamb, re-encoded to PNG
    (decompression-bomb capped); returns its URL, or None."""
    if not file or not file.filename:
        return None
    from utils.uploads import ext_ok, open_image
    if not ext_ok(file.filename, _MOCK_IMG_EXTS):
        return None
    try:
        im = open_image(file).convert('RGB')
    except Exception:
        return None
    name = secrets.token_hex(8) + '.png'
    folder = os.path.join(current_app.root_path, 'static', 'uploads', 'mock_jamb')
    os.makedirs(folder, exist_ok=True)
    im.save(os.path.join(folder, name), 'PNG')
    return url_for('static', filename='uploads/mock_jamb/' + name)


def _mock_subjects():
    return Subject.query.filter_by(is_active=True).order_by(Subject.name).all()


# =============================================================================
# CENTRAL QUESTION BANK — subject-scoped, section-tagged questions & passages
# (mock_exam_id NULL) that mocks draw from per the JAMB blueprint.
# =============================================================================

def _bank_url(subject_id, section=None):
    return url_for('mock_jamb.bank', subject_id=subject_id or '', section=section or '')


def _bank_coverage(subject):
    """Per-section stock vs the JAMB blueprint need for a subject: a list of
    ``{section, label, have, need, passage, short}`` plus the totals."""
    from utils.jamb_blueprint import blueprint_for, sections_for
    from models import MockJAMBQuestion
    counts = dict(db.session.query(MockJAMBQuestion.section, func.count(MockJAMBQuestion.id))
                  .filter(MockJAMBQuestion.subject_id == subject.id,
                          MockJAMBQuestion.mock_exam_id.is_(None))
                  .group_by(MockJAMBQuestion.section).all())
    bp = {s['section']: s['count'] for s in blueprint_for(subject.name)['sections']}
    rows = []
    for s in sections_for(subject.name):
        have = counts.get(s['section'], 0)
        need = bp.get(s['section'], 0)
        rows.append({'section': s['section'], 'label': s['label'], 'passage': s['passage'],
                     'have': have, 'need': need, 'short': need > 0 and have < need})
    untagged = counts.get(None, 0) + counts.get('', 0)
    return {'rows': rows, 'total': sum(counts.values()), 'untagged': untagged,
            'need': sum(bp.values())}


def _bank_filter_options(subject_id):
    """Distinct topics / sub-topics / years / exam bodies present in a subject's
    bank — the option lists that drive the filter drop-downs. Sub-topics are
    grouped by their topic so the picker can narrow as the topic changes."""
    from models import MockJAMBQuestion
    rows = (db.session.query(MockJAMBQuestion.topic, MockJAMBQuestion.subtopic)
            .filter(MockJAMBQuestion.subject_id == subject_id,
                    MockJAMBQuestion.mock_exam_id.is_(None)).distinct().all())
    topics, subs_by_topic = set(), {}
    for topic, sub in rows:
        t = (topic or '').strip()
        if t:
            topics.add(t)
            if (sub or '').strip():
                subs_by_topic.setdefault(t, set()).add(sub.strip())
    years = sorted({(y or '').strip() for (y,) in db.session.query(MockJAMBQuestion.exam_year)
                    .filter(MockJAMBQuestion.subject_id == subject_id,
                            MockJAMBQuestion.mock_exam_id.is_(None)).distinct().all()
                    if (y or '').strip()}, reverse=True)
    bodies = sorted({(b or '').strip() for (b,) in db.session.query(MockJAMBQuestion.exam_body)
                     .filter(MockJAMBQuestion.subject_id == subject_id,
                             MockJAMBQuestion.mock_exam_id.is_(None)).distinct().all()
                     if (b or '').strip()})
    return {
        'topics': sorted(topics, key=str.lower),
        'subtopics_by_topic': {t: sorted(v, key=str.lower) for t, v in subs_by_topic.items()},
        'all_subtopics': sorted({s for v in subs_by_topic.values() for s in v}, key=str.lower),
        'years': years, 'exam_bodies': bodies,
    }


@mock_jamb_bp.route('/bank')
@login_required
def bank():
    """Central question bank: author section-tagged questions & passages per
    subject that every mock draws from by the JAMB blueprint."""
    from routes.cbt import _subject_topic_tree
    from utils.jamb_blueprint import sections_for
    subjects = _mock_subjects()
    subject_id = request.args.get('subject_id', type=int) or (subjects[0].id if subjects else None)
    subject = db.session.get(Subject, subject_id) if subject_id else None
    section = (request.args.get('section') or '').strip() or None

    q_search = (request.args.get('q') or '').strip()
    # extra bank filters (all optional): topic, sub-topic, past-question year, exam body
    f_topic = (request.args.get('topic') or '').strip() or None
    f_subtopic = (request.args.get('subtopic') or '').strip() or None
    f_year = (request.args.get('year') or '').strip() or None
    f_body = (request.args.get('exam_body') or '').strip() or None
    page = max(1, request.args.get('page', 1, type=int))
    per_page = 25
    passages, standalone, coverage, sections = [], [], None, []
    pagination = None
    filter_options = None
    if subject:
        sections = sections_for(subject.name)
        coverage = _bank_coverage(subject)
        filter_options = _bank_filter_options(subject_id)

        def _apply(query):
            """Apply the shared bank filters (section/topic/subtopic/year/body/text)
            to a MockJAMBQuestion query."""
            if section:
                query = query.filter(MockJAMBQuestion.section == section)
            if f_topic:
                query = query.filter(MockJAMBQuestion.topic == f_topic)
            if f_subtopic:
                query = query.filter(MockJAMBQuestion.subtopic == f_subtopic)
            if f_year:
                query = query.filter(MockJAMBQuestion.exam_year == f_year)
            if f_body:
                query = query.filter(MockJAMBQuestion.exam_body == f_body)
            if q_search:
                query = query.filter(MockJAMBQuestion.question_text.ilike(f'%{q_search}%'))
            return query

        pq = MockJAMBPassage.query.filter_by(subject_id=subject_id).filter(
            MockJAMBPassage.mock_exam_id.is_(None))
        if section:
            pq = pq.filter(MockJAMBPassage.section == section)
        prows = pq.order_by(MockJAMBPassage.section, MockJAMBPassage.order, MockJAMBPassage.id).all()
        for p in prows:
            pqs = _apply(p.questions).order_by(
                MockJAMBQuestion.order, MockJAMBQuestion.id).all()
            # when a question-level filter is active, only show passages that still
            # have matching questions (so a topic/year filter doesn't list empty passages)
            if pqs or not (f_topic or f_subtopic or f_year or f_body or q_search):
                passages.append({'p': p, 'questions': pqs})
        sq = _apply(MockJAMBQuestion.query.filter_by(
            subject_id=subject_id, passage_id=None).filter(
            MockJAMBQuestion.mock_exam_id.is_(None)))
        sq = sq.order_by(MockJAMBQuestion.section, MockJAMBQuestion.order,
                         MockJAMBQuestion.id)
        pagination = sq.paginate(page=page, per_page=per_page, error_out=False)
        standalone = pagination.items
    from utils.myschool import on_jamb
    jamb_ok_ids = {s.id for s in subjects if on_jamb(s.name)}
    needs_image_count = MockJAMBQuestion.query.filter(
        MockJAMBQuestion.mock_exam_id.is_(None),
        MockJAMBQuestion.needs_image.is_(True)).count()
    # per-subject provenance breakdown for the bulk-delete panel
    bank_sources, bank_untagged = [], 0
    if subject:
        src_rows = (db.session.query(MockJAMBQuestion.source, func.count(MockJAMBQuestion.id))
                    .filter(MockJAMBQuestion.mock_exam_id.is_(None),
                            MockJAMBQuestion.subject_id == subject_id,
                            MockJAMBQuestion.passage_id.is_(None))
                    .group_by(MockJAMBQuestion.source).all())
        bank_sources = [{'source': s or '', 'label': (s or 'unknown'), 'count': n}
                        for s, n in sorted(src_rows, key=lambda r: -r[1])]
        bank_untagged = (MockJAMBQuestion.query
                         .filter(MockJAMBQuestion.mock_exam_id.is_(None),
                                 MockJAMBQuestion.subject_id == subject_id,
                                 MockJAMBQuestion.passage_id.is_(None),
                                 (MockJAMBQuestion.topic.is_(None)) | (MockJAMBQuestion.topic == '')).count())
    return render_template('mock_jamb/bank.html', subjects=subjects, subject=subject,
                           needs_image_count=needs_image_count,
                           subject_id=subject_id, section=section, sections=sections,
                           passages=passages, standalone=standalone, coverage=coverage,
                           pagination=pagination, q_search=q_search, jamb_ok_ids=jamb_ok_ids,
                           topic_tree=_subject_topic_tree(subject_id),
                           filter_options=filter_options,
                           f_topic=f_topic, f_subtopic=f_subtopic, f_year=f_year, f_body=f_body,
                           has_passage_sections=any(s['passage'] for s in sections),
                           novel_section=any(s['section'] == 'novel' for s in sections),
                           syllabus_url=url_for('cbt.syllabus', subject_id=subject_id or ''),
                           coded_syllabus_url=url_for('mock_jamb.bank_syllabus', subject_id=subject_id or ''),
                           analytics_url=url_for('mock_jamb.bank_analytics', subject_id=subject_id or ''),
                           index_url=url_for('mock_jamb.index'))


def _syllabus_tree(subject_id):
    """The stored coded syllabus for a subject as a nested list for display:
    [{code,name,count, topics:[{code,name, items:[{code,name}]}]}]."""
    from models import MockJAMBSyllabus, MockJAMBSyllabusNode
    syll = MockJAMBSyllabus.query.filter_by(subject_id=subject_id).first()
    if not syll:
        return None, None
    nodes = (MockJAMBSyllabusNode.query.filter_by(syllabus_id=syll.id)
             .order_by(MockJAMBSyllabusNode.sort_order, MockJAMBSyllabusNode.code).all())
    by_parent = {}
    for n in nodes:
        by_parent.setdefault(n.parent_id, []).append(n)

    def kids(pid, kind):
        return [n for n in by_parent.get(pid, []) if n.kind == kind]

    tree = []
    for sec in kids(None, 'section'):
        tree.append({
            'code': sec.code, 'name': sec.name, 'count': sec.question_count,
            'topics': [{
                'code': t.code, 'name': t.name,
                'items': [{'code': it.code, 'name': it.name} for it in kids(t.id, 'item')],
            } for t in kids(sec.id, 'topic')],
        })
    return syll, tree


def _bundled_syllabi():
    """Syllabus JSON files shipped in the repo (data, not code) that an admin can
    one-click import: {slug: {label, path}}."""
    base = os.path.join(current_app.root_path, 'data', 'jamb_syllabi')
    out = {}
    try:
        for fn in sorted(os.listdir(base)):
            if fn.endswith('.json'):
                out[fn[:-5]] = {'label': fn[:-5].replace('_', ' ').title(),
                                'path': os.path.join(base, fn)}
    except OSError:
        pass
    return out


def _match_subject(name):
    """Find the tenant's Subject row whose name matches an imported syllabus's
    subject, tolerant of the usual Nigerian exam aliases (Maths/Mathematics …)."""
    from utils.jamb_blueprint import norm_subject
    key = norm_subject(name)
    for s in _mock_subjects():
        if norm_subject(s.name) == key:
            return s
    return None


@mock_jamb_bp.route('/bank/syllabus')
@login_required
def bank_syllabus():
    """Coded-syllabus manager: view a subject's imported syllabus (stable codes)
    and import/replace it from JSON or CSV."""
    import json as _json
    subjects = _mock_subjects()
    subject_id = request.args.get('subject_id', type=int) or (subjects[0].id if subjects else None)
    subject = db.session.get(Subject, subject_id) if subject_id else None
    syll, tree = _syllabus_tree(subject_id) if subject_id else (None, None)
    blueprint = []
    if syll and syll.blueprint:
        try:
            blueprint = _json.loads(syll.blueprint)
        except (ValueError, TypeError):
            blueprint = []
    # Each bundled file labelled with the subject it will import into.
    bundled = _bundled_syllabi()
    for slug, b in bundled.items():
        try:
            with open(b['path'], encoding='utf-8') as fh:
                b['target'] = (_json.load(fh) or {}).get('subject') or b['label']
        except (OSError, ValueError):
            b['target'] = b['label']
    # Retag context: available years, whether an Anthropic key is set, and how
    # many stand-alone questions are already coded.
    years, tagged_count, total_q, key_set = [], 0, 0, False
    if subject_id:
        yr = (db.session.query(MockJAMBQuestion.exam_year)
              .filter(MockJAMBQuestion.subject_id == subject_id,
                      MockJAMBQuestion.mock_exam_id.is_(None),
                      MockJAMBQuestion.exam_year.isnot(None), MockJAMBQuestion.exam_year != '')
              .distinct().all())
        years = sorted({y[0] for y in yr}, reverse=True)
        base = MockJAMBQuestion.query.filter(
            MockJAMBQuestion.subject_id == subject_id,
            MockJAMBQuestion.mock_exam_id.is_(None),
            MockJAMBQuestion.passage_id.is_(None))
        total_q = base.count()
        tagged_count = base.filter(MockJAMBQuestion.syllabus_item_code.isnot(None),
                                   MockJAMBQuestion.syllabus_item_code != '').count()
        try:
            from utils.waec_ocr import _vision_config
            key_set = _vision_config().get('has_key', False)
        except Exception:
            key_set = False
    return render_template('mock_jamb/bank_syllabus.html',
                           subjects=subjects, subject=subject, subject_id=subject_id,
                           syllabus=syll, tree=tree, blueprint=blueprint, bundled=bundled,
                           years=years, tagged_count=tagged_count, total_questions_q=total_q,
                           key_set=key_set,
                           bank_url=url_for('mock_jamb.bank', subject_id=subject_id or ''))


@mock_jamb_bp.route('/bank/syllabus/import', methods=['POST'])
@login_required
@csrf_protect
def bank_syllabus_import():
    """Import a syllabus (pasted text, uploaded file, or a bundled data file) and
    reconcile it into the coded tables for a subject."""
    from utils.jamb_syllabus_import import import_syllabus, parse, SyllabusImportError
    subject_id = request.form.get('subject_id', type=int)
    subject = db.session.get(Subject, subject_id) if subject_id else None

    fmt = (request.form.get('format') or 'auto').lower()
    version = (request.form.get('version') or '').strip() or None
    text = None
    bundled = (request.form.get('bundled') or '').strip()
    if bundled:
        b = _bundled_syllabi().get(bundled)
        if not b:
            flash('That bundled syllabus is unavailable.', 'error')
            return redirect(url_for('mock_jamb.bank_syllabus', subject_id=subject_id or ''))
        with open(b['path'], encoding='utf-8') as fh:
            text = fh.read()
        fmt = 'json'
        # A bundled file prefers the subject IT declares (e.g. the Mathematics
        # file -> the Mathematics subject) so clicking "Mathematics" can't land
        # on Accounting. But schools rename subjects (Accounting for "Principles
        # of Accounts", Digital Technologies for "Computer Studies"), so when the
        # declared subject has no match here, fall back to the SELECTED subject.
        try:
            declared = parse(text, fmt='json').get('subject')
        except SyllabusImportError:
            declared = None
        matched = _match_subject(declared) if declared else None
        if matched is not None:
            subject = matched
            subject_id = subject.id
        elif subject is None:
            flash(f'No subject named "{declared}" exists here. Select the subject '
                  f'to import it into, then click the bundled button again.', 'error')
            return redirect(url_for('mock_jamb.bank_syllabus', subject_id=subject_id or ''))
        else:
            flash(f'No subject named "{declared}" here — imported into the selected '
                  f'subject "{subject.name}" instead.', 'warning')

    if not subject:
        flash('Pick a subject first.', 'error')
        return redirect(url_for('mock_jamb.bank_syllabus'))

    if text is None and request.files.get('file') and request.files['file'].filename:
        f = request.files['file']
        text = f.read().decode('utf-8', 'replace')
        if fmt == 'auto':
            fmt = 'csv' if f.filename.lower().endswith('.csv') else 'json'
    if text is None:
        text = request.form.get('text') or ''
    if not text.strip():
        flash('Paste a syllabus, upload a file, or choose a bundled one.', 'error')
        return redirect(url_for('mock_jamb.bank_syllabus', subject_id=subject_id))

    try:
        diff = import_syllabus(subject, text, fmt=fmt, version=version)
    except SyllabusImportError as e:
        db.session.rollback()
        flash(f'Import failed: {e}', 'error')
        return redirect(url_for('mock_jamb.bank_syllabus', subject_id=subject_id))

    flash(f"Syllabus imported for {subject.name}: "
          f"{len(diff['added'])} added, {len(diff['renamed'])} renamed, "
          f"{len(diff['removed'])} removed "
          f"({diff['sections']} sections, {diff['topics']} topics, {diff['items']} items).",
          'success')
    return redirect(url_for('mock_jamb.bank_syllabus', subject_id=subject_id))


@mock_jamb_bp.route('/bank/syllabus/clear', methods=['POST'])
@login_required
@csrf_protect
def bank_syllabus_clear():
    """Remove a subject's imported syllabus entirely (its nodes cascade). Bank
    questions and their draw blueprint are untouched — only the coded syllabus
    tree is deleted, so it can be re-imported cleanly."""
    from models import MockJAMBSyllabus
    subject_id = request.form.get('subject_id', type=int)
    subject = db.session.get(Subject, subject_id) if subject_id else None
    syll = MockJAMBSyllabus.query.filter_by(subject_id=subject_id).first() if subject_id else None
    if not syll:
        flash('No syllabus to clear for this subject.', 'warning')
        return redirect(url_for('mock_jamb.bank_syllabus', subject_id=subject_id or ''))
    db.session.delete(syll)            # nodes cascade via the relationship
    db.session.commit()
    flash(f'Syllabus cleared for {subject.name if subject else "the subject"}.', 'success')
    return redirect(url_for('mock_jamb.bank_syllabus', subject_id=subject_id or ''))


@mock_jamb_bp.route('/bank/syllabus/retag', methods=['POST'])
@login_required
@csrf_protect
def bank_syllabus_retag():
    """AI-retag a subject's bank questions to its CODED syllabus items, using the
    Anthropic key in Settings. The model may only pick from the imported syllabus
    codes — it can never invent one."""
    from utils.mock_bank_coded_retag import coded_retag
    subject_id = request.form.get('subject_id', type=int)
    subject = db.session.get(Subject, subject_id) if subject_id else None
    if not subject:
        flash('Pick a subject first.', 'error')
        return redirect(url_for('mock_jamb.bank_syllabus'))
    year = (request.form.get('year') or '').strip() or None
    exam_body = (request.form.get('exam_body') or '').strip() or None
    mode = 'untagged' if request.form.get('mode') == 'untagged' else 'all'

    res = coded_retag(subject, year=year, exam_body=exam_body, mode=mode)
    err = res.get('error')
    scope = ' '.join(filter(None, [exam_body, year]))
    scope = f' ({scope})' if scope else ''
    if err == 'no_key':
        flash('No Anthropic API key is set. Add one under Settings → AI Vision OCR to use AI tagging.', 'error')
    elif err == 'not_installed':
        flash('The "anthropic" package is not installed on the server, so AI tagging is unavailable.', 'error')
    elif err == 'no_syllabus':
        flash(f'Import a coded syllabus for {subject.name} first — the retag classifies against it.', 'error')
    elif res['tagged'] or res['outside']:
        more = (f" {res['total'] - res['scanned']} more remain — run again to continue."
                if res.get('capped') else '')
        outside = f"; {res['outside']} marked outside the syllabus" if res['outside'] else ''
        flash(f"AI classified {res['tagged']} of {res['scanned']} {subject.name}{scope} "
              f"question(s) to a syllabus code{outside}.{more}", 'success')
    elif res['scanned']:
        flash(f"AI reviewed {res['scanned']} {subject.name}{scope} question(s) but set no codes — "
              f"check the key and model under Settings → AI Vision OCR.", 'info')
    else:
        flash(f'No stand-alone {subject.name}{scope} bank questions to tag.', 'info')
    return redirect(url_for('mock_jamb.bank_syllabus', subject_id=subject_id))


@mock_jamb_bp.route('/bank/analytics')
@login_required
def bank_analytics():
    """Per-subject breakdown of the question bank: how questions distribute across
    topics, sub-topics, years and exam bodies, plus the cold/never-tested syllabus
    areas — so teachers can see what to focus students on."""
    from utils.mock_bank_analytics import subject_breakdown
    subjects = _mock_subjects()
    subject_id = request.args.get('subject_id', type=int) or (subjects[0].id if subjects else None)
    subject = db.session.get(Subject, subject_id) if subject_id else None
    data = subject_breakdown(subject) if subject else None
    # Past-question years present in this subject's bank, for the tag-scope picker.
    years = []
    if subject_id:
        years = sorted({(y or '').strip() for (y,) in db.session.query(MockJAMBQuestion.exam_year)
                        .filter(MockJAMBQuestion.subject_id == subject_id,
                                MockJAMBQuestion.mock_exam_id.is_(None),
                                MockJAMBQuestion.exam_year.isnot(None)).distinct().all()
                        if (y or '').strip()}, reverse=True)
    try:
        from utils.waec_ocr import _vision_config
        _vc = _vision_config()
        ai_ready = bool(_vc['installed'] and _vc['has_key'])
    except Exception:
        ai_ready = False
    # Local (offline) tagger: package present, and the latest run for this subject.
    try:
        from utils.mock_bank_local_tag import available as _local_available
        local_ready = _local_available()
    except Exception:
        local_ready = False
    local_job = None
    if subject_id:
        try:
            from models import BackgroundJob
            local_job = (BackgroundJob.query
                         .filter(BackgroundJob.kind == 'bank_local_tag',
                                 BackgroundJob.params.like(f'%"subject_id": {subject_id}%'))
                         .order_by(BackgroundJob.id.desc()).first())
        except Exception:
            local_job = None
    return render_template('mock_jamb/bank_analytics.html', subjects=subjects,
                           subject=subject, subject_id=subject_id, data=data, years=years,
                           ai_ready=ai_ready, local_ready=local_ready, local_job=local_job,
                           bank_url=url_for('mock_jamb.bank', subject_id=subject_id or ''),
                           index_url=url_for('mock_jamb.index'))


@mock_jamb_bp.route('/bank/retag', methods=['POST'])
@login_required
def bank_retag():
    """Auto-tag / re-tag a subject's bank questions with the keyword classifier.
    ``mode=all`` re-classifies already-tagged questions too (upgrading them with
    improved keywords); ``ensure_section`` makes still-unmatched questions drawable."""
    from utils.mock_bank_retag import retag_untagged
    subject_id = request.form.get('subject_id', type=int)
    subject = db.session.get(Subject, subject_id) if subject_id else None
    if not subject:
        flash('Pick a subject first.', 'error')
        return redirect(url_for('mock_jamb.bank_analytics'))
    mode = 'all' if request.form.get('mode') == 'all' else 'untagged'
    ensure_section = bool(request.form.get('ensure_section'))
    year = (request.form.get('year') or '').strip() or None
    res = retag_untagged(subject, mode=mode, ensure_section=ensure_section, year=year)
    verb = 'Re-classified' if mode == 'all' else 'Auto-tagged'
    if res['topic_set'] or res['section_ensured']:
        parts = []
        if res['topic_set']:
            parts.append(f"{verb} {res['topic_set']} question(s)")
        if res['section_set']:
            parts.append(f"{res['section_set']} got a section from the match")
        if res['section_ensured']:
            parts.append(f"{res['section_ensured']} unmatched made drawable")
        msg = ' · '.join(parts) + f" in {subject.name}."
        if res['still_untagged'] and not ensure_section:
            msg += (f" {res['still_untagged']} still have no topic (notation-only, figure-only "
                    f"or out-of-syllabus) — tick 'make drawable' or tag them by hand.")
        flash(msg, 'success')
    else:
        flash(f"No {subject.name} question matched the syllabus keywords — nothing changed.", 'info')
    return redirect(url_for('mock_jamb.bank_analytics', subject_id=subject_id))


@mock_jamb_bp.route('/bank/ai-retag', methods=['POST'])
@login_required
def bank_ai_retag():
    """Tag a subject's bank questions with AI, using the Anthropic key linked in
    Settings → OCR. Assigns each question a real JAMB-syllabus topic (never an
    invented one), across the whole subject or one past-question year."""
    from utils.mock_bank_ai_retag import ai_retag
    subject_id = request.form.get('subject_id', type=int)
    subject = db.session.get(Subject, subject_id) if subject_id else None
    if not subject:
        flash('Pick a subject first.', 'error')
        return redirect(url_for('mock_jamb.bank_analytics'))
    year = (request.form.get('year') or '').strip() or None
    res = ai_retag(subject, year=year)
    err = res.get('error')
    scope = f' ({year})' if year else ''
    if err == 'no_key':
        flash('No Anthropic API key is set. Add one under Settings → OCR to use AI tagging.', 'error')
    elif err == 'not_installed':
        flash('The "anthropic" package is not installed on the server, so AI tagging is unavailable.', 'error')
    elif err == 'no_syllabus':
        flash(f'No JAMB syllabus is available for {subject.name}, so AI tagging has no topics to assign.', 'error')
    elif res['topic_set']:
        more = (f" {res['total'] - res['scanned']} more remain — run again"
                f"{' or pick a year' if not year else ''} to continue." if res.get('capped') else '')
        flash(f"AI tagged {res['topic_set']} of {res['scanned']} {subject.name}{scope} "
              f"question(s) with a syllabus topic.{more}", 'success')
    elif res['scanned']:
        flash(f"AI reviewed {res['scanned']} {subject.name}{scope} question(s) but set no topics — "
              f"check the key and model under Settings → OCR.", 'info')
    else:
        flash(f'No stand-alone {subject.name}{scope} bank questions to tag.', 'info')
    return redirect(url_for('mock_jamb.bank_analytics', subject_id=subject_id))


@mock_jamb_bp.route('/bank/local-tag', methods=['POST'])
@login_required
def bank_local_tag():
    """Queue local (offline, no-key) embedding-based topic + sub-topic tagging.
    It runs in the background jobs worker (loads a model) — never inline in a web
    request, so torch stays out of the web workers."""
    from utils.mock_bank_local_tag import available
    from utils import jobs
    subject_id = request.form.get('subject_id', type=int)
    subject = db.session.get(Subject, subject_id) if subject_id else None
    if not subject:
        flash('Pick a subject first.', 'error')
        return redirect(url_for('mock_jamb.bank_analytics'))
    if not available():
        flash('Local tagging needs the "sentence-transformers" package installed on the server.', 'error')
        return redirect(url_for('mock_jamb.bank_analytics', subject_id=subject_id))
    if not jobs.async_enabled():
        flash('Local tagging runs in the background — enable the jobs worker '
              '(ASYNC_JOBS / the edusyncra-jobs service) to use it.', 'error')
        return redirect(url_for('mock_jamb.bank_analytics', subject_id=subject_id))
    year = (request.form.get('year') or '').strip() or None
    mode = 'all' if request.form.get('mode') == 'all' else 'untagged'
    jobs.enqueue('bank_local_tag', {'subject_id': subject_id, 'mode': mode, 'year': year})
    flash(f'Local tagging of {subject.name} started in the background — questions get tagged '
          f'as it runs. The first full run can take a while; refresh to see progress.', 'success')
    return redirect(url_for('mock_jamb.bank_analytics', subject_id=subject_id))


def _valid_section(subject, section):
    from utils.jamb_blueprint import sections_for
    keys = {s['section'] for s in sections_for(subject.name)}
    return section if section in keys else None


@mock_jamb_bp.route('/bank/passage/add', methods=['POST'])
@login_required
@csrf_protect
def bank_add_passage():
    subject_id = request.form.get('subject_id', type=int)
    subject = db.session.get(Subject, subject_id) if subject_id else None
    if not subject:
        flash('Choose a subject first.', 'error')
        return redirect(_bank_url(subject_id))
    kind = (request.form.get('kind') or 'comprehension').strip()
    if kind not in MockJAMBPassage.KINDS:
        kind = 'comprehension'
    section = _valid_section(subject, (request.form.get('section') or '').strip())
    body = (request.form.get('body') or '').strip()
    if not body:
        flash('The passage text is required.', 'error')
        return redirect(_bank_url(subject_id))
    nextord = (db.session.query(func.coalesce(func.max(MockJAMBPassage.order), 0))
               .filter(MockJAMBPassage.mock_exam_id.is_(None),
                       MockJAMBPassage.subject_id == subject_id).scalar()) + 1
    db.session.add(MockJAMBPassage(
        mock_exam_id=None, subject_id=subject_id, kind=kind, section=section,
        title=(request.form.get('title') or '').strip() or None, body=body,
        image_url=_save_mock_image(request.files.get('image')), order=nextord))
    db.session.commit()
    flash('Passage added to the bank — now add its questions.', 'success')
    return redirect(_bank_url(subject_id, section))


@mock_jamb_bp.route('/bank/passage/<int:passage_id>/edit', methods=['POST'])
@login_required
@csrf_protect
def bank_edit_passage(passage_id):
    p = db.get_or_404(MockJAMBPassage, passage_id)
    if p.mock_exam_id is not None:
        flash('Not a bank passage.', 'error')
        return redirect(_bank_url(p.subject_id))
    p.title = (request.form.get('title') or '').strip() or None
    if (request.form.get('body') or '').strip():
        p.body = request.form.get('body').strip()
    kind = (request.form.get('kind') or '').strip()
    if kind in MockJAMBPassage.KINDS:
        p.kind = kind
    subject = db.session.get(Subject, p.subject_id)
    sec = _valid_section(subject, (request.form.get('section') or '').strip())
    if sec:
        p.section = sec
    img = _save_mock_image(request.files.get('image'))
    if img:
        p.image_url = img
    db.session.commit()
    flash('Passage updated.', 'success')
    return redirect(_bank_url(p.subject_id, p.section))


@mock_jamb_bp.route('/bank/passage/<int:passage_id>/delete', methods=['POST'])
@login_required
@csrf_protect
def bank_delete_passage(passage_id):
    p = db.get_or_404(MockJAMBPassage, passage_id)
    if p.mock_exam_id is not None:
        flash('Not a bank passage.', 'error')
        return redirect(_bank_url(p.subject_id))
    sid = p.subject_id
    MockJAMBQuestion.query.filter_by(passage_id=p.id).delete()
    db.session.delete(p)
    db.session.commit()
    flash('Passage and its questions removed from the bank.', 'success')
    return redirect(_bank_url(sid))


def _read_bank_question(form, q, files, subject):
    """Populate a bank MockJAMBQuestion (section/exam_body/difficulty in addition
    to the shared fields). Returns an error string or None."""
    err = _read_question(form, q, files)
    if err:
        return err
    if not q.passage_id:   # a passage's questions inherit the passage section
        q.section = _valid_section(subject, (form.get('section') or '').strip())
    eb = (form.get('exam_body') or 'JAMB').strip()
    q.exam_body = eb if eb in ('JAMB', 'WAEC', 'Both') else 'JAMB'
    diff = (form.get('difficulty') or '').strip().lower()
    q.difficulty = diff if diff in ('easy', 'medium', 'hard') else None
    q.exam_year = (form.get('exam_year') or '').strip()[:8] or None
    return None


@mock_jamb_bp.route('/bank/question/add', methods=['POST'])
@login_required
@csrf_protect
def bank_add_question():
    subject_id = request.form.get('subject_id', type=int)
    subject = db.session.get(Subject, subject_id) if subject_id else None
    if not subject:
        flash('Choose a subject first.', 'error')
        return redirect(_bank_url(subject_id))
    passage_id = request.form.get('passage_id', type=int)
    passage = db.session.get(MockJAMBPassage, passage_id) if passage_id else None
    if passage and (passage.mock_exam_id is not None or passage.subject_id != subject_id):
        passage = None
    q = MockJAMBQuestion(mock_exam_id=None, subject_id=subject_id,
                         passage_id=(passage.id if passage else None))
    if passage:
        q.section = passage.section
    err = _read_bank_question(request.form, q, request.files, subject)
    if err:
        flash(err, 'error')
        return redirect(_bank_url(subject_id))
    q.order = (db.session.query(func.coalesce(func.max(MockJAMBQuestion.order), 0))
               .filter(MockJAMBQuestion.mock_exam_id.is_(None),
                       MockJAMBQuestion.subject_id == subject_id).scalar()) + 1
    db.session.add(q)
    db.session.commit()
    flash('Question added to the bank.', 'success')
    return redirect(_bank_url(subject_id, q.section))


@mock_jamb_bp.route('/bank/question/<int:question_id>/edit', methods=['GET', 'POST'])
@login_required
@csrf_protect
def bank_edit_question(question_id):
    from routes.cbt import _subject_topic_tree
    from utils.jamb_blueprint import sections_for
    q = db.get_or_404(MockJAMBQuestion, question_id)
    if q.mock_exam_id is not None:
        flash('Not a bank question.', 'error')
        return redirect(_bank_url(q.subject_id))
    subject = db.session.get(Subject, q.subject_id)
    if request.method == 'POST':
        err = _read_bank_question(request.form, q, request.files, subject)
        if err:
            flash(err, 'error')
            return redirect(url_for('mock_jamb.bank_edit_question', question_id=question_id))
        db.session.commit()
        flash('Question updated.', 'success')
        return redirect(_bank_url(q.subject_id, q.section))
    return render_template('mock_jamb/bank_edit_question.html', q=q, subject=subject,
                           sections=sections_for(subject.name),
                           topic_tree=_subject_topic_tree(q.subject_id),
                           back_url=_bank_url(q.subject_id, q.section))


@mock_jamb_bp.route('/bank/question/<int:question_id>/delete', methods=['POST'])
@login_required
@csrf_protect
def bank_delete_question(question_id):
    q = db.get_or_404(MockJAMBQuestion, question_id)
    if q.mock_exam_id is not None:
        flash('Not a bank question.', 'error')
        return redirect(_bank_url(q.subject_id))
    sid, sec = q.subject_id, q.section
    db.session.delete(q)
    db.session.commit()
    flash('Question deleted from the bank.', 'success')
    return redirect(_bank_url(sid, sec))


@mock_jamb_bp.route('/bank/import', methods=['POST'])
@login_required
@csrf_protect
def bank_import():
    """Bulk-add stand-alone bank questions from pasted rows. Each line (tab- or
    pipe-separated):
    ``question | A | B | C | D | correct | [section] | [topic] | [subtopic] | [year]``
    A default section / year applies when a row omits it. Duplicate questions
    (by text, within this subject's bank) are skipped."""
    import re as _re
    subject_id = request.form.get('subject_id', type=int)
    subject = db.session.get(Subject, subject_id) if subject_id else None
    if not subject:
        flash('Choose a subject first.', 'error')
        return redirect(_bank_url(subject_id))
    default_section = _valid_section(subject, (request.form.get('default_section') or '').strip())
    default_year = (request.form.get('default_year') or '').strip()[:8] or None
    exam_body = (request.form.get('exam_body') or 'JAMB').strip()
    if exam_body not in ('JAMB', 'WAEC', 'Both'):
        exam_body = 'JAMB'
    raw = request.form.get('rows') or ''
    base = (db.session.query(func.coalesce(func.max(MockJAMBQuestion.order), 0))
            .filter(MockJAMBQuestion.mock_exam_id.is_(None),
                    MockJAMBQuestion.subject_id == subject_id).scalar())

    def _norm(t):
        return _re.sub(r'\s+', ' ', (t or '').lower()).strip()
    seen = {_norm(q.question_text) for q in MockJAMBQuestion.query.filter_by(
        subject_id=subject_id, mock_exam_id=None).all()}

    added = skipped = duplicates = 0
    for line in raw.splitlines():
        line = line.strip()
        if not line:
            continue
        parts = [c.strip() for c in (line.split('\t') if '\t' in line else line.split('|'))]
        if len(parts) < 6:
            skipped += 1
            continue
        text, a, b, c, d, correct = parts[:6]
        correct = correct.strip().upper()
        if not text or correct not in ('A', 'B', 'C', 'D'):
            skipped += 1
            continue
        nt = _norm(text)
        if nt in seen:
            duplicates += 1
            continue
        seen.add(nt)
        section = _valid_section(subject, parts[6]) if len(parts) > 6 and parts[6] else default_section
        topic = parts[7] if len(parts) > 7 and parts[7] else None
        subtopic = parts[8] if len(parts) > 8 and parts[8] else None
        year = (parts[9].strip()[:8] if len(parts) > 9 and parts[9].strip() else default_year)
        # optional 11th column: a figure URL (kept as-is; may be hotlink-blocked).
        image = (parts[10].strip()[:300] if len(parts) > 10 and parts[10].strip()
                 and parts[10].strip().lower().startswith(('http://', 'https://', '/')) else None)
        base += 1
        db.session.add(MockJAMBQuestion(
            mock_exam_id=None, subject_id=subject_id, section=section, exam_body=exam_body,
            question_text=text, option_a=a, option_b=b, option_c=c, option_d=d,
            correct_option=correct, marks=1, topic=topic, subtopic=subtopic,
            exam_year=year, image_url=image, source='paste', order=base))
        added += 1
    db.session.commit()
    msg = f'Imported {added} question(s).'
    extra = []
    if duplicates:
        extra.append(f'{duplicates} duplicate(s) skipped')
    if skipped:
        extra.append(f'{skipped} malformed line(s)')
    if extra:
        msg += ' (' + ', '.join(extra) + ')'
    flash(msg, 'success' if added else 'warning')
    return redirect(_bank_url(subject_id, default_section))


# ---------------------------------------------------------------------------
# One-click myschool.ng harvest: scrape questions server-side into the bank.
# Resumable, chunked (the browser calls /step repeatedly).
# ---------------------------------------------------------------------------
@mock_jamb_bp.route('/bank/scrape/start', methods=['POST'])
@login_required
@csrf_protect
def bank_scrape_start():
    from utils.myschool_harvest import start_harvest
    raw = (request.form.get('subject_ids') or '').strip()
    ids = [int(x) for x in raw.split(',') if x.strip().isdigit()]
    subjects = [{'id': s.id, 'name': s.name}
                for s in Subject.query.filter(Subject.id.in_(ids or [-1])).all()]
    if not subjects:
        return jsonify({'error': 'Pick at least one subject to download.'}), 400
    exam = (request.form.get('exam') or 'jamb').strip().lower()
    if exam not in ('jamb', 'waec', 'neco', 'post-utme'):
        exam = 'jamb'
    y1 = request.form.get('year_from', type=int)
    y2 = request.form.get('year_to', type=int)
    return jsonify(start_harvest(subjects, exam=exam, year_min=y1, year_max=y2))


@mock_jamb_bp.route('/bank/scrape/step', methods=['POST'])
@login_required
@csrf_protect
def bank_scrape_step():
    from utils.myschool_harvest import harvest_step
    return jsonify(harvest_step(max_questions=6))


@mock_jamb_bp.route('/bank/scrape/stop', methods=['POST'])
@login_required
@csrf_protect
def bank_scrape_stop():
    from utils.myschool_harvest import pause_harvest
    return jsonify(pause_harvest())


@mock_jamb_bp.route('/bank/scrape/resume', methods=['POST'])
@login_required
@csrf_protect
def bank_scrape_resume():
    from utils.myschool_harvest import resume_harvest
    return jsonify(resume_harvest())


@mock_jamb_bp.route('/bank/scrape/status')
@login_required
def bank_scrape_status():
    from utils.myschool_harvest import _public, get_state
    return jsonify(_public(get_state()))


@mock_jamb_bp.route('/bank/delete-bulk', methods=['POST'])
@login_required
@csrf_protect
def bank_delete_bulk():
    """Bulk-delete stand-alone bank questions for one subject, optionally scoped
    to a source (e.g. old 'aloc' imports), the untagged ones (no topic), or a
    year — so you can clear a subject and re-download cleanly."""
    from models import MockJAMBAnswer
    subject_id = request.form.get('subject_id', type=int)
    subject = db.session.get(Subject, subject_id) if subject_id else None
    if not subject:
        flash('Choose a subject first.', 'error')
        return redirect(_bank_url(subject_id))
    scope = (request.form.get('scope') or 'all').strip()
    base = MockJAMBQuestion.query.filter(
        MockJAMBQuestion.mock_exam_id.is_(None),
        MockJAMBQuestion.subject_id == subject_id,
        MockJAMBQuestion.passage_id.is_(None))
    label = 'all bank'
    if scope == 'untagged':
        base = base.filter((MockJAMBQuestion.topic.is_(None)) | (MockJAMBQuestion.topic == ''))
        label = 'untagged (no topic)'
    elif scope == 'source':
        src = (request.form.get('source') or '').strip()
        base = base.filter(MockJAMBQuestion.source == (src or None))
        label = f"'{src or 'unknown'}'-source"
    elif scope == 'year':
        yr = (request.form.get('year') or '').strip()
        base = base.filter(MockJAMBQuestion.exam_year == yr)
        label = f'year {yr}'

    ids = [qid for (qid,) in base.with_entities(MockJAMBQuestion.id).all()]
    if not ids:
        flash('No matching questions to delete.', 'warning')
        return redirect(_bank_url(subject_id))
    # Delete in chunks (SQLite caps bound variables), clearing dependent answers
    # first so a question served in a past attempt can still be removed.
    deleted = 0
    for i in range(0, len(ids), 400):
        chunk = ids[i:i + 400]
        MockJAMBAnswer.query.filter(MockJAMBAnswer.question_id.in_(chunk)).delete(synchronize_session=False)
        deleted += MockJAMBQuestion.query.filter(MockJAMBQuestion.id.in_(chunk)).delete(synchronize_session=False)
    db.session.commit()
    from utils.audit import log_action
    log_action('mock_jamb.bank_delete_bulk',
               detail=f'subject={subject_id} scope={label} deleted={deleted}',
               target_type='subject', target_id=subject_id)
    flash(f'Deleted {deleted} {label} question(s) from {subject.name}.', 'success')
    return redirect(_bank_url(subject_id))


@mock_jamb_bp.route('/bank/needs-images')
@login_required
def bank_needs_images():
    """Review queue of bank questions that refer to a figure we couldn't fetch.
    Filterable by subject, year and exam body; upload the diagram (or dismiss the
    flag) inline. Flagged questions are held out of exams until resolved."""
    subjects = _mock_subjects()
    subject_id = request.args.get('subject_id', type=int)
    year = (request.args.get('year') or '').strip()
    exam_body = (request.args.get('exam_body') or '').strip()
    page = max(1, request.args.get('page', 1, type=int))

    def _flagged():
        return MockJAMBQuestion.query.filter(
            MockJAMBQuestion.mock_exam_id.is_(None),
            MockJAMBQuestion.needs_image.is_(True))

    q = _flagged()
    if subject_id:
        q = q.filter(MockJAMBQuestion.subject_id == subject_id)
    if year:
        q = q.filter(MockJAMBQuestion.exam_year == year)
    if exam_body:
        q = q.filter(MockJAMBQuestion.exam_body == exam_body)
    q = q.order_by(MockJAMBQuestion.subject_id, MockJAMBQuestion.exam_year,
                   MockJAMBQuestion.id)
    pagination = q.paginate(page=page, per_page=20, error_out=False)

    years = [y for (y,) in db.session.query(MockJAMBQuestion.exam_year)
             .filter(MockJAMBQuestion.mock_exam_id.is_(None),
                     MockJAMBQuestion.needs_image.is_(True),
                     MockJAMBQuestion.exam_year.isnot(None))
             .distinct().order_by(MockJAMBQuestion.exam_year.desc()).all() if y]
    subj_name = {s.id: s.name for s in subjects}
    total = _flagged().count()
    return render_template('mock_jamb/needs_images.html',
                           subjects=subjects, subject_id=subject_id, year=year,
                           exam_body=exam_body, years=years, subj_name=subj_name,
                           pagination=pagination, questions=pagination.items,
                           total=total, index_url=url_for('mock_jamb.index'),
                           bank_url=url_for('mock_jamb.bank'))


@mock_jamb_bp.route('/bank/question/<int:question_id>/set-image', methods=['POST'])
@login_required
@csrf_protect
def bank_set_question_image(question_id):
    """Attach a figure to a flagged bank question (clearing needs_image so it
    re-enters the pool), or dismiss the flag when no image is actually needed."""
    q = db.get_or_404(MockJAMBQuestion, question_id)
    if q.mock_exam_id is not None:
        flash('Not a bank question.', 'error')
        return redirect(url_for('mock_jamb.bank_needs_images'))
    back = request.referrer or url_for('mock_jamb.bank_needs_images')
    if request.form.get('dismiss'):
        q.needs_image = False
        db.session.commit()
        flash('Marked as not needing an image — it can now be used in exams.', 'success')
        return redirect(back)
    url = _save_mock_image(request.files.get('image'))
    if not url:
        flash('Choose an image file to upload.', 'error')
        return redirect(back)
    q.image_url = url
    q.needs_image = False
    db.session.commit()
    flash('Image added — the question is now in the exam pool.', 'success')
    return redirect(back)


@mock_jamb_bp.route('/bank/needs-images/bulk', methods=['POST'])
@login_required
@csrf_protect
def bank_needs_images_bulk():
    """Act on several flagged questions at once from the needs-images queue:
    ``action`` = ``dismiss`` (mark selected as not needing an image),
    ``dismiss_all`` (every flagged question matching the current filters) or
    ``delete`` (remove the selected questions)."""
    from models import MockJAMBAnswer
    action = (request.form.get('action') or '').strip()
    back = request.referrer or url_for('mock_jamb.bank_needs_images')
    base = MockJAMBQuestion.query.filter(
        MockJAMBQuestion.mock_exam_id.is_(None),
        MockJAMBQuestion.needs_image.is_(True))

    if action == 'dismiss_all':
        subject_id = request.form.get('subject_id', type=int)
        year = (request.form.get('year') or '').strip()
        exam_body = (request.form.get('exam_body') or '').strip()
        if subject_id:
            base = base.filter(MockJAMBQuestion.subject_id == subject_id)
        if year:
            base = base.filter(MockJAMBQuestion.exam_year == year)
        if exam_body:
            base = base.filter(MockJAMBQuestion.exam_body == exam_body)
        ids = [qid for (qid,) in base.with_entities(MockJAMBQuestion.id).all()]
    else:
        raw = request.form.get('question_ids') or ''
        ids = [int(x) for x in raw.split(',') if x.strip().isdigit()]
        ids = [qid for (qid,) in base.filter(MockJAMBQuestion.id.in_(ids))
               .with_entities(MockJAMBQuestion.id).all()] if ids else []

    if not ids:
        flash('Select at least one question first.', 'error')
        return redirect(back)

    if action == 'delete':
        deleted = 0
        for i in range(0, len(ids), 400):
            chunk = ids[i:i + 400]
            MockJAMBAnswer.query.filter(MockJAMBAnswer.question_id.in_(chunk)).delete(synchronize_session=False)
            deleted += MockJAMBQuestion.query.filter(MockJAMBQuestion.id.in_(chunk)).delete(synchronize_session=False)
        db.session.commit()
        flash(f'Deleted {deleted} question(s).', 'success')
    else:
        n = 0
        for i in range(0, len(ids), 400):
            n += (MockJAMBQuestion.query.filter(MockJAMBQuestion.id.in_(ids[i:i + 400]))
                  .update({MockJAMBQuestion.needs_image: False}, synchronize_session=False))
        db.session.commit()
        flash(f'Marked {n} question(s) as not needing an image — they can now be used in exams.', 'success')
    return redirect(back)


@mock_jamb_bp.route('/bank/assign-novel', methods=['POST'])
@login_required
@csrf_protect
def bank_assign_novel():
    """Tag the bank's Novel-section questions for a subject with a novel title
    (stored in ``topic``), so a mock naming that novel serves only these."""
    subject_id = request.form.get('subject_id', type=int)
    subject = db.session.get(Subject, subject_id) if subject_id else None
    if not subject:
        flash('Choose a subject first.', 'error')
        return redirect(_bank_url(subject_id))
    novel = (request.form.get('novel_title') or '').strip()
    scope = request.form.get('scope')
    empty_topic = db.or_(MockJAMBQuestion.topic.is_(None), MockJAMBQuestion.topic == '')
    make_novel = (scope == 'untagged_any')
    if make_novel:
        # Recover the note-less novel questions: any stand-alone question with no
        # topic (e.g. "Who makes this statement?") is treated as a Novel question —
        # move it into the Novel section and tag it with the recommended text.
        q = (MockJAMBQuestion.query.filter_by(subject_id=subject_id, mock_exam_id=None)
             .filter(MockJAMBQuestion.passage_id.is_(None), empty_topic))
    else:
        q = MockJAMBQuestion.query.filter_by(
            subject_id=subject_id, mock_exam_id=None, section='novel')
        if scope == 'untagged':
            q = q.filter(empty_topic)
    n = 0
    for row in q.all():
        row.topic = novel or None
        if make_novel:
            row.section = 'novel'
        n += 1
    db.session.commit()
    if make_novel:
        flash(f'Moved {n} untagged question(s) into the Novel section and tagged them "{novel}".', 'success')
    else:
        flash(f'Tagged {n} novel question(s) with "{novel}".' if novel
              else f'Cleared the novel tag on {n} question(s).', 'success')
    return redirect(_bank_url(subject_id, 'novel'))


@mock_jamb_bp.route('/exam/<int:exam_id>/questions')
@login_required
def questions(exam_id):
    """Question manager for a mock exam, filtered to one subject: passages with
    their questions, plus stand-alone questions. Feeds the in-app sitting."""
    from routes.cbt import _subject_topic_tree
    exam = db.get_or_404(MockJAMBExam, exam_id)
    require_branch_access(exam.branch_id)
    subjects = _mock_subjects()
    subject_id = request.args.get('subject_id', type=int) or (subjects[0].id if subjects else None)
    subject = db.session.get(Subject, subject_id) if subject_id else None
    page = max(1, request.args.get('page', 1, type=int))
    passages, standalone, qcount, pagination = [], [], 0, None
    if subject_id:
        prows = (MockJAMBPassage.query.filter_by(mock_exam_id=exam_id, subject_id=subject_id)
                 .order_by(MockJAMBPassage.order, MockJAMBPassage.id).all())
        passages = [{'p': p, 'questions': p.questions.order_by(
            MockJAMBQuestion.order, MockJAMBQuestion.id).all()} for p in prows]
        pagination = (MockJAMBQuestion.query.filter_by(
            mock_exam_id=exam_id, subject_id=subject_id, passage_id=None)
            .order_by(MockJAMBQuestion.order, MockJAMBQuestion.id)
            .paginate(page=page, per_page=25, error_out=False))
        standalone = pagination.items
        qcount = MockJAMBQuestion.query.filter_by(
            mock_exam_id=exam_id, subject_id=subject_id).count()
    return render_template('mock_jamb/questions.html', exam=exam, subjects=subjects,
                           subject=subject, subject_id=subject_id, passages=passages,
                           standalone=standalone, qcount=qcount, pagination=pagination,
                           topic_tree=_subject_topic_tree(subject_id),
                           syllabus_url=url_for('cbt.syllabus', subject_id=subject_id or ''))


def _q_url(exam_id, subject_id):
    return url_for('mock_jamb.questions', exam_id=exam_id, subject_id=subject_id or '')


@mock_jamb_bp.route('/exam/<int:exam_id>/passages/add', methods=['POST'])
@login_required
@csrf_protect
def add_passage(exam_id):
    exam = db.get_or_404(MockJAMBExam, exam_id)
    require_branch_access(exam.branch_id)
    subject_id = request.form.get('subject_id', type=int)
    kind = (request.form.get('kind') or 'comprehension').strip()
    if kind not in MockJAMBPassage.KINDS:
        kind = 'comprehension'
    body = (request.form.get('body') or '').strip()
    if not subject_id or not body:
        flash('A subject and the passage text are required.', 'error')
        return redirect(_q_url(exam_id, subject_id))
    nextord = (db.session.query(func.coalesce(func.max(MockJAMBPassage.order), 0))
               .filter(MockJAMBPassage.mock_exam_id == exam_id,
                       MockJAMBPassage.subject_id == subject_id).scalar()) + 1
    db.session.add(MockJAMBPassage(
        mock_exam_id=exam_id, subject_id=subject_id, kind=kind,
        title=(request.form.get('title') or '').strip() or None, body=body,
        image_url=_save_mock_image(request.files.get('image')), order=nextord))
    db.session.commit()
    flash('Passage added — now add its questions.', 'success')
    return redirect(_q_url(exam_id, subject_id))


@mock_jamb_bp.route('/passage/<int:passage_id>/edit', methods=['POST'])
@login_required
@csrf_protect
def edit_passage(passage_id):
    p = db.get_or_404(MockJAMBPassage, passage_id)
    require_branch_access(p.exam.branch_id)
    p.title = (request.form.get('title') or '').strip() or None
    if (request.form.get('body') or '').strip():
        p.body = request.form.get('body').strip()
    kind = (request.form.get('kind') or '').strip()
    if kind in MockJAMBPassage.KINDS:
        p.kind = kind
    img = _save_mock_image(request.files.get('image'))
    if img:
        p.image_url = img
    db.session.commit()
    flash('Passage updated.', 'success')
    return redirect(_q_url(p.mock_exam_id, p.subject_id))


@mock_jamb_bp.route('/passage/<int:passage_id>/delete', methods=['POST'])
@login_required
@csrf_protect
def delete_passage(passage_id):
    p = db.get_or_404(MockJAMBPassage, passage_id)
    require_branch_access(p.exam.branch_id)
    eid, sid = p.mock_exam_id, p.subject_id
    MockJAMBQuestion.query.filter_by(passage_id=p.id).delete()   # its questions go too
    db.session.delete(p)
    db.session.commit()
    flash('Passage and its questions removed.', 'success')
    return redirect(_q_url(eid, sid))


def _read_question(form, q, files):
    """Populate a MockJAMBQuestion from a submitted form. Returns an error string
    or None. A question attached to a passage is validated to belong to the same
    exam+subject, so a comprehension item can never be orphaned from its passage."""
    text = (form.get('question_text') or '').strip()
    correct = (form.get('correct_option') or '').strip().upper()
    if not text or correct not in ('A', 'B', 'C', 'D'):
        return 'Question text and a correct option (A–D) are required.'
    q.question_text = text
    q.correct_option = correct
    q.topic = (form.get('topic') or '').strip() or None
    q.subtopic = (form.get('subtopic') or '').strip() or None
    q.option_a = (form.get('option_a') or '').strip()
    q.option_b = (form.get('option_b') or '').strip()
    q.option_c = (form.get('option_c') or '').strip()
    q.option_d = (form.get('option_d') or '').strip()
    q.marks = form.get('marks', type=float) or 1
    img = _save_mock_image(files.get('image'))
    if img:
        q.image_url = img
    return None


@mock_jamb_bp.route('/exam/<int:exam_id>/questions/add', methods=['POST'])
@login_required
@csrf_protect
def add_mock_question(exam_id):
    exam = db.get_or_404(MockJAMBExam, exam_id)
    require_branch_access(exam.branch_id)
    subject_id = request.form.get('subject_id', type=int)
    passage_id = request.form.get('passage_id', type=int)
    if not subject_id:
        flash('Choose a subject first.', 'error')
        return redirect(_q_url(exam_id, subject_id))
    passage = db.session.get(MockJAMBPassage, passage_id) if passage_id else None
    if passage and (passage.mock_exam_id != exam_id or passage.subject_id != subject_id):
        passage = None
    q = MockJAMBQuestion(mock_exam_id=exam_id, subject_id=subject_id,
                         passage_id=(passage.id if passage else None))
    err = _read_question(request.form, q, request.files)
    if err:
        flash(err, 'error')
        return redirect(_q_url(exam_id, subject_id))
    q.order = (db.session.query(func.coalesce(func.max(MockJAMBQuestion.order), 0))
               .filter(MockJAMBQuestion.mock_exam_id == exam_id,
                       MockJAMBQuestion.subject_id == subject_id).scalar()) + 1
    db.session.add(q)
    db.session.commit()
    flash('Question added.', 'success')
    return redirect(_q_url(exam_id, subject_id))


@mock_jamb_bp.route('/question/<int:question_id>/edit', methods=['GET', 'POST'])
@login_required
@csrf_protect
def edit_mock_question(question_id):
    from routes.cbt import _subject_topic_tree
    q = db.get_or_404(MockJAMBQuestion, question_id)
    require_branch_access(q.exam.branch_id)
    if request.method == 'POST':
        err = _read_question(request.form, q, request.files)
        if err:
            flash(err, 'error')
            return redirect(url_for('mock_jamb.edit_mock_question', question_id=question_id))
        db.session.commit()
        flash('Question updated.', 'success')
        return redirect(_q_url(q.mock_exam_id, q.subject_id))
    return render_template('mock_jamb/edit_question.html', q=q, exam=q.exam,
                           topic_tree=_subject_topic_tree(q.subject_id),
                           back_url=_q_url(q.mock_exam_id, q.subject_id))


@mock_jamb_bp.route('/question/<int:question_id>/delete', methods=['POST'])
@login_required
@csrf_protect
def delete_mock_question(question_id):
    q = db.get_or_404(MockJAMBQuestion, question_id)
    require_branch_access(q.exam.branch_id)
    eid, sid = q.mock_exam_id, q.subject_id
    db.session.delete(q)
    db.session.commit()
    flash('Question deleted.', 'success')
    return redirect(_q_url(eid, sid))


# =============================================================================
# EXPORT FUNCTIONALITY
# =============================================================================

@mock_jamb_bp.route('/exam/<int:exam_id>/export')
@login_required
def export_results(exam_id):
    """Export exam results to Excel"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

    exam = db.get_or_404(MockJAMBExam, exam_id)
    require_branch_access(exam.branch_id)
    results = MockJAMBResult.query.filter_by(mock_exam_id=exam_id).join(Student).order_by(
        MockJAMBResult.total_score.desc()
    ).all()
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Mock JAMB Results"
    
    # Styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1a5f4a", end_color="1a5f4a", fill_type="solid")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Title
    ws.merge_cells('A1:M1')
    ws['A1'] = f"{exam.name} - Results"
    ws['A1'].font = Font(bold=True, size=14)
    ws['A1'].alignment = Alignment(horizontal='center')
    
    ws.merge_cells('A2:M2')
    ws['A2'] = f"Date: {exam.exam_date.strftime('%d %B %Y')}"
    ws['A2'].alignment = Alignment(horizontal='center')
    
    # Headers
    headers = ['Rank', 'Student ID', 'Name', 'Subject 1', 'Score 1', 'Subject 2', 'Score 2',
               'Subject 3', 'Score 3', 'Subject 4', 'Score 4', 'Total Score', 'Performance']
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border
        cell.alignment = Alignment(horizontal='center')
    
    # Data
    for row_idx, result in enumerate(results, 5):
        ws.cell(row=row_idx, column=1, value=row_idx - 4).border = border
        ws.cell(row=row_idx, column=2, value=result.student.student_id).border = border
        ws.cell(row=row_idx, column=3, value=result.student.full_name).border = border
        ws.cell(row=row_idx, column=4, value=result.subject1 or '').border = border
        ws.cell(row=row_idx, column=5, value=result.subject1_score or 0).border = border
        ws.cell(row=row_idx, column=6, value=result.subject2 or '').border = border
        ws.cell(row=row_idx, column=7, value=result.subject2_score or 0).border = border
        ws.cell(row=row_idx, column=8, value=result.subject3 or '').border = border
        ws.cell(row=row_idx, column=9, value=result.subject3_score or 0).border = border
        ws.cell(row=row_idx, column=10, value=result.subject4 or '').border = border
        ws.cell(row=row_idx, column=11, value=result.subject4_score or 0).border = border
        ws.cell(row=row_idx, column=12, value=result.total_score).border = border
        ws.cell(row=row_idx, column=13, value=result.performance_level).border = border
    
    # Adjust column widths
    ws.column_dimensions['A'].width = 6
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 25
    for col in ['D', 'F', 'H', 'J']:
        ws.column_dimensions[col].width = 18
    for col in ['E', 'G', 'I', 'K']:
        ws.column_dimensions[col].width = 8
    ws.column_dimensions['L'].width = 12
    ws.column_dimensions['M'].width = 15
    
    filename = f"mock_jamb_{exam.exam_number}_{exam.session.name.replace('/', '_')}.xlsx"

    ws.delete_cols(2)        # drop the admission-number column from the printout
    return xlsx_response(wb, filename)


# =============================================================================
# API ENDPOINTS
# =============================================================================

@mock_jamb_bp.route('/api/exam/<int:exam_id>/stats')
@login_required
def api_exam_stats(exam_id):
    """API endpoint for exam statistics (for charts)"""
    stats = MockJAMBAnalytics.get_exam_statistics(exam_id)
    
    if not stats or not stats['statistics']:
        return jsonify({'error': 'No data available'}), 404
    
    return jsonify({
        'exam_name': stats['exam'].name,
        'student_count': stats['student_count'],
        'statistics': stats['statistics'],
        'distribution': stats['distribution'],
        'subject_analysis': stats['subject_analysis']
    })


@mock_jamb_bp.route('/api/student/<int:student_id>/progress')
@login_required
def api_student_progress(student_id):
    """API endpoint for student progress (for charts)"""
    require_branch_access(db.get_or_404(Student, student_id).branch_id)
    session_id = request.args.get('session_id', type=int)
    progress = MockJAMBAnalytics.get_student_progress(student_id, session_id)
    
    if not progress:
        return jsonify({'error': 'No data available'}), 404
    
    return jsonify(progress)


# =============================================================================
# ADMIN: publish an exam for the online sitting + set its duration
# =============================================================================

@mock_jamb_bp.route('/exam/<int:exam_id>/blueprint', methods=['GET', 'POST'])
@login_required
@csrf_protect
def exam_blueprint(exam_id):
    """Per-mock JAMB blueprint editor: tune how many questions each subject draws
    from each section. A value equal to the JAMB default stores nothing; real
    deviations are kept as a compact override on the mock."""
    import json
    from utils.jamb_blueprint import JAMB_BLUEPRINT
    exam = db.get_or_404(MockJAMBExam, exam_id)
    require_branch_access(exam.branch_id)
    if request.method == 'POST':
        override = {}
        for subj_key, bp in JAMB_BLUEPRINT.items():
            for s in bp['sections']:
                raw = request.form.get(f'{subj_key}__{s["section"]}')
                if raw in (None, ''):
                    continue
                try:
                    n = max(0, int(raw))
                except (TypeError, ValueError):
                    continue
                if n != s['count']:               # only store real deviations
                    override.setdefault(subj_key, {})[s['section']] = n
        exam.blueprint = json.dumps(override) if override else None
        exam.novel_title = (request.form.get('novel_title') or '').strip() or None
        sm = (request.form.get('source_mode') or 'bank').strip()
        exam.source_mode = sm if sm in ('bank', 'manual') else 'bank'
        levels = request.form.getlist('eligible_levels')
        exam.eligible_levels = ','.join(x.strip() for x in levels if x.strip()) or None
        db.session.commit()
        flash('Blueprint saved.' if override else 'Reset to the JAMB default.', 'success')
        return redirect(url_for('mock_jamb.exam_blueprint', exam_id=exam_id))

    current = {}
    if exam.blueprint:
        try:
            current = json.loads(exam.blueprint)
        except Exception:
            current = {}
    subjects_view = []
    for subj_key, bp in JAMB_BLUEPRINT.items():
        ov = current.get(subj_key, {})
        secs = [{'section': s['section'], 'label': s['label'], 'passage': s['passage'],
                 'count': ov.get(s['section'], s['count']), 'default': s['count']}
                for s in bp['sections']]
        subjects_view.append({'key': subj_key, 'name': subj_key.title(),
                              'total': sum(x['count'] for x in secs),
                              'default_total': sum(x['default'] for x in secs),
                              'sections': secs, 'customised': subj_key in current})
    # Well-known JAMB recommended novels (suggestions; the field is free text).
    known_novels = ['The Life Changer — Khadija Abubakar Jalli',
                    'Sweet Sixteen — Bolaji Abdullahi',
                    'In Dependence — Sarah Ladipo Manyika',
                    'The Last Days at Forcados High School — A.H. Mohammed']
    from models import SchoolClass
    from utils.helpers import get_sss3_class
    all_classes = SchoolClass.query.order_by(SchoolClass.level, SchoolClass.name).all()
    selected_levels = [x.strip() for x in (exam.eligible_levels or '').split(',') if x.strip()]
    sss3 = get_sss3_class()
    return render_template('mock_jamb/blueprint.html', exam=exam, subjects=subjects_view,
                           novel_title=exam.novel_title or '', known_novels=known_novels,
                           source_mode=exam.source_mode or 'bank',
                           all_classes=all_classes, selected_levels=selected_levels,
                           default_level=(sss3.name if sss3 else 'SSS3'),
                           urls={'questions': url_for('mock_jamb.questions', exam_id=exam.id),
                                 'bank': url_for('mock_jamb.bank'),
                                 'view': url_for('mock_jamb.view_exam', exam_id=exam.id),
                                 'self': url_for('mock_jamb.exam_blueprint', exam_id=exam.id)})


@mock_jamb_bp.route('/exam/<int:exam_id>/publish', methods=['POST'])
@login_required
@csrf_protect
def toggle_publish(exam_id):
    exam = db.get_or_404(MockJAMBExam, exam_id)
    require_branch_access(exam.branch_id)
    if not exam.is_published and exam.questions.count() == 0:
        # A bank-source mock owns no questions — it draws a unique paper per student
        # from the central bank — so it can publish as long as the bank holds
        # questions. Only a 'manual' mock (or an empty bank) blocks publishing.
        bank_has = db.session.query(MockJAMBQuestion.id).filter(
            MockJAMBQuestion.mock_exam_id.is_(None)).first() is not None
        if not (exam.source_mode == 'bank' and bank_has):
            if exam.source_mode != 'bank' and bank_has:
                msg = ('This mock is set to use its OWN questions but has none. Set its source to '
                       '"Question bank" so it draws a different paper per student from the bank, '
                       'or add questions to the mock.')
            else:
                msg = ('The question bank has no questions yet — add or download questions in the '
                       'Question Bank first, then publish. The mock draws each student a unique '
                       'paper from the bank.')
            flash(msg, 'error')
            return redirect(url_for('mock_jamb.questions', exam_id=exam_id))
    dur = request.form.get('duration_minutes', type=int)
    if dur and 5 <= dur <= 300:
        exam.duration_minutes = dur
    if 'questions_per_subject' in request.form:
        qps = request.form.get('questions_per_subject', type=int)
        exam.questions_per_subject = qps if (qps and qps > 0) else None
    exam.is_published = not exam.is_published
    db.session.commit()
    flash('Online sitting opened for students.' if exam.is_published
          else 'Online sitting closed.', 'success')
    return redirect(url_for('mock_jamb.questions', exam_id=exam_id))


# =============================================================================
# STUDENT PORTAL: sit the mock JAMB online (shares the CBT portal login)
# =============================================================================

def _portal_guard(exam):
    """Ensure the exam is published and the student may sit it: same branch and
    an eligible class (SSS3 by default, configurable per mock)."""
    from routes.cbt import _current_student
    from utils.mock_jamb_sitting import student_eligible
    student = _current_student()
    if not student:
        return None, None
    if not exam or not exam.is_published or not exam.is_active:
        return student, None
    if exam.branch_id and student.branch_id and exam.branch_id != student.branch_id:
        return student, None
    if not student_eligible(exam, student):
        return student, None
    return student, exam


@mock_jamb_portal_bp.route('/')
def portal_list():
    from routes.cbt import cbt_login_required, _current_student
    @cbt_login_required
    def _inner():
        from models import MockJAMBAttempt
        from utils.mock_jamb_sitting import candidate_subject_ids, student_eligible
        student = _current_student()
        exams = (MockJAMBExam.query.filter_by(is_published=True, is_active=True)
                 .order_by(MockJAMBExam.exam_date.desc()).all())
        rows = []
        for e in exams:
            if e.branch_id and student.branch_id and e.branch_id != student.branch_id:
                continue
            if not student_eligible(e, student):
                continue
            subs = candidate_subject_ids(e, student)
            if not subs:
                continue
            att = MockJAMBAttempt.query.filter_by(mock_exam_id=e.id, student_id=student.id).first()
            rows.append({'exam': e, 'subjects': len(subs),
                         'submitted': bool(att and att.status == 'Submitted'),
                         'in_progress': bool(att and att.status != 'Submitted'),
                         'score': att.total_score if att and att.status == 'Submitted' else None})
        return render_template('mock_jamb/portal_list.html', student=student, rows=rows)
    return _inner()


@mock_jamb_portal_bp.route('/<int:exam_id>')
def portal_sit(exam_id):
    from routes.cbt import cbt_login_required
    @cbt_login_required
    def _inner():
        from models import MockJAMBAttempt
        from utils.mock_jamb_sitting import candidate_subject_ids, sitting_payload
        exam = db.session.get(MockJAMBExam, exam_id)
        student, ok = _portal_guard(exam)
        if not ok:
            flash('This mock is not open for you.', 'error')
            return redirect(url_for('mock_jamb_portal.portal_list'))
        subject_ids = candidate_subject_ids(exam, student)
        if not subject_ids:
            flash('You have no subjects to sit in this mock.', 'error')
            return redirect(url_for('mock_jamb_portal.portal_list'))
        # Safety net: if the student's timer already elapsed while the tab was
        # closed, force-submit (grade) it now instead of resuming a dead attempt.
        from utils.mock_jamb_sitting import auto_submit_expired
        auto_submit_expired(exam=exam, student=student)
        att = MockJAMBAttempt.query.filter_by(mock_exam_id=exam.id, student_id=student.id).first()
        if att and (att.status == 'Submitted' or att.submitted_at):
            return redirect(url_for('mock_jamb_portal.portal_done', exam_id=exam.id))
        if not att:
            att = MockJAMBAttempt(mock_exam_id=exam.id, student_id=student.id,
                                  duration_minutes=exam.duration_minutes or 120)
            db.session.add(att); db.session.commit()
        saved = {a.question_id: a.selected_option for a in att.answers}
        # seconds left = duration - elapsed
        import datetime as _dt
        elapsed = (_dt.datetime.now() - att.started_at).total_seconds() if att.started_at else 0
        remaining = max(0, int((att.duration_minutes or 120) * 60 - elapsed))
        from utils.mock_jamb_sitting import is_calculation_subject
        payload = sitting_payload(exam, subject_ids, att)
        # Persist the freshly-drawn paper (cached on the attempt) so reloads/resume
        # and grading reuse it instead of re-scanning the bank on every load.
        try:
            db.session.commit()
        except Exception:
            db.session.rollback()
        show_calc = any(s.get('subject') and is_calculation_subject(s['subject'].name)
                        for s in payload)
        return render_template('mock_jamb/portal_sit.html', exam=exam, student=student,
                               subjects=payload, saved=saved, show_calc=show_calc,
                               attempt=att, remaining=remaining)
    return _inner()


@mock_jamb_portal_bp.route('/<int:exam_id>/save', methods=['POST'])
def portal_save(exam_id):
    from routes.cbt import cbt_login_required
    @cbt_login_required
    def _inner():
        from models import MockJAMBAttempt, MockJAMBAnswer, MockJAMBQuestion
        exam = db.session.get(MockJAMBExam, exam_id)
        student, ok = _portal_guard(exam)
        if not ok:
            return jsonify({'error': 'closed'}), 403
        att = MockJAMBAttempt.query.filter_by(mock_exam_id=exam.id, student_id=student.id).first()
        if not att or att.status == 'Submitted':
            return jsonify({'error': 'not-active'}), 400
        qid = request.form.get('question_id', type=int)
        opt = (request.form.get('option') or '').strip().upper()
        q = db.session.get(MockJAMBQuestion, qid)
        if not q or opt not in ('A', 'B', 'C', 'D'):
            return jsonify({'error': 'bad'}), 400
        # The question must belong to THIS mock's pool — its own rows for a legacy
        # mock, or the shared bank (mock_exam_id NULL) for a bank-drawn mock. (The
        # earlier `q.mock_exam_id == exam.id` check silently rejected every
        # bank-drawn answer, since those questions are owned by no exam.)
        from utils.mock_jamb_sitting import _pool_condition
        in_pool = db.session.query(MockJAMBQuestion.id).filter(
            MockJAMBQuestion.id == qid, _pool_condition(MockJAMBQuestion, exam)).first()
        if not in_pool:
            return jsonify({'error': 'bad'}), 400
        ans = MockJAMBAnswer.query.filter_by(attempt_id=att.id, question_id=qid).first()
        if not ans:
            ans = MockJAMBAnswer(attempt_id=att.id, question_id=qid)
            db.session.add(ans)
        ans.selected_option = opt
        ans.is_correct = (opt == (q.correct_option or '').upper())
        db.session.commit()
        return jsonify({'ok': True})
    return _inner()


@mock_jamb_portal_bp.route('/<int:exam_id>/save-batch', methods=['POST'])
def portal_save_batch(exam_id):
    """Save several answers in one request (the sitting batches queued changes).
    Accepts ``answers`` as ``qid:opt,qid:opt`` (form) or a JSON ``{qid: opt}`` map.
    Each question is validated against THIS mock's pool (own rows or the bank)."""
    from routes.cbt import cbt_login_required
    @cbt_login_required
    def _inner():
        from models import MockJAMBAttempt, MockJAMBAnswer, MockJAMBQuestion
        from utils.mock_jamb_sitting import _pool_condition
        exam = db.session.get(MockJAMBExam, exam_id)
        student, ok = _portal_guard(exam)
        if not ok:
            return jsonify({'error': 'closed'}), 403
        att = MockJAMBAttempt.query.filter_by(mock_exam_id=exam.id, student_id=student.id).first()
        if not att or att.status == 'Submitted':
            return jsonify({'error': 'not-active'}), 400

        raw = {}
        payload = request.get_json(silent=True) if request.is_json else None
        if isinstance(payload, dict):
            raw = payload.get('answers') if isinstance(payload.get('answers'), dict) else payload
        else:
            for tok in (request.form.get('answers') or '').split(','):
                if ':' in tok:
                    k, v = tok.split(':', 1)
                    raw[k.strip()] = v.strip()

        norm = {}
        for k, v in (raw or {}).items():
            try:
                qid = int(k)
            except (TypeError, ValueError):
                continue
            opt = (str(v) or '').strip().upper()
            if opt in ('A', 'B', 'C', 'D'):
                norm[qid] = opt
        if not norm:
            return jsonify({'ok': True, 'saved': 0})

        qs = {q.id: q for q in MockJAMBQuestion.query.filter(
            MockJAMBQuestion.id.in_(list(norm)), _pool_condition(MockJAMBQuestion, exam)).all()}
        existing = {a.question_id: a for a in MockJAMBAnswer.query.filter(
            MockJAMBAnswer.attempt_id == att.id,
            MockJAMBAnswer.question_id.in_(list(norm))).all()}
        saved = 0
        for qid, opt in norm.items():
            q = qs.get(qid)
            if not q:
                continue
            ans = existing.get(qid)
            if not ans:
                ans = MockJAMBAnswer(attempt_id=att.id, question_id=qid)
                db.session.add(ans)
            ans.selected_option = opt
            ans.is_correct = (opt == (q.correct_option or '').upper())
            saved += 1
        db.session.commit()
        return jsonify({'ok': True, 'saved': saved})
    return _inner()


@mock_jamb_portal_bp.route('/<int:exam_id>/submit', methods=['POST'])
def portal_submit(exam_id):
    from routes.cbt import cbt_login_required
    @cbt_login_required
    def _inner():
        from models import MockJAMBAttempt
        from utils.mock_jamb_sitting import grade_attempt
        exam = db.session.get(MockJAMBExam, exam_id)
        from routes.cbt import _current_student
        student = _current_student()
        att = (MockJAMBAttempt.query.filter_by(mock_exam_id=exam_id, student_id=student.id).first()
               if (exam and student) else None)
        if not att:
            flash('No attempt to submit.', 'error')
            return redirect(url_for('mock_jamb_portal.portal_list'))
        if att.status != 'Submitted':
            grade_attempt(att)
        return redirect(url_for('mock_jamb_portal.portal_done', exam_id=exam_id))
    return _inner()


@mock_jamb_portal_bp.route('/<int:exam_id>/done')
def portal_done(exam_id):
    from routes.cbt import cbt_login_required, _current_student
    @cbt_login_required
    def _inner():
        from models import MockJAMBAttempt, MockJAMBResult
        student = _current_student()
        exam = db.session.get(MockJAMBExam, exam_id)
        att = (MockJAMBAttempt.query.filter_by(mock_exam_id=exam_id, student_id=student.id).first()
               if (exam and student) else None)
        if not att or att.status != 'Submitted':
            return redirect(url_for('mock_jamb_portal.portal_sit', exam_id=exam_id))
        result = MockJAMBResult.query.filter_by(student_id=student.id, mock_exam_id=exam_id).first()
        return render_template('mock_jamb/portal_done.html', exam=exam, student=student,
                               attempt=att, result=result)
    return _inner()
