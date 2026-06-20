"""
Student Contributions Module - SSS3 Graduation Fund Tracking
Hidden module accessible only with special access code
"""
from flask import Blueprint, render_template, request, redirect, url_for, flash, jsonify, session
from utils.helpers import get_active_term, get_active_session
from models import (
    db, Student, StudentEnrollment, ClassArmAssignment, SchoolClass, ContributionSettings, ContributionPayment,
    ContributionExpense
)
from functools import wraps
from datetime import datetime, date, timedelta
from sqlalchemy import func

contributions_bp = Blueprint('contributions', __name__, url_prefix='/contributions')

DEFAULT_ACCESS_CODE = "64665842"
SESSION_KEY = "contributions_access"


def _access_code():
    """The current access code: a configurable value stored in settings, or the
    built-in default when none has been set (so nobody is ever locked out)."""
    return ContributionSettings.get('access_code', DEFAULT_ACCESS_CODE) or DEFAULT_ACCESS_CODE


def contributions_access_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if not session.get(SESSION_KEY):
            return redirect(url_for('contributions.access_page'))
        return f(*args, **kwargs)
    return decorated_function


# --- SPA helpers (no-reload React shell + JSON-aware action responses) -------
from utils.spa import section_responders
_wants_json, _render, _ok, _err = section_responders(
    'contributions/app.html', 'contrib_json', 'contributions.dashboard',
    enrich=lambda p: p.setdefault('urls', _urls()))


def _urls():
    """Section-wide links used across the React shell."""
    return {
        'dashboard': url_for('contributions.dashboard'),
        'quick_entry': url_for('contributions.quick_entry'),
        'add_payment': url_for('contributions.add_payment'),
        'payments': url_for('contributions.payments_list'),
        'expenses': url_for('contributions.expenses_list'),
        'add_expense': url_for('contributions.add_expense'),
        'report': url_for('contributions.report'),
        'defaulters': url_for('contributions.defaulters'),
        'daily_summary': url_for('contributions.daily_summary'),
        'history': url_for('contributions.session_history'),
        'settings': url_for('contributions.settings'),
        'import': url_for('contributions.import_excel'),
        'export': url_for('contributions.export_excel'),
        'export_defaulters': url_for('contributions.export_defaulters'),
        'clear_all': url_for('contributions.clear_all_data'),
        'logout': url_for('contributions.logout_contributions'),
    }



@contributions_bp.route('/access', methods=['GET', 'POST'])
def access_page():
    if request.method == 'POST':
        code = request.form.get('access_code', '').strip()
        if code and code == _access_code():
            session[SESSION_KEY] = True
            flash('Access granted!', 'success')
            return redirect(url_for('contributions.dashboard'))
        else:
            flash('Invalid access code', 'error')
    return render_template('contributions/access.html')


@contributions_bp.route('/logout')
def logout_contributions():
    session.pop(SESSION_KEY, None)
    flash('Logged out from contributions module', 'info')
    return redirect(url_for('main.dashboard'))


@contributions_bp.route('/')
@contributions_access_required
def dashboard():
    """Main contributions dashboard with enhanced stats"""
    active_term = get_active_term()
    if not active_term:
        flash('No active term found', 'error')
        return redirect(url_for('main.dashboard'))
    
    sss3_class = SchoolClass.query.filter_by(name='SSS3').first()
    if not sss3_class:
        flash('SSS3 class not found', 'error')
        return redirect(url_for('main.dashboard'))
    
    # Get active session
    active_session = get_active_session()
    if not active_session:
        flash('No active session found', 'error')
        return redirect(url_for('main.dashboard'))
    
    sss3_assignments = ClassArmAssignment.query.filter_by(class_id=sss3_class.id, term_id=active_term.id).all()
    assignment_ids = [a.id for a in sss3_assignments]
    enrollments = StudentEnrollment.query.filter(StudentEnrollment.class_arm_assignment_id.in_(assignment_ids), StudentEnrollment.is_active == True).all()
    
    max_due = float(ContributionSettings.get('max_due', 20000))
    
    students_data = []
    total_expected = 0
    total_received = 0
    fully_paid_count = 0
    arms_summary = {}

    # One batched query for every student's total (was a SUM per student → N+1).
    from utils.services.contributions import paid_by_student
    paid_map = paid_by_student(active_session.id, [e.student_id for e in enrollments])

    for enrollment in enrollments:
        student = enrollment.student
        assignment = enrollment.class_arm_assignment
        arm_name = assignment.arm.name

        total_paid = paid_map.get(student.id, 0)
        remaining = max_due - total_paid
        status = 'Paid' if remaining <= 0 else 'Ongoing'
        
        if remaining <= 0:
            fully_paid_count += 1
        
        total_expected += max_due
        total_received += total_paid
        
        # Arm summary
        if arm_name not in arms_summary:
            arms_summary[arm_name] = {'total': 0, 'paid': 0, 'collected': 0}
        arms_summary[arm_name]['total'] += 1
        arms_summary[arm_name]['collected'] += total_paid
        if remaining <= 0:
            arms_summary[arm_name]['paid'] += 1
        
        students_data.append({
            'id': student.id, 'student_id': student.student_id, 'name': student.full_name,
            'arm': arm_name, 'max_due': max_due, 'total_paid': total_paid,
            'remaining': max(0, remaining), 'status': status,
            'percentage': min(100, (total_paid / max_due * 100)) if max_due > 0 else 0
        })
    
    # Calculate arm percentages
    for arm_name in arms_summary:
        expected = arms_summary[arm_name]['total'] * max_due
        arms_summary[arm_name]['percentage'] = (arms_summary[arm_name]['collected'] / expected * 100) if expected > 0 else 0
    
    students_data.sort(key=lambda x: (x['arm'], x['name']))
    
    # Additional stats
    total_expenses = db.session.query(func.sum(ContributionExpense.amount)).filter(
        ContributionExpense.session_id == active_session.id
    ).scalar() or 0
    expense_count = ContributionExpense.query.count()
    total_payments = ContributionPayment.query.count()
    total_outstanding = total_expected - total_received
    avg_paid = total_received / len(students_data) if students_data else 0
    collection_rate = (total_received / total_expected * 100) if total_expected > 0 else 0
    
    # Today and this week collections
    from datetime import timedelta
    today = date.today()
    week_start = today - timedelta(days=today.weekday())
    
    today_collections = db.session.query(func.sum(ContributionPayment.amount)).filter(
        ContributionPayment.payment_date == today
    ).scalar() or 0
    
    week_collections = db.session.query(func.sum(ContributionPayment.amount)).filter(
        ContributionPayment.payment_date >= week_start
    ).scalar() or 0
    
    # Recent payments
    recent_payments = ContributionPayment.query.filter_by(session_id=active_session.id).order_by(
        ContributionPayment.payment_date.desc(), ContributionPayment.created_at.desc()
    ).limit(10).all()
    
    # Top contributors
    top_contributors = sorted(students_data, key=lambda x: x['total_paid'], reverse=True)[:10]

    def _student_url(sid):
        return url_for('contributions.student_detail', student_id=sid)

    for s in students_data:
        s['detail_url'] = _student_url(s['id'])
        s['add_payment_url'] = url_for('contributions.add_payment', student_id=s['id'])

    arms_list = [{'name': name, 'paid': v['paid'], 'total': v['total'],
                  'collected': v['collected'], 'percentage': v['percentage']}
                 for name, v in sorted(arms_summary.items())]

    return _render({
        'page': 'dashboard',
        'stats': {
            'total_students': len(students_data), 'total_expected': total_expected,
            'total_received': total_received, 'total_expenses': total_expenses,
            'net_balance': total_received - total_expenses, 'fully_paid_count': fully_paid_count,
            'total_outstanding': total_outstanding, 'avg_paid': avg_paid,
            'total_payments': total_payments, 'today_collections': today_collections,
            'week_collections': week_collections, 'collection_rate': collection_rate,
            'expense_count': expense_count, 'max_due': max_due,
        },
        'students': students_data,
        'arms_summary': arms_list,
        'recent_payments': [{'student_name': p.student.full_name, 'amount': p.amount,
                             'date_short': p.payment_date.strftime('%d %b'),
                             'detail_url': _student_url(p.student_id)}
                            for p in recent_payments[:7]],
        'top_contributors': [{'name': s['name'], 'total_paid': s['total_paid'],
                              'detail_url': _student_url(s['id'])}
                             for s in top_contributors[:7]],
    })


@contributions_bp.route('/quick-entry', methods=['GET', 'POST'])
@contributions_access_required
def quick_entry():
    active_term = get_active_term()
    sss3_class = SchoolClass.query.filter_by(name='SSS3').first()
    active_session = get_active_session()
    if not active_term or not sss3_class or not active_session:
        flash('Configuration error', 'error')
        return redirect(url_for('contributions.dashboard'))
    
    sss3_assignments = ClassArmAssignment.query.filter_by(class_id=sss3_class.id, term_id=active_term.id).all()
    assignment_ids = [a.id for a in sss3_assignments]
    enrollments = StudentEnrollment.query.filter(StudentEnrollment.class_arm_assignment_id.in_(assignment_ids), StudentEnrollment.is_active == True).all()
    max_due = float(ContributionSettings.get('max_due', 20000))
    
    from utils.services.contributions import paid_by_student
    paid_map = paid_by_student(active_session.id, [e.student_id for e in enrollments])
    students_list = []
    for enrollment in enrollments:
        student = enrollment.student
        assignment = enrollment.class_arm_assignment
        total_paid = paid_map.get(student.id, 0)
        students_list.append({'id': student.id, 'name': student.full_name, 'arm': assignment.arm.name, 'total_paid': total_paid, 'remaining': max(0, max_due - total_paid)})
    students_list.sort(key=lambda x: (x['arm'], x['name']))
    
    if request.method == 'POST':
        try:
            payment_date_str = request.form.get('payment_date')
            received_by = request.form.get('received_by', '').strip()
            if not payment_date_str:
                return _err('Payment date is required', url_for('contributions.quick_entry'))
            payment_date = datetime.strptime(payment_date_str, '%Y-%m-%d').date()
            count = 0
            total_amount = 0
            for key, value in request.form.items():
                if key.startswith('amount_') and value.strip():
                    student_id = int(key.replace('amount_', ''))
                    amount = float(value.strip())
                    if amount > 0:
                        payment = ContributionPayment(session_id=active_session.id, student_id=student_id, amount=amount, payment_date=payment_date, received_by=received_by)
                        db.session.add(payment)
                        count += 1
                        total_amount += amount
            if count > 0:
                db.session.commit()
                return _ok(f'Added {count} payment(s) totaling ₦{total_amount:,.0f}',
                           url_for('contributions.quick_entry'))
            return _err('No payments entered', url_for('contributions.quick_entry'))
        except Exception as e:
            db.session.rollback()
            return _err(f'Error: {str(e)}', url_for('contributions.quick_entry'))

    return _render({
        'page': 'quick_entry',
        'students': students_list,
        'today': date.today().strftime('%Y-%m-%d'),
        'max_due': max_due,
        'submit_url': url_for('contributions.quick_entry'),
    })


@contributions_bp.route('/add-payment', methods=['GET', 'POST'])
@contributions_access_required
def add_payment():
    active_term = get_active_term()
    sss3_class = SchoolClass.query.filter_by(name='SSS3').first()
    active_session = get_active_session()
    if not active_term or not sss3_class or not active_session:
        flash('Configuration error', 'error')
        return redirect(url_for('contributions.dashboard'))
    
    sss3_assignments = ClassArmAssignment.query.filter_by(class_id=sss3_class.id, term_id=active_term.id).all()
    assignment_ids = [a.id for a in sss3_assignments]
    enrollments = StudentEnrollment.query.filter(StudentEnrollment.class_arm_assignment_id.in_(assignment_ids), StudentEnrollment.is_active == True).all()
    students_list = [{'id': e.student.id, 'name': f"{e.student.full_name} ({e.class_arm_assignment.arm.name})"} for e in enrollments]
    students_list.sort(key=lambda x: x['name'])
    
    if request.method == 'POST':
        try:
            student_id = int(request.form.get('student_id'))
            amount = float(request.form.get('amount'))
            payment_date = datetime.strptime(request.form.get('payment_date'), '%Y-%m-%d').date()
            received_by = request.form.get('received_by', '').strip()
            notes = request.form.get('notes', '').strip()
            payment = ContributionPayment(session_id=active_session.id, student_id=student_id, amount=amount, payment_date=payment_date, received_by=received_by, notes=notes)
            db.session.add(payment)
            db.session.commit()
            student = db.session.get(Student, student_id)
            return _ok(f'Payment of ₦{amount:,.0f} added for {student.full_name}',
                       url_for('contributions.dashboard'))
        except Exception as e:
            db.session.rollback()
            return _err(f'Error: {str(e)}', url_for('contributions.add_payment'))

    return _render({
        'page': 'add_payment',
        'students': students_list,
        'today': date.today().strftime('%Y-%m-%d'),
        'preselect': request.args.get('student_id', type=int),
        'submit_url': url_for('contributions.add_payment'),
        'info_url_base': url_for('contributions.api_student_info', student_id=0),
    })


@contributions_bp.route('/payments')
@contributions_access_required
def payments_list():
    active_session = get_active_session()
    filter_date = request.args.get('date')
    query = ContributionPayment.query.filter_by(session_id=active_session.id if active_session else None)
    if filter_date:
        try:
            payment_date = datetime.strptime(filter_date, '%Y-%m-%d').date()
            query = query.filter(ContributionPayment.payment_date == payment_date)
        except Exception:
            pass
    payments = query.order_by(ContributionPayment.payment_date.desc(), ContributionPayment.created_at.desc()).all()
    unique_dates = db.session.query(ContributionPayment.payment_date).distinct().order_by(ContributionPayment.payment_date.desc()).all()
    return _render({
        'page': 'payments',
        'filter_date': filter_date,
        'unique_dates': [{'value': d[0].strftime('%Y-%m-%d'),
                          'label': d[0].strftime('%a, %d %b %Y')} for d in unique_dates if d[0]],
        'payments': [{'id': p.id, 'date': p.payment_date.strftime('%a, %d %b %Y'),
                      'student_name': p.student.full_name, 'amount': p.amount,
                      'received_by': p.received_by or '-', 'notes': p.notes or '-',
                      'detail_url': url_for('contributions.student_detail', student_id=p.student_id),
                      'delete_url': url_for('contributions.delete_payment', payment_id=p.id)}
                     for p in payments],
    })


@contributions_bp.route('/payments/<int:payment_id>/delete', methods=['POST'])
@contributions_access_required
def delete_payment(payment_id):
    payment = db.get_or_404(ContributionPayment, payment_id)
    student_name = payment.student.full_name
    amount = payment.amount
    db.session.delete(payment)
    db.session.commit()
    return _ok(f'Deleted payment of ₦{amount:,.0f} from {student_name}',
               url_for('contributions.payments_list'))


@contributions_bp.route('/student/<int:student_id>')
@contributions_access_required
def student_detail(student_id):
    student = db.get_or_404(Student, student_id)
    active_session = get_active_session()
    max_due = float(ContributionSettings.get('max_due', 20000))
    payments = ContributionPayment.query.filter_by(student_id=student_id, session_id=active_session.id if active_session else None).order_by(ContributionPayment.payment_date.desc()).all()
    total_paid = sum(p.amount for p in payments)
    remaining = max(0, max_due - total_paid)
    start_date_str = ContributionSettings.get('start_date', '2025-01-06')
    try:
        start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date()
    except Exception:
        start_date = date(2025, 1, 6)
    weekly_data = []
    for week_num in range(1, 36):
        week_start = start_date + timedelta(weeks=week_num - 1)
        week_end = week_start + timedelta(days=6)
        week_payments = [p for p in payments if week_start <= p.payment_date <= week_end]
        week_total = sum(p.amount for p in week_payments)
        if week_total > 0 or week_start <= date.today():
            weekly_data.append({'week': week_num,
                                'period': f"{week_start.strftime('%d %b')} - {week_end.strftime('%d %b')}",
                                'amount': week_total})
    return _render({
        'page': 'student_detail',
        'student': {'id': student.id, 'name': student.full_name},
        'total_paid': total_paid, 'remaining': remaining, 'max_due': max_due,
        'status': 'Paid' if remaining <= 0 else 'Ongoing',
        'payments': [{'date': p.payment_date.strftime('%d %b %Y'), 'amount': p.amount,
                      'received_by': p.received_by or '-'} for p in payments],
        'weekly_data': weekly_data,
    })


@contributions_bp.route('/expenses')
@contributions_access_required
def expenses_list():
    """View all expenses with financial summary"""
    active_session = get_active_session()
    expenses = ContributionExpense.query.filter_by(session_id=active_session.id if active_session else None).order_by(ContributionExpense.expense_date.desc()).all()
    total_expenses = sum(e.amount for e in expenses)
    
    # Get total collected
    total_collected = db.session.query(func.sum(ContributionPayment.amount)).filter(
        ContributionPayment.session_id == active_session.id if active_session else None
    ).scalar() or 0
    
    # Calculate available balance
    available_balance = total_collected - total_expenses
    
    # Expense rate (percentage of collected that has been spent)
    expense_rate = (total_expenses / total_collected * 100) if total_collected > 0 else 0
    
    # Group expenses by description (simple categorization)
    category_summary = {}
    for expense in expenses:
        category = expense.description.split()[0] if expense.description else 'Other'
        category_summary[category] = category_summary.get(category, 0) + expense.amount
    
    category_summary = dict(sorted(category_summary.items(), key=lambda x: x[1], reverse=True))

    return _render({
        'page': 'expenses',
        'total_expenses': total_expenses, 'total_collected': total_collected,
        'available_balance': available_balance, 'expense_rate': expense_rate,
        'category_summary': ([{'name': k, 'amount': v} for k, v in category_summary.items()]
                             if len(category_summary) > 1 else []),
        'expenses': [{'id': e.id, 'date': e.expense_date.strftime('%a, %d %b %Y'),
                      'description': e.description, 'amount': e.amount, 'notes': e.notes or '-',
                      'delete_url': url_for('contributions.delete_expense', expense_id=e.id)}
                     for e in expenses],
    })



@contributions_bp.route('/expenses/add', methods=['GET', 'POST'])
@contributions_access_required
def add_expense():
    if request.method == 'POST':
        try:
            expense_date = datetime.strptime(request.form.get('expense_date'), '%Y-%m-%d').date()
            description = request.form.get('description', '').strip()
            amount = float(request.form.get('amount'))
            notes = request.form.get('notes', '').strip()
            active_session = get_active_session()
            expense = ContributionExpense(session_id=active_session.id if active_session else None, expense_date=expense_date, description=description, amount=amount, notes=notes)
            db.session.add(expense)
            db.session.commit()
            return _ok(f'Expense of ₦{amount:,.0f} added', url_for('contributions.expenses_list'))
        except Exception as e:
            db.session.rollback()
            return _err(f'Error: {str(e)}', url_for('contributions.add_expense'))
    return _render({
        'page': 'add_expense',
        'today': date.today().strftime('%Y-%m-%d'),
        'submit_url': url_for('contributions.add_expense'),
        'back_url': url_for('contributions.expenses_list'),
    })


@contributions_bp.route('/expenses/<int:expense_id>/delete', methods=['POST'])
@contributions_access_required
def delete_expense(expense_id):
    expense = db.get_or_404(ContributionExpense, expense_id)
    db.session.delete(expense)
    db.session.commit()
    return _ok('Expense deleted', url_for('contributions.expenses_list'))


@contributions_bp.route('/settings', methods=['GET', 'POST'])
@contributions_access_required
def settings():
    if request.method == 'POST':
        try:
            max_due = float(request.form.get('max_due', 20000))
            start_date = request.form.get('start_date', '2025-01-06')
            ContributionSettings.set('max_due', max_due)
            ContributionSettings.set('start_date', start_date)
            # Optional: change the access password. Blank = keep the current one.
            new_code = (request.form.get('access_code') or '').strip()
            if new_code:
                ContributionSettings.set('access_code', new_code)
        except Exception as e:
            return _err(f'Error: {str(e)}', url_for('contributions.settings'))
        return _ok('Settings updated', url_for('contributions.settings'))
    max_due = ContributionSettings.get('max_due', '20000')
    start_date = ContributionSettings.get('start_date', '2025-01-06')
    return _render({
        'page': 'settings', 'max_due': max_due, 'start_date': start_date,
        'has_custom_code': ContributionSettings.get('access_code') is not None,
        'submit_url': url_for('contributions.settings'),
        'back_url': url_for('contributions.dashboard'),
    })


@contributions_bp.route('/report')
@contributions_access_required
def report():
    active_term = get_active_term()
    sss3_class = SchoolClass.query.filter_by(name='SSS3').first()
    active_session = get_active_session()
    if not active_term or not sss3_class or not active_session:
        flash('Configuration error', 'error')
        return redirect(url_for('contributions.dashboard'))
    max_due = float(ContributionSettings.get('max_due', 20000))
    arms_data = {}
    from utils.services.contributions import paid_by_student
    paid_map = paid_by_student(active_session.id)   # one query for the whole session
    sss3_assignments = ClassArmAssignment.query.filter_by(class_id=sss3_class.id, term_id=active_term.id).all()
    for assignment in sss3_assignments:
        arm_name = assignment.arm.name
        enrollments = StudentEnrollment.query.filter_by(class_arm_assignment_id=assignment.id, is_active=True).all()
        arm_students = []
        arm_total_paid = 0
        arm_fully_paid = 0
        for enrollment in enrollments:
            student = enrollment.student
            total_paid = paid_map.get(student.id, 0)
            remaining = max(0, max_due - total_paid)
            if remaining <= 0:
                arm_fully_paid += 1
            arm_total_paid += total_paid
            arm_students.append({'name': student.full_name, 'total_paid': total_paid, 'remaining': remaining, 'status': 'Paid' if remaining <= 0 else 'Ongoing'})
        arm_students.sort(key=lambda x: x['name'])
        arms_data[arm_name] = {'students': arm_students, 'total_students': len(arm_students), 'total_expected': len(arm_students) * max_due, 'total_paid': arm_total_paid, 'fully_paid': arm_fully_paid}
    return _render({
        'page': 'report', 'max_due': max_due,
        'arms': [{'name': name, **data} for name, data in sorted(arms_data.items())],
        'back_url': url_for('contributions.dashboard'),
    })


@contributions_bp.route('/export')
@contributions_access_required
def export_excel():
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill
        from io import BytesIO
        from flask import send_file
        active_term = get_active_term()
        sss3_class = SchoolClass.query.filter_by(name='SSS3').first()
        active_session = get_active_session()
        max_due = float(ContributionSettings.get('max_due', 20000))
        wb = Workbook()
        ws_students = wb.active
        ws_students.title = "Students"
        headers = ['#', 'Name', 'Arm', 'Max Due', 'Total Paid', 'Remaining', 'Status']
        ws_students.append(headers)
        header_fill = PatternFill(start_color='1a5f4a', end_color='1a5f4a', fill_type='solid')
        header_font = Font(color='FFFFFF', bold=True)
        for col, cell in enumerate(ws_students[1], 1):
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')
        sss3_assignments = ClassArmAssignment.query.filter_by(class_id=sss3_class.id, term_id=active_term.id).all()
        assignment_ids = [a.id for a in sss3_assignments]
        enrollments = StudentEnrollment.query.filter(StudentEnrollment.class_arm_assignment_id.in_(assignment_ids), StudentEnrollment.is_active == True).all()
        students_data = []
        for enrollment in enrollments:
            student = enrollment.student
            assignment = enrollment.class_arm_assignment
            total_paid = db.session.query(func.sum(ContributionPayment.amount)).filter(
            ContributionPayment.student_id == student.id,
            ContributionPayment.session_id == active_session.id
        ).scalar() or 0
            remaining = max(0, max_due - total_paid)
            students_data.append({'name': student.full_name, 'arm': assignment.arm.name, 'max_due': max_due, 'total_paid': total_paid, 'remaining': remaining, 'status': 'Paid' if remaining <= 0 else 'Ongoing'})
        students_data.sort(key=lambda x: (x['arm'], x['name']))
        for i, s in enumerate(students_data, 1):
            ws_students.append([i, s['name'], s['arm'], s['max_due'], s['total_paid'], s['remaining'], s['status']])
        ws_payments = wb.create_sheet("Payments")
        ws_payments.append(['#', 'Student', 'Date', 'Amount', 'Received By', 'Notes'])
        for col, cell in enumerate(ws_payments[1], 1):
            cell.fill = header_fill
            cell.font = header_font
        payments = ContributionPayment.query.order_by(ContributionPayment.payment_date.desc()).all()
        for i, p in enumerate(payments, 1):
            ws_payments.append([i, p.student.full_name, p.payment_date.strftime('%Y-%m-%d'), p.amount, p.received_by or '', p.notes or ''])
        ws_expenses = wb.create_sheet("Expenses")
        ws_expenses.append(['#', 'Date', 'Description', 'Amount', 'Notes'])
        for col, cell in enumerate(ws_expenses[1], 1):
            cell.fill = header_fill
            cell.font = header_font
        expenses = ContributionExpense.query.order_by(ContributionExpense.expense_date.desc()).all()
        for i, e in enumerate(expenses, 1):
            ws_expenses.append([i, e.expense_date.strftime('%Y-%m-%d'), e.description, e.amount, e.notes or ''])
        ws_summary = wb.create_sheet("Summary")
        total_expected = len(students_data) * max_due
        total_received = sum(s['total_paid'] for s in students_data)
        total_expenses = sum(e.amount for e in expenses)
        ws_summary.append(['Summary', ''])
        ws_summary.append(['Total Students', len(students_data)])
        ws_summary.append(['Max Due Per Student', f'₦{max_due:,.0f}'])
        ws_summary.append(['Total Expected', f'₦{total_expected:,.0f}'])
        ws_summary.append(['Total Received', f'₦{total_received:,.0f}'])
        ws_summary.append(['Total Expenses', f'₦{total_expenses:,.0f}'])
        ws_summary.append(['Net Balance', f'₦{total_received - total_expenses:,.0f}'])
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        filename = f'contributions_report_{date.today().strftime("%Y%m%d")}.xlsx'
        return send_file(output, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', as_attachment=True, download_name=filename)
    except Exception as e:
        flash(f'Export error: {str(e)}', 'error')
        return redirect(url_for('contributions.dashboard'))


@contributions_bp.route('/api/student/<int:student_id>/info')
@contributions_access_required
def api_student_info(student_id):
    student = db.get_or_404(Student, student_id)
    active_session = get_active_session()
    max_due = float(ContributionSettings.get('max_due', 20000))
    total_paid = db.session.query(func.sum(ContributionPayment.amount)).filter(
        ContributionPayment.student_id == student_id,
        ContributionPayment.session_id == active_session.id if active_session else None
    ).scalar() or 0
    return jsonify({'name': student.full_name, 'total_paid': total_paid, 'remaining': max(0, max_due - total_paid), 'max_due': max_due})


@contributions_bp.route('/import', methods=['GET', 'POST'])
@contributions_access_required
def import_excel():
    """Import contributions data from Excel file"""
    if request.method == 'POST':
        if 'file' not in request.files:
            flash('No file uploaded', 'error')
            return redirect(url_for('contributions.import_excel'))
        
        file = request.files['file']
        if file.filename == '':
            flash('No file selected', 'error')
            return redirect(url_for('contributions.import_excel'))
        
        if not file.filename.endswith('.xlsx'):
            flash('Please upload an Excel file (.xlsx)', 'error')
            return redirect(url_for('contributions.import_excel'))
        
        try:
            import openpyxl
            from io import BytesIO
            import re
            
            wb = openpyxl.load_workbook(BytesIO(file.read()), data_only=True)
            
            active_term = get_active_term()
            sss3_class = SchoolClass.query.filter_by(name='SSS3').first()
            
            if not active_term or not sss3_class:
                flash('No active term or SSS3 class found', 'error')
                return redirect(url_for('contributions.import_excel'))
            
            sss3_assignments = ClassArmAssignment.query.filter_by(
                class_id=sss3_class.id, term_id=active_term.id
            ).all()
            assignment_ids = [a.id for a in sss3_assignments]
            enrollments = StudentEnrollment.query.filter(
                StudentEnrollment.class_arm_assignment_id.in_(assignment_ids),
                StudentEnrollment.is_active == True
            ).all()
            
            def normalize_name(name):
                if name is None:
                    return ''
                name = str(name).upper().strip()
                name = re.sub(r"['\-\.]", "", name)
                return ' '.join(name.split())
            
            def get_name_parts(name):
                return set(normalize_name(name).split())
            
            def names_match(name1, name2):
                n1 = normalize_name(name1)
                n2 = normalize_name(name2)
                if n1 == n2:
                    return True
                parts1 = get_name_parts(name1)
                parts2 = get_name_parts(name2)
                if len(parts1) >= 2 and len(parts2) >= 2:
                    common = parts1.intersection(parts2)
                    if len(common) >= 2:
                        return True
                if parts1 and parts2:
                    if parts1.issubset(parts2) or parts2.issubset(parts1):
                        return True
                return False
            
            db_students = []
            for enrollment in enrollments:
                student = enrollment.student
                db_students.append({
                    'id': student.id,
                    'full_name': student.full_name,
                    'surname': student.surname,
                    'first_name': student.first_name
                })
            
            def find_db_student(excel_name):
                for s in db_students:
                    if names_match(excel_name, s['full_name']):
                        return s['id']
                    if names_match(excel_name, f"{s['surname']} {s['first_name']}"):
                        return s['id']
                    if names_match(excel_name, f"{s['first_name']} {s['surname']}"):
                        return s['id']
                return None
            
            ws_students = wb['Students']
            excel_id_to_name = {}
            for row in range(2, ws_students.max_row + 1):
                eid = ws_students.cell(row, 1).value
                name = ws_students.cell(row, 2).value
                if eid is not None and name is not None:
                    try:
                        eid_int = int(eid)
                        name_str = str(name).strip()
                        if name_str.isdigit():
                            continue
                        excel_id_to_name[eid_int] = name_str
                    except (ValueError, TypeError):
                        continue
            
            ws_payments = wb['Payments']
            imported_count = 0
            skipped_count = 0
            unmatched_names = set()
            
            for row in range(2, ws_payments.max_row + 1):
                excel_student_id = ws_payments.cell(row, 2).value
                payment_date = ws_payments.cell(row, 4).value
                received_by = ws_payments.cell(row, 5).value
                amount = ws_payments.cell(row, 6).value
                
                if excel_student_id is None or str(excel_student_id).strip() == '-':
                    continue
                if amount is None or str(amount).strip() == '-':
                    continue
                
                try:
                    excel_student_id = int(excel_student_id)
                    amount = float(amount)
                except (ValueError, TypeError):
                    continue
                
                if amount <= 0:
                    continue
                
                excel_name = excel_id_to_name.get(excel_student_id)
                if not excel_name:
                    skipped_count += 1
                    continue
                
                db_student_id = find_db_student(excel_name)
                
                if not db_student_id:
                    unmatched_names.add(excel_name)
                    skipped_count += 1
                    continue
                
                if isinstance(payment_date, datetime):
                    pay_date = payment_date.date()
                elif isinstance(payment_date, date):
                    pay_date = payment_date
                else:
                    try:
                        pay_date = datetime.strptime(str(payment_date), '%Y-%m-%d').date()
                    except Exception:
                        skipped_count += 1
                        continue
                
                active_sess = get_active_session()
                payment = ContributionPayment(
                    session_id=active_sess.id if active_sess else None,
                    student_id=db_student_id,
                    amount=amount,
                    payment_date=pay_date,
                    received_by=str(received_by).strip() if received_by and str(received_by).strip() != '-' else None
                )
                db.session.add(payment)
                imported_count += 1
            
            db.session.commit()
            
            if unmatched_names:
                flash(f'Unmatched students: {", ".join(list(unmatched_names)[:5])}{"..." if len(unmatched_names) > 5 else ""}', 'warning')
            
            flash(f'Successfully imported {imported_count} payments. Skipped {skipped_count} rows.', 'success')
            return redirect(url_for('contributions.dashboard'))
            
        except Exception as e:
            db.session.rollback()
            import traceback
            traceback.print_exc()
            flash(f'Import error: {str(e)}', 'error')
            return redirect(url_for('contributions.import_excel'))
    
    return _render({
        'page': 'import',
        'submit_url': url_for('contributions.import_excel'),
        'clear_url': url_for('contributions.clear_all_data'),
        'back_url': url_for('contributions.dashboard'),
    })


@contributions_bp.route('/clear-all', methods=['POST'])
@contributions_access_required
def clear_all_data():
    """Clear all contribution data"""
    try:
        ContributionPayment.query.delete()
        ContributionExpense.query.delete()
        db.session.commit()
    except Exception as e:
        db.session.rollback()
        return _err(f'Error: {str(e)}', url_for('contributions.import_excel'))
    return _ok('All contribution data cleared', url_for('contributions.import_excel'))


@contributions_bp.route('/defaulters')
@contributions_access_required
def defaulters():
    """List students with outstanding balance"""
    active_term = get_active_term()
    sss3_class = SchoolClass.query.filter_by(name='SSS3').first()
    active_session = get_active_session()
    
    if not active_term or not sss3_class or not active_session:
        flash('Configuration error', 'error')
        return redirect(url_for('contributions.dashboard'))
    
    max_due = float(ContributionSettings.get('max_due', 20000))
    
    sss3_assignments = ClassArmAssignment.query.filter_by(class_id=sss3_class.id, term_id=active_term.id).all()
    assignment_ids = [a.id for a in sss3_assignments]
    enrollments = StudentEnrollment.query.filter(
        StudentEnrollment.class_arm_assignment_id.in_(assignment_ids),
        StudentEnrollment.is_active == True
    ).all()
    
    defaulters_list = []
    total_outstanding = 0

    from utils.services.contributions import paid_by_student
    paid_map = paid_by_student(active_session.id, [e.student_id for e in enrollments])

    for enrollment in enrollments:
        student = enrollment.student
        assignment = enrollment.class_arm_assignment

        total_paid = paid_map.get(student.id, 0)

        remaining = max_due - total_paid
        
        if remaining > 0:
            last_payment = ContributionPayment.query.filter_by(student_id=student.id, session_id=active_session.id).order_by(
                ContributionPayment.payment_date.desc()
            ).first()
            
            defaulters_list.append({
                'id': student.id,
                'name': student.full_name,
                'arm': assignment.arm.name,
                'total_paid': total_paid,
                'remaining': remaining,
                'percentage': (total_paid / max_due * 100) if max_due > 0 else 0,
                'last_payment': last_payment.payment_date if last_payment else None
            })
            total_outstanding += remaining
    
    defaulters_list.sort(key=lambda x: x['remaining'], reverse=True)
    avg_outstanding = total_outstanding / len(defaulters_list) if defaulters_list else 0

    for s in defaulters_list:
        s['last_payment'] = s['last_payment'].strftime('%d %b %Y') if s['last_payment'] else 'Never'
        s['detail_url'] = url_for('contributions.student_detail', student_id=s['id'])
        s['add_payment_url'] = url_for('contributions.add_payment', student_id=s['id'])

    return _render({
        'page': 'defaulters',
        'defaulters': defaulters_list,
        'total_outstanding': total_outstanding,
        'avg_outstanding': avg_outstanding,
        'max_due': max_due,
        'export_url': url_for('contributions.export_defaulters'),
        'back_url': url_for('contributions.dashboard'),
    })


@contributions_bp.route('/daily-summary')
@contributions_access_required
def daily_summary():
    """Show collections grouped by date"""
    active_session = get_active_session()
    daily_data = db.session.query(
        ContributionPayment.payment_date,
        func.count(ContributionPayment.id).label('count'),
        func.sum(ContributionPayment.amount).label('amount')
    ).filter(ContributionPayment.session_id == active_session.id if active_session else None
    ).group_by(ContributionPayment.payment_date).order_by(ContributionPayment.payment_date.desc()).all()
    
    result = []
    total_amount = 0
    best_day = None
    
    for row in daily_data:
        collectors = db.session.query(ContributionPayment.received_by).filter(
            ContributionPayment.payment_date == row.payment_date,
            ContributionPayment.session_id == active_session.id if active_session else None,
            ContributionPayment.received_by != None,
            ContributionPayment.received_by != ''
        ).distinct().all()
        
        day_data = {
            'date': row.payment_date,
            'count': row.count,
            'amount': row.amount or 0,
            'collectors': [c[0] for c in collectors if c[0]]
        }
        result.append(day_data)
        total_amount += day_data['amount']

        if not best_day or day_data['amount'] > best_day['amount']:
            best_day = day_data

    avg_daily = total_amount / len(result) if result else 0

    return _render({
        'page': 'daily_summary',
        'avg_daily': avg_daily,
        'days_count': len(result),
        'best_day': ({'amount': best_day['amount'], 'date_short': best_day['date'].strftime('%d %b')}
                     if best_day else None),
        'daily_data': [{'date': d['date'].strftime('%d %b %Y'),
                        'day': d['date'].strftime('%A'),
                        'count': d['count'], 'amount': d['amount'],
                        'collectors': ', '.join(d['collectors']) if d['collectors'] else '-',
                        'view_url': url_for('contributions.payments_list', date=d['date'].strftime('%Y-%m-%d'))}
                       for d in result],
        'back_url': url_for('contributions.dashboard'),
    })


@contributions_bp.route('/export-defaulters')
@contributions_access_required
def export_defaulters():
    """Export defaulters list to Excel"""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill
        from io import BytesIO
        from flask import send_file
        
        active_term = get_active_term()
        sss3_class = SchoolClass.query.filter_by(name='SSS3').first()
        active_session = get_active_session()
        max_due = float(ContributionSettings.get('max_due', 20000))
        
        sss3_assignments = ClassArmAssignment.query.filter_by(class_id=sss3_class.id, term_id=active_term.id).all()
        assignment_ids = [a.id for a in sss3_assignments]
        enrollments = StudentEnrollment.query.filter(
            StudentEnrollment.class_arm_assignment_id.in_(assignment_ids),
            StudentEnrollment.is_active == True
        ).all()
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Defaulters"
        
        headers = ['#', 'Name', 'Arm', 'Amount Paid', 'Outstanding', 'Progress %']
        ws.append(headers)
        
        header_fill = PatternFill(start_color='dc3545', end_color='dc3545', fill_type='solid')
        header_font = Font(color='FFFFFF', bold=True)
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
        
        idx = 1
        for enrollment in enrollments:
            student = enrollment.student
            assignment = enrollment.class_arm_assignment
            
            total_paid = db.session.query(func.sum(ContributionPayment.amount)).filter(
                ContributionPayment.student_id == student.id,
                ContributionPayment.session_id == active_session.id
            ).scalar() or 0
            
            remaining = max_due - total_paid
            
            if remaining > 0:
                ws.append([
                    idx,
                    student.full_name,
                    assignment.arm.name,
                    total_paid,
                    remaining,
                    f"{(total_paid/max_due*100):.1f}%"
                ])
                idx += 1
        
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        return send_file(output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=f'defaulters_{date.today().strftime("%Y%m%d")}.xlsx'
        )
    except Exception as e:
        flash(f'Export error: {str(e)}', 'error')
        return redirect(url_for('contributions.defaulters'))


@contributions_bp.route('/history')
@contributions_access_required
def session_history():
    """View contributions from previous sessions"""
    from models import AcademicSession
    
    # Get all sessions that have contribution data
    sessions_with_data = db.session.query(
        AcademicSession,
        func.count(ContributionPayment.id).label('payment_count'),
        func.sum(ContributionPayment.amount).label('total_collected')
    ).outerjoin(ContributionPayment).group_by(AcademicSession.id).order_by(
        AcademicSession.start_date.desc()
    ).all()
    
    session_data = []
    for session, payment_count, total_collected in sessions_with_data:
        if payment_count and payment_count > 0:
            # Get expense total for this session
            expenses = db.session.query(func.sum(ContributionExpense.amount)).filter(
                ContributionExpense.session_id == session.id
            ).scalar() or 0
            
            session_data.append({
                'id': session.id,
                'name': session.name,
                'is_active': session.is_active,
                'payment_count': payment_count,
                'total_collected': total_collected or 0,
                'total_expenses': expenses,
                'net_balance': (total_collected or 0) - expenses
            })
    
    for s in session_data:
        s['view_url'] = url_for('contributions.view_session', session_id=s['id'])
    return _render({'page': 'history', 'sessions': session_data,
                    'back_url': url_for('contributions.dashboard')})


@contributions_bp.route('/session/<int:session_id>')
@contributions_access_required
def view_session(session_id):
    """View contributions for a specific session"""
    from models import AcademicSession
    
    session = db.get_or_404(AcademicSession, session_id)
    
    # Get all payments for this session
    payments = ContributionPayment.query.filter_by(session_id=session_id).order_by(
        ContributionPayment.payment_date.desc()
    ).all()
    
    # Group by student
    student_totals = {}
    for payment in payments:
        sid = payment.student_id
        if sid not in student_totals:
            student_totals[sid] = {
                'student': payment.student,
                'total': 0,
                'payment_count': 0
            }
        student_totals[sid]['total'] += payment.amount
        student_totals[sid]['payment_count'] += 1
    
    # Get expenses
    expenses = ContributionExpense.query.filter_by(session_id=session_id).order_by(
        ContributionExpense.expense_date.desc()
    ).all()
    
    total_collected = sum(p.amount for p in payments)
    total_expenses = sum(e.amount for e in expenses)
    ordered = sorted(student_totals.values(), key=lambda x: x['total'], reverse=True)

    return _render({
        'page': 'view_session',
        'session': {'name': session.name, 'is_active': session.is_active},
        'payments_count': len(payments),
        'total_collected': total_collected, 'total_expenses': total_expenses,
        'net_balance': total_collected - total_expenses,
        'student_totals': [{'name': it['student'].full_name, 'payment_count': it['payment_count'],
                            'total': it['total']} for it in ordered],
        'expenses': [{'description': e.description,
                      'date': e.expense_date.strftime('%d %b %Y'), 'amount': e.amount}
                     for e in expenses],
        'back_url': url_for('contributions.session_history'),
    })
