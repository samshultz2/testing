"""
Reports and data management routes
"""
from flask import Blueprint, render_template, request, redirect, url_for, flash, Response, jsonify
from utils.helpers import get_active_term, get_active_session
from models import db, Student, ParentContact, Term
from utils.helpers import login_required
from utils.excel_utils import (
    export_students_to_excel, create_student_import_template,
    import_students_from_excel
)
from utils.web_exports import xlsx_response
from utils.branch_scope import scope_query, scope_by_student, viewing_branch_id

reports_bp = Blueprint('reports', __name__, url_prefix='/reports')


def _wants_json():
    return request.headers.get('X-Requested-With') == 'fetch' or request.is_json


def _render(payload):
    from utils.spa import render_or_json
    return render_or_json('reports/app.html', 'reports_json', payload)


@reports_bp.route('/')
@login_required
def index():
    """Reports main page"""
    return _render({'page': 'index', 'urls': {
        'summary': url_for('reports.summary_report'),
        'import': url_for('reports.import_students'),
        'export_students': url_for('reports.export_students'),
        'export_template': url_for('reports.export_template'),
    }})


# ============================================================================
# EXPORT FUNCTIONS
# ============================================================================

@reports_bp.route('/export/students')
@login_required
def export_students():
    """Export all students to Excel"""
    students = scope_query(Student.query.filter_by(is_active=True), Student).order_by(Student.surname).all()
    
    excel_file = export_students_to_excel(students)

    return xlsx_response(excel_file, 'students_export.xlsx')


@reports_bp.route('/export/class-students')
@login_required
def export_class_students():
    """Export students for a specific class"""
    from models import ClassArmAssignment, StudentEnrollment
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Border, Side
    import io
    
    assignment_id = request.args.get('assignment_id', type=int)
    term_id = request.args.get('term_id', type=int)
    
    if not assignment_id:
        # Show selection form
        terms = Term.query.order_by(Term.id.desc()).all()
        
        if not term_id:
            active_term = get_active_term()
            if active_term:
                term_id = active_term.id
        
        assignments = []
        if term_id:
            assignments = scope_query(
                ClassArmAssignment.query.filter_by(term_id=term_id), ClassArmAssignment).all()

        return _render({
            'page': 'export_class', 'term_id': term_id or '',
            'terms': [{'id': t.id, 'label': t.full_name} for t in terms],
            'assignments': [{'id': a.id, 'display_name': a.display_name,
                             'student_count': sum(1 for e in a.enrollments if e.is_active),
                             'export_url': url_for('reports.export_class_students', assignment_id=a.id)}
                            for a in assignments],
            'urls': {'self': url_for('reports.export_class_students')},
        })
    
    # Export the class
    assignment = db.get_or_404(ClassArmAssignment, assignment_id)
    
    enrollments = StudentEnrollment.query.filter_by(
        class_arm_assignment_id=assignment_id,
        is_active=True
    ).join(Student).order_by(Student.surname, Student.first_name).all()
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Students"
    
    # Styles
    header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF')
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    
    # Title
    ws['A1'] = f"{assignment.display_name} - Student List"
    ws['A1'].font = Font(bold=True, size=14)
    ws['A2'] = f"{assignment.term.full_name}"
    
    # Headers
    headers = ['S/N', 'Student ID', 'Surname', 'First Name', 'Other Names', 'Gender', 'Date of Birth', 'Religion']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = thin_border
    
    # Data
    for idx, enrollment in enumerate(enrollments, 1):
        student = enrollment.student
        row = idx + 4
        ws.cell(row=row, column=1, value=idx).border = thin_border
        ws.cell(row=row, column=2, value=student.student_id).border = thin_border
        ws.cell(row=row, column=3, value=student.surname).border = thin_border
        ws.cell(row=row, column=4, value=student.first_name).border = thin_border
        ws.cell(row=row, column=5, value=student.other_names or '').border = thin_border
        ws.cell(row=row, column=6, value=student.gender).border = thin_border
        ws.cell(row=row, column=7, value=student.date_of_birth.strftime('%Y-%m-%d') if student.date_of_birth else '').border = thin_border
        ws.cell(row=row, column=8, value=student.religion or '').border = thin_border
    
    # Auto-width
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except Exception:
                pass
        ws.column_dimensions[column].width = min(max_length + 2, 30)
    
    filename = f"students_{assignment.school_class.name}_{assignment.arm.name}.xlsx"

    return xlsx_response(wb, filename)


@reports_bp.route('/export/template')
@login_required
def export_template():
    """Download student import template"""
    template = create_student_import_template()

    return xlsx_response(template, 'student_import_template.xlsx')


# ============================================================================
# IMPORT FUNCTIONS
# ============================================================================

@reports_bp.route('/import/students', methods=['GET', 'POST'])
@login_required
def import_students():
    """Import students from Excel or CSV, optionally enrolling into a class arm."""
    from models import ClassArmAssignment, SchoolClass, ClassArm

    def _fail(msg):
        if _wants_json():
            return jsonify({'ok': False, 'error': msg}), 400
        flash(msg, 'error')
        return redirect(url_for('reports.import_students'))

    if request.method == 'POST':
        file = request.files.get('file')
        if not file or file.filename == '':
            return _fail('No file selected.')
        if not file.filename.lower().endswith(('.xlsx', '.xls', '.csv')):
            return _fail('Please upload a .xlsx or .csv file.')

        try:
            from utils.branch_scope import branch_for_new
            from utils.helpers import get_active_term
            from utils.excel_utils import import_students_from_upload

            # Optional: enrol all imported students into a chosen class + arm.
            assignment_id = None
            class_id = request.form.get('class_id', type=int)
            arm_id = request.form.get('arm_id', type=int)
            if class_id and arm_id:
                term = get_active_term()
                if not term:
                    flash('No active term set — students imported without a class.', 'warning')
                else:
                    # The (class, arm, term) combo is unique; reuse it if class
                    # management already created it, otherwise make it now.
                    assignment = ClassArmAssignment.query.filter_by(
                        class_id=class_id, arm_id=arm_id, term_id=term.id).first()
                    if not assignment:
                        assignment = ClassArmAssignment(
                            class_id=class_id, arm_id=arm_id, term_id=term.id,
                            branch_id=branch_for_new())
                        db.session.add(assignment)
                        db.session.commit()
                    assignment_id = assignment.id

            success_count, errors = import_students_from_upload(
                file,
                db,
                Student,
                ParentContact,
                branch_id=branch_for_new(),
                class_arm_assignment_id=assignment_id,
            )

            if _wants_json():
                msg = f'Successfully imported {success_count} student(s).' if success_count else 'No students imported.'
                return jsonify({'ok': True, 'message': msg, 'imported': success_count,
                                'warnings': errors[:8], 'more': max(0, len(errors) - 8)})
            if success_count > 0:
                flash(f'Successfully imported {success_count} students!', 'success')
            if errors:
                for error in errors[:5]:  # Show first 5 errors
                    flash(error, 'warning')
                if len(errors) > 5:
                    flash(f'... and {len(errors) - 5} more errors', 'warning')

        except Exception as e:
            return _fail(f'Error importing file: {str(e)}')

        return redirect(url_for('reports.import_students'))

    classes = SchoolClass.query.filter_by(is_active=True).order_by(SchoolClass.level).all()
    arms = ClassArm.query.filter_by(is_active=True).order_by(ClassArm.name).all()
    return _render({
        'page': 'import',
        'classes': [{'id': c.id, 'name': c.name} for c in classes],
        'arms': [{'id': a.id, 'name': a.name} for a in arms],
        'upload_url': url_for('reports.import_students'),
        'template_url': url_for('reports.export_template'),
    })


# ============================================================================
# CHARTS DATA API
# ============================================================================

@reports_bp.route('/api/charts/gender-distribution')
@login_required
def api_gender_distribution():
    """Get gender distribution data for charts"""
    male_count = scope_query(Student.query.filter_by(is_active=True, gender='Male'), Student).count()
    female_count = scope_query(Student.query.filter_by(is_active=True, gender='Female'), Student).count()
    
    return jsonify({
        'labels': ['Male', 'Female'],
        'data': [male_count, female_count],
        'backgroundColor': ['#4CAF50', '#E91E63']
    })


@reports_bp.route('/api/charts/religion-distribution')
@login_required
def api_religion_distribution():
    """Get religion distribution data for charts"""
    from sqlalchemy import func
    
    _rq = db.session.query(
        Student.religion,
        func.count(Student.id)
    ).filter(
        Student.is_active == True,
        Student.religion != None
    )
    _bid = viewing_branch_id()
    if _bid is not None:
        _rq = _rq.filter(Student.branch_id == _bid)
    distribution = _rq.group_by(Student.religion).all()
    
    labels = [d[0] or 'Not Specified' for d in distribution]
    data = [d[1] for d in distribution]
    
    colors = ['#2196F3', '#4CAF50', '#FFC107', '#9C27B0', '#FF5722']
    
    return jsonify({
        'labels': labels,
        'data': data,
        'backgroundColor': colors[:len(labels)]
    })


@reports_bp.route('/api/charts/enrollment-by-class')
@login_required
def api_enrollment_by_class():
    """Get enrollment distribution by class"""
    from models import ClassArmAssignment, StudentEnrollment, SchoolClass
    
    active_term = get_active_term()
    
    if not active_term:
        return jsonify({'labels': [], 'data': []})
    
    # Get enrollment counts by class
    from sqlalchemy import func
    
    _enq = db.session.query(
        SchoolClass.name,
        func.count(StudentEnrollment.id)
    ).join(
        ClassArmAssignment,
        ClassArmAssignment.class_id == SchoolClass.id
    ).join(
        StudentEnrollment,
        StudentEnrollment.class_arm_assignment_id == ClassArmAssignment.id
    ).filter(
        ClassArmAssignment.term_id == active_term.id,
        StudentEnrollment.is_active == True
    )
    _bid = viewing_branch_id()
    if _bid is not None:
        _enq = _enq.filter(ClassArmAssignment.branch_id == _bid)
    enrollments = _enq.group_by(SchoolClass.name).order_by(SchoolClass.level).all()
    
    return jsonify({
        'labels': [e[0] for e in enrollments],
        'data': [e[1] for e in enrollments],
        'backgroundColor': '#2196F3'
    })


@reports_bp.route('/api/charts/attendance-trend')
@login_required
def api_attendance_trend():
    """Get attendance trend data"""
    from models import Attendance, Week
    from sqlalchemy import func
    
    active_term = get_active_term()
    
    if not active_term:
        return jsonify({'labels': [], 'data': []})
    
    weeks = Week.query.filter_by(term_id=active_term.id).order_by(Week.week_number).all()
    
    from sqlalchemy import case
    weekly_data = []
    for week in weeks:
        # Get total attendance for this week (morning + afternoon ticks)
        attendance = db.session.query(
            func.sum(
                case((Attendance.morning_present == True, 1), else_=0)
                + case((Attendance.afternoon_present == True, 1), else_=0)
            )
        ).filter(Attendance.week_id == week.id).scalar() or 0
        
        weekly_data.append({
            'week': f'Week {week.week_number}',
            'attendance': attendance
        })
    
    return jsonify({
        'labels': [w['week'] for w in weekly_data],
        'data': [w['attendance'] for w in weekly_data],
        'borderColor': '#4CAF50',
        'backgroundColor': 'rgba(76, 175, 80, 0.1)'
    })


@reports_bp.route('/api/charts/waec-grade-distribution')
@login_required
def api_waec_grade_distribution():
    """Get WAEC grade distribution"""
    from models import WAECResult
    from sqlalchemy import func
    from utils.helpers import WAEC_GRADES
    
    year = request.args.get('year', type=int)
    
    query = scope_by_student(db.session.query(
        WAECResult.grade,
        func.count(WAECResult.id)
    ), WAECResult)

    if year:
        query = query.filter(WAECResult.exam_year == year)

    distribution = query.group_by(WAECResult.grade).all()
    
    grade_counts = {g: 0 for g in WAEC_GRADES}
    for grade, count in distribution:
        if grade in grade_counts:
            grade_counts[grade] = count
    
    colors = {
        'A1': '#4CAF50', 'B2': '#8BC34A', 'B3': '#CDDC39',
        'C4': '#FFEB3B', 'C5': '#FFC107', 'C6': '#FF9800',
        'D7': '#FF5722', 'E8': '#F44336', 'F9': '#B71C1C'
    }
    
    return jsonify({
        'labels': WAEC_GRADES,
        'data': [grade_counts[g] for g in WAEC_GRADES],
        'backgroundColor': [colors[g] for g in WAEC_GRADES]
    })


@reports_bp.route('/api/charts/jamb-score-distribution')
@login_required
def api_jamb_score_distribution():
    """Get JAMB score distribution"""
    from models import JAMBResult
    
    year = request.args.get('year', type=int)
    
    query = scope_by_student(JAMBResult.query, JAMBResult)
    if year:
        query = query.filter_by(exam_year=year)
    
    results = query.all()
    
    if not results:
        return jsonify({'labels': [], 'data': []})
    
    ranges = [
        ('0-100', 0, 100),
        ('101-150', 101, 150),
        ('151-200', 151, 200),
        ('201-250', 201, 250),
        ('251-300', 251, 300),
        ('301-350', 301, 350),
        ('351-400', 351, 400)
    ]
    
    distribution = []
    for label, low, high in ranges:
        count = sum(1 for r in results if low <= r.total_score <= high)
        distribution.append({'range': label, 'count': count})
    
    return jsonify({
        'labels': [d['range'] for d in distribution],
        'data': [d['count'] for d in distribution],
        'backgroundColor': '#2196F3'
    })


# ============================================================================
# SUMMARY REPORTS
# ============================================================================

@reports_bp.route('/summary')
@login_required
def summary_report():
    """Generate summary report"""
    from models import ClassArmAssignment, StudentEnrollment, WAECResult, JAMBResult
    
    active_session = get_active_session()
    active_term = get_active_term()
    
    # Student statistics (branch-scoped)
    _bid = viewing_branch_id()
    total_students = scope_query(Student.query.filter_by(is_active=True), Student).count()
    male_students = scope_query(Student.query.filter_by(is_active=True, gender='Male'), Student).count()
    female_students = scope_query(Student.query.filter_by(is_active=True, gender='Female'), Student).count()

    # Enrollment statistics
    active_enrollments = 0
    if active_term:
        _eq = StudentEnrollment.query.join(ClassArmAssignment).filter(
            ClassArmAssignment.term_id == active_term.id,
            StudentEnrollment.is_active == True)
        if _bid is not None:
            _eq = _eq.filter(ClassArmAssignment.branch_id == _bid)
        active_enrollments = _eq.count()

    # Results statistics
    students_with_waec = scope_by_student(db.session.query(WAECResult.student_id), WAECResult).distinct().count()
    students_with_jamb = scope_by_student(db.session.query(JAMBResult.student_id), JAMBResult).distinct().count()
    
    return _render({
        'page': 'summary',
        'kpis': {'total': total_students, 'male': male_students, 'female': female_students,
                 'enrolled': active_enrollments},
        'quick_info': {
            'session': active_session.name if active_session else 'Not Set',
            'term': active_term.name if active_term else 'Not Set',
            'with_waec': students_with_waec, 'with_jamb': students_with_jamb,
        },
        'chart_urls': {'gender': url_for('reports.api_gender_distribution'),
                       'attendance': url_for('reports.api_attendance_trend')},
    })
