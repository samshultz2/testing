"""results blueprint — analytics routes (split from the former routes/results.py)."""
from routes.results import *  # noqa: F401,F403
from utils.web_exports import csv_response


@results_bp.route('/')
@login_required
def index():
    """Results main page with overview"""
    waec_count = WAECResult.query.count()
    jamb_count = JAMBResult.query.count()
    
    waec_years = db.session.query(WAECResult.exam_year).distinct().order_by(WAECResult.exam_year.desc()).all()
    jamb_years = db.session.query(JAMBResult.exam_year).distinct().order_by(JAMBResult.exam_year.desc()).all()
    waec_years = [y[0] for y in waec_years]
    jamb_years = [y[0] for y in jamb_years]

    return _render({
        'page': 'index',
        'waec_count': waec_count, 'jamb_count': jamb_count,
        'waec_years': waec_years, 'jamb_years': jamb_years,
        'urls': {
            'waec_dashboard': url_for('results.waec_list'), 'add_waec': url_for('results.add_waec'),
            'jamb_dashboard': url_for('results.jamb_list'), 'add_jamb': url_for('results.add_jamb'),
            'export_waec': url_for('results.export_waec', year=waec_years[0]) if waec_years else '',
            'export_jamb': url_for('results.export_jamb', year=jamb_years[0]) if jamb_years else '',
        },
    })


@results_bp.route('/subject-enrolment')
@login_required
def subject_enrolment():
    """Report: how many students are enrolled for each WAEC / JAMB subject."""
    only_sss3 = request.args.get('scope', 'sss3') != 'all'
    if only_sss3:
        students = get_sss3_students()
    else:
        students = scope_query(Student.query.filter_by(is_active=True), Student).order_by(Student.surname).all()

    waec_counts = {}
    jamb_counts = {}
    waec_enrolled = 0
    jamb_enrolled = 0
    for s in students:
        wl = s.waec_subject_list
        jl = s.jamb_subject_list
        if wl:
            waec_enrolled += 1
        if jl:
            jamb_enrolled += 1
        for subj in wl:
            waec_counts[subj] = waec_counts.get(subj, 0) + 1
        for subj in jl:
            jamb_counts[subj] = jamb_counts.get(subj, 0) + 1

    waec_rows = sorted(waec_counts.items(), key=lambda x: (-x[1], x[0]))
    jamb_rows = sorted(jamb_counts.items(), key=lambda x: (-x[1], x[0]))
    scope = 'sss3' if only_sss3 else 'all'

    def _rows(rows, enrolled, exam):
        return [{'subject': subj, 'count': cnt,
                 'pct': round(cnt / enrolled * 100) if enrolled else 0,
                 'url': url_for('results.subject_enrolment_detail', exam=exam, subject=subj, scope=scope)}
                for subj, cnt in rows]

    return _render({
        'page': 'subject_enrolment', 'only_sss3': only_sss3, 'student_count': len(students),
        'waec_enrolled': waec_enrolled, 'jamb_enrolled': jamb_enrolled,
        'waec_rows': _rows(waec_rows, waec_enrolled, 'waec'),
        'jamb_rows': _rows(jamb_rows, jamb_enrolled, 'jamb'),
        'urls': {'sss3': url_for('results.subject_enrolment', scope='sss3'),
                 'all': url_for('results.subject_enrolment', scope='all')},
    })


@results_bp.route('/student/<int:student_id>/report')
@login_required
def student_report(student_id):
    """A consolidated, print/PDF-ready exam report for one student."""
    from models.mock_jamb import MockJAMBResult, MockJAMBExam, MockJAMBAnalytics

    student = db.get_or_404(Student, student_id)
    require_branch_access(student.branch_id)
    from utils.access_control import teacher_form_student_ids
    tids = teacher_form_student_ids()
    if tids is not None and student.id not in tids:
        abort(403)
    pass_grades = set(AcademicAnalytics.PASS_GRADES)
    distinction_grades = set(AcademicAnalytics.DISTINCTION_GRADES)

    # WAEC grouped by year with summary counts.
    waec = student.waec_results.order_by(WAECResult.exam_year.desc(), WAECResult.subject).all()
    waec_by_year = {}
    for r in waec:
        y = waec_by_year.setdefault(r.exam_year, {'rows': [], 'credits': 0, 'distinctions': 0})
        y['rows'].append(r)
        if r.grade in pass_grades:
            y['credits'] += 1
        if r.grade in distinction_grades:
            y['distinctions'] += 1
    waec_years = sorted(waec_by_year.items(), reverse=True)

    # JAMB results (most recent first), with subject breakdown.
    jamb = student.jamb_results.order_by(JAMBResult.exam_year.desc()).all()
    jamb_list = []
    for r in jamb:
        subs = []
        for n, sc in [(r.subject1, r.subject1_score), (r.subject2, r.subject2_score),
                      (r.subject3, r.subject3_score), (r.subject4, r.subject4_score)]:
            if n:
                subs.append({'subject': n, 'score': sc or 0})
        jamb_list.append({'year': r.exam_year, 'total': r.total_score, 'subjects': subs,
                          'level': AcademicAnalytics._jamb_performance_level(r.total_score)})

    # Mock JAMB progression + prediction (active session).
    mock_rows = MockJAMBResult.query.filter_by(student_id=student_id).join(MockJAMBExam).order_by(
        MockJAMBExam.exam_number).all()
    mock_progress = [{'name': m.exam.display_name, 'date': m.exam.exam_date, 'score': m.total_score}
                     for m in mock_rows]
    prediction = None
    active_session = get_active_session()
    if active_session:
        prediction = MockJAMBAnalytics.predict_real_jamb(student_id, active_session.id)

    from utils.admission import assess_admission
    admission = assess_admission(student)

    # Forward-looking readiness from the mock trajectory (actionable before the
    # real exams), plus the intended JAMB-combination check. The mock-based
    # prediction is corrected by the historical mock->real bias (cached).
    from utils import exam_insights, exam_trends
    bias = exam_insights.get_jamb_bias()
    if prediction:
        adj, applied = exam_insights.calibrate_jamb(prediction.get('predicted_score'), bias)
        if applied is not None:
            prediction['raw_score'] = prediction.get('predicted_score')
            prediction['predicted_score'] = adj
            prediction['calibrated_by'] = applied
    readiness = exam_insights.admission_readiness(
        student, active_session.id if active_session else None, calibration=bias)
    jamb_combo = exam_insights.jamb_subject_combo_check(student)
    # Difficulty-adjusted standing among peers (z-score / percentile per sitting).
    standing = exam_trends.standardized_mock_jamb_progress(
        student_id, active_session.id if active_session else None)

    # University aspiration: the student's OWN chosen-course verdict (gap to their
    # target + missing subjects) and the courses they're projected competitive
    # for. Reuses the readiness we already computed above — no recompute.
    sid = active_session.id if active_session else None
    aspiration = recommendations = subject_diag = None
    try:
        from utils.aspiration import (course_eligibility, recommend_courses,
                                       course_subject_diagnosis, ELIGIBILITY_LABELS)
        aspiration = course_eligibility(student, sid, readiness=readiness)
        if aspiration:
            aspiration['status_label'] = ELIGIBILITY_LABELS.get(aspiration.get('status'), aspiration.get('status'))
        recommendations = recommend_courses(student, sid, limit=8)
        # Attribute the target gap to specific course-relevant subjects.
        subject_diag = course_subject_diagnosis(student, sid)
    except Exception:
        aspiration = recommendations = subject_diag = None

    return render_template('results/student_report.html',
        student=student,
        waec_years=waec_years,
        jamb_list=jamb_list,
        mock_progress=mock_progress,
        prediction=prediction,
        admission=admission,
        readiness=readiness,
        jamb_combo=jamb_combo,
        standing=standing,
        aspiration=aspiration,
        recommendations=recommendations,
        subject_diag=subject_diag,
        generated=_date.today()
    )


@results_bp.route('/readiness')
@login_required
def readiness():
    """Actionable exam-readiness checklist for the SSS3 cohort."""
    students = get_sss3_students()
    total = len(students)

    # Readiness is judged from the MOCKS (Mock WAEC / Mock JAMB), since real JAMB
    # (2nd term) and WAEC (3rd term) don't exist yet for most of the cohort.
    no_stream, no_jamb, no_waec, no_jamb_subjects, no_waec_subjects = [], [], [], [], []
    below_target = []
    for s in students:
        if not s.stream:
            no_stream.append(s)
        if s.mock_jamb_results.count() == 0:
            no_jamb.append(s)
        if s.mock_waec_results.count() == 0:
            no_waec.append(s)
        if not s.jamb_subject_list:
            no_jamb_subjects.append(s)
        if not s.waec_subject_list:
            no_waec_subjects.append(s)
        if s.jamb_target:
            mocks = [m.total_score for m in s.mock_jamb_results.all()]
            best = max(mocks) if mocks else 0
            if best < s.jamb_target:
                below_target.append(s)

    groups = [
        {'key': 'no_jamb', 'title': 'No Mock JAMB result yet', 'icon': 'fa-file-contract', 'students': no_jamb},
        {'key': 'no_waec', 'title': 'No Mock WAEC result yet', 'icon': 'fa-file-alt', 'students': no_waec},
        {'key': 'below_target', 'title': 'Mock JAMB below their target', 'icon': 'fa-bullseye', 'students': below_target},
        {'key': 'no_stream', 'title': 'No stream / track set', 'icon': 'fa-route', 'students': no_stream},
        {'key': 'no_jamb_subjects', 'title': 'No JAMB subjects on profile', 'icon': 'fa-list', 'students': no_jamb_subjects},
        {'key': 'no_waec_subjects', 'title': 'No WAEC subjects on profile', 'icon': 'fa-list-check', 'students': no_waec_subjects},
    ]
    ready = total - len({s.id for g in groups for s in g['students']})

    def _action(key, sid):
        if key == 'no_jamb':
            return {'label': 'Mock JAMB', 'url': url_for('mock_jamb.index')}
        if key == 'no_waec':
            return {'label': 'Mock WAEC', 'url': url_for('mock_waec.index')}
        if key == 'no_jamb_subjects':
            return {'label': 'JAMB subjects', 'url': url_for('main.edit_student', student_id=sid)}
        if key == 'no_waec_subjects':
            return {'label': 'WAEC subjects', 'url': url_for('main.edit_student', student_id=sid)}
        return {'label': 'Edit', 'url': url_for('main.edit_student', student_id=sid)}

    return _render({
        'page': 'readiness', 'total': total, 'ready': ready,
        'groups': [{'key': g['key'], 'title': g['title'], 'icon': g['icon'],
                    'students': [{'id': s.id, 'full_name': s.full_name, 'student_id': s.student_id,
                                  'action': _action(g['key'], s.id)} for s in g['students']]}
                   for g in groups],
    })


@results_bp.route('/admission-readiness')
@login_required
def readiness_funnel():
    """Cohort admission-readiness funnel: how many SSS3 students are projected to
    get 5 credits incl. English & Maths, clear the JAMB baseline, and clear both
    (admission-ready) — from actual results where available, else the mocks."""
    from utils import exam_insights
    session = get_active_session()
    students = get_sss3_students()
    bias = exam_insights.get_jamb_bias()      # historical mock->real correction (cached)
    funnel = exam_insights.cohort_readiness(
        students, session.id if session else None, calibration=bias)
    # Rows sorted worst-first so intervention candidates surface at the top.
    order = {'NOT_READY': 0, 'AT_RISK': 1, 'CONDITIONAL': 2, 'READY': 3, 'NO_DATA': 4}
    rows = sorted(funnel['rows'], key=lambda r: order.get(r['readiness']['status'], 9))
    return render_template('results/readiness_funnel.html',
                           funnel=funnel, rows=rows, active_session=session, bias=bias)


@results_bp.route('/analytics')
@login_required
def analytics_hub():
    """One-stop analytics hub: every WAEC/JAMB stat, correlation and projection."""
    waec_years = [y[0] for y in db.session.query(WAECResult.exam_year).distinct().all()]
    jamb_years = [y[0] for y in db.session.query(JAMBResult.exam_year).distinct().all()]
    years = sorted(set(waec_years + jamb_years), reverse=True)

    year = resolve_exam_year(request.args.get('year', type=int), years)
    compare_year = request.args.get('compare', type=int)

    from utils.branch_scope import viewing_branch_id
    bid = viewing_branch_id()
    waec_stats = waec_school_stats(year, bid) if year else None
    jamb_stats = jamb_school_stats(year, bid) if year else None
    correlation = waec_jamb_correlation(year, bid) if year else None
    yoy = AcademicAnalytics.get_year_over_year_comparison(bid)

    # Gender breakdowns for the selected year.
    def gender_split(model):
        q = db.session.query(Student.gender, func.count(func.distinct(Student.id))).join(
            model, Student.id == model.student_id
        ).filter(model.exam_year == year)
        if bid is not None:
            q = q.filter(Student.branch_id == bid)
        rows = q.group_by(Student.gender).all()
        return {g or 'Unknown': c for g, c in rows}

    waec_gender = gender_split(WAECResult) if (year and waec_stats) else {}
    jamb_gender = gender_split(JAMBResult) if (year and jamb_stats) else {}

    # Gender-comparison breakdown: pass/distinction rates (WAEC) and mean score
    # / >=200 rate (JAMB) split by gender for the selected year.
    waec_gender_stats = []
    if year and waec_stats:
        _q = db.session.query(Student.gender, WAECResult.grade).join(
            WAECResult, Student.id == WAECResult.student_id
        ).filter(WAECResult.exam_year == year)
        if bid is not None:
            _q = _q.filter(Student.branch_id == bid)
        rows = _q.all()
        by_gender = defaultdict(list)
        for g, grade in rows:
            by_gender[g or 'Unknown'].append(grade)
        for g, grades in by_gender.items():
            total = len(grades)
            passes = sum(1 for x in grades if x in AcademicAnalytics.PASS_GRADES)
            dist = sum(1 for x in grades if x in AcademicAnalytics.DISTINCTION_GRADES)
            pts = [AcademicAnalytics.GRADE_POINTS.get(x, 9) for x in grades]
            waec_gender_stats.append({
                'gender': g,
                'entries': total,
                'pass_rate': round(passes / total * 100, 1) if total else 0,
                'distinction_rate': round(dist / total * 100, 1) if total else 0,
                'mean_points': round(sum(pts) / len(pts), 2) if pts else 0,
            })
        waec_gender_stats.sort(key=lambda x: x['gender'])

    jamb_gender_stats = []
    if year and jamb_stats:
        _q = db.session.query(Student.gender, JAMBResult.total_score).join(
            JAMBResult, Student.id == JAMBResult.student_id
        ).filter(JAMBResult.exam_year == year)
        if bid is not None:
            _q = _q.filter(Student.branch_id == bid)
        rows = _q.all()
        by_gender = defaultdict(list)
        for g, score in rows:
            by_gender[g or 'Unknown'].append(score)
        for g, scores in by_gender.items():
            total = len(scores)
            jamb_gender_stats.append({
                'gender': g,
                'candidates': total,
                'mean_score': round(sum(scores) / total, 1) if total else 0,
                'above_200': sum(1 for s in scores if s >= 200),
                'above_200_rate': round(sum(1 for s in scores if s >= 200) / total * 100, 1) if total else 0,
                'max_score': max(scores) if scores else 0,
            })
        jamb_gender_stats.sort(key=lambda x: x['gender'])

    # JAMB mean projection (simple linear fit) computed directly from JAMB years.
    projection = None
    jamb_means = []
    for jy in sorted(set(jamb_years)):
        rs = scope_by_student(JAMBResult.query.filter_by(exam_year=jy), JAMBResult).all()
        if rs:
            jamb_means.append({'year': jy, 'mean_score': round(sum(r.total_score for r in rs) / len(rs), 1)})
    if len(jamb_means) >= 2:
        xs = [t['year'] for t in jamb_means]
        ys = [t['mean_score'] for t in jamb_means]
        n = len(xs)
        mean_x = sum(xs) / n
        mean_y = sum(ys) / n
        denom = sum((x - mean_x) ** 2 for x in xs)
        slope = (sum((xs[i] - mean_x) * (ys[i] - mean_y) for i in range(n)) / denom) if denom else 0
        next_year = max(xs) + 1
        projected = round(mean_y + slope * (next_year - mean_x), 1)
        projected = max(0, min(400, projected))
        projection = {
            'next_year': next_year,
            'projected_mean': projected,
            'direction': 'up' if slope > 0.5 else 'down' if slope < -0.5 else 'flat',
            'slope_per_year': round(slope, 1),
            'latest_mean': ys[-1],
        }

    # University-cutoff readiness (JAMB >= 200) for the selected year.
    cutoff = None
    if year and jamb_stats:
        total = jamb_stats['total_students']
        cutoff = {
            'eligible_200': jamb_stats['above_200'],
            'eligible_200_pct': round(jamb_stats['above_200'] / total * 100, 1) if total else 0,
            'competitive_250': jamb_stats['above_250'],
            'competitive_250_pct': round(jamb_stats['above_250'] / total * 100, 1) if total else 0,
            'elite_300': jamb_stats['above_300'],
            'elite_300_pct': round(jamb_stats['above_300'] / total * 100, 1) if total else 0,
        }

    # Class/arm comparison + internal-vs-JAMB correlation for the selected year.
    class_compare = []
    internal_corr = None
    if year:
        active_term = get_active_term()
        arm_map = {}
        if active_term:
            enrs = StudentEnrollment.query.join(ClassArmAssignment).filter(
                ClassArmAssignment.term_id == active_term.id,
                StudentEnrollment.is_active == True
            ).all()
            for e in enrs:
                arm_map.setdefault(e.student_id, e.class_arm_assignment.display_name)

        jamb_by_arm = defaultdict(list)
        for r in scope_by_student(JAMBResult.query.filter_by(exam_year=year), JAMBResult).all():
            arm = arm_map.get(r.student_id)
            if arm:
                jamb_by_arm[arm].append(r.total_score)
        waec_by_arm = defaultdict(lambda: {'pass': 0, 'total': 0})
        for r in scope_by_student(WAECResult.query.filter_by(exam_year=year), WAECResult).all():
            arm = arm_map.get(r.student_id)
            if arm:
                waec_by_arm[arm]['total'] += 1
                if r.grade in AcademicAnalytics.PASS_GRADES:
                    waec_by_arm[arm]['pass'] += 1
        for arm in sorted(set(list(jamb_by_arm) + list(waec_by_arm))):
            js = jamb_by_arm.get(arm, [])
            w = waec_by_arm.get(arm, {'pass': 0, 'total': 0})
            class_compare.append({
                'arm': arm,
                'jamb_count': len(js),
                'jamb_mean': round(sum(js) / len(js), 1) if js else 0,
                'waec_pass_rate': round(w['pass'] / w['total'] * 100, 1) if w['total'] else 0,
                'waec_entries': w['total'],
            })

        pairs = []
        for r in scope_by_student(JAMBResult.query.filter_by(exam_year=year), JAMBResult).all():
            ts = TermSummary.query.filter_by(student_id=r.student_id).order_by(
                TermSummary.term_id.desc()).first()
            if ts and ts.average_score is not None:
                pairs.append((ts.average_score, r.total_score))
        if len(pairs) >= 5:
            xs = [p[0] for p in pairs]
            ys = [p[1] for p in pairs]
            internal_corr = {
                'n': len(pairs),
                'r': round(AcademicAnalytics._pearson_correlation(xs, ys), 3),
                'mean_internal': round(sum(xs) / len(xs), 1),
                'mean_jamb': round(sum(ys) / len(ys), 1),
            }

    # Trends from data we capture but didn't previously analyse.
    from utils import exam_trends
    active_sess = get_active_session()

    mock_trend = _mock_jamb_trend(bid)
    mock_waec_trend = _mock_waec_trend(bid)
    at_risk = _at_risk_register(limit=25)

    # Executive Smart Insights — synthesise the above stats into a ranked,
    # actionable "what / why / do next" summary (pure, adds no queries).
    from utils import exam_intelligence
    insights = exam_intelligence.school_insights(
        year=year, waec_stats=waec_stats, jamb_stats=jamb_stats,
        correlation=correlation, projection=projection, cutoff=cutoff,
        class_compare=class_compare, internal_corr=internal_corr,
        at_risk=at_risk, mock_trend=mock_trend,
        waec_gender_stats=waec_gender_stats, jamb_gender_stats=jamb_gender_stats,
        urls={'readiness': url_for('results.readiness_funnel'),
              'at_risk': url_for('results.api_at_risk')})

    from utils import exam_subjects
    scorecard = exam_subjects.subject_scorecard(
        waec_stats, jamb_stats, sss3_subject_teachers()) if year else []

    from utils.exam_refresh import refreshed_at
    return render_template('results/analytics_hub.html',
        insights=insights,
        analytics_refreshed_at=refreshed_at(),
        subject_scorecard=scorecard,
        scorecard_summary=exam_subjects.scorecard_summary(scorecard),
        branch_compare=branch_comparison(year) if year else [],
        year_compare=year_comparison(year, compare_year, bid) if year else None,
        compare_year=compare_year,
        years=years,
        selected_year=year,
        jamb_subjects=exam_trends.jamb_subject_breakdown(bid, year),
        waec_subject_gains=exam_trends.mock_waec_subject_gains(active_sess.id) if active_sess else {},
        attendance_corr=exam_trends.attendance_performance_correlation(get_sss3_students(), 'jamb'),
        class_compare=class_compare,
        internal_corr=internal_corr,
        waec_stats=waec_stats,
        jamb_stats=jamb_stats,
        correlation=correlation,
        yoy=yoy,
        waec_gender=waec_gender,
        jamb_gender=jamb_gender,
        waec_gender_stats=waec_gender_stats,
        jamb_gender_stats=jamb_gender_stats,
        projection=projection,
        cutoff=cutoff,
        at_risk=at_risk,
        mock_trend=mock_trend,
        mock_waec_trend=mock_waec_trend,
        recompute_url=url_for('results.recompute_analytics'),
    )


@results_bp.route('/analytics/by-class')
@login_required
def analytics_by_class():
    """External-exam performance rolled up to the class-arm level for a year —
    cohort-aware (maps each candidate to their latest senior-class arm)."""
    from utils.exam_class_league import exam_class_league
    from utils.branch_scope import viewing_branch_id
    waec_years = [y[0] for y in db.session.query(WAECResult.exam_year).distinct().all()]
    jamb_years = [y[0] for y in db.session.query(JAMBResult.exam_year).distinct().all()]
    years = sorted(set(waec_years + jamb_years), reverse=True)
    year = resolve_exam_year(request.args.get('year', type=int), years)
    data = exam_class_league(year, viewing_branch_id()) if year else None
    return render_template('results/analytics_by_class.html',
                           data=data, years=years, selected_year=year)


@results_bp.route('/analytics/trends')
@login_required
def analytics_trends():
    """Cross-year external-exam trends: how the WAEC 5-credits-incl-core / credit
    / F9 rates and the JAMB average & ≥200 rate move across every year on record.
    Branch-scoped to the branch being viewed."""
    from utils.branch_scope import viewing_branch_id
    bid = viewing_branch_id()
    waec = AcademicAnalytics.get_waec_multiyear_trends(bid)
    jamb = AcademicAnalytics.get_jamb_multiyear_trends(bid)
    return render_template('results/analytics_trends.html', waec=waec, jamb=jamb)


@results_bp.route('/analytics/watchlist')
@login_required
def at_risk_watchlist():
    """Live at-risk watchlist: SSS3 candidates projected off track for admission
    (from mock signals), grouped by class arm. Works before the risk engine runs."""
    from utils.at_risk_live import live_at_risk
    from utils.branch_scope import viewing_branch_id
    data = live_at_risk(branch_id=viewing_branch_id())
    return render_template('results/at_risk_watchlist.html', data=data,
                           action_plan_url='results.student_action_plan')


@results_bp.route('/analytics/aspirations')
@login_required
def aspiration_hub():
    """University-aspiration hub: target coverage, course-eligibility mix,
    most-wanted universities/courses, a JAMB subject-mismatch fix list, and the
    admission funnel with conversion. Projected live from mock signals."""
    from utils.aspiration_analytics import aspiration_overview
    from utils.branch_scope import viewing_branch_id
    data = aspiration_overview(branch_id=viewing_branch_id())
    return render_template('results/aspiration_hub.html', data=data)


@results_bp.route('/analytics/by-class/export')
@login_required
def analytics_by_class_export():
    """Export the class-arm external-exam league (Excel)."""
    from utils.exam_class_league import exam_class_league
    from utils.branch_scope import viewing_branch_id
    from utils.web_exports import xlsx_response
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    year = request.args.get('year', type=int)
    if not year:
        flash('Select a year first.', 'error')
        return redirect(url_for('results.analytics_by_class'))
    data = exam_class_league(year, viewing_branch_id())
    if not data or data['meta'].get('insufficient'):
        flash('No external results to group by class for that year.', 'warning')
        return redirect(url_for('results.analytics_by_class', year=year))
    wb = Workbook(); ws = wb.active; ws.title = f'By class {year}'
    head = ['Class arm', 'Students', 'JAMB candidates', 'JAMB mean', 'JAMB ≥ cutoff %',
            'WAEC students', 'Credit rate %', 'Distinction rate %', '5 credits incl. core %']
    ws.append(head)
    for c in ws[1]:
        c.fill = PatternFill('solid', fgColor='0D6A4E'); c.font = Font(bold=True, color='FFFFFF')
        c.alignment = Alignment(horizontal='center')
    for u in data['units']:
        ws.append([u['label'], u['students'], u['jamb_candidates'], u['jamb_mean'],
                   u['jamb_above_rate'], u['waec_students'], u['credit_rate'],
                   u['distinction_rate'], u['five_core_rate']])
    for col in ws.columns:
        w = max((len(str(c.value)) if c.value is not None else 0) for c in col) + 2
        ws.column_dimensions[col[0].column_letter].width = min(max(w, 12), 40)
    return xlsx_response(wb, f'exam_by_class_{year}.xlsx')


@results_bp.route('/analytics/export')
@login_required
def analytics_export():
    """Export the analytics hub for a year to a multi-sheet Excel workbook."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from io import BytesIO

    year = request.args.get('year', type=int)
    if not year:
        flash('Select a year to export.', 'error')
        return redirect(url_for('results.analytics_hub'))

    bid = viewing_branch_id()
    waec_stats = AcademicAnalytics.get_waec_school_statistics(year, bid)
    jamb_stats = AcademicAnalytics.get_jamb_school_statistics(year, bid)
    correlation = AcademicAnalytics.calculate_waec_jamb_correlation(year, bid)

    wb = Workbook()
    head_font = Font(bold=True, color='FFFFFF')
    head_fill = PatternFill(start_color='1a5f4a', end_color='1a5f4a', fill_type='solid')
    title_font = Font(bold=True, size=13)

    def style_header(ws, row=1):
        for cell in ws[row]:
            cell.font = head_font
            cell.fill = head_fill
            cell.alignment = Alignment(horizontal='center')

    # --- Overview sheet ---
    ws = wb.active
    ws.title = 'Overview'
    ws['A1'] = f'Exam Analytics — {year}'
    ws['A1'].font = title_font
    r = 3
    ws.cell(row=r, column=1, value='JAMB'); ws.cell(row=r, column=1).font = Font(bold=True); r += 1
    if jamb_stats:
        for label, key in [('Candidates', 'total_students'), ('Mean', 'mean_score'),
                           ('Median', 'median_score'), ('Highest', 'max_score'),
                           ('Lowest', 'min_score'), ('Std Dev', 'std_deviation'),
                           ('>=200', 'above_200'), ('>=250', 'above_250'), ('>=300', 'above_300')]:
            ws.cell(row=r, column=1, value=label); ws.cell(row=r, column=2, value=jamb_stats[key]); r += 1
    else:
        ws.cell(row=r, column=1, value='No JAMB data'); r += 1
    r += 1
    ws.cell(row=r, column=1, value='WAEC'); ws.cell(row=r, column=1).font = Font(bold=True); r += 1
    if waec_stats:
        for label, key in [('Students', 'unique_students'), ('Subject Entries', 'total_results'),
                           ('Pass Rate %', 'overall_pass_rate'), ('Distinction Rate %', 'overall_distinction_rate')]:
            ws.cell(row=r, column=1, value=label); ws.cell(row=r, column=2, value=waec_stats[key]); r += 1
    else:
        ws.cell(row=r, column=1, value='No WAEC data'); r += 1
    r += 1
    ws.cell(row=r, column=1, value='WAEC↔JAMB Correlation'); ws.cell(row=r, column=1).font = Font(bold=True); r += 1
    if correlation and not correlation.get('error'):
        for label, key in [('Pearson r', 'correlation_coefficient'), ('Predictive Power', 'predictive_power'),
                           ('Paired Students', 'sample_size')]:
            ws.cell(row=r, column=1, value=label); ws.cell(row=r, column=2, value=correlation[key]); r += 1
    else:
        ws.cell(row=r, column=1, value='Insufficient paired data'); r += 1
    ws.column_dimensions['A'].width = 22
    ws.column_dimensions['B'].width = 16

    # --- JAMB subjects ---
    if jamb_stats and jamb_stats['subject_analysis']:
        ws = wb.create_sheet('JAMB Subjects')
        ws.append(['Subject', 'Count', 'Mean', 'Max', 'Min', '>=50', '>=70'])
        style_header(ws)
        for s in jamb_stats['subject_analysis']:
            ws.append([s['subject'], s['count'], s['mean_score'], s['max_score'], s['min_score'], s['above_50'], s['above_70']])
        ws.column_dimensions['A'].width = 24

    # --- JAMB top 10 ---
    if jamb_stats and jamb_stats['top_10']:
        ws = wb.create_sheet('JAMB Top 10')
        ws.append(['Rank', 'Student', 'Score'])
        style_header(ws)
        for i, t in enumerate(jamb_stats['top_10'], 1):
            ws.append([i, t['student_name'], t['score']])
        ws.column_dimensions['B'].width = 28

    # --- WAEC subjects ---
    if waec_stats and waec_stats['subject_analysis']:
        ws = wb.create_sheet('WAEC Subjects')
        ws.append(['Subject', 'Entries', 'A1 %', 'Pass %', 'Below credit % (D7–F9)', 'Fail % (F9)'])
        style_header(ws)
        for s in sorted(waec_stats['subject_analysis'], key=lambda x: x['pass_rate'], reverse=True):
            ws.append([s['subject'], s['total_entries'], s['a1_rate'], s['pass_rate'],
                       s.get('below_credit_rate', 0), s['fail_rate']])
        ws.column_dimensions['A'].width = 24

    return xlsx_response(wb, f'exam_analytics_{year}.xlsx')


def _cutoff_from_jamb(jamb_stats):
    """University-readiness cut-off summary from JAMB school stats (mirrors the
    hub's inline computation), or None."""
    if not jamb_stats:
        return None
    total = jamb_stats['total_students']
    pct = lambda n: round(n / total * 100, 1) if total else 0
    return {'eligible_200': jamb_stats['above_200'], 'eligible_200_pct': pct(jamb_stats['above_200']),
            'competitive_250': jamb_stats['above_250'], 'competitive_250_pct': pct(jamb_stats['above_250']),
            'elite_300': jamb_stats['above_300'], 'elite_300_pct': pct(jamb_stats['above_300'])}


def _report_bundle(year, bid):
    """Shared stats + Smart Insights for the CSV / board-pack exports. Uses the
    cached school-stat wrappers so an export right after viewing the hub is free."""
    waec_stats = waec_school_stats(year, bid)
    jamb_stats = jamb_school_stats(year, bid)
    correlation = waec_jamb_correlation(year, bid)
    cutoff = _cutoff_from_jamb(jamb_stats)
    from utils import exam_intelligence
    insights = exam_intelligence.school_insights(
        year=year, waec_stats=waec_stats, jamb_stats=jamb_stats,
        correlation=correlation, cutoff=cutoff, at_risk=_at_risk_register(limit=25),
        urls={'readiness': url_for('results.readiness_funnel')})
    return waec_stats, jamb_stats, correlation, cutoff, insights


@results_bp.route('/analytics/export.csv')
@login_required
def analytics_export_csv():
    """Flat CSV of the year's key WAEC/JAMB stats + correlation (formula-guarded)."""
    import csv
    from io import StringIO
    from utils.web_exports import formula_guard as fg

    year = request.args.get('year', type=int)
    if not year:
        flash('Select a year to export.', 'error')
        return redirect(url_for('results.analytics_hub'))
    bid = viewing_branch_id()
    waec_stats, jamb_stats, correlation, cutoff, insights = _report_bundle(year, bid)

    out = StringIO()
    w = csv.writer(out)
    w.writerow(['Section', 'Metric', 'Value'])
    if jamb_stats:
        for label, key in [('Candidates', 'total_students'), ('Mean', 'mean_score'),
                           ('Median', 'median_score'), ('Highest', 'max_score'),
                           ('Lowest', 'min_score'), ('Std deviation', 'std_deviation'),
                           ('>=200', 'above_200'), ('>=250', 'above_250'), ('>=300', 'above_300')]:
            w.writerow(['JAMB', label, jamb_stats[key]])
    if cutoff:
        for label, key in [('Admissible (>=200) %', 'eligible_200_pct'),
                           ('Competitive (>=250) %', 'competitive_250_pct'),
                           ('Elite (>=300) %', 'elite_300_pct')]:
            w.writerow(['University readiness', label, cutoff[key]])
    if waec_stats:
        for label, key in [('Students', 'unique_students'), ('Subject entries', 'total_results'),
                           ('Pass rate %', 'overall_pass_rate'), ('Distinction rate %', 'overall_distinction_rate')]:
            w.writerow(['WAEC', label, waec_stats[key]])
    if correlation and not correlation.get('error'):
        for label, key in [('Pearson r', 'correlation_coefficient'),
                           ('Predictive power', 'predictive_power'), ('Paired students', 'sample_size')]:
            w.writerow(['WAEC<->JAMB', label, correlation[key]])
    if jamb_stats and jamb_stats.get('subject_analysis'):
        for s in jamb_stats['subject_analysis']:
            w.writerow(['JAMB subject', fg(s['subject']), s['mean_score']])
    if waec_stats and waec_stats.get('subject_analysis'):
        for s in sorted(waec_stats['subject_analysis'], key=lambda x: x['pass_rate'], reverse=True):
            w.writerow(['WAEC subject pass %', fg(s['subject']), s['pass_rate']])
    for i in insights:
        w.writerow(['Insight (%s)' % i['level'], fg(i['title']), fg(i['detail'])])

    return csv_response(out.getvalue(), f'exam_analytics_{year}.csv')


@results_bp.route('/analytics/board-pack')
@login_required
def analytics_board_pack():
    """One-page executive board-pack PDF: KPIs + Smart Insights + key stats."""
    from utils.web_exports import pdf_response
    from utils.exam_board_pack import board_pack_pdf
    from utils.school import school_profile

    year = request.args.get('year', type=int)
    if not year:
        flash('Select a year to export.', 'error')
        return redirect(url_for('results.analytics_hub'))
    bid = viewing_branch_id()
    waec_stats, jamb_stats, correlation, cutoff, insights = _report_bundle(year, bid)
    if not (waec_stats or jamb_stats):
        flash(f'No exam data for {year}.', 'error')
        return redirect(url_for('results.analytics_hub', year=year))

    branch_label = None
    if bid is not None:
        from models import Branch
        b = db.session.get(Branch, bid)
        branch_label = b.name if b else None

    pdf = board_pack_pdf(
        year=year, school_name=school_profile()['name'], generated=_date.today().isoformat(),
        insights=insights, jamb_stats=jamb_stats, waec_stats=waec_stats,
        cutoff=cutoff, correlation=correlation, branch_label=branch_label)
    log_action('analytics.board_pack', detail=f'year={year}, branch={bid or "all"}')
    return pdf_response(pdf, f'exam_board_pack_{year}.pdf', inline=True)


@results_bp.route('/analytics/deck.pptx')
@login_required
def analytics_deck():
    """Editable PowerPoint deck of the year's external-exam performance, for
    administration / parent / board meetings."""
    from flask import Response
    from utils.exam_deck import build_deck
    from utils.school import school_profile

    year = request.args.get('year', type=int)
    if not year:
        flash('Select a year to build the presentation.', 'error')
        return redirect(url_for('results.analytics_hub'))
    bid = viewing_branch_id()
    waec_stats, jamb_stats, correlation, cutoff, insights = _report_bundle(year, bid)
    if not (waec_stats or jamb_stats):
        flash(f'No exam data for {year}.', 'error')
        return redirect(url_for('results.analytics_hub', year=year))

    branch_label = None
    if bid is not None:
        from models import Branch
        b = db.session.get(Branch, bid)
        branch_label = b.name if b else None

    data = build_deck(
        year=year, school_name=school_profile()['name'], generated=_date.today().isoformat(),
        branch_label=branch_label, waec_stats=waec_stats, jamb_stats=jamb_stats,
        cutoff=cutoff, correlation=correlation, insights=insights)
    log_action('analytics.deck', detail=f'year={year}, branch={bid or "all"}')
    resp = Response(data, mimetype='application/vnd.openxmlformats-officedocument.presentationml.presentation')
    resp.headers['Content-Disposition'] = f'attachment; filename="exam_results_{year}.pptx"'
    return resp


@results_bp.route('/subject-enrolment/<exam>/<path:subject>')
@login_required
def subject_enrolment_detail(exam, subject):
    """List the students enrolled for a particular WAEC/JAMB subject."""
    exam = 'jamb' if exam.lower() == 'jamb' else 'waec'
    only_sss3 = request.args.get('scope', 'sss3') != 'all'
    if only_sss3:
        students = get_sss3_students()
    else:
        students = scope_query(Student.query.filter_by(is_active=True), Student).order_by(Student.surname).all()

    matched = []
    for s in students:
        enrolled = s.jamb_subject_list if exam == 'jamb' else s.waec_subject_list
        if subject in enrolled:
            matched.append(s)
    matched.sort(key=lambda s: (s.surname or '', s.first_name or ''))

    return _render({
        'page': 'subject_enrolment_detail', 'exam': exam,
        'exam_label': 'JAMB' if exam == 'jamb' else 'WAEC',
        'subject': subject, 'only_sss3': only_sss3,
        'students': [{'id': s.id, 'full_name': s.full_name, 'student_id': s.student_id,
                      'gender': s.gender or '', 'is_graduated': bool(s.is_graduated),
                      'view_url': url_for('main.view_student', student_id=s.id)} for s in matched],
        'back_url': url_for('results.subject_enrolment', scope='sss3' if only_sss3 else 'all'),
    })


@results_bp.route('/api/yoy-trends')
@login_required
def api_yoy_trends():
    """Get year-over-year performance trends"""
    data = AcademicAnalytics.get_year_over_year_comparison(viewing_branch_id())
    return jsonify(data)


@results_bp.route('/api/student-risk/<int:student_id>')
@login_required
def api_student_risk(student_id):
    """Get risk assessment for a student"""
    require_branch_access(db.get_or_404(Student, student_id).branch_id)
    risk = AcademicAnalytics.calculate_student_risk_score(student_id)
    return jsonify(risk)


@results_bp.route('/api/at-risk')
@login_required
def api_at_risk():
    out = _at_risk_register()
    return jsonify({'count': len(out), 'students': out})


@results_bp.route('/analytics/recompute', methods=['POST'])
@admin_required
def recompute_analytics():
    """Backfill/refresh persisted analytics for all in-scope students (and the
    WAEC↔JAMB correlation for recent years). Use after first deploy or a bulk
    import, since per-student rows are otherwise only written on results changes."""
    from utils.exam_refresh import run_exam_analytics_refresh
    bid = viewing_branch_id()
    # When the async-jobs flag is on, enqueue the (potentially slow) recompute so
    # the request returns immediately; the scheduler tick runs it in the
    # background. Otherwise run it synchronously exactly as before.
    from utils.jobs import async_enabled, enqueue
    if async_enabled():
        job = enqueue('analytics_recompute', {'branch_id': bid}, branch_id=bid)
        log_action('analytics.recompute', detail=f'queued job #{job.id}, branch={bid or "all"}')
        return _ok('Recompute queued — it will run in the background.',
                   url_for('results.tasks'))
    # Shared with the daily background job: recompute the SSS3 cohort, backfill
    # correlation, warm the hub caches, and stamp the refresh time.
    summary = run_exam_analytics_refresh(current_app, warm=True, branch_id=bid)
    log_action('analytics.recompute', detail=f'{summary["students"]} student(s), branch={bid or "all"}')
    return _ok(f'Recomputed analytics for {summary["students"]} student(s).',
               url_for('results.analytics_hub'))


@results_bp.route('/tasks')
@login_required
def tasks():
    """Background task list (async jobs). Empty/graceful when the feature is off
    or the table hasn't been created yet."""
    jobs = []
    try:
        from models import BackgroundJob
        jobs = (BackgroundJob.query.order_by(BackgroundJob.id.desc()).limit(50).all())
    except Exception:
        db.session.rollback()
    from utils.jobs import async_enabled
    return render_template('results/tasks.html', jobs=jobs, async_on=async_enabled())


@results_bp.route('/tasks/<int:job_id>.json')
@login_required
def task_status(job_id):
    """Poll a single job's status (for the Tasks page)."""
    from models import BackgroundJob
    job = db.session.get(BackgroundJob, job_id)
    if not job:
        abort(404)
    return jsonify(job.as_dict())


@results_bp.route('/api/waec-jamb-correlation/<int:year>')
@login_required
def api_waec_jamb_correlation(year):
    """Get WAEC-JAMB correlation data"""
    correlation = AcademicAnalytics.calculate_waec_jamb_correlation(year, viewing_branch_id())
    return jsonify(correlation)


@results_bp.route('/api/top-performers/<int:year>')
@login_required
def api_top_performers(year):
    """Get top performing students"""
    waec_stats = AcademicAnalytics.get_waec_school_statistics(year, viewing_branch_id())
    
    from utils.branch_scope import scope_by_student
    from utils.access_control import teacher_form_student_ids
    _jq = scope_by_student(JAMBResult.query.filter_by(exam_year=year).options(
        joinedload(JAMBResult.student)), JAMBResult)
    _tids = teacher_form_student_ids()
    if _tids is not None:
        _jq = _jq.filter(JAMBResult.student_id.in_(_tids or [-1]))
    jamb_results = _jq.order_by(JAMBResult.total_score.desc()).limit(10).all()
    
    jamb_top = [{
        'student_id': r.student_id,
        'name': r.student.full_name,
        'score': r.total_score
    } for r in jamb_results]
    
    return jsonify({
        'waec_top': waec_stats['top_performers'] if waec_stats else [],
        'jamb_top': jamb_top
    })


# =============================================================================
# WAEC BROADSHEET — the full grade matrix for an exam year (mirrors mock WAEC),
# viewable, printable (server PDF) and downloadable (PDF / Excel).
# =============================================================================

# Grade -> CSS badge tone, shared with the template.
_WAEC_GRADE_CLASS = {
    'A1': 'a1', 'B2': 'b', 'B3': 'b', 'C4': 'c', 'C5': 'c', 'C6': 'c',
    'D7': 'd', 'E8': 'e', 'F9': 'f',
}


def _waec_broadsheet_years():
    years = [y[0] for y in db.session.query(WAECResult.exam_year).distinct()
             .order_by(WAECResult.exam_year.desc()).all()]
    return years


def _waec_broadsheet_cached(year, branch_id):
    """The broadsheet, memoised in AnalyticsCache under the shared exam_hub
    namespace (so the existing bust/refresh invalidation covers it too)."""
    from routes.results import _cached_school_stats
    return _cached_school_stats(
        'waec_bs', year, branch_id,
        lambda: AcademicAnalytics.get_waec_broadsheet(year, branch_id))


def _broadsheet_etag(bs, year, branch_id, *extra):
    """A strong ETag derived from the broadsheet's actual grade matrix (every
    student's cells), so any grade edit changes it while an unchanged re-request
    is a cheap 304. ``extra`` carries output-shaping args (orientation, columns)."""
    import json
    from utils.http_cache import strong_etag
    fp = json.dumps([[r['student']['id'], r['cells']] for r in bs['rows']],
                    sort_keys=True, separators=(',', ':'))
    return strong_etag('waec_bs', year, branch_id, *extra, fp)


@results_bp.route('/waec/broadsheet')
@login_required
def waec_broadsheet():
    """On-screen WAEC broadsheet: grade matrix + per-subject and cohort summary."""
    years = _waec_broadsheet_years()
    year = resolve_exam_year(request.args.get('year', type=int), years)
    bs = _waec_broadsheet_cached(year, viewing_branch_id()) if year else None
    return render_template('results/waec_broadsheet.html', bs=bs, selected_year=year,
                           years=years, grade_classes=_WAEC_GRADE_CLASS)


@results_bp.route('/waec/broadsheet.pdf')
@login_required
@rate_limited('export', max_requests=40, window_minutes=10)
def waec_broadsheet_pdf():
    """Server-side WAEC broadsheet PDF. Previews inline; ?download=1 to save."""
    from flask import send_file
    from utils.school import school_profile, logo_path
    years = _waec_broadsheet_years()
    year = resolve_exam_year(request.args.get('year', type=int), years)
    bs = _waec_broadsheet_cached(year, viewing_branch_id()) if year else None
    if not bs or not bs['rows']:
        flash('No WAEC results recorded for that year.', 'warning')
        return redirect(url_for('results.waec_broadsheet', year=year))
    per = request.args.get('cols', default=0, type=int)
    orient = request.args.get('orient', 'landscape')
    # Cheap 304 when the browser already holds this exact broadsheet frame.
    from utils.http_cache import if_none_match, stamp
    etag = _broadsheet_etag(bs, year, viewing_branch_id(), 'pdf', per, orient)
    not_modified = if_none_match(etag)
    if not_modified is not None:
        return not_modified
    from utils.waec_broadsheet_pdf import waec_broadsheet_pdf as _mk
    school = dict(school_profile() or {})
    school.setdefault('logo_path', logo_path())
    buf = _mk(bs, year, school, opts={'title': False},
              per=(per if per and per > 0 else 0), orient=orient)
    name = f'waec_broadsheet_{year}.pdf'
    resp = send_file(buf, mimetype='application/pdf',
                     as_attachment=request.args.get('download') == '1', download_name=name)
    return stamp(resp, etag)


@results_bp.route('/waec/broadsheet/export')
@login_required
@rate_limited('export', max_requests=40, window_minutes=10)
def waec_broadsheet_export():
    """Wide WAEC broadsheet workbook: a column per subject (grade), with the
    per-subject offered/passed/failed/average-grade rows beneath."""
    from openpyxl import Workbook
    years = _waec_broadsheet_years()
    year = resolve_exam_year(request.args.get('year', type=int), years)
    bs = _waec_broadsheet_cached(year, viewing_branch_id()) if year else None
    if not bs or not bs['rows']:
        flash('No WAEC results recorded for that year.', 'warning')
        return redirect(url_for('results.waec_broadsheet', year=year))
    from utils.http_cache import if_none_match
    etag = _broadsheet_etag(bs, year, viewing_branch_id(), 'xlsx')
    not_modified = if_none_match(etag)
    if not_modified is not None:
        return not_modified
    subjects = bs['subjects']

    wb = Workbook()
    ws = wb.active
    ws.title = f'WAEC {year}'
    ws.append(['S/N', 'Student'] + subjects + ['Credits', 'Avg grade'])
    for i, row in enumerate(bs['rows'], 1):
        line = [i, row['student']['full_name']]
        line += [row['cells'].get(subj, '') for subj in subjects]
        line += [row['credits'], row['avg_grade']]
        ws.append(line)

    ws.append([])
    ss = bs['subject_summary']
    def _summary_row(label, fn):
        ws.append(['', label] + [fn(ss[s]) for s in subjects] + ['', ''])
    _summary_row('No. offered', lambda d: d['offered'])
    _summary_row('No. passed (C6+)', lambda d: d['passed'])
    _summary_row('No. failed', lambda d: d['failed'])
    _summary_row('Average grade', lambda d: d['avg_grade'])

    from utils.http_cache import stamp
    return stamp(xlsx_response(wb, f'waec_broadsheet_{year}.xlsx'), etag)
