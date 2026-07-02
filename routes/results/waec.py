"""results blueprint — waec routes (split from the former routes/results.py)."""
from routes.results import *  # noqa: F401,F403
from utils.search import like_term


@results_bp.route('/waec')
@login_required
def waec_list():
    """Comprehensive WAEC dashboard with filtering and analytics"""
    exam_year = request.args.get('year', type=int)
    min_a1 = request.args.get('min_a1', type=int)
    min_credits = request.args.get('min_credits', type=int)
    subject_filter = request.args.get('subject', '')
    grade_filter = request.args.get('grade', '')
    search = request.args.get('search', '').strip()
    sort_by = request.args.get('sort', 'name')
    sort_order = request.args.get('order', 'asc')
    view_mode = request.args.get('view', 'students')
    
    years = db.session.query(WAECResult.exam_year).distinct().order_by(WAECResult.exam_year.desc()).all()
    years = [y[0] for y in years]
    
    if not exam_year and years:
        exam_year = years[0]
    
    students_data = []
    subject_stats = []
    school_stats = None
    grade_distribution = {g: 0 for g in WAEC_GRADES}
    top_by_grade = {g: [] for g in WAEC_GRADES}
    subjects_by_grade = []
    
    if exam_year:
        from utils.branch_scope import viewing_branch_id
        school_stats = AcademicAnalytics.get_waec_school_statistics(exam_year, viewing_branch_id())

        if school_stats:
            grade_distribution = school_stats['grade_distribution']
            subject_stats = school_stats['subject_analysis']
        
        # Get all results for the year for detailed analysis (branch-scoped)
        from utils.branch_scope import scope_by_student
        all_results = scope_by_student(
            WAECResult.query.filter_by(exam_year=exam_year), WAECResult).all()
        
        # Build subject-grade matrix
        subject_grade_counts = defaultdict(lambda: {g: 0 for g in WAEC_GRADES})
        for r in all_results:
            subject_grade_counts[r.subject][r.grade] += 1
        
        # Convert to list for template
        subjects_by_grade = []
        for subject, grades in subject_grade_counts.items():
            total = sum(grades.values())
            subjects_by_grade.append({
                'subject': subject,
                'grades': grades,
                'total': total,
                'a1_count': grades.get('A1', 0),
                'b2_count': grades.get('B2', 0),
                'b3_count': grades.get('B3', 0),
                'credit_count': sum(grades.get(g, 0) for g in ['A1', 'B2', 'B3', 'C4', 'C5', 'C6']),
                'pass_rate': round(sum(grades.get(g, 0) for g in ['A1', 'B2', 'B3', 'C4', 'C5', 'C6']) / total * 100, 1) if total > 0 else 0
            })
        
        # Sort subjects based on parameters
        if sort_by == 'subject_a1':
            subjects_by_grade.sort(key=lambda x: x['a1_count'], reverse=(sort_order == 'desc'))
        elif sort_by == 'subject_b2':
            subjects_by_grade.sort(key=lambda x: x['b2_count'], reverse=(sort_order == 'desc'))
        elif sort_by == 'subject_credits':
            subjects_by_grade.sort(key=lambda x: x['credit_count'], reverse=(sort_order == 'desc'))
        elif sort_by == 'subject_pass':
            subjects_by_grade.sort(key=lambda x: x['pass_rate'], reverse=(sort_order == 'desc'))
        else:
            subjects_by_grade.sort(key=lambda x: x['a1_count'], reverse=True)
        
        from utils.branch_scope import scope_query
        base_query = scope_query(
            db.session.query(Student).join(WAECResult).filter(
                WAECResult.exam_year == exam_year),
            Student)

        if search:
            term = like_term(search)
            base_query = base_query.filter(
                db.or_(
                    Student.first_name.ilike(term, escape='\\'),
                    Student.surname.ilike(term, escape='\\'),
                    Student.middle_name.ilike(term, escape='\\'),
                    Student.student_id.ilike(term, escape='\\'),
                )
            )

        if subject_filter and grade_filter:
            base_query = base_query.filter(
                WAECResult.subject == subject_filter,
                WAECResult.grade == grade_filter
            )
        elif subject_filter:
            base_query = base_query.filter(WAECResult.subject == subject_filter)
        elif grade_filter:
            base_query = base_query.filter(WAECResult.grade == grade_filter)
        
        students = base_query.distinct().all()

        # Batch-load every matching student's results for the year in one query
        # (instead of one query per student).
        results_by_student = {}
        if students:
            for r in WAECResult.query.filter(
                    WAECResult.student_id.in_([s.id for s in students]),
                    WAECResult.exam_year == exam_year).all():
                results_by_student.setdefault(r.student_id, []).append(r)

        for student in students:
            results = results_by_student.get(student.id, [])

            # Count each grade
            grade_counts = {g: sum(1 for r in results if r.grade == g) for g in WAEC_GRADES}
            a1_count = grade_counts['A1']
            b2_count = grade_counts['B2']
            b3_count = grade_counts['B3']
            credit_count = sum(grade_counts[g] for g in ['A1', 'B2', 'B3', 'C4', 'C5', 'C6'])
            total_points = sum(WAECResult.grade_to_points(r.grade) for r in results)
            
            if min_a1 and a1_count < min_a1:
                continue
            if min_credits and credit_count < min_credits:
                continue
            
            student_data = {
                'id': student.id,
                'student_id': student.student_id,
                'name': student.full_name,
                'total_subjects': len(results),
                'a1_count': a1_count,
                'b2_count': b2_count,
                'b3_count': b3_count,
                'grade_counts': grade_counts,
                'credit_count': credit_count,
                'total_points': total_points,
                'avg_points': round(total_points / len(results), 2) if results else 0,
                'results': [{'subject': r.subject, 'grade': r.grade} for r in results]
            }
            students_data.append(student_data)
        
        # Build top performers by each grade
        for grade in WAEC_GRADES:
            grade_key = f'{grade.lower()}_count'
            sorted_students = sorted(
                [s for s in students_data if s['grade_counts'].get(grade, 0) > 0],
                key=lambda x: x['grade_counts'].get(grade, 0),
                reverse=True
            )[:10]
            top_by_grade[grade] = sorted_students
        
        # Sort students based on parameters
        if sort_by == 'a1':
            students_data.sort(key=lambda x: x['a1_count'], reverse=(sort_order == 'desc'))
        elif sort_by == 'b2':
            students_data.sort(key=lambda x: x['b2_count'], reverse=(sort_order == 'desc'))
        elif sort_by == 'b3':
            students_data.sort(key=lambda x: x['b3_count'], reverse=(sort_order == 'desc'))
        elif sort_by == 'credits':
            students_data.sort(key=lambda x: x['credit_count'], reverse=(sort_order == 'desc'))
        elif sort_by == 'points':
            students_data.sort(key=lambda x: x['avg_points'], reverse=(sort_order != 'desc'))
        else:
            students_data.sort(key=lambda x: x['name'], reverse=(sort_order == 'desc'))
    
    from utils.branch_scope import viewing_branch_id
    yoy_data = AcademicAnalytics.get_year_over_year_comparison(viewing_branch_id())
    
    return render_template('results/waec_dashboard.html',
        students=students_data,
        years=years,
        selected_year=exam_year,
        subjects=WAEC_SUBJECTS,
        grades=WAEC_GRADES,
        subject_filter=subject_filter,
        grade_filter=grade_filter,
        search=search,
        min_a1=min_a1,
        min_credits=min_credits,
        sort_by=sort_by,
        sort_order=sort_order,
        view_mode=view_mode,
        school_stats=school_stats,
        subject_stats=subject_stats,
        grade_distribution=grade_distribution,
        top_by_grade=top_by_grade,
        subjects_by_grade=subjects_by_grade,
        yoy_data=yoy_data
    )


@results_bp.route('/waec/add', methods=['GET', 'POST'])
@login_required
def add_waec():
    """Add WAEC results for a student"""
    students = get_sss3_students()

    if request.method == 'POST':
        try:
            student_id = request.form.get('student_id', type=int)
            exam_year = request.form.get('exam_year', type=int)
            
            if not student_id or not exam_year:
                flash('Student and exam year are required.', 'error')
                return redirect(url_for('results.add_waec'))
            
            subjects = request.form.getlist('subject[]')
            grades = request.form.getlist('grade[]')
            
            results_added = 0
            for i, subject in enumerate(subjects):
                if subject and i < len(grades) and grades[i]:
                    existing = WAECResult.query.filter_by(
                        student_id=student_id,
                        exam_year=exam_year,
                        subject=subject
                    ).first()
                    
                    if existing:
                        existing.grade = grades[i]
                    else:
                        result = WAECResult(
                            student_id=student_id,
                            exam_year=exam_year,
                            subject=subject,
                            grade=grades[i]
                        )
                        db.session.add(result)
                    
                    results_added += 1
            
            db.session.commit()
            recompute_student_safe(student_id)   # refresh persisted analytics
            flash(f'{results_added} WAEC results saved!', 'success')
            return redirect(url_for('results.view_waec_student', student_id=student_id))
            
        except Exception as e:
            db.session.rollback()
            flash(f'Error saving results: {str(e)}', 'error')
    
    return render_template('results/add_waec.html',
        students=students,
        subjects=WAEC_SUBJECTS,
        grades=WAEC_GRADES,
        default_subjects=WAEC_DEFAULT_SUBJECTS,
        stream_defaults=STREAM_WAEC_SUBJECTS,
        subject_map=student_subject_map(students),
        current_year=_date.today().year
    )


@results_bp.route('/waec/scan', methods=['GET', 'POST'])
@login_required
@rate_limited('ocr', max_requests=30, window_minutes=10)
def scan_waec():
    """Upload a WAEC result image, OCR it, and review before saving."""
    from utils.waec_ocr import (
        tesseract_available, extract_text, parse_waec_result, match_student,
        pdf_available, extract_text_from_pdf, pdf_first_page_png,
        vision_available, vision_extract,
    )
    import base64

    students = get_sss3_students()

    if request.method == 'POST':
        file = request.files.get('result_image')
        if not file or not file.filename:
            flash('Please choose a file to upload.', 'error')
            return redirect(url_for('results.scan_waec'))

        filename = (file.filename or '').lower()
        is_pdf = (file.mimetype == 'application/pdf') or filename.endswith('.pdf')
        file_bytes = file.read()

        try:
            if is_pdf:
                if not pdf_available():
                    flash('PDF support (PyMuPDF) is not installed on the server.', 'error')
                    return redirect(url_for('results.scan_waec'))
                text = extract_text_from_pdf(file_bytes)
                try:
                    preview_bytes = pdf_first_page_png(file_bytes)
                    preview = 'data:image/png;base64,' + base64.b64encode(preview_bytes).decode()
                except Exception:
                    preview = ''
            else:
                if not tesseract_available():
                    flash('OCR engine (Tesseract) is not installed on the server. '
                          'Install "tesseract-ocr" to scan images.', 'error')
                    return redirect(url_for('results.scan_waec'))
                text = extract_text(file_bytes)
                mime = file.mimetype or 'image/png'
                preview = f"data:{mime};base64,{base64.b64encode(file_bytes).decode()}"
        except Exception as e:
            flash(f'Could not read the file: {e}', 'error')
            return redirect(url_for('results.scan_waec'))

        # Prefer the Claude-vision reading when enabled; fall back to Tesseract.
        parsed = None
        if not is_pdf and vision_available():
            parsed = vision_extract(file_bytes, 'waec', file.mimetype or 'image/png')
            if parsed:
                text = '(read by Claude vision)\n' + text
        if not parsed:
            parsed = parse_waec_result(text)
        matched, score = match_student(parsed['name'], students)

        return render_template('results/waec_scan_review.html',
            students=students,
            subjects=WAEC_SUBJECTS,
            grades=WAEC_GRADES,
            parsed=parsed,
            matched=matched,
            match_score=score,
            preview=preview,
            current_year=parsed.get('year') or _date.today().year,
            raw_text=text
        )

    return render_template('results/waec_scan.html',
        ocr_ready=tesseract_available(),
        pdf_ready=pdf_available()
    )


@results_bp.route('/waec/paste', methods=['GET', 'POST'])
@login_required
def paste_waec():
    """Enter a WAEC result by pasting it as text (e.g. a result screenshot read by
    an external AI) — no OCR, no API key. Reuses the scan review + save flow."""
    from utils.waec_ocr import parse_waec_result, match_student
    students = get_sss3_students()
    if request.method == 'POST':
        text = (request.form.get('data') or '').strip()
        if not text:
            flash('Paste the result text first.', 'error')
            return render_template('results/waec_paste.html', subjects=WAEC_SUBJECTS, current_year=_date.today().year)
        parsed = parse_waec_result(text)
        if not parsed.get('subjects'):
            flash('No subjects/grades could be read from the pasted text.', 'warning')
            return render_template('results/waec_paste.html', subjects=WAEC_SUBJECTS, pasted=text, current_year=_date.today().year)
        year = request.form.get('exam_year', type=int)
        matched, score = match_student(parsed.get('name'), students)
        return render_template('results/waec_scan_review.html',
            students=students, subjects=WAEC_SUBJECTS, grades=WAEC_GRADES,
            parsed=parsed, matched=matched, match_score=score, preview='',
            current_year=year or parsed.get('year') or _date.today().year, raw_text=text)
    return render_template('results/waec_paste.html', subjects=WAEC_SUBJECTS, current_year=_date.today().year)


@results_bp.route('/waec/student/<int:student_id>')
@login_required
def view_waec_student(student_id):
    """View comprehensive WAEC profile for a student"""
    student = db.get_or_404(Student, student_id)
    require_branch_access(student.branch_id)
    waec_summary = AcademicAnalytics.get_student_waec_summary(student_id)
    risk_assessment = AcademicAnalytics.calculate_student_risk_score(student_id)
    jamb_prediction = AcademicAnalytics.predict_jamb_score(student_id)
    recommendations = AcademicAnalytics.get_subject_recommendations(student_id)
    
    return render_template('results/view_waec_student.html',
        student=student,
        waec_summary=waec_summary,
        risk_assessment=risk_assessment,
        jamb_prediction=jamb_prediction,
        recommendations=recommendations,
        grades=WAEC_GRADES
    )


@results_bp.route('/waec/student/<int:student_id>/edit/<int:year>', methods=['GET', 'POST'])
@login_required
def edit_waec(student_id, year):
    """Edit WAEC results for a student/year"""
    student = db.get_or_404(Student, student_id)
    require_branch_access(student.branch_id)
    
    if request.method == 'POST':
        try:
            WAECResult.query.filter_by(student_id=student_id, exam_year=year).delete()
            
            subjects = request.form.getlist('subject[]')
            grades = request.form.getlist('grade[]')
            
            for i, subject in enumerate(subjects):
                if subject and i < len(grades) and grades[i]:
                    result = WAECResult(
                        student_id=student_id,
                        exam_year=year,
                        subject=subject,
                        grade=grades[i]
                    )
                    db.session.add(result)
            
            db.session.commit()
            recompute_student_safe(student_id)
            log_action('results.waec_edit', detail=f'{year}', target=student)
            flash('WAEC results updated!', 'success')
            return redirect(url_for('results.view_waec_student', student_id=student_id))
            
        except Exception as e:
            db.session.rollback()
            flash(f'Error: {str(e)}', 'error')
    
    results = WAECResult.query.filter_by(student_id=student_id, exam_year=year).all()
    results_dict = {r.subject: r.grade for r in results}
    
    return render_template('results/edit_waec.html',
        student=student,
        year=year,
        results_dict=results_dict,
        subjects=WAEC_SUBJECTS,
        grades=WAEC_GRADES
    )


@results_bp.route('/waec/student/<int:student_id>/delete/<int:year>', methods=['POST'])
@admin_required
def delete_waec(student_id, year):
    """Delete all WAEC results for a student in a given year"""
    student = db.get_or_404(Student, student_id)
    require_branch_access(student.branch_id)
    
    try:
        deleted = WAECResult.query.filter_by(student_id=student_id, exam_year=year).delete()
        db.session.commit()
        recompute_student_safe(student_id)

        if deleted:
            log_action('results.waec_delete', detail=f'{deleted} result(s), {year}', target=student)
            flash(f'Deleted {deleted} WAEC results for {student.full_name} ({year}).', 'success')
        else:
            flash('No results found to delete.', 'warning')
            
    except Exception as e:
        db.session.rollback()
        flash(f'Error deleting results: {str(e)}', 'error')
    
    return redirect(url_for('results.waec_list', year=year))


@results_bp.route('/waec/result/<int:result_id>/delete', methods=['POST'])
@admin_required
def delete_waec_single(result_id):
    """Delete a single WAEC result entry"""
    result = db.get_or_404(WAECResult, result_id)
    student_id = result.student_id
    year = result.exam_year
    subject = result.subject
    
    try:
        db.session.delete(result)
        db.session.commit()
        recompute_student_safe(student_id)
        flash(f'Deleted {subject} result.', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Error: {str(e)}', 'error')
    
    return redirect(url_for('results.view_waec_student', student_id=student_id))


@results_bp.route('/api/waec/grade-distribution/<int:year>')
@login_required
def api_waec_grade_distribution(year):
    """Get WAEC grade distribution for charts"""
    results = scope_by_student(WAECResult.query.filter_by(exam_year=year), WAECResult).all()
    distribution = {g: sum(1 for r in results if r.grade == g) for g in WAEC_GRADES}
    return jsonify(distribution)


@results_bp.route('/api/waec/subject-stats/<int:year>')
@login_required
def api_waec_subject_stats(year):
    """Get subject-level statistics"""
    stats = AcademicAnalytics.get_waec_school_statistics(year)
    if stats:
        return jsonify(stats['subject_analysis'])
    return jsonify([])


@results_bp.route('/waec/export')
@login_required
@rate_limited('export', max_requests=40, window_minutes=10)
def export_waec():
    """Export WAEC results to Excel"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from io import BytesIO
    
    year = request.args.get('year', type=int)
    if not year:
        flash('Please select a year to export.', 'error')
        return redirect(url_for('results.waec_list'))
    log_action('data.export_results', detail=f'WAEC {year}')

    from utils.branch_scope import scope_query
    students_with_results = scope_query(
        db.session.query(Student).join(WAECResult).filter(WAECResult.exam_year == year),
        Student).distinct().order_by(Student.surname).all()
    
    wb = Workbook()
    ws = wb.active
    ws.title = f"WAEC {year}"
    
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF')
    border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    
    ws.merge_cells('A1:L1')
    ws['A1'] = f'WAEC RESULTS - {year}'
    ws['A1'].font = Font(bold=True, size=14)
    ws['A1'].alignment = Alignment(horizontal='center')
    
    headers = ['S/N', 'Student ID', 'Name'] + WAEC_SUBJECTS[:9] + ['Total Points', 'Credits', 'A1s']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
    
    for idx, student in enumerate(students_with_results, 1):
        row = idx + 3
        results = {r.subject: r.grade for r in student.waec_results.filter_by(exam_year=year).all()}
        
        ws.cell(row=row, column=1, value=idx).border = border
        ws.cell(row=row, column=2, value=student.student_id).border = border
        ws.cell(row=row, column=3, value=student.full_name).border = border
        
        total_points = 0
        credit_count = 0
        a1_count = 0
        
        for col, subject in enumerate(WAEC_SUBJECTS[:9], 4):
            grade = results.get(subject, '-')
            cell = ws.cell(row=row, column=col, value=grade)
            cell.border = border
            
            if grade != '-':
                points = WAECResult.grade_to_points(grade)
                total_points += points
                if grade in ['A1', 'B2', 'B3', 'C4', 'C5', 'C6']:
                    credit_count += 1
                if grade == 'A1':
                    a1_count += 1
        
        ws.cell(row=row, column=13, value=total_points).border = border
        ws.cell(row=row, column=14, value=credit_count).border = border
        ws.cell(row=row, column=15, value=a1_count).border = border
    
    ws.column_dimensions['A'].width = 5
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 25

    ws.delete_cols(2)        # drop the admission-number column from the printout
    return xlsx_response(wb, f'waec_results_{year}.xlsx')


@results_bp.route('/waec/analytics')
@login_required
def waec_analytics():
    """Enhanced WAEC analytics dashboard"""
    from utils.exam_analytics import WAECAnalytics
    
    year = request.args.get('year', type=int)
    
    years = db.session.query(WAECResult.exam_year).distinct().order_by(WAECResult.exam_year.desc()).all()
    years = [y[0] for y in years]
    
    if not year and years:
        year = years[0]
    
    stats = None
    year_comparison = None
    
    if year:
        stats = WAECAnalytics.get_year_statistics(year)
        year_comparison = WAECAnalytics.compare_years()
    
    return render_template('results/waec_analytics.html',
        years=years,
        selected_year=year,
        stats=stats,
        year_comparison=year_comparison
    )


@results_bp.route('/waec/student/<int:student_id>')
@login_required
def waec_student_analysis(student_id):
    """Detailed WAEC analysis for a specific student"""
    from utils.exam_analytics import WAECAnalytics, WAECJAMBCorrelation
    
    student = db.get_or_404(Student, student_id)
    require_branch_access(student.branch_id)
    year = request.args.get('year', type=int)
    
    waec_analysis = WAECAnalytics.get_student_waec_analysis(student_id, year)
    jamb_prediction = WAECJAMBCorrelation.predict_jamb_from_waec(student_id, year)
    
    return render_template('results/waec_student.html',
        student=student,
        waec_analysis=waec_analysis,
        jamb_prediction=jamb_prediction
    )
