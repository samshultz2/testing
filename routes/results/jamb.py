"""results blueprint — jamb routes (split from the former routes/results.py)."""
from routes.results import *  # noqa: F401,F403
from utils.search import like_term


@results_bp.route('/jamb')
@login_required
def jamb_list():
    """Comprehensive JAMB dashboard with filtering and analytics"""
    exam_year = request.args.get('year', type=int)
    min_score = request.args.get('min_score', type=int)
    max_score = request.args.get('max_score', type=int)
    search = request.args.get('search', '').strip()
    sort_by = request.args.get('sort', 'score')
    sort_order = request.args.get('order', 'desc')
    view_mode = request.args.get('view', 'students')
    
    years = db.session.query(JAMBResult.exam_year).distinct().order_by(JAMBResult.exam_year.desc()).all()
    years = [y[0] for y in years]
    
    if not exam_year and years:
        exam_year = years[0]
    
    students_data = []
    school_stats = None
    correlation = None
    subject_performance = []
    
    if exam_year:
        from utils.branch_scope import viewing_branch_id
        _bid = viewing_branch_id()
        school_stats = AcademicAnalytics.get_jamb_school_statistics(exam_year, _bid)
        correlation = AcademicAnalytics.calculate_waec_jamb_correlation(exam_year, _bid)
        
        # Get all results for subject analysis
        all_results = scope_by_student(JAMBResult.query.filter_by(exam_year=exam_year), JAMBResult).all()
        
        # Build subject performance data
        subject_scores = defaultdict(list)
        for r in all_results:
            if r.subject1 and r.subject1_score:
                subject_scores[r.subject1].append(r.subject1_score)
            if r.subject2 and r.subject2_score:
                subject_scores[r.subject2].append(r.subject2_score)
            if r.subject3 and r.subject3_score:
                subject_scores[r.subject3].append(r.subject3_score)
            if r.subject4 and r.subject4_score:
                subject_scores[r.subject4].append(r.subject4_score)
        
        # Calculate subject statistics
        for subject, scores in subject_scores.items():
            if scores:
                avg_score = sum(scores) / len(scores)
                above_50 = sum(1 for s in scores if s >= 50)
                above_70 = sum(1 for s in scores if s >= 70)
                subject_performance.append({
                    'subject': subject,
                    'count': len(scores),
                    'avg_score': round(avg_score, 1),
                    'max_score': max(scores),
                    'min_score': min(scores),
                    'above_50': above_50,
                    'above_50_pct': round(above_50 / len(scores) * 100, 1),
                    'above_70': above_70,
                    'above_70_pct': round(above_70 / len(scores) * 100, 1)
                })
        
        # Sort subjects by performance
        if sort_by == 'subject_avg':
            subject_performance.sort(key=lambda x: x['avg_score'], reverse=(sort_order == 'desc'))
        elif sort_by == 'subject_above50':
            subject_performance.sort(key=lambda x: x['above_50_pct'], reverse=(sort_order == 'desc'))
        elif sort_by == 'subject_above70':
            subject_performance.sort(key=lambda x: x['above_70_pct'], reverse=(sort_order == 'desc'))
        else:
            subject_performance.sort(key=lambda x: x['avg_score'], reverse=True)
        
        # Branch-scope the list exactly like the analytics query above (and the
        # WAEC list) so a branch user never sees other branches' JAMB results.
        query = (scope_by_student(JAMBResult.query.filter_by(exam_year=exam_year), JAMBResult)
                 .options(joinedload(JAMBResult.student)).join(Student))

        if search:
            term = like_term(search)
            query = query.filter(
                db.or_(
                    Student.first_name.ilike(term, escape='\\'),
                    Student.surname.ilike(term, escape='\\'),
                    Student.middle_name.ilike(term, escape='\\'),
                    Student.student_id.ilike(term, escape='\\'),
                )
            )

        if min_score:
            query = query.filter(JAMBResult.total_score >= min_score)
        if max_score:
            query = query.filter(JAMBResult.total_score <= max_score)
        
        if sort_by == 'score':
            if sort_order == 'desc':
                query = query.order_by(JAMBResult.total_score.desc())
            else:
                query = query.order_by(JAMBResult.total_score.asc())
        elif sort_by == 'name':
            if sort_order == 'desc':
                query = query.order_by(Student.surname.desc())
            else:
                query = query.order_by(Student.surname.asc())
        else:
            query = query.order_by(JAMBResult.total_score.desc())
        
        results = query.all()
        
        for idx, r in enumerate(results):
            subjects = []
            if r.subject1: subjects.append({'name': r.subject1, 'score': r.subject1_score or 0})
            if r.subject2: subjects.append({'name': r.subject2, 'score': r.subject2_score or 0})
            if r.subject3: subjects.append({'name': r.subject3, 'score': r.subject3_score or 0})
            if r.subject4: subjects.append({'name': r.subject4, 'score': r.subject4_score or 0})
            
            students_data.append({
                'id': r.student.id,
                'jamb_id': r.id,
                'student_id': r.student.student_id,
                'name': r.student.full_name,
                'total_score': r.total_score,
                'rank': idx + 1,
                'subjects': subjects,
                'performance_level': AcademicAnalytics._jamb_performance_level(r.total_score)
            })
    
    from utils.branch_scope import viewing_branch_id
    yoy_data = AcademicAnalytics.get_year_over_year_comparison(viewing_branch_id())
    
    return render_template('results/jamb_dashboard.html',
        students=students_data,
        years=years,
        selected_year=exam_year,
        min_score=min_score,
        max_score=max_score,
        search=search,
        sort_by=sort_by,
        sort_order=sort_order,
        view_mode=view_mode,
        school_stats=school_stats,
        correlation=correlation,
        subject_performance=subject_performance,
        yoy_data=yoy_data
    )


@results_bp.route('/jamb/add', methods=['GET', 'POST'])
@login_required
def add_jamb():
    """Add JAMB results for a student"""
    students = get_sss3_students()
    
    if request.method == 'POST':
        try:
            student_id = request.form.get('student_id', type=int)
            exam_year = request.form.get('exam_year', type=int)
            total_score = request.form.get('total_score', type=int)
            
            if not student_id or not exam_year or total_score is None:
                flash('All required fields must be filled.', 'error')
                return redirect(url_for('results.add_jamb'))
            
            if total_score < 0 or total_score > 400:
                flash('Total score must be between 0 and 400.', 'error')
                return redirect(url_for('results.add_jamb'))
            
            existing = JAMBResult.query.filter_by(student_id=student_id, exam_year=exam_year).first()
            
            if existing:
                existing.total_score = total_score
                existing.subject1 = request.form.get('subject1')
                existing.subject1_score = request.form.get('subject1_score', type=int)
                existing.subject2 = request.form.get('subject2')
                existing.subject2_score = request.form.get('subject2_score', type=int)
                existing.subject3 = request.form.get('subject3')
                existing.subject3_score = request.form.get('subject3_score', type=int)
                existing.subject4 = request.form.get('subject4')
                existing.subject4_score = request.form.get('subject4_score', type=int)
            else:
                result = JAMBResult(
                    student_id=student_id,
                    exam_year=exam_year,
                    total_score=total_score,
                    subject1=request.form.get('subject1'),
                    subject1_score=request.form.get('subject1_score', type=int),
                    subject2=request.form.get('subject2'),
                    subject2_score=request.form.get('subject2_score', type=int),
                    subject3=request.form.get('subject3'),
                    subject3_score=request.form.get('subject3_score', type=int),
                    subject4=request.form.get('subject4'),
                    subject4_score=request.form.get('subject4_score', type=int)
                )
                db.session.add(result)
            
            db.session.commit()
            recompute_student_safe(student_id)
            flash('JAMB result saved!', 'success')
            return redirect(url_for('results.jamb_list'))
            
        except Exception as e:
            db.session.rollback()
            flash(f'Error: {str(e)}', 'error')
    
    return render_template('results/add_jamb.html',
        students=students,
        subjects=WAEC_SUBJECTS,
        subject_map=student_subject_map(students),
        current_year=_date.today().year
    )


@results_bp.route('/jamb/scan', methods=['GET', 'POST'])
@login_required
@rate_limited('ocr', max_requests=30, window_minutes=10)
def scan_jamb():
    """Upload a JAMB result image/PDF, OCR it, and review before saving."""
    from utils.waec_ocr import (
        tesseract_available, extract_text, parse_jamb_result, match_student,
        pdf_available, extract_text_from_pdf, pdf_first_page_png,
        vision_available, vision_extract,
    )
    import base64

    students = get_sss3_students()

    if request.method == 'POST':
        file = request.files.get('result_image')
        if not file or not file.filename:
            flash('Please choose a file to upload.', 'error')
            return redirect(url_for('results.scan_jamb'))

        filename = (file.filename or '').lower()
        is_pdf = (file.mimetype == 'application/pdf') or filename.endswith('.pdf')
        file_bytes = file.read()

        try:
            if is_pdf:
                if not pdf_available():
                    flash('PDF support (PyMuPDF) is not installed on the server.', 'error')
                    return redirect(url_for('results.scan_jamb'))
                text = extract_text_from_pdf(file_bytes)
                try:
                    preview = 'data:image/png;base64,' + base64.b64encode(pdf_first_page_png(file_bytes)).decode()
                except Exception:
                    preview = ''
            else:
                if not tesseract_available():
                    flash('OCR engine (Tesseract) is not installed on the server. '
                          'Install "tesseract-ocr" to scan images.', 'error')
                    return redirect(url_for('results.scan_jamb'))
                text = extract_text(file_bytes)
                mime = file.mimetype or 'image/png'
                preview = f"data:{mime};base64,{base64.b64encode(file_bytes).decode()}"
        except Exception as e:
            flash(f'Could not read the file: {e}', 'error')
            return redirect(url_for('results.scan_jamb'))

        parsed = None
        if not is_pdf and vision_available():
            parsed = vision_extract(file_bytes, 'jamb', file.mimetype or 'image/png')
            if parsed:
                text = '(read by Claude vision)\n' + text
        if not parsed:
            parsed = parse_jamb_result(text)
        matched, score = match_student(parsed['name'], students)

        return render_template('results/jamb_scan_review.html',
            students=students,
            subjects=WAEC_SUBJECTS,
            parsed=parsed,
            matched=matched,
            match_score=score,
            preview=preview,
            current_year=parsed.get('year') or _date.today().year,
            raw_text=text
        )

    return render_template('results/jamb_scan.html',
        ocr_ready=tesseract_available(),
        pdf_ready=pdf_available()
    )


@results_bp.route('/jamb/paste', methods=['GET', 'POST'])
@login_required
def paste_jamb():
    """Enter a JAMB result by pasting it as text (e.g. read by an external AI) —
    no OCR. Reuses the JAMB scan review + save flow."""
    from utils.waec_ocr import parse_jamb_result, match_student
    students = get_sss3_students()
    if request.method == 'POST':
        text = (request.form.get('data') or '').strip()
        if not text:
            flash('Paste the result text first.', 'error')
            return render_template('results/jamb_paste.html', subjects=WAEC_SUBJECTS, current_year=_date.today().year)
        parsed = parse_jamb_result(text)
        if parsed.get('total_score') is None and not parsed.get('subjects'):
            flash('No JAMB score/subjects could be read from the pasted text.', 'warning')
            return render_template('results/jamb_paste.html', subjects=WAEC_SUBJECTS, pasted=text, current_year=_date.today().year)
        year = request.form.get('exam_year', type=int)
        matched, score = match_student(parsed.get('name'), students)
        return render_template('results/jamb_scan_review.html',
            students=students, subjects=WAEC_SUBJECTS, parsed=parsed, matched=matched,
            match_score=score, preview='',
            current_year=year or parsed.get('year') or _date.today().year, raw_text=text)
    return render_template('results/jamb_paste.html', subjects=WAEC_SUBJECTS, current_year=_date.today().year)


@results_bp.route('/scan/batch', methods=['GET', 'POST'])
@login_required
@rate_limited('ocr', max_requests=30, window_minutes=10)
def scan_batch():
    """Batch-scan several WAEC or JAMB results, each routed to its student."""
    from utils.waec_ocr import parse_waec_result, parse_jamb_result, match_student
    import json

    students = get_sss3_students()
    exam = request.values.get('exam', 'waec')
    exam = 'jamb' if exam == 'jamb' else 'waec'

    if request.method == 'POST' and request.form.get('action') == 'save':
        count = int(request.form.get('count', 0))
        saved = 0
        touched = set()
        for i in range(count):
            if not request.form.get(f'include_{i}'):
                continue
            student_id = request.form.get(f'student_id_{i}', type=int)
            year = request.form.get(f'year_{i}', type=int)
            if not student_id or not year:
                continue
            touched.add(student_id)
            data = json.loads(request.form.get(f'data_{i}', '{}'))
            if exam == 'waec':
                for row in data.get('subjects', []):
                    if not row.get('subject') or not row.get('grade'):
                        continue
                    existing = WAECResult.query.filter_by(student_id=student_id, exam_year=year, subject=row['subject']).first()
                    if existing:
                        existing.grade = row['grade']
                    else:
                        db.session.add(WAECResult(student_id=student_id, exam_year=year, subject=row['subject'], grade=row['grade']))
                saved += 1
            else:
                subs = data.get('subjects', [])[:4]
                total = data.get('total_score') or sum(r.get('score', 0) for r in subs)
                existing = JAMBResult.query.filter_by(student_id=student_id, exam_year=year).first()
                target = existing or JAMBResult(student_id=student_id, exam_year=year, total_score=total)
                target.total_score = total
                for n, row in enumerate(subs, 1):
                    setattr(target, f'subject{n}', row.get('subject'))
                    setattr(target, f'subject{n}_score', row.get('score'))
                if not existing:
                    db.session.add(target)
                saved += 1
        db.session.commit()
        for sid in touched:
            recompute_student_safe(sid)
        flash(f'Saved {saved} {exam.upper()} result(s) from batch scan.', 'success')
        return redirect(url_for('results.waec_list' if exam == 'waec' else 'results.jamb_list'))

    if request.method == 'POST':
        files = request.files.getlist('result_images')
        items = []
        for f in files:
            if not f or not f.filename:
                continue
            try:
                text = _read_uploaded_text(f)
            except Exception:
                current_app.logger.exception('OCR read failed for upload %s', f.filename)
                text = None
            if text is None:
                items.append({'filename': f.filename, 'error': 'Unreadable / engine missing', 'data': {}})
                continue
            parsed = parse_jamb_result(text) if exam == 'jamb' else parse_waec_result(text)
            matched, score = match_student(parsed.get('name'), students)
            items.append({
                'filename': f.filename,
                'parsed': parsed,
                'matched_id': matched.id if matched else None,
                'matched_name': matched.full_name if matched else None,
                'score': score,
                'data_json': json.dumps(parsed),
            })
        return render_template('results/scan_batch_review.html',
            exam=exam, items=items, students=students, current_year=_date.today().year)

    return render_template('results/scan_batch.html', exam=exam)


@results_bp.route('/jamb/student/<int:student_id>')
@login_required
def view_jamb_student(student_id):
    """View JAMB results for a student"""
    student = db.get_or_404(Student, student_id)
    require_branch_access(student.branch_id)
    jamb_summary = AcademicAnalytics.get_student_jamb_summary(student_id)
    waec_summary = AcademicAnalytics.get_student_waec_summary(student_id)
    
    return render_template('results/view_jamb_student.html',
        student=student,
        jamb_summary=jamb_summary,
        waec_summary=waec_summary
    )


@results_bp.route('/jamb/student/<int:student_id>/edit/<int:year>', methods=['GET', 'POST'])
@login_required
def edit_jamb(student_id, year):
    """Edit JAMB results for a student/year"""
    student = db.get_or_404(Student, student_id)
    require_branch_access(student.branch_id)
    result = JAMBResult.query.filter_by(student_id=student_id, exam_year=year).first_or_404()
    
    if request.method == 'POST':
        try:
            result.total_score = request.form.get('total_score', type=int)
            result.subject1 = request.form.get('subject1')
            result.subject1_score = request.form.get('subject1_score', type=int)
            result.subject2 = request.form.get('subject2')
            result.subject2_score = request.form.get('subject2_score', type=int)
            result.subject3 = request.form.get('subject3')
            result.subject3_score = request.form.get('subject3_score', type=int)
            result.subject4 = request.form.get('subject4')
            result.subject4_score = request.form.get('subject4_score', type=int)
            
            # Validate total score
            if result.total_score < 0 or result.total_score > 400:
                flash('Total score must be between 0 and 400.', 'error')
                return redirect(url_for('results.edit_jamb', student_id=student_id, year=year))
            
            db.session.commit()
            log_action('results.jamb_edit', detail=f'{year}', target=student)
            flash('JAMB result updated!', 'success')
            return redirect(url_for('results.view_jamb_student', student_id=student_id))
            
        except Exception as e:
            db.session.rollback()
            flash(f'Error: {str(e)}', 'error')
    
    return render_template('results/edit_jamb.html',
        student=student,
        result=result,
        year=year,
        subjects=WAEC_SUBJECTS
    )


@results_bp.route('/jamb/student/<int:student_id>/delete/<int:year>', methods=['POST'])
@admin_required
def delete_jamb(student_id, year):
    """Delete JAMB result for a student in a given year"""
    student = db.get_or_404(Student, student_id)
    require_branch_access(student.branch_id)
    
    try:
        result = JAMBResult.query.filter_by(student_id=student_id, exam_year=year).first()
        
        if result:
            db.session.delete(result)
            db.session.commit()
            log_action('results.jamb_delete', detail=f'{year}', target=student)
            flash(f'Deleted JAMB result for {student.full_name} ({year}).', 'success')
        else:
            flash('No result found to delete.', 'warning')
            
    except Exception as e:
        db.session.rollback()
        flash(f'Error deleting result: {str(e)}', 'error')
    
    return redirect(url_for('results.jamb_list', year=year))


@results_bp.route('/api/jamb/score-distribution/<int:year>')
@login_required
def api_jamb_score_distribution(year):
    """Get JAMB score distribution for charts"""
    results = scope_by_student(JAMBResult.query.filter_by(exam_year=year), JAMBResult).all()
    scores = [r.total_score for r in results]
    
    distribution = {
        '0-100': sum(1 for s in scores if s <= 100),
        '101-150': sum(1 for s in scores if 101 <= s <= 150),
        '151-200': sum(1 for s in scores if 151 <= s <= 200),
        '201-250': sum(1 for s in scores if 201 <= s <= 250),
        '251-300': sum(1 for s in scores if 251 <= s <= 300),
        '301-350': sum(1 for s in scores if 301 <= s <= 350),
        '351-400': sum(1 for s in scores if 351 <= s <= 400)
    }
    
    return jsonify(distribution)


@results_bp.route('/jamb/export')
@login_required
@rate_limited('export', max_requests=40, window_minutes=10)
def export_jamb():
    """Export JAMB results to Excel"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Border, Side
    from io import BytesIO
    
    year = request.args.get('year', type=int)
    if not year:
        flash('Please select a year to export.', 'error')
        return redirect(url_for('results.jamb_list'))
    log_action('data.export_results', detail=f'JAMB {year}')

    results = (scope_by_student(JAMBResult.query.filter_by(exam_year=year), JAMBResult)
               .options(joinedload(JAMBResult.student)).join(Student)
               .order_by(JAMBResult.total_score.desc()).all())

    wb = Workbook()
    ws = wb.active
    ws.title = f"JAMB {year}"
    
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF')
    border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    
    ws.merge_cells('A1:L1')
    ws['A1'] = f'JAMB RESULTS - {year}'
    ws['A1'].font = Font(bold=True, size=14)
    
    headers = ['Rank', 'Student ID', 'Name', 'Total Score', 'Subject 1', 'Score', 'Subject 2', 'Score', 'Subject 3', 'Score', 'Subject 4', 'Score']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
    
    for idx, r in enumerate(results, 1):
        row = idx + 3
        ws.cell(row=row, column=1, value=idx).border = border
        ws.cell(row=row, column=2, value=r.student.student_id).border = border
        ws.cell(row=row, column=3, value=r.student.full_name).border = border
        ws.cell(row=row, column=4, value=r.total_score).border = border
        ws.cell(row=row, column=5, value=r.subject1 or '-').border = border
        ws.cell(row=row, column=6, value=r.subject1_score or '-').border = border
        ws.cell(row=row, column=7, value=r.subject2 or '-').border = border
        ws.cell(row=row, column=8, value=r.subject2_score or '-').border = border
        ws.cell(row=row, column=9, value=r.subject3 or '-').border = border
        ws.cell(row=row, column=10, value=r.subject3_score or '-').border = border
        ws.cell(row=row, column=11, value=r.subject4 or '-').border = border
        ws.cell(row=row, column=12, value=r.subject4_score or '-').border = border
    
    ws.delete_cols(2)        # drop the admission-number column from the printout
    return xlsx_response(wb, f'jamb_results_{year}.xlsx')
