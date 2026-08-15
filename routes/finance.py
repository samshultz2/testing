"""
Finance / Fees routes — fee catalogue, fee structure per class/term, student
payments with printable receipts, discounts/waivers, defaulters and a finance
dashboard.
"""
from datetime import date, datetime
from utils import timeutil
from utils.helpers import get_active_term, session_terms

from flask import (Blueprint, render_template, request, redirect, url_for,
                   flash, jsonify, Response, session)
from sqlalchemy import func
from sqlalchemy.orm import joinedload
from utils.web_exports import xlsx_response, csv_response
from utils.branch_scope import require_branch_access, scope_query, scope_by_student, viewing_branch_id

from models import (
    db, FeeItem, FeeStructure, FeePayment, FeeDiscount, ExpenseCategory, Expense,
    Student, Term, SchoolClass, ClassArm, StudentEnrollment,
    ClassArmAssignment, AdditionalCharge, InstallmentPlan,
)
from utils.access_control import (
    login_required, admin_required, filter_classes_for_user,
)
from utils.finance import (
    student_bill, class_fee_total, next_receipt_no, collection_trend,
    fee_item_breakdown,
)
from utils.search import like_term

finance_bp = Blueprint('finance', __name__, url_prefix='/finance')

PAYMENT_METHODS = ['Cash', 'Bank Transfer', 'POS', 'Cheque', 'Online']


def _active_term_id():
    t = get_active_term()
    return t.id if t else None


def _parse_date(value, default=None):
    try:
        return datetime.strptime(value, '%Y-%m-%d').date()
    except (ValueError, TypeError):
        return default or timeutil.today()


# --- SPA helpers (no-reload React shell + JSON-aware action responses) ---
from utils.spa import section_responders
_wants_json, _render, _ok, _err = section_responders(
    'finance/app.html', 'fin_json', 'finance.dashboard')


def _is_admin():
    from utils.access_control import is_admin
    return is_admin()


def _bill_json(bill):
    if not bill:
        return None
    return {'lines': [{'name': it.name, 'amount': amt} for it, amt in bill['lines']],
            'billed': bill['billed'], 'discount': bill['discount'], 'paid': bill['paid'],
            'payable': bill['payable'], 'balance': bill['balance'],
            'class_id': bill['class_id'], 'arm_id': bill['arm_id']}


def _nav_urls():
    return {'dashboard': url_for('finance.dashboard'), 'record_payment': url_for('finance.record_payment'),
            'payments': url_for('finance.payments_list'), 'structure': url_for('finance.structure'),
            'items': url_for('finance.items_list'), 'expenses': url_for('finance.expenses_list'),
            'defaulters': url_for('finance.defaulters'), 'collections': url_for('finance.collections'),
            'reports': url_for('finance.reports')}


def _collections_query(from_date, to_date, term_id=None):
    from utils.branch_scope import scope_query
    q = scope_query(
        FeePayment.query.filter(FeePayment.payment_date >= from_date,
                                FeePayment.payment_date <= to_date),
        FeePayment)
    if term_id:
        q = q.filter(FeePayment.term_id == term_id)
    # Eager-load the student — every caller renders p.student.* in a loop (was N+1).
    return q.options(joinedload(FeePayment.student)).order_by(
        FeePayment.payment_date.desc(), FeePayment.id.desc())


@finance_bp.route('/collections')
@login_required
def collections():
    """Day-book: payments collected within a date range, with daily/method totals.

    Spans all terms by default (cash reconciliation), with an optional filter
    to narrow the day-book to a single term.
    """
    today = timeutil.today()
    from_date = _parse_date(request.args.get('from'), today.replace(day=1))
    to_date = _parse_date(request.args.get('to'), today)
    term_id = request.args.get('term_id', type=int)
    payments = _collections_query(from_date, to_date, term_id).all()
    total = sum(p.amount for p in payments)

    by_method = {}
    by_day = {}
    for p in payments:
        by_method[p.method or 'Other'] = by_method.get(p.method or 'Other', 0.0) + p.amount
        key = p.payment_date.isoformat()
        by_day[key] = by_day.get(key, 0.0) + p.amount
    method_rows = sorted(by_method.items(), key=lambda x: x[1], reverse=True)
    day_chart = [{'label': d, 'amount': round(v, 2)} for d, v in sorted(by_day.items())]

    return _render({
        'page': 'collections', 'nav': _nav_urls(),
        'from_date': from_date.isoformat(), 'to_date': to_date.isoformat(),
        'from_label': from_date.strftime('%d %b %Y'), 'to_label': to_date.strftime('%d %b %Y'),
        'term_id': term_id or '', 'total': total, 'count': len(payments),
        'method_rows': [{'method': m, 'amount': round(v, 2)} for m, v in method_rows],
        'day_chart': day_chart,
        'payments': [{'id': p.id, 'date': p.payment_date.strftime('%d %b %Y'),
                      'receipt_no': p.receipt_no, 'student': p.student.full_name if p.student else '—',
                      'method': p.method, 'amount': p.amount,
                      'receipt_url': url_for('finance.receipt', payment_id=p.id)} for p in payments],
        'terms': [{'id': t.id, 'full_name': t.full_name or t.name} for t in
                  session_terms()],
        'self_url': url_for('finance.collections'),
        'export_url': url_for('finance.collections_export', **{'from': from_date.isoformat(),
                              'to': to_date.isoformat(), 'term_id': term_id or ''}),
    })


@finance_bp.route('/collections/export')
@login_required
def collections_export():
    import csv, io
    today = timeutil.today()
    from_date = _parse_date(request.args.get('from'), today.replace(day=1))
    to_date = _parse_date(request.args.get('to'), today)
    term_id = request.args.get('term_id', type=int)
    out = io.StringIO()
    w = csv.writer(out)
    w.writerow(['Date', 'Receipt', 'Student', 'Student ID', 'Term', 'Method',
                'Reference', 'Received By', 'Amount'])
    from utils.web_exports import formula_guard as _fg
    for p in _collections_query(from_date, to_date, term_id).all():
        w.writerow([p.payment_date.strftime('%Y-%m-%d'), _fg(p.receipt_no),
                    _fg(p.student.full_name if p.student else ''),
                    _fg(p.student.student_id if p.student else ''),
                    _fg(p.term.full_name if p.term else ''), _fg(p.method),
                    _fg(p.reference or ''), _fg(p.received_by or ''), p.amount])
    fname = f'collections_{from_date}_{to_date}.csv'
    return csv_response(out.getvalue(), f'{fname}')


# ============================================================================
# DASHBOARD
# ============================================================================

@finance_bp.route('/')
@login_required
def dashboard():
    term_id = request.args.get('term_id', type=int) or _active_term_id()
    terms = session_terms()
    selected_term = db.session.get(Term, term_id) if term_id else None

    fees = _term_fee_summary(term_id)
    exp = _term_expense_summary(term_id)

    payable_total = max(fees['expected'] - fees['discounts'], 0)
    outstanding = payable_total - fees['collected']
    rate = round(fees['collected'] / payable_total * 100, 1) if payable_total > 0 else 0.0
    net = fees['collected'] - exp['expenses']

    class_chart = [{'name': k, 'expected': round(v['expected'], 2),
                    'collected': round(v['collected'], 2)}
                   for k, v in sorted(fees['by_class'].items())]
    method_chart = [{'method': k, 'amount': round(v, 2)}
                    for k, v in fees['method_breakdown'].items()]
    item_chart = fee_item_breakdown(term_id) if term_id else []
    trend_chart = collection_trend(term_id) if term_id else []
    expense_chart = [{'name': k, 'amount': round(v, 2)} for k, v in
                     sorted(exp['expense_breakdown'].items(), key=lambda x: x[1], reverse=True)]

    return _render({
        'page': 'dashboard', 'nav': _nav_urls(),
        'term_id': term_id or '', 'selected_term': selected_term.full_name if selected_term else '',
        'terms': [{'id': t.id, 'full_name': t.full_name} for t in terms],
        'expected': fees['expected'], 'collected': fees['collected'], 'discounts': fees['discounts'],
        'payable_total': payable_total, 'outstanding': outstanding, 'rate': rate,
        'expenses': exp['expenses'], 'net': net,
        'enrolled': fees['enrolled'], 'defaulter_count': fees['defaulter_count'],
        'class_chart': class_chart, 'method_chart': method_chart,
        'item_chart': item_chart, 'trend_chart': trend_chart, 'expense_chart': expense_chart,
        'has_structure': fees['expected'] > 0,
        'recent': [{'id': p.id, 'date': p.payment_date.strftime('%d %b'),
                    'student': p.student.full_name if p.student else '—', 'receipt_no': p.receipt_no,
                    'method': p.method, 'amount': p.amount,
                    'receipt_url': url_for('finance.receipt', payment_id=p.id)} for p in fees['recent']],
        'urls': {'record_payment': url_for('finance.record_payment', term_id=term_id or ''),
                 'defaulters': url_for('finance.defaulters', term_id=term_id or ''),
                 'items': url_for('finance.items_list'),
                 'structure': url_for('finance.structure', term_id=term_id or ''),
                 'payments': url_for('finance.payments_list', term_id=term_id or '')},
        'self_url': url_for('finance.dashboard'),
    })


def _term_fee_summary(term_id):
    """Fee expectation, collection, method split, recent payments + defaulter count."""
    summary = {'expected': 0.0, 'collected': 0.0, 'discounts': 0.0,
               'by_class': {}, 'method_breakdown': {}, 'recent': [],
               'enrolled': 0, 'defaulter_count': 0}
    if not term_id:
        return summary
    from utils.branch_scope import scope_query
    by_class = summary['by_class']
    method_breakdown = summary['method_breakdown']

    # Map each enrolled student to their class/arm for this term. Eager-load the
    # assignment + its class so the loop below doesn't lazy-load per enrollment (N+1).
    from sqlalchemy.orm import joinedload as _joinedload
    enrollments = scope_query(
        StudentEnrollment.query
        .join(ClassArmAssignment,
              StudentEnrollment.class_arm_assignment_id == ClassArmAssignment.id)
        .filter(StudentEnrollment.is_active == True,
                ClassArmAssignment.term_id == term_id),
        ClassArmAssignment).options(
            _joinedload(StudentEnrollment.class_arm_assignment)
            .joinedload(ClassArmAssignment.school_class)).all()

    placement = {}        # student_id -> (class_id, arm_id, class_name)
    total_cache = {}      # (class_id, arm_id) -> per-student fee total
    expected = 0.0
    for e in enrollments:
        asg = e.class_arm_assignment
        key = (asg.class_id, asg.arm_id)
        if key not in total_cache:
            total_cache[key] = class_fee_total(term_id, asg.class_id, asg.arm_id)
        cname = asg.school_class.name if asg.school_class else '—'
        placement[e.student_id] = (asg.class_id, asg.arm_id, cname)
        slot = by_class.setdefault(cname, {'expected': 0.0, 'collected': 0.0})
        slot['expected'] += total_cache[key]
        expected += total_cache[key]
    summary['expected'] = expected
    summary['enrolled'] = len(placement)

    from utils.branch_scope import scope_by_student
    summary['discounts'] = (scope_by_student(
        db.session.query(func.coalesce(func.sum(FeeDiscount.amount), 0.0))
        .filter(FeeDiscount.term_id == term_id), FeeDiscount).scalar()) or 0.0

    payments = scope_query(FeePayment.query.filter_by(term_id=term_id), FeePayment).all()
    collected = 0.0
    paid_by_student = {}
    for p in payments:
        collected += p.amount
        method_breakdown[p.method or 'Other'] = \
            method_breakdown.get(p.method or 'Other', 0.0) + p.amount
        paid_by_student[p.student_id] = paid_by_student.get(p.student_id, 0.0) + p.amount
        plc = placement.get(p.student_id)
        if plc:
            by_class.setdefault(plc[2], {'expected': 0.0, 'collected': 0.0})['collected'] += p.amount
    summary['collected'] = collected
    summary['recent'] = (scope_query(FeePayment.query.filter_by(term_id=term_id), FeePayment)
                         .order_by(FeePayment.created_at.desc()).limit(8).all())

    # Count students with an outstanding balance.
    disc_by_student = {}
    for d in scope_by_student(FeeDiscount.query.filter_by(term_id=term_id), FeeDiscount).all():
        disc_by_student[d.student_id] = disc_by_student.get(d.student_id, 0.0) + d.amount
    defaulters = 0
    for sid, (cid, aid, _cn) in placement.items():
        payable = max(total_cache[(cid, aid)] - disc_by_student.get(sid, 0.0), 0)
        if payable - paid_by_student.get(sid, 0.0) > 0.005:
            defaulters += 1
    summary['defaulter_count'] = defaulters
    return summary


def _term_expense_summary(term_id):
    """Total expenses, per-category breakdown and the latest few expenses."""
    summary = {'expenses': 0.0, 'expense_breakdown': {}, 'recent_expenses': []}
    if not term_id:
        return summary
    from utils.branch_scope import scope_query
    exp_rows = scope_query(
        Expense.query.filter_by(term_id=term_id).options(joinedload(Expense.category)),
        Expense).order_by(Expense.created_at.desc()).all()
    breakdown = summary['expense_breakdown']
    total = 0.0
    for e in exp_rows:
        total += e.amount
        cat = e.category.name if e.category else 'Uncategorised'
        breakdown[cat] = breakdown.get(cat, 0.0) + e.amount
    summary['expenses'] = total
    summary['recent_expenses'] = exp_rows[:6]
    return summary


# ============================================================================
# FEE ITEMS (catalogue)
# ============================================================================

@finance_bp.route('/items')
@login_required
def items_list():
    items = FeeItem.query.order_by(FeeItem.is_active.desc(), FeeItem.name).all()
    return _render({
        'page': 'items', 'nav': _nav_urls(), 'is_admin': _is_admin(),
        'items': [{'id': i.id, 'name': i.name, 'description': i.description or '',
                   'is_active': bool(i.is_active),
                   'edit_url': url_for('finance.edit_item', item_id=i.id),
                   'delete_url': url_for('finance.delete_item', item_id=i.id)} for i in items],
        'add_url': url_for('finance.add_item'),
        'structure_url': url_for('finance.structure'),
    })


@finance_bp.route('/items/add', methods=['POST'])
@login_required   # finance CRUD: gated to finance 'edit' users (bursars/accountants/admins) by enforce_write_level / subsection access
def add_item():
    name = (request.form.get('name') or '').strip()
    if not name:
        return _err('Fee item name is required.', url_for('finance.items_list'))
    if FeeItem.query.filter(func.lower(FeeItem.name) == name.lower()).first():
        return _err(f'"{name}" already exists.', url_for('finance.items_list'))
    db.session.add(FeeItem(name=name, description=(request.form.get('description') or '').strip()))
    db.session.commit()
    return _ok(f'Added fee item "{name}".', url_for('finance.items_list'))


@finance_bp.route('/items/<int:item_id>/edit', methods=['POST'])
@login_required   # finance CRUD: gated to finance 'edit' users (bursars/accountants/admins) by enforce_write_level / subsection access
def edit_item(item_id):
    item = db.get_or_404(FeeItem, item_id)
    name = (request.form.get('name') or '').strip()
    if name:
        item.name = name
    item.description = (request.form.get('description') or '').strip()
    item.is_active = bool(request.form.get('is_active'))
    db.session.commit()
    return _ok('Fee item updated.', url_for('finance.items_list'))


@finance_bp.route('/items/<int:item_id>/delete', methods=['POST'])
@login_required   # finance CRUD: gated to finance 'edit' users (bursars/accountants/admins) by enforce_write_level / subsection access
def delete_item(item_id):
    item = db.get_or_404(FeeItem, item_id)
    from utils.audit import log_action
    used = FeeStructure.query.filter_by(fee_item_id=item_id).count()
    if used:
        item.is_active = False
        db.session.commit()
        log_action('finance.item_deactivate', detail=item.name, target=item)
        return _ok(f'"{item.name}" is used in {used} fee structure row(s); deactivated instead of deleted.',
                   url_for('finance.items_list'))
    log_action('finance.item_delete', detail=item.name, target=item)
    db.session.delete(item)
    db.session.commit()
    return _ok('Fee item deleted.', url_for('finance.items_list'))


# ============================================================================
# FEE STRUCTURE (amounts per class/term)
# ============================================================================

@finance_bp.route('/structure')
@login_required
def structure():
    term_id = request.args.get('term_id', type=int) or _active_term_id()
    class_id = request.args.get('class_id', type=int)
    arm_id = request.args.get('arm_id', type=int)

    terms = session_terms()
    classes = SchoolClass.query.filter_by(is_active=True).order_by(SchoolClass.level).all()
    arms = ClassArm.query.filter_by(is_active=True, is_default=False).order_by(ClassArm.name).all()
    items = FeeItem.query.filter_by(is_active=True).order_by(FeeItem.name).all()

    current = {}
    total = 0.0
    if term_id and class_id:
        rows = FeeStructure.query.filter_by(
            term_id=term_id, class_id=class_id, arm_id=arm_id).all()
        for r in rows:
            current[r.fee_item_id] = r.amount
        total = sum(current.values())

    return _render({
        'page': 'structure', 'nav': _nav_urls(), 'is_admin': _is_admin(),
        'term_id': term_id or '', 'class_id': class_id or '', 'arm_id': arm_id or '',
        'terms': [{'id': t.id, 'full_name': t.full_name} for t in terms],
        'classes': [{'id': c.id, 'name': c.name} for c in classes],
        'arms': [{'id': a.id, 'name': a.name} for a in arms],
        'items': [{'id': i.id, 'name': i.name, 'description': i.description or '',
                   'amount': current.get(i.id, '')} for i in items],
        'has_current': bool(current), 'total': total,
        'self_url': url_for('finance.structure'), 'items_url': url_for('finance.items_list'),
        'urls': {'save': url_for('finance.save_structure'), 'copy': url_for('finance.copy_structure'),
                 'clear': url_for('finance.clear_structure')},
    })


@finance_bp.route('/structure/save', methods=['POST'])
@login_required   # finance CRUD: gated to finance 'edit' users (bursars/accountants/admins) by enforce_write_level / subsection access
def save_structure():
    term_id = request.form.get('term_id', type=int)
    class_id = request.form.get('class_id', type=int)
    arm_id = request.form.get('arm_id', type=int)  # None -> all arms
    if not (term_id and class_id):
        return _err('Select a term and class first.', url_for('finance.structure'))

    items = FeeItem.query.filter_by(is_active=True).all()
    changed = 0
    for item in items:
        raw = (request.form.get(f'amount_{item.id}') or '').strip()
        existing = FeeStructure.query.filter_by(
            term_id=term_id, class_id=class_id, arm_id=arm_id,
            fee_item_id=item.id).first()
        if raw == '':
            if existing:
                db.session.delete(existing)
                changed += 1
            continue
        try:
            amount = float(raw)
        except ValueError:
            continue
        if amount < 0:
            amount = 0
        if existing:
            if existing.amount != amount:
                existing.amount = amount
                existing.is_active = True
                changed += 1
        else:
            db.session.add(FeeStructure(
                term_id=term_id, class_id=class_id, arm_id=arm_id,
                fee_item_id=item.id, amount=amount))
            changed += 1

    db.session.commit()
    return _ok(f'Fee structure saved ({changed} change(s)).',
               url_for('finance.structure', term_id=term_id, class_id=class_id, arm_id=arm_id or ''))


@finance_bp.route('/structure/copy', methods=['POST'])
@login_required   # finance CRUD: gated to finance 'edit' users (bursars/accountants/admins) by enforce_write_level / subsection access
def copy_structure():
    """Copy a whole term's fee structure into another term."""
    from_term_id = request.form.get('from_term_id', type=int)
    to_term_id = request.form.get('to_term_id', type=int)
    if not (from_term_id and to_term_id) or from_term_id == to_term_id:
        return _err('Choose two different terms.', url_for('finance.structure', term_id=to_term_id or ''))
    src = FeeStructure.query.filter_by(term_id=from_term_id, is_active=True).all()
    copied = skipped = 0
    for r in src:
        exists = FeeStructure.query.filter_by(
            term_id=to_term_id, class_id=r.class_id, arm_id=r.arm_id,
            fee_item_id=r.fee_item_id).first()
        if exists:
            skipped += 1
            continue
        db.session.add(FeeStructure(
            term_id=to_term_id, class_id=r.class_id, arm_id=r.arm_id,
            fee_item_id=r.fee_item_id, amount=r.amount))
        copied += 1
    db.session.commit()
    return _ok(f'Copied {copied} fee row(s); skipped {skipped} already set.',
               url_for('finance.structure', term_id=to_term_id))


@finance_bp.route('/structure/clear', methods=['POST'])
@login_required   # finance CRUD: gated to finance 'edit' users (bursars/accountants/admins) by enforce_write_level / subsection access
def clear_structure():
    """Remove every fee row for a term + class (+ arm)."""
    term_id = request.form.get('term_id', type=int)
    class_id = request.form.get('class_id', type=int)
    arm_id = request.form.get('arm_id', type=int)
    if not (term_id and class_id):
        return _err('Select a term and class first.', url_for('finance.structure'))
    deleted = FeeStructure.query.filter_by(
        term_id=term_id, class_id=class_id, arm_id=arm_id).delete()
    db.session.commit()
    return _ok(f'Cleared {deleted} fee row(s) for this class.',
               url_for('finance.structure', term_id=term_id, class_id=class_id, arm_id=arm_id or ''))


# ============================================================================
# PAYMENTS
# ============================================================================

@finance_bp.route('/payments')
@login_required
def payments_list():
    term_id = request.args.get('term_id', type=int) or _active_term_id()
    class_id = request.args.get('class_id', type=int)
    q = (request.args.get('q') or '').strip()

    terms = session_terms()
    classes = SchoolClass.query.filter_by(is_active=True).order_by(SchoolClass.level).all()

    from utils.branch_scope import scope_query
    from sqlalchemy.orm import contains_eager
    query = FeePayment.query.filter_by(term_id=term_id) if term_id else FeePayment.query
    # Eager-load the already-joined Student so the render loop doesn't re-query it (N+1).
    query = (scope_query(query, FeePayment).join(Student, FeePayment.student_id == Student.id)
             .options(contains_eager(FeePayment.student)))
    if q:
        like = like_term(q)
        query = query.filter(db.or_(Student.surname.ilike(like, escape='\\'),
                                    Student.first_name.ilike(like, escape='\\'),
                                    Student.student_id.ilike(like, escape='\\'),
                                    FeePayment.receipt_no.ilike(like, escape='\\')))
    if class_id and term_id:
        sids = [sid for (sid,) in (StudentEnrollment.query
                .join(ClassArmAssignment,
                      StudentEnrollment.class_arm_assignment_id == ClassArmAssignment.id)
                .filter(StudentEnrollment.is_active == True,
                        ClassArmAssignment.term_id == term_id,
                        ClassArmAssignment.class_id == class_id)
                .with_entities(StudentEnrollment.student_id).all())]
        query = query.filter(FeePayment.student_id.in_(sids or [-1]))

    payments = query.order_by(FeePayment.payment_date.desc(), FeePayment.id.desc()).all()
    total = sum(p.amount for p in payments)

    return _render({
        'page': 'payments', 'nav': _nav_urls(), 'is_admin': _is_admin(),
        'term_id': term_id or '', 'class_id': class_id or '', 'q': q, 'total': total,
        'terms': [{'id': t.id, 'full_name': t.full_name} for t in terms],
        'classes': [{'id': c.id, 'name': c.name} for c in classes],
        'payments': [{'id': p.id, 'date': p.payment_date.strftime('%d %b %Y'), 'receipt_no': p.receipt_no,
                      'student': p.student.full_name if p.student else '—',
                      'student_id': p.student.student_id if p.student else '',
                      'method': p.method, 'received_by': p.received_by or '—', 'amount': p.amount,
                      'receipt_url': url_for('finance.receipt', payment_id=p.id),
                      'edit_url': url_for('finance.edit_payment', payment_id=p.id),
                      'delete_url': url_for('finance.delete_payment', payment_id=p.id),
                      'statement_url': url_for('finance.statement', student_id=p.student_id, term_id=p.term_id)}
                     for p in payments],
        'self_url': url_for('finance.payments_list'),
        'record_url': url_for('finance.record_payment', term_id=term_id or ''),
    })


@finance_bp.route('/payments/record', methods=['GET', 'POST'])
@login_required
def record_payment():
    term_id = request.values.get('term_id', type=int) or _active_term_id()
    student_id = request.values.get('student_id', type=int)

    if request.method == 'POST':
        student_id = request.form.get('student_id', type=int)
        term_id = request.form.get('term_id', type=int)
        amount = request.form.get('amount', type=float)
        if not (student_id and term_id and amount and amount > 0):
            return _err('Select a student, term and a positive amount.',
                        url_for('finance.record_payment', term_id=term_id, student_id=student_id))
        # Idempotency: when a transaction reference is supplied, refuse a second
        # payment with the same reference for the same student+term (double-submit
        # / retried POST). References are unique per real transaction.
        reference = (request.form.get('reference') or '').strip() or None
        if reference and FeePayment.query.filter_by(
                student_id=student_id, term_id=term_id, reference=reference).first():
            return _err(f'A payment with reference "{reference}" is already recorded '
                        f'for this student and term.',
                        url_for('finance.record_payment', term_id=term_id, student_id=student_id))
        payment = FeePayment(
            student_id=student_id,
            term_id=term_id,
            amount=amount,
            payment_date=_parse_date(request.form.get('payment_date')),
            method=request.form.get('method') or 'Cash',
            reference=reference,
            received_by=(request.form.get('received_by') or '').strip() or None,
            notes=(request.form.get('notes') or '').strip() or None,
            receipt_no=next_receipt_no(),
        )
        # Inherit the paying student's branch.
        stu = db.session.get(Student, student_id)
        if stu:
            require_branch_access(stu.branch_id)    # no recording fees across branches
        payment.branch_id = stu.branch_id if stu else None
        db.session.add(payment)
        db.session.commit()
        from utils import query_cache
        query_cache.bump('dash')              # finance KPIs changed
        from utils.audit import log_action
        log_action('finance.payment',
                   detail=f'{payment.amount:g} from {stu.full_name if stu else "—"}',
                   target=payment)
        return _ok(f'Payment recorded — receipt {payment.receipt_no}.',
                   url_for('finance.receipt', payment_id=payment.id))

    terms = session_terms()
    student = db.session.get(Student, student_id) if student_id else None
    bill = student_bill(student_id, term_id) if (student and term_id) else None

    # Class/arm roster picker — lets the user browse a class arm and pick a
    # student (disambiguates students who share a name across arms).
    assignment_id = request.values.get('assignment_id', type=int)
    assignments = []
    if term_id:
        all_assignments = ClassArmAssignment.query.filter_by(term_id=term_id).all()
        assignments = filter_classes_for_user(all_assignments)
        assignments.sort(key=lambda a: a.display_name)
    roster = []
    if term_id and assignment_id and not student:
        enrollments = (StudentEnrollment.query.filter_by(
            class_arm_assignment_id=assignment_id, is_active=True)
            .join(Student).order_by(Student.surname, Student.first_name).all())
        for e in enrollments:
            b = student_bill(e.student_id, term_id)
            roster.append({'student': e.student, 'balance': b['balance'],
                           'paid': b['paid'], 'billed': b['payable']})

    return _render({
        'page': 'record_payment', 'nav': _nav_urls(),
        'term_id': term_id or '', 'assignment_id': assignment_id or '',
        'terms': [{'id': t.id, 'full_name': t.full_name} for t in terms],
        'methods': PAYMENT_METHODS, 'today': timeutil.today().isoformat(),
        'current_user': session.get('user', ''),
        'student': ({'id': student.id, 'full_name': student.full_name,
                     'student_id': student.student_id} if student else None),
        'bill': _bill_json(bill),
        'assignments': [{'id': a.id, 'display_name': a.display_name} for a in assignments],
        'roster': [{'id': r['student'].id, 'full_name': r['student'].full_name,
                    'student_id': r['student'].student_id, 'balance': r['balance'],
                    'pick_url': url_for('finance.record_payment', student_id=r['student'].id, term_id=term_id)}
                   for r in roster],
        'self_url': url_for('finance.record_payment'),
        'submit_url': url_for('finance.record_payment'),
        'search_url': url_for('finance.search_students'),
        'urls': {'statement': (url_for('finance.statement', student_id=student.id, term_id=term_id) if student else ''),
                 'structure': (url_for('finance.structure', term_id=term_id, class_id=bill['class_id']) if bill else ''),
                 'change_student': url_for('finance.record_payment', term_id=term_id or '', assignment_id=assignment_id or '')},
    })


@finance_bp.route('/payments/search')
@login_required
def search_students():
    """JSON student lookup for the payment form."""
    q = (request.args.get('q') or '').strip()
    if len(q) < 2:
        return jsonify([])
    term_id = request.args.get('term_id', type=int) or _active_term_id()
    like = like_term(q)
    students = (Student.query.filter_by(is_active=True)
                .filter(db.or_(Student.surname.ilike(like, escape='\\'),
                               Student.first_name.ilike(like, escape='\\'),
                               Student.student_id.ilike(like, escape='\\')))
                .order_by(Student.surname).limit(15).all())
    out = []
    for s in students:
        # Class/arm in the selected term, to tell same-named students apart.
        cls = ''
        enr = (StudentEnrollment.query
               .join(ClassArmAssignment,
                     StudentEnrollment.class_arm_assignment_id == ClassArmAssignment.id)
               .filter(StudentEnrollment.student_id == s.id,
                       StudentEnrollment.is_active == True,
                       ClassArmAssignment.term_id == term_id).first()) if term_id else None
        if enr:
            cls = enr.class_arm_assignment.display_name
        out.append({'id': s.id, 'name': s.full_name, 'sid': s.student_id, 'cls': cls})
    return jsonify(out)


@finance_bp.route('/payments/<int:payment_id>/receipt')
@login_required
def receipt(payment_id):
    payment = db.get_or_404(FeePayment, payment_id)
    from utils.branch_scope import require_branch_access
    require_branch_access(payment.branch_id)
    bill = student_bill(payment.student_id, payment.term_id)
    from models import SchoolSettings
    school = {
        'name': SchoolSettings.get('school_name', 'My School'),
        'address': SchoolSettings.get('school_address', ''),
        'phone': SchoolSettings.get('school_phone', ''),
        'motto': SchoolSettings.get('school_motto', ''),
    }
    return render_template('finance/receipt.html', payment=payment, bill=bill, school=school)


@finance_bp.route('/payments/<int:payment_id>/receipt.pdf')
@login_required
def receipt_pdf(payment_id):
    """Download the receipt as a real PDF (reliable on any device)."""
    payment = db.get_or_404(FeePayment, payment_id)
    from utils.branch_scope import require_branch_access
    require_branch_access(payment.branch_id)
    bill = student_bill(payment.student_id, payment.term_id)
    from models import SchoolSettings
    school = {
        'name': SchoolSettings.get('school_name', 'My School'),
        'address': SchoolSettings.get('school_address', ''),
        'phone': SchoolSettings.get('school_phone', ''),
        'motto': SchoolSettings.get('school_motto', ''),
    }
    from utils.receipt_pdf import receipt_pdf as _make_pdf
    buf, filename = _make_pdf(payment, bill, school)
    from flask import send_file
    return send_file(buf, mimetype='application/pdf', as_attachment=True,
                     download_name=filename)


@finance_bp.route('/payments/<int:payment_id>/edit', methods=['GET', 'POST'])
@login_required
def edit_payment(payment_id):
    """Correct a recorded payment (amount, date, method, reference, notes)."""
    payment = db.get_or_404(FeePayment, payment_id)
    from utils.branch_scope import require_branch_access
    require_branch_access(payment.branch_id)

    if request.method == 'POST':
        amount = request.form.get('amount', type=float)
        if not amount or amount <= 0:
            return _err('Enter a positive amount.', url_for('finance.edit_payment', payment_id=payment_id))
        before_amount = payment.amount        # capture for the audit trail
        payment.amount = amount
        payment.payment_date = _parse_date(request.form.get('payment_date'), payment.payment_date)
        payment.method = request.form.get('method') or payment.method
        payment.reference = (request.form.get('reference') or '').strip() or None
        payment.received_by = (request.form.get('received_by') or '').strip() or None
        payment.notes = (request.form.get('notes') or '').strip() or None
        db.session.commit()
        from utils.audit import log_action
        # Record BOTH the old and new amount — a bare "new amount" hides a
        # down-edit used to skim recorded cash.
        log_action('finance.payment_edit',
                   detail=f'{before_amount:g}→{payment.amount:g}', target=payment)
        return _ok(f'Payment {payment.receipt_no} updated.',
                   url_for('finance.receipt', payment_id=payment.id))

    bill = student_bill(payment.student_id, payment.term_id)
    return _render({
        'page': 'edit_payment', 'nav': _nav_urls(), 'is_admin': _is_admin(),
        'methods': PAYMENT_METHODS, 'bill': _bill_json(bill),
        'payment': {'id': payment.id, 'amount': payment.amount, 'method': payment.method,
                    'payment_date': payment.payment_date.isoformat() if payment.payment_date else '',
                    'reference': payment.reference or '', 'received_by': payment.received_by or '',
                    'notes': payment.notes or '', 'receipt_no': payment.receipt_no,
                    'term_id': payment.term_id, 'term': payment.term.full_name if payment.term else '',
                    'student': payment.student.full_name if payment.student else '—',
                    'student_id_label': payment.student.student_id if payment.student else ''},
        'submit_url': url_for('finance.edit_payment', payment_id=payment.id),
        'receipt_url': url_for('finance.receipt', payment_id=payment.id),
        'delete_url': url_for('finance.delete_payment', payment_id=payment.id),
        'payments_url': url_for('finance.payments_list', term_id=payment.term_id),
    })


@finance_bp.route('/payments/<int:payment_id>/delete', methods=['POST'])
@login_required   # finance CRUD: gated to finance 'edit' users (bursars/accountants/admins) by enforce_write_level / subsection access
def delete_payment(payment_id):
    payment = db.get_or_404(FeePayment, payment_id)
    require_branch_access(payment.branch_id)   # no cross-branch payment deletion
    term_id = payment.term_id
    reason = (request.form.get('reason') or '').strip()
    # Audited reversal: record the full before-state + who/why. The ledger keeps a
    # permanent reversing entry (FeePayment.after_delete), so the money movement is
    # never lost even though the payment row is removed.
    from utils.audit import log_action
    from utils.access_control import get_current_user
    who = getattr(get_current_user(), 'full_name', None) or session.get('user') or 'unknown'
    before = (f'₦{payment.amount:g} · {payment.method} · {payment.receipt_no or "-"} · '
              f'{payment.payment_date} · student {payment.student_id}')
    log_action('finance.payment_reversed',
               detail=f'reversed by {who}: {reason or "(no reason given)"} | was: {before}',
               target_type='feepayment', target_id=payment.id, target_label=payment.receipt_no)
    db.session.delete(payment)
    db.session.commit()
    return _ok('Payment reversed (audit logged, ledger updated).',
               url_for('finance.payments_list', term_id=term_id))


# ============================================================================
# STUDENT STATEMENT + DISCOUNTS
# ============================================================================

@finance_bp.route('/students/<int:student_id>/statement')
@login_required
def statement(student_id):
    student = db.get_or_404(Student, student_id)
    require_branch_access(student.branch_id)        # block cross-branch IDOR
    term_id = request.args.get('term_id', type=int) or _active_term_id()
    terms = session_terms()
    bill = student_bill(student_id, term_id) if term_id else None
    payments = (FeePayment.query.filter_by(student_id=student_id, term_id=term_id)
                .order_by(FeePayment.payment_date).all()) if term_id else []
    discounts = (FeeDiscount.query.filter_by(student_id=student_id, term_id=term_id)
                 .order_by(FeeDiscount.created_at).all()) if term_id else []
    from utils import payments as pay_gw
    return _render({
        'page': 'statement', 'nav': _nav_urls(), 'is_admin': _is_admin(),
        'term_id': term_id or '', 'paylink': request.args.get('paylink') or '',
        'pay_enabled': pay_gw.is_configured(),
        'student': {'id': student.id, 'full_name': student.full_name,
                    'student_id': student.student_id, 'gender': student.gender or ''},
        'terms': [{'id': t.id, 'full_name': t.full_name} for t in terms],
        'bill': _bill_json(bill),
        'payments': [{'id': p.id, 'date': p.payment_date.strftime('%d %b %Y'), 'receipt_no': p.receipt_no,
                      'method': p.method, 'amount': p.amount,
                      'receipt_url': url_for('finance.receipt', payment_id=p.id),
                      'edit_url': url_for('finance.edit_payment', payment_id=p.id)} for p in payments],
        'discounts': [{'id': d.id, 'reason': d.reason or '', 'amount': d.amount,
                       'edit_url': url_for('finance.edit_discount', discount_id=d.id),
                       'delete_url': url_for('finance.delete_discount', discount_id=d.id)} for d in discounts],
        'self_url': url_for('finance.statement', student_id=student.id),
        'urls': {'record_payment': url_for('finance.record_payment', student_id=student.id, term_id=term_id or ''),
                 'payment_link': url_for('finance.payment_link', student_id=student.id),
                 'add_discount': url_for('finance.add_discount')},
        'student_id': student.id,
    })


@finance_bp.route('/students/<int:student_id>/payment-link', methods=['POST'])
@login_required
def payment_link(student_id):
    """Staff: generate a Paystack link to send to a parent (recorded via webhook)."""
    from utils import payments as pay_gw
    student = db.get_or_404(Student, student_id)
    require_branch_access(student.branch_id)   # no cross-branch payment links
    term_id = request.form.get('term_id', type=int) or _active_term_id()
    if not pay_gw.is_configured():
        return _err('Online payment is not configured (set Paystack keys).',
                    url_for('finance.statement', student_id=student_id, term_id=term_id))
    bill = student_bill(student_id, term_id) if term_id else None
    amount = request.form.get('amount', type=float) or (bill['balance'] if bill else 0)
    if not term_id or amount <= 0:
        return _err('No outstanding balance to collect.',
                    url_for('finance.statement', student_id=student_id, term_id=term_id))
    res = pay_gw.initialize(
        email=request.form.get('email') or '', amount_naira=amount,
        reference=pay_gw.new_reference('STF'),
        callback_url=url_for('parent.pay_callback', _external=True),
        metadata={'student_id': student_id, 'term_id': term_id})
    if res.get('ok'):
        return _ok('Payment link created.', url_for('finance.statement', student_id=student_id,
                   term_id=term_id, paylink=res['authorization_url']))
    return _err(res.get('error', 'Could not create payment link.'),
                url_for('finance.statement', student_id=student_id, term_id=term_id))


@finance_bp.route('/discounts/add', methods=['POST'])
@login_required   # finance CRUD: gated to finance 'edit' users (bursars/accountants/admins) by enforce_write_level / subsection access
def add_discount():
    student_id = request.form.get('student_id', type=int)
    term_id = request.form.get('term_id', type=int)
    amount = request.form.get('amount', type=float)
    if not (student_id and term_id and amount and amount > 0):
        return _err('A student, term and positive amount are required.',
                    url_for('finance.statement', student_id=student_id, term_id=term_id))
    # A waiver reduces what a student owes — gate it to the student's branch, just
    # like recording a payment, so a branch admin can't waive fees org-wide.
    student = db.get_or_404(Student, student_id)
    require_branch_access(student.branch_id)
    db.session.add(FeeDiscount(
        student_id=student_id, term_id=term_id, amount=amount,
        reason=(request.form.get('reason') or '').strip() or None))
    db.session.commit()
    from utils.audit import log_action
    log_action('finance.discount_add', detail=f'{amount:g} for {student.full_name}',
               target_type='student', target_id=student.id, target_label=student.full_name)
    return _ok('Discount / waiver applied.',
               url_for('finance.statement', student_id=student_id, term_id=term_id))


@finance_bp.route('/discounts/<int:discount_id>/edit', methods=['POST'])
@login_required   # finance CRUD: gated to finance 'edit' users (bursars/accountants/admins) by enforce_write_level / subsection access
def edit_discount(discount_id):
    d = db.get_or_404(FeeDiscount, discount_id)
    require_branch_access(d.student.branch_id if d.student else None)
    amount = request.form.get('amount', type=float)
    if not amount or amount <= 0:
        return _err('Enter a positive amount.',
                    url_for('finance.statement', student_id=d.student_id, term_id=d.term_id))
    before = d.amount
    d.amount = amount
    d.reason = (request.form.get('reason') or '').strip() or None
    db.session.commit()
    from utils.audit import log_action
    log_action('finance.discount_edit', detail=f'{before:g}→{amount:g}',
               target_type='student', target_id=d.student_id)
    return _ok('Discount updated.', url_for('finance.statement', student_id=d.student_id, term_id=d.term_id))


@finance_bp.route('/discounts/<int:discount_id>/delete', methods=['POST'])
@login_required   # finance CRUD: gated to finance 'edit' users (bursars/accountants/admins) by enforce_write_level / subsection access
def delete_discount(discount_id):
    d = db.get_or_404(FeeDiscount, discount_id)
    require_branch_access(d.student.branch_id if d.student else None)
    student_id, term_id, amount = d.student_id, d.term_id, d.amount
    db.session.delete(d)
    db.session.commit()
    from utils.audit import log_action
    log_action('finance.discount_delete', detail=f'{amount:g}',
               target_type='student', target_id=student_id)
    return _ok('Discount removed.', url_for('finance.statement', student_id=student_id, term_id=term_id))


# ============================================================================
# DEFAULTERS / OUTSTANDING
# ============================================================================

@finance_bp.route('/defaulters')
@login_required
def defaulters():
    term_id = request.args.get('term_id', type=int) or _active_term_id()
    class_id = request.args.get('class_id', type=int)
    terms = session_terms()
    classes = SchoolClass.query.filter_by(is_active=True).order_by(SchoolClass.level).all()

    rows = []
    totals = {'billed': 0.0, 'paid': 0.0, 'balance': 0.0}
    if term_id:
        from utils.branch_scope import scope_query
        enr_q = scope_query(
            StudentEnrollment.query
            .join(ClassArmAssignment,
                  StudentEnrollment.class_arm_assignment_id == ClassArmAssignment.id)
            .filter(StudentEnrollment.is_active == True,
                    ClassArmAssignment.term_id == term_id),
            ClassArmAssignment)
        if class_id:
            enr_q = enr_q.filter(ClassArmAssignment.class_id == class_id)
        enrollments = enr_q.all()

        # Pre-aggregate payments + discounts for the term.
        paid_map = dict(scope_query(db.session.query(FeePayment.student_id, func.sum(FeePayment.amount))
                        .filter(FeePayment.term_id == term_id), FeePayment)
                        .group_by(FeePayment.student_id).all())
        disc_map = dict(scope_by_student(db.session.query(FeeDiscount.student_id, func.sum(FeeDiscount.amount))
                        .filter(FeeDiscount.term_id == term_id), FeeDiscount)
                        .group_by(FeeDiscount.student_id).all())
        from utils.finance import charges_map
        extra_map = charges_map(term_id)                 # net additional charges / credit notes
        total_cache = {}
        for e in enrollments:
            asg = e.class_arm_assignment
            key = (asg.class_id, asg.arm_id)
            if key not in total_cache:
                total_cache[key] = class_fee_total(term_id, asg.class_id, asg.arm_id)
            billed = total_cache[key]
            payable = max(billed + (extra_map.get(e.student_id) or 0.0) - (disc_map.get(e.student_id) or 0.0), 0)
            paid = paid_map.get(e.student_id) or 0.0
            balance = payable - paid
            if balance > 0.005:
                rows.append({
                    'student': e.student,
                    'class_name': asg.school_class.name if asg.school_class else '—',
                    'arm_name': asg.arm_label,
                    'billed': payable, 'paid': paid, 'balance': balance,
                })
                totals['billed'] += payable
                totals['paid'] += paid
                totals['balance'] += balance
        rows.sort(key=lambda r: r['balance'], reverse=True)

    return _render({
        'page': 'defaulters', 'nav': _nav_urls(),
        'term_id': term_id or '', 'class_id': class_id or '', 'totals': totals,
        'terms': [{'id': t.id, 'full_name': t.full_name} for t in terms],
        'classes': [{'id': c.id, 'name': c.name} for c in classes],
        'rows': [{'student': r['student'].full_name, 'student_id': r['student'].student_id,
                  'class_name': r['class_name'], 'arm_name': r['arm_name'],
                  'billed': r['billed'], 'paid': r['paid'], 'balance': r['balance'],
                  'statement_url': url_for('finance.statement', student_id=r['student'].id, term_id=term_id),
                  'record_url': url_for('finance.record_payment', student_id=r['student'].id, term_id=term_id)}
                 for r in rows],
        'self_url': url_for('finance.defaulters'),
        'message_url': url_for('comms.compose', audience='defaulters', class_id=class_id or '', term_id=term_id or ''),
    })


# ============================================================================
# EXPENSES / EXPENDITURE
# ============================================================================

@finance_bp.route('/expenses')
@login_required
def expenses_list():
    term_id = request.args.get('term_id', type=int) or _active_term_id()
    category_id = request.args.get('category_id', type=int)
    terms = session_terms()
    categories = ExpenseCategory.query.filter_by(is_active=True).order_by(ExpenseCategory.name).all()

    from utils.branch_scope import scope_query
    query = scope_query(Expense.query, Expense)
    if term_id:
        query = query.filter_by(term_id=term_id)
    if category_id:
        query = query.filter_by(category_id=category_id)
    expenses = query.order_by(Expense.expense_date.desc(), Expense.id.desc()).all()
    total = sum(e.amount for e in expenses)

    # Per-category totals for the quick summary.
    by_cat = {}
    for e in expenses:
        cat = e.category.name if e.category else 'Uncategorised'
        by_cat[cat] = by_cat.get(cat, 0.0) + e.amount
    cat_summary = sorted(by_cat.items(), key=lambda x: x[1], reverse=True)

    return _render({
        'page': 'expenses', 'nav': _nav_urls(), 'is_admin': _is_admin(),
        'term_id': term_id or '', 'category_id': category_id or '', 'total': total,
        'today': timeutil.today().isoformat(), 'methods': PAYMENT_METHODS,
        'terms': [{'id': t.id, 'full_name': t.full_name} for t in terms],
        'categories': [{'id': c.id, 'name': c.name,
                        'delete_url': url_for('finance.delete_expense_category', category_id=c.id)}
                       for c in categories],
        'cat_summary': [{'name': n, 'amount': v} for n, v in cat_summary],
        'expenses': [{'id': e.id, 'date': e.expense_date.strftime('%d %b %Y'),
                      'expense_date': e.expense_date.isoformat() if e.expense_date else '',
                      'category': e.category.name if e.category else '', 'category_id': e.category_id or '',
                      'description': e.description, 'payee': e.payee or '', 'method': e.method, 'amount': e.amount,
                      'edit_url': url_for('finance.edit_expense', expense_id=e.id),
                      'delete_url': url_for('finance.delete_expense', expense_id=e.id)} for e in expenses],
        'self_url': url_for('finance.expenses_list'), 'reports_url': url_for('finance.reports', term_id=term_id or ''),
        'urls': {'add': url_for('finance.add_expense'), 'add_category': url_for('finance.add_expense_category')},
    })


@finance_bp.route('/expenses/add', methods=['POST'])
@login_required
def add_expense():
    term_id = request.form.get('term_id', type=int)
    description = (request.form.get('description') or '').strip()
    amount = request.form.get('amount', type=float)
    if not (description and amount and amount > 0):
        return _err('A description and positive amount are required.',
                    url_for('finance.expenses_list', term_id=term_id or ''))
    from utils.branch_scope import branch_for_new
    db.session.add(Expense(
        term_id=term_id or None,
        branch_id=branch_for_new(),
        category_id=request.form.get('category_id', type=int) or None,
        description=description,
        amount=amount,
        expense_date=_parse_date(request.form.get('expense_date')),
        payee=(request.form.get('payee') or '').strip() or None,
        method=request.form.get('method') or 'Cash',
        reference=(request.form.get('reference') or '').strip() or None,
        notes=(request.form.get('notes') or '').strip() or None,
    ))
    db.session.commit()
    from utils.audit import log_action
    log_action('finance.expense', detail=f'{amount:g} — {description}',
               target_type='expense', target_label=description)
    return _ok('Expense recorded.', url_for('finance.expenses_list', term_id=term_id or ''))


@finance_bp.route('/expenses/<int:expense_id>/edit', methods=['POST'])
@login_required
def edit_expense(expense_id):
    e = db.get_or_404(Expense, expense_id)
    from utils.branch_scope import require_branch_access
    require_branch_access(e.branch_id)
    amount = request.form.get('amount', type=float)
    description = (request.form.get('description') or '').strip()
    if not (description and amount and amount > 0):
        return _err('A description and positive amount are required.',
                    url_for('finance.expenses_list', term_id=e.term_id or ''))
    e.description = description
    e.amount = amount
    e.category_id = request.form.get('category_id', type=int) or None
    e.expense_date = _parse_date(request.form.get('expense_date'), e.expense_date)
    e.payee = (request.form.get('payee') or '').strip() or None
    e.method = request.form.get('method') or e.method
    e.reference = (request.form.get('reference') or '').strip() or None
    e.notes = (request.form.get('notes') or '').strip() or None
    db.session.commit()
    from utils.audit import log_action
    log_action('finance.expense_edit', detail=f'{amount:g} — {description}',
               target_type='expense', target_id=e.id, target_label=description)
    return _ok('Expense updated.', url_for('finance.expenses_list', term_id=e.term_id or ''))


@finance_bp.route('/expenses/<int:expense_id>/delete', methods=['POST'])
@login_required   # finance CRUD: gated to finance 'edit' users (bursars/accountants/admins) by enforce_write_level / subsection access
def delete_expense(expense_id):
    e = db.get_or_404(Expense, expense_id)
    require_branch_access(e.branch_id)   # no cross-branch expense deletion
    term_id = e.term_id
    from utils.audit import log_action
    log_action('finance.expense_delete', detail=f'{e.amount:g} — {e.description}',
               target_type='expense', target_id=e.id, target_label=e.description)
    db.session.delete(e)
    db.session.commit()
    return _ok('Expense deleted.', url_for('finance.expenses_list', term_id=term_id or ''))


@finance_bp.route('/expense-categories/add', methods=['POST'])
@login_required   # finance CRUD: gated to finance 'edit' users (bursars/accountants/admins) by enforce_write_level / subsection access
def add_expense_category():
    name = (request.form.get('name') or '').strip()
    if not name:
        return _err('Enter a category name.', url_for('finance.expenses_list'))
    if ExpenseCategory.query.filter(func.lower(ExpenseCategory.name) == name.lower()).first():
        return _err(f'Category "{name}" already exists.', url_for('finance.expenses_list'))
    db.session.add(ExpenseCategory(name=name))
    db.session.commit()
    return _ok(f'Added category "{name}".', url_for('finance.expenses_list'))


@finance_bp.route('/expense-categories/<int:category_id>/delete', methods=['POST'])
@login_required   # finance CRUD: gated to finance 'edit' users (bursars/accountants/admins) by enforce_write_level / subsection access
def delete_expense_category(category_id):
    cat = db.get_or_404(ExpenseCategory, category_id)
    from utils.audit import log_action
    if Expense.query.filter_by(category_id=category_id).count():
        cat.is_active = False
        db.session.commit()
        log_action('finance.expense_category_deactivate', detail=getattr(cat, 'name', None), target=cat)
        return _ok('Category is in use; deactivated instead of deleted.', url_for('finance.expenses_list'))
    log_action('finance.expense_category_delete', detail=getattr(cat, 'name', None), target=cat)
    db.session.delete(cat)
    db.session.commit()
    return _ok('Category deleted.', url_for('finance.expenses_list'))


# ============================================================================
# REPORTS + EXPORT
# ============================================================================

@finance_bp.route('/reports')
@login_required
def reports():
    term_id = request.args.get('term_id', type=int) or _active_term_id()
    terms = session_terms()
    selected_term = db.session.get(Term, term_id) if term_id else None

    # Expected (per-class) and collected.
    expected = collected = discounts = expenses = 0.0
    by_class = {}
    if term_id:
        _bid = viewing_branch_id()
        _enq = (StudentEnrollment.query
                .join(ClassArmAssignment,
                      StudentEnrollment.class_arm_assignment_id == ClassArmAssignment.id)
                .filter(StudentEnrollment.is_active == True,
                        ClassArmAssignment.term_id == term_id))
        if _bid is not None:
            _enq = _enq.filter(ClassArmAssignment.branch_id == _bid)
        enrollments = _enq.all()
        placement = {}
        total_cache = {}
        for e in enrollments:
            asg = e.class_arm_assignment
            key = (asg.class_id, asg.arm_id)
            if key not in total_cache:
                total_cache[key] = class_fee_total(term_id, asg.class_id, asg.arm_id)
            cname = asg.school_class.name if asg.school_class else '—'
            placement[e.student_id] = cname
            slot = by_class.setdefault(cname, {'expected': 0.0, 'collected': 0.0, 'students': 0})
            slot['expected'] += total_cache[key]
            slot['students'] += 1
            expected += total_cache[key]
        for p in scope_query(FeePayment.query.filter_by(term_id=term_id), FeePayment).all():
            collected += p.amount
            # Payments from withdrawn students (not in the active placement) go to
            # an "Unassigned" bucket so the class table reconciles to the total.
            cn = placement.get(p.student_id, 'Unassigned / withdrawn')
            by_class.setdefault(cn, {'expected': 0.0, 'collected': 0.0,
                                     'discount': 0.0, 'students': 0})['collected'] += p.amount
        # Per-class discounts (mapped via the student's placement) so each row's
        # outstanding matches the headline payable figure.
        for d in scope_by_student(FeeDiscount.query.filter_by(term_id=term_id), FeeDiscount).all():
            discounts += d.amount
            cn = placement.get(d.student_id)
            if cn and cn in by_class:
                by_class[cn].setdefault('discount', 0.0)
                by_class[cn]['discount'] += d.amount
        expenses = (scope_query(db.session.query(func.coalesce(func.sum(Expense.amount), 0.0))
                    .filter(Expense.term_id == term_id), Expense).scalar()) or 0.0

    payable = max(expected - discounts, 0)
    class_rows = []
    for name, v in sorted(by_class.items()):
        exp = max(v['expected'] - v.get('discount', 0.0), 0)
        col = v['collected']
        class_rows.append({'name': name, 'students': v['students'], 'expected': exp,
                           'collected': col, 'outstanding': max(exp - col, 0),
                           'rate': round(col / exp * 100, 1) if exp > 0 else 0.0})

    item_breakdown = fee_item_breakdown(term_id) if term_id else []
    method_rows = []
    if term_id:
        method_rows = (scope_query(db.session.query(FeePayment.method, func.sum(FeePayment.amount),
                                        func.count(FeePayment.id))
                       .filter(FeePayment.term_id == term_id), FeePayment)
                       .group_by(FeePayment.method).all())

    return _render({
        'page': 'reports', 'nav': _nav_urls(),
        'term_id': term_id or '', 'selected_term': selected_term.full_name if selected_term else '',
        'terms': [{'id': t.id, 'full_name': t.full_name} for t in terms],
        'expected': expected, 'payable': payable, 'collected': collected, 'discounts': discounts,
        'outstanding': payable - collected, 'expenses': expenses, 'net': collected - expenses,
        'rate': round(collected / payable * 100, 1) if payable > 0 else 0.0,
        'class_rows': class_rows, 'item_breakdown': item_breakdown,
        'method_rows': [{'method': m, 'amount': amt, 'count': cnt} for m, amt, cnt in method_rows],
        'self_url': url_for('finance.reports'),
        'export_url': url_for('finance.export_report', term_id=term_id) if term_id else '',
    })


@finance_bp.route('/reports/export')
@login_required
def export_report():
    """Export the term's payments and expenses to a multi-sheet Excel workbook."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    from flask import Response
    import io

    term_id = request.args.get('term_id', type=int) or _active_term_id()
    term = db.session.get(Term, term_id) if term_id else None

    wb = Workbook()
    head_fill = PatternFill(start_color='0d6a4e', end_color='0d6a4e', fill_type='solid')
    head_font = Font(bold=True, color='FFFFFF')

    def write_head(ws, headers):
        for col, h in enumerate(headers, 1):
            c = ws.cell(row=1, column=col, value=h)
            c.fill = head_fill
            c.font = head_font

    # Payments sheet
    ws = wb.active
    ws.title = 'Payments'
    write_head(ws, ['Receipt', 'Date', 'Student', 'Student ID', 'Method', 'Reference', 'Received By', 'Amount'])
    from utils.branch_scope import scope_query
    payments = (scope_query(FeePayment.query.filter_by(term_id=term_id), FeePayment)
                .order_by(FeePayment.payment_date).all()) if term_id else []
    for p in payments:
        ws.append([p.receipt_no, p.payment_date.strftime('%Y-%m-%d'), p.student.full_name,
                   p.student.student_id, p.method, p.reference or '', p.received_by or '', p.amount])

    # Expenses sheet
    ws2 = wb.create_sheet('Expenses')
    write_head(ws2, ['Date', 'Category', 'Description', 'Payee', 'Method', 'Reference', 'Amount'])
    for e in (scope_query(Expense.query.filter_by(term_id=term_id), Expense).order_by(Expense.expense_date).all() if term_id else []):
        ws2.append([e.expense_date.strftime('%Y-%m-%d'), e.category.name if e.category else '',
                    e.description, e.payee or '', e.method, e.reference or '', e.amount])

    # Outstanding sheet
    ws3 = wb.create_sheet('Outstanding')
    write_head(ws3, ['Student', 'Student ID', 'Class', 'Payable', 'Paid', 'Balance'])
    if term_id:
        paid_map = dict(db.session.query(FeePayment.student_id, func.sum(FeePayment.amount))
                        .filter(FeePayment.term_id == term_id).group_by(FeePayment.student_id).all())
        disc_map = dict(db.session.query(FeeDiscount.student_id, func.sum(FeeDiscount.amount))
                        .filter(FeeDiscount.term_id == term_id).group_by(FeeDiscount.student_id).all())
        enrollments = scope_query(
            StudentEnrollment.query
            .join(ClassArmAssignment,
                  StudentEnrollment.class_arm_assignment_id == ClassArmAssignment.id)
            .filter(StudentEnrollment.is_active == True,
                    ClassArmAssignment.term_id == term_id),
            ClassArmAssignment).all()
        cache = {}
        for e in enrollments:
            asg = e.class_arm_assignment
            key = (asg.class_id, asg.arm_id)
            if key not in cache:
                cache[key] = class_fee_total(term_id, asg.class_id, asg.arm_id)
            payable = max(cache[key] - (disc_map.get(e.student_id) or 0.0), 0)
            paid = paid_map.get(e.student_id) or 0.0
            if payable - paid > 0.005:
                ws3.append([e.student.full_name, e.student.student_id,
                            asg.display_name if asg else '', payable, paid, payable - paid])

    fname = f"finance_report_{(term.full_name if term else 'all').replace(' ', '_').replace('/', '-')}.xlsx"
    return xlsx_response(wb, fname)


# ============================================================================
# FINANCIAL OVERVIEW — ledger-backed consolidated reporting (Phase 1)
# ============================================================================
def _overview_date(v):
    try:
        return datetime.strptime(v, '%Y-%m-%d').date()
    except (ValueError, TypeError):
        return None


def _overview_filters():
    """Parse the overview filters, enforcing branch scope (a branch user can only
    ever see their own branch; a central owner may pick one, several, or all)."""
    from utils.branch_scope import is_central, viewing_branch_id
    args = request.args
    if is_central():
        branch_ids = [b for b in args.getlist('branch', type=int) if b]   # empty = all
    else:
        bid = viewing_branch_id()
        branch_ids = [bid] if bid not in (None, -1) else [-1]
    return {
        'branch_ids': branch_ids,
        'session_id': args.get('session_id', type=int),
        'term_id': args.get('term_id', type=int),
        'date_from': _overview_date(args.get('from')),
        'date_to': _overview_date(args.get('to')),
        'source_module': args.get('source') or None,
        'method': args.get('method') or None,
        'direction': args.get('direction') or None,
    }


@finance_bp.route('/overview')
@login_required
def overview():
    """Consolidated financial overview reading the central ledger — every revenue
    source and expense across the platform, filterable by branch/session/term/
    date/source/method, with a multi-branch owner view."""
    from utils import finance_ledger as L
    from utils.branch_scope import is_central
    filters = _overview_filters()
    return render_template('finance/overview.html',
                           d=L.summary(filters), opts=L.filter_options(),
                           sel=request.args, is_central=is_central(),
                           selected_branches=set(filters['branch_ids']),
                           nav=_nav_urls(),
                           export_url=url_for('finance.overview_export'),
                           sync_url=url_for('finance.ledger_sync'))


@finance_bp.route('/overview/export')
@login_required
def overview_export():
    from utils import finance_ledger as L
    rows = L.export_rows(_overview_filters())
    headers = ['Date', 'Type', 'Source', 'Category', 'Method', 'Branch', 'Amount', 'Reference', 'Description']
    if request.args.get('format') == 'xlsx':
        from openpyxl import Workbook
        wb = Workbook(); ws = wb.active; ws.title = 'Ledger'
        ws.append(headers)
        for r in rows:
            ws.append(r)
        return xlsx_response(wb, 'finance_ledger.xlsx')
    import csv
    import io
    buf = io.StringIO()
    w = csv.writer(buf)
    w.writerow(headers)
    w.writerows(rows)
    from flask import Response
    return csv_response(buf.getvalue(), 'finance_ledger.csv')


@finance_bp.route('/ledger/sync', methods=['POST'])
@login_required
def ledger_sync():
    """One-click backfill: mirror any pre-existing payments/expenses/sales into
    the ledger. Idempotent — safe to run repeatedly."""
    from utils import finance_ledger as L
    added = L.backfill()
    flash(f'Ledger synced — {added} existing record(s) added.'
          if added else 'Ledger already up to date — nothing to add.', 'success')
    return redirect(url_for('finance.overview'))


@finance_bp.route('/search')
@login_required
def search():
    """One search box across the whole finance module — matches students,
    fee payments, expenses and charges/credits (branch-scoped) and links each
    hit to where it lives."""
    from utils.search import like_term
    q = (request.args.get('q') or '').strip()
    results = {'students': [], 'payments': [], 'expenses': [], 'charges': []}
    total = 0
    if len(q) >= 2:
        pat = like_term(q)
        esc = '\\'

        students = (scope_query(Student.query.filter(Student.is_active.is_(True)), Student)
                    .filter(db.or_(Student.student_id.ilike(pat, escape=esc),
                                   Student.first_name.ilike(pat, escape=esc),
                                   Student.surname.ilike(pat, escape=esc)))
                    .order_by(Student.surname).limit(15).all())
        results['students'] = [{
            'name': s.full_name, 'sid': s.student_id,
            'url': url_for('finance.record_payment', student_id=s.id)} for s in students]

        pays = (scope_query(FeePayment.query, FeePayment)
                .options(joinedload(FeePayment.student))
                .filter(db.or_(FeePayment.receipt_no.ilike(pat, escape=esc),
                               FeePayment.reference.ilike(pat, escape=esc),
                               FeePayment.received_by.ilike(pat, escape=esc)))
                .order_by(FeePayment.payment_date.desc()).limit(15).all())
        results['payments'] = [{
            'receipt': p.receipt_no or '—', 'ref': p.reference or '',
            'student': p.student.full_name if p.student else '—',
            'amount': round(float(p.amount or 0), 2),
            'date': p.payment_date.strftime('%d %b %Y') if p.payment_date else '',
            'url': url_for('finance.record_payment', student_id=p.student_id)} for p in pays]

        exps = (scope_query(Expense.query, Expense)
                .filter(db.or_(Expense.description.ilike(pat, escape=esc),
                               Expense.reference.ilike(pat, escape=esc),
                               Expense.payee.ilike(pat, escape=esc)))
                .order_by(Expense.expense_date.desc()).limit(15).all())
        results['expenses'] = [{
            'description': e.description or '—', 'payee': e.payee or '',
            'amount': round(float(e.amount or 0), 2),
            'date': e.expense_date.strftime('%d %b %Y') if e.expense_date else '',
            'url': url_for('finance.expenses_list')} for e in exps]

        charges = (scope_query(AdditionalCharge.query, AdditionalCharge)
                   .filter(db.or_(AdditionalCharge.category.ilike(pat, escape=esc),
                                  AdditionalCharge.description.ilike(pat, escape=esc)))
                   .order_by(AdditionalCharge.id.desc()).limit(15).all())
        cmap = {s.id: s.full_name for s in Student.query.filter(
            Student.id.in_([c.student_id for c in charges])).all()} if charges else {}
        results['charges'] = [{
            'kind': c.kind, 'category': c.category or '—',
            'description': c.description or '', 'student': cmap.get(c.student_id, '—'),
            'amount': round(float(c.amount or 0), 2),
            'url': url_for('finance.billing_tools', term_id=c.term_id)} for c in charges]

        total = sum(len(v) for v in results.values())
    return render_template('finance/search.html', q=q, results=results,
                           total=total, nav=_nav_urls())


# ============================================================================
# ACCOUNTING — double-entry statements derived from the ledger (Phase 2)
# ============================================================================
@finance_bp.route('/accounting')
@login_required
def accounting():
    """Trial Balance, Income Statement (P&L), Balance Sheet, Cash Flow and Chart
    of Accounts — all auto-derived from the ledger, filterable like the overview."""
    from utils import finance_accounting as A
    from utils.branch_scope import is_central
    filters = _overview_filters()
    return render_template('finance/accounting.html',
                           s=A.all_statements(filters), sel=request.args,
                           is_central=is_central(), opts=_accounting_options(),
                           selected_branches=set(filters['branch_ids']),
                           nav=_nav_urls(),
                           export_url=url_for('finance.accounting_export'))


def _accounting_options():
    from utils import finance_ledger as L
    return L.filter_options()


@finance_bp.route('/accounting/export')
@login_required
def accounting_export():
    """Every statement as an Excel workbook (one sheet each) — auditor-ready."""
    from utils import finance_accounting as A
    from openpyxl import Workbook
    filters = _overview_filters()
    s = A.all_statements(filters)
    wb = Workbook()

    tb = wb.active; tb.title = 'Trial Balance'
    tb.append(['Code', 'Account', 'Type', 'Debit', 'Credit'])
    for a in s['trial_balance']['accounts']:
        tb.append([a['code'], a['name'], a['type'], a['debit'], a['credit']])
    tb.append(['', 'TOTAL', '', s['trial_balance']['total_debit'], s['trial_balance']['total_credit']])

    inc = s['income_statement']
    pl = wb.create_sheet('Income Statement')
    pl.append(['Income', 'Amount'])
    for r in inc['income']:
        pl.append([r['name'], r['amount']])
    pl.append(['Total Income', inc['total_income']])
    if inc.get('cogs'):
        pl.append([]); pl.append(['Cost of Goods Sold', inc['cogs']])
        pl.append(['Gross Profit', inc['gross_profit']])
    pl.append([]); pl.append(['Operating Expenses', 'Amount'])
    for r in inc.get('operating_expense', inc['expense']):
        pl.append([r['name'], r['amount']])
    pl.append(['Total Expenses', inc['total_expense']])
    pl.append([]); pl.append(['Net Surplus / (Deficit)', inc['net']])

    bs = wb.create_sheet('Balance Sheet')
    bs.append(['Assets', 'Amount'])
    for r in s['balance_sheet']['assets']:
        bs.append([r['name'], r['amount']])
    bs.append(['Total Assets', s['balance_sheet']['total_assets']])
    bs.append([]); bs.append(['Equity', 'Amount'])
    for r in s['balance_sheet']['equity']:
        bs.append([r['name'], r['amount']])
    bs.append(['Total Equity', s['balance_sheet']['total_equity']])

    cf = wb.create_sheet('Cash Flow')
    cf.append(['Inflows', 'Amount'])
    for r in s['cash_flow']['inflows']:
        cf.append([r['name'], r['amount']])
    cf.append(['Total In', s['cash_flow']['total_in']])
    cf.append([]); cf.append(['Outflows', 'Amount'])
    for r in s['cash_flow']['outflows']:
        cf.append([r['name'], r['amount']])
    cf.append(['Total Out', s['cash_flow']['total_out']])
    cf.append([]); cf.append(['Net Cash Movement', s['cash_flow']['net_change']])

    return xlsx_response(wb, 'financial_statements.xlsx')


# ============================================================================
# BILLING TOOLS — additional charges, bulk billing, penalties, credit notes
# (Phase 3). Server-rendered; results flow straight into student_bill.
# ============================================================================
def _enrolled_students(term_id, class_id=None):
    """Active students enrolled in a term (optionally one class), branch-scoped,
    with their branch — for bulk billing."""
    from utils.branch_scope import scope_query
    q = scope_query(
        StudentEnrollment.query
        .join(ClassArmAssignment, StudentEnrollment.class_arm_assignment_id == ClassArmAssignment.id)
        .filter(StudentEnrollment.is_active == True, ClassArmAssignment.term_id == term_id),
        ClassArmAssignment)
    if class_id:
        q = q.filter(ClassArmAssignment.class_id == class_id)
    out = []
    for e in q.all():
        if e.student and e.student.is_active:
            out.append((e.student, e.class_arm_assignment))
    return out


def _who():
    from utils.access_control import get_current_user
    return getattr(get_current_user(), 'full_name', None) or session.get('user') or 'staff'


@finance_bp.route('/billing-tools')
@login_required
def billing_tools():
    """Bulk billing, late-payment penalties and credit notes — the advanced
    billing actions, kept off the everyday screens."""
    term_id = request.args.get('term_id', type=int) or _active_term_id()
    terms = session_terms()
    classes = SchoolClass.query.filter_by(is_active=True).order_by(SchoolClass.level).all()
    recent = (AdditionalCharge.query.filter_by(term_id=term_id)
              .order_by(AdditionalCharge.id.desc()).limit(50).all() if term_id else [])
    charge_rows = [{
        'id': c.id, 'student': (c.student.full_name if c.student else '—'),
        'kind': c.kind, 'category': c.category or '', 'description': c.description or '',
        'amount': c.amount, 'by': c.created_by or '',
        'date': c.created_at.strftime('%d %b %Y') if c.created_at else '',
    } for c in recent]
    return render_template('finance/billing_tools.html',
                           term_id=term_id, sel=request.args,
                           terms=[{'id': t.id, 'full_name': t.full_name} for t in terms],
                           classes=[{'id': c.id, 'name': c.name} for c in classes],
                           charges=charge_rows, nav=_nav_urls())


@finance_bp.route('/billing-tools/bulk-charge', methods=['POST'])
@login_required
def bulk_charge():
    """Add one extra charge (or credit) to every enrolled student in a term/class."""
    term_id = request.form.get('term_id', type=int)
    class_id = request.form.get('class_id', type=int)
    kind = 'credit' if request.form.get('kind') == 'credit' else 'charge'
    amount = request.form.get('amount', type=float) or 0
    category = (request.form.get('category') or ('Credit Note' if kind == 'credit' else 'Additional Charge')).strip()
    description = (request.form.get('description') or category).strip()
    if not term_id or amount <= 0:
        return _err('Pick a term and a positive amount.', url_for('finance.billing_tools', term_id=term_id))
    students = _enrolled_students(term_id, class_id)
    for student, caa in students:
        db.session.add(AdditionalCharge(student_id=student.id, term_id=term_id,
                                        branch_id=caa.branch_id, kind=kind, category=category,
                                        description=description, amount=amount, created_by=_who()))
    db.session.commit()
    from utils.audit import log_action
    log_action('finance.bulk_charge', detail=f'{kind} ₦{amount:g} x{len(students)} — {description}')
    verb = 'Credit note' if kind == 'credit' else 'Charge'
    return _ok(f'{verb} of ₦{amount:,.0f} applied to {len(students)} student(s).',
               url_for('finance.billing_tools', term_id=term_id))


@finance_bp.route('/billing-tools/penalties', methods=['POST'])
@login_required
def apply_penalties():
    """Apply a late-payment penalty to students who still owe for the term. Fixed
    amount or a percentage of the outstanding balance. Skips students who already
    have a penalty this term, so a re-run never double-charges."""
    term_id = request.form.get('term_id', type=int)
    if not term_id:
        return _err('Pick a term.', url_for('finance.billing_tools'))
    ptype = request.form.get('ptype')                       # 'fixed' | 'percent'
    value = request.form.get('value', type=float) or 0
    if value <= 0:
        return _err('Enter a penalty amount / percentage.', url_for('finance.billing_tools', term_id=term_id))
    existing = {c.student_id for c in AdditionalCharge.query
                .filter_by(term_id=term_id, category='Penalty').all()}
    applied = 0
    for student, caa in _enrolled_students(term_id):
        if student.id in existing:
            continue
        if ptype == 'installment':
            # penalise only the overdue installment shortfall
            from utils import finance_installments as I
            stt = I.student_status(student.id, term_id)
            if not stt['has_plan'] or stt['behind'] <= 0.005:
                continue
            amount = value                             # flat penalty per behind student
        else:
            bill = student_bill(student.id, term_id)
            if bill['balance'] <= 0.005:
                continue
            amount = round(bill['balance'] * value / 100.0, 2) if ptype == 'percent' else value
        if amount <= 0:
            continue
        db.session.add(AdditionalCharge(student_id=student.id, term_id=term_id,
                                        branch_id=caa.branch_id, kind='charge', category='Penalty',
                                        description=('Late payment penalty (%.0f%%)' % value) if ptype == 'percent'
                                        else 'Late payment penalty', amount=amount, created_by=_who()))
        applied += 1
    db.session.commit()
    from utils.audit import log_action
    log_action('finance.penalties', detail=f'{ptype} {value} -> {applied} student(s), term {term_id}')
    return _ok(f'Penalty applied to {applied} student(s) with an outstanding balance.',
               url_for('finance.billing_tools', term_id=term_id))


@finance_bp.route('/billing-tools/charge/<int:charge_id>/remove', methods=['POST'])
@login_required
def remove_charge(charge_id):
    c = db.get_or_404(AdditionalCharge, charge_id)
    require_branch_access(c.branch_id)
    term_id = c.term_id
    from utils.audit import log_action
    log_action('finance.charge_removed',
               detail=f'{c.kind} ₦{c.amount:g} — {c.description}',
               target_type='additionalcharge', target_id=c.id)
    db.session.delete(c)
    db.session.commit()
    return _ok('Removed.', url_for('finance.billing_tools', term_id=term_id))


# ============================================================================
# INSTALLMENT PLANS — a term's payment schedule (Phase 3)
# ============================================================================
@finance_bp.route('/installments')
@login_required
def installments():
    """Set a term's installment schedule (a class-specific one overrides the
    term-wide one) and see which students are on track or behind."""
    from utils import finance_installments as I
    term_id = request.args.get('term_id', type=int) or _active_term_id()
    class_id = request.args.get('class_id', type=int)
    terms = session_terms()
    classes = SchoolClass.query.filter_by(is_active=True).order_by(SchoolClass.level).all()

    plan = I.get_plan(term_id, class_id) if term_id else []
    plan_rows = [{'label': r.label, 'due': r.due_date.strftime('%Y-%m-%d') if r.due_date else '',
                  'percent': r.percent} for r in plan]
    total_pct = round(sum(r.percent for r in plan), 2)

    roster = []
    if term_id and plan:
        for r in I.roster(term_id, _enrolled_students(term_id, class_id)):
            st = r['status']
            roster.append({
                'name': r['student'].full_name, 'sid': r['student'].student_id,
                'payable': st['payable'], 'paid': st['paid'],
                'expected': st['expected_to_date'], 'behind': st['behind'],
                'on_track': st['on_track'],
                'next_due': (st['next_due']['due_date'].strftime('%d %b') if st['next_due'] and st['next_due']['due_date'] else '—'),
            })
    return render_template('finance/installments.html',
                           term_id=term_id, class_id=class_id or '', sel=request.args,
                           terms=[{'id': t.id, 'full_name': t.full_name} for t in terms],
                           classes=[{'id': c.id, 'name': c.name} for c in classes],
                           plan_rows=plan_rows or [{'label': '', 'due': '', 'percent': ''}],
                           total_pct=total_pct, roster=roster, nav=_nav_urls())


@finance_bp.route('/installments/save', methods=['POST'])
@login_required
def installments_save():
    from utils import finance_installments as I
    from utils.branch_scope import branch_for_new
    term_id = request.form.get('term_id', type=int)
    class_id = request.form.get('class_id', type=int)
    if not term_id:
        return _err('Pick a term.', url_for('finance.installments'))
    labels = request.form.getlist('label')
    dues = request.form.getlist('due')
    pcts = request.form.getlist('percent')
    rows = []
    for i, label in enumerate(labels):
        try:
            pct = float(pcts[i]) if i < len(pcts) and pcts[i] else 0.0
        except ValueError:
            pct = 0.0
        due = _overview_date(dues[i]) if i < len(dues) else None
        rows.append({'label': label, 'due_date': due, 'percent': pct})
    total = sum(r['percent'] for r in rows if (r['label'] or '').strip() and r['percent'] > 0)
    if total > 100.5:
        return _err(f'Installments add up to {total:g}% — must not exceed 100%.',
                    url_for('finance.installments', term_id=term_id, class_id=class_id or ''))
    I.save_plan(term_id, class_id, rows, branch_id=branch_for_new())
    from utils.audit import log_action
    log_action('finance.installment_plan', detail=f'term {term_id} class {class_id or "all"}: {total:g}%')
    return _ok('Installment schedule saved.',
               url_for('finance.installments', term_id=term_id, class_id=class_id or ''))


@finance_bp.route('/installments/clear', methods=['POST'])
@login_required
def installments_clear():
    from utils import finance_installments as I
    term_id = request.form.get('term_id', type=int)
    class_id = request.form.get('class_id', type=int)
    if term_id:
        I.clear_plan(term_id, class_id)
    return _ok('Installment schedule cleared.',
               url_for('finance.installments', term_id=term_id, class_id=class_id or ''))


@finance_bp.route('/installments/notify-overdue', methods=['POST'])
@login_required
def notify_overdue():
    """Alert admins in-app about students behind on fees for the term."""
    from utils import finance_notify
    term_id = request.form.get('term_id', type=int) or _active_term_id()
    s = finance_notify.run_fee_reminders(
        term_id, url=url_for('finance.defaulters', term_id=term_id) if term_id else None)
    if s['count']:
        return _ok(f"Alerted admins — {s['count']} student(s) behind, ₦{s['total']:,.0f} outstanding.",
                   url_for('finance.installments', term_id=term_id))
    return _ok('No students are behind — nothing to alert.',
               url_for('finance.installments', term_id=term_id))


@finance_bp.route('/reminders/draft', methods=['POST'])
@login_required
def draft_reminders():
    """Queue an SMS fee reminder to every defaulter's parent for the term as a
    Draft campaign, then hand off to Communication for review + send. Nothing is
    sent to parents here — a human confirms in the campaign editor."""
    from utils import finance_notify
    term_id = request.form.get('term_id', type=int) or _active_term_id()
    class_id = request.form.get('class_id', type=int) or None
    channel = 'Email' if (request.form.get('channel') or '').lower() == 'email' else 'SMS'
    term = db.session.get(Term, term_id) if term_id else None
    if not term:
        return _err('Pick a term first.', url_for('finance.installments'))
    msg = finance_notify.draft_parent_reminders(term, channel=channel,
                                                class_id=class_id, created_by=_who())
    if not msg:
        reach = 'email address' if channel == 'Email' else 'phone number'
        return _ok(f'No defaulters with a parent {reach} — nothing to draft.',
                   url_for('finance.installments', term_id=term_id, class_id=class_id or ''))
    from utils.audit import log_action
    log_action('finance.fee_reminders_drafted',
               detail=f'term {term_id} via {channel}: {msg.recipient_count} parent(s)')
    return _ok(f'Drafted {channel} reminders for {msg.recipient_count} parent(s) — review and send.',
               url_for('comms.message_detail', message_id=msg.id))
