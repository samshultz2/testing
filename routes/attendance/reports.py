"""attendance blueprint — reports routes (split from the former routes/attendance.py)."""
from routes.attendance import *  # noqa: F401,F403


@attendance_bp.route('/daily')
@login_required
def daily_summary():
    """View daily attendance summary"""
    assignment_id = request.args.get('assignment_id', type=int)
    target_date = request.args.get('date')
    
    if target_date:
        target_date = datetime.strptime(target_date, '%Y-%m-%d').date()
    else:
        target_date = date.today()
    
    # Get assignments for dropdown
    active_term = get_active_term()
    assignments = []
    if active_term:
        assignments = filter_classes_for_user(
            ClassArmAssignment.query.filter_by(term_id=active_term.id).all(),
            form_only=True)

    # A form teacher lands on their own class without picking it.
    if not assignment_id:
        assignment_id = auto_select_assignment(assignments)

    summary = None
    selected_assignment = None
    week = None
    week_days = []   # per-day breakdown for the whole school week, at a glance

    if assignment_id:
        if not can_view_attendance(assignment_id):   # teachers: their form class only
            flash('You do not have access to this class.', 'error')
            return redirect(url_for('attendance.daily_summary'))
        selected_assignment = db.session.get(ClassArmAssignment, assignment_id)
        summary = get_daily_attendance_summary(assignment_id, target_date)

        if active_term:
            weeks = Week.query.filter_by(term_id=active_term.id).order_by(Week.week_number).all()
            week = pick_current_week(weeks, on=target_date)
            if week:
                holiday_dates = {h.date for h in Holiday.query.filter_by(term_id=active_term.id).all()}
                today = date.today()
                d = week.start_date
                while d <= week.end_date:
                    if d.weekday() < 5 and d not in holiday_dates and d <= today:
                        s = get_daily_attendance_summary(assignment_id, d)
                        week_days.append({'date': d, 'summary': s})
                    d += timedelta(days=1)

    return render_template('attendance/daily.html',
        assignments=assignments,
        selected_assignment=selected_assignment,
        target_date=target_date,
        summary=summary,
        week=week,
        week_days=week_days,
    )


@attendance_bp.route('/weekly')
@login_required
def weekly_summary():
    """View weekly attendance summary"""
    assignment_id = request.args.get('assignment_id', type=int)
    week_id = request.args.get('week_id', type=int)
    
    # Check class access
    if assignment_id and not can_view_attendance(assignment_id):
        flash('You do not have access to this class.', 'error')
        return redirect(url_for('attendance.weekly_summary'))
    
    # Get active term
    active_term = get_active_term()
    
    assignments = []
    weeks = []
    if active_term:
        all_assignments = ClassArmAssignment.query.filter_by(term_id=active_term.id).all()
        assignments = filter_classes_for_user(all_assignments, form_only=True)
        weeks = Week.query.filter_by(term_id=active_term.id).order_by(Week.week_number).all()

    # Default to the form teacher's class and the current week.
    if not assignment_id:
        assignment_id = auto_select_assignment(assignments)
    if not week_id:
        cw = pick_current_week(weeks)
        week_id = cw.id if cw else None

    summary = None
    selected_assignment = None
    selected_week = None

    if assignment_id and week_id:
        selected_assignment = db.session.get(ClassArmAssignment, assignment_id)
        selected_week = db.session.get(Week, week_id)
        summary = get_weekly_attendance_summary(assignment_id, week_id)
    
    return render_template('attendance/weekly.html',
        assignments=assignments,
        weeks=weeks,
        selected_assignment=selected_assignment,
        selected_week=selected_week,
        summary=summary
    )


@attendance_bp.route('/weekly/export')
@login_required
@rate_limited('export', max_requests=40, window_minutes=10)
def export_weekly():
    """Export weekly attendance to Excel"""
    assignment_id = request.args.get('assignment_id', type=int)
    week_id = request.args.get('week_id', type=int)
    
    if not assignment_id or not week_id:
        flash('Please select a class and week.', 'error')
        return redirect(url_for('attendance.weekly_summary'))

    assignment = db.get_or_404(ClassArmAssignment, assignment_id)
    require_branch_access(assignment.branch_id)
    if not can_view_attendance(assignment_id):   # teachers: their form class only
        abort(403)
    week = db.get_or_404(Week, week_id)

    summary = get_weekly_attendance_summary(assignment_id, week_id)
    
    excel_file = export_attendance_to_excel(
        summary,
        assignment.display_name,
        {
            'week_number': week.week_number,
            'start_date': week.start_date.strftime('%Y-%m-%d'),
            'end_date': week.end_date.strftime('%Y-%m-%d')
        }
    )
    
    filename = f"attendance_{assignment.display_name.replace(' ', '_')}_week{week.week_number}.xlsx"

    return xlsx_response(excel_file, filename)


@attendance_bp.route('/termly')
@login_required
def termly_summary():
    """View termly attendance summary"""
    assignment_id = request.args.get('assignment_id', type=int)
    term_id = request.args.get('term_id', type=int)
    
    # Check class access
    if assignment_id and not can_view_attendance(assignment_id):
        flash('You do not have access to this class.', 'error')
        return redirect(url_for('attendance.termly_summary'))
    
    # Get terms
    terms = Term.query.join(AcademicSession).order_by(
        AcademicSession.name.desc(),
        Term.term_number
    ).all()
    
    if not term_id:
        active_term = get_active_term()
        if active_term:
            term_id = active_term.id
    
    assignments = []
    selected_term = None
    if term_id:
        selected_term = db.session.get(Term, term_id)
        all_assignments = ClassArmAssignment.query.filter_by(term_id=term_id).all()
        assignments = filter_classes_for_user(all_assignments, form_only=True)
    
    summary = None
    selected_assignment = None
    
    if assignment_id and term_id:
        selected_assignment = db.session.get(ClassArmAssignment, assignment_id)
        summary = get_termly_attendance_summary(assignment_id, term_id)
    
    return render_template('attendance/termly.html',
        terms=terms,
        assignments=assignments,
        selected_term=selected_term,
        selected_assignment=selected_assignment,
        summary=summary
    )


@attendance_bp.route('/termly/export')
@login_required
@rate_limited('export', max_requests=40, window_minutes=10)
def export_termly():
    """Export termly attendance to Excel"""
    assignment_id = request.args.get('assignment_id', type=int)
    term_id = request.args.get('term_id', type=int)
    
    if not assignment_id or not term_id:
        flash('Please select a class and term.', 'error')
        return redirect(url_for('attendance.termly_summary'))

    assignment = db.get_or_404(ClassArmAssignment, assignment_id)
    require_branch_access(assignment.branch_id)
    if not can_view_attendance(assignment_id):   # teachers: their form class only
        abort(403)
    term = db.get_or_404(Term, term_id)

    summary = get_termly_attendance_summary(assignment_id, term_id)
    
    if not summary:
        flash('No attendance data found.', 'error')
        return redirect(url_for('attendance.termly_summary'))
    
    # Create Excel file
    import openpyxl
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    from io import BytesIO
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Termly Attendance"
    
    # Styles
    header_font = Font(bold=True, size=14)
    subheader_font = Font(bold=True, size=11)
    center_align = Alignment(horizontal='center', vertical='center')
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    green_fill = PatternFill(start_color='90EE90', end_color='90EE90', fill_type='solid')
    yellow_fill = PatternFill(start_color='FFFF99', end_color='FFFF99', fill_type='solid')
    red_fill = PatternFill(start_color='FFB6C1', end_color='FFB6C1', fill_type='solid')
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    header_font_white = Font(bold=True, size=11, color='FFFFFF')
    
    term_info = summary.get('term_info', {})
    total_school_days = term_info.get('total_school_days', 0)
    total_times_opened = term_info.get('total_times_opened', 0)
    class_totals = summary.get('class_totals', {})
    
    # Title
    ws.merge_cells('A1:I1')
    ws['A1'] = f"TERMLY ATTENDANCE REPORT - {assignment.display_name}"
    ws['A1'].font = header_font
    ws['A1'].alignment = center_align
    
    ws.merge_cells('A2:I2')
    ws['A2'] = f"{term.session.name} - Term {term.term_number}"
    ws['A2'].font = subheader_font
    ws['A2'].alignment = center_align
    
    # Term info row
    ws.merge_cells('A3:I3')
    ws['A3'] = f"Total School Days: {total_school_days} | Times Opened (AM+PM): {total_times_opened} | Total Students: {summary.get('total_students', 0)}"
    ws['A3'].alignment = center_align
    
    # Headers
    headers = ['S/N', 'Student Name', 'Gender', 'AM Present', 'PM Present', 'Total Present', 'Times Opened', 'Attendance %', 'Status']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=5, column=col, value=header)
        cell.font = header_font_white
        cell.alignment = center_align
        cell.border = thin_border
        cell.fill = header_fill
    
    # Data rows
    row = 6
    for idx, student in enumerate(summary.get('students', []), 1):
        morning = student.get('morning_total', 0)
        afternoon = student.get('afternoon_total', 0)
        total_present = student.get('termly_total', morning + afternoon)
        percentage = student.get('percentage', 0)
        
        ws.cell(row=row, column=1, value=idx).border = thin_border
        ws.cell(row=row, column=1).alignment = center_align
        
        ws.cell(row=row, column=2, value=student.get('student_name', '')).border = thin_border
        
        ws.cell(row=row, column=3, value=student.get('gender', '')).border = thin_border
        ws.cell(row=row, column=3).alignment = center_align
        
        ws.cell(row=row, column=4, value=morning).border = thin_border
        ws.cell(row=row, column=4).alignment = center_align
        
        ws.cell(row=row, column=5, value=afternoon).border = thin_border
        ws.cell(row=row, column=5).alignment = center_align
        
        ws.cell(row=row, column=6, value=total_present).border = thin_border
        ws.cell(row=row, column=6).alignment = center_align
        
        ws.cell(row=row, column=7, value=total_times_opened).border = thin_border
        ws.cell(row=row, column=7).alignment = center_align
        
        pct_cell = ws.cell(row=row, column=8, value=f"{percentage:.1f}%")
        pct_cell.border = thin_border
        pct_cell.alignment = center_align
        
        # Status based on percentage
        if percentage >= 75:
            status = "Good"
            status_fill = green_fill
        elif percentage >= 50:
            status = "Fair"
            status_fill = yellow_fill
        else:
            status = "At Risk"
            status_fill = red_fill
        
        status_cell = ws.cell(row=row, column=9, value=status)
        status_cell.border = thin_border
        status_cell.alignment = center_align
        status_cell.fill = status_fill
        
        row += 1
    
    # Summary rows
    row += 1
    ws.merge_cells(f'A{row}:C{row}')
    ws.cell(row=row, column=1, value="CLASS TOTALS").font = subheader_font
    ws.cell(row=row, column=4, value=class_totals.get('total_morning', 0)).font = subheader_font
    ws.cell(row=row, column=4).alignment = center_align
    ws.cell(row=row, column=5, value=class_totals.get('total_afternoon', 0)).font = subheader_font
    ws.cell(row=row, column=5).alignment = center_align
    ws.cell(row=row, column=6, value=class_totals.get('total_attendance', 0)).font = subheader_font
    ws.cell(row=row, column=6).alignment = center_align
    
    row += 1
    ws.merge_cells(f'A{row}:C{row}')
    ws.cell(row=row, column=1, value="CLASS AVERAGE").font = subheader_font
    ws.cell(row=row, column=8, value=f"{class_totals.get('termly_percentage', 0):.1f}%").font = subheader_font
    ws.cell(row=row, column=8).alignment = center_align
    
    # Gender breakdown
    row += 2
    ws.cell(row=row, column=1, value="GENDER BREAKDOWN:").font = subheader_font
    row += 1
    ws.cell(row=row, column=1, value=f"Male Students: {summary.get('total_male_students', 0)}")
    ws.cell(row=row, column=3, value=f"Male Attendance: {class_totals.get('male_attendance', 0)}")
    row += 1
    ws.cell(row=row, column=1, value=f"Female Students: {summary.get('total_female_students', 0)}")
    ws.cell(row=row, column=3, value=f"Female Attendance: {class_totals.get('female_attendance', 0)}")
    
    # Column widths
    ws.column_dimensions['A'].width = 6
    ws.column_dimensions['B'].width = 28
    ws.column_dimensions['C'].width = 10
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['F'].width = 14
    ws.column_dimensions['G'].width = 14
    ws.column_dimensions['H'].width = 14
    ws.column_dimensions['I'].width = 12
    
    filename = f"attendance_{assignment.display_name.replace(' ', '_')}_term{term.term_number}.xlsx"

    return xlsx_response(wb, filename)


@attendance_bp.route('/alerts')
@login_required
def attendance_alerts():
    """View students with poor attendance"""
    
    term_id = request.args.get('term_id', type=int)
    threshold = request.args.get('threshold', type=float)
    
    terms = Term.query.order_by(Term.id.desc()).all()
    
    # Get active term if not specified
    if not term_id:
        active_term = get_active_term()
        if active_term:
            term_id = active_term.id
    
    selected_term = db.session.get(Term, term_id) if term_id else None
    
    # Default threshold from settings or 75%
    if threshold is None:
        threshold = 75.0
    
    alerts = []
    
    if selected_term:
        # Only enrolments in classes this user may access (branch + role scoped).
        accessible = filter_classes_for_user(
            ClassArmAssignment.query.filter_by(term_id=term_id).all(), form_only=True)
        class_ids = [c.id for c in accessible]
        enrollments = StudentEnrollment.query.filter(
            StudentEnrollment.class_arm_assignment_id.in_(class_ids or [-1]),
            StudentEnrollment.is_active == True
        ).all() if class_ids else []
        
        # Get weeks for this term
        weeks = Week.query.filter_by(term_id=term_id).all()
        week_ids = [w.id for w in weeks]
        
        if week_ids:
            # Get holidays
            holidays = Holiday.query.filter_by(term_id=term_id).all()
            holiday_dates = set(h.date for h in holidays)
            
            # Calculate total times opened
            total_school_days = 0
            for week in weeks:
                current = week.start_date
                while current <= week.end_date:
                    if current.weekday() < 5 and current not in holiday_dates:
                        total_school_days += 1
                    current += timedelta(days=1)
            
            total_times_opened = total_school_days * 2  # Morning + Afternoon
            
            if total_times_opened > 0:
                for enrollment in enrollments:
                    # Count attendance
                    attendance_records = Attendance.query.filter(
                        Attendance.enrollment_id == enrollment.id,
                        Attendance.week_id.in_(week_ids)
                    ).all()
                    
                    total_present = sum(
                        (1 if a.morning_present else 0) + (1 if a.afternoon_present else 0)
                        for a in attendance_records
                    )
                    
                    percentage = round((total_present / total_times_opened) * 100, 2)
                    
                    if percentage < threshold:
                        alerts.append({
                            'student': enrollment.student,
                            'class': enrollment.class_arm_assignment.display_name,
                            'present': total_present,
                            'total': total_times_opened,
                            'percentage': percentage,
                            'status': 'critical' if percentage < 50 else 'warning'
                        })
        
        # Sort by percentage (lowest first)
        alerts.sort(key=lambda x: x['percentage'])
    
    return render_template('attendance/alerts.html',
        terms=terms, term_id=term_id, selected_term=selected_term,
        threshold=threshold, alerts=alerts
    )


@attendance_bp.route('/alerts/export')
@login_required
def export_alerts():
    """Export attendance alerts to Excel"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    import io
    
    term_id = request.args.get('term_id', type=int)
    threshold = request.args.get('threshold', type=float) or 75.0
    
    if not term_id:
        flash('Select a term first.', 'error')
        return redirect(url_for('attendance.attendance_alerts'))
    
    selected_term = db.session.get(Term, term_id)

    # Get alerts (same logic as the page) — scoped to the user's accessible classes
    accessible = filter_classes_for_user(
        ClassArmAssignment.query.filter_by(term_id=term_id).all(), form_only=True)
    class_ids = [c.id for c in accessible]
    enrollments = StudentEnrollment.query.filter(
        StudentEnrollment.class_arm_assignment_id.in_(class_ids or [-1]),
        StudentEnrollment.is_active == True
    ).all() if class_ids else []
    
    weeks = Week.query.filter_by(term_id=term_id).all()
    week_ids = [w.id for w in weeks]
    
    holidays = Holiday.query.filter_by(term_id=term_id).all()
    holiday_dates = set(h.date for h in holidays)
    
    total_school_days = 0
    for week in weeks:
        current = week.start_date
        while current <= week.end_date:
            if current.weekday() < 5 and current not in holiday_dates:
                total_school_days += 1
            current += timedelta(days=1)
    
    total_times_opened = total_school_days * 2
    
    alerts = []
    if total_times_opened > 0:
        for enrollment in enrollments:
            attendance_records = Attendance.query.filter(
                Attendance.enrollment_id == enrollment.id,
                Attendance.week_id.in_(week_ids)
            ).all()
            
            total_present = sum(
                (1 if a.morning_present else 0) + (1 if a.afternoon_present else 0)
                for a in attendance_records
            )
            
            percentage = round((total_present / total_times_opened) * 100, 2)
            
            if percentage < threshold:
                alerts.append({
                    'student': enrollment.student,
                    'class': enrollment.class_arm_assignment.display_name,
                    'present': total_present,
                    'total': total_times_opened,
                    'percentage': percentage
                })
    
    alerts.sort(key=lambda x: x['percentage'])
    
    # Create Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "Attendance Alerts"
    
    # Title
    ws['A1'] = f"Attendance Alerts - {selected_term.full_name}"
    ws['A1'].font = Font(bold=True, size=14)
    ws['A2'] = f"Students below {threshold}% attendance"
    
    # Headers
    headers = ['S/N', 'Student Name', 'Student ID', 'Class', 'Present', 'Total', 'Percentage']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col, value=header)
        cell.font = Font(bold=True)
    
    # Data
    for idx, alert in enumerate(alerts, 1):
        row = idx + 4
        ws.cell(row=row, column=1, value=idx)
        ws.cell(row=row, column=2, value=alert['student'].full_name)
        ws.cell(row=row, column=3, value=alert['student'].student_id)
        ws.cell(row=row, column=4, value=alert['class'])
        ws.cell(row=row, column=5, value=alert['present'])
        ws.cell(row=row, column=6, value=alert['total'])
        ws.cell(row=row, column=7, value=f"{alert['percentage']}%")
        
        # Highlight critical
        if alert['percentage'] < 50:
            for col in range(1, 8):
                ws.cell(row=row, column=col).fill = PatternFill(start_color='FFCDD2', fill_type='solid')
    
    return xlsx_response(wb, f'attendance_alerts_{selected_term.name}.xlsx')


@attendance_bp.route('/print-register')
@login_required
def print_register():
    """Print-ready attendance register"""
    from models import SchoolSettings
    
    term_id = request.args.get('term_id', type=int)
    assignment_id = request.args.get('assignment_id', type=int)
    week_id = request.args.get('week_id', type=int)
    
    terms = Term.query.order_by(Term.id.desc()).all()
    
    # Get active term if not specified
    if not term_id:
        active_term = get_active_term()
        if active_term:
            term_id = active_term.id
    
    selected_term = db.session.get(Term, term_id) if term_id else None
    
    # Get assignments for term (only those this user may access)
    assignments = []
    if term_id:
        assignments = filter_classes_for_user(
            ClassArmAssignment.query.filter_by(term_id=term_id).all(), form_only=True)

    if assignment_id and not can_view_attendance(assignment_id):
        flash('You do not have access to this class.', 'error')
        return redirect(url_for('attendance.print_register'))
    selected_assignment = db.session.get(ClassArmAssignment, assignment_id) if assignment_id else None
    
    # Get weeks for term
    weeks = []
    if term_id:
        weeks = Week.query.filter_by(term_id=term_id).order_by(Week.week_number).all()
    
    selected_week = db.session.get(Week, week_id) if week_id else None
    
    register_data = []
    school_days = []
    school_name = SchoolSettings.get('school_name', 'School Name')
    
    if selected_assignment and selected_week:
        # Get holidays
        holidays = Holiday.query.filter_by(term_id=term_id).all()
        holiday_dates = set(h.date for h in holidays)
        
        # Calculate school days in the week
        current = selected_week.start_date
        while current <= selected_week.end_date:
            if current.weekday() < 5 and current not in holiday_dates:
                school_days.append(current)
            current += timedelta(days=1)
        
        # Get enrolled students
        enrollments = StudentEnrollment.query.filter_by(
            class_arm_assignment_id=assignment_id,
            is_active=True
        ).join(Student).order_by(*roster_order()).all()
        
        for enrollment in enrollments:
            # Get attendance for the week
            attendance_records = Attendance.query.filter_by(
                enrollment_id=enrollment.id,
                week_id=week_id
            ).all()
            
            # Build attendance lookup by date
            attendance_lookup = {}
            total_present = 0
            
            for record in attendance_records:
                attendance_lookup[record.date.isoformat()] = {
                    'morning_present': record.morning_present,
                    'afternoon_present': record.afternoon_present
                }
                if record.morning_present:
                    total_present += 1
                if record.afternoon_present:
                    total_present += 1
            
            register_data.append({
                'student': enrollment.student,
                'attendance': attendance_lookup,
                'total_present': total_present
            })
    
    return render_template('attendance/print_register.html',
        terms=terms, term_id=term_id, selected_term=selected_term,
        assignments=assignments, assignment_id=assignment_id, selected_assignment=selected_assignment,
        weeks=weeks, week_id=week_id, selected_week=selected_week,
        register_data=register_data, school_days=school_days, school_name=school_name
    )


@attendance_bp.route('/analytics/export')
@login_required
def analytics_export():
    """Presentation-ready attendance analytics as a multi-section Excel workbook
    or a flattened CSV (executive summary, class ranking, distribution, chronic
    absentees, most improved). Branch/role-scoped like the analytics screen."""
    if not can_access_module('attendance'):
        abort(403)
    req_term = request.args.get('term_id', type=int)
    term = (db.session.get(Term, req_term) if req_term else None) or get_active_term()
    if not term:
        abort(400)
    accessible = filter_classes_for_user(
        ClassArmAssignment.query.filter_by(term_id=term.id).all(), form_only=True)
    from utils.branch_scope import is_central
    from utils import attendance_analytics as AA
    d = AA.build(term, accessible, is_central=is_central(), use_cache=False)
    fmt = request.args.get('format', 'xlsx')
    fname = f'attendance_analytics_{term.name}'.replace(' ', '_')
    from utils.audit import log_action
    log_action('attendance.analytics_export', detail=f'{term.name} ({fmt})')

    k = d['kpis']
    summary_rows = [
        ('Term', term.name), ('Overall attendance rate', f"{k['overall']}%"),
        ('Students', k['students']), ('Classes', k['classes']),
        ('School days', k['school_days']),
        (f"Chronic (< {d['critical']}%)", k['chronic']),
        ('Best class', k['best_class']), ('Needs attention', k['worst_class']),
    ]

    if fmt == 'csv':
        import csv, io
        from utils.web_exports import formula_guard as _fg
        out = io.StringIO(); w = csv.writer(out)
        w.writerow(['Attendance analytics', term.name])
        w.writerow([])
        w.writerow(['Executive summary'])
        for label, val in summary_rows:
            w.writerow([label, val])
        w.writerow([]); w.writerow(['Class ranking', 'Students', 'Attendance %'])
        for r in d['class_rank']:
            w.writerow([_fg(r['class']), r['students'], r['percentage']])
        w.writerow([]); w.writerow(['Distribution', 'Students'])
        for label, key in (('Excellent (>=90%)', 'excellent'), ('Good (75-89%)', 'good'),
                           ('Fair (50-74%)', 'fair'), ('Poor (<50%)', 'poor')):
            w.writerow([label, d['distribution'][key]])
        w.writerow([]); w.writerow(['Chronic absentees', 'Class', 'Attendance %'])
        for r in d['chronic_list']:
            w.writerow([_fg(r['name']), _fg(r['class']), r['percentage']])
        if d['most_improved']:
            w.writerow([]); w.writerow(['Most improved', 'Class', 'From %', 'To %', 'Change'])
            for r in d['most_improved']:
                w.writerow([_fg(r['name']), _fg(r['class']), r['from'], r['to'], r['delta']])
        return Response(out.getvalue(), mimetype='text/csv',
                        headers={'Content-Disposition': f'attachment; filename={fname}.csv'})

    from openpyxl import Workbook
    from utils.web_exports import xlsx_response
    wb = Workbook()
    ws = wb.active; ws.title = 'Summary'
    ws.append(['Attendance analytics', term.name]); ws.append([])
    ws.append(['Executive summary'])
    for label, val in summary_rows:
        ws.append([label, val])
    ws.append([]); ws.append(['Attendance distribution', 'Students'])
    for label, key in (('Excellent (>=90%)', 'excellent'), ('Good (75-89%)', 'good'),
                       ('Fair (50-74%)', 'fair'), ('Poor (<50%)', 'poor')):
        ws.append([label, d['distribution'][key]])

    wr = wb.create_sheet('Class ranking')
    wr.append(['Class', 'Students', 'Attendance %'])
    for r in d['class_rank']:
        wr.append([r['class'], r['students'], r['percentage']])
    if d['branch_rank']:
        wb2 = wb.create_sheet('Branch ranking')
        wb2.append(['Branch', 'Attendance %'])
        for r in d['branch_rank']:
            wb2.append([r['branch'], r['percentage']])

    wc = wb.create_sheet('Chronic absentees')
    wc.append(['Student', 'Class', 'Attendance %'])
    for r in d['chronic_list']:
        wc.append([r['name'], r['class'], r['percentage']])

    if d['most_improved']:
        wi = wb.create_sheet('Most improved')
        wi.append(['Student', 'Class', 'From %', 'To %', 'Change'])
        for r in d['most_improved']:
            wi.append([r['name'], r['class'], r['from'], r['to'], r['delta']])

    return xlsx_response(wb, f'{fname}.xlsx')
