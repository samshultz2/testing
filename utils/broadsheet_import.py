"""Parse an uploaded broadsheet (Excel or CSV) of subject *totals* per student,
and help map its columns to our subjects and its rows to our students.

The sheet looks like the printed broadsheet: one row per student, one column per
subject, each cell the student's final total for that subject (0–100). A couple
of columns are identity (name / id) and some may be non-subject (e.g. "AV" for
average, "P/W" for Project Work) — the caller decides, via the review screen,
which columns to import and what each maps to.
"""
import csv
import io
import re


def parse_table(file_bytes, filename):
    """Return ``{'headers': [...], 'rows': [[...]]}`` from an xlsx/xls/csv upload.

    The header is the first row that has at least two non-empty cells; rows above
    it (stray titles like "MERIT LIST …") are skipped. Everything is returned as
    trimmed strings.
    """
    name = (filename or '').lower()
    if name.endswith(('.xlsx', '.xlsm', '.xls')):
        grid = _read_xlsx(file_bytes)
    else:
        grid = _read_csv(file_bytes)
    # Find the header row: first with >= 2 non-empty cells.
    hdr_idx = 0
    for i, row in enumerate(grid):
        if sum(1 for c in row if str(c).strip()) >= 2:
            hdr_idx = i
            break
    headers = [str(c).strip() for c in grid[hdr_idx]] if grid else []
    # Trim trailing empty header columns.
    while headers and not headers[-1]:
        headers.pop()
    ncol = len(headers)
    rows = []
    for row in grid[hdr_idx + 1:]:
        cells = [str(c).strip() if c is not None else '' for c in row][:ncol]
        cells += [''] * (ncol - len(cells))
        if any(cells):
            rows.append(cells)
    return {'headers': headers, 'rows': rows}


def _read_csv(file_bytes):
    text = file_bytes.decode('utf-8-sig', errors='replace')
    return [list(r) for r in csv.reader(io.StringIO(text))]


def _read_xlsx(file_bytes):
    from openpyxl import load_workbook
    wb = load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    ws = wb.active
    grid = []
    for row in ws.iter_rows(values_only=True):
        grid.append(['' if v is None else v for v in row])
    wb.close()
    return grid


def _norm(s):
    return re.sub(r'[^a-z0-9]', '', str(s or '').lower())


# Header aliases → canonical subject name fragments, for the auto-guess only.
_SUBJECT_ALIASES = {
    'pw': 'project', 'projectwork': 'project',
    'fm': 'further mathematics', 'fmaths': 'further mathematics',
    'chm': 'chemistry', 'phy': 'physics', 'bio': 'biology', 'eng': 'english',
    'lit': 'literature', 'crs': 'christian religious', 'irs': 'islamic religious',
    'civ': 'civic', 'com': 'commerce', 'gov': 'government', 'eco': 'economics',
    'geo': 'geography', 'agr': 'agric', 'liv': 'livestock', 'dit': 'digital',
    'dpr': 'data processing', 'his': 'history', 'pho': 'photography',
    'mth': 'mathematics', 'math': 'mathematics', 'maths': 'mathematics',
}

# Columns that are clearly NOT subjects (identity / derived), so the guesser can
# leave them unmapped by default.
NON_SUBJECT_HEADERS = {'sn', 'sno', 'sn0', 'no', 'num', 'id', 'studentid', 'student',
                       'studentname', 'name', 'surname', 'firstname', 'av', 'avg',
                       'average', 'total', 'position', 'pos', 'remark', 'remarks'}


def _round_half_up(x):
    import math
    return int(math.floor(x + 0.5))


def reconstruct_missing_scores(average, shown_scores, num_missing=1):
    """Work back the score(s) of subject(s) left off the sheet but still counted
    in a student's average.

    ``average`` is the value in the sheet's Average column; ``shown_scores`` the
    student's scores for the subjects that ARE on the sheet. The denominator of
    the average is ``N = len(shown) + num_missing`` (the missing subjects are
    counted), so the missing total is ``round(average*N) - sum(shown)``. Returns
    a list of length ``num_missing`` (evenly split, remainder on the first) — for
    the common single-missing case that's the exact reconstructed score.

    This mirrors the client-side reconstruction in the review page; it's kept
    here as the tested reference for the formula.
    """
    shown = [float(s) for s in shown_scores if s is not None and str(s) != '']
    m = len(shown)
    n = m + max(1, int(num_missing))
    total_missing = _round_half_up(float(average) * n) - sum(shown)
    k = max(1, int(num_missing))
    per = _round_half_up(total_missing / k)
    out = [per] * k
    out[0] += total_missing - per * k
    return out


def guess_name_column(headers):
    """Index of the most likely student-name column, or None."""
    for i, h in enumerate(headers):
        if _norm(h) in ('studentname', 'name', 'student', 'fullname'):
            return i
    for i, h in enumerate(headers):
        if 'name' in _norm(h):
            return i
    return None


def match_subject(header, subjects):
    """Best-guess ``Subject`` for a column header, or None. ``subjects`` is a list
    of Subject rows (with .name and .short_name)."""
    h = _norm(header)
    if not h or h in NON_SUBJECT_HEADERS:
        return None
    # exact name / short_name
    for s in subjects:
        if _norm(s.name) == h or (s.short_name and _norm(s.short_name) == h):
            return s
    # alias fragment
    frag = _SUBJECT_ALIASES.get(h)
    if frag:
        for s in subjects:
            if frag in _norm(s.name).replace(' ', '') or _norm(frag).replace(' ', '') in _norm(s.name):
                return s
    # prefix / contains
    for s in subjects:
        ns = _norm(s.name)
        if ns.startswith(h) or h.startswith(ns[:4]) and len(h) >= 3:
            return s
    return None
