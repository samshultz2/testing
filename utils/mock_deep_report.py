"""PDF (board pack), Excel (multi-sheet) and HD-image exports of the mock
deep-analytics payload produced by ``utils.mock_deep_analytics.deep_analytics``.

The PDF is a decision-ready pack: KPI band, subject league, teacher
effectiveness, arm league, students needing attention and the three
recommendation buckets. The HD image is the same pack rendered to one tall PNG.
"""
import io

from utils.numfmt import fmt_num as _n


def _theme():
    from reportlab.lib.colors import HexColor
    return (HexColor('#0D6A4E'), HexColor('#F4F7F5'), HexColor('#6B7A74'))


def _school_name():
    try:
        from utils.school import school_profile
        return school_profile().get('name') or 'School'
    except Exception:
        return 'School'


def _val(v):
    return '—' if v is None else str(v)


def deep_stem(meta):
    kind = meta.get('kind', 'mock')
    who = ('mock_jamb' if kind == 'jamb' else 'mock_waec')
    name = (meta.get('exam_name') or 'exam').replace(' ', '_')
    return f"{who}_deep_{name}"


def deep_filename(meta, ext='pdf'):
    return f"{deep_stem(meta)}.{ext}"


def deep_pdf(data):
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.units import mm
    from reportlab.lib import colors
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.platypus import (SimpleDocTemplate, Table, TableStyle, Paragraph,
                                    Spacer, KeepTogether)
    from utils.web_exports import pdf_escape

    primary, light, muted = _theme()
    meta = data.get('meta') or {}
    kind = meta.get('kind', 'jamb')
    styles = getSampleStyleSheet()
    h = ParagraphStyle('h', parent=styles['Title'], fontSize=16, textColor=primary, spaceAfter=1)
    sub = ParagraphStyle('sub', parent=styles['Normal'], fontSize=10, textColor=muted)
    hh = ParagraphStyle('hh', parent=styles['Heading2'], fontSize=12.5, textColor=primary,
                        spaceBefore=10, spaceAfter=4)
    body = ParagraphStyle('b', parent=styles['Normal'], fontSize=9, leading=12)
    muted_s = ParagraphStyle('mu', parent=styles['Normal'], fontSize=9, textColor=muted)

    buf = io.BytesIO()
    label = 'Mock JAMB' if kind == 'jamb' else 'Mock WAEC'
    doc = SimpleDocTemplate(buf, pagesize=A4, topMargin=14 * mm, bottomMargin=14 * mm,
                            leftMargin=13 * mm, rightMargin=13 * mm,
                            title=f"{label} deep analytics — {meta.get('exam_name', '')}")
    W = A4[0] - 26 * mm
    elems = [Paragraph(pdf_escape(_school_name()), h),
             Paragraph(f"{label} · Deep Analytics · <b>{pdf_escape(meta.get('exam_name', ''))}</b> · "
                       f"{pdf_escape(meta.get('session_name', ''))} · {pdf_escape(meta.get('exam_date', ''))} · "
                       f"{meta.get('students', 0)} candidate(s)", sub),
             Spacer(1, 8)]

    # KPI band
    kpis = data.get('kpis') or []
    krows, row = [], []
    for k in kpis:
        row.append(Paragraph(
            f"<b><font size=13 color='#0D6A4E'>{pdf_escape(str(k['value']))}</font></b><br/>"
            f"<font size=8 color='#14211C'>{pdf_escape(k['label'])}</font><br/>"
            f"<font size=7 color='#6B7A74'>{pdf_escape(k.get('sub', ''))}</font>", muted_s))
        if len(row) == 4:
            krows.append(row); row = []
    if row:
        row += [''] * (4 - len(row)); krows.append(row)
    if krows:
        kt = Table(krows, colWidths=[W / 4] * 4)
        kt.setStyle(TableStyle([('BACKGROUND', (0, 0), (-1, -1), light),
                                ('INNERGRID', (0, 0), (-1, -1), 3, colors.white),
                                ('BOX', (0, 0), (-1, -1), 3, colors.white),
                                ('TOPPADDING', (0, 0), (-1, -1), 8), ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
                                ('LEFTPADDING', (0, 0), (-1, -1), 9)]))
        elems.append(kt)

    def _tbl(head, rows, widths, aligns=None):
        t = Table([head] + rows, colWidths=widths, repeatRows=1)
        style = [
            ('BACKGROUND', (0, 0), (-1, 0), primary), ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'), ('FONTSIZE', (0, 0), (-1, -1), 8),
            ('ALIGN', (1, 0), (-1, -1), 'CENTER'), ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('GRID', (0, 0), (-1, -1), 0.4, colors.HexColor('#D5DED9')),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, light]),
            ('TOPPADDING', (0, 0), (-1, -1), 3), ('BOTTOMPADDING', (0, 0), (-1, -1), 3)]
        if aligns:
            for col, al in aligns.items():
                style.append(('ALIGN', (col, 0), (col, -1), al))
        t.setStyle(TableStyle(style))
        return t

    # Subject league
    subs = data.get('subjects') or []
    if subs:
        rows = [[pdf_escape(s['subject']), str(s['n']), _val(s['mean']), _val(s['sd']),
                 f"{_val(s['pass_rate'])}%", f"{_val(s['distinction_rate'])}%", s['band_label']]
                for s in subs]
        elems.append(KeepTogether([Paragraph('Subject league — weakest first', hh),
                     _tbl(['Subject', 'N', 'Mean', 'SD', 'Pass', 'Distinction', 'Verdict'], rows,
                          [W * 0.28, W * 0.08, W * 0.12, W * 0.1, W * 0.12, W * 0.15, W * 0.15],
                          {0: 'LEFT', 6: 'LEFT'})]))

    # Teacher effectiveness
    teachers = data.get('teachers') or []
    if teachers:
        rows = [[pdf_escape(t['teacher']), pdf_escape(', '.join(t['subjects'])[:34]),
                 str(t['students']), _val(t['mean']), f"{_val(t['pass_rate'])}%",
                 (f"{t['delta']:+}" if t['delta'] is not None else '—'), t['verdict']]
                for t in teachers]
        elems.append(KeepTogether([Paragraph('Teacher effectiveness (evidence-based · one mock)', hh),
                     _tbl(['Teacher', 'Subjects', 'Students', 'Mean', 'Pass', 'Δ vs cohort', 'Verdict'], rows,
                          [W * 0.18, W * 0.24, W * 0.1, W * 0.1, W * 0.1, W * 0.11, W * 0.17],
                          {0: 'LEFT', 1: 'LEFT', 6: 'LEFT'})]))

    # Arm league
    arms = data.get('arms') or []
    if arms:
        if kind == 'jamb':
            head = ['Class arm', 'Students', 'JAMB mean', '≥200', 'Subject pass']
            rows = [[pdf_escape(a['arm']), str(a['students']), _val(a.get('jamb_mean')),
                     f"{_val(a.get('above_200_rate'))}%", f"{_val(a['pass_rate'])}%"] for a in arms]
        else:
            head = ['Class arm', 'Students', 'Avg credits', 'Credit rate']
            rows = [[pdf_escape(a['arm']), str(a['students']), _val(a.get('avg_credits')),
                     f"{_val(a['pass_rate'])}%"] for a in arms]
        w = [W / len(head)] * len(head)
        elems.append(KeepTogether([Paragraph('Class-arm league — best first', hh),
                     _tbl(head, rows, w, {0: 'LEFT'})]))

    # Students needing attention
    seg = data.get('segments') or {}
    for key, title, tone in [('critical', 'Critical — urgent intervention', '#B43A2E'),
                             ('at_risk', 'At risk — near the threshold', '#9A7B0A')]:
        lst = seg.get(key) or []
        if lst:
            rows = [[pdf_escape(x['name']), pdf_escape(str(x['metric'])), pdf_escape(x['note'])]
                    for x in lst[:30]]
            elems.append(KeepTogether([Paragraph(f"{title} ({len(lst)})", hh),
                         _tbl(['Candidate', 'Result', 'Recommended action'], rows,
                              [W * 0.28, W * 0.14, W * 0.58], {0: 'LEFT', 2: 'LEFT'})]))

    # Recommendations
    recs = data.get('recommendations') or {}
    tone_hex = {'positive': '#0D6A4E', 'negative': '#B43A2E', 'warning': '#9A7B0A', 'insight': '#1F6FB2'}
    for bucket, title in [('students', 'Recommendations · Students'),
                          ('teachers', 'Recommendations · Teachers'),
                          ('management', 'Recommendations · Management')]:
        items = recs.get(bucket) or []
        if items:
            block = [Paragraph(title, hh)]
            for r in items:
                c = tone_hex.get(r['tone'], '#14211C')
                block.append(Paragraph(f"<font color='{c}'><b>&#9632; {pdf_escape(r['title'])}.</b></font> "
                                       f"{pdf_escape(r['text'])}", body))
                block.append(Spacer(1, 3))
            elems.append(KeepTogether(block))

    elems.append(Spacer(1, 6))
    elems.append(Paragraph(
        "Method: subject entries are attributed to the SSS3 subject teacher for each candidate's arm; "
        "verdicts weigh mean, pass rate, spread and sample size. This is one mock — read teacher and "
        "student judgements as part of a trend, not in isolation.", muted_s))
    doc.build(elems)
    return buf.getvalue()


def deep_png(data, dpi=200):
    from utils.analytics_org_pdf import _pdf_to_png
    return _pdf_to_png(deep_pdf(data), dpi)


# ---------------------------------------------------------------------------
# progress trends (longitudinal across many mocks)
# ---------------------------------------------------------------------------

def trends_stem(meta):
    who = 'mock_jamb' if meta.get('kind') == 'jamb' else 'mock_waec'
    scope = 'all_sessions' if meta.get('session_id') is None else 'session'
    return f"{who}_progress_{scope}"


def trends_filename(meta, ext='pdf'):
    return f"{trends_stem(meta)}.{ext}"


_ARROW = {'up': '↑', 'down': '↓', 'flat': '→'}


def trends_pdf(data):
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.units import mm
    from reportlab.lib import colors
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.platypus import (SimpleDocTemplate, Table, TableStyle, Paragraph,
                                    Spacer, KeepTogether)
    from utils.web_exports import pdf_escape

    primary, light, muted = _theme()
    meta = data.get('meta') or {}
    kind = meta.get('kind', 'jamb')
    head = data.get('headline') or {}
    periods = data.get('periods') or []
    styles = getSampleStyleSheet()
    h = ParagraphStyle('h', parent=styles['Title'], fontSize=16, textColor=primary, spaceAfter=1)
    sub = ParagraphStyle('sub', parent=styles['Normal'], fontSize=10, textColor=muted)
    hh = ParagraphStyle('hh', parent=styles['Heading2'], fontSize=12.5, textColor=primary,
                        spaceBefore=10, spaceAfter=4)
    body = ParagraphStyle('b', parent=styles['Normal'], fontSize=9, leading=12)
    muted_s = ParagraphStyle('mu', parent=styles['Normal'], fontSize=9, textColor=muted)

    buf = io.BytesIO()
    label = 'Mock JAMB' if kind == 'jamb' else 'Mock WAEC'
    scope_txt = 'across all sessions' if meta.get('session_id') is None else 'this session'
    doc = SimpleDocTemplate(buf, pagesize=landscape(A4), topMargin=13 * mm, bottomMargin=13 * mm,
                            leftMargin=13 * mm, rightMargin=13 * mm,
                            title=f"{label} progress — {scope_txt}")
    W = landscape(A4)[0] - 26 * mm
    elems = [Paragraph(pdf_escape(_school_name()), h),
             Paragraph(f"{label} · Progress across {meta.get('periods_count', 0)} mocks · "
                       f"{pdf_escape(meta.get('span', ''))} · {scope_txt}", sub),
             Spacer(1, 8)]

    # headline
    def _hb(lbl, first, last, delta, direction):
        arrow = _ARROW.get(direction, '')
        col = '#0D6A4E' if direction == 'up' else '#B43A2E' if direction == 'down' else '#6B7A74'
        return Paragraph(f"<b><font size=13 color='{col}'>{_val(first)} &#8594; {_val(last)} "
                         f"{arrow}</font></b><br/><font size=8 color='#14211C'>{pdf_escape(lbl)}</font>"
                         f"<br/><font size=7 color='#6B7A74'>{'+' if (delta or 0) >= 0 else ''}"
                         f"{_val(delta)} first&#8594;latest</font>", muted_s)
    hb = Table([[_hb(head.get('primary_label', ''), head.get('primary_first'), head.get('primary_last'),
                     head.get('primary_delta'), head.get('primary_direction')),
                 _hb(head.get('secondary_label', ''), head.get('secondary_first'), head.get('secondary_last'),
                     head.get('secondary_delta'), head.get('secondary_direction'))]],
                colWidths=[W / 2] * 2)
    hb.setStyle(TableStyle([('BACKGROUND', (0, 0), (-1, -1), light),
                            ('INNERGRID', (0, 0), (-1, -1), 3, colors.white),
                            ('BOX', (0, 0), (-1, -1), 3, colors.white),
                            ('TOPPADDING', (0, 0), (-1, -1), 8), ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
                            ('LEFTPADDING', (0, 0), (-1, -1), 10)]))
    elems.append(hb)

    def _tbl(headrow, rows, widths, aligns=None):
        t = Table([headrow] + rows, colWidths=widths, repeatRows=1)
        style = [('BACKGROUND', (0, 0), (-1, 0), primary), ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                 ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'), ('FONTSIZE', (0, 0), (-1, -1), 8),
                 ('ALIGN', (1, 0), (-1, -1), 'CENTER'), ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                 ('GRID', (0, 0), (-1, -1), 0.4, colors.HexColor('#D5DED9')),
                 ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, light]),
                 ('TOPPADDING', (0, 0), (-1, -1), 3), ('BOTTOMPADDING', (0, 0), (-1, -1), 3)]
        if aligns:
            for col, al in aligns.items():
                style.append(('ALIGN', (col, 0), (col, -1), al))
        t.setStyle(TableStyle(style))
        return t

    # cohort per-period table
    plabels = [p['label'] for p in periods]
    cohort = data.get('cohort') or []
    if cohort:
        headrow = ['Metric'] + plabels
        rows = [[head.get('primary_label', 'Primary')] + [_val(c['primary']) for c in cohort],
                [head.get('secondary_label', 'Secondary')] + [_val(c['secondary']) for c in cohort],
                ['Cohort ' + ('pass' if kind == 'jamb' else 'credit') + ' %'] + [_val(c['pass_rate']) for c in cohort],
                ['Candidates'] + [_val(c['students']) for c in cohort]]
        cw = [W * 0.24] + [(W * 0.76) / max(1, len(plabels))] * len(plabels)
        elems.append(KeepTogether([Paragraph('Cohort trajectory', hh), _tbl(headrow, rows, cw, {0: 'LEFT'})]))

    def _trend_table(title, rows_src):
        if not rows_src:
            return
        rows = [[pdf_escape(r['name'])] + [_val(v) for v in r['points']] +
                [f"{'+' if (r['delta'] or 0) >= 0 else ''}{_val(r['delta'])} {_ARROW.get(r['direction'], '')}"]
                for r in rows_src]
        headrow = ['Name'] + plabels + ['Trend']
        np = len(plabels)
        cw = [W * 0.24] + [(W * 0.56) / max(1, np)] * np + [W * 0.2]
        elems.append(KeepTogether([Paragraph(title, hh), _tbl(headrow, rows, cw, {0: 'LEFT'})]))

    _trend_table('Subject trend — credit/pass rate per mock (weakest movement first)', data.get('subject_trends'))
    _trend_table('Teacher trend — pass rate per mock', data.get('teacher_trends'))
    _trend_table('Class-arm trend', data.get('arm_trends'))

    recs = data.get('recommendations') or {}
    tone_hex = {'positive': '#0D6A4E', 'negative': '#B43A2E', 'warning': '#9A7B0A', 'insight': '#1F6FB2'}
    for bucket, title in [('students', 'Recommendations · Students'),
                          ('teachers', 'Recommendations · Teachers'),
                          ('management', 'Recommendations · Management')]:
        items = recs.get(bucket) or []
        if items:
            block = [Paragraph(title, hh)]
            for r in items:
                c = tone_hex.get(r['tone'], '#14211C')
                block.append(Paragraph(f"<font color='{c}'><b>&#9632; {pdf_escape(r['title'])}.</b></font> "
                                       f"{pdf_escape(r['text'])}", body))
                block.append(Spacer(1, 3))
            elems.append(KeepTogether(block))

    doc.build(elems)
    return buf.getvalue()


def trends_png(data, dpi=170):
    from utils.analytics_org_pdf import _pdf_to_png
    return _pdf_to_png(trends_pdf(data), dpi)


def trends_xlsx(data):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    meta = data.get('meta') or {}
    head = data.get('headline') or {}
    periods = data.get('periods') or []
    plabels = [p['label'] for p in periods]
    wb = Workbook()
    head_fill = PatternFill('solid', fgColor='0D6A4E'); head_font = Font(bold=True, color='FFFFFF')

    def sheet(title, headers, rows, first=False):
        ws = wb.active if first else wb.create_sheet()
        ws.title = title[:31]
        ws.append(headers)
        for c in ws[1]:
            c.fill = head_fill; c.font = head_font; c.alignment = Alignment(horizontal='center')
        for r in rows:
            ws.append(r)
        for col in ws.columns:
            width = max((len(str(c.value)) if c.value is not None else 0) for c in col) + 2
            ws.column_dimensions[col[0].column_letter].width = min(max(width, 10), 44)
        return ws

    ws = wb.active
    ws.title = 'Trajectory'
    ws['A1'] = f"{'Mock JAMB' if meta.get('kind') == 'jamb' else 'Mock WAEC'} — Progress"
    ws['A1'].font = Font(bold=True, size=14, color='0D6A4E')
    ws['A2'] = f"{meta.get('periods_count', 0)} mocks · {meta.get('span', '')}"
    ws.append([])
    ws.append(['Metric', 'First', 'Latest', 'Change', 'Direction'])
    for c in ws[4]:
        c.fill = head_fill; c.font = head_font
    ws.append([head.get('primary_label'), head.get('primary_first'), head.get('primary_last'),
               head.get('primary_delta'), head.get('primary_direction')])
    ws.append([head.get('secondary_label'), head.get('secondary_first'), head.get('secondary_last'),
               head.get('secondary_delta'), head.get('secondary_direction')])
    ws.column_dimensions['A'].width = 24

    cohort = data.get('cohort') or []
    if cohort:
        sheet('Cohort by mock', ['Mock', 'Primary', 'Secondary', 'Pass/Credit %', 'Candidates'],
              [[plabels[i], c['primary'], c['secondary'], c['pass_rate'], c['students']]
               for i, c in enumerate(cohort)])

    def trend_sheet(title, rows_src):
        if not rows_src:
            return
        sheet(title, ['Name'] + plabels + ['Change', 'Direction'],
              [[r['name']] + list(r['points']) + [r['delta'], r['direction']] for r in rows_src])

    trend_sheet('Subject trend', data.get('subject_trends'))
    trend_sheet('Teacher trend', data.get('teacher_trends'))
    trend_sheet('Class-arm trend', data.get('arm_trends'))

    recs = data.get('recommendations') or {}
    rec_rows = []
    for bucket in ('students', 'teachers', 'management'):
        for r_ in (recs.get(bucket) or []):
            rec_rows.append([bucket.capitalize(), r_['title'], r_['text']])
    if rec_rows:
        sheet('Recommendations', ['Audience', 'Headline', 'Detail'], rec_rows)

    out = io.BytesIO(); wb.save(out); return out.getvalue()


def deep_xlsx(data):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    meta = data.get('meta') or {}
    wb = Workbook()
    head_fill = PatternFill('solid', fgColor='0D6A4E'); head_font = Font(bold=True, color='FFFFFF')

    def sheet(title, headers, rows, first=False):
        ws = wb.active if first else wb.create_sheet()
        ws.title = title[:31]
        ws.append(headers)
        for c in ws[1]:
            c.fill = head_fill; c.font = head_font; c.alignment = Alignment(horizontal='center')
        for r in rows:
            ws.append(r)
        for col in ws.columns:
            width = max((len(str(c.value)) if c.value is not None else 0) for c in col) + 2
            ws.column_dimensions[col[0].column_letter].width = min(max(width, 10), 48)
        return ws

    ws = wb.active
    ws.title = 'Overview'
    ws['A1'] = f"{'Mock JAMB' if meta.get('kind') == 'jamb' else 'Mock WAEC'} — Deep Analytics"
    ws['A1'].font = Font(bold=True, size=14, color='0D6A4E')
    ws['A2'] = f"{meta.get('exam_name', '')} · {meta.get('session_name', '')} · {meta.get('exam_date', '')}"
    r = 4
    for k in (data.get('kpis') or []):
        ws[f'A{r}'] = k['label']; ws[f'A{r}'].font = Font(bold=True)
        ws[f'B{r}'] = k['value']; ws[f'C{r}'] = k.get('sub', ''); r += 1
    ws.column_dimensions['A'].width = 22; ws.column_dimensions['B'].width = 14
    ws.column_dimensions['C'].width = 26

    subs = data.get('subjects') or []
    if subs:
        sheet('Subjects', ['Subject', 'N', 'Mean', 'SD', 'Min', 'Max', 'Pass %', 'Distinction %',
                           'Verdict', 'Recommendation'],
              [[s['subject'], s['n'], s['mean'], s['sd'], s['min'], s['max'], s['pass_rate'],
                s['distinction_rate'], s['band_label'], s['recommendation']] for s in subs])

    teachers = data.get('teachers') or []
    if teachers:
        sheet('Teachers', ['Teacher', 'Subjects', 'Students', 'Entries', 'Mean', 'Pass %',
                           'Distinction %', 'Δ vs cohort', 'Verdict', 'Recommendation'],
              [[t['teacher'], ', '.join(t['subjects']), t['students'], t['entries'], t['mean'],
                t['pass_rate'], t['distinction_rate'], t['delta'], t['verdict'], t['recommendation']]
               for t in teachers])

    arms = data.get('arms') or []
    if arms:
        if meta.get('kind') == 'jamb':
            sheet('Class arms', ['Arm', 'Students', 'JAMB mean', '≥200 %', 'Subject pass %'],
                  [[a['arm'], a['students'], a.get('jamb_mean'), a.get('above_200_rate'), a['pass_rate']]
                   for a in arms])
        else:
            sheet('Class arms', ['Arm', 'Students', 'Avg credits', 'Credit rate %'],
                  [[a['arm'], a['students'], a.get('avg_credits'), a['pass_rate']] for a in arms])

    seg = data.get('segments') or {}
    seg_rows = []
    for key, band in [('critical', 'Critical'), ('at_risk', 'At risk'), ('honour', 'Honour roll')]:
        for x in (seg.get(key) or []):
            seg_rows.append([band, x['name'], x['metric'], x['note']])
    if seg_rows:
        sheet('Students', ['Segment', 'Candidate', 'Result', 'Note'], seg_rows)

    recs = data.get('recommendations') or {}
    rec_rows = []
    for bucket in ('students', 'teachers', 'management'):
        for r_ in (recs.get(bucket) or []):
            rec_rows.append([bucket.capitalize(), r_['title'], r_['text']])
    if rec_rows:
        sheet('Recommendations', ['Audience', 'Headline', 'Detail'], rec_rows)

    out = io.BytesIO(); wb.save(out); return out.getvalue()
