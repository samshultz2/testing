"""Institution analytics → a board-pack PDF.

A professional, branded multi-section report for owners / principals / boards:
KPI band, unit league, subject league, teacher effectiveness league, the
decision-oriented recommendations and the intervention / honour lists.
"""
import io

from utils.numfmt import fmt_num as _n


def _theme():
    from reportlab.lib.colors import HexColor
    return (HexColor('#0D6A4E'), HexColor('#C9A227'), HexColor('#F4F7F5'),
            HexColor('#14211C'), HexColor('#6B7A74'), HexColor('#B43A2E'))


def _school_name():
    try:
        from utils.school import school_profile
        return school_profile().get('name') or 'School'
    except Exception:
        return 'School'


def institution_stem(data, term):
    base = (data.get('scope_label') or 'school').replace(' ', '_')
    tname = (term.name if term else 'term').replace(' ', '_')
    return f"analytics_{base}_{tname}"


def institution_filename(data, term, ext='pdf'):
    return f"{institution_stem(data, term)}.{ext}"


def institution_png(data, term, dpi=200):
    """Render the board-pack PDF to a single tall HD PNG (every page stitched)."""
    pdf = institution_pdf(data, term)
    if not pdf:
        return None
    import fitz
    doc = fitz.open(stream=pdf, filetype='pdf')
    matrix = fitz.Matrix(dpi / 72.0, dpi / 72.0)
    pixmaps = [page.get_pixmap(matrix=matrix) for page in doc]
    if len(pixmaps) == 1:
        return pixmaps[0].tobytes('png')
    from PIL import Image
    imgs = [Image.frombytes('RGB', (p.width, p.height), p.samples) for p in pixmaps]
    gap = max(1, dpi // 20)
    width = max(im.width for im in imgs)
    height = sum(im.height for im in imgs) + gap * (len(imgs) - 1)
    canvas = Image.new('RGB', (width, height), 'white')
    y = 0
    for im in imgs:
        canvas.paste(im, ((width - im.width) // 2, y))
        y += im.height + gap
    out = io.BytesIO()
    canvas.save(out, format='PNG')
    return out.getvalue()


def institution_xlsx(data, term):
    """Multi-sheet workbook of the institution analytics (KPIs + every league +
    trend + the roll lists) for analysts who want the raw numbers."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    s = data.get('summary') or {}
    wb = Workbook()
    head_fill = PatternFill('solid', fgColor='0D6A4E')
    head_font = Font(bold=True, color='FFFFFF')
    title_font = Font(bold=True, size=14, color='0D6A4E')

    def sheet(title, headers, rows, first=False):
        ws = wb.active if first else wb.create_sheet()
        ws.title = title[:31]
        ws.append(headers)
        for c in ws[1]:
            c.fill = head_fill; c.font = head_font; c.alignment = Alignment(horizontal='center')
        for r in rows:
            ws.append(r)
        for col in ws.columns:
            width = max((len(str(c.value)) if c.value is not None else 0) for c in col) + 2
            ws.column_dimensions[col[0].column_letter].width = min(max(width, 10), 44)
        return ws

    # Summary sheet
    ws = wb.active
    ws.title = 'Summary'
    ws['A1'] = f"Academic Performance — {data.get('scope_label', '')}"
    ws['A1'].font = title_font
    ws['A2'] = f"{term.full_name if term else ''}"
    r = 4
    for lbl, val in [
        ('Students', s.get('students')), ('Assessed', s.get('assessed')),
        ('Average score', s.get('class_average')), ('Pass rate %', s.get('pass_rate')),
        ('Distinction rate %', s.get('distinction_rate')), ('Highest', s.get('highest')),
        ('Lowest', s.get('lowest')), ('Entry completion %', s.get('completion')),
        ('Units', s.get('units')), ('Subjects', s.get('subjects_count')),
        ('Teachers', s.get('teachers_count')), ('Pass mark', s.get('pass_mark')),
    ]:
        ws[f'A{r}'] = lbl; ws[f'A{r}'].font = Font(bold=True); ws[f'B{r}'] = val
        r += 1
    ws.column_dimensions['A'].width = 22; ws.column_dimensions['B'].width = 14

    units = data.get('units') or []
    if units:
        sheet(f"{data.get('unit_kind','Unit')} league",
              [data.get('unit_kind', 'Unit'), 'Average', 'Pass %', 'Students'],
              [[u['label'], u['average'], u['pass_rate'], u['students']] for u in units])
    subjects = data.get('subjects') or []
    if subjects:
        sheet('Subjects', ['Subject', 'Average', 'Pass %', 'Highest', 'Lowest', 'Assessed'],
              [[x['name'], x['average'], x['pass_rate'], x['highest'], x['lowest'], x['assessed']]
               for x in subjects])
    teachers = data.get('teachers') or []
    if teachers:
        sheet('Teachers', ['Teacher', 'Average', 'Pass %', 'Subjects', 'Classes', 'Entries',
                           'Completion %', 'Verdict'],
              [[t['name'], t['average'], t['pass_rate'], t['subject_count'], t['class_count'],
                t['entries'], t['completion'], t['verdict']] for t in teachers])
    branches = data.get('branches') or []
    if branches:
        sheet('Branches', ['Branch', 'Average', 'Pass %', 'Students'],
              [[b['label'], b['average'], b['pass_rate'], b['students']] for b in branches])
    att = data.get('attendance') or {}
    if att.get('bands'):
        rows = [[b['band'], b['count'], b['average']] for b in att['bands']]
        ws2 = sheet('Attendance vs results', ['Attendance band', 'Students', 'Avg score'], rows)
        if att.get('correlation') is not None:
            ws2.append([]); ws2.append(['Correlation (r)', att['correlation']])
    tr = data.get('trends') or {}
    if tr.get('term_names'):
        rows = [['Average score'] + list(tr.get('averages', [])),
                ['Pass rate'] + list(tr.get('pass_rates', []))]
        sheet('Trend', ['Metric'] + tr['term_names'], rows)
    interv = data.get('intervention') or []
    if interv:
        sheet('Intervention', ['Student', 'Class', 'Average', 'Failing'],
              [[x['name'], x.get('class', ''), x['average'], x['failed']] for x in interv])
    honour = data.get('honour_roll') or []
    if honour:
        sheet('Honour roll', ['Student', 'Class', 'Average'],
              [[x['name'], x.get('class', ''), x['average']] for x in honour])

    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()


def institution_pdf(data, term):
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.units import mm
    from reportlab.lib import colors
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.platypus import (SimpleDocTemplate, Table, TableStyle, Paragraph,
                                    Spacer, KeepTogether)
    from utils.web_exports import pdf_escape

    primary, accent, light, ink, muted, danger = _theme()
    s = data.get('summary') or {}
    styles = getSampleStyleSheet()
    h = ParagraphStyle('h', parent=styles['Title'], fontSize=17, textColor=primary, spaceAfter=1)
    sub = ParagraphStyle('sub', parent=styles['Normal'], fontSize=10, textColor=muted)
    hh = ParagraphStyle('hh', parent=styles['Heading2'], fontSize=12.5, textColor=primary,
                        spaceBefore=10, spaceAfter=4)
    body = ParagraphStyle('b', parent=styles['Normal'], fontSize=9, leading=12)
    muted_s = ParagraphStyle('mu', parent=styles['Normal'], fontSize=9, textColor=muted)

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4, topMargin=14 * mm, bottomMargin=14 * mm,
                            leftMargin=15 * mm, rightMargin=15 * mm,
                            title=f"Analytics — {data.get('scope_label', '')}")
    W = A4[0] - 30 * mm
    elems = [Paragraph(pdf_escape(_school_name()), h),
             Paragraph(f"Academic Performance Report · <b>{pdf_escape(data.get('scope_label',''))}</b> · "
                       f"{pdf_escape(term.full_name if term else '')}", sub),
             Spacer(1, 8)]

    # ---- KPI band ---------------------------------------------------------
    kpis = [
        ('Class average', _n(s.get('class_average'))),
        ('Pass rate', f"{_n(s.get('pass_rate'))}%"),
        ('Distinctions', f"{_n(s.get('distinction_rate'))}%"),
        ('Students', f"{s.get('assessed')}/{s.get('students')}"),
        ('Completion', f"{_n(s.get('completion'))}%"),
        ('Subjects', str(s.get('subjects_count', 0))),
        ('Teachers', str(s.get('teachers_count', 0))),
        ('Units', str(s.get('units', 0))),
    ]
    krows, row = [], []
    for lbl, val in kpis:
        row.append(Paragraph(f"<b><font size=15 color='#0D6A4E'>{pdf_escape(val)}</font></b>"
                             f"<br/><font size=8 color='#6B7A74'>{pdf_escape(lbl)}</font>", muted_s))
        if len(row) == 4:
            krows.append(row); row = []
    if row:
        row += [''] * (4 - len(row)); krows.append(row)
    kt = Table(krows, colWidths=[W / 4] * 4)
    kt.setStyle(TableStyle([('BACKGROUND', (0, 0), (-1, -1), light),
                            ('INNERGRID', (0, 0), (-1, -1), 3, colors.white),
                            ('BOX', (0, 0), (-1, -1), 3, colors.white),
                            ('TOPPADDING', (0, 0), (-1, -1), 9), ('BOTTOMPADDING', (0, 0), (-1, -1), 9),
                            ('LEFTPADDING', (0, 0), (-1, -1), 10)]))
    elems.append(kt)

    def league(title, head, rows, widths, hi=None):
        block = [Paragraph(title, hh)]
        data_rows = [head] + rows
        t = Table(data_rows, colWidths=widths, repeatRows=1)
        style = [
            ('BACKGROUND', (0, 0), (-1, 0), primary), ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'), ('FONTSIZE', (0, 0), (-1, -1), 8.5),
            ('ALIGN', (1, 0), (-1, -1), 'CENTER'), ('ALIGN', (0, 0), (0, -1), 'LEFT'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('GRID', (0, 0), (-1, -1), 0.4, colors.HexColor('#D5DED9')),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, light]),
            ('TOPPADDING', (0, 0), (-1, -1), 3.5), ('BOTTOMPADDING', (0, 0), (-1, -1), 3.5),
        ]
        if hi:
            for r, colour in hi:
                style.append(('TEXTCOLOR', (0, r), (-1, r), colour))
        t.setStyle(TableStyle(style))
        block.append(t)
        return KeepTogether(block)

    # ---- unit league ------------------------------------------------------
    units = data.get('units') or []
    if units:
        kind = data.get('unit_kind', 'Unit')
        rows, hi = [], []
        for i, u in enumerate(units, 1):
            rows.append([f"{i}. {pdf_escape(u['label'])}", _n(u['average']),
                         f"{_n(u['pass_rate'])}%", str(u['students'])])
        hi = [(1, primary)]
        if len(units) > 1:
            hi.append((len(units), danger))
        elems.append(league(f"{kind} league (best → worst)",
                            [kind, 'Avg', 'Pass %', 'Students'], rows,
                            [W * 0.46, W * 0.18, W * 0.18, W * 0.18], hi))

    # ---- branch (campus) league ------------------------------------------
    branches = data.get('branches') or []
    if branches:
        rows = [[f"{i}. {pdf_escape(b['label'])}", _n(b['average']), f"{_n(b['pass_rate'])}%",
                 str(b['students'])] for i, b in enumerate(branches, 1)]
        hi = [(1, primary)]
        if len(branches) > 1:
            hi.append((len(branches), danger))
        elems.append(league("Campus league (best → worst)",
                            ['Branch', 'Avg', 'Pass %', 'Students'], rows,
                            [W * 0.46, W * 0.18, W * 0.18, W * 0.18], hi))

    # ---- subject league ---------------------------------------------------
    subjects = data.get('subjects') or []
    if subjects:
        rows = [[pdf_escape(x['name']), _n(x['average']), f"{_n(x['pass_rate'])}%",
                 _n(x['highest']), _n(x['lowest']), str(x['assessed'])] for x in subjects]
        elems.append(league("Subject league (hardest → easiest)",
                            ['Subject', 'Avg', 'Pass %', 'High', 'Low', 'N'], rows,
                            [W * 0.34, W * 0.13, W * 0.15, W * 0.12, W * 0.12, W * 0.14]))

    # ---- teacher league ---------------------------------------------------
    teachers = data.get('teachers') or []
    if teachers:
        flag_colour = {'strong': primary, 'good': ink, 'watch': accent,
                       'review': danger, 'compliance': danger, 'insufficient': muted}
        rows, hi = [], []
        for i, t in enumerate(teachers, 1):
            rows.append([pdf_escape(t['name']), _n(t['average']), f"{_n(t['pass_rate'])}%",
                         str(t['subject_count']), str(t['class_count']),
                         f"{_n(t['completion'])}%", pdf_escape(t['verdict'])])
            hi.append((i, flag_colour.get(t['flag'], ink)))
        elems.append(league("Teacher effectiveness league",
                            ['Teacher', 'Avg', 'Pass %', 'Subj', 'Cls', 'Entry', 'Verdict'], rows,
                            [W * 0.20, W * 0.09, W * 0.11, W * 0.08, W * 0.07, W * 0.10, W * 0.35], hi))

    # ---- recommendations --------------------------------------------------
    recs = data.get('recommendations') or []
    if recs:
        tone_hex = {'positive': '#0D6A4E', 'negative': '#B43A2E', 'watch': '#9A7B0A',
                    'insight': '#14211C'}
        block = [Paragraph('Insights &amp; recommendations', hh)]
        for r in recs:
            c = tone_hex.get(r['tone'], '#14211C')
            block.append(Paragraph(
                f"<font color='{c}'><b>&#9632; {pdf_escape(r['title'])}.</b></font> "
                f"{pdf_escape(r['text'])}", body))
            block.append(Spacer(1, 3))
        elems.append(KeepTogether(block[:2]))
        elems += block[2:]

    # ---- attendance vs results -------------------------------------------
    att = data.get('attendance') or {}
    if att.get('bands'):
        rows = [[pdf_escape(b['band']), str(b['count']), _n(b['average'])] for b in att['bands']]
        head = ['Attendance band', 'Students', 'Avg score']
        blk = [Paragraph('Attendance vs results', hh)]
        if att.get('correlation') is not None:
            blk.append(Paragraph(f"Correlation (r) = <b>{_n(att['correlation'])}</b> "
                                 f"across {att.get('coverage', 0)} students with attendance records.", body))
        t = Table([head] + rows, colWidths=[W * 0.4, W * 0.3, W * 0.3], repeatRows=1)
        t.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), primary), ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'), ('FONTSIZE', (0, 0), (-1, -1), 8.5),
            ('ALIGN', (1, 0), (-1, -1), 'CENTER'), ('GRID', (0, 0), (-1, -1), 0.4, colors.HexColor('#D5DED9')),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, light]),
            ('TOPPADDING', (0, 0), (-1, -1), 3.5), ('BOTTOMPADDING', (0, 0), (-1, -1), 3.5)]))
        blk.append(t)
        elems.append(KeepTogether(blk))

    # ---- term-on-term trend ----------------------------------------------
    tr = data.get('trends') or {}
    tnames = tr.get('term_names') or []
    if len(tnames) > 1 and any(v is not None for v in (tr.get('averages') or [])):
        head = ['Metric'] + tnames
        rows = [['Average score'] + [_n(v) if v is not None else '—' for v in tr.get('averages', [])],
                ['Pass rate'] + [(f"{_n(v)}%" if v is not None else '—') for v in tr.get('pass_rates', [])]]
        w0 = W * 0.28
        col = (W - w0) / len(tnames)
        elems.append(league("Performance trend across terms (this session)",
                            head, rows, [w0] + [col] * len(tnames)))

    # ---- intervention & honour -------------------------------------------
    interv = data.get('intervention') or []
    if interv:
        rows = [[pdf_escape(x['name']), pdf_escape(x.get('class', '')), _n(x['average']),
                 str(x['failed'])] for x in interv[:25]]
        elems.append(league(f"Students needing intervention ({len(interv)})",
                            ['Student', 'Class', 'Avg', 'Failing'], rows,
                            [W * 0.42, W * 0.28, W * 0.15, W * 0.15],
                            hi=[(i, danger) for i in range(1, len(rows) + 1)]))
    honour = data.get('honour_roll') or []
    if honour:
        rows = [[f"{i}. " + pdf_escape(x['name']), pdf_escape(x.get('class', '')), _n(x['average'])]
                for i, x in enumerate(honour[:25], 1)]
        elems.append(league(f"Honour roll — distinctions ({len(honour)})",
                            ['Student', 'Class', 'Avg'], rows,
                            [W * 0.5, W * 0.3, W * 0.2],
                            hi=[(i, primary) for i in range(1, len(rows) + 1)]))

    doc.build(elems)
    return buf.getvalue()
