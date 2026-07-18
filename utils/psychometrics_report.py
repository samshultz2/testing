"""PDF & Excel exports for the psychometric item analysis."""
import io


def _theme():
    from reportlab.lib.colors import HexColor
    return (HexColor('#0D6A4E'), HexColor('#F4F7F5'), HexColor('#6B7A74'),
            HexColor('#B43A2E'), HexColor('#9A7B0A'))


_VERDICT_HEX = {'keep': '#0D6A4E', 'review': '#9A7B0A', 'reject': '#B43A2E'}


def item_analysis_pdf(data):
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.units import mm
    from reportlab.lib import colors
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.platypus import (SimpleDocTemplate, Table, TableStyle, Paragraph,
                                    Spacer, KeepTogether)
    from utils.web_exports import pdf_escape
    from utils.numfmt import fmt_num as _n

    primary, light, muted, danger, warn = _theme()
    meta = data.get('meta') or {}
    s = data.get('summary') or {}
    styles = getSampleStyleSheet()
    h = ParagraphStyle('h', parent=styles['Title'], fontSize=16, textColor=primary, spaceAfter=1)
    sub = ParagraphStyle('sub', parent=styles['Normal'], fontSize=10, textColor=muted)
    hh = ParagraphStyle('hh', parent=styles['Heading2'], fontSize=12.5, textColor=primary,
                        spaceBefore=10, spaceAfter=4)
    body = ParagraphStyle('b', parent=styles['Normal'], fontSize=9, leading=12)
    muted_s = ParagraphStyle('mu', parent=styles['Normal'], fontSize=9, textColor=muted)

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4, topMargin=14 * mm, bottomMargin=14 * mm,
                            leftMargin=14 * mm, rightMargin=14 * mm,
                            title=f"Item analysis — {meta.get('title', '')}")
    W = A4[0] - 28 * mm
    elems = [Paragraph('Psychometric Item Analysis', h),
             Paragraph(f"<b>{pdf_escape(meta.get('title', ''))}</b> · "
                       f"{meta.get('respondents', 0)} candidates · {meta.get('question_count', 0)} items", sub),
             Spacer(1, 8)]

    if meta.get('insufficient'):
        elems.append(Paragraph(pdf_escape(meta.get('reason', 'Not enough data.')), body))
        doc.build(elems)
        return buf.getvalue()

    kpis = [('KR-20', _n(s.get('kr20'))), ('Reliability', s.get('kr20_label', '')),
            ('Mean score', f"{_n(s.get('mean_score'))}/{s.get('items')}"),
            ('Mean %', f"{_n(s.get('mean_pct'))}%"), ('SEM', _n(s.get('sem'))),
            ('Avg difficulty', _n(s.get('mean_difficulty'))),
            ('Avg discrimination', _n(s.get('mean_discrimination'))),
            ('Keep / Review / Reject', f"{s.get('keep')} / {s.get('review')} / {s.get('reject')}")]
    krows, row = [], []
    for lbl, val in kpis:
        row.append(Paragraph(f"<b><font size=12 color='#0D6A4E'>{pdf_escape(str(val))}</font></b>"
                             f"<br/><font size=8 color='#6B7A74'>{pdf_escape(lbl)}</font>", muted_s))
        if len(row) == 4:
            krows.append(row); row = []
    if row:
        row += [''] * (4 - len(row)); krows.append(row)
    kt = Table(krows, colWidths=[W / 4] * 4)
    kt.setStyle(TableStyle([('BACKGROUND', (0, 0), (-1, -1), light),
                            ('INNERGRID', (0, 0), (-1, -1), 3, colors.white),
                            ('BOX', (0, 0), (-1, -1), 3, colors.white),
                            ('TOPPADDING', (0, 0), (-1, -1), 8), ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
                            ('LEFTPADDING', (0, 0), (-1, -1), 9)]))
    elems.append(kt)

    recs = data.get('recommendations') or []
    if recs:
        tone_hex = {'positive': '#0D6A4E', 'negative': '#B43A2E', 'watch': '#9A7B0A'}
        block = [Paragraph('Findings &amp; recommendations', hh)]
        for r in recs:
            c = tone_hex.get(r['tone'], '#14211C')
            block.append(Paragraph(f"<font color='{c}'><b>&#9632; {pdf_escape(r['title'])}.</b></font> "
                                   f"{pdf_escape(r['text'])}", body))
            block.append(Spacer(1, 3))
        elems += block

    topics = (data.get('topics') or {})
    if topics.get('has_topics') and topics.get('items'):
        trows = [[pdf_escape(t['topic']), str(t['questions']), f"{_n(t['mastery'])}%",
                  t['band'].title(), f"{t['below_half']} ({_n(t['below_half_pct'])}%)"]
                 for t in topics['items']]
        tt = Table([['Topic', 'Items', 'Mastery', 'Band', 'Below half']] + trows,
                   colWidths=[W * 0.4, W * 0.1, W * 0.15, W * 0.15, W * 0.2], repeatRows=1)
        tt.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), primary), ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'), ('FONTSIZE', (0, 0), (-1, -1), 8.5),
            ('ALIGN', (1, 0), (-1, -1), 'CENTER'), ('GRID', (0, 0), (-1, -1), 0.4, colors.HexColor('#D5DED9')),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, light]),
            ('TOPPADDING', (0, 0), (-1, -1), 3), ('BOTTOMPADDING', (0, 0), (-1, -1), 3)]))
        elems.append(KeepTogether([Paragraph('Topic mastery (weakest first)', hh), tt]))

    items = data.get('items') or []
    if items:
        head = ['Q', 'Key', 'p%', 'Difficulty', 'D', 'r_pb', 'Dead', 'Verdict']
        rows = [head]
        hi = []
        for it in items:
            rows.append([str(it['number']), it['key'], _n(it['p_pct']), it['p_label'],
                         _n(it['d']) if it['d'] is not None else '—',
                         _n(it['rpb']) if it['rpb'] is not None else '—',
                         str(it['dead_distractors']) if it['dead_distractors'] else '',
                         it['verdict'].upper()])
            hi.append(_VERDICT_HEX.get(it['verdict']))
        t = Table(rows, colWidths=[W * x for x in (0.06, 0.07, 0.09, 0.22, 0.1, 0.1, 0.08, 0.28)], repeatRows=1)
        style = [
            ('BACKGROUND', (0, 0), (-1, 0), primary), ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'), ('FONTSIZE', (0, 0), (-1, -1), 8),
            ('ALIGN', (2, 0), (-2, -1), 'CENTER'), ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('GRID', (0, 0), (-1, -1), 0.4, colors.HexColor('#D5DED9')),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, light]),
            ('TOPPADDING', (0, 0), (-1, -1), 2.5), ('BOTTOMPADDING', (0, 0), (-1, -1), 2.5)]
        for i, hx in enumerate(hi, 1):
            if hx:
                style.append(('TEXTCOLOR', (7, i), (7, i), colors.HexColor(hx)))
        t.setStyle(TableStyle(style))
        elems.append(KeepTogether([Paragraph('Item statistics', hh)]))
        elems.append(t)
        elems.append(Paragraph("p% = difficulty (percent correct). D = discrimination "
                               "(upper 27% − lower 27%). r_pb = point-biserial item-total "
                               "correlation. Dead = non-functioning distractors.", muted_s))

    doc.build(elems)
    return buf.getvalue()


def item_analysis_xlsx(data):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    meta = data.get('meta') or {}
    s = data.get('summary') or {}
    wb = Workbook()
    head_fill = PatternFill('solid', fgColor='0D6A4E')
    head_font = Font(bold=True, color='FFFFFF')

    def sheet(title, headers, rows, first=False):
        ws = wb.active if first else wb.create_sheet()
        ws.title = title[:31]
        ws.append(headers)
        for c in ws[1]:
            c.fill = head_fill; c.font = head_font; c.alignment = Alignment(horizontal='center')
        for r in rows:
            ws.append(r)
        for col in ws.columns:
            width = max((len(str(c.value)) if c.value is not None else 0) for c in col) + 2
            ws.column_dimensions[col[0].column_letter].width = min(max(width, 10), 50)
        return ws

    ws = wb.active
    ws.title = 'Summary'
    ws['A1'] = f"Item Analysis — {meta.get('title', '')}"
    ws['A1'].font = Font(bold=True, size=14, color='0D6A4E')
    r = 3
    for lbl, val in [('Candidates', meta.get('respondents')), ('Items', meta.get('question_count')),
                     ('KR-20 reliability', s.get('kr20')), ('Reliability band', s.get('kr20_label')),
                     ('SEM', s.get('sem')), ('Mean score', s.get('mean_score')),
                     ('SD', s.get('sd_score')), ('Mean %', s.get('mean_pct')),
                     ('Avg difficulty (p)', s.get('mean_difficulty')),
                     ('Avg discrimination (D)', s.get('mean_discrimination')),
                     ('Keep', s.get('keep')), ('Review', s.get('review')), ('Reject', s.get('reject'))]:
        ws[f'A{r}'] = lbl; ws[f'A{r}'].font = Font(bold=True); ws[f'B{r}'] = val; r += 1
    ws.column_dimensions['A'].width = 24; ws.column_dimensions['B'].width = 30

    topics = (data.get('topics') or {})
    if topics.get('has_topics') and topics.get('items'):
        sheet('Topics', ['Topic', 'Items', 'Mastery %', 'Band', 'Below half', 'Below half %'],
              [[t['topic'], t['questions'], t['mastery'], t['band'], t['below_half'],
                t['below_half_pct']] for t in topics['items']])
        cols = topics.get('columns') or []
        if topics.get('students') and cols:
            sheet('Student topics', ['Student', 'Overall %'] + cols + ['Weakest'],
                  [[s['name'], s['overall']] + [s['cells'].get(c) for c in cols] + [s['weakest']]
                   for s in topics['students']])

    items = data.get('items') or []
    if items:
        sheet('Items', ['Q', 'Key', 'Difficulty p', 'p %', 'Difficulty band',
                        'Discrimination D', 'Point-biserial', 'Blank', 'Dead distractors', 'Verdict'],
              [[it['number'], it['key'], it['p'], it['p_pct'], it['p_label'], it['d'],
                it['rpb'], it['blank'], it['dead_distractors'], it['verdict']] for it in items])
        drows = []
        for it in items:
            for o in it['options']:
                drows.append([it['number'], o['option'], 'KEY' if o['is_key'] else '',
                              o['picks'], o['rate'], o['upper'], o['lower'], o['flag'] or ''])
        sheet('Distractors', ['Q', 'Option', 'Key?', 'Picks', 'Pick %', 'Upper 27%', 'Lower 27%', 'Flag'], drows)
    out = io.BytesIO(); wb.save(out); return out.getvalue()
