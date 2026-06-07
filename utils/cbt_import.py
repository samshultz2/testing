"""
Parse a question file (Excel .xlsx or CSV) into question dicts for the CBT
question bank / an exam.

Recognised columns (case-insensitive, flexible names):
  question | a/option a | b | c | d | correct (A-D) | marks | topic | difficulty
"""
import csv
import io


_ALIASES = {
    'question': 'question', 'question text': 'question', 'q': 'question',
    'a': 'a', 'option a': 'a', 'optiona': 'a',
    'b': 'b', 'option b': 'b', 'optionb': 'b',
    'c': 'c', 'option c': 'c', 'optionc': 'c',
    'd': 'd', 'option d': 'd', 'optiond': 'd',
    'correct': 'correct', 'answer': 'correct', 'correct option': 'correct',
    'correct answer': 'correct', 'key': 'correct',
    'marks': 'marks', 'mark': 'marks', 'score': 'marks',
    'topic': 'topic', 'difficulty': 'difficulty', 'level': 'difficulty',
}


def _normalise_correct(value, row):
    """Return 'A'-'D'. Accepts a letter, or the full text of the correct option."""
    v = (str(value or '')).strip()
    if v[:1].upper() in ('A', 'B', 'C', 'D') and len(v) <= 2:
        return v[:1].upper()
    # Maybe they wrote the answer text — match it to an option.
    for letter in ('a', 'b', 'c', 'd'):
        if row.get(letter) and str(row[letter]).strip().lower() == v.lower():
            return letter.upper()
    return v[:1].upper() if v[:1].upper() in ('A', 'B', 'C', 'D') else ''


def _build(rows):
    """rows: list of header-keyed dicts (already aliased)."""
    out = []
    for r in rows:
        q = (str(r.get('question') or '')).strip()
        if not q:
            continue
        correct = _normalise_correct(r.get('correct'), r)
        if correct not in ('A', 'B', 'C', 'D'):
            continue
        try:
            marks = float(r.get('marks') or 1)
        except (ValueError, TypeError):
            marks = 1
        out.append({
            'question_text': q,
            'option_a': (str(r.get('a') or '')).strip(),
            'option_b': (str(r.get('b') or '')).strip(),
            'option_c': (str(r.get('c') or '')).strip(),
            'option_d': (str(r.get('d') or '')).strip(),
            'correct_option': correct,
            'marks': marks if marks > 0 else 1,
            'topic': (str(r.get('topic') or '')).strip() or None,
            'difficulty': (str(r.get('difficulty') or 'Medium')).strip().title() or 'Medium',
        })
    return out


def _alias_headers(headers):
    return [_ALIASES.get((h or '').strip().lower(), (h or '').strip().lower()) for h in headers]


def parse_csv(data):
    text = data.decode('utf-8-sig', errors='ignore')
    reader = csv.reader(io.StringIO(text))
    rows = list(reader)
    if not rows:
        return []
    keys = _alias_headers(rows[0])
    dicts = [dict(zip(keys, r)) for r in rows[1:]]
    return _build(dicts)


def parse_xlsx(data):
    from openpyxl import load_workbook
    wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    keys = _alias_headers([str(c) if c is not None else '' for c in rows[0]])
    dicts = [dict(zip(keys, r)) for r in rows[1:]]
    return _build(dicts)


def parse_file(data, filename):
    name = (filename or '').lower()
    if name.endswith('.csv'):
        return parse_csv(data)
    if name.endswith(('.xlsx', '.xlsm')):
        return parse_xlsx(data)
    # try xlsx then csv
    try:
        return parse_xlsx(data)
    except Exception:
        return parse_csv(data)
