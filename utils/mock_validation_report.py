"""PDF & Excel exports for mock→actual validation (JAMB numeric, WAEC grades)."""
import io

from utils.numfmt import fmt_num as _n


def _theme():
    from reportlab.lib.colors import HexColor
    return (HexColor('#0D6A4E'), HexColor('#F4F7F5'), HexColor('#6B7A74'))


def _school_name():
    try:
        from utils.school import school_profile
        return school_profile().get('name') or 'School'
    except Exception:
        return 'School'


def validation_pdf(data):
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.units import mm
    from reportlab.lib import colors
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.platypus import (SimpleDocTemplate, Table, TableStyle, Paragraph,
                                    Spacer, KeepTogether)
    from utils.web_exports import pdf_escape

    primary, light, muted = _theme()
    meta = data.get('meta') or {}
    s = data.get('summary') or {}
    kind = meta.get('kind', 'jamb')
    styles = getSampleStyleSheet()
    h = ParagraphStyle('h', parent=styles['Title'], fontSize=16, textColor=primary, spaceAfter=1)
    sub = ParagraphStyle('sub', parent=styles['Normal'], fontSize=10, textColor=muted)
    hh = ParagraphStyle('hh', parent=styles['Heading2'], fontSize=12.5, textColor=primary,
                        spaceBefore=10, spaceAfter=4)
    body = ParagraphStyle('b', parent=styles['Normal'], fontSize=9, leading=12)
    muted_s = ParagraphStyle('mu', parent=styles['Normal'], fontSize=9, textColor=muted)

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4, topMargin=14 * mm, bottomMargin=14 * mm,
                            leftMargin=14 * mm, rightMargin=14 * mm,
                            title=f"Mock validation — {meta.get('mock_name', '')}")
    W = A4[0] - 28 * mm
    label = 'Mock JAMB → Actual JAMB' if kind == 'jamb' else 'Mock WAEC → Actual WAEC'
    elems = [Paragraph(pdf_escape(_school_name()), h),
             Paragraph(f"Mock Validation · <b>{label}</b> · {pdf_escape(meta.get('mock_name', ''))} · "
                       f"{s.get('matched', 0)} matched candidate(s)", sub),
             Spacer(1, 8)]

    if kind == 'jamb':
        kpis = [('Correlation', _n(s.get('correlation'))), ('R²', _n(s.get('r_squared'))),
                ('Mean abs error', f"±{_n(s.get('mae'))}"), ('Bias', f"{_n(s.get('bias'))}"),
                ('Within 40 pts', f"{_n(s.get('within40_pct'))}%"),
                ('Mock mean', _n(s.get('mock_mean'))), ('Actual mean', _n(s.get('actual_mean'))),
                ('Calibration', s.get('bias_direction', ''))]
    else:
        kpis = [('Exact grade', f"{_n(s.get('exact_pct'))}%"),
                ('Within 1 grade', f"{_n(s.get('within1_pct'))}%"),
                ('Correlation', _n(s.get('correlation'))),
                ('Grade MAE', _n(s.get('grade_mae'))), ('Grade bias', _n(s.get('grade_bias'))),
                ('Credit accuracy', f"{_n(s.get('credit_accuracy'))}%"),
                ('Credit sensitivity', f"{_n(s.get('credit_sensitivity'))}%"),
                ('Calibration', s.get('bias_direction', ''))]
    krows, row = [], []
    for lbl, val in kpis:
        row.append(Paragraph(f"<b><font size=12 color='#0D6A4E'>{pdf_escape(str(val))}</font></b>"
                             f"<br/><font size=8 color='#6B7A74'>{pdf_escape(lbl)}</font>", muted_s))
        if len(row) == 4:
            krows.append(row); row = []
    if row:
        row += [''] * (4 - len(row)); krows.append(row)
    kt = Table(krows, colWidths=[W / 4] * 4)
    kt.setStyle(TableStyle([('BACKGROUND', (0, 0), (-1, -1), light),
                            ('INNERGRID', (0, 0), (-1, -1), 3, colors.white),
                            ('BOX', (0, 0), (-1, -1), 3, colors.white),
                            ('TOPPADDING', (0, 0), (-1, -1), 8), ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
                            ('LEFTPADDING', (0, 0), (-1, -1), 9)]))
    elems.append(kt)

    recs = data.get('recommendations') or []
    if recs:
        tone_hex = {'positive': '#0D6A4E', 'negative': '#B43A2E', 'watch': '#9A7B0A'}
        block = [Paragraph('Findings &amp; recommendations', hh)]
        for r in recs:
            c = tone_hex.get(r['tone'], '#14211C')
            block.append(Paragraph(f"<font color='{c}'><b>&#9632; {pdf_escape(r['title'])}.</b></font> "
                                   f"{pdf_escape(r['text'])}", body))
            block.append(Spacer(1, 3))
        elems += block

    def _tbl(head, rows, widths):
        t = Table([head] + rows, colWidths=widths, repeatRows=1)
        t.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), primary), ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'), ('FONTSIZE', (0, 0), (-1, -1), 8.5),
            ('ALIGN', (1, 0), (-1, -1), 'CENTER'), ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('GRID', (0, 0), (-1, -1), 0.4, colors.HexColor('#D5DED9')),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, light]),
            ('TOPPADDING', (0, 0), (-1, -1), 3), ('BOTTOMPADDING', (0, 0), (-1, -1), 3)]))
        return t

    if kind == 'jamb':
        pm = data.get('per_mock') or []
        if pm:
            rows = [[pdf_escape(x['name']), str(x['matched']), _n(x['correlation']), f"±{_n(x['mae'])}"] for x in pm]
            elems.append(KeepTogether([Paragraph('Which mock predicts best', hh),
                         _tbl(['Mock', 'Matched', 'Correlation', 'MAE'], rows,
                              [W * 0.4, W * 0.2, W * 0.2, W * 0.2])]))
        pairs = data.get('pairs') or []
        if pairs:
            rows = [[pdf_escape(p['name']), str(p['mock']), str(p['actual']),
                     ('+' if p['actual'] - p['mock'] >= 0 else '') + str(p['actual'] - p['mock'])]
                    for p in sorted(pairs, key=lambda x: x['actual'] - x['mock'])[:40]]
            elems.append(KeepTogether([Paragraph('Candidate mock vs actual (largest gaps first)', hh),
                         _tbl(['Candidate', 'Mock', 'Actual', 'Δ'], rows,
                              [W * 0.5, W * 0.17, W * 0.17, W * 0.16])]))
    else:
        subs = data.get('subjects') or []
        if subs:
            rows = [[pdf_escape(x['subject']), str(x['matched']), f"{_n(x['exact_pct'])}%",
                     f"{_n(x['within1_pct'])}%", _n(x['correlation']),
                     f"{_n(x['credit_accuracy'])}%"] for x in subs]
            elems.append(KeepTogether([Paragraph('Per-subject grade agreement (weakest first)', hh),
                         _tbl(['Subject', 'N', 'Exact', 'Within 1', 'Corr', 'Credit acc.'], rows,
                              [W * 0.3, W * 0.1, W * 0.15, W * 0.15, W * 0.15, W * 0.15])]))

    doc.build(elems)
    return buf.getvalue()


def validation_xlsx(data):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    meta = data.get('meta') or {}
    s = data.get('summary') or {}
    kind = meta.get('kind', 'jamb')
    wb = Workbook()
    head_fill = PatternFill('solid', fgColor='0D6A4E'); head_font = Font(bold=True, color='FFFFFF')

    def sheet(title, headers, rows, first=False):
        ws = wb.active if first else wb.create_sheet()
        ws.title = title[:31]
        ws.append(headers)
        for c in ws[1]:
            c.fill = head_fill; c.font = head_font; c.alignment = Alignment(horizontal='center')
        for r in rows:
            ws.append(r)
        for col in ws.columns:
            width = max((len(str(c.value)) if c.value is not None else 0) for c in col) + 2
            ws.column_dimensions[col[0].column_letter].width = min(max(width, 10), 44)
        return ws

    ws = wb.active
    ws.title = 'Summary'
    ws['A1'] = f"Mock Validation — {meta.get('mock_name', '')}"
    ws['A1'].font = Font(bold=True, size=14, color='0D6A4E')
    r = 3
    for lbl, val in list(s.items()):
        if isinstance(val, dict):
            continue
        ws[f'A{r}'] = lbl; ws[f'A{r}'].font = Font(bold=True); ws[f'B{r}'] = val; r += 1
    ws.column_dimensions['A'].width = 22; ws.column_dimensions['B'].width = 22

    if kind == 'jamb':
        if data.get('per_mock'):
            sheet('Per mock', ['Mock', 'Matched', 'Correlation', 'MAE'],
                  [[x['name'], x['matched'], x['correlation'], x['mae']] for x in data['per_mock']])
        if data.get('pairs'):
            sheet('Candidates', ['Candidate', 'Mock', 'Actual', 'Error', 'Year'],
                  [[p['name'], p['mock'], p['actual'], p['actual'] - p['mock'], p.get('year')]
                   for p in data['pairs']])
    else:
        if data.get('subjects'):
            sheet('Subjects', ['Subject', 'Matched', 'Exact %', 'Within 1 %', 'Correlation',
                               'Grade MAE', 'Grade bias', 'Credit accuracy %', 'Credit sensitivity %'],
                  [[x['subject'], x['matched'], x['exact_pct'], x['within1_pct'], x['correlation'],
                    x['grade_mae'], x['grade_bias'], x['credit_accuracy'], x['credit_sensitivity']]
                   for x in data['subjects']])
    out = io.BytesIO(); wb.save(out); return out.getvalue()
