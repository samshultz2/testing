"""
Excel import/export utilities for the Student Management System
Uses openpyxl for Excel file operations
"""
import io
import re
from datetime import datetime
from utils.web_exports import formula_guard as _fg
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter


# Style definitions
HEADER_FONT = Font(bold=True, color='FFFFFF', size=11)
HEADER_FILL = PatternFill(start_color='2E7D32', end_color='2E7D32', fill_type='solid')
HEADER_ALIGNMENT = Alignment(horizontal='center', vertical='center', wrap_text=True)

CELL_ALIGNMENT = Alignment(horizontal='left', vertical='center')
THIN_BORDER = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)


def create_styled_workbook():
    """Create a new workbook with default styling"""
    wb = Workbook()
    return wb


def style_header_row(ws, row_num=1, num_cols=None):
    """Apply header styling to a row"""
    if num_cols is None:
        num_cols = ws.max_column
    
    for col in range(1, num_cols + 1):
        cell = ws.cell(row=row_num, column=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGNMENT
        cell.border = THIN_BORDER


def auto_adjust_columns(ws):
    """Auto-adjust column widths based on content"""
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except Exception:
                pass
        
        adjusted_width = min(max_length + 2, 50)  # Cap at 50
        ws.column_dimensions[column_letter].width = adjusted_width


def export_students_to_excel(students):
    """
    Export student data to Excel
    
    Args:
        students: List of Student model objects
    
    Returns:
        BytesIO object containing the Excel file
    """
    wb = create_styled_workbook()
    ws = wb.active
    ws.title = "Students"
    
    # Define headers
    headers = [
        'Student ID', 'Surname', 'First Name', 'Middle Name', 'Gender',
        'Date of Birth', 'Religion', 'Home Address', 'Hobbies',
        'Primary Contact', 'Created Date'
    ]
    
    # Write headers
    for col, header in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=header)
    
    style_header_row(ws, 1, len(headers))
    
    # Write data
    for row, student in enumerate(students, 2):
        primary_contact = student.parent_contacts.filter_by(is_primary=True).first()
        
        ws.cell(row=row, column=1, value=_fg(student.student_id))
        ws.cell(row=row, column=2, value=_fg(student.surname))
        ws.cell(row=row, column=3, value=_fg(student.first_name))
        ws.cell(row=row, column=4, value=_fg(student.middle_name or ''))
        ws.cell(row=row, column=5, value=_fg(student.gender))
        ws.cell(row=row, column=6, value=student.date_of_birth.strftime('%Y-%m-%d') if student.date_of_birth else '')
        ws.cell(row=row, column=7, value=_fg(student.religion or ''))
        ws.cell(row=row, column=8, value=_fg(student.home_address or ''))
        ws.cell(row=row, column=9, value=_fg(student.hobbies or ''))
        ws.cell(row=row, column=10, value=_fg(primary_contact.phone_number if primary_contact else ''))
        ws.cell(row=row, column=11, value=student.created_at.strftime('%Y-%m-%d'))
        
        # Apply border to data cells
        for col in range(1, len(headers) + 1):
            ws.cell(row=row, column=col).border = THIN_BORDER
    
    auto_adjust_columns(ws)
    
    # Save to BytesIO
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output


def create_student_import_template():
    """
    Create a template Excel file for importing students
    
    Returns:
        BytesIO object containing the Excel template
    """
    wb = create_styled_workbook()
    ws = wb.active
    ws.title = "Student Import Template"

    # One clean header row (row 1) — columns are matched by name, in any order.
    # Only Surname and First Name are required. The importer also accepts a sheet
    # typed straight from a class register with the same column names.
    headers = [
        'Surname', 'First Name', 'Middle Name', 'Gender', 'Date of Birth',
        'Religion', 'Address', 'Hobbies', 'Name of Primary Contact',
        'Phone Number', 'Relationship',
    ]
    for col, header in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=header)
    style_header_row(ws, 1, len(headers))

    # A single example row. Its surname is 'EXAMPLE', which the importer always
    # skips — so the file imports cleanly even if you forget to delete it.
    example = [
        'EXAMPLE', 'John', 'Oluwaseun', 'Male', '2010-05-15',
        'Christianity', '25 Broad Street, Lagos', 'Football, Reading',
        'Mrs. Jane Adeyemi', '08012345678', 'Mother',
    ]
    for col, value in enumerate(example, 1):
        cell = ws.cell(row=2, column=col, value=value)
        cell.font = Font(italic=True, color='999999')

    # Instructions live in a comment on the first header cell (doesn't affect import).
    from openpyxl.comments import Comment
    ws.cell(row=1, column=1).comment = Comment(
        "Fill one student per row from row 2 down.\n"
        "Required: Surname, First Name.\n"
        "Gender: Male/Female (optional — blank becomes 'Unknown').\n"
        "Date of Birth: e.g. 2010-05-15 or 15/05/2010.\n"
        "Delete the grey EXAMPLE row (or leave it — it is ignored).",
        "PosyHub")

    auto_adjust_columns(ws)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output


def _norm_header(value):
    """Normalise a header cell to a comparison key: lowercased, no '*'/spaces/punct."""
    if value is None:
        return ''
    s = str(value).strip().lower().replace('*', '')
    for ch in ('_', '-', '.', '/'):
        s = s.replace(ch, ' ')
    return ' '.join(s.split())


# Accepted column-name aliases → canonical field. Header matching is order-free,
# so the spreadsheet's columns can be in any order and named loosely.
_HEADER_ALIASES = {
    'surname': 'surname', 'last name': 'surname', 'lastname': 'surname',
    'first name': 'first_name', 'firstname': 'first_name', 'first': 'first_name',
    'middle name': 'middle_name', 'middlename': 'middle_name',
    'other name': 'middle_name', 'other names': 'middle_name',
    'gender': 'gender', 'sex': 'gender',
    'date of birth': 'dob', 'dob': 'dob', 'birth date': 'dob', 'birthdate': 'dob',
    'religion': 'religion',
    'address': 'address', 'home address': 'address', 'house address': 'address',
    'hobbies': 'hobbies', 'hobby': 'hobbies', 'interests': 'hobbies',
    'name of primary contact': 'parent_name', 'primary contact': 'parent_name',
    'parent name': 'parent_name', 'guardian name': 'parent_name',
    'parent': 'parent_name', 'guardian': 'parent_name', 'parent guardian': 'parent_name',
    'phone number': 'parent_phone', 'phone': 'parent_phone', 'phone no': 'parent_phone',
    'parent phone': 'parent_phone', 'guardian phone': 'parent_phone',
    'contact': 'parent_phone', 'contact number': 'parent_phone', 'mobile': 'parent_phone',
    'relationship': 'parent_rel', 'relation': 'parent_rel',
}

# First-column values that mark a non-data row (instruction/sample artefacts from
# older templates) and should be skipped rather than imported.
_SKIP_FIRST_CELL = {'instructions:', 'instruction', 'required', 'optional',
                    'surname', 'last name', 'example'}


def _cell_str(value):
    """Excel cell → trimmed string, dropping the spurious '.0' on integer floats."""
    if value is None:
        return None
    if isinstance(value, float) and value.is_integer():
        value = int(value)
    s = str(value).strip()
    return s or None


def _cap(value, max_len):
    """Truncate a string to a column's length so a length-enforcing database
    (PostgreSQL) never rejects an over-long value and rolls back the whole row."""
    if value is None:
        return None
    return value[:max_len]


def _parse_dob(value, errors, row_num):
    """Parse a date-of-birth cell that may be a real date or a string."""
    from datetime import date as _date
    if value is None or (isinstance(value, str) and not value.strip()):
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, _date):
        return value
    s = str(value).strip()
    for fmt in ('%Y-%m-%d', '%d/%m/%Y', '%d/%m/%y', '%d-%m-%Y', '%m/%d/%Y'):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    errors.append(f"Row {row_num}: couldn't read date of birth '{s}' — left blank")
    return None


def _normalise_phone(value):
    """Nigerian phone tidy-up: keep digits, restore a leading 0 lost by Excel."""
    s = _cell_str(value)
    if not s:
        return None
    digits = ''.join(c for c in s if c.isdigit())
    if not digits:
        return s
    # Excel stores 08012345678 as the integer 8012345678 — restore the 0.
    if len(digits) == 10 and not digits.startswith('0'):
        digits = '0' + digits
    return digits


def _normalise_gender(value):
    """Map loose gender values to Male/Female; return None if unrecognised."""
    s = _cell_str(value)
    if not s:
        return None
    t = s.strip().lower()
    if t in ('m', 'male', 'boy'):
        return 'Male'
    if t in ('f', 'female', 'girl'):
        return 'Female'
    return None


def _dup_key(surname, first_name, dob, extra=''):
    """Identity key for duplicate detection.

    Name + date of birth is the primary key. When the sheet has **no date of
    birth** (common for hand-typed registers), the name alone is not a reliable
    identity — two different students often share a name — so we add the home
    address as a discriminator. That way distinct same-name students still
    import, while a true re-import (same name + same address) is still skipped.
    """
    return ((surname or '').strip().lower(),
            (first_name or '').strip().lower(),
            dob,
            extra if dob is None else '')


def _rows_from_xlsx(file_stream):
    wb = load_workbook(file_stream, data_only=True)
    return list(wb.active.iter_rows(values_only=True))


def _rows_from_csv(file_stream):
    import csv
    data = file_stream.read()
    if isinstance(data, bytes):
        data = data.decode('utf-8-sig', errors='replace')   # tolerate a BOM
    # skipinitialspace so '..., "No 1, Main St", ...' (a space before the quote)
    # is still read as one quoted field rather than splitting on the inner comma.
    return [tuple(r) for r in csv.reader(io.StringIO(data), skipinitialspace=True)]


def _split_phones(value):
    """A phone cell may hold several numbers, e.g. "0803…, 0701…". Split on the
    usual separators and normalise each, dropping blanks/dupes."""
    s = _cell_str(value)
    if not s:
        return []
    out = []
    for part in re.split(r'[,;/\n]| and | & ', s):
        n = _normalise_phone(part)
        if n and any(c.isdigit() for c in n) and n not in out:
            out.append(n)
    return out


def rows_from_pasted_text(text):
    """Parse copy-pasted tabular text into row tuples (header row first).

    The delimiter is auto-detected from the heading line: tab if it looks
    tab-separated (a spreadsheet paste), otherwise comma. Quoted values are
    handled by the csv module. Blank lines are dropped. This feeds the same
    header-matching importer as the .xlsx/.csv paths, so a plain paste of just
    'Surname, First Name' works like a full sheet.
    """
    import csv
    lines = [ln for ln in (text or '').splitlines() if ln.strip()]
    if not lines:
        return []
    head = lines[0]
    delim = '\t' if (head.count('\t') and head.count('\t') >= head.count(',')) else ','
    # skipinitialspace so '..., "No 1, Main St", ...' (note the space before the
    # opening quote, common in hand-typed/pasted lists) is read as one quoted
    # field instead of splitting on the comma inside the quotes.
    return [tuple(r) for r in csv.reader(io.StringIO('\n'.join(lines)),
                                         delimiter=delim, skipinitialspace=True)]


def preview_student_rows(rows):
    """Dry-run a parsed sheet: which columns were recognised, which ignored,
    and a per-row breakdown — without touching the database.

    Returns a dict mirroring what ``import_student_rows`` would do for the
    name/skip checks, so the UI can show a faithful preview before committing.
    """
    if not rows:
        return {'recognised': [], 'ignored': [], 'total': 0, 'valid': 0,
                'invalid': 0, 'rows': []}
    header = rows[0]
    colmap, ignored = {}, []
    for cell in header:
        field = _HEADER_ALIASES.get(_norm_header(cell))
        if field and field not in colmap:
            colmap[field] = True
        elif cell not in (None, ''):
            ignored.append(str(cell).strip())

    def col(row, field):
        # Recompute the index the same way import_student_rows does.
        for idx, cell in enumerate(header):
            if _HEADER_ALIASES.get(_norm_header(cell)) == field:
                return _cell_str(row[idx]) if idx < len(row) else None
        return None

    out, valid = [], 0
    for row_num, row in enumerate(rows[1:], start=2):
        surname = col(row, 'surname')
        first_name = col(row, 'first_name')
        if not surname and not first_name:
            continue   # blank line — not shown, not counted
        if surname and surname.lower() in _SKIP_FIRST_CELL:
            continue   # leftover template/instruction row
        # At least one name present → importable (the other is left blank).
        valid += 1
        partial = not (surname and first_name)
        out.append({'row': row_num,
                    'error': 'Only one name — will import, please complete it' if partial else None,
                    'name': ' '.join(p for p in [surname, first_name] if p),
                    'details': {'gender': _normalise_gender(col(row, 'gender')) or 'Unknown',
                                'religion': col(row, 'religion'),
                                'date_of_birth': col(row, 'dob'),
                                'phone_number': ', '.join(_split_phones(col(row, 'parent_phone'))) or None,
                                'address': col(row, 'address')}})
    return {'recognised': sorted(colmap.keys()), 'ignored': ignored,
            'total': len(out), 'valid': valid, 'invalid': len(out) - valid,
            'rows': out}


def import_students_from_excel(file_stream, db, Student, ParentContact,
                               branch_id=None, skip_duplicates=True,
                               class_arm_assignment_id=None):
    """Import students from an .xlsx stream (thin wrapper over import_student_rows)."""
    return import_student_rows(_rows_from_xlsx(file_stream), db, Student, ParentContact,
                               branch_id=branch_id, skip_duplicates=skip_duplicates,
                               class_arm_assignment_id=class_arm_assignment_id)


def import_students_from_upload(file_storage, db, Student, ParentContact, **kwargs):
    """Import from an uploaded .xlsx/.xls/.csv file (auto-detected by name)."""
    name = (getattr(file_storage, 'filename', '') or '').lower()
    if name.endswith('.csv'):
        rows = _rows_from_csv(file_storage.stream)
    elif name.endswith(('.xlsx', '.xls')):
        rows = _rows_from_xlsx(file_storage.stream)
    else:
        return 0, ['Unsupported file type — please upload a .xlsx or .csv file.']
    return import_student_rows(rows, db, Student, ParentContact, **kwargs)


def import_student_rows(rows, db, Student, ParentContact,
                        branch_id=None, skip_duplicates=True,
                        class_arm_assignment_id=None):
    """
    Import students from a list of rows (header row + data rows).

    Columns are matched by **header name** (row 1), in any order, so a register
    typed as Surname / First Name / Middle Name / Date of Birth / Address /
    Phone Number / Religion / Name of Primary Contact imports cleanly. Only
    Surname and First Name are required; a missing Gender column defaults to
    'Unknown' (flagged so you can fill it in later).

    Args:
        rows: list of row tuples (first row is the header)
        db, Student, ParentContact: SQLAlchemy db + model classes
        branch_id: branch to stamp on every imported student (optional)
        skip_duplicates: skip rows whose name + DOB already exist (default True)
        class_arm_assignment_id: if set, enroll every imported student into this
            class-arm assignment for the active term

    Returns:
        Tuple of (success_count, messages) — messages are per-row notes/errors.
    """
    success_count = 0
    errors = []
    enrolled = 0
    StudentEnrollment = None
    if class_arm_assignment_id is not None:
        from models import StudentEnrollment as _SE
        StudentEnrollment = _SE

    # Build {canonical_field: column_index} from the header row.
    if not rows:
        return 0, ['The sheet is empty.']
    header = rows[0]
    colmap = {}
    for idx, cell in enumerate(header):
        field = _HEADER_ALIASES.get(_norm_header(cell))
        if field and field not in colmap:
            colmap[field] = idx

    if 'surname' not in colmap and 'first_name' not in colmap:
        return 0, [
            "Couldn't find a name column. Row 1 must include at least a "
            "'Surname' or 'First Name' column. Found: "
            + ', '.join(str(h) for h in header if h) + '.'
        ]

    def get(row, field):
        idx = colmap.get(field)
        if idx is None or idx >= len(row):
            return None
        return row[idx]

    # Preload existing students' identity keys so re-running an import doesn't
    # create duplicates. Only *active* students count: a student you soft-deleted
    # (moved to trash) must not block re-importing them, otherwise deleting a
    # class to re-add it would report every row as a duplicate. Scoped to the
    # target branch. Keys added during this run are tracked too, so duplicate
    # rows within the same file are also caught.
    seen_keys = set()
    if skip_duplicates:
        q = Student.query.filter(Student.is_active == True)
        if branch_id is not None:
            q = q.filter(Student.branch_id == branch_id)
        for sn, fn, d, addr in q.with_entities(
                Student.surname, Student.first_name, Student.date_of_birth,
                Student.home_address):
            seen_keys.add(_dup_key(sn, fn, d, (addr or '').strip().lower()))

    gender_defaulted = 0
    duplicates_skipped = 0

    for row_num, row in enumerate(rows[1:], start=2):
        surname = _cell_str(get(row, 'surname'))
        first_name = _cell_str(get(row, 'first_name'))

        # Skip blank rows and leftover instruction/sample rows.
        if not surname and not first_name:
            continue
        if surname and surname.lower() in _SKIP_FIRST_CELL:
            continue
        # A row needs at least one of the two names; the missing one is stored
        # blank (the columns are NOT NULL) so a record with only a surname —
        # or only a first name — still imports and can be completed later.
        if not surname or not first_name:
            errors.append(f"Row {row_num}: only one name given — imported, please complete it")
            surname = surname or ''
            first_name = first_name or ''

        dob = _parse_dob(get(row, 'dob'), errors, row_num)
        address = _cell_str(get(row, 'address'))

        if skip_duplicates:
            key = _dup_key(surname, first_name, dob, (address or '').strip().lower())
            if key in seen_keys:
                duplicates_skipped += 1
                errors.append(
                    f"Row {row_num}: {surname} {first_name} already exists — skipped")
                continue
            seen_keys.add(key)

        gender = _normalise_gender(get(row, 'gender'))
        defaulted = gender is None
        if defaulted:
            gender = 'Unknown'

        try:
            # Per-row savepoint: a bad row rolls back only itself, never the
            # rows already imported in this batch.
            with db.session.begin_nested():
                # Cap each value to its column length. Length-enforcing
                # databases (PostgreSQL) reject an over-long value and would
                # otherwise roll back the whole row — e.g. a mis-shifted column
                # that drops a long address into the 30-char religion field.
                student = Student(
                    student_id=Student.generate_student_id(),
                    surname=_cap(surname, 50),
                    first_name=_cap(first_name, 50),
                    middle_name=_cap(_cell_str(get(row, 'middle_name')), 50),
                    gender=_cap(gender, 10),
                    date_of_birth=dob,
                    religion=_cap(_cell_str(get(row, 'religion')), 30),
                    home_address=_cell_str(get(row, 'address')),   # TEXT — unbounded
                    hobbies=_cell_str(get(row, 'hobbies')),         # TEXT — unbounded
                    branch_id=branch_id,
                    is_active=True,
                )
                db.session.add(student)
                db.session.flush()  # assign student.id

                # A cell may carry several phone numbers — create one contact
                # each (first is primary), all sharing the parent name/relation.
                # A name on its own (no phone) is dropped, as phone is required.
                parent_name = _cap(_cell_str(get(row, 'parent_name')), 100)
                parent_rel = _cap(_cell_str(get(row, 'parent_rel')) or 'Guardian', 20)
                for i, phone in enumerate(_split_phones(get(row, 'parent_phone'))):
                    db.session.add(ParentContact(
                        student_id=student.id,
                        name=parent_name,
                        phone_number=_cap(phone, 15),
                        relationship=parent_rel,
                        is_primary=(i == 0),
                    ))

                # Optionally enrol the new student in the chosen class arm.
                if StudentEnrollment is not None:
                    db.session.add(StudentEnrollment(
                        student_id=student.id,
                        class_arm_assignment_id=class_arm_assignment_id,
                        is_active=True,
                    ))

            success_count += 1
            if defaulted:
                gender_defaulted += 1
            if StudentEnrollment is not None:
                enrolled += 1

        except Exception as e:
            errors.append(f"Row {row_num}: {str(e)}")

    if success_count > 0:
        db.session.commit()

    if enrolled:
        errors.insert(0, f"{enrolled} student(s) enrolled in the selected class.")
    if duplicates_skipped:
        errors.insert(0, f"{duplicates_skipped} row(s) skipped as duplicates — a student "
                         "with that name is already on record (matched on name + date of "
                         "birth, or name + address when no DOB is given).")
    if gender_defaulted:
        errors.insert(0, f"{gender_defaulted} student(s) had no Gender column/value "
                         "and were set to 'Unknown' — edit them to set Male/Female.")

    return success_count, errors



def export_attendance_to_excel(attendance_data, class_name, week_info):
    """
    Export weekly attendance data to Excel
    
    Args:
        attendance_data: Dict from get_weekly_attendance_summary
        class_name: Name of the class arm
        week_info: Week details
    
    Returns:
        BytesIO object containing the Excel file
    """
    wb = create_styled_workbook()
    ws = wb.active
    ws.title = "Weekly Attendance"
    
    # Title section
    ws.cell(row=1, column=1, value=f"Weekly Attendance Report - {class_name}")
    ws.cell(row=1, column=1).font = Font(bold=True, size=14)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=8)
    
    ws.cell(row=2, column=1, value=f"Week {week_info['week_number']}: {week_info['start_date']} to {week_info['end_date']}")
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=8)
    
    # Headers
    headers = ['S/N', 'Student Name', 'Gender']
    for day in attendance_data['school_days']:
        headers.append(f"{day.strftime('%a')} AM")
        headers.append(f"{day.strftime('%a')} PM")
    headers.extend(['Weekly Total', 'Percentage'])
    
    for col, header in enumerate(headers, 1):
        ws.cell(row=4, column=col, value=header)
    
    style_header_row(ws, 4, len(headers))
    
    # Student data
    times_opened = attendance_data['times_opened']
    for idx, student in enumerate(attendance_data['students'], 1):
        row = idx + 4
        ws.cell(row=row, column=1, value=idx)
        ws.cell(row=row, column=2, value=student['student_name'])
        ws.cell(row=row, column=3, value=student['gender'])
        
        col = 4
        for daily in student['daily']:
            ws.cell(row=row, column=col, value='✓' if daily['morning'] else '✗')
            ws.cell(row=row, column=col + 1, value='✓' if daily['afternoon'] else '✗')
            col += 2
        
        ws.cell(row=row, column=col, value=student['weekly_total'])
        percentage = round((student['weekly_total'] / times_opened * 100), 2) if times_opened > 0 else 0
        ws.cell(row=row, column=col + 1, value=f"{percentage}%")
        
        # Apply borders
        for c in range(1, len(headers) + 1):
            ws.cell(row=row, column=c).border = THIN_BORDER
    
    # Summary section
    summary_row = len(attendance_data['students']) + 6
    totals = attendance_data['class_totals']
    
    ws.cell(row=summary_row, column=1, value="CLASS SUMMARY")
    ws.cell(row=summary_row, column=1).font = Font(bold=True)
    
    ws.cell(row=summary_row + 1, column=1, value=f"Total Students: {attendance_data['total_students']}")
    ws.cell(row=summary_row + 2, column=1, value=f"Male: {attendance_data['total_male']}, Female: {attendance_data['total_female']}")
    ws.cell(row=summary_row + 3, column=1, value=f"Times Opened: {times_opened}")
    ws.cell(row=summary_row + 4, column=1, value=f"Total Morning Attendance: {totals['total_morning']}")
    ws.cell(row=summary_row + 5, column=1, value=f"Total Afternoon Attendance: {totals['total_afternoon']}")
    ws.cell(row=summary_row + 6, column=1, value=f"Total Attendance: {totals['total_attendance']}")
    ws.cell(row=summary_row + 7, column=1, value=f"Weekly Percentage: {totals['weekly_percentage']}%")
    
    auto_adjust_columns(ws)
    
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output


def export_waec_results_to_excel(results_data):
    """
    Export WAEC results to Excel
    
    Args:
        results_data: List of result dictionaries
    
    Returns:
        BytesIO object containing the Excel file
    """
    wb = create_styled_workbook()
    ws = wb.active
    ws.title = "WAEC Results"
    
    # Get all unique subjects
    all_subjects = set()
    for result in results_data:
        all_subjects.update(result.get('subjects', {}).keys())
    subjects = sorted(all_subjects)
    
    # Headers
    headers = ['S/N', 'Student Name', 'Exam Year'] + subjects + ['Total Points', 'Credit Passes']
    
    for col, header in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=header)
    
    style_header_row(ws, 1, len(headers))
    
    # Data rows
    for idx, result in enumerate(results_data, 1):
        row = idx + 1
        ws.cell(row=row, column=1, value=idx)
        ws.cell(row=row, column=2, value=result['student_name'])
        ws.cell(row=row, column=3, value=result['exam_year'])
        
        col = 4
        total_points = 0
        credit_passes = 0
        
        for subject in subjects:
            grade = result.get('subjects', {}).get(subject, '-')
            ws.cell(row=row, column=col, value=grade)
            
            if grade != '-':
                from models import WAECResult
                total_points += WAECResult.grade_to_points(grade)
                if WAECResult.is_pass(grade):
                    credit_passes += 1
            col += 1
        
        ws.cell(row=row, column=col, value=total_points)
        ws.cell(row=row, column=col + 1, value=credit_passes)
        
        for c in range(1, len(headers) + 1):
            ws.cell(row=row, column=c).border = THIN_BORDER
    
    auto_adjust_columns(ws)
    
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output
